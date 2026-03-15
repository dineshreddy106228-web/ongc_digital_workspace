import numpy as np, pandas as pd, pickle, warnings
warnings.filterwarnings('ignore')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

with open('/home/claude/forecast_final.pkl','rb') as f:
    results = pickle.load(f)
with open('/home/claude/stitched_data.pkl','rb') as f:
    d = pickle.load(f)
stitched = d['stitched']

OUTPUT = '/home/claude/xylene_forecast_v1.xlsx'
wb = Workbook()

FN = 'Arial'
def fnt(bold=False, sz=10, col='000000'):
    return Font(name=FN, bold=bold, size=sz, color=col)

def fl(c):
    return PatternFill('solid', fgColor=str(c)[:6], start_color=str(c)[:6])

C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
L  = Alignment(horizontal='left',   vertical='center', wrap_text=False)
R  = Alignment(horizontal='right',  vertical='center')
TH = Side(style='thin', color='BFBFBF')
BD = Border(left=TH, right=TH, top=TH, bottom=TH)

C_HDR='1A4A6B'; C_GRN='1D7A55'; C_AMB='BA7517'; C_RED='A32D2D'
C_OFF='185FA5'; C_ON='3B6D11'; C_TOT='EAF6F1'; C_ALT='F7FAFA'
OFFSHORE=['11A1','12A1','13A1']
SCORE_LIGHT={C_GRN:'C7E6D5', C_AMB:'FFF3CD', C_RED:'FFCDD2'}
PLANTS=['11A1','12A1','13A1','20A1','21A1','22A1','25A1','41A1','50A1','51A1']

def score_color(s):
    return C_GRN if s>=80 else C_AMB if s>=70 else C_RED

def score_label(s):
    return ('High confidence — proceed' if s>=80
            else 'Medium — field review advised' if s>=70
            else 'Low — manual validation required')

def title_row(ws, r, val, nc, bg=C_HDR, sz=13, h=28):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    c = ws.cell(r, 1, val)
    c.font=Font(name=FN,bold=True,size=sz,color='FFFFFF')
    c.fill=fl(bg); c.alignment=C; ws.row_dimensions[r].height=h

def sub_row(ws, r, val, nc, bg='E8F4FD', h=18):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
    c = ws.cell(r, 1, val)
    c.font=Font(name=FN,size=9,italic=True)
    c.fill=fl(bg); c.alignment=C; ws.row_dimensions[r].height=h

def hdr(ws, r, vals, bg, fc='FFFFFF', h=28, sc=1):
    ws.row_dimensions[r].height=h
    for i,v in enumerate(vals, sc):
        c=ws.cell(r,i,v)
        c.font=Font(name=FN,bold=True,size=10,color=fc)
        c.fill=fl(bg); c.alignment=C; c.border=BD

def wcol(ws, col_letter, w):
    ws.column_dimensions[col_letter].width=w

# ══════════════════════════════════════════════════════════════════════
# SHEET 1 — FORECAST SUMMARY
# ══════════════════════════════════════════════════════════════════════
ws1=wb.active; ws1.title='1_Forecast Summary'
ws1.freeze_panes='C5'

fd=list(results['22A1']['forecast_dates'])
NC=2+3*len(fd)+4

title_row(ws1,1,'ONGC Xylene — Demand Forecast  |  Apr-2026 to Sep-2026  |  Litres  |  Confidence threshold: ≥75 for procurement planning',NC,h=30)
sub_row(ws1,2,'Movements 201+221+261 net of reversals  |  Rollup: 11F*/12F*/13F* consumption → Shore Base  |  Confidence capped at 88% (3-year data; 90%+ needs 5+ years)',NC,h=18)

ws1.row_dimensions[3].height=22
ws1.merge_cells(start_row=3,start_column=1,end_row=3,end_column=2)
c=ws1.cell(3,1,'Forecast Unit')
c.font=Font(name=FN,bold=True,size=10,color='FFFFFF')
c.fill=fl(C_HDR); c.alignment=C; c.border=BD

for fi,fdt in enumerate(fd):
    sc=3+fi*3
    ws1.merge_cells(start_row=3,start_column=sc,end_row=3,end_column=sc+2)
    c=ws1.cell(3,sc,fdt.strftime('%b-%Y'))
    c.font=Font(name=FN,bold=True,size=10,color='FFFFFF')
    c.fill=fl(C_GRN); c.alignment=C; c.border=BD

sc_start=3+len(fd)*3
for j,lbl in enumerate(['Conf Score','Rating','Model','Action'],sc_start):
    c=ws1.cell(3,j,lbl)
    c.font=Font(name=FN,bold=True,size=10,color='FFFFFF')
    c.fill=fl(C_HDR); c.alignment=C; c.border=BD

h4=['Unit','Description']
for fdt in fd:
    h4+=['Point (L)','Lower 90%','Upper 90%']
h4+=['Score','Rating','Model','Action']
hdr(ws1,4,h4,C_HDR,h=32)

for ri,plant in enumerate(PLANTS):
    r=results[plant]; er=ri+5
    uc=C_OFF if plant in OFFSHORE else C_ON
    sc=score_color(r['score'])
    action=('Proceed with plan' if r['score']>=80
            else 'Review before ordering' if r['score']>=70
            else 'Manual validation')

    row_data=[plant, r['label']]
    for fi in range(len(fd)):
        row_data+=[round(r['forecast'][fi],0),
                   round(r['lower_90'][fi],0),
                   round(r['upper_90'][fi],0)]
    row_data+=[r['score'],score_label(r['score']),r['model'],action]

    for ci,val in enumerate(row_data,1):
        cell=ws1.cell(er,ci,val); cell.border=BD
        if ci==1:
            cell.font=Font(name=FN,bold=True,size=11,color='FFFFFF')
            cell.fill=fl(uc); cell.alignment=C
        elif ci==2:
            cell.font=fnt(sz=9)
            cell.fill=fl('E8F0FE' if plant in OFFSHORE else 'EAF6E8')
            cell.alignment=L
        elif ci==2+len(fd)*3+1:
            cell.font=Font(name=FN,bold=True,size=11,color='FFFFFF')
            cell.fill=fl(sc); cell.alignment=C
        elif ci==2+len(fd)*3+2:
            cell.font=Font(name=FN,bold=True,size=9)
            cell.fill=fl(SCORE_LIGHT[sc]); cell.alignment=C
        elif ci>=2+len(fd)*3+3:
            cell.font=fnt(sz=9)
            cell.fill=fl(SCORE_LIGHT[sc]); cell.alignment=L
        else:
            col_in_grp=(ci-3)%3
            if col_in_grp==0: cell.fill=fl('EAF6EE'); cell.font=Font(name=FN,bold=True,size=10)
            elif col_in_grp==1: cell.fill=fl('EBF5FB'); cell.font=fnt(sz=10)
            else: cell.fill=fl('FEF9E7'); cell.font=fnt(sz=10)
            cell.alignment=R
            if isinstance(val,(int,float)): cell.number_format='#,##0'
    ws1.row_dimensions[er].height=22

nr=len(PLANTS)+7
ws1.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=NC)
n=ws1.cell(nr,1,'📌  Confidence note: 3 years of data = 3 observations per seasonal position. Score ≥75 = fit for procurement planning with field confirmation. Score <70 = validate manually. 90%+ confidence requires 5+ years of stable data.')
n.font=Font(name=FN,size=9,italic=True,color='5D4E37')
n.fill=fl('FFF9E6'); n.alignment=C; ws1.row_dimensions[nr].height=32

wcol(ws1,'A',12); wcol(ws1,'B',38)
for i in range(3,3+len(fd)*3): wcol(ws1,get_column_letter(i),13)
for i in range(3+len(fd)*3,3+len(fd)*3+4): wcol(ws1,get_column_letter(i),22)

# ══════════════════════════════════════════════════════════════════════
# SHEET 2 — DETAILED FORECAST
# ══════════════════════════════════════════════════════════════════════
ws2=wb.create_sheet('2_Detailed Forecast')

SRC_SHORT={'mb51_23-24':'MB51 23-24','mb51_24-25':'MB51 24-25',
           'mb51_25-26':'MB51 25-26','cons_23-24 (gap-fill)':'Cons 23-24','zero_fill':'—'}
src_df_all=stitched.copy()
src_df_all['src_short']=src_df_all['source'].map(SRC_SHORT).fillna('—')

row=1
for plant in PLANTS:
    r=results[plant]
    uc=C_OFF if plant in OFFSHORE else C_ON
    sc=score_color(r['score'])
    hist_dates=pd.DatetimeIndex(r['history_dates'])
    hist_vals=r['history']
    fd_dates=r['forecast_dates']
    src_lookup=src_df_all[src_df_all['plant_code']==plant].set_index('date')['src_short'].to_dict()

    ws2.merge_cells(start_row=row,start_column=1,end_row=row,end_column=11)
    c=ws2.cell(row,1,f"  {plant}  |  {r['label']}  |  Model: {r['model']}  |  Score: {r['score']}  |  {score_label(r['score'])}")
    c.font=Font(name=FN,bold=True,size=11,color='FFFFFF')
    c.fill=fl(uc); c.alignment=L; ws2.row_dimensions[row].height=24; row+=1

    ws2.merge_cells(start_row=row,start_column=1,end_row=row,end_column=11)
    c2=ws2.cell(row,1,f"  Active months: {r['nz_months']}/36  |  CV: {r['cv']:.2f}  |  Level ratio FY25-26/FY24-25: {r['level_ratio']:.2f}×  |  Avg monthly (active): {r['history'][r['history']>0].mean():.0f} L")
    c2.font=Font(name=FN,size=9,italic=True,color='FFFFFF')
    c2.fill=fl(sc); c2.alignment=L; ws2.row_dimensions[row].height=18; row+=1

    hdr(ws2,row,['Month','FY','Actual (L)','Source','','Fcst Month','Point (L)','Lower 90%','Upper 90%','vs Prior Yr','PI Width'],C_HDR,h=24); row+=1

    for i in range(36):
        dt=hist_dates[i]; v=hist_vals[i]
        fy=(f"FY{str(dt.year)[2:]}-{str(dt.year+1)[2:]}" if dt.month>=4
            else f"FY{str(dt.year-1)[2:]}-{str(dt.year)[2:]}")
        src=src_lookup.get(pd.Timestamp(dt),'—')
        zf=fl('F5F5F5') if v==0 else fl(C_ALT if i%2 else 'FFFFFF')

        for ci,val in enumerate([dt.strftime('%b-%Y'),fy,v,src],1):
            cell=ws2.cell(row,ci,val); cell.border=BD; cell.fill=zf
            cell.font=Font(name=FN,size=9,color='BFBFBF' if v==0 else '222222')
            cell.alignment=C if ci!=4 else L
            if ci==3: cell.number_format='#,##0'
        ws2.cell(row,5,'').border=BD

        if i<len(fd_dates):
            fdt=fd_dates[i]; fv=r['forecast'][i]; fl_=r['lower_90'][i]; fh_=r['upper_90'][i]
            prior_val=hist_vals[i+24] if i+24<36 else 0
            vs_prior=((fv-prior_val)/prior_val*100) if prior_val>0 else 0
            pi_width=fh_-fl_
            for ci,val in enumerate([fdt.strftime('%b-%Y'),round(fv,0),round(fl_,0),round(fh_,0),f"{vs_prior:+.1f}%",round(pi_width,0)],6):
                cell=ws2.cell(row,ci,val); cell.border=BD
                cell.font=Font(name=FN,size=9,bold=(ci==7)); cell.alignment=R if ci!=6 else C
                fills=['F0F0F0','EAF6EE','EBF5FB','FEF9E7',
                       'FFCDD2' if vs_prior<-20 else 'C8E6C9' if vs_prior>20 else 'F5F5F5',
                       'F5F5F5']
                cell.fill=fl(fills[ci-6])
                if ci==10: cell.alignment=C
                elif isinstance(val,(int,float)): cell.number_format='#,##0'
        else:
            for ci in range(6,12):
                ws2.cell(row,ci,'').border=BD; ws2.cell(row,ci).fill=fl('FAFAFA')
        ws2.row_dimensions[row].height=16; row+=1
    row+=2

for c,w in zip('ABCDEFGHIJK',[12,10,14,18,3,12,16,14,14,12,14]):
    wcol(ws2,c,w)

# ══════════════════════════════════════════════════════════════════════
# SHEET 3 — SEASONAL PATTERN
# ══════════════════════════════════════════════════════════════════════
ws3=wb.create_sheet('3_Seasonal Pattern')
title_row(ws3,1,'Seasonal Pattern — Median Consumption by Calendar Month  |  Basis for Forecast  |  Litres',17,h=26)
sub_row(ws3,2,'Each cell = median Litres consumed in that calendar month across 3 FYs. Green intensity = relative demand.',17,h=18)

month_names=['Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','Jan','Feb','Mar']
month_nums=[4,5,6,7,8,9,10,11,12,1,2,3]
hdr(ws3,3,['Unit','Description']+month_names+['Annual (L)','Peak Month','Trough Month'],C_HDR,h=28)

for ri,plant in enumerate(PLANTS):
    r=results[plant]; er=ri+4
    uc=C_OFF if plant in OFFSHORE else C_ON
    seas=[r['seas_med'].get(m,0) for m in month_nums]
    total=sum(seas)
    peak_idx=int(np.argmax(seas))
    trough_idx=int(np.argmin(seas))

    ws3.cell(er,1,plant).font=Font(name=FN,bold=True,size=10,color='FFFFFF')
    ws3.cell(er,1).fill=fl(uc); ws3.cell(er,1).alignment=C; ws3.cell(er,1).border=BD
    ws3.cell(er,2,r['label']).font=fnt(sz=9)
    ws3.cell(er,2).fill=fl('E8F0FE' if plant in OFFSHORE else 'EAF6E8')
    ws3.cell(er,2).alignment=L; ws3.cell(er,2).border=BD

    for ci,v in enumerate(seas,3):
        cell=ws3.cell(er,ci,round(v,0))
        cell.number_format='#,##0'; cell.font=fnt(sz=10)
        cell.alignment=R; cell.border=BD
        cell.fill=fl('F5F5F5') if v==0 else fl('FFFFFF')

    ws3.cell(er,15,round(total,0)).number_format='#,##0'
    ws3.cell(er,15).font=Font(name=FN,bold=True,size=10)
    ws3.cell(er,15).fill=fl(C_TOT); ws3.cell(er,15).border=BD; ws3.cell(er,15).alignment=R

    ws3.cell(er,16,month_names[peak_idx]).fill=fl('C7E6D5')
    ws3.cell(er,16).font=Font(name=FN,bold=True,size=9)
    ws3.cell(er,16).alignment=C; ws3.cell(er,16).border=BD

    ws3.cell(er,17,month_names[trough_idx]).fill=fl('FFE0B2')
    ws3.cell(er,17).font=Font(name=FN,bold=True,size=9)
    ws3.cell(er,17).alignment=C; ws3.cell(er,17).border=BD
    ws3.row_dimensions[er].height=20

ws3.conditional_formatting.add(
    f'C4:{get_column_letter(14)}{len(PLANTS)+3}',
    ColorScaleRule(start_type='min',start_color='FFFFFF',
                   mid_type='percentile',mid_value=50,mid_color='B7DEC8',
                   end_type='max',end_color='1D7A55'))

wcol(ws3,'A',12); wcol(ws3,'B',40)
for i in range(3,18): wcol(ws3,get_column_letter(i),11)
wcol(ws3,get_column_letter(15),14); wcol(ws3,get_column_letter(16),12); wcol(ws3,get_column_letter(17),12)

# ══════════════════════════════════════════════════════════════════════
# SHEET 4 — CONFIDENCE BREAKDOWN
# ══════════════════════════════════════════════════════════════════════
ws4=wb.create_sheet('4_Confidence & Data Quality')
title_row(ws4,1,'Confidence Score Breakdown & Data Quality — Xylene Pilot',9,h=26)
sub_row(ws4,2,'Score components: data richness, volatility (CV), stability (level ratio), model fit. Honest ceiling with 3-year data = 88%.',9,h=18)
hdr(ws4,3,['Unit','Description','Active Months','% Active','CV','Level Ratio','Model','Score','Rating'],C_HDR,h=32)

for ri,plant in enumerate(PLANTS):
    r=results[plant]; er=ri+4
    uc=C_OFF if plant in OFFSHORE else C_ON
    sc=score_color(r['score'])

    row_vals=[plant,r['label'],r['nz_months'],round(r['nz_months']/36*100,1),
              r['cv'],r['level_ratio'],r['model'],r['score'],score_label(r['score'])]

    for ci,val in enumerate(row_vals,1):
        cell=ws4.cell(er,ci,val); cell.border=BD
        if ci==1:
            cell.font=Font(name=FN,bold=True,size=10,color='FFFFFF')
            cell.fill=fl(uc); cell.alignment=C
        elif ci==8:
            cell.font=Font(name=FN,bold=True,size=11,color='FFFFFF')
            cell.fill=fl(sc); cell.alignment=C; cell.number_format='0.0'
        elif ci==9:
            cell.font=Font(name=FN,bold=True,size=9)
            cell.fill=fl(SCORE_LIGHT[sc]); cell.alignment=C
        else:
            cell.font=fnt(sz=10)
            cell.fill=fl(C_ALT if ri%2 else 'FFFFFF')
            cell.alignment=C if ci in (3,4,5,6) else L
        if ci in (4,5,6): cell.number_format='0.00'
    ws4.row_dimensions[er].height=22

mr=len(PLANTS)+6
for off,txt in enumerate([
    'Score Methodology:',
    'Base = 55  |  +12 if active months ≥30  |  +6 if ≥20',
    '+10 if CV < 0.6 (low volatility)  |  +5 if CV < 0.9',
    '+8 if level ratio 0.8–1.2 (stable)  |  +4 if 0.6–1.4',
    '+5 bonus for Damped Holt-Winters (dense, low-CV series)',
    'Hard cap = 88. Score 90%+ requires 5+ years of stable data.',
]):
    ws4.merge_cells(start_row=mr+off,start_column=1,end_row=mr+off,end_column=9)
    c=ws4.cell(mr+off,1,txt)
    c.font=Font(name=FN,size=9,bold=(off==0),italic=(off>0))
    c.fill=fl('FFF9E6' if off>0 else 'F5E6CC'); c.alignment=L
    ws4.row_dimensions[mr+off].height=16

wcol(ws4,'A',12); wcol(ws4,'B',40)
for c,w in zip('CDEFGHI',[16,10,12,14,20,12,28]): wcol(ws4,c,w)

# ══════════════════════════════════════════════════════════════════════
# SHEET 5 — PROCUREMENT SIGNALS
# ══════════════════════════════════════════════════════════════════════
ws5=wb.create_sheet('5_Procurement Signals')
title_row(ws5,1,'Procurement Signals — 6-Month Forward Demand & Reorder Indicators  |  Xylene  |  Litres',10,h=26)
sub_row(ws5,2,'Forecast-based signals. Actual ROP and Days of Cover require MB52 stock-on-hand. Lead times = placeholders — update from ME2M.',10,h=24)
hdr(ws5,3,['Unit','Description','6-Month Fcst (L)','Monthly Avg (L)','Lead Time (mo)*','Safety Stock (L)**','ROP (L)***','FY25-26 Avg (L)','Trend vs FY24-25','Action Flag'],C_HDR,h=36)

for ri,plant in enumerate(PLANTS):
    r=results[plant]; er=ri+4
    uc=C_OFF if plant in OFFSHORE else C_ON
    sc=score_color(r['score'])

    fc6=r['forecast'].sum(); avg6=fc6/6
    fy2526=r['history'][24:]; fy2526_avg=fy2526[fy2526>0].mean() if (fy2526>0).any() else 0
    fy2425=r['history'][12:24]; fy2425_avg=fy2425[fy2425>0].mean() if (fy2425>0).any() else 0
    trend_pct=((fy2526_avg-fy2425_avg)/fy2425_avg*100) if fy2425_avg>0 else 0

    lead_time=2.0 if plant in OFFSHORE else 1.5
    std_dev=r['history'].std()
    safety_stk=round(1.645*std_dev*np.sqrt(lead_time),0)
    rop=round(avg6*lead_time+safety_stk,0)

    action=('⚠️  Validate first' if r['score']<70
            else '📉  Demand declining' if trend_pct<-20
            else '📈  Demand rising — raise safety stock' if trend_pct>30
            else '✅  Standard plan')

    row_vals=[plant,r['label'],round(fc6,0),round(avg6,0),lead_time,
              safety_stk,rop,round(fy2526_avg,0),f"{trend_pct:+.1f}%",action]

    for ci,val in enumerate(row_vals,1):
        cell=ws5.cell(er,ci,val); cell.border=BD
        if ci==1:
            cell.font=Font(name=FN,bold=True,size=10,color='FFFFFF')
            cell.fill=fl(uc); cell.alignment=C
        elif ci==2:
            cell.font=fnt(sz=9)
            cell.fill=fl('E8F0FE' if plant in OFFSHORE else 'EAF6E8'); cell.alignment=L
        elif ci==9:
            t_fill='FFCDD2' if trend_pct<-20 else 'C8E6C9' if trend_pct>20 else 'F5F5F5'
            cell.font=Font(name=FN,bold=True,size=10)
            cell.fill=fl(t_fill); cell.alignment=C
        elif ci==10:
            cell.font=fnt(sz=9); cell.fill=fl(SCORE_LIGHT[sc]); cell.alignment=L
        else:
            cell.font=fnt(sz=10)
            cell.fill=fl(C_ALT if ri%2 else 'FFFFFF'); cell.alignment=R
        if isinstance(val,(int,float)) and ci not in (5,9): cell.number_format='#,##0'
        if ci==5: cell.number_format='0.0'
    ws5.row_dimensions[er].height=22

for off,txt in enumerate([
    '*  Lead time = placeholder. Offshore default=2 months, Onshore=1.5 months. Update from ME2M open PO data.',
    '** Safety stock = 1.645 × σ_monthly × √(lead_time).  Z=1.645 → 95% service level.',
    '*** ROP = Avg monthly forecast × lead time + safety stock. Subtract MB52 stock-on-hand for live net ROP.',
]):
    ws5.merge_cells(start_row=len(PLANTS)+6+off,start_column=1,end_row=len(PLANTS)+6+off,end_column=10)
    c=ws5.cell(len(PLANTS)+6+off,1,txt)
    c.font=Font(name=FN,size=9,italic=True,color='5D4E37')
    c.fill=fl('FFF9E6'); c.alignment=L; ws5.row_dimensions[len(PLANTS)+6+off].height=16

wcol(ws5,'A',12); wcol(ws5,'B',38)
for c,w in zip('CDEFGHIJ',[18,16,14,16,14,16,14,30]): wcol(ws5,c,w)

# ── Save ──────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
