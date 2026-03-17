"""CSC Drafting Workspace — Streamlit entrypoint.

Run: streamlit run csc_app.py
"""

from __future__ import annotations

import io
import logging
import logging.handlers
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import html as _html_module
import pandas as pd
import streamlit as st

# ---------------------------------------------------------------------------
# Path constants — resolved before any imports that depend on CWD
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent.resolve()
DATA_DIR = BASE_DIR / "data"
LOG_DIR  = BASE_DIR / "logs"
EXPORTS_DIR = BASE_DIR / "exports"
BACKUPS_DIR = BASE_DIR / "backups"
CSC_DB_PATH = DATA_DIR / "CSC_Data.db"
HCC_LOGO_PATH = BASE_DIR / "_hcc_logo_source.jpg"


# ---------------------------------------------------------------------------
# Logging setup (rotating, runs once per process)
# ---------------------------------------------------------------------------

def _ensure_runtime_dirs() -> None:
    for path in (DATA_DIR, LOG_DIR, EXPORTS_DIR, BACKUPS_DIR):
        path.mkdir(parents=True, exist_ok=True)


def _setup_logging() -> None:
    _ensure_runtime_dirs()
    root = logging.getLogger()
    formatter = logging.Formatter(
        "%(asctime)s %(levelname)-8s %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    target_log = (LOG_DIR / "app.log").resolve()

    has_file_handler = any(
        isinstance(h, logging.handlers.RotatingFileHandler)
        and Path(getattr(h, "baseFilename", "")).resolve() == target_log
        for h in root.handlers
    )
    if not has_file_handler:
        fh = logging.handlers.RotatingFileHandler(
            target_log,
            maxBytes=5 * 1024 * 1024,
            backupCount=3,
            encoding="utf-8",
        )
        fh.setFormatter(formatter)
        root.addHandler(fh)

    has_stdout_handler = any(
        isinstance(h, logging.StreamHandler) and h.stream is sys.stdout
        for h in root.handlers
    )
    if not has_stdout_handler:
        sh = logging.StreamHandler(sys.stdout)
        sh.setFormatter(formatter)
        root.addHandler(sh)

    root.setLevel(logging.INFO)


_setup_logging()
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Late imports (after logging is ready)
# ---------------------------------------------------------------------------

from csc_repository import (
    CSCRepository,
    normalize_parameter_type_label,
    normalize_test_procedure_type,
)
from csc_pages import render_active_section
from csc_admin_pages import render_admin_workspace
from csc_utils import (
    SECTION_ORDER,
    SECTION_BACKGROUND,
    ROLE_ADMIN,
    ROLE_USER,
    ADMIN_PASSPHRASE,
    DEFAULT_PREPARED_BY,
    fmt_ist,
    parse_spec_number,
)

# ---------------------------------------------------------------------------
# Session-state defaults
# ---------------------------------------------------------------------------

_SESSION_DEFAULTS: dict[str, Any] = {
    "csc_active_draft_id":   None,
    "csc_active_section":    SECTION_BACKGROUND,
    "csc_user_name":         "committee_1",
    "csc_role":              ROLE_USER,      # "Admin" | "User"
    "csc_show_new_form":     False,
    "csc_show_load_form":    False,
    "csc_draft_list_cache":  None,
    "csc_last_cache_epoch":  0.0,
    "csc_show_meta_edit":    False,
    "csc_admin_pass_ok":     False,          # True once passphrase verified
}

_COMMITTEE_USERS: dict[str, dict[str, Any]] = {
    "committee_1": {
        "label": "Committee 1 (DFC, CCA)",
        "subsets": {"DFC", "CCA"},
    },
    "committee_2": {
        "label": "Committee 2 (WS)",
        "subsets": {"WS"},
    },
    "committee_3": {
        "label": "Committee 3 (PC, WIC)",
        "subsets": {"PC", "WIC"},
    },
    "committee_4": {
        "label": "Committee 4 (WM, UTL, LPG)",
        "subsets": {"WM", "UTL", "LPG"},
    },
}
_PDF_PARSE_CACHE_VER = "v3"
TYPE_CLASS_EXERCISE_TYPES = ["Essential", "Desirable"]
TYPE_CLASS_ANY_MORE_OPTIONS = ["Yes", "No"]
TYPE_CLASS_TEST_PROC_TYPES = [
    "ASTM",
    "API",
    "IS",
    "Inhouse",
    "Other National / International",
]


def _committee_label(user_name: str) -> str:
    profile = _COMMITTEE_USERS.get(user_name, {})
    return str(profile.get("label") or user_name or "Committee User")


def _committee_subsets(user_name: str) -> set[str]:
    profile = _COMMITTEE_USERS.get(user_name, {})
    return set(profile.get("subsets") or set())


def _spec_subset(spec_number: str) -> str | None:
    subset, _seq, _year = parse_spec_number(spec_number or "")
    return subset


def _filter_specs_for_subsets(
    specs: list[dict[str, Any]],
    allowed_subsets: set[str],
) -> list[dict[str, Any]]:
    if not allowed_subsets:
        return []
    filtered = []
    for spec in specs:
        subset = _spec_subset(str(spec.get("spec_number", "") or ""))
        if subset in allowed_subsets:
            filtered.append(spec)
    return filtered


def _is_draft_allowed_for_subsets(
    draft: dict[str, Any] | None,
    allowed_subsets: set[str],
) -> bool:
    if not draft:
        return False
    subset = _spec_subset(str(draft.get("spec_number", "") or ""))
    return subset in allowed_subsets


def _ensure_session_state() -> None:
    for key, val in _SESSION_DEFAULTS.items():
        if key not in st.session_state:
            st.session_state[key] = val
    if st.session_state.get("csc_user_name") not in _COMMITTEE_USERS:
        st.session_state["csc_user_name"] = "committee_1"


# ---------------------------------------------------------------------------
# Repository factory
# ---------------------------------------------------------------------------

@st.cache_resource
def _get_repo() -> CSCRepository:
    """Initialize the repository exactly once per Streamlit server process.

    st.cache_resource persists the object across reruns and sessions, so
    ensure_db() runs only once instead of on every rerun.
    """
    _ensure_runtime_dirs()
    repo = CSCRepository(str(CSC_DB_PATH))
    repo.ensure_db()
    logger.info("Database initialized at %s", CSC_DB_PATH)
    return repo


def _repo() -> CSCRepository:
    return _get_repo()


# ---------------------------------------------------------------------------
# CSS injection
# ---------------------------------------------------------------------------

_CSS = """
<style>
:root {
    --csc-copper:    #A54B23;
    --csc-copper-dk: #7A3318;
    --csc-bg:        #F8F6F4;
    --csc-bg2:       #FAF6F2;
    --csc-ink:       #1F2937;
    --csc-muted:     #6B7280;
    --csc-border:    #E5E7EB;
}

/* ---- Sidebar brand strip ---- */
section[data-testid="stSidebar"] {
    background: var(--csc-bg2);
    border-right: 1px solid var(--csc-border);
}

/* ---- Metadata card ---- */
.csc-meta-card {
    background: linear-gradient(180deg, #fff7f3 0%, #fffcfa 100%);
    border: 1px solid var(--csc-border);
    border-left: 4px solid var(--csc-copper);
    border-radius: 10px;
    padding: 0.7rem 0.9rem;
    margin: 0.5rem 0 0.8rem;
}
.csc-meta-spec {
    font-size: 0.72rem;
    font-weight: 800;
    letter-spacing: 0.07em;
    text-transform: uppercase;
    color: var(--csc-copper-dk);
    margin-bottom: 0.2rem;
}
.csc-meta-name {
    font-size: 0.98rem;
    font-weight: 700;
    color: var(--csc-ink);
    margin-bottom: 0.15rem;
}
.csc-meta-row {
    font-size: 0.78rem;
    color: var(--csc-muted);
}
.csc-meta-status {
    display: inline-block;
    margin-top: 0.35rem;
    padding: 0.1rem 0.5rem;
    border-radius: 999px;
    font-size: 0.7rem;
    font-weight: 700;
    letter-spacing: 0.04em;
    text-transform: uppercase;
    background: #FEE2D5;
    color: var(--csc-copper-dk);
}
.csc-meta-status.final {
    background: #D1FAE5;
    color: #065F46;
}
.csc-meta-status.review {
    background: #FEF9C3;
    color: #713F12;
}

/* ---- Section navigator ---- */
.csc-nav-label {
    font-size: 0.7rem;
    font-weight: 700;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    color: var(--csc-muted);
    margin: 0.6rem 0 0.3rem;
}

/* ---- Main content area ---- */
.main .block-container {
    padding-top: 1.5rem;
    max-width: 100%;
}

/* ---- Section header divider ---- */
hr {
    border: none;
    border-top: 1px solid var(--csc-border);
    margin: 0.3rem 0 1rem;
}

/* ---- Buttons ---- */
.stButton > button[kind="primary"] {
    background: var(--csc-copper) !important;
    border-color: var(--csc-copper) !important;
    color: white !important;
    border-radius: 6px !important;
}
.stButton > button[kind="primary"]:hover {
    background: var(--csc-copper-dk) !important;
    border-color: var(--csc-copper-dk) !important;
}

/* ---- Data editor/text wrapping ---- */
div[data-testid="stDataEditor"] [role="gridcell"],
div[data-testid="stDataFrame"] [role="gridcell"] {
    white-space: normal !important;
    word-break: break-word !important;
    line-height: 1.3 !important;
}
div[data-testid="stDataEditor"] [role="gridcell"] > div,
div[data-testid="stDataFrame"] [role="gridcell"] > div {
    overflow: visible !important;
    text-overflow: clip !important;
}
div[data-testid="stDataEditor"] {
    overflow-x: auto !important;
}
div[data-testid="stDataEditor"] [role="grid"] {
    min-width: 1400px !important;
}

/* ---- App header ---- */
.csc-app-header {
    display: flex;
    align-items: center;
    gap: 0.8rem;
    padding: 0.2rem 0 0.8rem;
    border-bottom: 2px solid var(--csc-copper);
    margin-bottom: 1.2rem;
}
.csc-app-title {
    font-size: 1.3rem;
    font-weight: 800;
    color: var(--csc-copper);
    letter-spacing: -0.01em;
}
.csc-app-subtitle {
    font-size: 0.8rem;
    color: var(--csc-muted);
    margin-top: -0.2rem;
}
/* ---- Role badge ---- */
.csc-role-admin {
    display:inline-block;padding:2px 10px;border-radius:999px;
    background:#FEE2D5;color:#7A3318;font-size:0.72rem;font-weight:700;
    letter-spacing:.04em;text-transform:uppercase;margin-top:.2rem;
}
.csc-role-user {
    display:inline-block;padding:2px 10px;border-radius:999px;
    background:#DBEAFE;color:#1E40AF;font-size:0.72rem;font-weight:700;
    letter-spacing:.04em;text-transform:uppercase;margin-top:.2rem;
}
</style>
"""


# ---------------------------------------------------------------------------
# Sidebar: metadata card
# ---------------------------------------------------------------------------

def _render_meta_card(draft: dict[str, Any]) -> None:
    def _e(val: Any, fallback: str = "—") -> str:
        return _html_module.escape(str(val or fallback), quote=True)

    status = draft.get("status", "Draft")
    status_cls = (
        "final"  if status == "Final" else
        "review" if status == "Under Review" else ""
    )
    st.markdown(
        f"""
<div class="csc-meta-card">
  <div class="csc-meta-spec">{_e(draft.get('spec_number'))}</div>
  <div class="csc-meta-name">{_e(draft.get('chemical_name'))}</div>
  <div class="csc-meta-row">Prepared by: {_e(draft.get('prepared_by'))}</div>
  <div class="csc-meta-row">Date: {_e(draft.get('meeting_date'))}</div>
  <span class="csc-meta-status {_html_module.escape(status_cls)}">{_e(status)}</span>
</div>
""",
        unsafe_allow_html=True,
    )


# ---------------------------------------------------------------------------
# Sidebar: new draft form
# ---------------------------------------------------------------------------

def _render_new_draft_form(repo: CSCRepository) -> None:
    """New draft form with optional PDF spec upload for auto-population."""

    st.markdown("##### New Draft")

    # ── Step 1: Optional PDF upload ──────────────────────────────────────
    st.markdown("**Step 1 — Upload Existing Specification PDF** *(optional)*")
    uploaded_pdf = st.file_uploader(
        "Upload the current ONGC spec PDF to auto-fill parameters",
        type=["pdf"],
        key="new_draft_pdf_upload",
        help="Upload the existing Corporate Specification PDF. "
             "Metadata and all parameters will be extracted automatically.",
    )

    # Parse PDF and store in session state to survive form submission
    extracted_spec = None
    if uploaded_pdf is not None:
        pdf_key = f"csc_extracted_spec_{_PDF_PARSE_CACHE_VER}_{uploaded_pdf.name}_{uploaded_pdf.size}"
        if pdf_key not in st.session_state:
            with st.spinner("Extracting parameters from PDF…"):
                try:
                    from csc_pdf_extractor import extract_spec_from_pdf
                    pdf_bytes = uploaded_pdf.read()
                    spec_doc  = extract_spec_from_pdf(pdf_bytes)
                    st.session_state[pdf_key] = spec_doc
                    if spec_doc.parse_warnings:
                        for w in spec_doc.parse_warnings:
                            st.warning(w)
                    else:
                        st.success(
                            f"✓ Extracted **{len(spec_doc.parameters)}** parameters "
                            f"from **{spec_doc.chemical_name}**"
                        )
                except Exception as exc:
                    st.error(f"PDF extraction failed: {exc}")
                    logger.exception("PDF extraction error: %s", exc)
                    st.session_state[pdf_key] = None
        extracted_spec = st.session_state.get(pdf_key)

    # Show extraction summary if available
    if extracted_spec and extracted_spec.parameters:
        with st.expander(f"Preview: {len(extracted_spec.parameters)} extracted parameters", expanded=False):
            for p in extracted_spec.parameters:
                grp = f"[{p.group}] " if p.group else ""
                st.markdown(
                    f"**{p.number}.** {grp}{p.name}  \n"
                    f"<span style='color:#6B7280;font-size:0.85rem;'>{p.existing_value}</span>",
                    unsafe_allow_html=True,
                )

    st.markdown("**Step 2 — Confirm / Edit Draft Details**")

    # Pre-fill fields from extracted spec if available
    default_spec      = extracted_spec.spec_number      if extracted_spec else ""
    default_chem      = extracted_spec.chemical_name    if extracted_spec else ""
    default_test_proc = extracted_spec.test_procedure   if extracted_spec else ""
    default_mat_code  = extracted_spec.material_code    if extracted_spec else ""

    with st.form("new_draft_form", clear_on_submit=True):
        spec = st.text_input(
            "Specification Number *",
            value=default_spec,
            placeholder="e.g. ONGC/DFC/73/2026",
            max_chars=200,
        )
        chem = st.text_input(
            "Chemical Name *",
            value=default_chem,
            placeholder="e.g. XC Polymer for Drilling Fluids",
            max_chars=300,
        )
        committee = st.text_input(
            "Committee Name",
            value="Chemical Specification Committee",
            max_chars=300,
        )
        col_tp, col_mc = st.columns(2)
        with col_tp:
            test_proc = st.text_input(
                "Test Procedure",
                value=default_test_proc,
                max_chars=300,
                help="e.g. ONGC/ Test Procedure Vol - I / DFC / 73",
            )
        with col_mc:
            mat_code = st.text_input(
                "Material Code",
                value=default_mat_code,
                max_chars=100,
                help="e.g. 100101210",
            )
        date_str  = st.text_input(
            "Meeting Date",
            placeholder="e.g. 15 March 2026",
            max_chars=50,
        )
        prepared  = st.text_input("Prepared By", value=DEFAULT_PREPARED_BY, max_chars=200)
        reviewed  = st.text_input("Reviewed By", max_chars=200)

        submitted = st.form_submit_button("Create Draft", type="primary")

    if submitted:
        if not spec.strip() or not chem.strip():
            st.error("Specification Number and Chemical Name are required.")
            return
        user = st.session_state.get("csc_user_name") or "Unknown"
        try:
            draft_id = repo.create_draft(
                spec_number    = spec.strip(),
                chemical_name  = chem.strip(),
                committee_name = committee.strip() or "Chemical Specification Committee",
                meeting_date   = date_str.strip(),
                prepared_by    = prepared.strip(),
                reviewed_by    = reviewed.strip(),
                user_name      = user,
                test_procedure = test_proc.strip(),
                material_code  = mat_code.strip(),
            )

            # If PDF was extracted, bulk-insert parameters pre-populated
            if extracted_spec and extracted_spec.parameters:
                from csc_pdf_extractor import spec_to_parameter_dicts
                param_dicts = spec_to_parameter_dicts(extracted_spec)
                repo.upsert_parameters(draft_id, param_dicts, user_name=user)
                repo.log_audit(
                    draft_id, "PDF_IMPORT", user,
                    new_value=f"Imported {len(param_dicts)} parameters from PDF",
                )

            st.session_state["csc_active_draft_id"]  = draft_id
            st.session_state["csc_active_section"]   = SECTION_BACKGROUND
            st.session_state["csc_show_new_form"]    = False
            st.session_state["csc_draft_list_cache"] = None
            # Clear the extracted spec from session
            for k in list(st.session_state.keys()):
                if k.startswith("csc_extracted_spec_"):
                    del st.session_state[k]

            n_params = len(extracted_spec.parameters) if extracted_spec else 0
            msg = f"Draft created: **{spec.strip()}**"
            if n_params:
                msg += f" · {n_params} parameters imported from PDF"
            st.success(msg)
            st.rerun()
        except Exception as exc:
            st.error(f"Failed to create draft: {exc}")
            logger.exception("Draft creation error: %s", exc)


# ---------------------------------------------------------------------------
# Sidebar: load draft selector
# ---------------------------------------------------------------------------

def _render_load_draft(
    repo: CSCRepository,
    user_name: str,
    allowed_subsets: set[str],
) -> None:
    import time

    now = time.time()
    cache_age = now - st.session_state.get("csc_last_cache_epoch", 0.0)
    cache_scope = f"{user_name}|{','.join(sorted(allowed_subsets))}"
    cached_payload = st.session_state.get("csc_draft_list_cache") or {}
    stale_scope = cached_payload.get("scope") != cache_scope
    if st.session_state["csc_draft_list_cache"] is None or cache_age > 30 or stale_scope:
        published_all = repo.list_reviewable_specs()
        published = _filter_specs_for_subsets(published_all, allowed_subsets)

        revisions_all = repo.list_user_revision_drafts(user_name)
        revisions = _filter_specs_for_subsets(revisions_all, allowed_subsets)

        st.session_state["csc_draft_list_cache"] = {
            "scope": cache_scope,
            "published": published,
            "revisions": revisions,
        }
        st.session_state["csc_last_cache_epoch"] = now

    cached = st.session_state["csc_draft_list_cache"] or {}
    published = cached.get("published", [])
    revisions = cached.get("revisions", [])

    if not published and not revisions:
        subset_text = ", ".join(sorted(allowed_subsets))
        st.sidebar.info(
            f"No published specs are assigned to this committee yet. Allowed subsets: {subset_text}"
        )
        return

    st.sidebar.caption(
        "Assigned subsets: " + ", ".join(sorted(allowed_subsets))
    )

    all_display = []
    if published:
        all_display += [
            (
                f"📋 {d['spec_number']} — {d['chemical_name']}  [Master v{d.get('spec_version', 1)}]",
                ("master", d["draft_id"]),
            )
            for d in published
        ]
    if revisions:
        all_display += [
            (
                f"📝 {d['spec_number']} — {d['chemical_name']}  [{d.get('status', 'Draft')}]",
                ("revision", d["draft_id"]),
            )
            for d in revisions
        ]

    if not all_display:
        st.sidebar.info("No specs available.")
        return

    options = dict(all_display)
    choice = st.sidebar.selectbox(
        "Assigned Specifications",
        list(options.keys()),
        key=f"load_draft_selector_{user_name}",
    )
    mode, selected_id = options[choice]
    btn_label = "Open for Modification" if mode == "master" else "Continue Modification"
    if st.sidebar.button(btn_label, key=f"load_draft_btn_{user_name}", type="primary"):
        active_draft_id = selected_id
        if mode == "master":
            active_draft_id, _created = repo.create_or_get_revision_draft(selected_id, user_name)

        st.session_state["csc_active_draft_id"] = active_draft_id
        st.session_state["csc_active_section"]  = SECTION_BACKGROUND
        st.session_state["csc_draft_list_cache"] = None
        st.rerun()


# ---------------------------------------------------------------------------
# Sidebar: metadata editor (inline)
# ---------------------------------------------------------------------------

def _render_meta_editor(repo: CSCRepository, draft: dict[str, Any]) -> None:
    with st.form("meta_edit_form"):
        st.markdown("##### Edit Draft Details")
        spec = st.text_input("Specification Number", value=draft.get("spec_number", ""), max_chars=200)
        chem = st.text_input("Chemical Name", value=draft.get("chemical_name", ""), max_chars=300)
        committee = st.text_input("Committee Name", value=draft.get("committee_name", ""), max_chars=300)
        date_str  = st.text_input("Meeting Date", value=draft.get("meeting_date", ""), max_chars=50)
        prepared  = st.text_input(
            "Prepared By",
            value=draft.get("prepared_by", DEFAULT_PREPARED_BY) or DEFAULT_PREPARED_BY,
            max_chars=200,
        )
        reviewed  = st.text_input("Reviewed By", value=draft.get("reviewed_by", ""), max_chars=200)
        st.caption(f"Status: **{draft.get('status', 'Draft')}**")
        col1, col2 = st.columns(2)
        with col1:
            save = st.form_submit_button("Save", type="primary")
        with col2:
            cancel = st.form_submit_button("Cancel")

    if save:
        user = st.session_state.get("csc_user_name") or "Unknown"
        try:
            repo.update_draft_metadata(
                draft_id       = draft["draft_id"],
                spec_number    = spec,
                chemical_name  = chem,
                committee_name = committee,
                meeting_date   = date_str,
                prepared_by    = prepared,
                reviewed_by    = reviewed,
                status         = draft.get("status", "Draft"),
                user_name      = user,
            )
            st.session_state["csc_show_meta_edit"]  = False
            st.session_state["csc_draft_list_cache"] = None
            st.success("Details updated.")
            st.rerun()
        except Exception as exc:
            st.error(f"Update failed: {exc}")

    if cancel:
        st.session_state["csc_show_meta_edit"] = False
        st.rerun()


# ---------------------------------------------------------------------------
# Sidebar renderer
# ---------------------------------------------------------------------------

def _render_role_switcher() -> None:
    """Sidebar widget to toggle between User and Admin mode."""
    current_role = st.session_state.get("csc_role", ROLE_USER)
    role_label_cls = "csc-role-admin" if current_role == ROLE_ADMIN else "csc-role-user"
    st.sidebar.markdown(
        f'<span class="{role_label_cls}">{current_role} Mode</span>',
        unsafe_allow_html=True,
    )

    if current_role == ROLE_USER:
        if st.sidebar.button(
            "Switch to Admin Mode",
            key="btn_switch_admin",
            use_container_width=True,
        ):
            st.session_state["csc_show_admin_gate"] = True
            st.rerun()

        # Passphrase gate
        if st.session_state.get("csc_show_admin_gate"):
            with st.sidebar.form("admin_gate_form"):
                entered = st.text_input(
                    "Admin Passphrase",
                    type="password",
                    placeholder="Enter passphrase",
                    max_chars=100,
                )
                ok = st.form_submit_button("Unlock", type="primary")
            if ok:
                if entered == ADMIN_PASSPHRASE:
                    st.session_state["csc_role"]            = ROLE_ADMIN
                    st.session_state["csc_admin_pass_ok"]   = True
                    st.session_state["csc_show_admin_gate"] = False
                    # Close any open committee draft
                    st.session_state["csc_active_draft_id"] = None
                    st.rerun()
                else:
                    st.sidebar.error("Incorrect passphrase.")
    else:
        # Currently in admin mode
        if st.sidebar.button(
            "Switch to User Mode",
            key="btn_switch_user",
            use_container_width=True,
        ):
            st.session_state["csc_role"]          = ROLE_USER
            st.session_state["csc_admin_pass_ok"] = False
            st.session_state["csc_active_draft_id"] = None
            st.rerun()


def _render_sidebar(repo: CSCRepository) -> None:
    with st.sidebar:
        # App brand
        st.markdown(
            '<div class="csc-app-title" style="font-size:1.05rem;font-weight:800;'
            'color:#A54B23;padding:0.4rem 0;">CSC Drafting Workspace</div>',
            unsafe_allow_html=True,
        )
        st.caption("ONGC · Corporate Chemistry")
        if HCC_LOGO_PATH.exists():
            st.image(str(HCC_LOGO_PATH), width=140)
        st.divider()

        # Committee user login
        user_ids = list(_COMMITTEE_USERS.keys())
        current_user = st.session_state.get("csc_user_name", "committee_1")
        if current_user not in user_ids:
            current_user = "committee_1"
        selected_user = st.selectbox(
            "Committee User",
            options=user_ids,
            index=user_ids.index(current_user),
            format_func=_committee_label,
            key="sidebar_committee_user",
        )
        if selected_user != current_user:
            st.session_state["csc_user_name"] = selected_user
            st.session_state["csc_active_draft_id"] = None
            st.session_state["csc_active_section"] = SECTION_BACKGROUND
            st.session_state["csc_draft_list_cache"] = None
            st.rerun()
        user_name = selected_user
        st.caption("Logged in as: " + _committee_label(user_name))

        # Role switcher
        st.divider()
        _render_role_switcher()

        current_role = st.session_state.get("csc_role", ROLE_USER)

        # ── Admin mode: no draft controls; admin workspace is in main panel ──
        if current_role == ROLE_ADMIN:
            st.divider()
            st.info(
                "You are in **Admin Mode**.\n\n"
                "Use the main panel to ingest PDFs, manage specs, and export.",
                icon="🛠",
            )
            return

        # ── User mode: committee-filtered direct spec picker ───────────────
        st.divider()
        allowed_subsets = _committee_subsets(user_name)
        _render_load_draft(repo, user_name, allowed_subsets)

        # Active draft card
        draft_id = st.session_state.get("csc_active_draft_id")
        if draft_id:
            draft = repo.load_draft(draft_id)
            if draft and _is_draft_allowed_for_subsets(draft, allowed_subsets):
                st.divider()
                _render_meta_card(draft)

                col_edit, col_del = st.columns(2)
                with col_edit:
                    if st.button("Edit Details", key="btn_edit_meta", use_container_width=True):
                        st.session_state["csc_show_meta_edit"] = not st.session_state.get("csc_show_meta_edit", False)
                with col_del:
                    if st.button("Close Draft", key="btn_close_draft", use_container_width=True):
                        st.session_state["csc_active_draft_id"] = None
                        st.session_state["csc_active_section"]  = SECTION_BACKGROUND
                        st.rerun()

                if st.session_state.get("csc_show_meta_edit"):
                    _render_meta_editor(repo, draft)
            elif draft:
                st.warning("This spec is not assigned to the selected committee user.")
                st.session_state["csc_active_draft_id"] = None
                st.session_state["csc_active_section"] = SECTION_BACKGROUND
                st.rerun()

        # Section navigator
        if draft_id:
            st.divider()
            st.markdown('<div class="csc-nav-label">Sections</div>', unsafe_allow_html=True)
            active_section = st.session_state.get("csc_active_section", SECTION_BACKGROUND)
            for i, section in enumerate(SECTION_ORDER, start=1):
                is_active = section == active_section
                btn_label = f"{i}. {section}"
                if st.button(
                    btn_label,
                    key=f"nav_{section}",
                    use_container_width=True,
                    type="primary" if is_active else "secondary",
                ):
                    st.session_state["csc_active_section"] = section
                    st.rerun()


# ---------------------------------------------------------------------------
# Main panel header
# ---------------------------------------------------------------------------

def _render_main_header(role: str = ROLE_USER) -> None:
    subtitle = (
        "Admin Workspace &nbsp;·&nbsp; ONGC Corporate Chemistry"
        if role == ROLE_ADMIN
        else "Chemical Specification Committee &nbsp;·&nbsp; ONGC Corporate Chemistry"
    )
    icon = "🛠" if role == ROLE_ADMIN else "📋"
    col_title, col_logo = st.columns([6, 1], vertical_alignment="center")
    with col_title:
        st.markdown(
            f"""
<div class="csc-app-header">
  <div>
    <div class="csc-app-title">{icon}&nbsp; CSC Drafting Workspace</div>
    <div class="csc-app-subtitle">{subtitle}</div>
  </div>
</div>
""",
            unsafe_allow_html=True,
        )
    with col_logo:
        if HCC_LOGO_PATH.exists():
            st.image(str(HCC_LOGO_PATH), width=110)


def _type_to_workbook_label(raw: str) -> str:
    return normalize_parameter_type_label(raw or "")


def _parse_yes_no_cell(raw: Any) -> bool:
    txt = str(raw or "").strip().lower()
    return txt in {"yes", "y", "true", "1"}


def _excel_sheet_title(raw: str, used_titles: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", str(raw or "").strip()) or "Specification"
    base = base[:31]
    candidate = base
    idx = 1
    while candidate in used_titles:
        suffix = f"_{idx}"
        candidate = f"{base[:max(1, 31 - len(suffix))]}{suffix}"
        idx += 1
    used_titles.add(candidate)
    return candidate


def _build_type_classification_workbook(
    repo: CSCRepository,
    published_specs: list[dict[str, Any]],
) -> tuple[bytes, int, int]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()
    spec_count = 0
    row_count = 0

    headers = [
        "Spec Number",
        "Chemical Name",
        "Material Code",
        "Draft ID",
        "Parameter ID",
        "S. No.",
        "Parameter Name",
        "Requirement",
        "Test Method",
        "Current Type",
        "Type",
        "Any more changes proposed ?",
        "Test Procedure Type",
        "Test Procedure Text",
    ]

    for spec in published_specs:
        draft_id = str(spec.get("draft_id", "") or "").strip()
        if not draft_id:
            continue
        spec_no = str(spec.get("spec_number", "") or "").strip()
        chem = str(spec.get("chemical_name", "") or "").strip()
        draft_meta = repo.load_draft(draft_id) or {}
        material_code = str(draft_meta.get("material_code", "") or "").strip()
        sheet_name = _excel_sheet_title(spec_no or chem, used_titles)
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)

        header_fill = PatternFill(start_color="E9F2FF", end_color="E9F2FF", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = header_fill

        params = repo.load_parameters(draft_id)
        for idx, p in enumerate(params, start=1):
            current_type = _type_to_workbook_label(str(p.get("parameter_type", "Essential") or "Essential"))
            ws.append(
                [
                    spec_no,
                    chem,
                    material_code,
                    draft_id,
                    str(p.get("parameter_id", "") or ""),
                    idx,
                    str(p.get("parameter_name", "") or ""),
                    str(p.get("existing_value", "") or ""),
                    str(p.get("test_method", "") or ""),
                    current_type,
                    current_type,
                    "No",
                    normalize_test_procedure_type(str(p.get("test_procedure_type", "") or "")),
                    str(
                        p.get("test_procedure_text", "")
                        or p.get("test_method", "")
                        or ""
                    ),
                ]
            )
            row_count += 1

        last_row = ws.max_row

        if last_row >= 2:
            type_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(TYPE_CLASS_EXERCISE_TYPES)}"',
                allow_blank=False,
            )
            any_more_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(TYPE_CLASS_ANY_MORE_OPTIONS)}"',
                allow_blank=False,
            )
            test_proc_type_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(TYPE_CLASS_TEST_PROC_TYPES)}"',
                allow_blank=False,
            )
            ws.add_data_validation(type_validation)
            ws.add_data_validation(any_more_validation)
            ws.add_data_validation(test_proc_type_validation)
            type_validation.add(f"K2:K{last_row}")
            any_more_validation.add(f"L2:L{last_row}")
            test_proc_type_validation.add(f"M2:M{last_row}")

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].hidden = True
        ws.column_dimensions["E"].hidden = True
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 38
        ws.column_dimensions["H"].width = 38
        ws.column_dimensions["I"].width = 28
        ws.column_dimensions["J"].width = 20
        ws.column_dimensions["K"].width = 24
        ws.column_dimensions["L"].width = 30
        ws.column_dimensions["M"].width = 36
        ws.column_dimensions["N"].width = 42

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:N{max(1, ws.max_row)}"
        spec_count += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), spec_count, row_count


def _parse_type_classification_workbook(
    workbook_bytes: bytes,
    allowed_draft_ids: set[str],
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    updates: dict[str, list[dict[str, Any]]] = {}
    summary: dict[str, Any] = {
        "sheets_seen": 0,
        "rows_seen": 0,
        "rows_loaded": 0,
        "rows_skipped": 0,
        "errors": [],
    }

    required = {
        "draft id",
        "parameter id",
        "current type",
        "type",
        "any more changes proposed ?",
        "test procedure type",
        "test procedure text",
    }

    for ws in wb.worksheets:
        summary["sheets_seen"] += 1
        header_map: dict[str, int] = {}
        for idx, cell in enumerate(ws[1], start=1):
            header_map[str(cell.value or "").strip().lower()] = idx
        missing = [name for name in required if name not in header_map]
        if missing:
            summary["errors"].append(
                f"{ws.title}: missing required columns ({', '.join(sorted(missing))})"
            )
            continue

        for row in range(2, ws.max_row + 1):
            draft_id = str(ws.cell(row=row, column=header_map["draft id"]).value or "").strip()
            parameter_id = str(ws.cell(row=row, column=header_map["parameter id"]).value or "").strip()
            if not draft_id and not parameter_id:
                continue

            summary["rows_seen"] += 1
            if not draft_id or draft_id not in allowed_draft_ids or not parameter_id:
                summary["rows_skipped"] += 1
                continue

            current_type = str(ws.cell(row=row, column=header_map["current type"]).value or "").strip()
            edited_type = str(ws.cell(row=row, column=header_map["type"]).value or "").strip()
            any_more_raw = ws.cell(
                row=row,
                column=header_map["any more changes proposed ?"],
            ).value
            updates.setdefault(draft_id, []).append(
                {
                    "parameter_id": parameter_id,
                    "type": normalize_parameter_type_label(edited_type or current_type),
                    "any_more_changes": _parse_yes_no_cell(any_more_raw),
                    "test_procedure_type": normalize_test_procedure_type(
                        str(
                            ws.cell(
                                row=row,
                                column=header_map["test procedure type"],
                            ).value
                            or ""
                        ).strip()
                    ),
                    "test_procedure_text": str(
                        ws.cell(
                            row=row,
                            column=header_map["test procedure text"],
                        ).value
                        or ""
                    ).strip(),
                }
            )
            summary["rows_loaded"] += 1

    return updates, summary


def _summarize_type_classification_preview(
    repo: CSCRepository,
    updates: dict[str, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    preview_rows: list[dict[str, Any]] = []
    for draft_id, rows in updates.items():
        draft = repo.load_draft(draft_id)
        if not draft:
            continue
        params = repo.load_parameters(draft_id)
        by_pid = {str(p.get("parameter_id", "") or ""): p for p in params}
        type_changes = 0
        proc_changes = 0
        flagged = 0
        for row in rows:
            pid = str(row.get("parameter_id", "") or "")
            if not pid:
                continue
            current = by_pid.get(pid, {})
            old_type = normalize_parameter_type_label(
                str(current.get("parameter_type", "Essential") or "Essential")
            )
            new_type = normalize_parameter_type_label(str(row.get("type", "") or ""))
            if old_type != new_type:
                type_changes += 1
            old_proc_type = normalize_test_procedure_type(
                str(current.get("test_procedure_type", "") or "")
            )
            new_proc_type = normalize_test_procedure_type(
                str(row.get("test_procedure_type", "") or "")
            )
            old_proc_text = str(current.get("test_procedure_text", "") or "")
            new_proc_text = str(row.get("test_procedure_text", "") or "")
            if old_proc_type != new_proc_type or old_proc_text != new_proc_text:
                proc_changes += 1
            if bool(row.get("any_more_changes", False)):
                flagged += 1
        preview_rows.append(
            {
                "Spec No.": str(draft.get("spec_number", "") or "—"),
                "Chemical": str(draft.get("chemical_name", "") or "—"),
                "Rows in Upload": len(rows),
                "Type Changes": type_changes,
                "Procedure Changes": proc_changes,
                "Any More = Yes": flagged,
            }
        )
    return preview_rows


def _render_type_classification_exercise(
    repo: CSCRepository,
    user_name: str,
    allowed_subsets: set[str],
) -> None:
    published = _filter_specs_for_subsets(repo.list_reviewable_specs(), allowed_subsets)
    st.markdown("---")
    st.markdown("#### Type Classification Exercise")
    st.caption(
        "Download one workbook with one sheet per specification. Update only "
        "`Type`, `Any more changes proposed ?`, `Test Procedure Type`, and "
        "`Test Procedure Text`, then upload it to apply updates."
    )

    if not published:
        st.info("No published specs are currently available for this committee.")
        return

    prep_key = f"type_cls_prepare_{user_name}"
    if st.button("Prepare Type Classification Workbook", key=prep_key, type="primary"):
        try:
            workbook_bytes, spec_count, row_count = _build_type_classification_workbook(repo, published)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.session_state["type_cls_wb_bytes"] = workbook_bytes
            st.session_state["type_cls_wb_name"] = f"Type_Classification_Exercise_{stamp}.xlsx"
            st.success(f"Workbook prepared: {spec_count} spec(s), {row_count} parameter row(s).")
        except Exception as exc:
            st.error(f"Workbook generation failed: {exc}")
            logger.exception("Type classification workbook generation failed: %s", exc)

    wb_bytes = st.session_state.get("type_cls_wb_bytes")
    wb_name = st.session_state.get("type_cls_wb_name") or "Type_Classification_Exercise.xlsx"
    if wb_bytes:
        st.download_button(
            label="Download Workbook",
            data=wb_bytes,
            file_name=wb_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"type_cls_download_{user_name}",
        )

    uploaded = st.file_uploader(
        "Upload completed Type Classification workbook",
        type=["xlsx"],
        key=f"type_cls_upload_{user_name}",
    )
    if uploaded is None:
        return

    try:
        allowed_ids = {str(s.get("draft_id", "") or "") for s in published}
        updates, parse_summary = _parse_type_classification_workbook(uploaded.getvalue(), allowed_ids)
    except Exception as exc:
        st.error(f"Workbook parsing failed: {exc}")
        logger.exception("Type classification workbook parse failed: %s", exc)
        return

    st.caption(
        f"Parsed sheets: {parse_summary['sheets_seen']} · Rows loaded: {parse_summary['rows_loaded']} · "
        f"Rows skipped: {parse_summary['rows_skipped']}"
    )
    if parse_summary["errors"]:
        for err in parse_summary["errors"]:
            st.warning(err)

    preview_rows = _summarize_type_classification_preview(repo, updates)
    if preview_rows:
        st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

    can_apply = bool(updates) and int(parse_summary.get("rows_loaded", 0) or 0) > 0
    if st.button(
        "Apply Uploaded Type Classification",
        key=f"type_cls_apply_{user_name}",
        disabled=not can_apply,
        type="primary",
    ):
        result = repo.apply_type_classification_exercise(updates, user_name)
        st.session_state["csc_draft_list_cache"] = None
        updated_specs = result.get("updated_specs", []) or []
        flagged_specs = result.get("flagged_specs", []) or []
        errors = result.get("errors", []) or []

        st.success(
            f"Processed {int(result.get('processed_specs', 0) or 0)} spec(s). "
            f"Updated {len(updated_specs)} spec(s)."
        )

        if updated_specs:
            st.markdown("**Version updates applied**")
            st.dataframe(
                pd.DataFrame(updated_specs).rename(
                    columns={
                        "spec_label": "Specification",
                        "changed_count": "Rows Updated",
                        "new_version": "New Version",
                    }
                ),
                use_container_width=True,
                hide_index=True,
            )

        if flagged_specs:
            flagged_rows = []
            for row in flagged_specs:
                flagged_rows.append(
                    {
                        "Specification": row.get("spec_label", ""),
                        "Flagged Parameters": int(row.get("flagged_count", 0) or 0),
                        "Parameters": ", ".join(row.get("flagged_parameters", [])[:20]),
                    }
                )
            st.warning("Specs with additional proposed changes (Any more changes proposed ? = Yes):")
            st.dataframe(pd.DataFrame(flagged_rows), use_container_width=True, hide_index=True)

        if errors:
            st.error("Some updates were skipped:")
            for err in errors:
                st.caption(f"- {err}")


def _render_user_landing(
    repo: CSCRepository,
    user_name: str,
    allowed_subsets: set[str],
) -> None:
    """Landing page for committee users when no draft is open."""
    _render_main_header(ROLE_USER)
    st.markdown("")
    st.info(
        "Select a published spec from the sidebar list to open a revision "
        "and directly propose modifications.",
        icon="📋",
    )
    st.caption(
        f"Active user: {_committee_label(user_name)}  ·  Allowed subsets: "
        + ", ".join(sorted(allowed_subsets))
    )
    if HCC_LOGO_PATH.exists():
        c1, c2, c3 = st.columns([2, 1, 2])
        with c2:
            st.image(str(HCC_LOGO_PATH), width=150)

    # Show published admin specs prominently
    try:
        published = _filter_specs_for_subsets(repo.list_reviewable_specs(), allowed_subsets)
        own_drafts = _filter_specs_for_subsets(repo.list_user_revision_drafts(user_name), allowed_subsets)

        if published:
            st.markdown("---")
            st.markdown("#### 📋 Specs Ready for Committee Review")
            st.caption("Only specs assigned to this committee are shown.")
            df_pub = pd.DataFrame(published)[[
                "spec_number", "chemical_name", "spec_version", "updated_at"
            ]].rename(columns={
                "spec_number":   "Spec No.",
                "chemical_name": "Chemical Name",
                "spec_version":  "Version",
                "updated_at":    "Last Updated",
            })
            df_pub["Version"] = df_pub["Version"].apply(lambda v: f"v{v}")
            df_pub["Last Updated"] = df_pub["Last Updated"].apply(fmt_ist)
            st.dataframe(df_pub, use_container_width=True, hide_index=True)
        else:
            st.caption("No published specs currently mapped to this committee.")

        if own_drafts:
            st.markdown("---")
            st.markdown("#### My Revision Requests")
            df_own = pd.DataFrame(own_drafts)[[
                "spec_number", "chemical_name", "status", "updated_at", "submitted_for_approval_at"
            ]].rename(columns={
                "spec_number":   "Spec No.",
                "chemical_name": "Chemical",
                "status":        "Status",
                "updated_at":    "Last Updated",
                "submitted_for_approval_at": "Submitted",
            })
            df_own["Last Updated"] = df_own["Last Updated"].apply(fmt_ist)
            df_own["Submitted"] = df_own["Submitted"].apply(fmt_ist)
            st.dataframe(df_own, use_container_width=True, hide_index=True)

    except Exception as exc:
        logger.exception("Landing page spec listing failed: %s", exc)
        st.warning("Could not load spec listing. Check logs for details.")

    _render_type_classification_exercise(repo, user_name, allowed_subsets)


# ---------------------------------------------------------------------------
# Main application
# ---------------------------------------------------------------------------

def main() -> None:
    st.set_page_config(
        page_title="CSC Drafting Workspace",
        page_icon="📋",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    st.markdown(_CSS, unsafe_allow_html=True)
    _ensure_session_state()
    logger.info("CSC app start")

    repo = _repo()

    # Health check (once per session is fine)
    if not repo.health_check():
        st.error(
            "Database connection failed. "
            "Check that the `data/` directory is writable and try again."
        )
        st.stop()

    _render_sidebar(repo)

    current_role = st.session_state.get("csc_role", ROLE_USER)
    user_name    = st.session_state.get("csc_user_name") or "Unknown"

    # ── Admin mode: full admin workspace in main panel ────────────────────
    if current_role == ROLE_ADMIN:
        _render_main_header(ROLE_ADMIN)
        render_admin_workspace(repo, user_name)
        return

    # ── User / committee mode ─────────────────────────────────────────────
    draft_id = st.session_state.get("csc_active_draft_id")
    allowed_subsets = _committee_subsets(user_name)

    if draft_id:
        active_draft = repo.load_draft(draft_id)
        if not _is_draft_allowed_for_subsets(active_draft, allowed_subsets):
            st.session_state["csc_active_draft_id"] = None
            st.session_state["csc_active_section"] = SECTION_BACKGROUND
            draft_id = None

    if not draft_id:
        _render_user_landing(repo, user_name, allowed_subsets)
    else:
        render_active_section(repo, draft_id, user_name)


if __name__ == "__main__":
    main()
