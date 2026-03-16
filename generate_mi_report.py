#!/usr/bin/env python3
"""
ONGC Material Intelligence — Professional Excel Report Generator

Reads Consumption_data.xlsx and Procurement_data.xlsx from the workspace,
processes them through the plant-rollup pipeline, and writes a 5-sheet
workbook styled after the xylene_stitched_v2 reference.

Sheets
------
  0_Plant Rollup Map      Plant groupings and hierarchy (from JSON config)
  1_Quality Summary       Per-material data quality metrics
  2_Monthly Pivot         Material × Month consumption (FY-grouped columns)
  3_Long Format           One row per material per month (forecast input)
  4_Source Audit          Color-coded FY provenance per material × month
  5_Plant Pivot           Plant × Month consumption by reporting plant
  6_Charts                Visual analytics (6 charts inc. Pareto combos)

Usage
-----
  python generate_mi_report.py [material_filter]

  material_filter  optional — partial case-insensitive name to limit output
                   e.g.  "xylene"  or  "grease"

Output
------
  mi_report_<timestamp>.xlsx   written to the same directory as this script
"""

from __future__ import annotations

import json
import os
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.legend import Legend as ChartLegend
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.drawing.text import RichTextProperties
from openpyxl.chart.text import RichText

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).resolve().parent
CONS_FILE   = BASE_DIR / "Consumption_data.xlsx"
PROC_FILE   = BASE_DIR / "Procurement_data.xlsx"
PLANT_JSON  = BASE_DIR / "inventory_plant_grouping.json"
MASTER_FILE = BASE_DIR / "Master_data.xlsx"

# ── Style palette (matches xylene_stitched_v2 reference exactly) ───────────────
_FN      = "Arial"
_C_NAVY  = "1A4A6B"   # main title / column-header bg
_C_BLUE  = "185FA5"   # plant / material ID cell bg
_C_TEAL  = "2E8B6E"   # FY-section & pivot sub-header bg
_C_SUB   = "E8F4FD"   # subtitle row bg
_C_ALT   = "F7FAFA"   # alternating data row bg
_C_GRN_D = "1D7A55"   # confidence: High (dark)
_C_GRN_L = "C7E6D5"   # confidence: High (light bg)
_C_AMB_D = "BA7517"   # confidence: Medium (dark)
_C_AMB_L = "FFF3CD"   # confidence: Medium (light bg)
_C_RED_D = "A32D2D"   # confidence: Low (dark)
_C_RED_L = "FFCDD2"   # confidence: Low (light bg)

# Source-audit colour per financial year
_FY_COLOURS: dict[str, tuple[str, str]] = {
    "FY23-24": ("C8E6C9", "222222"),  # green
    "FY24-25": ("BBDEFB", "222222"),  # blue
    "FY25-26": ("E1BEE7", "222222"),  # purple
    "zero":    ("F0F0F0", "888888"),  # grey  (no data that month)
}

_TH = Side(style="thin", color="BFBFBF")
_BD = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)
_CA = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CL = Alignment(horizontal="left",   vertical="center", wrap_text=False)
_CR = Alignment(horizontal="right",  vertical="center")

# Movement-type sign map (MB51)
_MVT_SIGNS = {
    "201": 1.0, "221": 1.0, "261": 1.0,
    "202": -1.0, "222": -1.0, "262": -1.0,
}

# Known plant descriptions
_PLANT_META: dict[str, tuple[str, str]] = {
    "11A1": ("Offshore — Shore Base",
             "UT Offshore — Mumbai High North  (ICP / SHP / BHS / MHN / NQO + Sagar Samrat)"),
    "12A1": ("Offshore — Shore Base",
             "UT Offshore — Heera / Ratna & R-Series"),
    "13A1": ("Offshore — Shore Base",
             "UT Offshore — Mumbai High South  (SBS / NQS / MHS / HDP)"),
    "22A1": ("Onshore", "Onshore — Ahmedabad / Ankleshwar Asset"),
    "25A1": ("Onshore", "Onshore — Rajasthan Asset"),
    "50A1": ("Onshore", "Onshore — Tripura Asset"),
    "51A1": ("Onshore", "Onshore — Assam Asset"),
}

# Short display name for each reporting plant (used beside plant code in workbook)
_PLANT_SHORT_NAME: dict[str, str] = {
    "11A1": "Mumbai High North",
    "12A1": "Heera / Ratna",
    "13A1": "Mumbai High South",
    "22A1": "Ahmedabad / Ankleshwar",
    "25A1": "Rajasthan",
    "50A1": "Tripura",
    "51A1": "Assam",
}


def _plant_label(code: str) -> str:
    """Return 'CODE — Short Name' if a short name is known, else just the code."""
    name = _PLANT_SHORT_NAME.get(code)
    return f"{code} — {name}" if name else code


# ═══════════════════════════════════════════════════════════════════════════════
# Data loading & processing
# ═══════════════════════════════════════════════════════════════════════════════

def _load_plant_config() -> dict:
    if PLANT_JSON.exists():
        return json.loads(PLANT_JSON.read_text())
    return {
        "prefix_groups": {"11": "11A1", "12": "12A1", "13": "13A1"},
        "explicit_groups": {"22W1": "22A1", "25W1": "25A1", "50W1": "50A1", "51W1": "51A1"},
        "group_members": {
            "11A1": ["11A1", "11F1", "11F2", "11F3", "11F4", "11F5", "11F9", "11FA"],
            "12A1": ["12A1", "12F1", "12F5"],
            "13A1": ["13A1", "13F1", "13F2", "13F3", "13F9", "13FB", "13FC", "13U9"],
            "22A1": ["22A1", "22W1"], "25A1": ["25A1", "25W1"],
            "50A1": ["50A1", "50W1"], "51A1": ["51A1", "51W1"],
        },
    }


def _load_master_data() -> pd.DataFrame:
    """Load material master (density + physical state) from MySQL DB.

    Primary source: the live *material_master* table used by the app.
    Fallback: Master_data.xlsx in the same directory as this script.

    Returns a DataFrame with columns:
        material_code   – SAP material number as string
        short_text      – material description
        physical_state  – "Liquid" | "Solid" | None
        density         – float (kg/L = g/mL = MT/kL) or None
    """
    # ── Try MySQL DB ──────────────────────────────────────────────────────
    try:
        import pymysql as _pym  # type: ignore
        import json as _json

        # Read credentials from .env (same directory)
        env_path = BASE_DIR.parent / ".env"
        if not env_path.exists():
            env_path = BASE_DIR / ".env"
        env_vars: dict[str, str] = {}
        if env_path.exists():
            for line in env_path.read_text().splitlines():
                line = line.strip()
                if "=" in line and not line.startswith("#"):
                    k, _, v = line.partition("=")
                    env_vars[k.strip()] = v.strip().strip('"').strip("'")

        conn = _pym.connect(
            host=env_vars.get("DB_HOST", "localhost"),
            port=int(env_vars.get("DB_PORT", "3306")),
            user=env_vars.get("DB_USER", "root"),
            password=env_vars.get("DB_PASSWORD", ""),
            database=env_vars.get("DB_NAME", "ongc_digital_workspace"),
            charset="utf8mb4",
        )
        cur = conn.cursor()
        cur.execute(
            "SELECT material, short_text, physical_state, extra_data "
            "FROM material_master"
        )
        rows_db = cur.fetchall()
        conn.close()

        records = []
        for mat, stext, phys, extra_raw in rows_db:
            ed: dict = {}
            if extra_raw:
                try:
                    ed = _json.loads(extra_raw) if isinstance(extra_raw, str) else extra_raw
                except Exception:
                    pass
            density_raw = ed.get("Density") or ed.get("density")
            try:
                density = float(density_raw) if density_raw not in (None, "", "nan", "NaN") else None
            except Exception:
                density = None
            records.append({
                "material_code": str(mat).strip(),
                "short_text":    str(stext or "").strip(),
                "physical_state": str(phys or "").strip() or None,
                "density":       density,
            })
        df_md = pd.DataFrame(records)
        print(f"  Loaded {len(df_md)} master records from MySQL DB")
        return df_md

    except Exception as exc:
        print(f"  DB not available ({type(exc).__name__}): {exc}")
        print("  Falling back to Master_data.xlsx …")

    # ── Fallback: xlsx ────────────────────────────────────────────────────
    if not MASTER_FILE.exists():
        print("  Warning: Master_data.xlsx not found — no density conversion.")
        return pd.DataFrame(columns=["material_code", "short_text",
                                     "physical_state", "density"])
    raw = pd.read_excel(MASTER_FILE, dtype=str)
    raw.columns = [c.strip() for c in raw.columns]
    records = []
    for _, row in raw.iterrows():
        mat = str(row.get("Material", "") or "").strip()
        if not mat or mat.lower() == "nan":
            continue
        stext = str(row.get("Short Text", "") or "").strip()
        phys  = str(row.get("Physical State", "") or "").strip() or None
        den_raw = row.get("Density")
        try:
            density = float(den_raw) if den_raw not in (None, "nan", "") else None
        except Exception:
            density = None
        records.append({"material_code": mat, "short_text": stext,
                         "physical_state": phys, "density": density})
    df_md = pd.DataFrame(records)
    print(f"  Loaded {len(df_md)} master records from Master_data.xlsx (fallback)")
    return df_md


# Conversion factors to MT
_LITERS_PER_BBL = 158.987   # 1 US petroleum barrel = 158.987 L


def _qty_to_mt(qty: float, uom: str, physical_state: str | None,
               density: float | None) -> float:
    """Convert a consumption quantity (any UoM) to metric tons (MT).

    density is in kg/L  (= g/mL = MT/kL).
    For mass UoMs (MT, KG, G) density is not needed.
    For volume UoMs (L, KL, BBL) density is used (defaults to 1.0 if missing).
    """
    uom = str(uom).strip().upper()
    d = density if (density is not None and density > 0) else 1.0
    if uom in ("MT", "TO", "T"):
        return qty
    elif uom == "KG":
        return qty / 1_000.0
    elif uom == "G":
        return qty / 1_000_000.0
    elif uom == "L":
        return qty * d / 1_000.0          # L × kg/L ÷ 1000 = MT
    elif uom == "KL":
        return qty * d                     # kL × kg/L = kg/1000... wait: kL×(kg/L)=MT? no
        # kL × kg/L = kL × 1000 L/kL × kg/L = 1000 kg = 1 MT for d=1
        # Actually: 1 kL = 1000 L; qty_MT = qty_kL × 1000L × d kg/L ÷ 1000 = qty_kL × d
        # So qty * d is correct ✓
    elif uom in ("BBL", "BL"):
        return qty * _LITERS_PER_BBL * d / 1_000.0
    elif uom in ("M3", "CBM"):
        return qty * d                     # same ratio as KL
    else:
        return qty   # unknown UoM — return as-is


def _resolve_reporting_plant(plant: str, cfg: dict) -> str:
    explicit: dict[str, str] = cfg.get("explicit_groups", {})
    if plant in explicit:
        return explicit[plant]
    prefix_groups: dict[str, str] = cfg.get("prefix_groups", {})
    for prefix, parent in prefix_groups.items():
        if plant.startswith(prefix):
            return parent
    return plant


def _financial_year(year: int, month: int) -> str:
    if month >= 4:
        return f"FY{str(year)[2:]}-{str(year + 1)[2:]}"
    return f"FY{str(year - 1)[2:]}-{str(year)[2:]}"


def _month_label(year: int, month: int) -> str:
    return datetime(year, month, 1).strftime("%b-%Y")


def _load_consumption(material_filter: str | None, cfg: dict,
                      master_df: pd.DataFrame | None = None) -> pd.DataFrame:
    print("  Loading consumption data …")
    df = pd.read_excel(CONS_FILE)
    df.columns = [c.strip().lower().replace(" ", "_").replace(".", "") for c in df.columns]

    rename = {
        "material":                "material_code",
        "material_description":    "material_desc",
        "base_unit_of_measure":    "base_uom",
        "movement_type":           "movement_type",
        "posting_date":            "posting_date",
        "qty_in_unit_of_entry":    "entry_qty",
        "unit_of_entry":           "uom",
        "amtinloccur":             "usage_value",
        "plant":                   "plant",
        "storage_location":        "storage_location",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    # Drop rows without a posting date or movement type
    df = df.dropna(subset=["posting_date"])
    df["posting_date"] = pd.to_datetime(df["posting_date"], errors="coerce")
    df = df.dropna(subset=["posting_date"])

    df["year"]  = df["posting_date"].dt.year.astype(int)
    df["month"] = df["posting_date"].dt.month.astype(int)

    # Sign-adjust quantity by movement type
    mvt = df["movement_type"].astype(str).str.strip()
    signs = mvt.map(_MVT_SIGNS).fillna(0.0)
    df["_sign"]     = signs.values
    df["usage_qty"] = pd.to_numeric(df["entry_qty"], errors="coerce").fillna(0.0) * df["_sign"]
    df["usage_value"] = pd.to_numeric(df.get("usage_value", pd.Series(0.0, index=df.index)), errors="coerce").fillna(0.0) * df["_sign"]

    # Keep only known consumption movement types (sign != 0)
    df = df[df["_sign"] != 0.0].drop(columns=["_sign"])

    # Normalise text
    df["material_desc"] = df["material_desc"].astype(str).str.strip().str.upper()
    df["plant"] = df["plant"].astype(str).str.strip().str.upper()
    df["uom"]   = df.get("uom", pd.Series("", index=df.index)).astype(str).str.strip()

    df["reporting_plant"] = df["plant"].apply(lambda p: _resolve_reporting_plant(p, cfg))
    df["financial_year"]  = df.apply(lambda r: _financial_year(r["year"], r["month"]), axis=1)
    df["month_label"]     = df.apply(lambda r: _month_label(r["year"], r["month"]), axis=1)

    if material_filter:
        df = df[df["material_desc"].str.contains(material_filter.upper(), na=False)]

    # ── Unit conversion → MT using master-data density ────────────────────
    if master_df is not None and not master_df.empty:
        # Merge density / physical_state by material_code
        mref = master_df[["material_code", "physical_state", "density"]].copy()
        mref["material_code"] = mref["material_code"].astype(str).str.strip()
        df["material_code"]   = df["material_code"].astype(str).str.strip()
        df = df.merge(mref, on="material_code", how="left")
    else:
        df["physical_state"] = None
        df["density"]        = None

    # Preserve original UoM for reference (before normalising to MT)
    df["original_uom"] = df["uom"].astype(str).str.strip()

    # Apply conversion row-by-row only for non-MT rows (fast path skips MT rows)
    needs_conv = df["original_uom"].str.upper() != "MT"
    if needs_conv.any():
        def _conv(row):
            return _qty_to_mt(row["usage_qty"], row["original_uom"],
                              row.get("physical_state"), row.get("density"))
        df.loc[needs_conv, "usage_qty"] = df[needs_conv].apply(_conv, axis=1)
    # Normalise UoM to MT after conversion
    df["uom"] = "MT"

    return df


def _load_procurement(material_filter: str | None, cfg: dict) -> pd.DataFrame:
    print("  Loading procurement data …")
    df = pd.read_excel(PROC_FILE)
    df.columns = [c.strip().lower().replace(" ", "_").replace("/", "_") for c in df.columns]

    rename = {
        "material":                  "material_code",
        "short_text":                "material_desc",
        "plant":                     "plant",
        "purchasing_document":       "po_number",
        "document_date":             "doc_date",
        "supplier_supplying_plant":  "vendor",
        "order_quantity":            "order_qty",
        "order_unit":                "order_unit",
        "net_price":                 "net_price",
        "price_unit":                "price_unit",
        "effective_value":           "effective_value",
        "currency":                  "currency",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    df["doc_date"] = pd.to_datetime(df.get("doc_date", pd.Series(dtype="object")), errors="coerce")
    df["material_desc"] = df.get("material_desc", pd.Series("", index=df.index)).astype(str).str.strip().str.upper()
    df["plant"]         = df.get("plant", pd.Series("", index=df.index)).astype(str).str.strip().str.upper()
    df["effective_value"] = pd.to_numeric(df.get("effective_value", pd.Series(0.0, index=df.index)), errors="coerce").fillna(0.0)
    df["order_qty"]     = pd.to_numeric(df.get("order_qty", pd.Series(0.0, index=df.index)), errors="coerce").fillna(0.0)
    df["net_price"]     = pd.to_numeric(df.get("net_price",  pd.Series(0.0, index=df.index)), errors="coerce").fillna(0.0)

    df["reporting_plant"] = df["plant"].apply(lambda p: _resolve_reporting_plant(p, cfg))

    if material_filter:
        df = df[df["material_desc"].str.contains(material_filter.upper(), na=False)]

    return df


def _build_monthly_series(cons: pd.DataFrame) -> pd.DataFrame:
    """Aggregate to one row per (material_desc, year, month)."""
    if cons.empty:
        return pd.DataFrame(columns=["material_desc", "material_code", "uom",
                                      "year", "month", "financial_year",
                                      "month_label", "usage_qty", "usage_value"])
    agg = (
        cons.groupby(["material_desc", "year", "month"], as_index=False)
        .agg(
            material_code=("material_code", "first"),
            uom=("uom", "first"),
            financial_year=("financial_year", "first"),
            month_label=("month_label", "first"),
            usage_qty=("usage_qty", "sum"),
            usage_value=("usage_value", "sum"),
        )
    )
    return agg.sort_values(["material_desc", "year", "month"])


def _build_proc_monthly(proc: pd.DataFrame) -> pd.DataFrame:
    """Monthly aggregated procurement value per material."""
    if proc.empty or "doc_date" not in proc.columns:
        return pd.DataFrame(columns=["material_desc", "year", "month", "effective_value"])
    df = proc.dropna(subset=["doc_date"]).copy()
    df["year"]  = df["doc_date"].dt.year.astype(int)
    df["month"] = df["doc_date"].dt.month.astype(int)
    agg = (
        df.groupby(["material_desc", "year", "month"], as_index=False)
        .agg(effective_value=("effective_value", "sum"))
    )
    return agg.sort_values(["year", "month"])


def _quality_metrics(monthly: pd.DataFrame) -> pd.DataFrame:
    """Compute per-material quality metrics."""
    rows = []
    for mat, grp in monthly.groupby("material_desc"):
        qtys = grp["usage_qty"].values
        nonzero = qtys[qtys != 0]           # months with any net movement
        active = int((qtys != 0).sum())
        total_months = len(qtys)
        zero_months  = total_months - active
        pct_active   = round(100.0 * active / total_months, 1) if total_months else 0.0
        total_qty    = float(qtys.sum())
        abs_nz       = np.abs(nonzero)
        avg_monthly  = float(abs_nz.mean()) if len(abs_nz) else 0.0
        peak_month   = float(abs_nz.max())  if len(abs_nz) else 0.0
        std_dev      = float(abs_nz.std())  if len(abs_nz) > 1 else 0.0
        cv           = round(std_dev / avg_monthly, 2) if avg_monthly > 0 else 0.0

        if cv >= 1.0:
            volatility = "High"
        elif cv >= 0.5:
            volatility = "Medium"
        else:
            volatility = "Low"

        sorted_grp = grp[grp["usage_qty"] != 0].sort_values(["year", "month"])
        first_active = sorted_grp["month_label"].iloc[0]  if not sorted_grp.empty else "—"
        last_active  = sorted_grp["month_label"].iloc[-1] if not sorted_grp.empty else "—"

        uom          = grp["uom"].iloc[0]
        mat_code     = grp["material_code"].iloc[0]

        rows.append(dict(
            material_desc=mat,
            material_code=mat_code,
            uom=uom,
            active_months=active,
            zero_months=zero_months,
            total_months=total_months,
            pct_active=pct_active,
            total_qty=round(total_qty, 0),
            avg_monthly=round(avg_monthly, 0),
            peak_month=round(peak_month, 0),
            std_dev=round(std_dev, 0),
            cv=cv,
            volatility=volatility,
            first_active=first_active,
            last_active=last_active,
        ))

    qdf = pd.DataFrame(rows).sort_values("avg_monthly", ascending=False)
    return qdf.reset_index(drop=True)


def _all_months(monthly: pd.DataFrame) -> list[tuple[int, int, str, str]]:
    """Return sorted list of (year, month, label, fy) across the whole dataset."""
    seen = monthly[["year", "month", "month_label", "financial_year"]].drop_duplicates()
    seen = seen.sort_values(["year", "month"])
    return list(seen.itertuples(index=False, name=None))


# ═══════════════════════════════════════════════════════════════════════════════
# Workbook helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _fl(hex6: str) -> PatternFill:
    # openpyxl requires full 8-char ARGB; prepend FF (fully opaque)
    c = "FF" + hex6[:6]
    return PatternFill("solid", fgColor=c, start_color=c, end_color=c)


def _fnt(bold=False, sz=10, col="000000") -> Font:
    return Font(name=_FN, bold=bold, size=sz, color=col)


def _title(ws, row: int, text: str, ncols: int, h: int = 28):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font  = Font(name=_FN, bold=True, size=13, color="FFFFFF")
    c.fill  = _fl(_C_NAVY)
    c.alignment = _CA
    ws.row_dimensions[row].height = h


def _subtitle(ws, row: int, text: str, ncols: int, h: int = 18, bg: str = _C_SUB):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font  = Font(name=_FN, italic=True, size=9, color="444444")
    c.fill  = _fl(bg)
    c.alignment = _CL
    ws.row_dimensions[row].height = h


def _hdr(ws, row: int, values: list, bg=_C_NAVY, fc="FFFFFF", h=26, col_start=1):
    ws.row_dimensions[row].height = h
    for i, v in enumerate(values, col_start):
        c = ws.cell(row, i, v)
        c.font      = Font(name=_FN, bold=True, size=10, color=fc)
        c.fill      = _fl(bg)
        c.alignment = _CA
        c.border    = _BD


def _plant_id_cell(ws, row: int, col: int, value, h: int = 20):
    c = ws.cell(row, col, value)
    c.font      = Font(name=_FN, bold=True, size=10, color="FFFFFF")
    c.fill      = _fl(_C_BLUE)
    c.alignment = _CL
    c.border    = _BD
    ws.row_dimensions[row].height = h


def _data_cell(ws, row: int, col: int, value, alt: bool = False,
               bold=False, align=_CL, num_fmt: str | None = None):
    c = ws.cell(row, col, value)
    c.font      = Font(name=_FN, bold=bold, size=10)
    c.fill      = _fl(_C_ALT if alt else "FFFFFF")
    c.alignment = align
    c.border    = _BD
    if num_fmt:
        c.number_format = num_fmt
    return c


def _rotate_axis_labels(axis, degrees: int = -90) -> None:
    """Rotate axis tick-mark labels to *degrees* (negative = clockwise)."""
    rot = degrees * 60_000           # openpyxl uses 60 000ths of a degree
    body_pr = RichTextProperties()
    body_pr.rot = rot
    axis.txPr = RichText(bodyPr=body_pr)


def _data_labels_vertical(chart, pos: str = "outEnd", show_val: bool = True,
                           degrees: int = -90) -> None:
    """Attach value data-labels rotated *degrees* to *chart*."""
    dLbls = DataLabelList()
    dLbls.showVal        = show_val
    dLbls.showSerName    = False
    dLbls.showCatName    = False
    dLbls.showLegendKey  = False
    dLbls.position       = pos
    # Rotate the label text
    rot      = degrees * 60_000
    body_pr  = RichTextProperties()
    body_pr.rot = rot
    dLbls.txPr = RichText(bodyPr=body_pr)
    chart.dLbls = dLbls


def _chart_legend(pos: str = "b") -> ChartLegend:
    """Return a visible, non-overlapping legend at the given position.

    pos: 'b' bottom  |  'r' right  |  't' top  |  'l' left
    """
    lg = ChartLegend()
    lg.position = pos
    lg.overlay = False
    return lg


def _conf_badge(ws, row: int, col: int, conf: str):
    """Colour-coded confidence badge cell."""
    low = conf.lower()
    if "high" in low:
        bg, fc = _C_GRN_L, _C_GRN_D
    elif "medium" in low:
        bg, fc = _C_AMB_L, _C_AMB_D
    else:
        bg, fc = _C_RED_L, _C_RED_D
    c = ws.cell(row, col, conf)
    c.font      = Font(name=_FN, bold=True, size=9, color=fc)
    c.fill      = _fl(bg)
    c.alignment = _CA
    c.border    = _BD


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet builders
# ═══════════════════════════════════════════════════════════════════════════════

def _sheet0_plant_map(wb: Workbook, cfg: dict, cons: pd.DataFrame):
    """0_Plant Rollup Map"""
    ws = wb.create_sheet("0_Plant Rollup Map")
    members: dict[str, list[str]] = cfg.get("group_members", {})
    n_cols = 6

    _title(ws, 1,
           "ONGC Material Intelligence — Plant to Forecast Unit Mapping  |  "
           "Shore Warehouse → Offshore Platforms / Onshore Assets",
           n_cols)
    _subtitle(ws, 2,
              "Consumption at sub-plants is aggregated into the parent forecast unit.  "
              "Forecasting and procurement decisions are made at the forecast unit level.",
              n_cols)
    _hdr(ws, 3, ["Forecast Unit", "Type", "Description",
                  "Rolled-up Sub-plants", "# Sub-plants",
                  "Active in Data"])

    # Determine which parent plants appear in the data
    active_parents = set()
    if not cons.empty:
        active_parents = set(cons["reporting_plant"].unique())

    for ri, (parent, children) in enumerate(members.items()):
        rn = ri + 4
        alt = ri % 2 == 1
        ptype, pdesc = _PLANT_META.get(parent, ("—", "—"))
        sub_str = ", ".join(c for c in children if c != parent)
        in_data = "Yes" if parent in active_parents else "No"

        _plant_id_cell(ws, rn, 1, _plant_label(parent), h=36)
        ws.cell(rn, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        _data_cell(ws, rn, 2, ptype,     alt)
        _data_cell(ws, rn, 3, pdesc,     alt)
        _data_cell(ws, rn, 4, sub_str,   alt, align=_CL)
        _data_cell(ws, rn, 5, len(children), alt, align=_CA)
        _data_cell(ws, rn, 6, in_data,   alt, align=_CA)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "B4"


def _sheet1_quality(wb: Workbook, qdf: pd.DataFrame):
    """1_Quality Summary"""
    ws = wb.create_sheet("1_Quality Summary")
    n_cols = 15

    _title(ws, 1,
           "Material Intelligence — Data Quality Summary  |  All Financial Years",
           n_cols)
    _subtitle(ws, 2,
              "Movements included: 201 + 221 + 261 (net of reversals 202 / 222 / 262)  "
              "|  All sub-plant consumption aggregated to forecast unit level",
              n_cols)
    _hdr(ws, 3, [
        "Material", "Material Code", "UoM",
        "Active Months", "Zero Months", "Total Months", "% Active",
        "Total Consumption", "Avg Monthly", "Peak Month",
        "Std Dev", "CV", "Volatility",
        "First Active", "Last Active",
    ])

    for ri, row in qdf.iterrows():
        rn = ri + 4
        alt = ri % 2 == 1
        _plant_id_cell(ws, rn, 1, row["material_desc"])
        _data_cell(ws, rn,  2, row["material_code"],  alt, align=_CA)
        _data_cell(ws, rn,  3, row["uom"],            alt, align=_CA)
        _data_cell(ws, rn,  4, int(row["active_months"]),  alt, align=_CA)
        _data_cell(ws, rn,  5, int(row["zero_months"]),    alt, align=_CA)
        _data_cell(ws, rn,  6, int(row["total_months"]),   alt, align=_CA)
        _data_cell(ws, rn,  7, row["pct_active"],          alt, align=_CA, num_fmt='0.0"%"')
        _data_cell(ws, rn,  8, row["total_qty"],           alt, align=_CR, num_fmt="#,##0")
        _data_cell(ws, rn,  9, row["avg_monthly"],         alt, align=_CR, num_fmt="#,##0")
        _data_cell(ws, rn, 10, row["peak_month"],          alt, align=_CR, num_fmt="#,##0")
        _data_cell(ws, rn, 11, row["std_dev"],             alt, align=_CR, num_fmt="#,##0")
        _data_cell(ws, rn, 12, row["cv"],                  alt, align=_CA, num_fmt="0.00")

        # Volatility badge
        vol_bg = {"High": _C_RED_L, "Medium": _C_AMB_L, "Low": _C_GRN_L}.get(row["volatility"], "FFFFFF")
        vol_fc = {"High": _C_RED_D, "Medium": _C_AMB_D, "Low": _C_GRN_D}.get(row["volatility"], "000000")
        vc = ws.cell(rn, 13, row["volatility"])
        vc.font      = Font(name=_FN, bold=True, size=9, color=vol_fc)
        vc.fill      = _fl(vol_bg)
        vc.alignment = _CA
        vc.border    = _BD

        _data_cell(ws, rn, 14, row["first_active"], alt, align=_CA)
        _data_cell(ws, rn, 15, row["last_active"],  alt, align=_CA)
        ws.row_dimensions[rn].height = 20

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    for col in "DEFGHIJKLMNO":
        ws.column_dimensions[col].width = 14
    ws.freeze_panes = "D4"


def _sheet2_pivot(wb: Workbook, monthly: pd.DataFrame, qdf: pd.DataFrame):
    """2_Monthly Pivot"""
    ws = wb.create_sheet("2_Monthly Pivot")

    all_months = _all_months(monthly)  # [(year, month, label, fy), ...]
    # Group months by FY
    fy_groups: dict[str, list[tuple]] = {}
    for m in all_months:
        fy = m[3]
        fy_groups.setdefault(fy, []).append(m)

    fys = sorted(fy_groups.keys())

    # FY total columns: one per FY (desc), then 3-year total
    fy_desc  = sorted(fys, reverse=True)   # most recent first
    n_fy     = len(fy_desc)
    n_months = len(all_months)
    n_cols   = 1 + n_fy + 1 + n_months  # plant + FY totals + 3yr total + monthly

    _title(ws, 1,
           f"Material Intelligence — Monthly Consumption Pivot  |  "
           f"Materials × {all_months[0][2] if all_months else '?'} to "
           f"{all_months[-1][2] if all_months else '?'}  |  Rollup applied",
           n_cols)
    _subtitle(ws, 2,
              "Sub-plant consumption rolled up to forecast-unit level.  "
              "Zeros included for months with no recorded movement.",
              n_cols)

    # Row 3: FY group super-headers (merged across months in each FY)
    ws.row_dimensions[3].height = 22
    fy_summary_col_start = 2
    monthly_col_start    = 2 + n_fy + 1  # after fy totals + 3yr total

    # A3 — "Forecast Unit" label (mimics xylene reference)
    c3a = ws.cell(3, 1, "Forecast Unit")
    c3a.font      = Font(name=_FN, bold=True, size=10, color="FFFFFF")
    c3a.fill      = _fl(_C_NAVY)
    c3a.alignment = _CA
    c3a.border    = _BD

    # Merge "FY Summary Totals" across the FY total columns
    ws.merge_cells(start_row=3, start_column=2,
                   end_row=3, end_column=1 + n_fy + 1)
    c = ws.cell(3, 2, "FY Summary Totals")
    c.font      = Font(name=_FN, bold=True, size=10, color="FFFFFF")
    c.fill      = _fl(_C_NAVY)
    c.alignment = _CA
    c.border    = _BD

    # Merge each FY label across its monthly columns
    col_cursor = monthly_col_start
    for fy in fys:
        months_in_fy = fy_groups[fy]
        start = col_cursor
        end   = col_cursor + len(months_in_fy) - 1
        label = f"{fy}  ({months_in_fy[0][2]} – {months_in_fy[-1][2]})"
        ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
        c = ws.cell(3, start, label)
        c.font      = Font(name=_FN, bold=True, size=10, color="FFFFFF")
        c.fill      = _fl(_C_NAVY)
        c.alignment = _CA
        c.border    = _BD
        col_cursor += len(months_in_fy)

    # Row 4: column sub-headers
    hdr4 = ["Material"]
    for fy in fy_desc:
        short = fy.replace("FY", "FY ").replace("-", "-")
        hdr4.append(f"{short} Total")
    hdr4.append("All-Year Total")
    for m in all_months:
        hdr4.append(m[2])   # "Apr-2023" etc.
    _hdr(ws, 4, hdr4, bg=_C_TEAL, h=26)

    # Build a lookup: material → {(year,month): qty}
    mat_order = qdf["material_desc"].tolist()
    qty_map: dict[str, dict[tuple, float]] = {}
    for _, r in monthly.iterrows():
        mat = r["material_desc"]
        qty_map.setdefault(mat, {})[(int(r["year"]), int(r["month"]))] = float(r["usage_qty"])

    for ri, mat in enumerate(mat_order):
        rn  = ri + 5
        alt = ri % 2 == 1
        _plant_id_cell(ws, rn, 1, mat)

        q_by_fy: dict[str, float] = {}
        for (yr, mo), qty in qty_map.get(mat, {}).items():
            fy = _financial_year(yr, mo)
            q_by_fy[fy] = q_by_fy.get(fy, 0.0) + qty

        # FY total columns (desc order)
        for ci, fy in enumerate(fy_desc, 2):
            val = round(q_by_fy.get(fy, 0.0), 0)
            c = ws.cell(rn, ci, val if val else "—")
            c.font      = Font(name=_FN, bold=True, size=10)
            c.fill      = _fl("EAF6F1" if not alt else "D4EDE5")
            c.alignment = _CR
            c.border    = _BD
            c.number_format = "#,##0"

        # All-year total
        all_total = round(sum(q_by_fy.values()), 0)
        c = ws.cell(rn, 2 + n_fy, all_total if all_total else "—")
        c.font      = Font(name=_FN, bold=True, size=10)
        c.fill      = _fl("D5E8D4" if not alt else "C5D8C4")
        c.alignment = _CR
        c.border    = _BD
        c.number_format = "#,##0"

        # Monthly cells
        for mi, (yr, mo, lbl, fy) in enumerate(all_months):
            col = monthly_col_start + mi
            qty = qty_map.get(mat, {}).get((yr, mo), 0.0)
            val = round(qty, 0) if qty else 0
            c   = ws.cell(rn, col, val if val else "—")
            c.font      = Font(name=_FN, size=9)
            c.fill      = _fl("EAF6F1" if not alt else "F7FAFA")
            c.alignment = _CR
            c.border    = _BD
            if val:
                c.number_format = "#,##0"

        ws.row_dimensions[rn].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 42
    for ci in range(2, 2 + n_fy + 1 + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    for ci in range(monthly_col_start, monthly_col_start + n_months):
        ws.column_dimensions[get_column_letter(ci)].width = 11

    ws.freeze_panes = f"B5"


def _sheet3_long(wb: Workbook, monthly: pd.DataFrame):
    """3_Long Format (Forecast Input)"""
    ws = wb.create_sheet("3_Long Format")
    n_cols = 8

    _title(ws, 1,
           "Material Intelligence — Stitched Long-Format Dataset  |  "
           "Forecast pipeline input  |  Rollup applied",
           n_cols)
    _subtitle(ws, 2,
              "One row per material per month.  Zero-filled months included "
              "(SAP suppresses zero-consumption months).  "
              "Source column traces data provenance.",
              n_cols)
    _hdr(ws, 3, ["Material", "Material Code", "Date", "Month",
                  "Financial Year", "Consumption", "Value (INR)", "UoM"])

    # Build complete month grid for every material
    all_months = _all_months(monthly)
    mat_list   = monthly["material_desc"].unique()

    # Lookup for quick access
    lookup: dict[tuple, dict] = {}
    for _, r in monthly.iterrows():
        lookup[(r["material_desc"], int(r["year"]), int(r["month"]))] = r.to_dict()

    rn = 4
    for mat in sorted(mat_list):
        mat_info = monthly[monthly["material_desc"] == mat].iloc[0]
        mat_code = str(mat_info["material_code"])
        uom      = str(mat_info["uom"])

        for yr, mo, lbl, fy in all_months:
            key  = (mat, yr, mo)
            data = lookup.get(key)
            qty  = round(float(data["usage_qty"]), 2)   if data else 0.0
            val  = round(float(data["usage_value"]), 2) if data else 0.0
            alt  = (rn % 2 == 0)
            bg   = "E8F5E9" if not alt else "F7FAFA"

            # Material column with blue style
            c = ws.cell(rn, 1, mat)
            c.font      = Font(name=_FN, bold=True, size=9, color="FFFFFF")
            c.fill      = _fl(_C_BLUE)
            c.alignment = _CL
            c.border    = _BD

            for col_idx, (value, fmt) in enumerate([
                (mat_code,                                None),
                (datetime(yr, mo, 1).strftime("%Y-%m-%d"), None),
                (lbl,                                     None),
                (fy,                                      None),
                (qty,                                     "#,##0.##"),
                (val,                                     "#,##0.##"),
                (uom,                                     None),
            ], 2):
                dc = ws.cell(rn, col_idx, value)
                dc.font      = Font(name=_FN, size=9)
                dc.fill      = _fl(bg)
                dc.alignment = _CA if col_idx in (3, 4, 5, 8) else _CL
                dc.border    = _BD
                if fmt:
                    dc.number_format = fmt

            ws.row_dimensions[rn].height = 15
            rn += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 8
    ws.freeze_panes = "C4"


def _sheet4_audit(wb: Workbook, cons_raw: pd.DataFrame, monthly: pd.DataFrame):
    """4_Source Audit — colour-coded FY provenance per material × month"""
    ws = wb.create_sheet("4_Source Audit")

    all_months = _all_months(monthly)
    n_cols     = 1 + len(all_months)
    mat_list   = sorted(monthly["material_desc"].unique())

    # Build (material, year, month) → financial_year map from raw data
    fy_map: dict[tuple, str] = {}
    if not cons_raw.empty:
        for _, r in cons_raw.iterrows():
            key = (str(r["material_desc"]), int(r["year"]), int(r["month"]))
            fy  = str(r["financial_year"])
            # Normalise fy to one of our known keys
            norm = fy.replace("FY", "FY").strip()
            fy_map[key] = norm

    _title(ws, 1,
           "Source Audit — Data provenance per material × month  "
           "(use to validate stitching logic)",
           n_cols)

    # Build legend subtitle
    legend_parts = []
    for fy, (bg, _) in sorted(_FY_COLOURS.items()):
        if fy != "zero":
            legend_parts.append(f"{fy}=#{bg}")
    _subtitle(ws, 2,
              "  |  ".join([
                  "Green=FY23-24",
                  "Blue=FY24-25",
                  "Purple=FY25-26",
                  "Grey=zero-fill (no movement that month)",
              ]), n_cols, bg="F0F0F0")

    # Row 3 header
    _hdr(ws, 3, ["Material"] + [m[2] for m in all_months], bg=_C_TEAL, h=26)

    for ri, mat in enumerate(mat_list):
        rn = ri + 4
        _plant_id_cell(ws, rn, 1, mat)
        ws.row_dimensions[rn].height = 16

        for ci, (yr, mo, lbl, fy_col) in enumerate(all_months, 2):
            key   = (mat, yr, mo)
            fy    = fy_map.get(key)
            if fy and fy in _FY_COLOURS:
                bg, fc = _FY_COLOURS[fy]
                label  = fy
            else:
                bg, fc = _FY_COLOURS["zero"]
                label  = "—"

            c = ws.cell(rn, ci, label)
            c.font      = Font(name=_FN, size=8, color=fc)
            c.fill      = _fl(bg)
            c.alignment = _CA
            c.border    = _BD

    ws.column_dimensions["A"].width = 42
    for ci in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 11
    ws.freeze_panes = "B4"


# ═══════════════════════════════════════════════════════════════════════════════
# Plant Pivot sheet (5)
# ═══════════════════════════════════════════════════════════════════════════════

def _sheet5_plant_pivot(wb: Workbook, cons_raw: pd.DataFrame, cfg: dict, monthly: pd.DataFrame):
    """5_Plant Pivot — Plant × Month consumption data table."""
    ws = wb.create_sheet("5_Plant Pivot")

    if cons_raw.empty:
        ws.cell(1, 1, "No consumption data available")
        return

    all_months_list = _all_months(monthly)
    reporting_plants = sorted(cons_raw["reporting_plant"].unique().tolist())
    fys = sorted(monthly["financial_year"].unique())

    # Aggregate |consumption| per plant × month
    plant_monthly = (
        cons_raw.groupby(["reporting_plant", "year", "month"])["usage_qty"]
        .apply(lambda x: float(x.abs().sum()))
        .reset_index()
    )
    # Attach month_label and financial_year
    month_meta = monthly[["year", "month", "month_label", "financial_year"]].drop_duplicates()
    plant_monthly = plant_monthly.merge(month_meta, on=["year", "month"], how="left")

    n_fys = len(fys)
    n_months = len(all_months_list)
    # Columns: Plant | FY totals… | Overall Total | Monthly cols…
    monthly_col_start = 1 + n_fys + 1 + 1   # 1-based: plant=1, then n_fys FY cols, then overall
    n_cols = 1 + n_fys + 1 + n_months

    _title(ws, 1,
           f"Material Intelligence — Plant-wise Consumption Pivot  |  "
           f"{all_months_list[0][2] if all_months_list else '?'} to "
           f"{all_months_list[-1][2] if all_months_list else '?'}",
           n_cols)
    _subtitle(ws, 2,
              "Sub-plant consumption rolled up to reporting plant (forecast unit) level.  "
              "Values are absolute (reversals excluded from totals).  "
              "Plant names shown beside codes.",
              n_cols)

    # Row 3: group super-headers
    ws.row_dimensions[3].height = 22

    # A3 label
    c3a = ws.cell(3, 1, "Reporting Plant")
    c3a.font      = Font(name=_FN, bold=True, size=10, color="FFFFFF")
    c3a.fill      = _fl(_C_NAVY)
    c3a.alignment = _CA
    c3a.border    = _BD

    # FY Summary block header (cols 2 … 2+n_fys-1 + overall col)
    fy_hdr_end = 1 + n_fys + 1   # last col of summary block
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=fy_hdr_end)
    c_fy_hdr = ws.cell(3, 2, "FY Summary Totals")
    c_fy_hdr.font      = Font(name=_FN, bold=True, size=10, color="FFFFFF")
    c_fy_hdr.fill      = _fl(_C_TEAL)
    c_fy_hdr.alignment = _CA
    c_fy_hdr.border    = _BD

    # Monthly FY group headers
    fy_groups_map: dict[str, list] = {}
    for m in all_months_list:
        fy_groups_map.setdefault(m[3], []).append(m)

    running_mc = monthly_col_start
    for fy in fys:
        months_in_fy = fy_groups_map.get(fy, [])
        if not months_in_fy:
            continue
        end_mc = running_mc + len(months_in_fy) - 1
        ws.merge_cells(start_row=3, start_column=running_mc, end_row=3, end_column=end_mc)
        c_fy = ws.cell(3, running_mc, fy)
        c_fy.font      = Font(name=_FN, bold=True, size=10, color="FFFFFF")
        c_fy.fill      = _fl(_C_TEAL)
        c_fy.alignment = _CA
        c_fy.border    = _BD
        running_mc = end_mc + 1

    # Row 4: column headers
    hdr_values = (
        ["Reporting Plant"]
        + list(fys)
        + ["Overall Total"]
        + [lbl for _, _, lbl, _ in all_months_list]
    )
    _hdr(ws, 4, hdr_values)

    # Data rows
    for ri, plant in enumerate(reporting_plants):
        rn = ri + 5
        alt = ri % 2 == 1

        # Plant label cell (code + short name with wrap)
        c_plant = ws.cell(rn, 1, _plant_label(plant))
        c_plant.font      = Font(name=_FN, bold=True, size=10, color="FFFFFF")
        c_plant.fill      = _fl(_C_BLUE)
        c_plant.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c_plant.border    = _BD
        ws.row_dimensions[rn].height = 36

        # FY totals
        overall = 0.0
        for ci, fy in enumerate(fys, 2):
            fy_qty = plant_monthly[
                (plant_monthly["reporting_plant"] == plant) &
                (plant_monthly["financial_year"] == fy)
            ]["usage_qty"].sum()
            _data_cell(ws, rn, ci, round(fy_qty, 0), alt, align=_CR, num_fmt="#,##0")
            overall += fy_qty

        # Overall total
        overall_col = 2 + n_fys
        _data_cell(ws, rn, overall_col, round(overall, 0), alt, bold=True,
                   align=_CR, num_fmt="#,##0")

        # Monthly breakdown
        for mci, (yr, mo, lbl, fy) in enumerate(all_months_list, monthly_col_start):
            qty = plant_monthly[
                (plant_monthly["reporting_plant"] == plant) &
                (plant_monthly["year"] == yr) &
                (plant_monthly["month"] == mo)
            ]["usage_qty"].sum()
            val = round(qty, 0) if qty > 0 else ""
            _data_cell(ws, rn, mci, val, alt, align=_CR,
                       num_fmt="#,##0" if val != "" else None)

    # Column widths
    ws.column_dimensions["A"].width = 28
    for ci in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.freeze_panes = "B5"


# ═══════════════════════════════════════════════════════════════════════════════
# Charts sheet (6)
# ═══════════════════════════════════════════════════════════════════════════════

def _sheet5_charts(wb: Workbook, qdf: pd.DataFrame, monthly: pd.DataFrame,
                   proc_monthly: pd.DataFrame, cons_raw: pd.DataFrame, cfg: dict,
                   master_df: pd.DataFrame | None = None):
    """6_Charts — Visual analytics (6 charts, Pareto combos + plant chart)."""

    # ── Build hidden data sheet ────────────────────────────────────────────
    wd = wb.create_sheet("_chart_data")
    wd.sheet_state = "hidden"

    all_months_list = _all_months(monthly)
    fys = sorted(monthly["financial_year"].unique())

    # ── Section 1: Top 10 by Consumption (rows 1–11) ──────────────────────
    wd.cell(1, 1, "Material"); wd.cell(1, 2, "Avg Monthly Consumption")
    for i, row in enumerate(qdf.head(10).itertuples(), 2):
        wd.cell(i, 1, row.material_desc[:30])
        wd.cell(i, 2, float(row.avg_monthly))

    # ── Section 2: Top 10 by Procurement Value (rows 14–24) ───────────────
    proc_by_mat = (
        proc_monthly.groupby("material_desc")["effective_value"].sum()
        .sort_values(ascending=False).head(10)
    )
    wd.cell(14, 1, "Material"); wd.cell(14, 2, "Total Procurement Value (INR)")
    for i, (mat, val) in enumerate(proc_by_mat.items(), 15):
        wd.cell(i, 1, mat[:30])
        wd.cell(i, 2, float(val))

    # ── Section 3: Pareto — Monthly Consumption (sorted high → low) ───────
    # Monthly totals across all materials, sorted descending
    month_label_map = {
        (int(r["year"]), int(r["month"])): r["month_label"]
        for _, r in monthly.drop_duplicates(["year", "month"]).iterrows()
    }
    monthly_totals = (
        monthly.groupby(["year", "month"])["usage_qty"]
        .apply(lambda x: float(x.abs().sum()))
        .reset_index()
    )
    monthly_totals["month_label"] = monthly_totals.apply(
        lambda r: month_label_map.get((int(r["year"]), int(r["month"])), ""), axis=1
    )
    pareto_cons = monthly_totals.sort_values("usage_qty", ascending=False).reset_index(drop=True)
    grand_cons = pareto_cons["usage_qty"].sum()
    pareto_cons["cum_pct"] = (
        pareto_cons["usage_qty"].cumsum() / grand_cons * 100
    ).round(1) if grand_cons > 0 else 0.0

    row28 = 28
    wd.cell(row28, 1, "Month")
    wd.cell(row28, 2, "Total Consumption")
    wd.cell(row28, 3, "Cumulative %")
    for ri, row in enumerate(pareto_cons.itertuples(), row28 + 1):
        wd.cell(ri, 1, row.month_label)
        wd.cell(ri, 2, float(row.usage_qty))
        wd.cell(ri, 3, float(row.cum_pct))
    cons_data_end_row = row28 + len(pareto_cons)

    # ── Section 4: Pareto — Monthly Procurement Value (sorted high → low) ─
    if not proc_monthly.empty and "year" in proc_monthly.columns:
        monthly_proc_totals = (
            proc_monthly.groupby(["year", "month"])["effective_value"]
            .sum()
            .reset_index()
        )
        monthly_proc_totals["month_label"] = monthly_proc_totals.apply(
            lambda r: month_label_map.get((int(r["year"]), int(r["month"])), ""), axis=1
        )
        pareto_proc = monthly_proc_totals.sort_values("effective_value", ascending=False).reset_index(drop=True)
        grand_proc = pareto_proc["effective_value"].sum()
        pareto_proc["cum_pct"] = (
            pareto_proc["effective_value"].cumsum() / grand_proc * 100
        ).round(1) if grand_proc > 0 else 0.0
    else:
        pareto_proc = pd.DataFrame({"month_label": [], "effective_value": [], "cum_pct": []})

    val_start = cons_data_end_row + 5
    wd.cell(val_start, 1, "Month")
    wd.cell(val_start, 2, "Total Proc Value (INR)")
    wd.cell(val_start, 3, "Cumulative %")
    for ri, row in enumerate(pareto_proc.itertuples(), val_start + 1):
        wd.cell(ri, 1, row.month_label)
        wd.cell(ri, 2, float(row.effective_value))
        wd.cell(ri, 3, float(row.cum_pct))
    val_end_row = val_start + max(len(pareto_proc), 1)

    # ── Section 5: FY Year-on-Year — top 10 ───────────────────────────────
    yoy_start = val_end_row + 5
    wd.cell(yoy_start, 1, "Material")
    for ci, fy in enumerate(fys, 2):
        wd.cell(yoy_start, ci, fy)
    for ri, mat in enumerate(qdf.head(10)["material_desc"].tolist(), yoy_start + 1):
        wd.cell(ri, 1, mat[:30])
        for ci, fy in enumerate(fys, 2):
            fy_total = monthly[
                (monthly["material_desc"] == mat) &
                (monthly["financial_year"] == fy)
            ]["usage_qty"].apply(abs).sum()
            wd.cell(ri, ci, float(fy_total))
    yoy_end_row = yoy_start + 10

    # ── Section 6: Plant × FY pivot (for plant chart) ─────────────────────
    reporting_plants: list[str] = []
    if not cons_raw.empty:
        reporting_plants = sorted(cons_raw["reporting_plant"].unique().tolist())

    # Build plant × month consumption (absolute values)
    if not cons_raw.empty:
        plant_monthly_agg = (
            cons_raw.groupby(["reporting_plant", "financial_year"])["usage_qty"]
            .apply(lambda x: float(x.abs().sum()))
            .reset_index()
        )
    else:
        plant_monthly_agg = pd.DataFrame(
            columns=["reporting_plant", "financial_year", "usage_qty"])

    plant_pivot_start = yoy_end_row + 5
    wd.cell(plant_pivot_start, 1, "Reporting Plant")
    for ci, fy in enumerate(fys, 2):
        wd.cell(plant_pivot_start, ci, fy)
    for ri, plant in enumerate(reporting_plants, plant_pivot_start + 1):
        wd.cell(ri, 1, _plant_label(plant))
        for ci, fy in enumerate(fys, 2):
            total = plant_monthly_agg[
                (plant_monthly_agg["reporting_plant"] == plant) &
                (plant_monthly_agg["financial_year"] == fy)
            ]["usage_qty"].sum()
            wd.cell(ri, ci, float(total))
    plant_pivot_end = plant_pivot_start + max(len(reporting_plants), 1)

    # ── Build charts sheet ─────────────────────────────────────────────────
    ws = wb.create_sheet("6_Charts")
    n_cols_title = 32
    _title(ws, 1, "Material Intelligence — Visual Analytics", n_cols_title)
    _subtitle(ws, 2,
              "Charts 3 & 4: bars sorted highest → lowest + cumulative % on secondary axis.  "
              "Consumption quantities converted to MT using master-data density values.", n_cols_title)

    # ── Chart dimensions & layout constants ───────────────────────────────
    _W  = 26    # standard chart width  (cm)
    _H  = 16    # standard chart height (cm)
    _WP = 28    # plant chart width     (cm)
    _HP = 18    # plant chart height    (cm)
    _R1, _R2, _R3 = 4, 38, 72
    _COL_L = "A"
    _COL_R = "O"

    # ── Chart 1: Top 10 by Consumption (MT) — horizontal bar ──────────────
    c1 = BarChart()
    c1.type = "bar"; c1.grouping = "clustered"
    c1.title = "Top 10 Materials — Avg Monthly Consumption (MT)"
    c1.y_axis.title = "Material"
    c1.x_axis.title = "Avg Monthly Consumption (MT)"
    c1.style = 10; c1.width = _W; c1.height = _H
    c1.legend = _chart_legend("b")
    _rotate_axis_labels(c1.x_axis)
    c1.add_data(Reference(wd, min_col=2, max_col=2, min_row=1, max_row=11),
                titles_from_data=True)
    c1.set_categories(Reference(wd, min_col=1, min_row=2, max_row=11))
    c1.series[0].graphicalProperties.solidFill = "1A4A6B"
    ws.add_chart(c1, f"{_COL_L}{_R1}")

    # ── Chart 2: Top 10 by Procurement Value — horizontal bar ─────────────
    c2 = BarChart()
    c2.type = "bar"; c2.grouping = "clustered"
    c2.title = "Top 10 Materials — Total Procurement Value (INR)"
    c2.y_axis.title = "Material"
    c2.x_axis.title = "Procurement Value (INR)"
    c2.style = 10; c2.width = _W; c2.height = _H
    c2.legend = _chart_legend("b")
    _rotate_axis_labels(c2.x_axis)
    c2.add_data(Reference(wd, min_col=2, max_col=2, min_row=14, max_row=24),
                titles_from_data=True)
    c2.set_categories(Reference(wd, min_col=1, min_row=15, max_row=24))
    c2.series[0].graphicalProperties.solidFill = "2E8B6E"
    ws.add_chart(c2, f"{_COL_R}{_R1}")

    # ── Chart 3: Monthly Consumption — sorted high→low (column + cumulative %)
    c3_bar = BarChart()
    c3_bar.type = "col"; c3_bar.grouping = "clustered"
    c3_bar.title = "Monthly Total Consumption (MT) — Sorted Highest to Lowest"
    c3_bar.y_axis.title = "Total Consumption (MT)"
    c3_bar.x_axis.title = "Month"
    c3_bar.style = 10; c3_bar.width = _W; c3_bar.height = _H
    _rotate_axis_labels(c3_bar.x_axis)
    c3_bar.add_data(Reference(wd, min_col=2, max_col=2,
                               min_row=row28, max_row=cons_data_end_row),
                    titles_from_data=True)
    c3_bar.set_categories(Reference(wd, min_col=1,
                                     min_row=row28 + 1, max_row=cons_data_end_row))
    c3_bar.series[0].graphicalProperties.solidFill = "1A4A6B"
    _data_labels_vertical(c3_bar, pos="outEnd")

    c3_line = LineChart()
    c3_line.add_data(Reference(wd, min_col=3, max_col=3,
                                min_row=row28, max_row=cons_data_end_row),
                     titles_from_data=True)
    c3_line.series[0].graphicalProperties.line.solidFill = "D97706"
    c3_line.series[0].graphicalProperties.line.width = 28000
    c3_line.y_axis.axId = 200
    c3_line.y_axis.crosses = "max"
    c3_line.y_axis.title = "Cumulative %"
    c3_bar += c3_line
    c3_bar.legend = _chart_legend("b")
    ws.add_chart(c3_bar, f"{_COL_L}{_R2}")

    # ── Chart 4: Monthly Procurement Value — sorted high→low (column + cumulative %)
    c4_bar = BarChart()
    c4_bar.type = "col"; c4_bar.grouping = "clustered"
    c4_bar.title = "Monthly Total Procurement Value (INR) — Sorted Highest to Lowest"
    c4_bar.y_axis.title = "Total Proc Value (INR)"
    c4_bar.x_axis.title = "Month"
    c4_bar.style = 10; c4_bar.width = _W; c4_bar.height = _H
    _rotate_axis_labels(c4_bar.x_axis)
    c4_bar.add_data(Reference(wd, min_col=2, max_col=2,
                               min_row=val_start, max_row=val_end_row),
                    titles_from_data=True)
    c4_bar.set_categories(Reference(wd, min_col=1,
                                     min_row=val_start + 1, max_row=val_end_row))
    c4_bar.series[0].graphicalProperties.solidFill = "2E8B6E"
    _data_labels_vertical(c4_bar, pos="outEnd")

    c4_line = LineChart()
    c4_line.add_data(Reference(wd, min_col=3, max_col=3,
                                min_row=val_start, max_row=val_end_row),
                     titles_from_data=True)
    c4_line.series[0].graphicalProperties.line.solidFill = "D97706"
    c4_line.series[0].graphicalProperties.line.width = 28000
    c4_line.y_axis.axId = 200
    c4_line.y_axis.crosses = "max"
    c4_line.y_axis.title = "Cumulative %"
    c4_bar += c4_line
    c4_bar.legend = _chart_legend("b")
    ws.add_chart(c4_bar, f"{_COL_R}{_R2}")

    # ── Chart 5: FY Year-on-Year — horizontal grouped bar ─────────────────
    n_fys = len(fys)
    c5 = BarChart()
    c5.type = "bar"; c5.grouping = "clustered"
    c5.title = "FY Year-on-Year Consumption (MT) — Top 10 Materials"
    c5.y_axis.title = "Material"
    c5.x_axis.title = "Consumption (MT)"
    c5.style = 10; c5.width = _W; c5.height = _H
    c5.legend = _chart_legend("b")
    _rotate_axis_labels(c5.x_axis)
    fy_colours5 = ["1A4A6B", "2E8B6E", "D97706"]
    for ci in range(2, 2 + n_fys):
        c5.add_data(Reference(wd, min_col=ci, max_col=ci,
                               min_row=yoy_start, max_row=yoy_end_row),
                    titles_from_data=True)
        c5.series[-1].graphicalProperties.solidFill = fy_colours5[min(ci - 2, 2)]
    c5.set_categories(Reference(wd, min_col=1, min_row=yoy_start + 1, max_row=yoy_end_row))
    ws.add_chart(c5, f"{_COL_L}{_R3}")

    # ── Chart 6: Plant-wise Consumption by FY — horizontal grouped bar ─────
    if reporting_plants:
        c6 = BarChart()
        c6.type = "bar"; c6.grouping = "clustered"
        c6.title = "Plant-wise Total Consumption (MT) by Financial Year"
        c6.y_axis.title = "Reporting Plant"
        c6.x_axis.title = "Total Consumption (MT)"
        c6.style = 10; c6.width = _WP; c6.height = _HP
        c6.legend = _chart_legend("b")
        _rotate_axis_labels(c6.x_axis)
        fy_colours6 = ["1A4A6B", "2E8B6E", "D97706"]
        for ci in range(2, 2 + n_fys):
            c6.add_data(Reference(wd, min_col=ci, max_col=ci,
                                   min_row=plant_pivot_start, max_row=plant_pivot_end),
                        titles_from_data=True)
            c6.series[-1].graphicalProperties.solidFill = fy_colours6[min(ci - 2, 2)]
        c6.set_categories(Reference(wd, min_col=1,
                                     min_row=plant_pivot_start + 1, max_row=plant_pivot_end))
        ws.add_chart(c6, f"{_COL_R}{_R3}")

    # ── Density reference table (below all charts) ─────────────────────────
    ref_start_row = _R3 + 38    # well below the last chart row
    _title(ws, ref_start_row,
           "Density Reference — Values Used for Consumption → MT Conversion",
           n_cols_title, h=22)
    _subtitle(ws, ref_start_row + 1,
              "Liquids: qty_MT = qty_L × density ÷ 1000  |  qty_MT = qty_KL × density  "
              "|  Solids in KG: qty_MT = qty_KG ÷ 1000  |  Already MT: no conversion",
              n_cols_title)
    _hdr(ws, ref_start_row + 2,
         ["Material Description", "Physical State", "Density (kg/L)", "Source UoM(s)", "Conversion Applied"],
         h=20)

    # Build density reference from master_df (or fallback)
    if master_df is not None and not master_df.empty:
        md_ref = master_df.copy()
    else:
        md_ref = pd.DataFrame(columns=["material_code", "short_text",
                                        "physical_state", "density"])

    # Enrich with ORIGINAL UoMs from cons_raw (before they were normalised to MT)
    uom_col = "original_uom" if "original_uom" in cons_raw.columns else "uom"
    if not cons_raw.empty:
        mat_uoms = (
            cons_raw.groupby("material_desc")[uom_col]
            .apply(lambda s: ", ".join(sorted(s.unique())))
            .reset_index(name="src_uom")
        )
    else:
        mat_uoms = pd.DataFrame(columns=["material_desc", "src_uom"])

    # Cross-reference material_desc with master short_text (normalised match)
    if not md_ref.empty and "short_text" in md_ref.columns:
        # Build display by joining on normalised text
        md_ref["_norm"] = md_ref["short_text"].astype(str).str.strip().str.upper()
        mat_uoms["_norm"] = mat_uoms["material_desc"].astype(str).str.strip().str.upper()
        merged = mat_uoms.merge(md_ref[["_norm", "physical_state", "density"]], on="_norm", how="left")
    else:
        merged = mat_uoms.copy()
        merged["physical_state"] = None
        merged["density"] = None

    # Write the density reference rows
    for ri_d, row_d in enumerate(merged.sort_values("material_desc").itertuples()):
        rn_d = ref_start_row + 3 + ri_d
        alt_d = ri_d % 2 == 1
        mat_d  = str(row_d.material_desc)
        state  = str(getattr(row_d, "physical_state", "") or "—")
        dens   = getattr(row_d, "density", None)
        dens_s = f"{float(dens):.4f}" if dens not in (None, "", "nan") else "1.0 (default)"
        src_u  = str(getattr(row_d, "src_uom", "MT"))
        # Determine conversion description
        uoms_set = {u.strip().upper() for u in src_u.split(",")}
        if uoms_set == {"MT"}:
            conv = "No conversion (already MT)"
        elif "KG" in uoms_set and uoms_set.issubset({"KG", "MT"}):
            conv = "KG ÷ 1000 = MT"
        elif state == "Liquid" and uoms_set & {"L", "KL", "BBL"}:
            conv = f"L × {dens_s} ÷ 1000 = MT"
        else:
            conv = "KG ÷ 1000 = MT" if "KG" in uoms_set else "As-is"

        _plant_id_cell(ws, rn_d, 1, mat_d[:50], h=18)
        _data_cell(ws, rn_d, 2, state,  alt_d, align=_CA)
        _data_cell(ws, rn_d, 3, dens_s, alt_d, align=_CA)
        _data_cell(ws, rn_d, 4, src_u,  alt_d, align=_CA)
        _data_cell(ws, rn_d, 5, conv,   alt_d, align=_CL)

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 36


# ═══════════════════════════════════════════════════════════════════════════════
# Public entry-point for the web application
# ═══════════════════════════════════════════════════════════════════════════════

def generate_workbook_from_dataframes(
    cons_raw: pd.DataFrame,
    proc_raw: pd.DataFrame,
    master_df: pd.DataFrame | None = None,
    plant_filter: str | None = None,
) -> "BytesIO":
    """Build the professional MI workbook from pre-loaded DataFrames.

    This is the entry-point used by the Flask web application so that the app's
    already-loaded in-memory data is reused rather than re-reading the xlsx files.

    Parameters
    ----------
    cons_raw   : Consumption DataFrame as returned by the app's
                 ``_DataStore._filtered_consumption()``.  Must contain at least:
                 reporting_plant, plant, material_code, material_desc,
                 storage_location, year (int), month (int), financial_year,
                 usage_qty, uom.
    proc_raw   : Procurement DataFrame as returned by the app's
                 ``_DataStore._filtered_procurement()``.  Must contain at least:
                 reporting_plant, plant, material_code, material_desc,
                 doc_date, effective_value.
    master_df  : Optional DataFrame with columns material_code, short_text,
                 physical_state, density — used for unit → MT conversion.
                 If None no conversion is applied.
    plant_filter: Human-readable label used only in the report filename.

    Returns
    -------
    BytesIO   : In-memory Excel workbook stream (seek(0) already called).
    """
    from io import BytesIO as _BytesIO

    cons = cons_raw.copy()
    proc = proc_raw.copy()

    # ── Ensure month_label column (app uses month_raw="MM.YYYY") ──────────
    if "month_label" not in cons.columns:
        cons["month_label"] = cons.apply(
            lambda r: _month_label(int(r["year"]), int(r["month"])), axis=1
        )

    # ── Ensure usage_value column ─────────────────────────────────────────
    if "usage_value" not in cons.columns:
        cons["usage_value"] = 0.0
    else:
        cons["usage_value"] = pd.to_numeric(cons["usage_value"], errors="coerce").fillna(0.0)

    # ── Density / unit → MT conversion (mirrors _load_consumption logic) ──
    if master_df is not None and not master_df.empty:
        mref = master_df[["material_code", "physical_state", "density"]].copy()
        mref["material_code"] = mref["material_code"].astype(str).str.strip()
        cons["material_code"] = cons["material_code"].astype(str).str.strip()
        # Only merge if columns not already present
        if "density" not in cons.columns:
            cons = cons.merge(mref, on="material_code", how="left")
    else:
        if "physical_state" not in cons.columns:
            cons["physical_state"] = None
        if "density" not in cons.columns:
            cons["density"] = None

    # Preserve original UoM for the density reference table
    if "original_uom" not in cons.columns:
        cons["original_uom"] = cons["uom"].astype(str).str.strip()

    needs_conv = cons["original_uom"].str.upper() != "MT"
    if needs_conv.any():
        def _conv(row):
            return _qty_to_mt(
                row["usage_qty"], row["original_uom"],
                row.get("physical_state"), row.get("density"),
            )
        cons.loc[needs_conv, "usage_qty"] = cons[needs_conv].apply(_conv, axis=1)
    cons["uom"] = "MT"

    # ── Load plant grouping config ────────────────────────────────────────
    cfg = _load_plant_config()

    # ── Build derived tables ──────────────────────────────────────────────
    monthly      = _build_monthly_series(cons)
    proc_monthly = _build_proc_monthly(proc)
    qdf          = _quality_metrics(monthly)

    # ── Build workbook ────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    _sheet0_plant_map(wb, cfg, cons)
    _sheet1_quality(wb, qdf)
    _sheet2_pivot(wb, monthly, qdf)
    _sheet3_long(wb, monthly)
    _sheet4_audit(wb, cons, monthly)
    _sheet5_plant_pivot(wb, cons, cfg, monthly)
    _sheet5_charts(wb, qdf, monthly, proc_monthly, cons, cfg, master_df)

    stream = _BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    material_filter = sys.argv[1] if len(sys.argv) > 1 else None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = BASE_DIR / f"mi_report_{ts}.xlsx"

    print(f"ONGC Material Intelligence — Report Generator")
    print(f"  Filter : {material_filter or 'all materials'}")
    print(f"  Output : {out_path.name}")
    print()

    cfg       = _load_plant_config()
    master_df = _load_master_data()
    cons_raw  = _load_consumption(material_filter, cfg, master_df)
    proc_raw  = _load_procurement(material_filter, cfg)

    print(f"  Consumption rows : {len(cons_raw):,}")
    print(f"  Procurement rows : {len(proc_raw):,}")

    monthly = _build_monthly_series(cons_raw)
    proc_monthly = _build_proc_monthly(proc_raw)
    qdf     = _quality_metrics(monthly)

    print(f"  Materials found  : {len(qdf)}")
    print(f"  Months in data   : {len(monthly[['year','month']].drop_duplicates())}")
    print()
    print("  Building workbook …")

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    _sheet0_plant_map(wb, cfg, cons_raw)
    _sheet1_quality(wb, qdf)
    _sheet2_pivot(wb, monthly, qdf)
    _sheet3_long(wb, monthly)
    _sheet4_audit(wb, cons_raw, monthly)
    _sheet5_plant_pivot(wb, cons_raw, cfg, monthly)
    _sheet5_charts(wb, qdf, monthly, proc_monthly, cons_raw, cfg, master_df)

    wb.save(out_path)
    print(f"  ✓  Saved → {out_path}")
    print()
    print("Done.")


if __name__ == "__main__":
    main()
