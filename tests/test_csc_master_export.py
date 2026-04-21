from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.modules.csc.routes import _build_material_master_export_workbook


def test_build_material_master_export_workbook_has_published_and_submitted_sheets() -> None:
    stream = _build_material_master_export_workbook(
        subset_summary_rows=[
            {
                "subset_code": "DFC",
                "subset": "Drilling Fluid Chemicals",
                "open_drafts": 2,
                "submitted_drafts": 1,
                "saved_not_submitted": 1,
                "not_yet_saved": 1,
            }
        ],
        published_rows=[
            {
                "material": "090001043",
                "short_text": "BARYTES",
                "group": "Weighing Agent",
                "material_type": "Performance",
                "updated_at": "25 Mar 2026",
                "updated_by": "editor_1",
                "extra_data": {
                    "Density": "4.2",
                    "impact_flag_1": "YES",
                    "impact_classification": "Critical",
                    "impact_confidence": "High",
                },
            }
        ],
        submitted_rows=[
            {
                "parent_spec_number": "ONGC/CCA/37/2026",
                "draft_spec_number": "ONGC/CCA/37/2026",
                "chemical_name": "BARYTES",
                "subset": "Cementing Additives",
                "revision_status": "Submitted",
                "submitted_at": "2026-03-25 11:30",
                "submitted_by": "committee_user",
                "material_code": "090001043",
                "short_text": "BARYTES",
                "group": "Weighing Agent",
                "material_type": "Performance",
                "extra__Density": "4.2",
                "impact_flag_1": "PROVISIONAL",
                "impact_classification": "High",
                "impact_confidence": "Medium",
            }
        ],
        committee_user_rows=[
            {
                "parent_spec_number": "ONGC/DFC/07/2026",
                "draft_spec_number": "ONGC/DFC/07/2026",
                "chemical_name": "CALCIUM CARBONATE",
                "subset": "Drilling Fluid Chemicals",
                "revision_status": "Open",
                "updated_at": "2026-03-26 09:15",
                "updated_by": "committee_user_2",
                "material_code": "090001044",
                "short_text": "CALCIUM CARBONATE",
                "group": "Other",
                "material_type": "Performance",
                "extra__Density": "",
                "impact_flag_1": "REVIEW",
                "impact_classification": "Moderate",
                "impact_confidence": "Low",
            }
        ],
    )

    workbook = load_workbook(BytesIO(stream.getvalue()))
    assert workbook.sheetnames == [
        "Subset Summary",
        "Published",
        "Submitted Drafts",
        "Committee User Drafts",
    ]

    summary = workbook["Subset Summary"]
    assert summary["A1"].value == "Subset Code"
    assert summary["A2"].value == "DFC"
    assert summary["C2"].value == 2
    assert summary["F2"].value == 1

    published = workbook["Published"]
    assert published["A1"].value == "Updated At"
    assert published["C1"].value == "Material Code"
    assert published["A2"].value == "25 Mar 2026"
    assert published["C2"].value == "090001043"
    published_headers = [cell.value for cell in published[1]]
    assert "Impact Flag 1 — HSE / Hazard" in published_headers
    assert "Impact Classification" in published_headers
    assert "Impact Confidence" in published_headers
    published_header_index = {header: idx for idx, header in enumerate(published_headers, start=1)}
    assert published.cell(row=2, column=published_header_index["Impact Flag 1 — HSE / Hazard"]).value == "YES"
    assert published.cell(row=2, column=published_header_index["Impact Classification"]).value == "Critical"
    assert published.cell(row=2, column=published_header_index["Impact Confidence"]).value == "High"

    submitted = workbook["Submitted Drafts"]
    assert submitted["A1"].value == "Published Spec Number"
    assert submitted["B2"].value == "ONGC/CCA/37/2026"
    assert submitted["C2"].value == "BARYTES"
    assert submitted["G2"].value == "committee_user"
    assert submitted["H2"].value == "090001043"
    submitted_headers = [cell.value for cell in submitted[1]]
    submitted_header_index = {header: idx for idx, header in enumerate(submitted_headers, start=1)}
    assert submitted.cell(row=2, column=submitted_header_index["Impact Flag 1 — HSE / Hazard"]).value == "PROVISIONAL"
    assert submitted.cell(row=2, column=submitted_header_index["Impact Classification"]).value == "High"
    assert submitted.cell(row=2, column=submitted_header_index["Impact Confidence"]).value == "Medium"

    committee_user = workbook["Committee User Drafts"]
    assert committee_user["A1"].value == "Published Spec Number"
    assert committee_user["A2"].value == "ONGC/DFC/07/2026"
    assert committee_user["E2"].value == "Open"
    assert committee_user["G2"].value == "committee_user_2"
    assert committee_user["H2"].value == "090001044"
    committee_headers = [cell.value for cell in committee_user[1]]
    committee_header_index = {header: idx for idx, header in enumerate(committee_headers, start=1)}
    assert committee_user.cell(row=2, column=committee_header_index["Impact Flag 1 — HSE / Hazard"]).value == "REVIEW"
    assert committee_user.cell(row=2, column=committee_header_index["Impact Classification"]).value == "Moderate"
    assert committee_user.cell(row=2, column=committee_header_index["Impact Confidence"]).value == "Low"
