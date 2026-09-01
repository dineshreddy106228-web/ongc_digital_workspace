from datetime import datetime
from types import SimpleNamespace

from app.core.services import corporate_specifications as cs
from app.core.services import csc_export


def _record(record_id, spec_number, chemical_name, material_code="", version=0):
    return SimpleNamespace(
        id=record_id,
        spec_number=spec_number,
        chemical_name=chemical_name,
        material_code=material_code,
        spec_version=version,
        updated_at=datetime(2026, 3, 25, 15, 40),
    )


def _register_row(row_id, spec_number, chemical_name, material_code="", standard_days=None):
    return SimpleNamespace(
        id=row_id,
        specification_no=spec_number,
        chemical_name=chemical_name,
        material_code=material_code,
        standard_days=standard_days,
        remarks="",
    )


def _stub_catalogue(monkeypatch, register, records, counts):
    monkeypatch.setattr(cs, "_register_rows", lambda: register)
    monkeypatch.setattr(cs, "_specification_records", lambda: records)
    monkeypatch.setattr(cs, "_parameter_counts", lambda: counts)


def test_category_of_reads_the_second_segment_and_rejects_junk():
    assert cs.category_of("ONGC / DFC / 01 / 2026") == "DFC"
    assert cs.category_of("ONGC/PC/11A/2015") == "PC"
    assert cs.category_of("") is None
    assert cs.category_of("ONGC / 01 / 2026") is None


def test_normalize_spec_number_bridges_the_two_spellings_in_use():
    # The register writes ONGC / DFC / 01 / 2026; specification records write ONGC/DFC/01/2026.
    assert cs.normalize_spec_number("ONGC / DFC / 01 / 2026") == cs.normalize_spec_number("ONGC/DFC/01/2026")


def test_sequence_orders_serials_numerically_with_letter_suffixes_after():
    assert cs._sequence_of("ONGC / PC / 06 / 2026") < cs._sequence_of("ONGC / PC / 11 / 2026")
    assert cs._sequence_of("ONGC / PC / 11 / 2026") < cs._sequence_of("ONGC / PC / 11A / 2015")


def test_requirement_prefers_the_structured_field_then_falls_back_to_the_legacy_column():
    legacy = SimpleNamespace(
        required_value_type="text", required_value_text=None, existing_value="1.0 (Maximum)",
        required_value_operator_1=None, required_value_value_1=None,
        required_value_operator_2=None, required_value_value_2=None,
    )
    assert cs.requirement_of(legacy) == "1.0 (Maximum)"

    current = SimpleNamespace(
        required_value_type="text", required_value_text="2.0 (Maximum)", existing_value="1.0 (Maximum)",
        required_value_operator_1=None, required_value_value_1=None,
        required_value_operator_2=None, required_value_value_2=None,
    )
    assert cs.requirement_of(current) == "2.0 (Maximum)"


def test_catalogue_matches_a_register_chemical_to_its_record_by_material_code(monkeypatch):
    _stub_catalogue(
        monkeypatch,
        register=[_register_row(1, "ONGC / DFC / 01 / 2026", "Aluminium Stearate", "090001043")],
        records=[_record(10, "ONGC/DFC/01/2015", "ALUMINUM STEARATE", "090001043")],
        counts={10: 7},
    )
    entry, = cs.catalogue()
    assert entry["ref"] == "r-1"
    assert entry["record_id"] == 10
    assert entry["parameter_count"] == 7
    assert entry["has_parameters"] is True
    assert entry["on_register"] is True


def test_catalogue_falls_back_to_the_specification_number_when_no_code_matches(monkeypatch):
    _stub_catalogue(
        monkeypatch,
        register=[_register_row(1, "ONGC / WM / 03 / 2026", "Well Maker Fluid")],
        records=[_record(10, "ONGC/WM/03/2026", "WELL MAKER FLUID")],
        counts={10: 4},
    )
    entry, = cs.catalogue()
    assert entry["record_id"] == 10


def test_catalogue_keeps_records_that_are_not_on_the_register(monkeypatch):
    _stub_catalogue(
        monkeypatch,
        register=[_register_row(1, "ONGC / DFC / 01 / 2026", "Aluminium Stearate", "090001043")],
        records=[
            _record(10, "ONGC/DFC/01/2026", "ALUMINUM STEARATE", "090001043"),
            _record(11, "ONGC/WS/21/2015", "GELLING AGENTS"),
        ],
        counts={10: 7, 11: 5},
    )
    off_register = [entry for entry in cs.catalogue() if not entry["on_register"]]
    assert [entry["ref"] for entry in off_register] == ["s-11"]


def test_catalogue_maps_one_record_onto_every_chemical_it_covers(monkeypatch):
    _stub_catalogue(
        monkeypatch,
        register=[
            _register_row(1, "ONGC / WS / 02 / 2026", "Guar Gum Grade I"),
            _register_row(2, "ONGC / WS / 02 / 2026", "Guar Gum Grade II"),
        ],
        records=[_record(10, "ONGC/WS/02/2026", "GUAR GUM")],
        counts={10: 6},
    )
    entries = cs.catalogue()
    assert len(entries) == 2
    assert {entry["record_id"] for entry in entries} == {10}


def test_category_tiles_split_specified_from_awaiting_and_follow_the_register_order(monkeypatch):
    _stub_catalogue(
        monkeypatch,
        register=[
            _register_row(1, "ONGC / PC / 06 / 2026", "Demulsifier"),
            _register_row(2, "ONGC / DFC / 01 / 2026", "Aluminium Stearate"),
            _register_row(3, "ONGC / DFC / 02 / 2026", "Barytes"),
            _register_row(4, "", "Unlisted chemical"),
        ],
        records=[_record(10, "ONGC/DFC/01/2026", "ALUMINUM STEARATE")],
        counts={10: 7},
    )
    tiles = cs.category_tiles()
    assert [tile["key"] for tile in tiles] == ["DFC", "PC", cs.UNCATEGORISED_KEY]
    dfc = tiles[0]
    assert (dfc["chemicals"], dfc["specified"], dfc["awaiting"]) == (2, 1, 1)
    assert tiles[-1]["is_unspecified"] is True


def test_catalogue_places_legacy_api_grade_register_rows_in_api_category(monkeypatch):
    _stub_catalogue(
        monkeypatch,
        register=[_register_row(1, "", "Baryte API Grade")],
        records=[],
        counts={},
    )

    entry = cs.catalogue()[0]
    assert entry["category"] == "API"
    assert entry["category_label"] == "API Grade Chemicals"
    assert cs.category_tiles()[0]["key"] == "API"


def test_parameter_payload_keeps_submitted_rows_and_drops_blank_ones():
    form = SimpleNamespace(
        getlist=lambda name: ["0", "1", "2"],
        get=lambda name, default="": {
            "parameter_name_0": "Physical state", "requirement_0": "Free flowing powder",
            "parameter_name_2": "  ", "requirement_2": "",
            "parameter_name_1": "Moisture", "requirement_1": "2.0 (Maximum)",
        }.get(name, default),
    )
    rows = cs.parameter_payload(form)
    assert [row["parameter_name"] for row in rows] == ["Physical state", "Moisture"]


def _snapshot(
    requirement="1.0 (Maximum)",
    background="Legacy specification",
    storage="Store in a cool, dry place",
    hse="NO",
    supply_flagged=False,
):
    """A whole-specification snapshot, in the shape the version check compares."""
    return {
        "spec_number": "ONGC/DFC/01/2026",
        "chemical_name": "ALUMINUM STEARATE",
        "material_code": "090001043",
        "sections": {"background": background},
        "parameters": [
            {
                "parameter_name": "Moisture", "parameter_type": "Vital", "requirement": requirement,
                "unit_of_measure": "", "conditions": "", "test_method": "", "remarks": "",
            }
        ],
        "identity": {"short_text": "ALUMINIUM STEARATE", "group": "", "material_type": "", "centralization": ""},
        "material_properties": {"physical_state": "Powder", "flammable": "No"},
        "storage": {"storage_conditions_general": storage, "storage_conditions_special": ""},
        "impact": {"hse_flag": hse, "supply_flag": "REVIEW"},
        "impact_classification": "LOW IMPACT",
        "issues": {"supply": {"is_present": supply_flagged, "note": ""}},
    }


def test_a_whitespace_only_edit_does_not_count_as_a_revision():
    assert cs._comparable(_snapshot()) == cs._comparable(_snapshot(requirement="  1.0 (Maximum)  "))
    assert cs._comparable(_snapshot()) == cs._comparable(_snapshot(storage="  Store in a cool, dry place "))


def test_changing_a_required_value_or_a_narrative_counts_as_a_revision():
    assert cs._comparable(_snapshot()) != cs._comparable(_snapshot(requirement="2.0 (Maximum)"))
    assert cs._comparable(_snapshot()) != cs._comparable(_snapshot(background="Revised in 2026"))


def test_supporting_data_changes_also_count_as_a_revision():
    # Storage conditions, impact flags and the issue register are part of the specification,
    # so editing any of them has to produce a new version rather than a silent save.
    assert cs._comparable(_snapshot()) != cs._comparable(_snapshot(storage="Refrigerate below 4 degC"))
    assert cs._comparable(_snapshot()) != cs._comparable(_snapshot(hse="YES"))
    assert cs._comparable(_snapshot()) != cs._comparable(_snapshot(supply_flagged=True))


def test_supporting_payload_reads_every_block_off_the_form():
    posted = {
        "master_short_text": "ALUMINIUM STEARATE",
        "master_physical_state": "Powder",
        "master_storage_conditions_special": "  Keep away from oxidisers.  ",
        "impact_hse_flag": "yes",
        "impact_supply_flag": "NONSENSE",
        "issue_supply_present": "1",
        "issue_supply_note": "Single qualified vendor.",
    }
    form = SimpleNamespace(getlist=lambda name: [], get=lambda name, default="": posted.get(name, default))
    payload = cs.supporting_payload(form)

    assert payload["master"]["short_text"] == "ALUMINIUM STEARATE"
    assert payload["master"]["storage_conditions_special"] == "Keep away from oxidisers."
    assert payload["impact"]["hse_flag"] == "YES"
    # An unrecognised answer falls back to REVIEW rather than being stored as-is.
    assert payload["impact"]["supply_flag"] == "REVIEW"
    # Flags nobody answered still come back, so the checklist is always complete.
    assert set(payload["impact"]) == {flag["id"] for flag in cs.IMPACT_CHECKLIST_FLAGS}
    assert payload["issues"]["supply"] == {"is_present": True, "note": "Single qualified vendor."}
    assert payload["issues"]["quality"] == {"is_present": False, "note": ""}


def test_change_status_distinguishes_an_addition_from_a_revision():
    assert cs._change_status("", "Powder", True) == "Added"
    assert cs._change_status("Liquid", "Powder", True) == "Revised"
    assert cs._change_status("Powder", "Powder", True) == "Retained"
    # With no earlier version to compare against, nothing is reported as changed.
    assert cs._change_status("", "Powder", False) == "Retained"


def _stub_master(monkeypatch, committed, staged, child=SimpleNamespace(id=99)):
    monkeypatch.setattr(cs, "revision_child", lambda record: child)
    monkeypatch.setattr(cs, "_staged_master_values", lambda c: staged)
    monkeypatch.setattr(
        "app.core.services.csc_master_data.get_master_form_values", lambda record: dict(committed)
    )


def test_committed_master_data_outranks_the_unpublished_draft(monkeypatch):
    _stub_master(monkeypatch, {"physical_state": "Liquid"}, {"physical_state": "Solid"})
    assert cs.master_values(SimpleNamespace(id=1))["physical_state"] == "Liquid"


def test_the_draft_fills_in_master_fields_the_record_never_captured(monkeypatch):
    # This is the whole point: the committee recorded these on the revision draft
    # and they were never committed to the material master.
    _stub_master(monkeypatch, {"physical_state": ""}, {"physical_state": "Solid", "flammable": "No"})
    values = cs.master_values(SimpleNamespace(id=1))
    assert values["physical_state"] == "Solid"
    assert values["flammable"] == "No"


def test_the_boilerplate_storage_default_yields_to_a_recorded_value(monkeypatch):
    from app.core.services.csc_master_data import STORAGE_CONDITIONS_GENERAL_DEFAULT

    # get_master_form_values injects this text whenever the column is blank, so it
    # must not out-rank storage guidance the committee actually wrote down.
    _stub_master(
        monkeypatch,
        {"storage_conditions_general": STORAGE_CONDITIONS_GENERAL_DEFAULT},
        {"storage_conditions_general": "Store below 25 degC away from oxidisers."},
    )
    assert cs.master_values(SimpleNamespace(id=1))["storage_conditions_general"] == (
        "Store below 25 degC away from oxidisers."
    )
    assert cs._is_placeholder("storage_conditions_general", STORAGE_CONDITIONS_GENERAL_DEFAULT)
    assert not cs._is_placeholder("storage_conditions_general", "Store below 25 degC.")
    assert cs._is_placeholder("physical_state", "   ")


def test_a_specification_with_no_revision_draft_is_unaffected(monkeypatch):
    _stub_master(monkeypatch, {"physical_state": "Powder"}, {}, child=None)
    assert cs.master_values(SimpleNamespace(id=1))["physical_state"] == "Powder"


def test_merged_sections_take_the_narrative_the_draft_holds(monkeypatch):
    record, child = SimpleNamespace(id=1), SimpleNamespace(id=2)
    sections = {
        1: {"background": "Root background", "justification": ""},
        2: {"background": "Draft background", "justification": "Because the 2015 limits are obsolete."},
    }
    monkeypatch.setattr(cs, "revision_child", lambda r: child)
    monkeypatch.setattr(cs, "_sections", lambda draft: dict(sections[draft.id]))
    merged = cs.merged_sections(record)
    # The record's own narrative wins where it has one...
    assert merged["background"] == "Root background"
    # ...and the draft supplies what the record never captured.
    assert merged["justification"] == "Because the 2015 limits are obsolete."


def test_narrative_sections_cover_the_full_dossier_set():
    assert [key for key, _label in cs.NARRATIVE_SECTIONS] == [
        "background", "existing_spec", "proposed_changes", "justification", "recommendation",
    ]


def _master_spec(spec_number, chemical_name, parameters=None):
    return {
        "draft": {
            "spec_number": spec_number,
            "chemical_name": chemical_name,
            "material_code": "",
        },
        "parameters": parameters or [{"parameter_name": "Appearance"}],
    }


def test_master_index_plan_uses_unique_position_bookmarks_and_group_page_breaks():
    specs = [
        _master_spec("ONGC/DFC/01/2026", "Alpha"),
        _master_spec("ONGC/DFC/01/2026", "Alpha duplicate"),
        _master_spec("ONGC/PC/02/2026", "Beta"),
    ]
    groups = csc_export._subgroup_sequence(specs)

    plan = csc_export._index_plan(specs, groups)

    assert plan["bookmarks"] == ["CSCSPEC_1", "CSCSPEC_2", "CSCSPEC_3"]
    assert len(set(plan["bookmarks"])) == len(plan["bookmarks"])
    assert [(row["kind"], row["page"]) for row in plan["rows"]] == [
        ("group", 2), ("spec", 3), ("spec", 4), ("group", 5), ("spec", 6),
    ]
    assert groups[0]["bookmark"] != groups[1]["bookmark"]


def test_dossier_bundle_returns_one_docx_or_a_zip_of_dossiers(monkeypatch):
    entries = [
        {"ref": "r-1", "record_id": 1, "spec_number": "ONGC/DFC/01/2026", "chemical_name": "Alpha", "version": 2},
        {"ref": "r-2", "record_id": 2, "spec_number": "ONGC/PC/02/2026", "chemical_name": "Beta", "version": 4},
        {"ref": "r-3", "record_id": None, "spec_number": "ONGC/PC/03/2026", "chemical_name": "Awaiting", "version": 0},
    ]
    monkeypatch.setattr(cs, "catalogue", lambda: entries)
    monkeypatch.setattr(cs, "specification_data", lambda ref: {"record": object(), "ref": ref})
    monkeypatch.setattr(cs, "dossier_context", lambda data: data)
    monkeypatch.setattr(
        "app.core.services.csc_dossier_export.build_enterprise_dossier",
        lambda data: f"dossier:{data['ref']}".encode(),
    )

    stream, filename, mimetype, skipped = cs.build_dossier_bundle(["r-1"])
    assert filename == "ONGC_DFC_01_2026_v2_dossier.docx"
    assert mimetype.endswith("wordprocessingml.document")
    assert stream.read() == b"dossier:r-1"
    assert skipped == []

    stream, filename, mimetype, skipped = cs.build_dossier_bundle(["r-1", "r-2", "r-3"])
    assert filename.startswith("ONGC_Specification_Dossiers_2_")
    assert mimetype == "application/zip"
    assert skipped == ["Awaiting"]
    from zipfile import ZipFile
    with ZipFile(stream) as archive:
        assert sorted(archive.namelist()) == [
            "ONGC_DFC_01_2026_v2_dossier.docx",
            "ONGC_PC_02_2026_v4_dossier.docx",
        ]


# ── SAP LABIMS comparison workbook ───────────────────────────────────────────


def _stub_parameters(monkeypatch, by_record_id):
    monkeypatch.setattr(
        cs, "parameter_rows",
        lambda record: [
            {
                "id": index,
                "parameter_name": name,
                "parameter_type": "Vital",
                "requirement": requirement,
                "unit_of_measure": "%",
                "conditions": "",
                "test_method": "IS 1234",
                "remarks": "",
                "sort_order": index,
            }
            for index, (name, requirement) in enumerate(by_record_id.get(record.id, []), start=1)
        ],
    )


def _comparison_workbook(monkeypatch, register, records, counts, parameters, category=None):
    from io import BytesIO
    from openpyxl import load_workbook

    _stub_catalogue(monkeypatch, register, records, counts)
    _stub_parameters(monkeypatch, parameters)
    stream, filename = cs.build_specification_comparison_workbook(category)
    return load_workbook(BytesIO(stream.getvalue())), filename


def test_comparison_workbook_gives_every_specification_its_own_sheet(monkeypatch):
    workbook, filename = _comparison_workbook(
        monkeypatch,
        register=[
            _register_row(1, "ONGC / DFC / 01 / 2026", "Barytes", "100101102", 5),
            _register_row(2, "ONGC / PC / 04 / 2026", "Xylene", "100200300"),
        ],
        records=[
            _record(11, "ONGC / DFC / 01 / 2026", "Barytes", "100101102"),
            _record(12, "ONGC / PC / 04 / 2026", "Xylene", "100200300"),
        ],
        counts={11: 2, 12: 1},
        parameters={
            11: [("Specific gravity", "4.20 (Minimum)"), ("Moisture", "1.0 (Maximum)")],
            12: [("Purity", "99.0 (Minimum)")],
        },
    )

    assert workbook.sheetnames == ["Index", "DFC-01 Barytes", "PC-04 Xylene"]
    assert filename.startswith("ONGC_Specifications_vs_SAP_LABIMS_ALL_")

    index = workbook["Index"]
    assert [row[0] for row in index.iter_rows(min_row=2, max_col=1, values_only=True)] == [
        "DFC-01 Barytes", "PC-04 Xylene",
    ]
    assert index["A2"].hyperlink.target == "#'DFC-01 Barytes'!A1"

    sheet = workbook["DFC-01 Barytes"]
    assert [row[0] for row in sheet.iter_rows(min_row=1, max_row=6, max_col=1, values_only=True)] == [
        "Specification No.", "Chemical Name", "Material Code",
        "Chemical Sub-group", "Specification Version", "Standard Testing Days",
    ]
    assert sheet["B1"].value == "ONGC / DFC / 01 / 2026"
    assert sheet["B6"].value == 5
    # A blank row separates the identity block from the table it heads.
    assert sheet["A7"].value is None
    assert [str(area) for area in sheet.merged_cells.ranges] == ["A8:H8", "I8:N8", "O8:P8"]
    assert sheet["A8"].value == "ONGC Corporate Specification"
    assert sheet["I8"].value.startswith("SAP LABIMS — Material Inspection Characteristics")
    assert sheet["O8"].value == "Reconciliation"


def test_comparison_sheet_writes_the_requirement_and_leaves_the_sap_columns_empty(monkeypatch):
    workbook, _ = _comparison_workbook(
        monkeypatch,
        register=[_register_row(1, "ONGC / DFC / 01 / 2026", "Barytes", "100101102")],
        records=[_record(11, "ONGC / DFC / 01 / 2026", "Barytes", "100101102")],
        counts={11: 2},
        parameters={11: [("Specific gravity", "4.20 (Minimum)"), ("Moisture", "1.0 (Maximum)")]},
    )
    sheet = workbook["DFC-01 Barytes"]

    assert [cell.value for cell in sheet[9]] == [
        "S. No.", "Parameter", "Type", "Required Value", "Unit", "Conditions",
        "Test Method", "Remarks",
        "Characteristic", "Description", "Lower Limit", "Upper Limit", "Unit",
        "Method / Procedure",
        "Match?", "Reconciliation Note",
    ]
    assert [cell.value for cell in sheet[10]][:4] == [1, "Specific gravity", "Vital", "4.20 (Minimum)"]
    # The SAP block and the reconciliation block are the reader's to fill in.
    assert [cell.value for cell in sheet[10]][8:] == [None] * 8
    # Frozen below the headings and right of the parameter name, so the SAP
    # columns can be filled without losing sight of what they answer to.
    assert sheet.freeze_panes == "C10"

    invitation = sheet.cell(row=13, column=1).value
    assert invitation.startswith("Characteristics held in SAP LABIMS only")


def test_comparison_sheet_names_stay_legal_and_unique(monkeypatch):
    long_name = "Calcium Carbonate Micronised Extra Fine Grade For Drilling Use"
    workbook, _ = _comparison_workbook(
        monkeypatch,
        register=[
            _register_row(1, "ONGC / DFC / 07 / 2026", long_name, "100000001"),
            _register_row(2, "ONGC / DFC / 07 / 2026", long_name, "100000002"),
        ],
        records=[
            _record(11, "ONGC / DFC / 07 / 2026", long_name, "100000001"),
            _record(12, "ONGC / DFC / 07 / 2026", long_name, "100000002"),
        ],
        counts={11: 1, 12: 1},
        parameters={11: [("Purity", "99")], 12: [("Purity", "99")]},
    )

    sheets = workbook.sheetnames[1:]
    assert len(sheets) == 2
    assert all(len(name) <= 31 for name in sheets)
    assert not any(set(name) & set("[]:*?/\\") for name in sheets)
    assert len({name.casefold() for name in sheets}) == 2
    assert sheets[1].endswith("(2)")


def test_comparison_workbook_covers_one_subgroup_and_skips_unparameterised_records(monkeypatch):
    register = [
        _register_row(1, "ONGC / DFC / 01 / 2026", "Barytes", "100101102"),
        _register_row(2, "ONGC / PC / 04 / 2026", "Xylene", "100200300"),
        _register_row(3, "ONGC / DFC / 09 / 2026", "Awaiting parameters", "100300400"),
    ]
    records = [
        _record(11, "ONGC / DFC / 01 / 2026", "Barytes", "100101102"),
        _record(12, "ONGC / PC / 04 / 2026", "Xylene", "100200300"),
        _record(13, "ONGC / DFC / 09 / 2026", "Awaiting parameters", "100300400"),
    ]
    counts = {11: 1, 12: 1}
    parameters = {11: [("Specific gravity", "4.20")], 12: [("Purity", "99.0")]}

    workbook, filename = _comparison_workbook(
        monkeypatch, register, records, counts, parameters, category="DFC",
    )
    assert workbook.sheetnames == ["Index", "DFC-01 Barytes"]
    assert "_DFC_" in filename

    import pytest

    _stub_catalogue(monkeypatch, register, records, counts)
    _stub_parameters(monkeypatch, parameters)
    with pytest.raises(ValueError, match="No specifications with recorded parameters"):
        cs.build_specification_comparison_workbook("WS")
