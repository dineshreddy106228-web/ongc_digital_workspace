from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.services.inventory_intelligence import _DataStore


def test_build_financial_year_summary_workbook_groups_and_sorts_by_fy() -> None:
    store = _DataStore()
    store._cons = pd.DataFrame(
        [
            {
                "reporting_plant": "10D1",
                "financial_year": "FY24-25",
                "financial_year_start": 2024,
                "material_code": "090001043",
                "material_desc": "BARYTES",
                "actual_usage_qty": 2.0,
                "usage_qty": 2000.0,
                "actual_uom": "MT",
                "uom": "KG",
                "month": 4,
            },
            {
                "reporting_plant": "10D1",
                "financial_year": "FY24-25",
                "financial_year_start": 2024,
                "material_code": "090001044",
                "material_desc": "CHEM-X",
                "actual_usage_qty": 500.0,
                "usage_qty": 500.0,
                "actual_uom": "L",
                "uom": "L",
                "month": 5,
            },
            {
                "reporting_plant": "10D1",
                "financial_year": "FY23-24",
                "financial_year_start": 2023,
                "material_code": "090001045",
                "material_desc": "OLD-CHEM",
                "actual_usage_qty": 1000.0,
                "usage_qty": 1000.0,
                "actual_uom": "KG",
                "uom": "KG",
                "month": 4,
            },
        ]
    )
    store._proc = pd.DataFrame(
        [
            {
                "reporting_plant": "10D1",
                "financial_year": "FY24-25",
                "financial_year_start": 2024,
                "material_code": "090001043",
                "material_desc": "BARYTES",
                "order_unit": "KG",
                "effective_value": 500000.0,
            },
            {
                "reporting_plant": "10D1",
                "financial_year": "FY24-25",
                "financial_year_start": 2024,
                "material_code": "090001044",
                "material_desc": "CHEM-X",
                "order_unit": "L",
                "effective_value": 200000.0,
            },
            {
                "reporting_plant": "10D1",
                "financial_year": "FY23-24",
                "financial_year_start": 2023,
                "material_code": "090001045",
                "material_desc": "OLD-CHEM",
                "order_unit": "KG",
                "effective_value": 100000.0,
            },
        ]
    )

    stream, filename = store.build_financial_year_summary_workbook(plant="10D1")

    assert filename == "material-intelligence-fy-summary-10d1.xlsx"
    workbook = load_workbook(BytesIO(stream.getvalue()))
    assert workbook.sheetnames == [
        "Consumption FY24-25",
        "Procurement FY24-25",
        "Consumption FY23-24",
        "Procurement FY23-24",
    ]

    cons_sheet = workbook["Consumption FY24-25"]
    headers = [cell.value for cell in cons_sheet[1]]
    assert headers[:5] == [
        "Material",
        "Name of Chemical",
        "Unit of Measurement (actual)",
        "Annual Consumption (Actual)",
        "Annual Consumption (MT)",
    ]
    assert cons_sheet["A2"].value == "090001043"
    assert cons_sheet["B2"].value == "BARYTES"
    assert cons_sheet["C2"].value == "MT"
    assert cons_sheet["D2"].value == 2.0
    assert cons_sheet["E2"].value == 2.0
    assert cons_sheet["F2"].value == 2.0
    assert cons_sheet["G2"].value == 2.0
    assert cons_sheet["H2"].value == 0.0
    assert cons_sheet["A3"].value == "090001044"
    assert cons_sheet["D3"].value == 500.0
    assert cons_sheet["E3"].value == 0.5
    assert cons_sheet["F3"].value == 0.0
    assert cons_sheet["G3"].value == 0.0
    assert cons_sheet["H3"].value == 500.0
    assert cons_sheet["I3"].value == 0.5

    proc_sheet = workbook["Procurement FY24-25"]
    assert proc_sheet["A2"].value == "090001043"
    assert proc_sheet["D2"].value == 500000.0
    assert proc_sheet["A3"].value == "090001044"
    assert proc_sheet["D3"].value == 200000.0
