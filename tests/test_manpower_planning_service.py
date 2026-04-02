from __future__ import annotations

from datetime import date
from pathlib import Path
import sys

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.services import manpower_planning


BASE_HEADERS = [
    "CPF NO",
    "NAME",
    "DESIGNATION TEXT",
    "LEVEL",
    "DISCIPLINE TEXT",
    "PERSONAL AREA",
    "LOCATION",
    "REGION",
    "ORG.UNIT TEXT",
    "POSITION TEXT",
    "GENDER KEY",
    "DATE OF BIRTH",
    "DATE OF JOIN ONGC",
    "DATE OF JOIN POST",
    "EFF DATE PROM",
    "DATE OF JOIN PER AREA",
    "DATE OF JOIN POSITION",
    "DATE OF RETIREMENT",
    "HANDICAP",
    "Mobile No",
]


def test_load_manpower_workspace_uses_sheet_reference_date_and_builds_location_summary(tmp_path) -> None:
    workbook_path = tmp_path / "Chemistry_Manpower_as on 26.03.2026.xlsx"
    workbook = Workbook()

    master_sheet = workbook.active
    master_sheet.title = "Manpower "
    master_sheet.append(BASE_HEADERS)
    master_sheet.append([99999, "Ignored Employee"] + [""] * 18)

    agartala = workbook.create_sheet("Agartala DFS")
    agartala.append(BASE_HEADERS + [None, "Vintage"])
    agartala.append(
        [
            105980,
            "Murali Krishna Beera",
            "Deputy General Manager ( Chemistry )",
            "E5",
            "CHEMISTRY",
            "AGAR",
            "AGARTALA",
            "Central Region",
            "TRI MUD",
            "LOCATION MANAGER - MUD -",
            "M",
            date(1984, 8, 7),
            date(2008, 9, 15),
            date(2025, 6, 12),
            date(2025, 1, 1),
            date(2025, 6, 3),
            date(2023, 5, 26),
            date(2044, 8, 31),
            "",
            "9969223025",
            date(2026, 5, 31),
            3,
        ]
    )
    agartala.append(
        [
            123176,
            "Kamleshwar Tiwari",
            "Manager (Chemistry)",
            "E3",
            "CHEMISTRY",
            "AGAR",
            "AGARTALA",
            "Central Region",
            "TRI MUD",
            "SHIFT CHEMIST-DFS-TRI",
            "M",
            date(1984, 4, 15),
            date(2012, 4, 23),
            date(2022, 1, 4),
            date(2022, 1, 1),
            date(2023, 5, 18),
            date(2025, 8, 1),
            date(2044, 4, 30),
            "",
            "8584891660",
            date(2026, 5, 31),
            0,
        ]
    )

    silchar = workbook.create_sheet("DFS Silchar")
    silchar.append(BASE_HEADERS)
    silchar.append(
        [
            81696,
            "Zabed Rahman",
            "Deputy General Manager ( Chemistry )",
            "E5",
            "CHEMISTRY",
            "SILC",
            "SILCHAR",
            "Eastern Region",
            "CAH MUD",
            "CAH MUD - Location Manage",
            "M",
            date(1967, 2, 16),
            date(1993, 2, 22),
            date(2020, 1, 15),
            date(2020, 1, 1),
            date(2024, 7, 5),
            date(2020, 6, 1),
            date(2027, 2, 28),
            "",
            "9969223363",
        ]
    )

    workbook.save(workbook_path)

    workspace = manpower_planning._load_manpower_workspace(workbook_path)

    assert workspace["available"] is True
    assert workspace["as_of_date_iso"] == "2026-05-31"
    assert workspace["total_locations"] == 2
    assert workspace["total_employees"] == 3
    assert workspace["largest_location"] == "Agartala DFS"
    assert len(workspace["retiring_soon_employees"]) == 1

    agartala_location = workspace["locations"][0]
    assert agartala_location["slug"] == "agartala-dfs"
    assert agartala_location["employee_count"] == 2
    assert agartala_location["average_vintage"] == 1.5
    assert agartala_location["employees"][0]["name"] == "Murali Krishna Beera"
    assert agartala_location["employees"][0]["vintage_years"] == 3

    silchar_location = next(
        location for location in workspace["locations"] if location["slug"] == "dfs-silchar"
    )
    assert silchar_location["employees"][0]["vintage_years"] == 5
    assert silchar_location["retirements_within_two_years"] == 1

    retirement_watch = workspace["retiring_soon_employees"][0]
    assert retirement_watch["name"] == "Zabed Rahman"
    assert retirement_watch["location_slug"] == "dfs-silchar"

    vintage_cohorts = {cohort["years"]: cohort for cohort in workspace["vintage_cohorts"]}
    assert [cohort["years"] for cohort in workspace["vintage_cohorts"]] == [10, 9, 8, 7]
    assert vintage_cohorts[10]["count"] == 0
    assert vintage_cohorts[9]["count"] == 0
    assert vintage_cohorts[8]["count"] == 0
    assert vintage_cohorts[7]["count"] == 0

    transfer_cohorts = {cohort["level"]: cohort for cohort in workspace["transfer_level_cohorts"]}
    assert [cohort["level"] for cohort in workspace["transfer_level_cohorts"]] == ["E5", "E4"]
    assert transfer_cohorts["E5"]["count"] == 2
    assert transfer_cohorts["E4"]["count"] == 0
    assert transfer_cohorts["E5"]["employees"][0]["name"] == "Murali Krishna Beera"

    northeast_candidates = workspace["northeast_transfer_candidates"]
    assert northeast_candidates["count"] == 2
    assert northeast_candidates["employees"][0]["name"] == "Murali Krishna Beera"
    assert northeast_candidates["employees"][1]["name"] == "Zabed Rahman"
