from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook
import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import app.core.services.inventory_monthly_report as monthly_report
from app.core.services.inventory_intelligence import _merge_monthly_seed_collection_frame


def _xlsx_bytes(frame: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False)
    return buffer.getvalue()


def _stub_pdf(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(monthly_report, "_build_pdf_report", lambda summary: b"%PDF-fake")


def _sample_mb51() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Plant": "10D1",
                "Material": "090001043",
                "Material Description": "BARYTES",
                "Movement Type": "201",
                "Posting Date": "2026-01-15",
                "Qty in unit of entry": 1.0,
                "Unit of Entry": "MT",
                "Amt.in Loc.Cur.": 100000.0,
                "Storage Location": "0001",
            },
            {
                "Plant": "10D1",
                "Material": "090001043",
                "Material Description": "BARYTES",
                "Movement Type": "201",
                "Posting Date": "2026-02-15",
                "Qty in unit of entry": 1.0,
                "Unit of Entry": "MT",
                "Amt.in Loc.Cur.": 100000.0,
                "Storage Location": "0001",
            },
            {
                "Plant": "10D1",
                "Material": "090001043",
                "Material Description": "BARYTES",
                "Movement Type": "201",
                "Posting Date": "2026-03-15",
                "Qty in unit of entry": 5.0,
                "Unit of Entry": "MT",
                "Amt.in Loc.Cur.": 500000.0,
                "Storage Location": "0001",
            },
            {
                "Plant": "10D1",
                "Material": "090001044",
                "Material Description": "SCALE INHIBITOR",
                "Movement Type": "201",
                "Posting Date": "2026-01-12",
                "Qty in unit of entry": 0.0,
                "Unit of Entry": "MT",
                "Amt.in Loc.Cur.": 0.0,
                "Storage Location": "0002",
            },
            {
                "Plant": "10D1",
                "Material": "090001044",
                "Material Description": "SCALE INHIBITOR",
                "Movement Type": "201",
                "Posting Date": "2026-02-12",
                "Qty in unit of entry": 0.0,
                "Unit of Entry": "MT",
                "Amt.in Loc.Cur.": 0.0,
                "Storage Location": "0002",
            },
            {
                "Plant": "10D1",
                "Material": "090001044",
                "Material Description": "SCALE INHIBITOR",
                "Movement Type": "201",
                "Posting Date": "2026-03-12",
                "Qty in unit of entry": 0.0,
                "Unit of Entry": "MT",
                "Amt.in Loc.Cur.": 0.0,
                "Storage Location": "0002",
            },
        ]
    )


def _sample_me2m() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Plant": "10D1",
                "Net Price": 100000.0,
                "Purchasing Document": "4500001",
                "Document Date": "2026-01-01",
                "Material": "090001043",
                "Item": "10",
                "Supplier/Supplying Plant": "Vendor A",
                "Short Text": "BARYTES",
                "Order Quantity": 5.0,
                "Still to be delivered (qty)": 4.0,
                "Order Unit": "MT",
                "Currency": "INR",
                "Price Unit": 1.0,
                "Effective value": 800000.0,
                "Release indicator": "R",
            },
            {
                "Plant": "10D1",
                "Net Price": 50000.0,
                "Purchasing Document": "4500002",
                "Document Date": "2026-03-05",
                "Material": "090001044",
                "Item": "10",
                "Supplier/Supplying Plant": "Vendor Solo",
                "Short Text": "SCALE INHIBITOR",
                "Order Quantity": 3.0,
                "Still to be delivered (qty)": 1.0,
                "Order Unit": "MT",
                "Currency": "INR",
                "Price Unit": 1.0,
                "Effective value": 300000.0,
                "Release indicator": "R",
            },
        ]
    )


def _sample_mc9() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Plant": "10D1 PLANT",
                "Material": "090001043 BARYTES",
                "Storage Location": "0001",
                "Month": "2.2026",
                "ValStckVal": 700000.0,
                "ValStckVal.1": "INR",
                "Val. stock": 6.0,
                "Val. stock.1": "MT",
                "Tot. usage": 1.0,
                "Tot. usage.1": "MT",
                "Tot.us.val": 100000.0,
                "Tot.us.val.1": "INR",
            },
            {
                "Plant": "10D1 PLANT",
                "Material": "090001043 BARYTES",
                "Storage Location": "0001",
                "Month": "3.2026",
                "ValStckVal": 100000.0,
                "ValStckVal.1": "INR",
                "Val. stock": 0.5,
                "Val. stock.1": "MT",
                "Tot. usage": 5.0,
                "Tot. usage.1": "MT",
                "Tot.us.val": 500000.0,
                "Tot.us.val.1": "INR",
            },
            {
                "Plant": "10D1 PLANT",
                "Material": "090001044 SCALE INHIBITOR",
                "Storage Location": "0002",
                "Month": "1.2026",
                "ValStckVal": 300000.0,
                "ValStckVal.1": "INR",
                "Val. stock": 3.0,
                "Val. stock.1": "MT",
                "Tot. usage": 0.0,
                "Tot. usage.1": "MT",
                "Tot.us.val": 0.0,
                "Tot.us.val.1": "INR",
            },
            {
                "Plant": "10D1 PLANT",
                "Material": "090001044 SCALE INHIBITOR",
                "Storage Location": "0002",
                "Month": "2.2026",
                "ValStckVal": 300000.0,
                "ValStckVal.1": "INR",
                "Val. stock": 3.0,
                "Val. stock.1": "MT",
                "Tot. usage": 0.0,
                "Tot. usage.1": "MT",
                "Tot.us.val": 0.0,
                "Tot.us.val.1": "INR",
            },
            {
                "Plant": "10D1 PLANT",
                "Material": "090001044 SCALE INHIBITOR",
                "Storage Location": "0002",
                "Month": "3.2026",
                "ValStckVal": 300000.0,
                "ValStckVal.1": "INR",
                "Val. stock": 3.0,
                "Val. stock.1": "MT",
                "Tot. usage": 0.0,
                "Tot. usage.1": "MT",
                "Tot.us.val": 0.0,
                "Tot.us.val.1": "INR",
            },
        ]
    )


def test_build_monthly_inventory_report_package_generates_intelligence_summary(monkeypatch: pytest.MonkeyPatch) -> None:
    _stub_pdf(monkeypatch)
    monkeypatch.setattr(
        monthly_report,
        "_load_previous_summary",
        lambda report_year, report_month: {
            "report_label": "Feb 2026",
            "kpis": {
                "total_mb51_consumption_value": 400000.0,
                "total_mb51_consumption_mt": 1.0,
                "total_mc9_usage_value": 100000.0,
                "total_closing_stock_value": 1000000.0,
                "total_procurement_value": 500000.0,
            },
        },
    )

    package = monthly_report.build_monthly_inventory_report_package(
        report_year=2026,
        report_month=3,
        mb51_bytes=_xlsx_bytes(_sample_mb51()),
        mb51_filename="mb51_mar_2026.xlsx",
        me2m_bytes=_xlsx_bytes(_sample_me2m()),
        me2m_filename="me2m_mar_2026.xlsx",
        mc9_bytes=_xlsx_bytes(_sample_mc9()),
        mc9_filename="mc9_mar_2026.xlsx",
    )

    assert package.report_label == "Mar 2026"
    assert package.pdf_bytes == b"%PDF-fake"
    assert package.summary["kpi_trends"]["total_mb51_consumption_value"]["available"] is True
    assert package.summary["kpi_trends"]["total_mb51_consumption_value"]["previous_label"] == "Feb 2026"
    assert package.summary["stock_health"]["band_counts"]["red"] >= 1
    assert package.summary["stock_health"]["dead_stock_materials"][0]["material_desc"] == "SCALE INHIBITOR"
    assert any(item["type"] == "Consumption spike" for item in package.summary["anomalies"])
    assert package.summary["open_order_risk"]["aged_order_count"] >= 1
    assert package.summary["top_consumption_by_value"][0]["abc_class"] == "A"
    assert "A-class materials contributed" in package.summary["executive_narrative"]
    assert "aged beyond 60 days" in package.summary["executive_narrative"]

    workbook = load_workbook(BytesIO(package.excel_bytes))
    assert workbook.sheetnames == [
        "Overview",
        "Material Rollup",
        "Stock Movement",
        "MB51 Consumption",
        "MC9 Stock",
        "Opening Stock",
        "ME2M Snapshot",
        "MB51 Raw",
        "ME2M Raw",
        "MC9 Raw",
    ]
    overview = workbook["Overview"]
    values = {str(overview[f"A{row}"].value): overview[f"B{row}"].value for row in range(1, 40) if overview[f"A{row}"].value}
    assert values["Red-zone materials"] >= 1
    rollup = workbook["Material Rollup"]
    assert rollup["A2"].value == "90001043"
    assert rollup["B2"].value == "BARYTES"
    assert rollup["AB2"].value == "A"


def test_build_monthly_inventory_report_package_rejects_missing_month_rows(monkeypatch: pytest.MonkeyPatch) -> None:
    _stub_pdf(monkeypatch)
    mb51 = pd.DataFrame(
        [
            {
                "Plant": "10D1",
                "Material": "090001043",
                "Material Description": "BARYTES",
                "Movement Type": "201",
                "Posting Date": "2026-02-15",
                "Qty in unit of entry": 2.0,
                "Unit of Entry": "MT",
                "Amt.in Loc.Cur.": 500000.0,
            }
        ]
    )

    with pytest.raises(ValueError, match="MB51 file does not contain rows for Mar 2026"):
        monthly_report.build_monthly_inventory_report_package(
            report_year=2026,
            report_month=3,
            mb51_bytes=_xlsx_bytes(mb51),
            mb51_filename="mb51_feb_2026.xlsx",
            me2m_bytes=_xlsx_bytes(_sample_me2m()),
            me2m_filename="me2m_mar_2026.xlsx",
            mc9_bytes=_xlsx_bytes(_sample_mc9()),
            mc9_filename="mc9_mar_2026.xlsx",
        )


def test_build_monthly_inventory_report_package_estimates_opening_stock_when_prior_mc9_month_missing(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    _stub_pdf(monkeypatch)
    mc9 = pd.DataFrame(
        [
            {
                "Plant": "10D1 PLANT",
                "Material": "090001043 BARYTES",
                "Storage Location": "0001",
                "Month": "3.2026",
                "ValStckVal": 900000.0,
                "ValStckVal.1": "INR",
                "Val. stock": 9.0,
                "Val. stock.1": "MT",
                "Tot. usage": 2.0,
                "Tot. usage.1": "MT",
                "Tot.us.val": 500000.0,
                "Tot.us.val.1": "INR",
            }
        ]
    )

    package = monthly_report.build_monthly_inventory_report_package(
        report_year=2026,
        report_month=3,
        mb51_bytes=_xlsx_bytes(
            pd.DataFrame(
                [
                    {
                        "Plant": "10D1",
                        "Material": "090001043",
                        "Material Description": "BARYTES",
                        "Movement Type": "201",
                        "Posting Date": "2026-03-15",
                        "Qty in unit of entry": 2.0,
                        "Unit of Entry": "MT",
                        "Amt.in Loc.Cur.": 500000.0,
                        "Storage Location": "0001",
                    }
                ]
            )
        ),
        mb51_filename="mb51_mar_2026.xlsx",
        me2m_bytes=_xlsx_bytes(_sample_me2m().iloc[[0]]),
        me2m_filename="me2m_mar_2026.xlsx",
        mc9_bytes=_xlsx_bytes(mc9),
        mc9_filename="mc9_mar_2026.xlsx",
    )

    workbook = load_workbook(BytesIO(package.excel_bytes))
    stock_movement = workbook["Stock Movement"]
    assert stock_movement["C2"].value == "MT"
    assert stock_movement["D2"].value == 7
    assert stock_movement["E2"].value == 7
    assert stock_movement["I2"].value == "MT"
    assert stock_movement["J2"].value == 9


def test_days_on_hand_and_kpi_trends_helpers() -> None:
    trends = monthly_report._build_kpi_trends(
        {
            "total_mb51_consumption_value": 500.0,
            "total_mb51_consumption_mt": 5.0,
            "total_mc9_usage_value": 400.0,
            "total_closing_stock_value": 100.0,
            "total_procurement_value": 800.0,
        },
        {
            "report_label": "Feb 2026",
            "kpis": {
                "total_mb51_consumption_value": 250.0,
                "total_mb51_consumption_mt": 4.0,
                "total_mc9_usage_value": 400.0,
                "total_closing_stock_value": 200.0,
                "total_procurement_value": 400.0,
            },
        },
    )
    assert trends["total_mb51_consumption_value"]["direction"] == "up"
    assert trends["total_closing_stock_value"]["direction"] == "down"
    assert trends["total_mc9_usage_value"]["direction"] == "flat"
    assert monthly_report._pdf_trend_label({"available": False, "pct_change": None, "previous_label": None}) == (
        "Baseline unavailable - upload prior month batch"
    )

    stock_health = monthly_report._build_stock_health(
        pd.DataFrame(
            [
                {
                    "material_code": "90001043",
                    "material_desc": "BARYTES",
                    "closing_stock_qty_mt": 1.0,
                    "closing_stock_value": 100.0,
                },
                {
                    "material_code": "90001044",
                    "material_desc": "SCALE INHIBITOR",
                    "closing_stock_qty_mt": 5.0,
                    "closing_stock_value": 300.0,
                },
            ]
        ),
        pd.DataFrame(
            [
                {
                    "material_code": "90001043",
                    "material_desc": "BARYTES",
                    "avg_monthly_usage_mt": 2.0,
                    "recent_zero_streak": 0,
                },
                {
                    "material_code": "90001044",
                    "material_desc": "SCALE INHIBITOR",
                    "avg_monthly_usage_mt": 0.0,
                    "recent_zero_streak": 3,
                },
            ]
        ),
    )
    assert stock_health["band_counts"]["red"] == 1
    assert stock_health["dead_stock_materials"][0]["material_desc"] == "SCALE INHIBITOR"


def test_monthly_seed_merge_replaces_only_uploaded_month() -> None:
    existing = pd.DataFrame(
        [
            {
                "Plant": "10D1",
                "Material": "090001043",
                "Material Description": "BARYTES",
                "Movement Type": "201",
                "Posting Date": "2026-01-15",
                "Qty in unit of entry": 1.0,
                "Unit of Entry": "MT",
                "Amt.in Loc.Cur.": 100000.0,
            },
            {
                "Plant": "10D1",
                "Material": "090001043",
                "Material Description": "BARYTES",
                "Movement Type": "201",
                "Posting Date": "2026-02-15",
                "Qty in unit of entry": 2.0,
                "Unit of Entry": "MT",
                "Amt.in Loc.Cur.": 200000.0,
            },
        ]
    )
    uploaded = pd.DataFrame(
        [
            {
                "Plant": "10D1",
                "Material": "090001043",
                "Material Description": "BARYTES",
                "Movement Type": "201",
                "Posting Date": "2026-03-15",
                "Qty in unit of entry": 3.0,
                "Unit of Entry": "MT",
                "Amt.in Loc.Cur.": 300000.0,
            }
        ]
    )

    merged = _merge_monthly_seed_collection_frame(existing, uploaded, "consumption")
    posting_dates = sorted(pd.to_datetime(merged["Posting Date"]).dt.strftime("%Y-%m").tolist())
    assert posting_dates == ["2026-01", "2026-02", "2026-03"]

    replacement = pd.DataFrame(
        [
            {
                "Plant": "10D1",
                "Material": "090001043",
                "Material Description": "BARYTES",
                "Movement Type": "201",
                "Posting Date": "2026-02-20",
                "Qty in unit of entry": 9.0,
                "Unit of Entry": "MT",
                "Amt.in Loc.Cur.": 900000.0,
            }
        ]
    )
    replaced = _merge_monthly_seed_collection_frame(existing, replacement, "consumption")
    replaced_dates = sorted(pd.to_datetime(replaced["Posting Date"]).dt.strftime("%Y-%m-%d").tolist())
    assert replaced_dates == ["2026-01-15", "2026-02-20"]
