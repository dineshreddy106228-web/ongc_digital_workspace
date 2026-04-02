from __future__ import annotations

"""Workbook-backed data access for the Manpower Planning module."""

from collections import Counter
from datetime import date, datetime
from pathlib import Path
import re

from flask import current_app
from openpyxl import load_workbook

from app.extensions import cache


MASTER_SHEET_NAME = "MANPOWER"
FILENAME_DATE_PATTERN = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")
BASE_HEADERS = {
    "CPF NO",
    "NAME",
    "DESIGNATION TEXT",
    "LEVEL",
    "DISCIPLINE TEXT",
    "PERSONAL AREA",
    "LOCATION",
    "REGION",
    "ORG.UNIT TEXT",
    "POSITION TEXT",
    "GENDER KEY",
    "DATE OF BIRTH",
    "DATE OF JOIN ONGC",
    "DATE OF JOIN POST",
    "EFF DATE PROM",
    "DATE OF JOIN PER AREA",
    "DATE OF JOIN POSITION",
    "DATE OF RETIREMENT",
    "HANDICAP",
    "MOBILE NO",
}
VINTAGE_BANDS = (
    ("0-1 Years", 0, 1),
    ("2-4 Years", 2, 4),
    ("5-9 Years", 5, 9),
    ("10+ Years", 10, None),
)
RETIREMENT_SOON_DAYS = 730
NORTHEAST_TRANSFER_KEYWORDS = (
    "agartala",
    "nazira",
    "jorhat",
    "silchar",
    "sivasagar",
    "svs",
)


def _normalize_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).upper()


def _clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text == "//":
        return ""
    return text


def _coerce_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _coerce_int(value) -> int | None:
    text = _clean_text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _format_date(value: date | None) -> str:
    if value is None:
        return "-"
    return value.strftime("%d %b %Y")


def _slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "location"


def _is_northeast_transfer_location(location_name: str) -> bool:
    normalized = location_name.strip().lower()
    return any(keyword in normalized for keyword in NORTHEAST_TRANSFER_KEYWORDS)


def _vintage_band(value: int | None) -> str:
    if value is None:
        return "Unknown"
    for label, start, end in VINTAGE_BANDS:
        if end is None and value >= start:
            return label
        if end is not None and start <= value <= end:
            return label
    return "Unknown"


def _full_years_between(start: date | None, end: date) -> int | None:
    if start is None or start > end:
        return None
    years = end.year - start.year
    if (end.month, end.day) < (start.month, start.day):
        years -= 1
    return max(years, 0)


def _days_until(target_date: date | None, as_of_date: date) -> int | None:
    if target_date is None:
        return None
    return (target_date - as_of_date).days


def _retirement_tone(days_until_retirement: int | None) -> str:
    if days_until_retirement is None:
        return "none"
    if days_until_retirement < 0:
        return "past"
    if days_until_retirement <= RETIREMENT_SOON_DAYS:
        return "urgent"
    return "future"


def _retirement_horizon_label(days_until_retirement: int | None) -> str:
    if days_until_retirement is None:
        return "Retirement date not available"
    if days_until_retirement < 0:
        return "Retirement date passed"
    if days_until_retirement == 0:
        return "Retiring today"
    months = max(1, round(days_until_retirement / 30.4))
    if months < 12:
        return f"Retiring in {months} month" if months == 1 else f"Retiring in {months} months"
    years = max(1, round(days_until_retirement / 365))
    return f"Retiring in {years} year" if years == 1 else f"Retiring in {years} years"


def _parse_filename_date(path: Path) -> date | None:
    match = FILENAME_DATE_PATTERN.search(path.name)
    if not match:
        return None
    day, month, year = (int(part) for part in match.groups())
    return date(year, month, day)


def _seed_path() -> Path:
    configured_path = current_app.config.get("MANPOWER_SEED_PATH", "")
    path = Path(configured_path)
    if not path.is_absolute():
        path = Path(current_app.root_path).parent / path
    return path


def _empty_workspace(path: Path, error_message: str) -> dict:
    return {
        "available": False,
        "error_message": error_message,
        "source_filename": path.name,
        "source_path": str(path),
        "as_of_date_label": "-",
        "as_of_date_iso": None,
        "total_locations": 0,
        "total_employees": 0,
        "average_vintage": 0.0,
        "max_vintage": 0,
        "retirements_within_two_years": 0,
        "largest_location": None,
        "locations": [],
    }


def _sheet_is_master(title: str) -> bool:
    return _normalize_header(title) == MASTER_SHEET_NAME


def _parse_location_sheet(worksheet) -> dict:
    rows = worksheet.iter_rows(values_only=True)
    header = list(next(rows, ()))
    header_map = {
        _normalize_header(cell): index
        for index, cell in enumerate(header)
        if _clean_text(cell)
    }
    missing_headers = BASE_HEADERS - set(header_map)
    if missing_headers:
        return {
            "name": worksheet.title.strip(),
            "slug": _slugify(worksheet.title.strip()),
            "employees": [],
            "reference_dates": [],
        }

    employees: list[dict] = []
    reference_dates: list[date] = []
    for row in rows:
        cpf_no = _coerce_int(row[header_map["CPF NO"]])
        name = _clean_text(row[header_map["NAME"]])
        if cpf_no is None and not name:
            continue

        reference_date = _coerce_date(row[20]) if len(row) > 20 else None
        source_vintage = _coerce_int(row[21]) if len(row) > 21 else None
        if reference_date is not None:
            reference_dates.append(reference_date)

        date_of_join_position = _coerce_date(row[header_map["DATE OF JOIN POSITION"]])
        date_of_join_per_area = _coerce_date(row[header_map["DATE OF JOIN PER AREA"]])
        date_of_join_post = _coerce_date(row[header_map["DATE OF JOIN POST"]])
        date_of_join_ongc = _coerce_date(row[header_map["DATE OF JOIN ONGC"]])

        employees.append(
            {
                "cpf_no": str(cpf_no) if cpf_no is not None else _clean_text(row[header_map["CPF NO"]]),
                "name": name or "Unknown",
                "designation": _clean_text(row[header_map["DESIGNATION TEXT"]]),
                "level": _clean_text(row[header_map["LEVEL"]]),
                "discipline": _clean_text(row[header_map["DISCIPLINE TEXT"]]),
                "personal_area": _clean_text(row[header_map["PERSONAL AREA"]]),
                "base_location": _clean_text(row[header_map["LOCATION"]]),
                "region": _clean_text(row[header_map["REGION"]]),
                "org_unit": _clean_text(row[header_map["ORG.UNIT TEXT"]]),
                "position": _clean_text(row[header_map["POSITION TEXT"]]),
                "gender": _clean_text(row[header_map["GENDER KEY"]]),
                "date_of_birth": _coerce_date(row[header_map["DATE OF BIRTH"]]),
                "date_of_join_ongc": date_of_join_ongc,
                "date_of_join_post": date_of_join_post,
                "date_of_join_per_area": date_of_join_per_area,
                "date_of_join_position": date_of_join_position,
                "date_of_retirement": _coerce_date(row[header_map["DATE OF RETIREMENT"]]),
                "mobile_no": _clean_text(row[header_map["MOBILE NO"]]),
                "_source_vintage": source_vintage,
                "_vintage_anchor": (
                    date_of_join_position
                    or date_of_join_per_area
                    or date_of_join_post
                    or date_of_join_ongc
                ),
            }
        )

    return {
        "name": worksheet.title.strip(),
        "slug": _slugify(worksheet.title.strip()),
        "employees": employees,
        "reference_dates": reference_dates,
    }


def _materialize_location(parsed_location: dict, as_of_date: date) -> dict:
    employees: list[dict] = []
    vintage_values: list[int] = []
    region_counts: Counter[str] = Counter()
    base_location_counts: Counter[str] = Counter()
    level_counts: Counter[str] = Counter()
    band_counts: Counter[str] = Counter()
    retirements_within_two_years = 0
    upcoming_retirements: list[dict] = []
    next_retirement_after_watch: dict | None = None

    for raw_employee in parsed_location["employees"]:
        vintage_years = raw_employee["_source_vintage"]
        if vintage_years is None:
            vintage_years = _full_years_between(raw_employee["_vintage_anchor"], as_of_date)

        if vintage_years is not None:
            vintage_values.append(vintage_years)
            band_counts[_vintage_band(vintage_years)] += 1
        else:
            band_counts["Unknown"] += 1

        region = raw_employee["region"] or "Unspecified"
        base_location = raw_employee["base_location"] or "Unspecified"
        level = raw_employee["level"] or "Unspecified"
        region_counts[region] += 1
        base_location_counts[base_location] += 1
        level_counts[level] += 1

        retirement_date = raw_employee["date_of_retirement"]
        retirement_days_out = _days_until(retirement_date, as_of_date)
        retirement_tone = _retirement_tone(retirement_days_out)
        retirement_horizon = _retirement_horizon_label(retirement_days_out)

        if retirement_date is not None and 0 <= (retirement_date - as_of_date).days <= RETIREMENT_SOON_DAYS:
            retirements_within_two_years += 1
        if retirement_date is not None and retirement_days_out is not None and retirement_days_out >= 0:
            retirement_preview = {
                "cpf_no": raw_employee["cpf_no"],
                "name": raw_employee["name"],
                "designation": raw_employee["designation"] or "-",
                "level": raw_employee["level"] or "-",
                "base_location": raw_employee["base_location"] or "-",
                "date_of_retirement": _format_date(retirement_date),
                "retirement_days_out": retirement_days_out,
                "retirement_tone": retirement_tone,
                "retirement_horizon": retirement_horizon,
            }
            if retirement_days_out <= RETIREMENT_SOON_DAYS:
                upcoming_retirements.append(retirement_preview)
            elif next_retirement_after_watch is None:
                next_retirement_after_watch = retirement_preview

        employees.append(
            {
                "cpf_no": raw_employee["cpf_no"],
                "name": raw_employee["name"],
                "designation": raw_employee["designation"] or "-",
                "level": raw_employee["level"] or "-",
                "discipline": raw_employee["discipline"] or "-",
                "personal_area": raw_employee["personal_area"] or "-",
                "base_location": raw_employee["base_location"] or "-",
                "region": raw_employee["region"] or "-",
                "org_unit": raw_employee["org_unit"] or "-",
                "position": raw_employee["position"] or "-",
                "gender": raw_employee["gender"] or "-",
                "mobile_no": raw_employee["mobile_no"] or "-",
                "date_of_birth": _format_date(raw_employee["date_of_birth"]),
                "date_of_join_ongc": _format_date(raw_employee["date_of_join_ongc"]),
                "date_of_join_post": _format_date(raw_employee["date_of_join_post"]),
                "date_of_join_per_area": _format_date(raw_employee["date_of_join_per_area"]),
                "date_of_join_position": _format_date(raw_employee["date_of_join_position"]),
                "date_of_retirement": _format_date(raw_employee["date_of_retirement"]),
                "retirement_days_out": retirement_days_out,
                "retirement_tone": retirement_tone,
                "retirement_horizon": retirement_horizon,
                "is_retiring_soon": bool(
                    retirement_days_out is not None and 0 <= retirement_days_out <= RETIREMENT_SOON_DAYS
                ),
                "vintage_years": vintage_years,
                "vintage_label": (
                    "-"
                    if vintage_years is None
                    else f"{vintage_years} yr" if vintage_years == 1 else f"{vintage_years} yrs"
                ),
            }
        )

    employees.sort(
        key=lambda employee: (
            employee["vintage_years"] is None,
            -(employee["vintage_years"] or 0),
            employee["name"].lower(),
        )
    )
    upcoming_retirements.sort(
        key=lambda employee: (
            employee["retirement_days_out"] is None,
            employee["retirement_days_out"] if employee["retirement_days_out"] is not None else 10**9,
            employee["name"].lower(),
        )
    )

    return {
        "name": parsed_location["name"],
        "slug": parsed_location["slug"],
        "employee_count": len(employees),
        "average_vintage": round(sum(vintage_values) / len(vintage_values), 1) if vintage_values else 0.0,
        "max_vintage": max(vintage_values) if vintage_values else 0,
        "retirements_within_two_years": retirements_within_two_years,
        "primary_region": region_counts.most_common(1)[0][0] if region_counts else "-",
        "reported_locations": ", ".join(sorted(base_location_counts)),
        "top_levels": ", ".join(level for level, _ in level_counts.most_common(3)),
        "vintage_bands": [
            {
                "label": label,
                "count": band_counts.get(label, 0),
            }
            for label, _start, _end in VINTAGE_BANDS
        ],
        "upcoming_retirements": upcoming_retirements[:8],
        "next_retirement_date": (
            upcoming_retirements[0]["date_of_retirement"]
            if upcoming_retirements
            else next_retirement_after_watch["date_of_retirement"] if next_retirement_after_watch
            else None
        ),
        "employees": employees,
    }


def _load_manpower_workspace(path: Path) -> dict:
    if not path.exists():
        return _empty_workspace(path, "Seed workbook not found.")

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        parsed_locations: list[dict] = []
        reference_dates: Counter[date] = Counter()
        for worksheet in workbook.worksheets:
            if _sheet_is_master(worksheet.title):
                continue
            parsed_location = _parse_location_sheet(worksheet)
            if not parsed_location["employees"]:
                continue
            parsed_locations.append(parsed_location)
            reference_dates.update(parsed_location["reference_dates"])

        as_of_date = (
            reference_dates.most_common(1)[0][0]
            if reference_dates
            else (_parse_filename_date(path) or date.today())
        )
        locations = [
            _materialize_location(parsed_location, as_of_date)
            for parsed_location in parsed_locations
        ]
    finally:
        workbook.close()

    locations.sort(
        key=lambda location: (
            -location["employee_count"],
            -location["average_vintage"],
            location["name"].lower(),
        )
    )

    all_vintages = [
        employee["vintage_years"]
        for location in locations
        for employee in location["employees"]
        if employee["vintage_years"] is not None
    ]
    total_employees = sum(location["employee_count"] for location in locations)
    largest_location = locations[0]["name"] if locations else None
    retiring_soon_employees: list[dict] = []
    exact_vintage_years = (10, 9, 8, 7)
    vintage_cohorts: dict[int, list[dict]] = {year: [] for year in exact_vintage_years}
    transfer_levels = ("E5", "E4")
    transfer_level_cohorts: dict[str, list[dict]] = {level: [] for level in transfer_levels}
    northeast_transfer_candidates: list[dict] = []

    for location in locations:
        for employee in location["employees"]:
            employee_snapshot = {
                "name": employee["name"],
                "cpf_no": employee["cpf_no"],
                "designation": employee["designation"],
                "level": employee["level"],
                "base_location": employee["base_location"],
                "region": employee["region"],
                "location_name": location["name"],
                "location_slug": location["slug"],
                "date_of_join_position": employee["date_of_join_position"],
                "date_of_retirement": employee["date_of_retirement"],
                "retirement_days_out": employee["retirement_days_out"],
                "retirement_horizon": employee["retirement_horizon"],
                "vintage_years": employee["vintage_years"],
                "vintage_label": employee["vintage_label"],
                "position": employee["position"],
            }
            if employee["is_retiring_soon"]:
                retiring_soon_employees.append(employee_snapshot)
            if employee["vintage_years"] in vintage_cohorts:
                vintage_cohorts[employee["vintage_years"]].append(employee_snapshot)
            if employee["level"] in transfer_level_cohorts:
                transfer_level_cohorts[employee["level"]].append(employee_snapshot)
            if (
                _is_northeast_transfer_location(location["name"])
                and employee["vintage_years"] is not None
                and employee["vintage_years"] >= 3
            ):
                northeast_transfer_candidates.append(employee_snapshot)

    retiring_soon_employees.sort(
        key=lambda employee: (
            employee["retirement_days_out"] is None,
            employee["retirement_days_out"] if employee["retirement_days_out"] is not None else 10**9,
            employee["location_name"].lower(),
            employee["name"].lower(),
        )
    )
    vintage_cohorts_summary = []
    for year in exact_vintage_years:
        cohort = vintage_cohorts[year]
        cohort.sort(
            key=lambda employee: (
                employee["location_name"].lower(),
                employee["name"].lower(),
            )
        )
        vintage_cohorts_summary.append(
            {
                "years": year,
                "label": f"{year} Years",
                "count": len(cohort),
                "employees": cohort,
            }
        )
    transfer_level_summary = []
    for level in transfer_levels:
        cohort = transfer_level_cohorts[level]
        cohort.sort(
            key=lambda employee: (
                employee["location_name"].lower(),
                employee["name"].lower(),
            )
        )
        transfer_level_summary.append(
            {
                "level": level,
                "label": f"{level} Officers",
                "count": len(cohort),
                "employees": cohort,
            }
        )
    northeast_transfer_candidates.sort(
        key=lambda employee: (
            employee["location_name"].lower(),
            -(employee["vintage_years"] or 0),
            employee["name"].lower(),
        )
    )

    return {
        "available": bool(locations),
        "error_message": None if locations else "No manpower rows were found in the workbook.",
        "source_filename": path.name,
        "source_path": str(path),
        "as_of_date_label": _format_date(as_of_date),
        "as_of_date_iso": as_of_date.isoformat(),
        "total_locations": len(locations),
        "total_employees": total_employees,
        "average_vintage": round(sum(all_vintages) / len(all_vintages), 1) if all_vintages else 0.0,
        "max_vintage": max(all_vintages) if all_vintages else 0,
        "retirements_within_two_years": sum(
            location["retirements_within_two_years"] for location in locations
        ),
        "largest_location": largest_location,
        "retiring_soon_employees": retiring_soon_employees,
        "vintage_cohorts": vintage_cohorts_summary,
        "transfer_level_cohorts": transfer_level_summary,
        "northeast_transfer_candidates": {
            "label": "North East Transfer Eligibility",
            "count": len(northeast_transfer_candidates),
            "locations": [
                location["name"]
                for location in locations
                if _is_northeast_transfer_location(location["name"])
            ],
            "employees": northeast_transfer_candidates,
        },
        "locations": locations,
    }


@cache.memoize(timeout=300)
def _get_manpower_workspace_cached(seed_path: str, mtime_ns: int) -> dict:
    del mtime_ns
    return _load_manpower_workspace(Path(seed_path))


def get_manpower_workspace() -> dict:
    """Return the parsed module dataset with cache invalidation on workbook change."""
    path = _seed_path()
    if not path.exists():
        return _empty_workspace(path, "Seed workbook not found.")
    return _get_manpower_workspace_cached(str(path), path.stat().st_mtime_ns)


def get_manpower_location(location_slug: str) -> dict | None:
    """Return one parsed location bucket from the workbook-backed dataset."""
    workspace = get_manpower_workspace()
    for location in workspace["locations"]:
        if location["slug"] == location_slug:
            return location
    return None
