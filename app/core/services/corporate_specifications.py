"""Catalogue, detail, editing and versioning for Corporate Specifications Management.

Every specification held in the system is current: there is no published/unpublished
split. The chemical list is the corporate specification register that QC Laboratory
Monitoring maintains and Inventory Monitoring already reads (``qc_testing_standards``),
unioned with any specification record that is not on that register yet. Parameters,
narrative sections and version history come from the CSC specification tables.
"""

from __future__ import annotations

import json
import logging
import re
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.services.csc_utils import (
    IMPACT_CHECKLIST_FLAGS,
    SPEC_SUBSET_LABELS,
    SPEC_SUBSET_ORDER,
    VALID_FLAG_VALUES,
    build_impact_legacy_payload,
    deserialize_impact_checklist_state,
    format_required_value,
    format_spec_version,
    increment_spec_version,
    normalize_spec_version,
    sanitize_multiline_text,
    sanitize_text,
    sequence_sort_key,
    summarize_impact_checklist_state,
)
from app.extensions import db
from app.models.csc import (
    CSCDraft, CSCImpactAnalysis, CSCIssueFlag, CSCParameter, CSCSection, CSCSpecVersion,
)

UNCATEGORISED_KEY = "unspecified"
UNCATEGORISED_LABEL = "Not on the corporate specification register"
TEST_PROCEDURE_SECTION = "Meta::test_procedure"
MATERIAL_CODE_SECTION = "Meta::material_code"
# Narrative sections shown on the specification page, in display order.
NARRATIVE_SECTIONS = (
    ("background", "Background"),
    ("existing_spec", "Existing Specification"),
    ("proposed_changes", "Proposed Changes"),
    ("justification", "Justification"),
    ("recommendation", "Recommendation"),
)
# Sections written by the retired committee workflow; never shown or edited here.
INTERNAL_SECTION_PREFIXES = ("__workflow", "Meta::")
PARAMETER_TYPES = ("Vital", "Essential", "Desirable")
# Material-master identity fields; the rest of the master row is covered by the
# material-property and storage field definitions in csc_master_data.
IDENTITY_FIELDS = (
    ("short_text", "Material Short Text"),
    ("group", "Group"),
    ("material_type", "Type"),
    ("centralization", "Centralization"),
)
ISSUE_FLAG_LABELS = (
    ("operational", "Operational"),
    ("quality", "Quality"),
    ("supply", "Supply"),
    ("testing", "Testing"),
)
IMPACT_FLAG_VALUES = tuple(VALID_FLAG_VALUES)
REVISION_ACTION = "SPECIFICATION_REVISED"
logger = logging.getLogger(__name__)
CATEGORY_ICONS = {
    "DFC": "bi-droplet-half", "CCA": "bi-bricks", "WCF": "bi-moisture", "WS": "bi-lightning-charge",
    "PC": "bi-funnel", "WIC": "bi-water", "WM": "bi-tools", "UTL": "bi-gear-wide-connected", "LPG": "bi-fire",
}


def category_of(specification_no: Any) -> str | None:
    """The category segment of a specification number, e.g. DFC in ONGC / DFC / 01 / 2026."""
    parts = [part.strip().upper() for part in str(specification_no or "").split("/")]
    if len(parts) < 2 or not re.fullmatch(r"[A-Z]{2,6}", parts[1]):
        return None
    return parts[1]


def category_label(code: str | None) -> str:
    if not code or code == UNCATEGORISED_KEY:
        return UNCATEGORISED_LABEL
    return SPEC_SUBSET_LABELS.get(code, f"{code} chemicals")


def _sequence_of(specification_no: Any) -> tuple:
    """Sort key for the serial within a category, e.g. 11A in ONGC / PC / 11A / 2015."""
    parts = [part.strip() for part in str(specification_no or "").split("/")]
    return sequence_sort_key(parts[2] if len(parts) > 2 else "")


def normalize_spec_number(value: Any) -> str:
    """Spec numbers are written both as ONGC/DFC/01/2026 and ONGC / DFC / 01 / 2026."""
    return re.sub(r"\s*/\s*", "/", str(value or "").strip().upper())


def _material_code(value: Any) -> str:
    text = str(value or "").strip()
    return text if text.isdigit() else ""


def requirement_of(parameter: CSCParameter) -> str:
    """The current required value, however the row happens to store it."""
    return format_required_value(
        parameter.required_value_type or "text",
        parameter.required_value_text or parameter.existing_value or "",
        parameter.required_value_operator_1,
        parameter.required_value_value_1,
        parameter.required_value_operator_2,
        parameter.required_value_value_2,
    ) or (parameter.existing_value or "")


# ── Catalogue ────────────────────────────────────────────────────────────────


def _register_rows() -> list[Any]:
    from app.models.quality_control.qc_testing_standard import QCTestingStandard

    return QCTestingStandard.query.all()


def _specification_records() -> list[CSCDraft]:
    """Every root specification record; revision children are workflow residue."""
    return CSCDraft.query.filter(CSCDraft.parent_draft_id.is_(None)).all()


def _parameter_counts() -> dict[int, int]:
    """Parameter counts for every specification record in one query."""
    return dict(
        db.session.query(CSCParameter.draft_id, db.func.count(CSCParameter.id))
        .group_by(CSCParameter.draft_id)
        .all()
    )


def _entry(
    *,
    ref: str,
    chemical_name: str,
    specification_no: str,
    material_code: str,
    record: CSCDraft | None,
    on_register: bool,
    parameter_count: int = 0,
    standard_days: Any = None,
    remarks: Any = None,
) -> dict[str, Any]:
    spec_number = specification_no or (record.spec_number if record else "")
    code = category_of(spec_number)
    return {
        "ref": ref,
        "chemical_name": chemical_name or (record.chemical_name if record else "") or "—",
        "spec_number": spec_number or "—",
        "material_code": material_code or (record.material_code or "" if record else ""),
        "category": code or UNCATEGORISED_KEY,
        "category_label": category_label(code),
        "sequence": _sequence_of(spec_number),
        "record": record,
        "record_id": record.id if record is not None else None,
        "parameter_count": parameter_count,
        "has_parameters": parameter_count > 0,
        "version": format_spec_version(record.spec_version) if record is not None else None,
        "updated_at": record.updated_at if record is not None else None,
        "standard_days": standard_days,
        "remarks": str(remarks or "").strip(),
        "on_register": on_register,
    }


def catalogue() -> list[dict[str, Any]]:
    """One entry per chemical: the register unioned with off-register specification records.

    A specification covering several chemicals is matched to each of them, so the same
    record can back more than one entry — editing it updates the specification once.
    """
    records = _specification_records()
    counts = _parameter_counts()
    by_code: dict[str, CSCDraft] = {}
    by_spec: dict[str, CSCDraft] = {}
    for record in records:
        code = _material_code(record.material_code)
        if code:
            by_code.setdefault(code, record)
        spec = normalize_spec_number(record.spec_number)
        if spec:
            by_spec.setdefault(spec, record)

    entries: list[dict[str, Any]] = []
    matched: set[int] = set()
    for row in _register_rows():
        code = _material_code(row.material_code)
        record = by_code.get(code) or by_spec.get(normalize_spec_number(row.specification_no))
        if record is not None:
            matched.add(record.id)
        entries.append(
            _entry(
                ref=f"r-{row.id}",
                chemical_name=row.chemical_name or "",
                specification_no=row.specification_no or "",
                material_code=code,
                record=record,
                on_register=True,
                parameter_count=counts.get(record.id, 0) if record is not None else 0,
                standard_days=row.standard_days,
                remarks=row.remarks,
            )
        )

    for record in records:
        if record.id in matched:
            continue
        entries.append(
            _entry(
                ref=f"s-{record.id}",
                chemical_name=record.chemical_name or "",
                specification_no=record.spec_number or "",
                material_code=_material_code(record.material_code),
                record=record,
                on_register=False,
                parameter_count=counts.get(record.id, 0),
            )
        )

    entries.sort(key=_entry_sort_key)
    return entries


def _entry_sort_key(entry: dict[str, Any]) -> tuple:
    category = entry["category"]
    rank = SPEC_SUBSET_ORDER.index(category) if category in SPEC_SUBSET_ORDER else len(SPEC_SUBSET_ORDER)
    return rank, entry["sequence"], entry["chemical_name"].casefold()


def resolve(ref: str, entries: list[dict[str, Any]] | None = None) -> dict[str, Any] | None:
    """Look one catalogue entry up by its stable reference."""
    for entry in catalogue() if entries is None else entries:
        if entry["ref"] == ref:
            return entry
    return None


def category_tiles(entries: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    entries = catalogue() if entries is None else entries
    tiles: dict[str, dict[str, Any]] = {}
    for entry in entries:
        tile = tiles.setdefault(
            entry["category"],
            {
                "key": entry["category"],
                "label": entry["category_label"],
                "icon": CATEGORY_ICONS.get(entry["category"], "bi-clipboard-data"),
                "chemicals": 0,
                "specified": 0,
                "awaiting": 0,
                "is_unspecified": entry["category"] == UNCATEGORISED_KEY,
            },
        )
        tile["chemicals"] += 1
        tile["specified" if entry["has_parameters"] else "awaiting"] += 1
    ordered = sorted(
        tiles.values(),
        key=lambda tile: (
            SPEC_SUBSET_ORDER.index(tile["key"]) if tile["key"] in SPEC_SUBSET_ORDER else 99,
            tile["key"],
        ),
    )
    largest = max((tile["chemicals"] for tile in ordered), default=0)
    for tile in ordered:
        tile["share"] = round(tile["chemicals"] / largest * 100) if largest else 0
    return ordered


def landing_data() -> dict[str, Any]:
    entries = catalogue()
    counted: dict[int, int] = {}
    for entry in entries:
        if entry["record_id"]:
            counted[entry["record_id"]] = entry["parameter_count"]
    return {
        "tiles": category_tiles(entries),
        "chemical_total": len(entries),
        "specification_total": len(counted),
        "parameter_total": sum(counted.values()),
        "awaiting_total": sum(1 for entry in entries if not entry["has_parameters"]),
        "off_register_total": sum(1 for entry in entries if not entry["on_register"]),
        "last_revised_at": max(
            (entry["updated_at"] for entry in entries if entry["updated_at"]), default=None
        ),
    }


def category_data(key: str, query: str = "") -> dict[str, Any]:
    entries = [entry for entry in catalogue() if entry["category"] == key]
    if not entries:
        return {}
    needle = query.strip().casefold()
    if needle:
        entries = [
            entry
            for entry in entries
            if needle in entry["chemical_name"].casefold()
            or needle in entry["spec_number"].casefold()
            or needle in entry["material_code"]
        ]
    return {
        "category_key": key,
        "category_label": category_label(None if key == UNCATEGORISED_KEY else key),
        "category_icon": CATEGORY_ICONS.get(key, "bi-clipboard-data"),
        "entries": entries,
        "query": query,
        "specified_count": sum(1 for entry in entries if entry["has_parameters"]),
    }


# ── Specification detail ─────────────────────────────────────────────────────


def _sections(record: CSCDraft) -> dict[str, str]:
    return {
        section.section_name: section.section_text or ""
        for section in record.sections.order_by(CSCSection.sort_order).all()
    }


def merged_sections(record: CSCDraft, child: CSCDraft | None = None) -> dict[str, str]:
    """Specification sections, filled in from the revision draft where blank.

    The narrative the committee wrote — existing specification, proposed changes,
    justification — only ever landed on the revision draft, so it is read from there
    when the specification record has nothing of its own.
    """
    sections = _sections(record)
    child = revision_child(record) if child is None else child
    if child is None:
        return sections
    for name, text in _sections(child).items():
        if str(text or "").strip() and not str(sections.get(name) or "").strip():
            sections[name] = text
    return sections


def parameter_rows(record: CSCDraft) -> list[dict[str, Any]]:
    return [
        {
            "id": parameter.id,
            "parameter_name": parameter.parameter_name or "",
            "parameter_type": (parameter.parameter_type or "").strip(),
            "requirement": requirement_of(parameter),
            "unit_of_measure": (parameter.unit_of_measure or "").strip(),
            "conditions": (parameter.parameter_conditions or "").strip(),
            "test_method": (parameter.test_procedure_text or parameter.test_method or "").strip(),
            "remarks": (parameter.remarks or "").strip(),
            "sort_order": parameter.sort_order,
        }
        for parameter in record.parameters.order_by(CSCParameter.sort_order, CSCParameter.id).all()
    ]


def version_history(record: CSCDraft) -> list[dict[str, Any]]:
    rows = []
    for snapshot in record.spec_versions.order_by(
        CSCSpecVersion.spec_version.desc(), CSCSpecVersion.created_at.desc()
    ).all():
        rows.append(
            {
                "version": format_spec_version(snapshot.spec_version),
                "created_at": snapshot.created_at,
                "created_by": snapshot.created_by,
                "action": snapshot.source_action,
                "reason": snapshot.remarks or "",
                "is_current": snapshot.spec_version == normalize_spec_version(record.spec_version),
            }
        )
    return rows


# ── Supporting data: material master, properties, storage, impact, issues ────


def _display(value: Any) -> str:
    text = str(value if value is not None else "").strip()
    return text


def revision_child(record: CSCDraft) -> CSCDraft | None:
    """The specification's revision draft, if one was ever opened.

    The retired workflow left one draft per specification. Its content was never
    published, but it is where the committee actually recorded the material
    properties, storage detail and impact assessment, so it is read here as a
    fallback wherever the specification record itself is blank.
    """
    if not record.id:
        return None
    return (
        CSCDraft.query.filter_by(parent_draft_id=record.id)
        .order_by(CSCDraft.updated_at.desc(), CSCDraft.id.desc())
        .first()
    )


def _staged_master_values(child: CSCDraft | None) -> dict[str, str]:
    """The master-data payload the revision draft staged but never committed."""
    if child is None:
        return {}
    from app.core.services.csc_master_data import _load_staged_master_payload

    try:
        return _load_staged_master_payload(child)
    except Exception:  # noqa: BLE001 - a malformed payload must not break the page
        logger.warning("Could not read staged master data for draft %s", child.id, exc_info=True)
        return {}


def _is_placeholder(field: str, value: Any) -> bool:
    """True when a value carries no real information and should yield to the draft."""
    from app.core.services.csc_master_data import STORAGE_CONDITIONS_GENERAL_DEFAULT

    text = str(value or "").strip()
    if not text:
        return True
    # get_master_form_values injects this boilerplate whenever the column is empty,
    # so it must not out-rank a value the committee actually recorded.
    return field == "storage_conditions_general" and text == STORAGE_CONDITIONS_GENERAL_DEFAULT


def master_values(record: CSCDraft, child: CSCDraft | None = None) -> dict[str, str]:
    """The material-master values, filled in from the revision draft where blank.

    Committed master-master data always wins; the draft only fills gaps.
    """
    from app.core.services.csc_master_data import get_master_form_values

    values = get_master_form_values(record)
    staged = _staged_master_values(revision_child(record) if child is None else child)
    for field, staged_value in staged.items():
        if str(staged_value or "").strip() and _is_placeholder(field, values.get(field)):
            values[field] = staged_value
    return values


def identity_rows(record: CSCDraft, values: dict[str, str] | None = None) -> list[dict[str, str]]:
    values = master_values(record) if values is None else values
    rows = [{"label": "Material Code", "value": _display(record.material_code)}]
    rows += [
        {"label": label, "value": _display(values.get(field))}
        for field, label in IDENTITY_FIELDS
    ]
    return [row for row in rows if row["value"]]


def _field_rows(definitions, values: dict[str, str]) -> list[dict[str, Any]]:
    """Field definitions carrying their current value, ready for display or editing."""
    rows = []
    for field in definitions:
        name = str(field["field_name"])
        rows.append(
            {
                "field_name": name,
                "label": str(field["label"]),
                "section": str(field.get("section") or ""),
                "dimension": str(field.get("dimension") or ""),
                "input_type": str(field.get("input_type") or "options"),
                "options": list(field.get("options") or []),
                "placeholder": str(field.get("placeholder") or ""),
                "value": _display(values.get(name)),
            }
        )
    return rows


def identity_fields(record: CSCDraft, values: dict[str, str] | None = None) -> list[dict[str, Any]]:
    """Editable material-master identity fields, shaped like the other field blocks."""
    from app.core.services.csc_master_data import (
        MASTER_CENTRALIZATION_OPTIONS, MASTER_GROUP_OPTIONS, MASTER_TYPE_OPTIONS,
    )

    options = {
        "group": MASTER_GROUP_OPTIONS,
        "material_type": MASTER_TYPE_OPTIONS,
        "centralization": MASTER_CENTRALIZATION_OPTIONS,
    }
    definitions = [
        {
            "field_name": field,
            "label": label,
            "section": "Material Identity",
            "dimension": "",
            "input_type": "options" if field in options else "text",
            "options": options.get(field, []),
        }
        for field, label in IDENTITY_FIELDS
    ]
    return _field_rows(definitions, master_values(record) if values is None else values)


def material_property_fields(record: CSCDraft, values: dict[str, str] | None = None) -> list[dict[str, Any]]:
    from app.core.services.csc_master_data import get_material_properties_fields

    return _field_rows(get_material_properties_fields(), master_values(record) if values is None else values)


def storage_fields(record: CSCDraft, values: dict[str, str] | None = None) -> list[dict[str, Any]]:
    from app.core.services.csc_master_data import get_storage_handling_fields

    return _field_rows(get_storage_handling_fields(), master_values(record) if values is None else values)


def group_by_section(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Group field rows under their section heading, preserving definition order."""
    groups: list[dict[str, Any]] = []
    index: dict[str, dict[str, Any]] = {}
    for row in rows:
        section = row["section"] or "Other"
        group = index.get(section)
        if group is None:
            group = {"section": section, "rows": []}
            index[section] = group
            groups.append(group)
        group["rows"].append(row)
    return groups


def impact_view(record: CSCDraft) -> dict[str, Any]:
    """The impact checklist for this specification, answered or awaiting assessment."""
    # Queried rather than read off the relationship so a checklist written earlier in
    # this transaction is seen by the post-save snapshot.
    analysis = (
        CSCImpactAnalysis.query.filter_by(draft_id=record.id).first() if record.id else None
    )
    from_revision = False
    if analysis is None or not (analysis.checklist_state_json or "").strip():
        child = revision_child(record)
        if child is not None:
            child_analysis = CSCImpactAnalysis.query.filter_by(draft_id=child.id).first()
            if child_analysis is not None and (child_analysis.checklist_state_json or "").strip():
                analysis, from_revision = child_analysis, True
    state = deserialize_impact_checklist_state(
        analysis.checklist_state_json if analysis is not None else None,
        record.chemical_name,
    )
    summary = summarize_impact_checklist_state(state)
    flags = [
        {
            "id": flag["id"],
            "order": flag["order"],
            "dimension": flag["dimension"],
            "type": flag["type"],
            "section_label": flag["section_label"],
            "question": flag["question"],
            "detail": flag["detail"],
            "source": flag["source"],
            "value": flag["answer"],
        }
        for flag in sorted(summary["flags"], key=lambda flag: flag["order"])
    ]
    return {
        "recorded": analysis is not None and bool((analysis.checklist_state_json or "").strip()),
        "from_revision": from_revision,
        "classification": summary["classification"],
        "grade": summary["grade"],
        "confidence": summary.get("confidence", ""),
        "rule": summary["rule"],
        "tone": summary.get("tone", "grey"),
        "red_yes_count": summary["red_yes_count"],
        "amber_yes_count": summary["amber_yes_count"],
        "answered_count": summary["answered_count"],
        "total_flags": len(IMPACT_CHECKLIST_FLAGS),
        "flags": flags,
        "red_flags": [flag for flag in flags if flag["type"] == "RED"],
        "amber_flags": [flag for flag in flags if flag["type"] == "AMBER"],
    }


def issue_flag_rows(record: CSCDraft) -> list[dict[str, Any]]:
    """The four-issue register, defaulting to unflagged for specifications that have none."""
    stored = {flag.issue_type: flag for flag in record.issue_flags.all()}
    child = revision_child(record)
    if child is not None:
        for flag in child.issue_flags.all():
            existing = stored.get(flag.issue_type)
            if existing is None or (not existing.is_present and not (existing.note or "").strip()):
                stored[flag.issue_type] = flag
    rows = []
    for issue_type, label in ISSUE_FLAG_LABELS:
        flag = stored.get(issue_type)
        rows.append(
            {
                "issue_type": issue_type,
                "label": label,
                "is_present": bool(flag.is_present) if flag is not None else False,
                "note": _display(flag.note) if flag is not None else "",
            }
        )
    return rows


def specification_data(ref: str) -> dict[str, Any] | None:
    entries = catalogue()
    entry = resolve(ref, entries)
    if entry is None:
        return None
    record = entry["record"]
    child = revision_child(record) if record is not None else None
    sections = merged_sections(record, child) if record is not None else {}
    values = master_values(record, child) if record is not None else {}
    return {
        "entry": entry,
        "record": record,
        "parameters": parameter_rows(record) if record is not None else [],
        "test_procedure": sections.get(TEST_PROCEDURE_SECTION, "").strip(),
        "covered_materials": sections.get(MATERIAL_CODE_SECTION, "").strip(),
        "narratives": [
            {"key": key, "label": label, "text": sections.get(key, "").strip()}
            for key, label in NARRATIVE_SECTIONS
        ],
        "versions": version_history(record) if record is not None else [],
        "shared_with": _shared_chemicals(entry, entries),
        "revision_draft": child,
        "identity": identity_rows(record, values) if record is not None else [],
        "identity_fields": identity_fields(record, values) if record is not None else [],
        "material_properties": material_property_fields(record, values) if record is not None else [],
        "storage": storage_fields(record, values) if record is not None else [],
        "impact": impact_view(record) if record is not None else None,
        "issues": issue_flag_rows(record) if record is not None else [],
    }


def _shared_chemicals(entry: dict[str, Any], entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Other register chemicals covered by the same specification record."""
    if not entry["record_id"]:
        return []
    return [
        other
        for other in entries
        if other["record_id"] == entry["record_id"] and other["ref"] != entry["ref"]
    ]


# ── Editing and versioning ───────────────────────────────────────────────────


class SpecificationError(Exception):
    """Raised when a specification edit cannot be applied."""


def create_record(entry: dict[str, Any], username: str) -> CSCDraft:
    """Open a specification record for a register chemical that has none yet."""
    if entry["record"] is not None:
        return entry["record"]
    record = CSCDraft(
        spec_number=entry["spec_number"] if entry["spec_number"] != "—" else "",
        chemical_name=entry["chemical_name"] if entry["chemical_name"] != "—" else "",
        material_code=entry["material_code"] or None,
        status="Published",
        admin_stage="published",
        is_admin_draft=True,
        prepared_by=username,
        spec_version=0,
    )
    db.session.add(record)
    db.session.flush()
    return record


def _snapshot(record: CSCDraft) -> dict[str, Any]:
    """The whole specification — everything the dossier prints — at this moment."""
    values = master_values(record)
    impact = impact_view(record)
    return {
        "spec_number": record.spec_number or "",
        "chemical_name": record.chemical_name or "",
        "material_code": record.material_code or "",
        "parameters": parameter_rows(record),
        "sections": {
            name: text
            for name, text in merged_sections(record).items()
            if not name.startswith(INTERNAL_SECTION_PREFIXES) or name == TEST_PROCEDURE_SECTION
        },
        "identity": {field: _display(values.get(field)) for field, _label in IDENTITY_FIELDS},
        "material_properties": {row["field_name"]: row["value"] for row in material_property_fields(record, values)},
        "storage": {row["field_name"]: row["value"] for row in storage_fields(record, values)},
        "impact": {flag["id"]: flag["value"] for flag in impact["flags"]},
        "impact_classification": impact["classification"],
        "issues": {row["issue_type"]: {"is_present": row["is_present"], "note": row["note"]} for row in issue_flag_rows(record)},
    }


_COMPARED_PARAMETER_KEYS = (
    "parameter_name", "parameter_type", "requirement", "unit_of_measure",
    "conditions", "test_method", "remarks",
)


def _comparable(snapshot: dict[str, Any]) -> str:
    """Snapshot reduced to the substance that decides whether a version is warranted."""
    return json.dumps(
        {
            "spec_number": snapshot["spec_number"].strip(),
            "chemical_name": snapshot["chemical_name"].strip(),
            "material_code": snapshot["material_code"].strip(),
            "sections": {name: text.strip() for name, text in sorted(snapshot["sections"].items())},
            "parameters": [
                {key: str(parameter.get(key) or "").strip() for key in _COMPARED_PARAMETER_KEYS}
                for parameter in snapshot["parameters"]
            ],
            "identity": {key: value.strip() for key, value in sorted(snapshot["identity"].items())},
            "material_properties": {key: value.strip() for key, value in sorted(snapshot["material_properties"].items())},
            "storage": {key: value.strip() for key, value in sorted(snapshot["storage"].items())},
            "impact": dict(sorted(snapshot["impact"].items())),
            "issues": {
                key: {"is_present": bool(value["is_present"]), "note": str(value["note"]).strip()}
                for key, value in sorted(snapshot["issues"].items())
            },
        },
        sort_keys=True,
        ensure_ascii=False,
    )


def _write_section(record: CSCDraft, name: str, text: str, sort_order: int) -> None:
    section = record.sections.filter_by(section_name=name).first()
    if section is None:
        section = CSCSection(draft_id=record.id, section_name=name, sort_order=sort_order)
        db.session.add(section)
    section.section_text = text


def parameter_payload(form: Any) -> list[dict[str, str]]:
    """Read the editor grid off the submitted form, dropping wholly blank rows."""
    rows: list[dict[str, str]] = []
    for index in form.getlist("parameter_index"):
        row = {
            "parameter_name": sanitize_multiline_text(form.get(f"parameter_name_{index}", ""), 2000),
            "parameter_type": sanitize_text(form.get(f"parameter_type_{index}", ""), 50),
            "requirement": sanitize_multiline_text(form.get(f"requirement_{index}", ""), 2000),
            "unit_of_measure": sanitize_text(form.get(f"unit_of_measure_{index}", ""), 100),
            "conditions": sanitize_multiline_text(form.get(f"conditions_{index}", ""), 2000),
            "test_method": sanitize_multiline_text(form.get(f"test_method_{index}", ""), 2000),
            "remarks": sanitize_multiline_text(form.get(f"remarks_{index}", ""), 2000),
        }
        if any(value.strip() for value in row.values()):
            rows.append(row)
    return rows


def supporting_payload(form: Any) -> dict[str, Any]:
    """Read the material-master, impact and issue blocks off the submitted form."""
    from app.core.services.csc_master_data import (
        get_material_properties_fields, get_storage_handling_fields,
    )

    values: dict[str, str] = {}
    for field, _label in IDENTITY_FIELDS:
        values[field] = sanitize_text(form.get(f"master_{field}", ""), 255)
    for field in list(get_material_properties_fields()) + list(get_storage_handling_fields()):
        name = str(field["field_name"])
        limit = 2000 if field.get("input_type") == "textarea" else 255
        values[name] = sanitize_multiline_text(form.get(f"master_{name}", ""), limit)

    impact = {}
    for flag in IMPACT_CHECKLIST_FLAGS:
        answer = sanitize_text(form.get(f"impact_{flag['id']}", ""), 20).upper()
        impact[flag["id"]] = answer if answer in IMPACT_FLAG_VALUES else "REVIEW"

    issues = {
        issue_type: {
            "is_present": form.get(f"issue_{issue_type}_present") in {"1", "on", "true", "yes"},
            "note": sanitize_multiline_text(form.get(f"issue_{issue_type}_note", ""), 2000),
        }
        for issue_type, _label in ISSUE_FLAG_LABELS
    }
    return {"master": values, "impact": impact, "issues": issues}


def _apply_master_values(record: CSCDraft, values: dict[str, str], user_id: int | None) -> None:
    """Write the material-master row behind this specification."""
    from app.core.services.csc_master_data import upsert_master_record_from_form

    payload = {"material_code": record.material_code or "", "chemical_name": record.chemical_name or "", **values}
    # Density only applies to liquids; the old editor cleared it the same way.
    if _display(values.get("physical_state")) != "Liquid":
        payload["extra__Density"] = ""
    master_record = upsert_master_record_from_form(record, payload, user_id=user_id)
    if master_record is not None and master_record not in db.session:
        db.session.add(master_record)


def _apply_impact(record: CSCDraft, answers: dict[str, str], username: str, user_id: int | None) -> None:
    """Write the impact checklist and mirror its outcome onto the material master."""
    from app.core.services.csc_master_data import sync_impact_fields_to_master_record

    now = datetime.now(timezone.utc).isoformat()
    state = deserialize_impact_checklist_state(
        {
            "version": 3,
            "chemical_name": record.chemical_name or "",
            "flags": {
                flag_id: {"value": value, "source_type": "MODULE_ADMIN", "answered_by": username, "answered_on": now}
                for flag_id, value in answers.items()
            },
        },
        record.chemical_name,
    )
    payload = build_impact_legacy_payload(state)

    analysis = record.impact_analysis
    if analysis is None:
        analysis = CSCImpactAnalysis(draft_id=record.id)
        db.session.add(analysis)
    analysis.operational_impact_score = payload["operational_impact_score"]
    analysis.safety_environment_score = payload["safety_environment_score"]
    analysis.supply_risk_score = payload["supply_risk_score"]
    analysis.no_substitute_flag = payload["no_substitute_flag"]
    analysis.impact_score_total = payload["impact_score_total"]
    analysis.impact_grade = payload["impact_grade"]
    analysis.checklist_state_json = payload["checklist_state_json"]
    analysis.operational_note = payload["operational_note"]
    analysis.safety_environment_note = payload["safety_environment_note"]
    analysis.supply_risk_note = payload["supply_risk_note"]
    analysis.reviewed_by = username

    master_record = sync_impact_fields_to_master_record(record, state, user_id=user_id)
    if master_record is not None and master_record not in db.session:
        db.session.add(master_record)


def _apply_issue_flags(record: CSCDraft, issues: dict[str, dict[str, Any]]) -> None:
    stored = {flag.issue_type: flag for flag in record.issue_flags.all()}
    for order, (issue_type, _label) in enumerate(ISSUE_FLAG_LABELS):
        submitted = issues.get(issue_type) or {"is_present": False, "note": ""}
        flag = stored.get(issue_type)
        if flag is None:
            flag = CSCIssueFlag(draft_id=record.id, issue_type=issue_type)
            db.session.add(flag)
        flag.is_present = bool(submitted["is_present"])
        flag.note = str(submitted["note"]).strip()
        flag.sort_order = order


def save_specification(
    record: CSCDraft,
    *,
    chemical_name: str,
    spec_number: str,
    material_code: str,
    test_procedure: str,
    narratives: dict[str, str],
    parameters: list[dict[str, str]],
    supporting: dict[str, Any],
    reason: str,
    username: str,
    user_id: int | None = None,
) -> tuple[bool, str]:
    """Apply an admin edit, versioning the specification only when something changed.

    ``supporting`` carries the material-master, impact-checklist and issue-register
    blocks from :func:`supporting_payload`. Returns ``(version_created,
    version_display)``. The caller commits.
    """
    if not chemical_name.strip():
        raise SpecificationError("A chemical name is required.")
    if not spec_number.strip():
        raise SpecificationError("A specification number is required.")
    if not parameters:
        raise SpecificationError("A specification needs at least one parameter.")
    for position, row in enumerate(parameters, start=1):
        if not row["parameter_name"].strip():
            raise SpecificationError(f"Parameter {position} has no name.")
        if not row["requirement"].strip():
            raise SpecificationError(f"Parameter {position} has no required value.")

    before = _snapshot(record)

    record.chemical_name = chemical_name.strip()
    record.spec_number = spec_number.strip()
    record.material_code = material_code.strip() or None

    existing = record.parameters.order_by(CSCParameter.sort_order, CSCParameter.id).all()
    for position, row in enumerate(parameters):
        parameter = existing[position] if position < len(existing) else CSCParameter(draft_id=record.id)
        if position >= len(existing):
            db.session.add(parameter)
        parameter.parameter_name = row["parameter_name"].strip()
        parameter.parameter_type = row["parameter_type"].strip() or "Essential"
        parameter.unit_of_measure = row["unit_of_measure"].strip()
        parameter.parameter_conditions = row["conditions"].strip()
        parameter.test_procedure_text = row["test_method"].strip()
        parameter.remarks = row["remarks"].strip()
        parameter.sort_order = position
        # Requirement is kept in both columns so exports written against either stay correct.
        parameter.required_value_type = "text"
        parameter.required_value_text = row["requirement"].strip()
        parameter.existing_value = row["requirement"].strip()
    for parameter in existing[len(parameters):]:
        db.session.delete(parameter)

    _write_section(record, TEST_PROCEDURE_SECTION, test_procedure.strip(), 90)
    for order, (key, _label) in enumerate(NARRATIVE_SECTIONS):
        _write_section(record, key, narratives.get(key, "").strip(), order)

    _apply_master_values(record, supporting.get("master") or {}, user_id)
    _apply_impact(record, supporting.get("impact") or {}, username, user_id)
    _apply_issue_flags(record, supporting.get("issues") or {})

    db.session.flush()
    after = _snapshot(record)
    if _comparable(before) == _comparable(after):
        return False, format_spec_version(record.spec_version)

    if not reason.strip():
        raise SpecificationError("Describe the reason for this revision before saving.")

    record.spec_version = increment_spec_version(record.spec_version, "whole")
    record.status = "Published"
    record.admin_stage = "published"
    db.session.add(
        CSCSpecVersion(
            draft_id=record.id,
            spec_version=record.spec_version,
            created_by=username,
            source_action="SPECIFICATION_REVISED",
            remarks=reason.strip(),
            payload_json=json.dumps({"before": before, "after": after}, default=str, ensure_ascii=False),
        )
    )
    return True, format_spec_version(record.spec_version)


# ── Dossier ──────────────────────────────────────────────────────────────────


def previous_snapshot(record: CSCDraft) -> dict[str, Any] | None:
    """The specification as it stood before the most recent revision, if there was one."""
    snapshot = (
        record.spec_versions.filter_by(source_action=REVISION_ACTION)
        .order_by(CSCSpecVersion.spec_version.desc(), CSCSpecVersion.created_at.desc())
        .first()
    )
    if snapshot is None:
        return None
    try:
        payload = json.loads(snapshot.payload_json or "{}")
    except json.JSONDecodeError:
        return None
    before = payload.get("before")
    return before if isinstance(before, dict) else None


def _comparison_rows(current: list[dict[str, str]], baseline: dict[str, str] | None) -> list[dict[str, str]]:
    """Label/value rows carrying the previous version alongside, for the dossier tables."""
    rows = []
    for row in current:
        was = _display(baseline.get(row["label"])) if baseline is not None else ""
        rows.append(
            {
                "label": row["label"],
                "value": row["value"] or "—",
                "source_value": (was or "—") if baseline is not None else "—",
                "change_status": _change_status(was, row["value"], baseline is not None),
            }
        )
    return rows


def _change_status(before: str, after: str, has_baseline: bool) -> str:
    if not has_baseline:
        return "Retained"
    if before.strip() == after.strip():
        return "Retained"
    return "Added" if not before.strip() else "Revised"


def _keyed_rows(pairs: list[tuple[str, str]]) -> list[dict[str, str]]:
    return [{"label": label, "value": _display(value)} for label, value in pairs]


def dossier_context(data: dict[str, Any]) -> dict[str, Any]:
    """Assemble the full review context the dossier document is built from.

    The comparison columns show the specification as it stood before its most recent
    revision, so the dossier reads the way it always did — baseline beside current.
    """
    record = data["record"]
    entry = data["entry"]
    baseline = previous_snapshot(record)
    impact = data["impact"]

    def baseline_map(key: str, labels: dict[str, str]) -> dict[str, str] | None:
        if baseline is None or key not in baseline:
            return None
        stored = baseline[key] or {}
        return {labels[name]: value for name, value in stored.items() if name in labels}

    property_labels = {row["field_name"]: row["label"] for row in data["material_properties"]}
    storage_labels = {row["field_name"]: row["label"] for row in data["storage"]}
    identity_labels = {field: label for field, label in IDENTITY_FIELDS}

    baseline_parameters = (baseline or {}).get("parameters") or []
    parameter_baseline = {row.get("parameter_name", ""): row for row in baseline_parameters}
    parameter_rows_out = []
    for parameter in data["parameters"]:
        was = parameter_baseline.get(parameter["parameter_name"], {})
        parameter_rows_out.append(
            {
                "parameter_name": parameter["parameter_name"] or "Untitled Parameter",
                "parameter_type": parameter["parameter_type"] or "—",
                "unit_of_measure": parameter["unit_of_measure"] or "—",
                "final_requirement": parameter["requirement"] or "—",
                "conditions": parameter["conditions"] or "—",
                "test_procedure_type": parameter["test_method"] or "—",
                "procedure_text": parameter["test_method"] or "—",
                "required_value": _display(was.get("requirement")) or "—",
                "source_parameter_type": _display(was.get("parameter_type")) or "—",
                "source_unit_of_measure": _display(was.get("unit_of_measure")) or "—",
                "source_conditions": _display(was.get("conditions")) or "—",
                "source_test_procedure_type": _display(was.get("test_method")) or "—",
                "source_test_method": _display(was.get("test_method")) or "—",
                "change_status": _change_status(
                    _display(was.get("requirement")), parameter["requirement"], bool(parameter_baseline)
                ),
            }
        )

    baseline_sections = (baseline or {}).get("sections") or {}
    section_rows = []
    for key, label in NARRATIVE_SECTIONS:
        text = next((row["text"] for row in data["narratives"] if row["key"] == key), "")
        was = _display(baseline_sections.get(key))
        section_rows.append(
            {
                "label": label,
                "text": text or "—",
                "source_text": (was or "—") if baseline is not None else "—",
                "change_status": _change_status(was, text, baseline is not None),
            }
        )

    impact_rows = _keyed_rows(
        [
            ("Impact Classification", f"{impact['classification']} ({impact['confidence']})" if impact["confidence"] else impact["classification"]),
            ("Decision Rule", impact["rule"]),
            ("Red YES Count", f"{impact['red_yes_count']}/5"),
            ("Amber YES Count", f"{impact['amber_yes_count']}/5"),
            ("Flags Answered", f"{impact['answered_count']}/{impact['total_flags']}"),
        ]
        + [(flag["dimension"], flag["value"]) for flag in impact["flags"]]
    )
    baseline_impact = None
    if baseline is not None and "impact" in baseline:
        dimensions = {flag["id"]: flag["dimension"] for flag in impact["flags"]}
        baseline_impact = {
            dimensions[flag_id]: value
            for flag_id, value in (baseline["impact"] or {}).items()
            if flag_id in dimensions
        }
        baseline_impact["Impact Classification"] = _display(baseline.get("impact_classification"))

    return {
        "draft": {
            "spec_number": record.spec_number or "—",
            "chemical_name": record.chemical_name or "—",
            "material_code": data["covered_materials"] or record.material_code or "—",
            "test_procedure": data["test_procedure"],
            "version_display": f"v{entry['version']}",
        },
        "summary_rows": _keyed_rows(
            [
                ("Specification", record.spec_number),
                ("Chemical", record.chemical_name),
                ("Category", entry["category_label"]),
                ("Version", f"v{entry['version']}"),
                ("Material Code", record.material_code),
                ("Physical State", next((row["value"] for row in data["material_properties"] if row["field_name"] == "physical_state"), "")),
                ("Impact Classification", impact["classification"]),
                ("Prepared By", record.prepared_by),
                ("Reviewed By", record.reviewed_by),
                ("Last Revised", record.updated_at.strftime("%d %b %Y") if record.updated_at else ""),
                ("On Corporate Register", "Yes" if entry["on_register"] else "No"),
                ("Standard Testing Time", f"{entry['standard_days']} days" if entry["standard_days"] else ""),
            ]
        ),
        "section_rows": section_rows,
        "parameter_rows": parameter_rows_out,
        "master_rows": _comparison_rows(data["identity"], baseline_map("identity", identity_labels)),
        "material_property_rows": _comparison_rows(
            [{"label": row["label"], "value": row["value"]} for row in data["material_properties"] if row["value"]],
            baseline_map("material_properties", property_labels),
        ),
        "storage_rows": _comparison_rows(
            [{"label": row["label"], "value": row["value"]} for row in data["storage"] if row["value"]],
            baseline_map("storage", storage_labels),
        ),
        "impact_rows": _comparison_rows(impact_rows, baseline_impact),
        "issue_rows": data["issues"],
        "revision": None,
        "latest_review_notes": _latest_reason(record),
    }


def _latest_reason(record: CSCDraft) -> str:
    snapshot = (
        record.spec_versions.order_by(
            CSCSpecVersion.spec_version.desc(), CSCSpecVersion.created_at.desc()
        ).first()
    )
    return (snapshot.remarks or "") if snapshot is not None else ""


# ── Master export ────────────────────────────────────────────────────────────


_HEADER_FILL = PatternFill("solid", fgColor="0E766E")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_sheet(sheet: Any, headers: list[str], rows: list[list[Any]], widths: list[int]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_register_workbook() -> tuple[BytesIO, str]:
    """The whole corporate specification register: categories, specifications, parameters."""
    entries = catalogue()
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Category Summary"
    _write_sheet(
        summary,
        ["Category", "Category Name", "Chemicals", "With Parameters", "Awaiting Parameters"],
        [
            [tile["key"], tile["label"], tile["chemicals"], tile["specified"], tile["awaiting"]]
            for tile in category_tiles(entries)
        ],
        [14, 42, 12, 18, 22],
    )

    _write_sheet(
        workbook.create_sheet("Specifications"),
        [
            "Category", "Specification No.", "Chemical Name", "Material Code", "Version",
            "Parameters", "Standard Testing Days", "On Register", "Last Revised",
        ],
        [
            [
                entry["category"],
                entry["spec_number"],
                entry["chemical_name"],
                entry["material_code"],
                entry["version"] or "",
                entry["parameter_count"],
                entry["standard_days"] if entry["standard_days"] is not None else "",
                "Yes" if entry["on_register"] else "No",
                entry["updated_at"].strftime("%Y-%m-%d %H:%M") if entry["updated_at"] else "",
            ]
            for entry in entries
        ],
        [12, 26, 52, 16, 10, 12, 20, 13, 18],
    )

    parameter_data: list[list[Any]] = []
    seen: set[int] = set()
    for entry in entries:
        record = entry["record"]
        if record is None or record.id in seen:
            continue
        seen.add(record.id)
        for position, parameter in enumerate(parameter_rows(record), start=1):
            parameter_data.append(
                [
                    entry["category"], record.spec_number or "", record.chemical_name or "", position,
                    parameter["parameter_name"], parameter["parameter_type"], parameter["requirement"],
                    parameter["unit_of_measure"], parameter["conditions"], parameter["test_method"],
                    parameter["remarks"],
                ]
            )
    _write_sheet(
        workbook.create_sheet("Parameters"),
        [
            "Category", "Specification No.", "Chemical Name", "S. No.", "Parameter", "Type",
            "Required Value", "Unit", "Conditions", "Test Method", "Remarks",
        ],
        parameter_data,
        [12, 26, 42, 8, 46, 12, 36, 12, 30, 26, 26],
    )

    _write_supporting_sheets(workbook, entries)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    return stream, f"ONGC_Corporate_Specification_Register_{stamp}.xlsx"


def _write_supporting_sheets(workbook: Workbook, entries: list[dict[str, Any]]) -> None:
    """Material master, properties, storage, impact and issue sheets — one row per specification."""
    from app.core.services.csc_master_data import (
        get_material_properties_fields, get_storage_handling_fields,
    )

    property_fields = [(str(f["field_name"]), str(f["label"])) for f in get_material_properties_fields()]
    storage_field_list = [(str(f["field_name"]), str(f["label"])) for f in get_storage_handling_fields()]
    impact_dimensions = [(flag["id"], flag["dimension"]) for flag in sorted(IMPACT_CHECKLIST_FLAGS, key=lambda f: f["order"])]

    master_data: list[list[Any]] = []
    property_data: list[list[Any]] = []
    storage_data: list[list[Any]] = []
    impact_data: list[list[Any]] = []
    issue_data: list[list[Any]] = []

    seen: set[int] = set()
    for entry in entries:
        record = entry["record"]
        if record is None or record.id in seen:
            continue
        seen.add(record.id)
        values = master_values(record)
        head = [entry["category"], record.spec_number or "", record.chemical_name or "", record.material_code or ""]
        master_data.append(head + [_display(values.get(field)) for field, _label in IDENTITY_FIELDS])
        property_data.append(head + [_display(values.get(field)) for field, _label in property_fields])
        storage_data.append(head + [_display(values.get(field)) for field, _label in storage_field_list])

        impact = impact_view(record)
        impact_data.append(
            head
            + [
                "Yes" if impact["recorded"] else "No",
                impact["classification"], impact["confidence"], impact["rule"],
                impact["red_yes_count"], impact["amber_yes_count"],
                f"{impact['answered_count']}/{impact['total_flags']}",
            ]
            + [flag["value"] for flag in impact["flags"]]
        )

        flags = {row["issue_type"]: row for row in issue_flag_rows(record)}
        issue_row = list(head)
        for issue_type, _label in ISSUE_FLAG_LABELS:
            row = flags[issue_type]
            issue_row += ["Yes" if row["is_present"] else "No", row["note"]]
        issue_data.append(issue_row)

    head_headers = ["Category", "Specification No.", "Chemical Name", "Material Code"]
    _write_sheet(
        workbook.create_sheet("Material Master"),
        head_headers + [label for _field, label in IDENTITY_FIELDS],
        master_data,
        [12, 26, 42, 16] + [26] * len(IDENTITY_FIELDS),
    )
    _write_sheet(
        workbook.create_sheet("Material Properties"),
        head_headers + [label for _field, label in property_fields],
        property_data,
        [12, 26, 42, 16] + [24] * len(property_fields),
    )
    _write_sheet(
        workbook.create_sheet("Storage and Handling"),
        head_headers + [label for _field, label in storage_field_list],
        storage_data,
        [12, 26, 42, 16] + [24] * len(storage_field_list),
    )
    _write_sheet(
        workbook.create_sheet("Impact Assessment"),
        head_headers
        + ["Assessment Recorded", "Classification", "Confidence", "Decision Rule", "Red YES", "Amber YES", "Flags Answered"]
        + [dimension for _flag_id, dimension in impact_dimensions],
        impact_data,
        [12, 26, 42, 16, 18, 18, 14, 46, 10, 12, 16] + [20] * len(impact_dimensions),
    )
    issue_headers = list(head_headers)
    for _issue_type, label in ISSUE_FLAG_LABELS:
        issue_headers += [f"{label} Flagged", f"{label} Note"]
    _write_sheet(
        workbook.create_sheet("Issue Register"),
        issue_headers,
        issue_data,
        [12, 26, 42, 16] + [16, 34] * len(ISSUE_FLAG_LABELS),
    )


def _impact_analysis_payload(record: CSCDraft) -> dict[str, Any] | None:
    analysis = CSCImpactAnalysis.query.filter_by(draft_id=record.id).first()
    return analysis.to_dict() if analysis is not None else None


def export_bundles(category: str | None = None) -> list[dict[str, Any]]:
    """Specification payloads for the master Word document, one per specification record."""
    bundles: list[dict[str, Any]] = []
    seen: set[int] = set()
    for entry in catalogue():
        record = entry["record"]
        if record is None or record.id in seen or not entry["has_parameters"]:
            continue
        if category and entry["category"] != category:
            continue
        seen.add(record.id)
        payload = record.to_dict()
        payload["version_display"] = f"v{format_spec_version(record.spec_version)}"
        bundles.append(
            {
                "draft": payload,
                "sections": {
                    name: text
                    for name, text in _sections(record).items()
                    if not name.startswith("__workflow")
                },
                "parameters": [
                    {
                        "id": parameter["id"],
                        "parameter_name": parameter["parameter_name"],
                        "parameter_type": parameter["parameter_type"],
                        "existing_value": parameter["requirement"],
                        "proposed_value": "",
                        "test_method": parameter["test_method"],
                        "sort_order": parameter["sort_order"],
                    }
                    for parameter in parameter_rows(record)
                ],
                "flags": [
                    {
                        "issue_type": row["issue_type"],
                        "is_present": row["is_present"],
                        "note": row["note"],
                        "sort_order": order,
                    }
                    for order, row in enumerate(issue_flag_rows(record))
                ],
                "impact_analysis": _impact_analysis_payload(record),
            }
        )
    return bundles


def export_categories() -> list[dict[str, Any]]:
    """Category options for the master document, counting the sheets it would print."""
    counts: dict[str, int] = {}
    seen: set[int] = set()
    for entry in catalogue():
        if not entry["has_parameters"] or entry["record_id"] in seen:
            continue
        seen.add(entry["record_id"])
        counts[entry["category"]] = counts.get(entry["category"], 0) + 1
    return [
        {"code": code, "label": category_label(code), "count": counts[code]}
        for code in SPEC_SUBSET_ORDER + sorted(set(counts) - set(SPEC_SUBSET_ORDER))
        if counts.get(code)
    ]
