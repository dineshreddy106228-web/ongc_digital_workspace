"""Director inventory workbook builder for SAP procurement and consumption exports."""

from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import IO, Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="top")
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="top")
HELPER_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
HELPER_HEADER_FONT = Font(color="203864", bold=True)
ONGC_NAVY = "0F3B63"
ONGC_BLUE = "1F5A92"
ONGC_TEAL = "14808A"
ONGC_AMBER = "C55A11"
ONGC_GOLD = "C9A227"
ONGC_SLATE = "5B6B7A"
NUMFMT_COUNT = "#,##0"
NUMFMT_QTY = "#,##0.000"
NUMFMT_VALUE = "#,##0.00"
EXEC_HEADER_FILL = PatternFill(fill_type="solid", fgColor=ONGC_NAVY)
EXEC_SUB_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")
EXEC_HEADER_FONT = Font(color="FFFFFF", bold=True, size=16)
EXEC_SUB_FONT = Font(color=ONGC_NAVY, italic=True, size=10)
SECTION_FONT = Font(color=ONGC_NAVY, bold=True, size=10)

PROCUREMENT_HEADERS = {
    "plant": "plant",
    "net price": "net_price",
    "purchasing document": "purchasing_document",
    "document date": "document_date",
    "material": "material",
    "item": "item",
    "supplier supplying plant": "supplier",
    "short text": "short_text",
    "order quantity": "order_quantity",
    "still to be delivered qty": "still_to_be_delivered_quantity",
    "still to be delivered quantity": "still_to_be_delivered_quantity",
    "still to be delivered": "still_to_be_delivered_quantity",
    "open quantity": "still_to_be_delivered_quantity",
    "open qty": "still_to_be_delivered_quantity",
    "order unit": "order_unit",
    "currency": "currency",
    "price unit": "price_unit",
    "effective value": "effective_value",
    "release indicator": "release_indicator",
}

CONSUMPTION_HEADERS = {
    "plant": "plant",
    "material": "material",
    "storage location": "storage_location",
    "month": "month",
    "usage quantity": "usage_qty",
    "usage uom": "usage_uom",
    "usage value": "usage_value",
    "usage currency": "currency",
}

PROCUREMENT_REQUIRED_FIELDS = {
    "plant",
    "net_price",
    "item",
    "material",
    "supplier",
    "short_text",
    "order_quantity",
    "still_to_be_delivered_quantity",
    "order_unit",
    "currency",
    "price_unit",
    "effective_value",
    "release_indicator",
    "document_date",
    "purchasing_document",
}
CONSUMPTION_REQUIRED_FIELDS = {
    "plant",
    "material",
    "storage_location",
    "month",
    "usage_qty",
    "usage_uom",
    "usage_value",
    "currency",
}

DATE_FORMATS = (
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%d.%m.%Y",
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d-%m-%y",
    "%d/%m/%y",
    "%d.%m.%y",
    "%m/%d/%Y",
    "%m-%d-%Y",
    "%b %Y",
    "%B %Y",
)

NUMBER_QUANTITY = Decimal("0.001")
NUMBER_VALUE = Decimal("0.01")

PROCUREMENT_FIELD_LABELS = {
    "plant": "Plant",
    "net_price": "Net Price",
    "purchasing_document": "Purchasing Document",
    "document_date": "Document Date",
    "material": "Material",
    "item": "Item",
    "supplier": "Supplier/Supplying Plant",
    "short_text": "Short Text",
    "order_quantity": "Order Quantity",
    "still_to_be_delivered_quantity": "Still to be delivered (qty)",
    "order_unit": "Order Unit",
    "currency": "Currency",
    "price_unit": "Price Unit",
    "effective_value": "Effective value",
    "release_indicator": "Release indicator",
}

CONSUMPTION_FIELD_LABELS = {
    "plant": "Plant",
    "material": "Material",
    "storage_location": "Storage Location",
    "month": "Month",
    "usage_qty": "Usage Quantity",
    "usage_uom": "Usage UoM",
    "usage_value": "Usage Value",
    "currency": "Usage Currency",
}


@dataclass(frozen=True)
class ReportingPeriod:
    label: str
    file_token: str
    start_date: date
    end_date: date
    month_count: int


@dataclass(frozen=True)
class DirectorReportResult:
    workbook_bytes: bytes
    filename: str
    consumption_row_count: int
    procurement_row_count: int
    issue_count: int


class ValidationCollector:
    """Collect unique validation issues without exploding workbook size."""

    def __init__(self) -> None:
        self._issues: list[tuple[str, str, str]] = []
        self._seen: set[tuple[str, str, str]] = set()

    def add(self, issue_type: str, material: str | None, details: str) -> None:
        key = (issue_type, (material or "").strip(), details.strip())
        if key in self._seen:
            return
        self._seen.add(key)
        self._issues.append(key)

    @property
    def rows(self) -> list[tuple[str, str, str]]:
        return list(self._issues)


def build_director_inventory_report(
    *,
    consumption_bytes: bytes,
    consumption_filename: str,
    procurement_bytes: bytes,
    procurement_filename: str,
) -> DirectorReportResult:
    """Parse the two SAP exports and return the Director workbook."""
    validations = ValidationCollector()

    consumption_rows = _parse_consumption_rows(
        _read_rows(io.BytesIO(consumption_bytes), consumption_filename),
        validations,
    )
    procurement_parse = _parse_procurement_rows(
        _read_rows(io.BytesIO(procurement_bytes), procurement_filename),
        validations,
    )
    procurement_rows = procurement_parse["rows"]
    if not consumption_rows and not procurement_rows:
        raise ValueError("No procurement or consumption rows were found in the uploaded files.")

    reporting_period = infer_reporting_period(consumption_rows, procurement_rows)

    workbook_bytes = _build_workbook(
        reporting_period=reporting_period,
        consumption_rows=consumption_rows,
        procurement_rows=procurement_rows,
        validations=validations,
        excluded_procurement_rows=procurement_parse["excluded_rows"],
    )

    filename = f"Director_Inventory_Chemical_Analysis_{reporting_period.file_token}.xlsx"
    return DirectorReportResult(
        workbook_bytes=workbook_bytes,
        filename=filename,
        consumption_row_count=len(consumption_rows),
        procurement_row_count=len(procurement_rows),
        issue_count=len(validations.rows),
    )


def infer_reporting_period(
    consumption_rows: list[dict[str, Any]],
    procurement_rows: list[dict[str, Any]],
) -> ReportingPeriod:
    """Infer the report period from the union of consumption and procurement data."""
    period_points = [row["month_date"] for row in consumption_rows]
    period_points.extend(date(row["document_date"].year, row["document_date"].month, 1) for row in procurement_rows)
    if not period_points:
        raise ValueError("Could not infer a reporting period from the uploaded files.")

    start_date = min(period_points)
    end_month = max(period_points)
    end_date = date(end_month.year, end_month.month, 1)
    month_count = _month_span(start_date, end_date)

    if month_count == 1:
        label = start_date.strftime("%b %Y")
        file_token = start_date.strftime("%Y_%m")
    elif (
        start_date.month == 4
        and end_date.month == 3
        and end_date.year == start_date.year + 1
        and month_count == 12
    ):
        label = f"FY{start_date.year}-{str(end_date.year)[-2:]}"
        file_token = f"FY{start_date.year}_{str(end_date.year)[-2:]}"
    else:
        label = f"{start_date.strftime('%b %Y')} to {end_date.strftime('%b %Y')}"
        file_token = f"{start_date.strftime('%Y_%m')}_to_{end_date.strftime('%Y_%m')}"

    return ReportingPeriod(
        label=label,
        file_token=file_token,
        start_date=start_date,
        end_date=end_date,
        month_count=month_count,
    )


def _read_rows(file_stream: IO[bytes], filename: str) -> list[list[Any]]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "csv":
        return _read_csv(file_stream)
    if ext == "xlsx":
        return _read_xlsx(file_stream)
    if ext == "xls":
        raise ValueError("Legacy .xls files are not supported. Save the SAP export as .xlsx or .csv.")
    raise ValueError(f"Unsupported file type '.{ext}'. Please upload a CSV or .xlsx file.")


def _read_csv(file_stream: IO[bytes]) -> list[list[str]]:
    raw_bytes = file_stream.read()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw_bytes.decode(encoding)
            return [row for row in csv.reader(io.StringIO(text))]
        except UnicodeDecodeError:
            continue
    raise ValueError("Could not decode CSV file. Ensure the export is UTF-8 or Windows-1252 encoded.")


def _read_xlsx(file_stream: IO[bytes]) -> list[list[Any]]:
    workbook = load_workbook(filename=io.BytesIO(file_stream.read()), read_only=True, data_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _normalize_header(raw_value: Any) -> str:
    text = str(raw_value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    tokens = text.split()
    if len(tokens) > 1 and len(tokens[0]) == 1 and tokens[0].isalpha():
        tokens = tokens[1:]
    return " ".join(tokens)


def _format_missing_fields(missing: set[str], field_labels: dict[str, str]) -> str:
    return ", ".join(field_labels.get(field, field) for field in sorted(missing))


def _map_headers(
    row: list[Any],
    aliases: dict[str, str],
    required_fields: set[str],
    field_labels: dict[str, str],
) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for index, raw in enumerate(row):
        field = aliases.get(_normalize_header(raw))
        if field:
            mapping[index] = field

    missing = required_fields.difference(mapping.values())
    if missing:
        expected = _format_missing_fields(missing, field_labels)
        raise ValueError(f"Missing required columns: {expected}.")
    return mapping


def _is_blank_row(row: list[Any]) -> bool:
    return not any(str(cell or "").strip() for cell in row)


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))

    raw = str(value).strip()
    if not raw:
        return None

    negative = False
    if raw.startswith("(") and raw.endswith(")"):
        negative = True
        raw = raw[1:-1]
    if raw.endswith("-"):
        negative = True
        raw = raw[:-1]

    cleaned = re.sub(r"[^\d,.\-]", "", raw)
    if cleaned.count(",") and cleaned.count("."):
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif cleaned.count(","):
        whole, _, fraction = cleaned.rpartition(",")
        if len(fraction) <= 2:
            cleaned = f"{whole.replace(',', '')}.{fraction}"
        else:
            cleaned = cleaned.replace(",", "")

    try:
        number = Decimal(cleaned)
    except InvalidOperation:
        return None
    return -number if negative else number


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 1 < float(value) < 80000:
        return date(1899, 12, 30) + timedelta(days=int(float(value)))

    raw = str(value).strip()
    if not raw:
        return None

    if re.fullmatch(r"\d{8}", raw):
        try:
            return datetime.strptime(raw, "%Y%m%d").date()
        except ValueError:
            return None

    for fmt in DATE_FORMATS:
        try:
            parsed = datetime.strptime(raw, fmt)
            return date(parsed.year, parsed.month, parsed.day)
        except ValueError:
            continue
    return None


def _parse_month(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    if isinstance(value, (int, float)):
        integer = int(float(value))
        if 200001 <= integer <= 210012:
            year = integer // 100
            month = integer % 100
            if 1 <= month <= 12:
                return date(year, month, 1)
        excel_date = _parse_date(value)
        if excel_date:
            return date(excel_date.year, excel_date.month, 1)

    raw = str(value).strip()
    if not raw:
        return None

    patterns = (
        (r"^(\d{4})[./-](\d{1,2})$", False),
        (r"^(\d{1,2})[/-](\d{4})$", True),
        (r"^(\d{1,2})\.(\d{4})$", True),
        (r"^(\d{4})(\d{2})$", False),
    )
    for pattern, month_first in patterns:
        match = re.match(pattern, raw)
        if not match:
            continue
        if month_first:
            month = int(match.group(1))
            year = int(match.group(2))
        else:
            year = int(match.group(1))
            month = int(match.group(2))
        if 1 <= month <= 12:
            return date(year, month, 1)

    parsed = _parse_date(raw)
    if parsed:
        return date(parsed.year, parsed.month, 1)
    return None


def _row_dict(row: list[Any], header_map: dict[int, str]) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for index, field in header_map.items():
        values[field] = row[index] if index < len(row) else None
    return values


def _map_consumption_headers(row: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}

    for index, raw in enumerate(row):
        normalized = _normalize_header(raw)
        field = CONSUMPTION_HEADERS.get(normalized)
        if field:
            mapping[index] = field

    missing = CONSUMPTION_REQUIRED_FIELDS.difference(mapping.values())
    if missing:
        expected = _format_missing_fields(missing, CONSUMPTION_FIELD_LABELS)
        raise ValueError(f"Missing required consumption columns: {expected}.")
    return mapping


def _extract_material_code(value: str) -> str:
    text = value.strip()
    if not text:
        return ""
    match = re.match(r"^(\d+)\b", text)
    if match:
        return match.group(1)
    return text


def _month_span(start_month: date, end_month: date) -> int:
    return (end_month.year - start_month.year) * 12 + (end_month.month - start_month.month) + 1


def _storage_prefix(storage_location: str) -> str:
    match = re.match(r"^(\d{2})", storage_location.strip())
    return match.group(1) if match else ""


def _select_reporting_plant(prefix: str, plants_by_prefix: dict[str, set[str]]) -> str:
    if not prefix:
        return ""

    plants = plants_by_prefix.get(prefix, set())
    if not plants:
        return f"{prefix}A1"

    preferred = f"{prefix}A1"
    if preferred in plants:
        return preferred

    alpha_plants = sorted(plant for plant in plants if re.match(rf"^{re.escape(prefix)}A\d+$", plant))
    if alpha_plants:
        return alpha_plants[0]

    return sorted(plants)[0]


def _parse_consumption_rows(rows: list[list[Any]], validations: ValidationCollector) -> list[dict[str, Any]]:
    if not rows:
        raise ValueError("Consumption file is empty.")

    header_map = _map_consumption_headers(rows[0])
    parsed_rows: list[dict[str, Any]] = []

    for line_number, row in enumerate(rows[1:], start=2):
        if _is_blank_row(row):
            continue
        values = _row_dict(row, header_map)

        material = _extract_material_code(_stringify(values.get("material")))
        plant = _stringify(values.get("plant"))
        storage_location = _stringify(values.get("storage_location"))
        month_date = _parse_month(values.get("month"))

        if not material:
            validations.add("MISSING_MATERIAL", "", f"Consumption row {line_number} was skipped because Material is blank.")
            continue
        if not month_date:
            validations.add("INVALID_MONTH", material, f"Consumption row {line_number} has an invalid Month value.")
            continue
        if not plant:
            validations.add("MISSING_PLANT", material, f"Consumption row {line_number} has a blank Plant value.")

        usage_qty = _parse_decimal(values.get("usage_qty")) or Decimal("0")
        usage_value = _parse_decimal(values.get("usage_value")) or Decimal("0")

        parsed_rows.append(
            {
                "plant": plant,
                "material": material,
                "storage_location": storage_location,
                "storage_prefix": _storage_prefix(storage_location),
                "month_date": month_date,
                "month_key": month_date.strftime("%Y-%m"),
                "usage_qty": usage_qty,
                "usage_uom": _stringify(values.get("usage_uom")),
                "usage_value": usage_value,
                "currency": _stringify(values.get("currency")),
            }
        )

    return parsed_rows


def _parse_procurement_rows(rows: list[list[Any]], validations: ValidationCollector) -> dict[str, Any]:
    if not rows:
        raise ValueError("Procurement file is empty.")

    header_map = _map_headers(
        rows[0],
        PROCUREMENT_HEADERS,
        PROCUREMENT_REQUIRED_FIELDS,
        PROCUREMENT_FIELD_LABELS,
    )
    parsed_rows: list[dict[str, Any]] = []
    excluded_rows = 0

    for line_number, row in enumerate(rows[1:], start=2):
        if _is_blank_row(row):
            continue
        values = _row_dict(row, header_map)

        material = _stringify(values.get("material"))
        document_date = _parse_date(values.get("document_date"))
        deletion_indicator = _stringify(values.get("deletion_indicator"))

        if deletion_indicator:
            excluded_rows += 1
            continue
        if not material:
            validations.add("MISSING_MATERIAL", "", f"Procurement row {line_number} was skipped because Material is blank.")
            continue
        if not document_date:
            validations.add("INVALID_PROCUREMENT_DATE", material, f"Procurement row {line_number} has an invalid Document Date.")
            continue

        short_text = _stringify(values.get("short_text"))
        if not short_text:
            validations.add("MISSING_CHEMICAL_NAME", material, "No Short Text in procurement data.")

        ordered_quantity = _parse_decimal(values.get("order_quantity")) or Decimal("0")
        still_to_be_delivered_quantity = _parse_decimal(values.get("still_to_be_delivered_quantity")) or Decimal("0")
        parsed_rows.append(
            {
                "plant": _stringify(values.get("plant")),
                "material": material,
                "storage_location": _stringify(values.get("storage_location")),
                "document_date": document_date,
                "document_month": date(document_date.year, document_date.month, 1),
                "purchasing_document": _stringify(values.get("purchasing_document")),
                "item": _stringify(values.get("item")),
                "supplier": _stringify(values.get("supplier")),
                "short_text": short_text,
                "order_quantity": ordered_quantity,
                "still_to_be_delivered_quantity": still_to_be_delivered_quantity,
                "procured_quantity": ordered_quantity - still_to_be_delivered_quantity,
                "order_unit": _stringify(values.get("order_unit")),
                "effective_value": _parse_decimal(values.get("effective_value")) or Decimal("0"),
                "currency": _stringify(values.get("currency")),
            }
        )

    return {"rows": parsed_rows, "excluded_rows": excluded_rows}


def _build_workbook(
    *,
    reporting_period: ReportingPeriod,
    consumption_rows: list[dict[str, Any]],
    procurement_rows: list[dict[str, Any]],
    validations: ValidationCollector,
    excluded_procurement_rows: int,
) -> bytes:
    procurement_by_material: dict[str, list[dict[str, Any]]] = defaultdict(list)
    chemical_name_lookup: dict[str, str] = {}
    procurement_uoms: dict[str, set[str]] = defaultdict(set)
    procurement_plants_by_prefix: dict[str, set[str]] = defaultdict(set)

    for row in sorted(
        procurement_rows,
        key=lambda item: (
            item["material"],
            item["document_date"],
            item["purchasing_document"],
            item["item"],
        ),
        reverse=True,
    ):
        material = row["material"]
        procurement_by_material[material].append(row)
        if row["order_unit"]:
            procurement_uoms[material].add(row["order_unit"])
        if row["short_text"] and material not in chemical_name_lookup:
            chemical_name_lookup[material] = row["short_text"]
        plant_prefix = _storage_prefix(row["plant"])
        if plant_prefix and row["plant"]:
            procurement_plants_by_prefix[plant_prefix].add(row["plant"])

    consumption_uoms: dict[str, set[str]] = defaultdict(set)
    material_summary: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "qty": Decimal("0"),
            "value": Decimal("0"),
            "plants": set(),
            "storage_locations": set(),
            "months": set(),
        }
    )
    monthly_detail_map: dict[tuple[str, str, str, str], dict[str, Any]] = {}

    for row in consumption_rows:
        material = row["material"]
        if row["usage_uom"]:
            consumption_uoms[material].add(row["usage_uom"])

        reporting_plant = _select_reporting_plant(row["storage_prefix"], procurement_plants_by_prefix)
        if not reporting_plant:
            reporting_plant = row["plant"]

        summary = material_summary[material]
        summary["qty"] += row["usage_qty"]
        summary["value"] += row["usage_value"]
        if reporting_plant:
            summary["plants"].add(reporting_plant)
        if row["storage_location"]:
            summary["storage_locations"].add(row["storage_location"])
        if row["usage_qty"] > 0:
            summary["months"].add(row["month_key"])

        detail_key = (
            reporting_plant,
            material,
            row["month_key"],
        )
        detail = monthly_detail_map.get(detail_key)
        if detail is None:
            detail = {
                "plant": reporting_plant,
                "material": material,
                "month": row["month_key"],
                "usage_qty": Decimal("0"),
                "usage_uom": row["usage_uom"],
                "usage_value": Decimal("0"),
                "currency": row["currency"],
            }
            monthly_detail_map[detail_key] = detail
        detail["usage_qty"] += row["usage_qty"]
        detail["usage_value"] += row["usage_value"]
        if not detail["usage_uom"] and row["usage_uom"]:
            detail["usage_uom"] = row["usage_uom"]
        if not detail["currency"] and row["currency"]:
            detail["currency"] = row["currency"]

    procurement_last3_rows: list[list[Any]] = []
    month_location_map: dict[tuple[str, str], dict[str, Decimal]] = defaultdict(
        lambda: {
            "consumption_qty": Decimal("0"),
            "consumption_value": Decimal("0"),
            "procurement_qty": Decimal("0"),
            "procurement_value": Decimal("0"),
        }
    )

    for detail in monthly_detail_map.values():
        month_location = month_location_map[(detail["month"], detail["plant"])]
        month_location["consumption_qty"] += detail["usage_qty"]
        month_location["consumption_value"] += detail["usage_value"]

    for material in sorted(procurement_by_material):
        rows = procurement_by_material[material]
        for rank, row in enumerate(rows[:3], start=1):
            procurement_last3_rows.append(
                [
                    material,
                    chemical_name_lookup.get(material, ""),
                    rank,
                    row["plant"],
                    row["storage_location"],
                    row["purchasing_document"],
                    row["document_date"].strftime("%Y-%m-%d"),
                    row["supplier"],
                    _excel_number(row["order_quantity"], NUMBER_QUANTITY),
                    _excel_number(row["still_to_be_delivered_quantity"], NUMBER_QUANTITY),
                    _excel_number(row["procured_quantity"], NUMBER_QUANTITY),
                    row["order_unit"],
                    _excel_number(row["effective_value"], NUMBER_VALUE),
                    row["currency"],
                ]
            )

        for row in rows:
            month_key = row["document_month"].strftime("%Y-%m")
            reporting_plant = _select_reporting_plant(_storage_prefix(row["plant"]), procurement_plants_by_prefix) or row["plant"]
            month_location = month_location_map[(month_key, reporting_plant)]
            month_location["procurement_qty"] += row["procured_quantity"]
            month_location["procurement_value"] += row["effective_value"]

    for material, units in consumption_uoms.items():
        procurement_units = procurement_uoms.get(material, set())
        if units and procurement_units and units.isdisjoint(procurement_units):
            validations.add(
                "UOM_MISMATCH",
                material,
                f"Procurement {'/'.join(sorted(procurement_units))} vs Consumption {'/'.join(sorted(units))}",
            )

    all_materials = sorted(
        set(material_summary) | set(procurement_by_material),
        key=lambda material: (chemical_name_lookup.get(material, "").lower(), material),
    )
    chemical_summary_rows: list[list[Any]] = []
    for material in all_materials:
        consumption = material_summary.get(material, {})
        procurement_list = procurement_by_material.get(material, [])
        latest_procurement = procurement_list[0] if procurement_list else None

        if material not in chemical_name_lookup:
            validations.add("MISSING_CHEMICAL_NAME", material, "No Short Text in procurement data.")

        summary_text = _last_three_procurement_summary(procurement_list)
        chemical_summary_rows.append(
            [
                chemical_name_lookup.get(material, ""),
                material,
                _preferred_uom(consumption_uoms.get(material, set()), procurement_uoms.get(material, set())),
                _excel_number(consumption.get("qty", Decimal("0")), NUMBER_QUANTITY),
                _excel_number(
                    _monthly_average(consumption.get("qty", Decimal("0")), NUMBER_QUANTITY, reporting_period.month_count),
                    NUMBER_QUANTITY,
                ),
                _excel_number(consumption.get("value", Decimal("0")), NUMBER_VALUE),
                _excel_number(
                    _monthly_average(consumption.get("value", Decimal("0")), NUMBER_VALUE, reporting_period.month_count),
                    NUMBER_VALUE,
                ),
                latest_procurement["document_date"].strftime("%d-%m-%Y") if latest_procurement else "",
                _excel_number(latest_procurement["procured_quantity"], NUMBER_QUANTITY) if latest_procurement else "",
                latest_procurement["purchasing_document"] if latest_procurement else "",
                summary_text,
                len(consumption.get("plants", set())),
                len(consumption.get("storage_locations", set())),
                len(consumption.get("months", set())),
            ]
        )

    consumption_detail_rows = [
        [
            row["plant"],
            row["material"],
            chemical_name_lookup.get(row["material"], ""),
            row["month"],
            _excel_number(row["usage_qty"], NUMBER_QUANTITY),
            row["usage_uom"],
            _excel_number(row["usage_value"], NUMBER_VALUE),
            row["currency"],
        ]
        for row in sorted(
            monthly_detail_map.values(),
            key=lambda item: (item["month"], item["plant"], item["material"]),
        )
    ]

    month_location_rows = [
        [
            month,
            plant,
            _excel_number(values["consumption_qty"], NUMBER_QUANTITY),
            _excel_number(values["consumption_value"], NUMBER_VALUE),
            _excel_number(values["procurement_qty"], NUMBER_QUANTITY),
            _excel_number(values["procurement_value"], NUMBER_VALUE),
        ]
        for (month, plant), values in sorted(month_location_map.items())
    ]

    methodology_rows = [
        ["Report Name", "Chemical Inventory Analysis"],
        ["Reporting Period", reporting_period.label],
        ["Reporting Period Start", reporting_period.start_date.strftime("%Y-%m")],
        ["Reporting Period End", reporting_period.end_date.strftime("%Y-%m")],
        ["Detected Month Count", reporting_period.month_count],
        ["Procurement Data Source", "SAP Procurement Export"],
        ["Consumption Data Source", "SAP Usage Report"],
        ["Average Monthly Calculation", f"Report Total ÷ {reporting_period.month_count} month(s)"],
        ["Procurement Ranking Logic", "Latest Document Date descending"],
        ["Excluded Procurement Rows", "None"],
        ["Excluded Procurement Row Count", excluded_procurement_rows],
        ["Consumption Plant Grouping", "Storage Location first two digits mapped to procurement plant prefix; prefer A1"],
        ["Primary Material Key", "Material"],
    ]

    validation_rows = validations.rows or [("INFO", "", "No validation issues detected.")]

    workbook = Workbook(write_only=True)

    _append_sheet(
        workbook,
        "Chemical_Summary",
        [
            "Chemical_Name",
            "Material_Code",
            "UoM",
            "Report_Consumption_Qty",
            "Avg_Monthly_Consumption_Qty",
            "Report_Consumption_Value",
            "Avg_Monthly_Consumption_Value",
            "Last_Procurement_Date",
            "Last_Procurement_Qty",
            "Last_Procurement_PO_No",
            "Last_3_Procurement_Summary",
            "Plant_Count",
            "Storage_Location_Count",
            "Active_Months_Count",
        ],
        chemical_summary_rows,
        widths=[34, 14, 8, 20, 22, 20, 22, 18, 18, 18, 44, 12, 16, 12],
        wrap_columns={1, 11},
        numeric_columns={4, 5, 6, 7, 9, 12, 13, 14},
    )
    _append_sheet(
        workbook,
        "Consumption_Monthly_Detail",
        [
            "Plant",
            "Material",
            "Chemical_Name",
            "Month",
            "Usage_Qty",
            "Usage_UoM",
            "Usage_Value",
            "Currency",
        ],
        consumption_detail_rows,
        widths=[22, 14, 34, 10, 14, 10, 14, 10],
        wrap_columns={3},
        numeric_columns={5, 7},
    )
    _append_sheet(
        workbook,
        "Procurement_Last3_Detail",
        [
            "Material",
            "Chemical_Name",
            "Procurement_Rank",
            "Plant",
            "Storage_Location",
            "Purchasing_Document",
            "Document_Date",
            "Supplier",
            "Order_Quantity",
            "Still_To_Be_Delivered_Qty",
            "Procured_Quantity",
            "Order_Unit",
            "Effective_Value",
            "Currency",
        ],
        procurement_last3_rows,
        widths=[14, 34, 10, 20, 16, 18, 14, 22, 14, 16, 14, 10, 14, 10],
        wrap_columns={2, 8},
        numeric_columns={3, 9, 10, 11, 13},
    )
    _append_sheet(
        workbook,
        "Month_Location_Summary",
        [
            "Month",
            "Plant",
            "Consumption_Qty",
            "Consumption_Value",
            "Procurement_Qty",
            "Procurement_Value",
        ],
        month_location_rows,
        widths=[10, 22, 16, 18, 16, 18],
        numeric_columns={3, 4, 5, 6},
    )
    _append_sheet(
        workbook,
        "Methodology",
        ["Field", "Value"],
        methodology_rows,
        widths=[28, 44],
        wrap_columns={2},
    )
    _append_sheet(
        workbook,
        "Validation_Issues",
        ["Issue_Type", "Material", "Details"],
        validation_rows,
        widths=[24, 16, 72],
        wrap_columns={3},
    )

    output = io.BytesIO()
    workbook.save(output)

    chart_workbook = load_workbook(io.BytesIO(output.getvalue()))
    _add_enterprise_charts(
        chart_workbook,
        chemical_summary_rows=chemical_summary_rows,
        consumption_detail_rows=consumption_detail_rows,
        procurement_last3_rows=procurement_last3_rows,
        month_location_rows=month_location_rows,
        methodology_rows=methodology_rows,
        validation_rows=validation_rows,
        material_count=len(all_materials),
        plant_count=len({row[1] for row in month_location_rows if row[1]}),
        consumption_row_count=len(consumption_detail_rows),
        procurement_row_count=len(procurement_last3_rows),
    )

    final_output = io.BytesIO()
    chart_workbook.save(final_output)
    return final_output.getvalue()


def _append_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    widths: list[int] | None = None,
    wrap_columns: set[int] | None = None,
    numeric_columns: set[int] | None = None,
) -> None:
    sheet = workbook.create_sheet(title=title)
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = True
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    wrap_columns = wrap_columns or set()
    numeric_columns = numeric_columns or set()
    if widths:
        for column_index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    header_cells = []
    for value in headers:
        cell = WriteOnlyCell(sheet, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        header_cells.append(cell)
    sheet.append(header_cells)

    for row in rows:
        body_cells = []
        for column_index, value in enumerate(row, start=1):
            cell = WriteOnlyCell(sheet, value=value)
            if column_index in numeric_columns:
                cell.alignment = NUMBER_ALIGNMENT
            elif column_index in wrap_columns:
                cell.alignment = WRAP_ALIGNMENT
            else:
                cell.alignment = BODY_ALIGNMENT
            body_cells.append(cell)
        sheet.append(body_cells)


def _add_enterprise_charts(
    workbook,
    *,
    chemical_summary_rows: list[list[Any]],
    consumption_detail_rows: list[list[Any]],
    procurement_last3_rows: list[list[Any]],
    month_location_rows: list[list[Any]],
    methodology_rows: list[list[Any]],
    validation_rows: list[tuple[str, str, str]],
    material_count: int,
    plant_count: int,
    consumption_row_count: int,
    procurement_row_count: int,
) -> None:
    helper_sheet = workbook.create_sheet("_chart_data")
    helper_sheet.sheet_state = "hidden"
    helper_cursor = 1

    chemical_sheet = workbook["Chemical_Summary"]
    _prepare_executive_layout(
        chemical_sheet,
        title="Chemical Summary Dashboard",
        subtitle="Executive view of chemical consumption, procurement recency, and activity footprint",
    )
    helper_cursor = _add_chemical_summary_chart(chemical_sheet, helper_sheet, helper_cursor, chemical_summary_rows)

    consumption_sheet = workbook["Consumption_Monthly_Detail"]
    _prepare_executive_layout(
        consumption_sheet,
        title="Consumption Monthly Dashboard",
        subtitle="Trend view of usage value by month, with plant-level monthly detail below",
    )
    helper_cursor = _add_consumption_detail_chart(consumption_sheet, helper_sheet, helper_cursor, consumption_detail_rows)

    procurement_sheet = workbook["Procurement_Last3_Detail"]
    _prepare_executive_layout(
        procurement_sheet,
        title="Procurement Dashboard",
        subtitle="Most recent procurement positions ranked by effective value",
    )
    helper_cursor = _add_procurement_detail_chart(procurement_sheet, helper_sheet, helper_cursor, procurement_last3_rows)

    month_location_sheet = workbook["Month_Location_Summary"]
    _prepare_executive_layout(
        month_location_sheet,
        title="Month and Plant Dashboard",
        subtitle="Normalized plant comparison of consumption and procurement value by month",
    )
    helper_cursor = _add_month_location_chart(month_location_sheet, helper_sheet, helper_cursor, month_location_rows)

    methodology_sheet = workbook["Methodology"]
    _prepare_executive_layout(
        methodology_sheet,
        title="Methodology Dashboard",
        subtitle="Coverage, scale, and rules used to produce the report",
        chart_rows=16,
    )
    _add_methodology_chart(
        methodology_sheet,
        helper_sheet=helper_sheet,
        start_row=helper_cursor,
        methodology_rows=methodology_rows,
        material_count=material_count,
        plant_count=plant_count,
        consumption_row_count=consumption_row_count,
        procurement_row_count=procurement_row_count,
        issue_count=max(0, len(validation_rows) - (1 if validation_rows and validation_rows[0][0] == "INFO" else 0)),
    )
    helper_cursor += 12

    validation_sheet = workbook["Validation_Issues"]
    _prepare_executive_layout(
        validation_sheet,
        title="Validation Dashboard",
        subtitle="Data quality and completeness findings from the ingestion step",
        chart_rows=16,
    )
    _add_validation_chart(validation_sheet, helper_sheet, helper_cursor, validation_rows)


def _write_helper_table(
    worksheet,
    *,
    start_row: int,
    headers: list[str],
    rows: list[list[Any]],
) -> tuple[int, int]:
    for offset, header in enumerate(headers):
        cell = worksheet.cell(row=start_row, column=1 + offset, value=header)
        cell.fill = HELPER_HEADER_FILL
        cell.font = HELPER_HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    for row_index, values in enumerate(rows, start=start_row + 1):
        for offset, value in enumerate(values):
            worksheet.cell(row=row_index, column=1 + offset, value=value)

    return start_row, start_row + max(len(rows), 1)


def _style_chart(chart, *, title: str, width: float = 11.5, height: float = 6.4) -> None:
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    chart.legend.position = "b"
    chart.roundedCorners = True


def _set_series_color(chart, series_index: int, color: str) -> None:
    if series_index >= len(chart.ser):
        return
    series = chart.ser[series_index]
    series.graphicalProperties.solidFill = color
    series.graphicalProperties.line.solidFill = color


def _configure_axes(
    chart,
    *,
    x_title: str | None = None,
    y_title: str | None = None,
    x_numfmt: str | None = None,
    y_numfmt: str | None = None,
) -> None:
    if x_title is not None and getattr(chart, "x_axis", None) is not None:
        chart.x_axis.title = x_title
    if y_title is not None and getattr(chart, "y_axis", None) is not None:
        chart.y_axis.title = y_title
    if x_numfmt and getattr(chart, "x_axis", None) is not None and hasattr(chart.x_axis, "numFmt"):
        chart.x_axis.numFmt = x_numfmt
    if y_numfmt and getattr(chart, "y_axis", None) is not None and hasattr(chart.y_axis, "numFmt"):
        chart.y_axis.numFmt = y_numfmt


def _show_data_labels(chart, *, value: bool = True, percent: bool = False, category: bool = False) -> None:
    labels = DataLabelList()
    labels.showVal = value
    labels.showPercent = percent
    labels.showCatName = category
    chart.dLbls = labels


def _prepare_executive_layout(
    worksheet,
    *,
    title: str,
    subtitle: str,
    chart_rows: int = 18,
) -> None:
    inserted_rows = chart_rows + 2
    worksheet.insert_rows(1, amount=inserted_rows)
    last_data_col = worksheet.max_column
    last_data_row = worksheet.max_row
    end_col_letter = get_column_letter(max(8, min(last_data_col, 12)))
    header_row = inserted_rows + 1
    first_data_row = header_row + 1
    section_row = header_row - 1

    worksheet.merge_cells(f"A1:{end_col_letter}1")
    worksheet["A1"] = title
    worksheet["A1"].fill = EXEC_HEADER_FILL
    worksheet["A1"].font = EXEC_HEADER_FONT
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 24

    worksheet.merge_cells(f"A2:{end_col_letter}2")
    worksheet["A2"] = subtitle
    worksheet["A2"].fill = EXEC_SUB_FILL
    worksheet["A2"].font = EXEC_SUB_FONT
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[2].height = 18

    worksheet[f"A{section_row}"] = "Detail Table"
    worksheet[f"A{section_row}"].font = SECTION_FONT
    worksheet.freeze_panes = f"A{first_data_row}"
    worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(last_data_col)}{last_data_row}"
    worksheet.sheet_view.zoomScale = 90


def _top_n(rows: list[list[Any]], *, label_index: int, value_index: int, fallback_index: int | None = None, limit: int = 10) -> list[list[Any]]:
    ranked = []
    for row in rows:
        value = row[value_index]
        if value in (None, "", 0, 0.0):
            continue
        label = row[label_index] or (row[fallback_index] if fallback_index is not None else "")
        ranked.append([label, value])
    ranked.sort(key=lambda item: item[1], reverse=True)
    return ranked[:limit]


def _aggregate_sum(rows: list[list[Any]], *, key_index: int, value_index: int) -> list[list[Any]]:
    totals: dict[Any, Decimal] = defaultdict(lambda: Decimal("0"))
    for row in rows:
        key = row[key_index]
        value = row[value_index]
        if key in (None, "") or value in (None, ""):
            continue
        totals[key] += Decimal(str(value))
    return [[key, float(total)] for key, total in sorted(totals.items())]


def _add_chemical_summary_chart(worksheet, helper_sheet, start_row: int, rows: list[list[Any]]) -> int:
    helper_rows = _top_n(rows, label_index=0, value_index=5, fallback_index=1, limit=10)
    if not helper_rows:
        helper_rows = [["No Data", 0]]
    start_row, end_row = _write_helper_table(
        helper_sheet,
        start_row=start_row,
        headers=["Chemical", "Consumption Value"],
        rows=helper_rows,
    )

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.gapWidth = 55
    chart.overlap = 0
    _style_chart(chart, title="Top Chemicals by Consumption Value")
    _configure_axes(chart, x_title="Chemical", y_title="Consumption Value", y_numfmt=NUMFMT_VALUE)
    data = Reference(helper_sheet, min_col=2, min_row=start_row, max_row=end_row)
    cats = Reference(helper_sheet, min_col=1, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _set_series_color(chart, 0, ONGC_NAVY)
    _show_data_labels(chart)
    worksheet.add_chart(chart, "A3")
    return end_row + 3


def _add_consumption_detail_chart(worksheet, helper_sheet, start_row: int, rows: list[list[Any]]) -> int:
    helper_rows = _aggregate_sum(rows, key_index=3, value_index=6)
    if not helper_rows:
        helper_rows = [["No Data", 0]]
    start_row, end_row = _write_helper_table(
        helper_sheet,
        start_row=start_row,
        headers=["Month", "Usage Value"],
        rows=helper_rows,
    )

    chart = LineChart()
    chart.grouping = "standard"
    chart.marker = "circle"
    chart.smooth = True
    _style_chart(chart, title="Monthly Consumption Value Trend")
    _configure_axes(chart, x_title="Month", y_title="Usage Value", y_numfmt=NUMFMT_VALUE)
    data = Reference(helper_sheet, min_col=2, min_row=start_row, max_row=end_row)
    cats = Reference(helper_sheet, min_col=1, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _set_series_color(chart, 0, ONGC_BLUE)
    worksheet.add_chart(chart, "A3")
    return end_row + 3


def _add_procurement_detail_chart(worksheet, helper_sheet, start_row: int, rows: list[list[Any]]) -> int:
    latest_rows = [row for row in rows if row[2] == 1 and row[10] not in (None, "", 0, 0.0)]
    helper_rows = _top_n(latest_rows, label_index=1, value_index=10, fallback_index=0, limit=10)
    if not helper_rows:
        helper_rows = [["No Data", 0]]
    start_row, end_row = _write_helper_table(
        helper_sheet,
        start_row=start_row,
        headers=["Chemical", "Latest Procurement Value"],
        rows=helper_rows,
    )

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.gapWidth = 50
    _style_chart(chart, title="Top Latest Procurement Values")
    _configure_axes(chart, x_title="Chemical", y_title="Effective Value", y_numfmt=NUMFMT_VALUE)
    data = Reference(helper_sheet, min_col=2, min_row=start_row, max_row=end_row)
    cats = Reference(helper_sheet, min_col=1, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _set_series_color(chart, 0, ONGC_TEAL)
    _show_data_labels(chart)
    worksheet.add_chart(chart, "A3")
    return end_row + 3


def _add_month_location_chart(worksheet, helper_sheet, start_row: int, rows: list[list[Any]]) -> int:
    consumption_by_month = _aggregate_sum(rows, key_index=0, value_index=3)
    procurement_by_month = dict((row[0], row[1]) for row in _aggregate_sum(rows, key_index=0, value_index=5))
    helper_rows = [
        [month, value, procurement_by_month.get(month, 0)]
        for month, value in consumption_by_month
    ]
    if not helper_rows:
        helper_rows = [["No Data", 0, 0]]
    start_row, end_row = _write_helper_table(
        helper_sheet,
        start_row=start_row,
        headers=["Month", "Consumption Value", "Procurement Value"],
        rows=helper_rows,
    )

    chart = LineChart()
    chart.grouping = "standard"
    chart.marker = "circle"
    chart.smooth = True
    _style_chart(chart, title="Consumption vs Procurement Value by Month")
    _configure_axes(chart, x_title="Month", y_title="Value", y_numfmt=NUMFMT_VALUE)
    data = Reference(helper_sheet, min_col=2, max_col=3, min_row=start_row, max_row=end_row)
    cats = Reference(helper_sheet, min_col=1, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _set_series_color(chart, 0, ONGC_NAVY)
    _set_series_color(chart, 1, ONGC_AMBER)
    worksheet.add_chart(chart, "A3")
    return end_row + 3


def _add_methodology_chart(
    worksheet,
    *,
    helper_sheet,
    start_row: int,
    methodology_rows: list[list[Any]],
    material_count: int,
    plant_count: int,
    consumption_row_count: int,
    procurement_row_count: int,
    issue_count: int,
) -> None:
    detected_month_count = 0
    for field, value in methodology_rows:
        if field == "Detected Month Count":
            detected_month_count = int(value)
            break
    helper_rows = [
        ["Detected Months", detected_month_count],
        ["Materials", material_count],
        ["Plants", plant_count],
        ["Consumption Rows", consumption_row_count],
        ["Procurement Rows", procurement_row_count],
        ["Validation Issues", issue_count],
    ]
    start_row, end_row = _write_helper_table(
        helper_sheet,
        start_row=start_row,
        headers=["Metric", "Count"],
        rows=helper_rows,
    )

    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.gapWidth = 45
    _style_chart(chart, title="Report Coverage Metrics", width=10.5, height=5.8)
    _configure_axes(chart, x_title="Count", y_title="Metric", x_numfmt=NUMFMT_COUNT)
    data = Reference(helper_sheet, min_col=2, min_row=start_row, max_row=end_row)
    cats = Reference(helper_sheet, min_col=1, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _set_series_color(chart, 0, ONGC_GOLD)
    _show_data_labels(chart)
    worksheet.add_chart(chart, "A3")


def _add_validation_chart(worksheet, helper_sheet, start_row: int, rows: list[tuple[str, str, str]]) -> None:
    issue_counts: dict[str, int] = defaultdict(int)
    for issue_type, _, _ in rows:
        issue_counts[issue_type] += 1
    helper_rows = [[issue_type, count] for issue_type, count in sorted(issue_counts.items(), key=lambda item: (-item[1], item[0]))]
    if not helper_rows:
        helper_rows = [["INFO", 0]]
    start_row, end_row = _write_helper_table(
        helper_sheet,
        start_row=start_row,
        headers=["Issue Type", "Issue Count"],
        rows=helper_rows,
    )

    chart = DoughnutChart()
    _style_chart(chart, title="Validation Issue Mix", width=9.5, height=5.8)
    chart.varyColors = True
    data = Reference(helper_sheet, min_col=2, min_row=start_row, max_row=end_row)
    cats = Reference(helper_sheet, min_col=1, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _show_data_labels(chart, value=False, percent=True, category=True)
    worksheet.add_chart(chart, "A3")


def _preferred_uom(consumption_units: set[str], procurement_units: set[str]) -> str:
    if consumption_units:
        return sorted(consumption_units)[0]
    if procurement_units:
        return sorted(procurement_units)[0]
    return ""


def _monthly_average(value: Decimal, quantum: Decimal, month_count: int) -> Decimal:
    divisor = Decimal(month_count or 1)
    return (value / divisor).quantize(quantum, rounding=ROUND_HALF_UP)


def _excel_number(value: Decimal, quantum: Decimal) -> int | float:
    quantized = value.quantize(quantum, rounding=ROUND_HALF_UP)
    if quantum == NUMBER_VALUE:
        return float(quantized)
    if quantized == quantized.to_integral():
        return int(quantized)
    return float(quantized)


def _last_three_procurement_summary(procurement_rows: list[dict[str, Any]]) -> str:
    parts = []
    for rank, row in enumerate(procurement_rows[:3], start=1):
        quantity = _excel_number(row["procured_quantity"], NUMBER_QUANTITY)
        parts.append(
            f"{rank}){row['purchasing_document']}|{row['document_date'].strftime('%d-%m-%y')}|{quantity} {row['order_unit']}".strip()
        )
    return "; ".join(parts)
