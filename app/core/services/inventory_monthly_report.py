"""Monthly inventory upload processing and report generation."""

from __future__ import annotations

import calendar
import json
import textwrap
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import undefer

from app.extensions import db
from app.models.inventory.monthly_upload_batch import InventoryMonthlyUploadBatch
from app.core.services.inventory_intelligence import (
    _detect_frame_variant,
    _normalize_consumption_frame,
    _normalize_mc9_frame,
    _normalize_procurement_frame,
    _normalize_text,
    _preferred_uom,
    _quantity_to_mt,
    _read_uploaded_seed_dataframe,
    _safe_float,
    _validate_seed_frame,
)


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="DCE6F1")
SECTION_FONT = Font(color="203864", bold=True)
ANOMALY_STDDEV_THRESHOLD = 1.5
ANOMALY_MIN_HISTORY_MONTHS = 2
DAYS_ON_HAND_RED = 30.0
DAYS_ON_HAND_AMBER = 90.0
DEAD_STOCK_MONTHS = 3
OPEN_ORDER_AGING_DAYS = 60
OPEN_ORDER_VALUE_RISK_MULTIPLE = 2.0
TOP_SECTION_LIMIT = 8


@dataclass(frozen=True)
class MonthlyInventoryReportPackage:
    report_year: int
    report_month: int
    report_label: str
    mb51_filename: str
    me2m_filename: str
    mc9_filename: str
    mb51_row_count: int
    me2m_row_count: int
    mc9_row_count: int
    excel_filename: str
    excel_bytes: bytes
    pdf_filename: str
    pdf_bytes: bytes
    summary: dict[str, Any]


def _validate_report_period(report_year: int, report_month: int) -> None:
    if report_year < 2000 or report_year > 2100:
        raise ValueError("Select a valid report year for the monthly update.")
    if report_month < 1 or report_month > 12:
        raise ValueError("Select a valid report month for the monthly update.")


def _month_label(report_year: int, report_month: int) -> str:
    return f"{calendar.month_abbr[int(report_month)]} {int(report_year)}"


def _month_token(report_year: int, report_month: int) -> str:
    return f"{int(report_year):04d}-{int(report_month):02d}"


def _available_month_labels(frame: pd.DataFrame) -> str:
    if frame.empty or "year" not in frame.columns or "month" not in frame.columns:
        return "none"
    months = (
        frame.loc[frame["year"].notna() & frame["month"].notna(), ["year", "month"]]
        .drop_duplicates()
        .sort_values(["year", "month"])
        .itertuples(index=False, name=None)
    )
    labels = [f"{calendar.month_abbr[int(month)]} {int(year)}" for year, month in months]
    return ", ".join(labels) if labels else "none"


def _filter_month(frame: pd.DataFrame, report_year: int, report_month: int) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()
    years = pd.to_numeric(frame.get("year"), errors="coerce")
    months = pd.to_numeric(frame.get("month"), errors="coerce")
    mask = years.eq(int(report_year)) & months.eq(int(report_month))
    return frame.loc[mask].copy()


def _previous_month(report_year: int, report_month: int) -> tuple[int, int]:
    if int(report_month) == 1:
        return int(report_year) - 1, 12
    return int(report_year), int(report_month) - 1


def _material_key_columns() -> list[str]:
    return ["material_code", "material_desc"]


def _canonical_material_code(value) -> str:
    text = _normalize_text(value).upper()
    if text.endswith(".0"):
        text = text[:-2]
    if text.isdigit():
        return text.lstrip("0") or "0"
    return text


def _material_display(code: str, desc: str) -> str:
    code_text = _normalize_text(code)
    desc_text = _normalize_text(desc)
    if code_text and desc_text:
        return f"{desc_text} ({code_text})"
    return desc_text or code_text or "Unknown Material"


def _rows_to_dicts(frame: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for _, row in frame.iterrows():
        item: dict[str, Any] = {}
        for column, value in row.items():
            if isinstance(value, pd.Timestamp):
                item[column] = value.strftime("%Y-%m-%d")
            elif pd.isna(value):
                item[column] = ""
            else:
                item[column] = value
        rows.append(item)
    return rows


def _format_inr_compact(value: float) -> str:
    value = _safe_float(value)
    if abs(value) >= 1_00_00_000:
        return f"Rs {value / 1_00_00_000:,.2f} Cr"
    if abs(value) >= 1_00_000:
        return f"Rs {value / 1_00_000:,.2f} L"
    return f"Rs {value:,.2f}"


def _format_pct(value: float | None) -> str:
    if value is None:
        return "n/a"
    return f"{value:+.1f}%"


def _month_end(report_year: int, report_month: int) -> pd.Timestamp:
    return pd.Timestamp(datetime(report_year, report_month, calendar.monthrange(report_year, report_month)[1]))


def _history_key(code: Any, desc: Any) -> tuple[str, str]:
    return _canonical_material_code(code), _normalize_text(desc)


def _days_on_hand_band(days_on_hand: float | None) -> str:
    if days_on_hand is None:
        return "no-run-rate"
    if days_on_hand < DAYS_ON_HAND_RED:
        return "red"
    if days_on_hand <= DAYS_ON_HAND_AMBER:
        return "amber"
    return "green"


def _summarize_consumption(consumption: pd.DataFrame) -> pd.DataFrame:
    if consumption.empty:
        return pd.DataFrame(
            columns=[
                "material_code",
                "material_desc",
                "actual_uom",
                "monthly_usage_qty_actual",
                "monthly_usage_qty_mt",
                "monthly_usage_value",
                "plant_count",
                "movement_rows",
            ]
        )

    frame = consumption.copy()
    frame["material_code"] = frame["material_code"].map(_canonical_material_code)
    frame["actual_usage_qty"] = pd.to_numeric(
        frame.get("actual_usage_qty", frame.get("usage_qty")),
        errors="coerce",
    )
    frame["usage_value"] = pd.to_numeric(frame.get("usage_value"), errors="coerce")
    frame["usage_qty_mt"] = frame.apply(
        lambda row: _quantity_to_mt(row.get("actual_usage_qty"), row.get("actual_uom") or row.get("uom")),
        axis=1,
    )
    grouped = (
        frame.groupby(_material_key_columns(), as_index=False)
        .agg(
            actual_uom=("actual_uom", _preferred_uom),
            monthly_usage_qty_actual=("actual_usage_qty", "sum"),
            monthly_usage_qty_mt=("usage_qty_mt", "sum"),
            monthly_usage_value=("usage_value", "sum"),
            plant_count=("reporting_plant", "nunique"),
            movement_rows=("material_desc", "size"),
        )
        .sort_values(["monthly_usage_value", "monthly_usage_qty_mt", "material_desc"], ascending=[False, False, True])
    )
    return grouped


def _summarize_mc9(mc9: pd.DataFrame) -> pd.DataFrame:
    if mc9.empty:
        return pd.DataFrame(
            columns=[
                "material_code",
                "material_desc",
                "usage_uom",
                "stock_uom",
                "monthly_usage_qty_actual",
                "monthly_usage_qty_mt",
                "monthly_usage_value",
                "closing_stock_qty_actual",
                "closing_stock_qty_mt",
                "closing_stock_value",
                "plant_count",
                "stock_rows",
            ]
        )

    frame = mc9.copy()
    frame["material_code"] = frame["material_code"].map(_canonical_material_code)
    for column in ["usage_qty", "usage_value", "stock_qty", "stock_value"]:
        frame[column] = pd.to_numeric(frame.get(column), errors="coerce")
    frame["usage_qty_mt"] = frame.apply(
        lambda row: _quantity_to_mt(row.get("usage_qty"), row.get("uom")),
        axis=1,
    )
    frame["stock_qty_mt"] = frame.apply(
        lambda row: _quantity_to_mt(row.get("stock_qty"), row.get("stock_uom")),
        axis=1,
    )
    grouped = (
        frame.groupby(_material_key_columns(), as_index=False)
        .agg(
            usage_uom=("actual_uom", _preferred_uom),
            stock_uom=("actual_stock_uom", _preferred_uom),
            monthly_usage_qty_actual=("actual_usage_qty", "sum"),
            monthly_usage_qty_mt=("usage_qty_mt", "sum"),
            monthly_usage_value=("usage_value", "sum"),
            closing_stock_qty_actual=("actual_stock_qty", "sum"),
            closing_stock_qty_mt=("stock_qty_mt", "sum"),
            closing_stock_value=("stock_value", "sum"),
            plant_count=("reporting_plant", "nunique"),
            stock_rows=("material_desc", "size"),
        )
        .sort_values(["closing_stock_value", "monthly_usage_value", "material_desc"], ascending=[False, False, True])
    )
    return grouped


def _summarize_opening_stock(mc9_all: pd.DataFrame, report_year: int, report_month: int) -> pd.DataFrame:
    if mc9_all.empty:
        return pd.DataFrame(
            columns=[
                "material_code",
                "material_desc",
                "opening_stock_uom",
                "opening_stock_qty_actual",
                "opening_stock_qty_mt",
                "opening_stock_value",
            ]
        )

    previous_year, previous_month = _previous_month(report_year, report_month)
    previous_frame = _filter_month(mc9_all, previous_year, previous_month)
    if previous_frame.empty:
        return pd.DataFrame(
            columns=[
                "material_code",
                "material_desc",
                "opening_stock_uom",
                "opening_stock_qty_actual",
                "opening_stock_qty_mt",
                "opening_stock_value",
            ]
        )

    frame = previous_frame.copy()
    frame["material_code"] = frame["material_code"].map(_canonical_material_code)
    for column in ["stock_qty", "stock_value"]:
        frame[column] = pd.to_numeric(frame.get(column), errors="coerce")
    frame["stock_qty_mt"] = frame.apply(
        lambda row: _quantity_to_mt(row.get("stock_qty"), row.get("stock_uom")),
        axis=1,
    )
    grouped = (
        frame.groupby(_material_key_columns(), as_index=False)
        .agg(
            opening_stock_uom=("actual_stock_uom", _preferred_uom),
            opening_stock_qty_actual=("actual_stock_qty", "sum"),
            opening_stock_qty_mt=("stock_qty_mt", "sum"),
            opening_stock_value=("stock_value", "sum"),
        )
        .sort_values(["opening_stock_value", "material_desc"], ascending=[False, True])
    )
    return grouped


def _summarize_procurement(procurement: pd.DataFrame) -> pd.DataFrame:
    if procurement.empty:
        return pd.DataFrame(
            columns=[
                "material_code",
                "material_desc",
                "order_unit",
                "ordered_qty",
                "procured_qty",
                "open_qty",
                "effective_value",
                "vendor_count",
                "po_count",
            ]
        )

    frame = procurement.copy()
    frame["material_code"] = frame["material_code"].map(_canonical_material_code)
    for column in ["order_qty", "procured_qty", "still_to_be_delivered_qty", "effective_value"]:
        frame[column] = pd.to_numeric(frame.get(column), errors="coerce")
    grouped = (
        frame.groupby(_material_key_columns(), as_index=False)
        .agg(
            order_unit=("order_unit", _preferred_uom),
            ordered_qty=("order_qty", "sum"),
            procured_qty=("procured_qty", "sum"),
            open_qty=("still_to_be_delivered_qty", "sum"),
            effective_value=("effective_value", "sum"),
            vendor_count=("vendor", "nunique"),
            po_count=("po_number", "nunique"),
        )
        .sort_values(["effective_value", "open_qty", "material_desc"], ascending=[False, False, True])
    )
    return grouped


def _build_material_rollup(
    opening_stock_summary: pd.DataFrame,
    consumption_summary: pd.DataFrame,
    mc9_summary: pd.DataFrame,
    procurement_summary: pd.DataFrame,
) -> pd.DataFrame:
    key_columns = _material_key_columns()
    rollup = opening_stock_summary.merge(
        consumption_summary,
        on=key_columns,
        how="outer",
    ).merge(
        mc9_summary,
        on=key_columns,
        how="outer",
        suffixes=("_mb51", "_mc9"),
    ).merge(
        procurement_summary,
        on=key_columns,
        how="outer",
    )

    if rollup.empty:
        return pd.DataFrame(
            columns=[
                "material_code",
                "material_desc",
                "opening_stock_uom",
                "opening_stock_qty_actual",
                "opening_stock_qty_mt",
                "opening_stock_value",
                "consumption_uom",
                "mb51_usage_qty_actual",
                "mb51_usage_qty_mt",
                "mb51_usage_value",
                "mc9_usage_uom",
                "mc9_usage_qty_actual",
                "mc9_usage_qty_mt",
                "mc9_usage_value",
                "stock_uom",
                "closing_stock_qty_actual",
                "closing_stock_qty_mt",
                "closing_stock_value",
                "procurement_uom",
                "ordered_qty",
                "procured_qty",
                "open_qty",
                "procurement_value",
                "vendor_count",
            ]
        )

    rename_map = {
        "actual_uom": "consumption_uom",
        "monthly_usage_qty_actual_mb51": "mb51_usage_qty_actual",
        "monthly_usage_qty_mt_mb51": "mb51_usage_qty_mt",
        "monthly_usage_value_mb51": "mb51_usage_value",
        "usage_uom": "mc9_usage_uom",
        "monthly_usage_qty_actual_mc9": "mc9_usage_qty_actual",
        "monthly_usage_qty_mt_mc9": "mc9_usage_qty_mt",
        "monthly_usage_value_mc9": "mc9_usage_value",
        "order_unit": "procurement_uom",
        "effective_value": "procurement_value",
    }

    # Align MC9 columns after the merge suffixes.
    for source, target in [
        ("monthly_usage_qty_actual_mc9", "mc9_usage_qty_actual"),
        ("monthly_usage_qty_mt_mc9", "mc9_usage_qty_mt"),
        ("monthly_usage_value_mc9", "mc9_usage_value"),
        ("plant_count_mc9", "mc9_plant_count"),
        ("stock_rows", "mc9_row_count"),
    ]:
        if source in rollup.columns:
            rename_map[source] = target

    if "plant_count_mb51" in rollup.columns:
        rename_map["plant_count_mb51"] = "mb51_plant_count"
    if "movement_rows" in rollup.columns:
        rename_map["movement_rows"] = "mb51_row_count"
    if "usage_uom_mc9" in rollup.columns:
        rename_map["usage_uom_mc9"] = "mc9_usage_uom"
    if "stock_uom_mc9" in rollup.columns:
        rename_map["stock_uom_mc9"] = "stock_uom"

    rollup = rollup.rename(columns=rename_map)

    numeric_columns = [
        "opening_stock_qty_actual",
        "opening_stock_qty_mt",
        "opening_stock_value",
        "mb51_usage_qty_actual",
        "mb51_usage_qty_mt",
        "mb51_usage_value",
        "mc9_usage_qty_actual",
        "mc9_usage_qty_mt",
        "mc9_usage_value",
        "closing_stock_qty_actual",
        "closing_stock_qty_mt",
        "closing_stock_value",
        "ordered_qty",
        "procured_qty",
        "open_qty",
        "procurement_value",
        "vendor_count",
        "mb51_plant_count",
        "mc9_plant_count",
    ]
    for column in numeric_columns:
        if column in rollup.columns:
            rollup[column] = pd.to_numeric(rollup[column], errors="coerce").fillna(0.0)

    for column in ["opening_stock_uom", "consumption_uom", "mc9_usage_uom", "stock_uom", "procurement_uom"]:
        if column not in rollup.columns:
            rollup[column] = ""
        rollup[column] = rollup[column].fillna("").astype(str)

    rollup["material_desc"] = rollup["material_desc"].fillna("")
    rollup["material_code"] = rollup["material_code"].fillna("")
    missing_opening_mask = (
        rollup["opening_stock_uom"].fillna("").astype(str).str.strip().eq("")
        & rollup["opening_stock_qty_actual"].eq(0.0)
        & rollup["opening_stock_qty_mt"].eq(0.0)
        & rollup["opening_stock_value"].eq(0.0)
    )
    if missing_opening_mask.any():
        for index in rollup.index[missing_opening_mask]:
            stock_uom = _normalize_text(rollup.at[index, "stock_uom"]).upper()
            consumption_uom = _normalize_text(rollup.at[index, "consumption_uom"]).upper()
            estimated_mt = (
                _safe_float(rollup.at[index, "closing_stock_qty_mt"])
                - _safe_float(rollup.at[index, "mb51_usage_qty_mt"])
            )
            estimated_value = (
                _safe_float(rollup.at[index, "closing_stock_value"])
                - _safe_float(rollup.at[index, "mb51_usage_value"])
            )
            rollup.at[index, "opening_stock_qty_mt"] = estimated_mt
            rollup.at[index, "opening_stock_value"] = estimated_value

            if stock_uom and (not consumption_uom or stock_uom == consumption_uom):
                rollup.at[index, "opening_stock_uom"] = stock_uom
                rollup.at[index, "opening_stock_qty_actual"] = (
                    _safe_float(rollup.at[index, "closing_stock_qty_actual"])
                    - _safe_float(rollup.at[index, "mb51_usage_qty_actual"])
                )
            elif consumption_uom and not stock_uom:
                rollup.at[index, "opening_stock_uom"] = consumption_uom
                rollup.at[index, "opening_stock_qty_actual"] = (
                    _safe_float(rollup.at[index, "closing_stock_qty_actual"])
                    - _safe_float(rollup.at[index, "mb51_usage_qty_actual"])
                )
            else:
                rollup.at[index, "opening_stock_uom"] = "MT"
                rollup.at[index, "opening_stock_qty_actual"] = estimated_mt

    rollup = rollup.sort_values(
        ["mb51_usage_value", "procurement_value", "closing_stock_value", "material_desc"],
        ascending=[False, False, False, True],
    )
    return rollup


def _build_stock_movement_sheet(material_rollup: pd.DataFrame) -> pd.DataFrame:
    movement = material_rollup.loc[
        :,
        [
            "material_code",
            "material_desc",
            "opening_stock_uom",
            "opening_stock_qty_actual",
            "opening_stock_qty_mt",
            "consumption_uom",
            "mb51_usage_qty_actual",
            "mb51_usage_qty_mt",
            "stock_uom",
            "closing_stock_qty_actual",
            "closing_stock_qty_mt",
        ],
    ].copy()
    movement = movement.rename(
        columns={
            "material_code": "Material",
            "material_desc": "Name of Chemical",
            "opening_stock_uom": "Opening Stock UoM",
            "opening_stock_qty_actual": "Opening Stock Qty (Actual)",
            "opening_stock_qty_mt": "Opening Stock Qty (MT)",
            "consumption_uom": "Consumption UoM",
            "mb51_usage_qty_actual": "Consumption Qty (Actual)",
            "mb51_usage_qty_mt": "Consumption Qty (MT)",
            "stock_uom": "Closing Stock UoM",
            "closing_stock_qty_actual": "Closing Stock Qty (Actual)",
            "closing_stock_qty_mt": "Closing Stock Qty (MT)",
        }
    )
    return movement.sort_values(
        ["Consumption Qty (MT)", "Consumption Qty (Actual)", "Name of Chemical"],
        ascending=[False, False, True],
    )


def _top_rows(
    frame: pd.DataFrame,
    value_column: str,
    limit: int = 10,
    *,
    classification_map: dict[tuple[str, str], str] | None = None,
) -> list[dict[str, Any]]:
    if frame.empty or value_column not in frame.columns:
        return []
    rows: list[dict[str, Any]] = []
    for _, row in frame.head(limit).iterrows():
        history_key = _history_key(row.get("material_code"), row.get("material_desc"))
        rows.append(
            {
                "material_code": _normalize_text(row.get("material_code")),
                "material_desc": _normalize_text(row.get("material_desc")),
                "material_label": _material_display(row.get("material_code"), row.get("material_desc")),
                "abc_class": (classification_map or {}).get(history_key, ""),
                value_column: round(_safe_float(row.get(value_column)), 2),
            }
        )
    return rows


def _build_consumption_history(consumption_all: pd.DataFrame) -> pd.DataFrame:
    if consumption_all.empty:
        return pd.DataFrame(
            columns=[
                "material_code",
                "material_desc",
                "year",
                "month",
                "monthly_usage_qty_mt",
                "monthly_usage_value",
            ]
        )

    frame = consumption_all.copy()
    frame["material_code"] = frame["material_code"].map(_canonical_material_code)
    frame["usage_qty_mt"] = frame.apply(
        lambda row: _quantity_to_mt(row.get("actual_usage_qty"), row.get("actual_uom") or row.get("uom")),
        axis=1,
    )
    frame["usage_value"] = pd.to_numeric(frame.get("usage_value"), errors="coerce").fillna(0.0)
    grouped = (
        frame.groupby(_material_key_columns() + ["year", "month"], as_index=False)
        .agg(
            monthly_usage_qty_mt=("usage_qty_mt", "sum"),
            monthly_usage_value=("usage_value", "sum"),
        )
        .sort_values(["material_code", "material_desc", "year", "month"])
    )
    return grouped


def _build_consumption_history_metrics(
    consumption_all: pd.DataFrame,
    report_year: int,
    report_month: int,
) -> pd.DataFrame:
    history = _build_consumption_history(consumption_all)
    if history.empty:
        return pd.DataFrame(
            columns=[
                "material_code",
                "material_desc",
                "history_months",
                "avg_monthly_usage_mt",
                "std_monthly_usage_mt",
                "avg_monthly_usage_value",
                "std_monthly_usage_value",
                "prior_month_usage_mt",
                "prior_month_usage_value",
                "current_month_usage_mt",
                "current_month_usage_value",
                "recent_zero_streak",
            ]
        )

    current_mask = history["year"].eq(report_year) & history["month"].eq(report_month)
    rows: list[dict[str, Any]] = []
    for (material_code, material_desc), frame in history.groupby(_material_key_columns(), dropna=False):
        frame = frame.sort_values(["year", "month"]).copy()
        frame["period_key"] = frame["year"].astype(int) * 100 + frame["month"].astype(int)
        current_frame = frame.loc[current_mask.reindex(frame.index, fill_value=False)]
        prior = frame.loc[frame["period_key"] < report_year * 100 + report_month].copy()
        current_qty = _safe_float(current_frame["monthly_usage_qty_mt"].sum()) if not current_frame.empty else 0.0
        current_value = _safe_float(current_frame["monthly_usage_value"].sum()) if not current_frame.empty else 0.0
        avg_qty = _safe_float(prior["monthly_usage_qty_mt"].mean()) if not prior.empty else 0.0
        std_qty = float(prior["monthly_usage_qty_mt"].std(ddof=0)) if len(prior) > 1 else 0.0
        avg_value = _safe_float(prior["monthly_usage_value"].mean()) if not prior.empty else 0.0
        std_value = float(prior["monthly_usage_value"].std(ddof=0)) if len(prior) > 1 else 0.0
        prior_month_qty = _safe_float(prior["monthly_usage_qty_mt"].iloc[-1]) if not prior.empty else 0.0
        prior_month_value = _safe_float(prior["monthly_usage_value"].iloc[-1]) if not prior.empty else 0.0

        positive_periods = frame.loc[frame["monthly_usage_qty_mt"].fillna(0.0) > 1e-9, ["year", "month"]]
        if current_qty > 1e-9:
            zero_streak = 0
        elif positive_periods.empty:
            zero_streak = max(int(len(prior)) + 1, 1)
        else:
            last_positive = positive_periods.iloc[-1]
            zero_streak = max(
                ((int(report_year) - int(last_positive["year"])) * 12)
                + (int(report_month) - int(last_positive["month"])),
                1,
            )

        rows.append(
            {
                "material_code": material_code,
                "material_desc": material_desc,
                "history_months": int(len(prior)),
                "avg_monthly_usage_mt": round(avg_qty, 2),
                "std_monthly_usage_mt": round(std_qty, 2),
                "avg_monthly_usage_value": round(avg_value, 2),
                "std_monthly_usage_value": round(std_value, 2),
                "prior_month_usage_mt": round(prior_month_qty, 2),
                "prior_month_usage_value": round(prior_month_value, 2),
                "current_month_usage_mt": round(current_qty, 2),
                "current_month_usage_value": round(current_value, 2),
                "recent_zero_streak": int(zero_streak),
            }
        )
    return pd.DataFrame(rows)


def _apply_abc_classification(frame: pd.DataFrame, value_column: str) -> tuple[dict[tuple[str, str], str], dict[str, Any]]:
    if frame.empty or value_column not in frame.columns:
        return {}, {"value_share": {"A": 0.0, "B": 0.0, "C": 0.0}, "material_count": {"A": 0, "B": 0, "C": 0}}

    working = frame.copy()
    working[value_column] = pd.to_numeric(working.get(value_column), errors="coerce").fillna(0.0)
    working = working.sort_values([value_column, "material_desc"], ascending=[False, True]).reset_index(drop=True)
    total_value = _safe_float(working[value_column].sum())
    classification_map: dict[tuple[str, str], str] = {}
    value_by_class = {"A": 0.0, "B": 0.0, "C": 0.0}
    count_by_class = {"A": 0, "B": 0, "C": 0}
    cumulative = 0.0
    for _, row in working.iterrows():
        value = _safe_float(row.get(value_column))
        share = (value / total_value) if total_value > 0 else 0.0
        cumulative += share
        if total_value <= 0:
            abc_class = "C"
        elif cumulative <= 0.80 or count_by_class["A"] == 0:
            abc_class = "A"
        elif cumulative <= 0.95:
            abc_class = "B"
        else:
            abc_class = "C"
        key = _history_key(row.get("material_code"), row.get("material_desc"))
        classification_map[key] = abc_class
        value_by_class[abc_class] += value
        count_by_class[abc_class] += 1

    summary = {
        "value_share": {
            label: round((value_by_class[label] / total_value * 100.0), 1) if total_value > 0 else 0.0
            for label in ("A", "B", "C")
        },
        "material_count": count_by_class,
    }
    return classification_map, summary


def _build_stock_health(
    material_rollup: pd.DataFrame,
    history_metrics: pd.DataFrame,
) -> dict[str, Any]:
    if material_rollup.empty:
        return {
            "band_counts": {"red": 0, "amber": 0, "green": 0, "no-run-rate": 0},
            "red_zone_materials": [],
            "dead_stock_materials": [],
        }

    frame = material_rollup.copy()
    if not history_metrics.empty:
        frame = frame.merge(
            history_metrics[
                [
                    "material_code",
                    "material_desc",
                    "avg_monthly_usage_mt",
                    "recent_zero_streak",
                ]
            ],
            on=_material_key_columns(),
            how="left",
        )
    else:
        frame["avg_monthly_usage_mt"] = 0.0
        frame["recent_zero_streak"] = 0

    band_counts = {"red": 0, "amber": 0, "green": 0, "no-run-rate": 0}
    red_zone_materials: list[dict[str, Any]] = []
    dead_stock_materials: list[dict[str, Any]] = []

    for _, row in frame.iterrows():
        avg_usage_mt = _safe_float(row.get("avg_monthly_usage_mt"))
        closing_stock_mt = _safe_float(row.get("closing_stock_qty_mt"))
        closing_stock_value = _safe_float(row.get("closing_stock_value"))
        days_on_hand = None
        if avg_usage_mt > 0:
            days_on_hand = round((closing_stock_mt / avg_usage_mt) * 30.0, 1)
        band = _days_on_hand_band(days_on_hand)
        band_counts[band] += 1
        material = {
            "material_code": _normalize_text(row.get("material_code")),
            "material_desc": _normalize_text(row.get("material_desc")),
            "days_on_hand": days_on_hand,
            "band": band,
            "closing_stock_value": round(closing_stock_value, 2),
            "closing_stock_qty_mt": round(closing_stock_mt, 2),
            "avg_monthly_usage_mt": round(avg_usage_mt, 2),
        }
        if band == "red":
            red_zone_materials.append(material)
        if (
            closing_stock_value > 0
            and int(_safe_float(row.get("recent_zero_streak"))) >= DEAD_STOCK_MONTHS
        ):
            dead_stock_materials.append(material)

    red_zone_materials.sort(
        key=lambda item: (
            999999.0 if item["days_on_hand"] is None else item["days_on_hand"],
            -_safe_float(item["closing_stock_value"]),
            item["material_desc"],
        )
    )
    dead_stock_materials.sort(
        key=lambda item: (-_safe_float(item["closing_stock_value"]), item["material_desc"])
    )
    return {
        "band_counts": band_counts,
        "red_zone_materials": red_zone_materials[:TOP_SECTION_LIMIT],
        "dead_stock_materials": dead_stock_materials[:TOP_SECTION_LIMIT],
        "dead_stock_value": round(
            sum(_safe_float(item["closing_stock_value"]) for item in dead_stock_materials),
            2,
        ),
    }


def _build_anomalies(
    history_metrics: pd.DataFrame,
    procurement_frame: pd.DataFrame,
) -> list[dict[str, Any]]:
    anomalies: list[dict[str, Any]] = []
    if not history_metrics.empty:
        for _, row in history_metrics.iterrows():
            history_months = int(_safe_float(row.get("history_months")))
            current_qty = _safe_float(row.get("current_month_usage_mt"))
            baseline_qty = _safe_float(row.get("avg_monthly_usage_mt"))
            std_qty = _safe_float(row.get("std_monthly_usage_mt"))
            current_value = _safe_float(row.get("current_month_usage_value"))
            baseline_value = _safe_float(row.get("avg_monthly_usage_value"))
            std_value = _safe_float(row.get("std_monthly_usage_value"))
            zero_streak = int(_safe_float(row.get("recent_zero_streak")))
            label = _material_display(row.get("material_code"), row.get("material_desc"))

            if history_months >= ANOMALY_MIN_HISTORY_MONTHS:
                qty_trigger = baseline_qty + (ANOMALY_STDDEV_THRESHOLD * std_qty)
                qty_spike = (std_qty > 0 and current_qty > qty_trigger) or (
                    std_qty <= 1e-9 and baseline_qty > 0 and current_qty >= baseline_qty * (1.0 + ANOMALY_STDDEV_THRESHOLD)
                )
                if qty_spike:
                    anomalies.append(
                        {
                            "type": "Consumption spike",
                            "material_code": _normalize_text(row.get("material_code")),
                            "material_desc": _normalize_text(row.get("material_desc")),
                            "detail": f"{label} consumed {current_qty:,.2f} MT versus baseline {baseline_qty:,.2f} MT.",
                            "severity_score": round(current_qty - baseline_qty, 2),
                            "delta_pct": round(((current_qty - baseline_qty) / baseline_qty) * 100.0, 1)
                            if baseline_qty > 0
                            else None,
                        }
                    )
                value_trigger = baseline_value + (ANOMALY_STDDEV_THRESHOLD * std_value)
                value_spike = (std_value > 0 and current_value > value_trigger) or (
                    std_value <= 1e-9 and baseline_value > 0 and current_value >= baseline_value * (1.0 + ANOMALY_STDDEV_THRESHOLD)
                )
                if value_spike:
                    anomalies.append(
                        {
                            "type": "Consumption value spike",
                            "material_code": _normalize_text(row.get("material_code")),
                            "material_desc": _normalize_text(row.get("material_desc")),
                            "detail": f"{label} usage value reached {_format_inr_compact(current_value)} against baseline {_format_inr_compact(baseline_value)}.",
                            "severity_score": round(current_value - baseline_value, 2),
                            "delta_pct": round(((current_value - baseline_value) / baseline_value) * 100.0, 1)
                            if baseline_value > 0
                            else None,
                        }
                    )
            if zero_streak >= 1 and baseline_qty > 0 and current_qty <= 1e-9 and history_months > 0:
                anomalies.append(
                    {
                        "type": "Consumption dropped to zero",
                        "material_code": _normalize_text(row.get("material_code")),
                        "material_desc": _normalize_text(row.get("material_desc")),
                        "detail": f"{label} recorded zero MB51 usage this month after averaging {baseline_qty:,.2f} MT.",
                        "severity_score": round(baseline_qty, 2),
                        "delta_pct": -100.0,
                    }
                )

    if not procurement_frame.empty:
        working = procurement_frame.copy()
        working["material_code"] = working["material_code"].map(_canonical_material_code)
        for column in ["effective_value", "order_qty", "still_to_be_delivered_qty"]:
            working[column] = pd.to_numeric(working.get(column), errors="coerce").fillna(0.0)
        working["open_value_est"] = working.apply(
            lambda row: (
                _safe_float(row.get("effective_value"))
                * (
                    _safe_float(row.get("still_to_be_delivered_qty"))
                    / _safe_float(row.get("order_qty"))
                )
                if _safe_float(row.get("order_qty")) > 0
                else (_safe_float(row.get("effective_value")) if _safe_float(row.get("still_to_be_delivered_qty")) > 0 else 0.0)
            ),
            axis=1,
        )
        grouped = (
            working.groupby(_material_key_columns(), as_index=False)
            .agg(
                open_value_est=("open_value_est", "sum"),
                effective_value=("effective_value", "sum"),
            )
        )
        if not history_metrics.empty:
            grouped = grouped.merge(
                history_metrics[
                    [
                        "material_code",
                        "material_desc",
                        "avg_monthly_usage_value",
                    ]
                ],
                on=_material_key_columns(),
                how="left",
            )
        for _, row in grouped.iterrows():
            baseline_value = _safe_float(row.get("avg_monthly_usage_value"))
            current_open_value = _safe_float(row.get("open_value_est"))
            if baseline_value > 0 and current_open_value >= baseline_value * OPEN_ORDER_VALUE_RISK_MULTIPLE:
                label = _material_display(row.get("material_code"), row.get("material_desc"))
                anomalies.append(
                    {
                        "type": "Open order exposure",
                        "material_code": _normalize_text(row.get("material_code")),
                        "material_desc": _normalize_text(row.get("material_desc")),
                        "detail": f"{label} has {_format_inr_compact(current_open_value)} open order exposure against average monthly consumption value {_format_inr_compact(baseline_value)}.",
                        "severity_score": round(current_open_value, 2),
                        "delta_pct": round(((current_open_value - baseline_value) / baseline_value) * 100.0, 1),
                    }
                )

    anomalies.sort(
        key=lambda item: (-_safe_float(item.get("severity_score")), item.get("material_desc", ""))
    )
    return anomalies[:TOP_SECTION_LIMIT]


def _build_open_order_risk(
    procurement_frame: pd.DataFrame,
    history_metrics: pd.DataFrame,
    stock_health: dict[str, Any],
    report_year: int,
    report_month: int,
) -> dict[str, Any]:
    if procurement_frame.empty:
        return {"rows": [], "aged_order_count": 0, "aged_open_value": 0.0}

    working = procurement_frame.copy()
    working["material_code"] = working["material_code"].map(_canonical_material_code)
    for column in ["effective_value", "order_qty", "still_to_be_delivered_qty"]:
        working[column] = pd.to_numeric(working.get(column), errors="coerce").fillna(0.0)
    working["doc_date"] = pd.to_datetime(working.get("doc_date"), errors="coerce")
    report_end = _month_end(report_year, report_month)
    working["pending_days"] = (report_end - working["doc_date"]).dt.days.fillna(0).astype(int)
    working["open_value_est"] = working.apply(
        lambda row: (
            _safe_float(row.get("effective_value"))
            * (
                _safe_float(row.get("still_to_be_delivered_qty"))
                / _safe_float(row.get("order_qty"))
            )
            if _safe_float(row.get("order_qty")) > 0
            else (_safe_float(row.get("effective_value")) if _safe_float(row.get("still_to_be_delivered_qty")) > 0 else 0.0)
        ),
        axis=1,
    )
    grouped = (
        working.groupby(["material_code", "material_desc", "vendor", "po_number"], as_index=False)
        .agg(
            pending_days=("pending_days", "max"),
            open_qty=("still_to_be_delivered_qty", "sum"),
            open_value_est=("open_value_est", "sum"),
        )
    )
    if not history_metrics.empty:
        grouped = grouped.merge(
            history_metrics[
                [
                    "material_code",
                    "material_desc",
                    "avg_monthly_usage_value",
                ]
            ],
            on=_material_key_columns(),
            how="left",
        )
    red_zone_codes = {
        item["material_code"]
        for item in stock_health.get("red_zone_materials", [])
        if _normalize_text(item.get("material_code"))
    }

    rows: list[dict[str, Any]] = []
    for _, row in grouped.iterrows():
        open_qty = _safe_float(row.get("open_qty"))
        open_value = _safe_float(row.get("open_value_est"))
        pending_days = int(_safe_float(row.get("pending_days")))
        if open_qty <= 0 and open_value <= 0:
            continue
        baseline_value = _safe_float(row.get("avg_monthly_usage_value"))
        exposure_multiple = (open_value / baseline_value) if baseline_value > 0 else None
        is_aged = pending_days >= OPEN_ORDER_AGING_DAYS
        is_large = exposure_multiple is not None and exposure_multiple >= OPEN_ORDER_VALUE_RISK_MULTIPLE
        is_stock_sensitive = _normalize_text(row.get("material_code")) in red_zone_codes
        if not any((is_aged, is_large, is_stock_sensitive)):
            continue
        risk_flags = []
        if is_aged:
            risk_flags.append("aged >60d")
        if is_large:
            risk_flags.append("high exposure")
        if is_stock_sensitive:
            risk_flags.append("red stock")
        rows.append(
            {
                "material_code": _normalize_text(row.get("material_code")),
                "material_desc": _normalize_text(row.get("material_desc")),
                "vendor": _normalize_text(row.get("vendor")),
                "po_number": _normalize_text(row.get("po_number")),
                "pending_days": pending_days,
                "open_qty": round(open_qty, 2),
                "open_value": round(open_value, 2),
                "exposure_multiple": round(exposure_multiple, 1) if exposure_multiple is not None else None,
                "risk_flags": ", ".join(risk_flags),
                "risk_score": int(is_aged) + int(is_large) + int(is_stock_sensitive),
            }
        )
    rows.sort(
        key=lambda item: (-int(item["risk_score"]), -_safe_float(item["open_value"]), -int(item["pending_days"]))
    )
    aged_rows = [row for row in rows if row["pending_days"] >= OPEN_ORDER_AGING_DAYS]
    return {
        "rows": rows[:TOP_SECTION_LIMIT],
        "aged_order_count": len(aged_rows),
        "aged_open_value": round(sum(_safe_float(row["open_value"]) for row in aged_rows), 2),
    }


def _load_previous_summary(report_year: int, report_month: int) -> dict[str, Any] | None:
    target_year, target_month = _previous_month(report_year, report_month)
    try:
        batch = InventoryMonthlyUploadBatch.query.filter_by(
            report_year=target_year,
            report_month=target_month,
        ).one_or_none()
        if batch is None:
            batch = (
                InventoryMonthlyUploadBatch.query.filter(
                    (InventoryMonthlyUploadBatch.report_year < report_year)
                    | (
                        (InventoryMonthlyUploadBatch.report_year == report_year)
                        & (InventoryMonthlyUploadBatch.report_month < report_month)
                    )
                )
                .order_by(
                    InventoryMonthlyUploadBatch.report_year.desc(),
                    InventoryMonthlyUploadBatch.report_month.desc(),
                )
                .first()
            )
    except Exception:
        return None
    if batch is None:
        return None
    try:
        return json.loads(batch.summary_json or "{}")
    except Exception:
        return None


def _build_kpi_trends(current_kpis: dict[str, Any], previous_summary: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    metrics = {
        "total_mb51_consumption_value": "MB51 consumption value",
        "total_mb51_consumption_mt": "MB51 consumption MT",
        "total_mc9_usage_value": "MC.9 usage value",
        "total_closing_stock_value": "MC.9 closing stock value",
        "total_procurement_value": "ME2M procurement value",
    }
    previous_kpis = (previous_summary or {}).get("kpis", {})
    previous_label = (previous_summary or {}).get("report_label")
    trends: dict[str, dict[str, Any]] = {}
    for key, label in metrics.items():
        current_value = _safe_float(current_kpis.get(key))
        previous_value = _safe_float(previous_kpis.get(key)) if previous_kpis else 0.0
        if not previous_kpis or key not in previous_kpis:
            trends[key] = {
                "label": label,
                "available": False,
                "direction": "flat",
                "pct_change": None,
                "previous_label": previous_label,
            }
            continue
        delta = current_value - previous_value
        if abs(delta) < 1e-9:
            direction = "flat"
        elif delta > 0:
            direction = "up"
        else:
            direction = "down"
        pct_change = None
        if previous_value > 0:
            pct_change = round((delta / previous_value) * 100.0, 1)
        trends[key] = {
            "label": label,
            "available": True,
            "direction": direction,
            "pct_change": pct_change,
            "previous_label": previous_label,
            "previous_value": round(previous_value, 2),
        }
    return trends


def _build_executive_narrative(summary: dict[str, Any]) -> str:
    kpis = summary["kpis"]
    stock_health = summary["stock_health"]
    anomalies = summary["anomalies"]
    open_order_risk = summary["open_order_risk"]
    abc_summary = summary["abc_summary"]
    top_consumption = summary["top_consumption_by_value"]
    mb51_trend = summary["kpi_trends"].get("total_mb51_consumption_value", {})

    intro = (
        f"Chemical inventory consumption for {summary['report_label']} totalled "
        f"{_format_inr_compact(kpis['total_mb51_consumption_value'])} across "
        f"{kpis['materials']} materials and {kpis['reporting_plants']} reporting plants"
    )
    if mb51_trend.get("available") and mb51_trend.get("pct_change") is not None and mb51_trend.get("previous_label"):
        direction_text = "up" if mb51_trend["pct_change"] > 0 else "down" if mb51_trend["pct_change"] < 0 else "flat"
        intro += f", {direction_text} {abs(_safe_float(mb51_trend['pct_change'])):.1f}% versus {mb51_trend['previous_label']}"
    intro += "."

    details = []
    if top_consumption and kpis["total_mb51_consumption_value"] > 0:
        lead = top_consumption[0]
        share = (_safe_float(lead["monthly_usage_value"]) / _safe_float(kpis["total_mb51_consumption_value"])) * 100.0
        details.append(
            f"{lead['material_desc']} accounted for {share:.1f}% of total MB51 consumption value."
        )
    details.append(
        f"A-class materials contributed {abc_summary['value_share'].get('A', 0.0):.1f}% of monthly consumption value."
    )
    red_count = stock_health["band_counts"].get("red", 0)
    if red_count > 0:
        details.append(
            f"{red_count} materials are in the red stock zone and {len(stock_health.get('dead_stock_materials', []))} materials show dead-stock risk."
        )
    if anomalies:
        details.append(f"Top exception: {anomalies[0]['detail']}")
    if open_order_risk.get("aged_order_count", 0) > 0:
        details.append(
            f"{open_order_risk['aged_order_count']} open orders are aged beyond {OPEN_ORDER_AGING_DAYS} days, with {_format_inr_compact(open_order_risk['aged_open_value'])} still exposed."
        )
    return " ".join([intro] + details).strip()


def _build_summary(
    *,
    report_year: int,
    report_month: int,
    mb51_filename: str,
    me2m_filename: str,
    mc9_filename: str,
    consumption_all: pd.DataFrame,
    consumption_frame: pd.DataFrame,
    procurement_frame: pd.DataFrame,
    mc9_frame: pd.DataFrame,
    consumption_summary: pd.DataFrame,
    procurement_summary: pd.DataFrame,
    mc9_summary: pd.DataFrame,
    material_rollup: pd.DataFrame,
    excluded_mb51_rows: int,
    excluded_mc9_rows: int,
) -> dict[str, Any]:
    report_label = _month_label(report_year, report_month)
    unique_plant_count = len(
        {
            _normalize_text(value)
            for value in pd.concat(
                [
                    consumption_frame.get("reporting_plant", pd.Series(dtype="object")),
                    procurement_frame.get("reporting_plant", pd.Series(dtype="object")),
                    mc9_frame.get("reporting_plant", pd.Series(dtype="object")),
                ],
                axis=0,
            ).tolist()
            if _normalize_text(value)
        }
    )

    kpis = {
        "materials": int(len(material_rollup)),
        "reporting_plants": int(unique_plant_count),
        "vendors": int(procurement_frame.get("vendor", pd.Series(dtype="object")).replace("", pd.NA).dropna().nunique()) if not procurement_frame.empty else 0,
        "total_mb51_consumption_mt": round(_safe_float(consumption_summary.get("monthly_usage_qty_mt", pd.Series(dtype="float")).sum()), 2),
        "total_mb51_consumption_value": round(_safe_float(consumption_summary.get("monthly_usage_value", pd.Series(dtype="float")).sum()), 2),
        "total_mc9_usage_value": round(_safe_float(mc9_summary.get("monthly_usage_value", pd.Series(dtype="float")).sum()), 2),
        "total_closing_stock_value": round(_safe_float(mc9_summary.get("closing_stock_value", pd.Series(dtype="float")).sum()), 2),
        "total_procurement_value": round(_safe_float(procurement_summary.get("effective_value", pd.Series(dtype="float")).sum()), 2),
        "total_open_qty": round(_safe_float(procurement_summary.get("open_qty", pd.Series(dtype="float")).sum()), 2),
    }
    history_metrics = _build_consumption_history_metrics(consumption_all, report_year, report_month)
    abc_map, abc_summary = _apply_abc_classification(consumption_summary, "monthly_usage_value")
    stock_health = _build_stock_health(material_rollup, history_metrics)
    anomalies = _build_anomalies(history_metrics, procurement_frame)
    open_order_risk = _build_open_order_risk(
        procurement_frame,
        history_metrics,
        stock_health,
        report_year,
        report_month,
    )
    previous_summary = _load_previous_summary(report_year, report_month)
    kpi_trends = _build_kpi_trends(kpis, previous_summary)

    top_consumption = _top_rows(
        consumption_summary,
        "monthly_usage_value",
        classification_map=abc_map,
    )
    top_procurement = _top_rows(procurement_summary, "effective_value", classification_map=abc_map)
    top_stock = _top_rows(mc9_summary, "closing_stock_value", classification_map=abc_map)

    observations = [
        f"Coverage this month: {len(material_rollup)} materials across {unique_plant_count} reporting plants.",
        (
            f"Stock health: {stock_health['band_counts'].get('red', 0)} red, "
            f"{stock_health['band_counts'].get('amber', 0)} amber, "
            f"{stock_health['band_counts'].get('green', 0)} green."
        ),
    ]
    if anomalies:
        observations.insert(0, anomalies[0]["detail"])
    if open_order_risk.get("aged_order_count", 0) > 0:
        observations.append(
            f"Aged open orders > {OPEN_ORDER_AGING_DAYS} days: {open_order_risk['aged_order_count']}."
        )

    summary = {
        "report_year": report_year,
        "report_month": report_month,
        "report_label": report_label,
        "source_files": {
            "mb51": mb51_filename,
            "me2m": me2m_filename,
            "mc9": mc9_filename,
        },
        "row_counts": {
            "mb51_in_month": int(len(consumption_frame)),
            "me2m_snapshot": int(len(procurement_frame)),
            "mc9_in_month": int(len(mc9_frame)),
            "mb51_excluded_other_months": int(excluded_mb51_rows),
            "mc9_excluded_other_months": int(excluded_mc9_rows),
        },
        "kpis": kpis,
        "kpi_trends": kpi_trends,
        "top_consumption_by_value": top_consumption,
        "top_procurement_by_value": top_procurement,
        "top_stock_by_value": top_stock,
        "abc_summary": abc_summary,
        "stock_health": stock_health,
        "anomalies": anomalies,
        "open_order_risk": open_order_risk,
        "observations": observations,
    }
    summary["executive_narrative"] = _build_executive_narrative(summary)
    return summary


def _write_sheet_headers(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"


def _autosize_sheet(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 32)


def _append_dataframe_sheet(workbook: Workbook, title: str, frame: pd.DataFrame, headers: list[str] | None = None) -> None:
    sheet = workbook.create_sheet(title[:31])
    if frame.empty:
        _write_sheet_headers(sheet, headers or [])
        if headers:
            sheet.append(["" for _ in headers])
            sheet["A2"] = "No rows available."
        _autosize_sheet(sheet)
        return

    data = frame.copy()
    if headers is None:
        headers = [str(column) for column in data.columns]
    _write_sheet_headers(sheet, headers)
    for _, row in data.iterrows():
        values = []
        for column in data.columns:
            value = row.get(column)
            if isinstance(value, pd.Timestamp):
                values.append(value.strftime("%Y-%m-%d"))
            elif pd.isna(value):
                values.append("")
            else:
                values.append(value)
        sheet.append(values)
    _autosize_sheet(sheet)


def _build_excel_report(
    *,
    summary: dict[str, Any],
    material_rollup: pd.DataFrame,
    opening_stock_summary: pd.DataFrame,
    consumption_summary: pd.DataFrame,
    procurement_summary: pd.DataFrame,
    mc9_summary: pd.DataFrame,
    consumption_frame: pd.DataFrame,
    procurement_frame: pd.DataFrame,
    mc9_frame: pd.DataFrame,
) -> bytes:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview["A1"] = "Inventory Monthly Update"
    overview["A1"].font = Font(size=16, bold=True, color="1F4E78")
    overview["A3"] = "Reporting Month"
    overview["B3"] = summary["report_label"]
    overview["A4"] = "MB51 File"
    overview["B4"] = summary["source_files"]["mb51"]
    overview["A5"] = "ME2M File"
    overview["B5"] = summary["source_files"]["me2m"]
    overview["A6"] = "MC.9 File"
    overview["B6"] = summary["source_files"]["mc9"]

    overview["A8"] = "Key Metrics"
    overview["A8"].fill = SECTION_FILL
    overview["A8"].font = SECTION_FONT
    metric_rows = [
        ("Materials in rollup", summary["kpis"]["materials"]),
        ("Reporting plants", summary["kpis"]["reporting_plants"]),
        ("Vendors", summary["kpis"]["vendors"]),
        ("MB51 monthly consumption (MT)", summary["kpis"]["total_mb51_consumption_mt"]),
        ("MB51 monthly consumption value", summary["kpis"]["total_mb51_consumption_value"]),
        ("MC.9 monthly usage value", summary["kpis"]["total_mc9_usage_value"]),
        ("MC.9 closing stock value", summary["kpis"]["total_closing_stock_value"]),
        ("ME2M procurement value", summary["kpis"]["total_procurement_value"]),
        ("ME2M open quantity", summary["kpis"]["total_open_qty"]),
        ("Red-zone materials", summary["stock_health"]["band_counts"].get("red", 0)),
        ("Dead-stock materials", len(summary["stock_health"].get("dead_stock_materials", []))),
    ]
    start_row = 9
    for offset, (label, value) in enumerate(metric_rows):
        overview.cell(row=start_row + offset, column=1, value=label)
        overview.cell(row=start_row + offset, column=2, value=value)

    obs_row = start_row + len(metric_rows) + 2
    overview.cell(row=obs_row, column=1, value="Executive Observations")
    overview.cell(row=obs_row, column=1).fill = SECTION_FILL
    overview.cell(row=obs_row, column=1).font = SECTION_FONT
    for index, text in enumerate(summary["observations"], start=1):
        overview.cell(row=obs_row + index, column=1, value=f"- {text}")
    narrative_row = obs_row + len(summary["observations"]) + 2
    overview.cell(row=narrative_row, column=1, value="Executive Narrative")
    overview.cell(row=narrative_row, column=1).fill = SECTION_FILL
    overview.cell(row=narrative_row, column=1).font = SECTION_FONT
    overview.cell(row=narrative_row + 1, column=1, value=summary.get("executive_narrative", ""))
    _autosize_sheet(overview)

    material_rollup_sheet = material_rollup.loc[
        :,
        [
            "material_code",
            "material_desc",
            "opening_stock_uom",
            "opening_stock_qty_actual",
            "opening_stock_qty_mt",
            "opening_stock_value",
            "consumption_uom",
            "mb51_usage_qty_actual",
            "mb51_usage_qty_mt",
            "mb51_usage_value",
            "mc9_usage_uom",
            "mc9_usage_qty_actual",
            "mc9_usage_qty_mt",
            "mc9_usage_value",
            "stock_uom",
            "closing_stock_qty_actual",
            "closing_stock_qty_mt",
            "closing_stock_value",
            "procurement_uom",
            "ordered_qty",
            "procured_qty",
            "open_qty",
            "procurement_value",
            "vendor_count",
            "avg_monthly_usage_mt",
            "days_on_hand",
            "stock_band",
            "abc_class",
        ],
    ].rename(
        columns={
            "material_code": "Material",
            "material_desc": "Name of Chemical",
            "opening_stock_uom": "Opening Stock UoM",
            "opening_stock_qty_actual": "Opening Stock Qty (Actual)",
            "opening_stock_qty_mt": "Opening Stock Qty (MT)",
            "opening_stock_value": "Opening Stock Value",
            "consumption_uom": "MB51 UoM",
            "mb51_usage_qty_actual": "MB51 Qty (Actual)",
            "mb51_usage_qty_mt": "MB51 Qty (MT)",
            "mb51_usage_value": "MB51 Value",
            "mc9_usage_uom": "MC.9 Usage UoM",
            "mc9_usage_qty_actual": "MC.9 Usage Qty (Actual)",
            "mc9_usage_qty_mt": "MC.9 Usage Qty (MT)",
            "mc9_usage_value": "MC.9 Usage Value",
            "stock_uom": "Closing Stock UoM",
            "closing_stock_qty_actual": "Closing Stock Qty (Actual)",
            "closing_stock_qty_mt": "Closing Stock Qty (MT)",
            "closing_stock_value": "Closing Stock Value",
            "procurement_uom": "ME2M UoM",
            "ordered_qty": "Ordered Qty",
            "procured_qty": "Procured Qty",
            "open_qty": "Open Qty",
            "procurement_value": "Procurement Value",
            "vendor_count": "Vendor Count",
            "avg_monthly_usage_mt": "Avg Monthly Consumption (MT)",
            "days_on_hand": "Days on Hand",
            "stock_band": "Stock Band",
            "abc_class": "ABC Class",
        }
    )
    _append_dataframe_sheet(workbook, "Material Rollup", material_rollup_sheet)
    _append_dataframe_sheet(workbook, "Stock Movement", _build_stock_movement_sheet(material_rollup))
    _append_dataframe_sheet(workbook, "MB51 Consumption", consumption_summary.rename(columns={
        "material_code": "Material",
        "material_desc": "Name of Chemical",
        "actual_uom": "Unit of Measurement (actual)",
        "monthly_usage_qty_actual": "Monthly Consumption (Actual)",
        "monthly_usage_qty_mt": "Monthly Consumption (MT)",
        "monthly_usage_value": "Monthly Consumption Value",
        "plant_count": "Reporting Plants",
        "movement_rows": "Movement Rows",
    }))
    _append_dataframe_sheet(workbook, "MC9 Stock", mc9_summary.rename(columns={
        "material_code": "Material",
        "material_desc": "Name of Chemical",
        "usage_uom": "Usage UoM",
        "stock_uom": "Stock UoM",
        "monthly_usage_qty_actual": "Monthly Usage (Actual)",
        "monthly_usage_qty_mt": "Monthly Usage (MT)",
        "monthly_usage_value": "Monthly Usage Value",
        "closing_stock_qty_actual": "Closing Stock Qty (Actual)",
        "closing_stock_qty_mt": "Closing Stock Qty (MT)",
        "closing_stock_value": "Closing Stock Value",
        "plant_count": "Reporting Plants",
        "stock_rows": "Rows",
    }))
    _append_dataframe_sheet(workbook, "Opening Stock", opening_stock_summary.rename(columns={
        "material_code": "Material",
        "material_desc": "Name of Chemical",
        "opening_stock_uom": "Unit of Measurement (actual)",
        "opening_stock_qty_actual": "Opening Stock Qty (Actual)",
        "opening_stock_qty_mt": "Opening Stock Qty (MT)",
        "opening_stock_value": "Opening Stock Value",
    }))
    _append_dataframe_sheet(workbook, "ME2M Snapshot", procurement_summary.rename(columns={
        "material_code": "Material",
        "material_desc": "Name of Chemical",
        "order_unit": "Unit of Measurement (actual)",
        "ordered_qty": "Ordered Quantity",
        "procured_qty": "Procured Quantity",
        "open_qty": "Open Quantity",
        "effective_value": "Procurement Value",
        "vendor_count": "Vendor Count",
        "po_count": "PO Count",
    }))

    _append_dataframe_sheet(
        workbook,
        "MB51 Raw",
        consumption_frame.loc[
            :,
            [
                "reporting_plant",
                "plant",
                "material_code",
                "material_desc",
                "posting_date",
                "movement_type",
                "actual_usage_qty",
                "actual_uom",
                "usage_qty",
                "uom",
                "usage_value",
                "storage_location",
                "po_number",
            ],
        ].rename(
            columns={
                "reporting_plant": "Reporting Plant",
                "plant": "Source Plant",
                "material_code": "Material",
                "material_desc": "Name of Chemical",
                "posting_date": "Posting Date",
                "movement_type": "Movement Type",
                "actual_usage_qty": "Qty (Actual)",
                "actual_uom": "Actual UoM",
                "usage_qty": "Qty (Normalized)",
                "uom": "Normalized UoM",
                "usage_value": "Value",
                "storage_location": "Storage Location",
                "po_number": "PO Number",
            }
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "ME2M Raw",
        procurement_frame.loc[
            :,
            [
                "reporting_plant",
                "plant",
                "material_code",
                "material_desc",
                "doc_date",
                "po_number",
                "vendor",
                "order_qty",
                "procured_qty",
                "still_to_be_delivered_qty",
                "order_unit",
                "effective_value",
                "release_indicator",
            ],
        ].rename(
            columns={
                "reporting_plant": "Reporting Plant",
                "plant": "Source Plant",
                "material_code": "Material",
                "material_desc": "Name of Chemical",
                "doc_date": "Document Date",
                "po_number": "PO Number",
                "vendor": "Vendor",
                "order_qty": "Ordered Qty",
                "procured_qty": "Procured Qty",
                "still_to_be_delivered_qty": "Open Qty",
                "order_unit": "Order Unit",
                "effective_value": "Effective Value",
                "release_indicator": "Release Indicator",
            }
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "MC9 Raw",
        mc9_frame.loc[
            :,
            [
                "reporting_plant",
                "plant",
                "material_code",
                "material_desc",
                "posting_date",
                "usage_qty",
                "uom",
                "usage_value",
                "stock_qty",
                "stock_uom",
                "stock_value",
                "storage_location",
            ],
        ].rename(
            columns={
                "reporting_plant": "Reporting Plant",
                "plant": "Source Plant",
                "material_code": "Material",
                "material_desc": "Name of Chemical",
                "posting_date": "Month",
                "usage_qty": "Usage Qty",
                "uom": "Usage UoM",
                "usage_value": "Usage Value",
                "stock_qty": "Closing Stock Qty",
                "stock_uom": "Closing Stock UoM",
                "stock_value": "Closing Stock Value",
                "storage_location": "Storage Location",
            }
        ),
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ── Enterprise Chairman-Level PDF Report (ReportLab) ─────────────────────────
import os as _os

try:
    from reportlab.lib.colors import HexColor as _RLHex, black as _RL_BLACK, white as _RL_WHITE
    from reportlab.lib.enums import TA_CENTER as _TA_C, TA_LEFT as _TA_L, TA_RIGHT as _TA_R
    from reportlab.lib.pagesizes import A4 as _RL_A4
    from reportlab.lib.styles import ParagraphStyle as _RLParaStyle
    from reportlab.lib.units import mm as _RL_MM
    from reportlab.platypus import (
        PageBreak as _RLPageBreak,
        Paragraph as _RLPara,
        SimpleDocTemplate as _RLDoc,
        Spacer as _RLSpacer,
        Table as _RLTable,
        TableStyle as _RLTableStyle,
    )
except ModuleNotFoundError as exc:  # pragma: no cover - depends on env packaging
    _REPORTLAB_IMPORT_ERROR = exc
else:
    _REPORTLAB_IMPORT_ERROR = None

    # ── Brand colours ─────────────────────────────────────────────────────────
    _PDF_NAVY = _RLHex("#0F3B63")
    _PDF_AMBER = _RLHex("#C55A11")
    _PDF_GOLD = _RLHex("#C9A227")
    _PDF_LIGHT = _RLHex("#DCE6F1")
    _PDF_SLATE = _RLHex("#5B6B7A")
    _PDF_GREY = _RLHex("#F5F6F8")
    _PDF_GLINE = _RLHex("#D0D5DD")

    _PDF_PW, _PDF_PH = _RL_A4
    _PDF_ML = 18 * _RL_MM
    _PDF_MR = 18 * _RL_MM
    _PDF_MT = 20 * _RL_MM
    _PDF_MB = 22 * _RL_MM
    _PDF_CW = _PDF_PW - _PDF_ML - _PDF_MR

    _PDF_LOGO = _os.path.abspath(
        _os.path.join(_os.path.dirname(__file__), "..", "..", "..", "..", "ONGC Logo.jpeg")
    )

    def _pdf_ps(name: str, **kw) -> _RLParaStyle:
        return _RLParaStyle(name, **kw)

    # ── Shared paragraph styles ───────────────────────────────────────────────
    _PS_COVER_TITLE = _pdf_ps(
        "CoverTitle", fontName="Helvetica-Bold", fontSize=24, leading=30, textColor=_RL_WHITE, alignment=_TA_C
    )
    _PS_COVER_SUB = _pdf_ps(
        "CoverSub", fontName="Helvetica", fontSize=13, leading=17, textColor=_PDF_LIGHT, alignment=_TA_C, spaceAfter=4
    )
    _PS_COVER_PERIOD = _pdf_ps(
        "CoverPeriod", fontName="Helvetica-Bold", fontSize=18, leading=24, textColor=_PDF_GOLD, alignment=_TA_C
    )
    _PS_SECTION_H = _pdf_ps(
        "SectionH", fontName="Helvetica-Bold", fontSize=13, leading=16, textColor=_PDF_NAVY, spaceBefore=8, spaceAfter=4
    )
    _PS_BODY = _pdf_ps("Body", fontName="Helvetica", fontSize=9, leading=13, textColor=_RL_BLACK)
    _PS_BODY_RIGHT = _pdf_ps(
        "BodyR", fontName="Helvetica", fontSize=9, leading=13, textColor=_RL_BLACK, alignment=_TA_R
    )
    _PS_SMALL = _pdf_ps("Small", fontName="Helvetica", fontSize=7.5, leading=11, textColor=_PDF_SLATE)
    _PS_OBS = _pdf_ps(
        "Obs", fontName="Helvetica", fontSize=9, leading=14, textColor=_RL_BLACK, leftIndent=6, spaceAfter=3
    )
    _PS_TH = _pdf_ps(
        "TH", fontName="Helvetica-Bold", fontSize=8.5, leading=11, textColor=_RL_WHITE, alignment=_TA_C
    )
    _PS_TD = _pdf_ps("TD", fontName="Helvetica", fontSize=8, leading=11, textColor=_RL_BLACK)
    _PS_TDR = _pdf_ps(
        "TDR", fontName="Helvetica", fontSize=8, leading=11, textColor=_RL_BLACK, alignment=_TA_R
    )


def _pdf_rule(color=None) -> _RLTable:
    """Full-width 1.2 pt colour rule."""
    c = color if color is not None else _PDF_NAVY
    return _RLTable([[""]], colWidths=[_PDF_CW], rowHeights=[1.2],
                    style=[("BACKGROUND", (0, 0), (-1, -1), c)])


def _pdf_fmt_inr(v: float) -> str:
    """Format a Rupee value as Cr / L / plain, with ₹ symbol."""
    return _format_inr_compact(v).replace("Rs", "\u20b9", 1)


def _pdf_header_footer(canvas, doc) -> None:
    """Persistent ONGC header band and confidentiality footer on every page."""
    canvas.saveState()
    # ── Header band (navy) ─────────────────────────────────────────────────
    canvas.setFillColor(_PDF_NAVY)
    canvas.rect(0, _PDF_PH - 14 * _RL_MM, _PDF_PW, 14 * _RL_MM, fill=1, stroke=0)
    if _os.path.isfile(_PDF_LOGO):
        try:
            canvas.drawImage(
                _PDF_LOGO,
                _PDF_ML, _PDF_PH - 13 * _RL_MM,
                width=26 * _RL_MM, height=11 * _RL_MM,
                preserveAspectRatio=True, mask="auto",
            )
        except Exception:
            pass
    canvas.setFillColor(_RL_WHITE)
    canvas.setFont("Helvetica-Bold", 9)
    canvas.drawCentredString(_PDF_PW / 2, _PDF_PH - 8 * _RL_MM,
                             "OIL AND NATURAL GAS CORPORATION LIMITED")
    canvas.setFont("Helvetica", 7.5)
    canvas.drawCentredString(_PDF_PW / 2, _PDF_PH - 12 * _RL_MM,
                             "Inventory Status  |  Monthly Executive Report")
    canvas.drawRightString(_PDF_PW - _PDF_MR, _PDF_PH - 9 * _RL_MM, f"Page {doc.page}")
    # ── Footer band ────────────────────────────────────────────────────────
    canvas.setFillColor(_PDF_NAVY)
    canvas.rect(0, 0, _PDF_PW, 9 * _RL_MM, fill=1, stroke=0)
    canvas.setFillColor(_PDF_GOLD)
    canvas.rect(0, 9 * _RL_MM, _PDF_PW, 0.8 * _RL_MM, fill=1, stroke=0)
    canvas.setFillColor(_RL_WHITE)
    canvas.setFont("Helvetica", 6.5)
    canvas.drawString(_PDF_ML, 3.5 * _RL_MM,
                      "CONFIDENTIAL \u2014 FOR INTERNAL MANAGEMENT USE ONLY")
    canvas.drawRightString(
        _PDF_PW - _PDF_MR, 3.5 * _RL_MM,
        f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y')}",
    )
    canvas.restoreState()


def _pdf_cover_page(story: list, summary: dict[str, Any]) -> None:
    story.append(_RLSpacer(1, 28 * _RL_MM))
    t = _RLTable([[_RLPara("INVENTORY STATUS", _PS_COVER_TITLE)]],
                 colWidths=[_PDF_CW])
    t.setStyle(_RLTableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _PDF_NAVY),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(t)
    story.append(_RLSpacer(1, 3 * _RL_MM))
    story.append(_RLPara("Monthly Executive Report", _PS_COVER_SUB))
    story.append(_RLSpacer(1, 5 * _RL_MM))

    period_t = _RLTable([[_RLPara(summary["report_label"], _PS_COVER_PERIOD)]],
                        colWidths=[_PDF_CW])
    period_t.setStyle(_RLTableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _PDF_LIGHT),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(period_t)
    story.append(_RLSpacer(1, 7 * _RL_MM))

    kpis = summary["kpis"]
    lw = 45 * _RL_MM
    org_rows = [
        ["Organisation",  "Oil and Natural Gas Corporation Limited"],
        ["Division",      "Office of Head Corporate Chemistry"],
        ["Report Type",   "Monthly Inventory Intelligence Update"],
        ["Data Sources",  "MB51 (Consumption)  \u00b7  ME2M (Procurement)  \u00b7  MC.9 (Stock)"],
        ["Coverage",      (f"{kpis['materials']} Materials  |  "
                           f"{kpis['reporting_plants']} Plants  |  "
                           f"{kpis['vendors']} Vendors")],
    ]
    org_tbl = _RLTable(org_rows, colWidths=[lw, _PDF_CW - lw])
    org_tbl.setStyle(_RLTableStyle([
        ("FONTNAME",  (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",  (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE",  (0, 0), (-1, -1), 9),
        ("LEADING",   (0, 0), (-1, -1), 14),
        ("TEXTCOLOR", (0, 0), (0, -1), _PDF_NAVY),
        ("TEXTCOLOR", (1, 0), (1, -1), _RL_BLACK),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [_PDF_GREY, _RL_WHITE]),
        ("BOX",       (0, 0), (-1, -1), 0.5, _PDF_GLINE),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, _PDF_GLINE),
    ]))
    story.append(org_tbl)
    story.append(_RLSpacer(1, 8 * _RL_MM))
    story.append(_pdf_rule(_PDF_AMBER))
    story.append(_RLSpacer(1, 4 * _RL_MM))
    story.append(_RLPara(
        "This report is prepared by the Inventory Intelligence system for review. "
        "All values are in Indian Rupees (\u20b9) unless stated.",
        _PS_SMALL,
    ))
    story.append(_RLPageBreak())


def _pdf_kpi_card(label: str, value: str, accent: bool = False) -> _RLTable:
    bg   = _PDF_NAVY if accent else _RL_WHITE
    v_c  = _RL_WHITE if accent else _PDF_NAVY
    l_c  = _PDF_GOLD if accent else _PDF_SLATE
    vps  = _pdf_ps(f"KV_{label[:6]}", fontName="Helvetica-Bold", fontSize=13,
                   leading=17, textColor=v_c, alignment=_TA_C)
    lps  = _pdf_ps(f"KL_{label[:6]}", fontName="Helvetica", fontSize=7.5,
                   leading=10, textColor=l_c, alignment=_TA_C)
    card = _RLTable(
        [[_RLPara(value, vps)], [_RLPara(label, lps)]],
        colWidths=[(_PDF_CW - 6 * _RL_MM) / 4],
    )
    bc = _PDF_NAVY if accent else _PDF_GLINE
    card.setStyle(_RLTableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), bg),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("BOX",           (0, 0), (-1, -1), 0.75, bc),
    ]))
    return card


def _pdf_kpi_row(pairs: list) -> _RLTable:
    cells = [_pdf_kpi_card(lbl, val, acc) for lbl, val, acc in pairs]
    outer = _RLTable([cells], colWidths=[(_PDF_CW + 2 * _RL_MM) / 4] * 4)
    outer.setStyle(_RLTableStyle([
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return outer


def _pdf_compact_table(
    headers: list[str],
    rows: list[list[str]],
    col_widths: list[float],
    *,
    right_align_cols: set[int] | None = None,
) -> _RLTable:
    table_rows = [[_RLPara(header, _PS_TH) for header in headers]]
    for row in rows:
        rendered = []
        for index, value in enumerate(row):
            style = _PS_TDR if right_align_cols and index in right_align_cols else _PS_TD
            rendered.append(_RLPara(str(value), style))
        table_rows.append(rendered)
    table = _RLTable(table_rows, colWidths=col_widths)
    table.setStyle(_RLTableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _PDF_NAVY),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_RL_WHITE, _PDF_GREY]),
        ("GRID", (0, 0), (-1, -1), 0.4, _PDF_GLINE),
        ("LINEBELOW", (0, 0), (-1, 0), 1.2, _PDF_AMBER),
    ]))
    return table


def _pdf_trend_label(trend: dict[str, Any]) -> str:
    if not trend.get("available") or trend.get("pct_change") is None or not trend.get("previous_label"):
        return "Baseline unavailable - upload prior month batch"
    arrow = "↑" if trend.get("direction") == "up" else "↓" if trend.get("direction") == "down" else "→"
    return f"{arrow} {abs(_safe_float(trend.get('pct_change'))):.1f}% vs {trend.get('previous_label')}"


def _pdf_contents_page(story: list) -> None:
    story.append(_RLSpacer(1, 5 * _RL_MM))
    story.append(_RLPara("Contents", _PS_SECTION_H))
    story.append(_pdf_rule(_PDF_NAVY))
    story.append(_RLSpacer(1, 4 * _RL_MM))
    story.append(
        _pdf_compact_table(
            ["Seq", "Section", "Purpose"],
            [
                ["1", "How to Read This Report", "Interpret trend arrows, RAG stock signals, and exception flags."],
                ["2", "Executive Summary", "Headline KPIs and the auto-written management narrative."],
                ["3", "Anomalies & Exceptions", "Highlights unexpected demand changes and abnormal exposures."],
                ["4", "Stock Health", "Shows days-on-hand, red-zone items, and dead-stock build-up."],
                ["5", "Open Order Risk", "Shows aging procurement exposure and unusually large open-order risk."],
                ["6", "Top Materials & ABC Lens", "Ranks current value concentration with ABC classification."],
                ["7", "Source Files & Notes", "Documents the data used and report caveats."],
            ],
            [12 * _RL_MM, 56 * _RL_MM, _PDF_CW - 68 * _RL_MM],
        )
    )
    story.append(_RLPageBreak())


def _pdf_reading_guide_page(story: list) -> None:
    story.append(_RLSpacer(1, 5 * _RL_MM))
    story.append(_RLPara("How to Read This Report", _PS_SECTION_H))
    story.append(_pdf_rule(_PDF_AMBER))
    story.append(_RLSpacer(1, 4 * _RL_MM))
    story.append(
        _RLPara(
            (
                "Start with the executive narrative and KPI movement, then move to exceptions, stock health, and "
                "procurement risk. The report is designed to surface management attention points first, with ranked "
                "tables kept intentionally short."
            ),
            _PS_BODY,
        )
    )
    story.append(_RLSpacer(1, 4 * _RL_MM))
    story.append(
        _pdf_compact_table(
            ["Signal", "Meaning"],
            [
                ["Trend arrows", "Month-on-month movement versus the previous stored monthly batch when available; otherwise backfill the prior month once to unlock comparison."],
                ["Red / Amber / Green", "Red is under 30 days on hand, amber is 30 to 90 days, green is above 90 days."],
                ["Anomalies", f"Triggered when current demand materially deviates from MB51 history or open-order exposure is unusually large."],
                ["Dead stock", f"Positive stock with zero MB51 consumption for {DEAD_STOCK_MONTHS}+ consecutive months."],
                ["ABC class", "A-class materials drive most of the month’s value and warrant tighter control."],
            ],
            [34 * _RL_MM, _PDF_CW - 34 * _RL_MM],
        )
    )
    story.append(_RLPageBreak())


def _pdf_executive_summary_page(story: list, summary: dict[str, Any]) -> None:
    kpis = summary["kpis"]
    trends = summary["kpi_trends"]
    story.append(_RLSpacer(1, 5 * _RL_MM))
    story.append(_RLPara("Executive Summary", _PS_SECTION_H))
    story.append(_pdf_rule(_PDF_NAVY))
    story.append(_RLSpacer(1, 4 * _RL_MM))

    row1 = [
        ("Materials Tracked", str(kpis["materials"]), True),
        ("Reporting Plants", str(kpis["reporting_plants"]), False),
        ("Active Vendors", str(kpis["vendors"]), False),
        (
            "MB51 Consumption (MT)",
            f"{kpis['total_mb51_consumption_mt']:,.2f}<br/><font size='7'>{_pdf_trend_label(trends.get('total_mb51_consumption_mt', {}))}</font>",
            False,
        ),
    ]
    row2 = [
        (
            "MB51 Consumption Value",
            f"{_pdf_fmt_inr(kpis['total_mb51_consumption_value'])}<br/><font size='7'>{_pdf_trend_label(trends.get('total_mb51_consumption_value', {}))}</font>",
            True,
        ),
        (
            "MC.9 Usage Value",
            f"{_pdf_fmt_inr(kpis['total_mc9_usage_value'])}<br/><font size='7'>{_pdf_trend_label(trends.get('total_mc9_usage_value', {}))}</font>",
            False,
        ),
        (
            "MC.9 Closing Stock",
            f"{_pdf_fmt_inr(kpis['total_closing_stock_value'])}<br/><font size='7'>{_pdf_trend_label(trends.get('total_closing_stock_value', {}))}</font>",
            False,
        ),
        (
            "ME2M Procurement",
            f"{_pdf_fmt_inr(kpis['total_procurement_value'])}<br/><font size='7'>{_pdf_trend_label(trends.get('total_procurement_value', {}))}</font>",
            False,
        ),
    ]
    story.append(_pdf_kpi_row(row1))
    story.append(_RLSpacer(1, 3 * _RL_MM))
    story.append(_pdf_kpi_row(row2))
    story.append(_RLSpacer(1, 6 * _RL_MM))
    story.append(_RLPara("Executive Narrative", _PS_SECTION_H))
    story.append(_pdf_rule(_PDF_AMBER))
    story.append(_RLSpacer(1, 3 * _RL_MM))
    story.append(_RLPara(summary.get("executive_narrative", ""), _PS_BODY))
    story.append(_RLSpacer(1, 5 * _RL_MM))
    story.append(_RLPara("Key Signals", _PS_SECTION_H))
    story.append(_pdf_rule(_PDF_NAVY))
    story.append(_RLSpacer(1, 3 * _RL_MM))
    for idx, obs in enumerate(summary["observations"], start=1):
        story.append(_RLPara(f"<b>{idx}.</b>&nbsp; {obs}", _PS_OBS))
    story.append(_RLPageBreak())


def _pdf_anomalies_page(story: list, summary: dict[str, Any]) -> None:
    story.append(_RLSpacer(1, 5 * _RL_MM))
    story.append(_RLPara("Anomalies & Exceptions", _PS_SECTION_H))
    story.append(_pdf_rule(_PDF_NAVY))
    story.append(_RLSpacer(1, 3 * _RL_MM))
    anomalies = summary.get("anomalies", [])
    if not anomalies:
        story.append(_RLPara("No material anomalies were triggered from the available monthly history.", _PS_SMALL))
        story.append(_RLPageBreak())
        return
    story.append(
        _pdf_compact_table(
            ["Type", "Code", "Material", "Delta", "Why it matters"],
            [
                [
                    item.get("type", ""),
                    item.get("material_code", ""),
                    item.get("material_desc", ""),
                    _format_pct(item.get("delta_pct")),
                    item.get("detail", ""),
                ]
                for item in anomalies
            ],
            [30 * _RL_MM, 18 * _RL_MM, 40 * _RL_MM, 18 * _RL_MM, _PDF_CW - 106 * _RL_MM],
        )
    )
    story.append(_RLPageBreak())


def _pdf_stock_health_page(story: list, summary: dict[str, Any]) -> None:
    stock_health = summary["stock_health"]
    band_counts = stock_health["band_counts"]
    story.append(_RLSpacer(1, 5 * _RL_MM))
    story.append(_RLPara("Stock Health & Days-on-Hand", _PS_SECTION_H))
    story.append(_pdf_rule(_PDF_NAVY))
    story.append(_RLSpacer(1, 3 * _RL_MM))
    story.append(
        _RLPara(
            (
                f"RAG distribution: {band_counts.get('red', 0)} red, {band_counts.get('amber', 0)} amber, "
                f"{band_counts.get('green', 0)} green, and {band_counts.get('no-run-rate', 0)} without a reliable run rate. "
                f"Dead-stock exposure totals {_pdf_fmt_inr(stock_health.get('dead_stock_value', 0))}."
            ),
            _PS_BODY,
        )
    )
    story.append(_RLSpacer(1, 4 * _RL_MM))
    if stock_health.get("red_zone_materials"):
        story.append(_RLPara("Red-Zone Materials", _PS_SECTION_H))
        story.append(
            _pdf_compact_table(
                ["Code", "Material", "Days", "Closing Stock", "Avg MT / Month"],
                [
                    [
                        item["material_code"],
                        item["material_desc"],
                        "n/a" if item["days_on_hand"] is None else f"{item['days_on_hand']:,.1f}",
                        _pdf_fmt_inr(item["closing_stock_value"]),
                        f"{item['avg_monthly_usage_mt']:,.2f}",
                    ]
                    for item in stock_health["red_zone_materials"]
                ],
                [18 * _RL_MM, 52 * _RL_MM, 16 * _RL_MM, 24 * _RL_MM, _PDF_CW - 110 * _RL_MM],
                right_align_cols={2, 3, 4},
            )
        )
        story.append(_RLSpacer(1, 4 * _RL_MM))
    if stock_health.get("dead_stock_materials"):
        story.append(_RLPara("Dead Stock Alerts", _PS_SECTION_H))
        story.append(
            _pdf_compact_table(
                ["Code", "Material", "Closing Stock", "Days"],
                [
                    [
                        item["material_code"],
                        item["material_desc"],
                        _pdf_fmt_inr(item["closing_stock_value"]),
                        "n/a" if item["days_on_hand"] is None else f"{item['days_on_hand']:,.1f}",
                    ]
                    for item in stock_health["dead_stock_materials"]
                ],
                [18 * _RL_MM, 60 * _RL_MM, 26 * _RL_MM, _PDF_CW - 104 * _RL_MM],
                right_align_cols={2, 3},
            )
        )
    story.append(_RLPageBreak())


def _pdf_risk_page(story: list, summary: dict[str, Any]) -> None:
    open_order_risk = summary["open_order_risk"]
    story.append(_RLSpacer(1, 5 * _RL_MM))
    story.append(_RLPara("Open Order Risk", _PS_SECTION_H))
    story.append(_pdf_rule(_PDF_NAVY))
    story.append(_RLSpacer(1, 3 * _RL_MM))
    story.append(
        _RLPara(
            f"{open_order_risk.get('aged_order_count', 0)} risky open orders are aged beyond {OPEN_ORDER_AGING_DAYS} days.",
            _PS_BODY,
        )
    )
    story.append(_RLSpacer(1, 4 * _RL_MM))
    if open_order_risk.get("rows"):
        story.append(_RLPara("Open Order Aging & Exposure", _PS_SECTION_H))
        story.append(
            _pdf_compact_table(
                ["Code", "Material", "Vendor", "Days", "Open Value", "Flags"],
                [
                    [
                        item["material_code"],
                        item["material_desc"],
                        item["vendor"],
                        str(item["pending_days"]),
                        _pdf_fmt_inr(item["open_value"]),
                        item["risk_flags"],
                    ]
                    for item in open_order_risk["rows"]
                ],
                [16 * _RL_MM, 42 * _RL_MM, 34 * _RL_MM, 12 * _RL_MM, 22 * _RL_MM, _PDF_CW - 126 * _RL_MM],
                right_align_cols={3, 4},
            )
        )
    story.append(_RLPageBreak())


def _pdf_top_table(title: str, rows: list, value_key: str, value_label: str, story: list) -> None:
    story.append(_RLPara(title, _PS_SECTION_H))
    story.append(_pdf_rule(_PDF_NAVY))
    story.append(_RLSpacer(1, 3 * _RL_MM))
    if not rows:
        story.append(_RLPara("No data available for this section.", _PS_SMALL))
        story.append(_RLSpacer(1, 4 * _RL_MM))
        return
    story.append(
        _pdf_compact_table(
            ["#", "Code", "Description", "ABC", value_label],
            [
                [
                    str(index),
                    row.get("material_code", ""),
                    row.get("material_desc", ""),
                    row.get("abc_class", ""),
                    _pdf_fmt_inr(float(row.get(value_key, 0))),
                ]
                for index, row in enumerate(rows[:10], start=1)
            ],
            [8 * _RL_MM, 20 * _RL_MM, 60 * _RL_MM, 12 * _RL_MM, _PDF_CW - 100 * _RL_MM],
            right_align_cols={4},
        )
    )
    story.append(_RLSpacer(1, 5 * _RL_MM))


def _pdf_top_materials_page(story: list, summary: dict[str, Any]) -> None:
    story.append(_RLSpacer(1, 4 * _RL_MM))
    story.append(
        _RLPara(
            f"A-class materials contributed {summary['abc_summary']['value_share'].get('A', 0.0):.1f}% of MB51 monthly consumption value.",
            _PS_BODY,
        )
    )
    story.append(_RLSpacer(1, 4 * _RL_MM))
    _pdf_top_table(
        "Top 10 Materials \u2014 MB51 Monthly Consumption Value",
        summary["top_consumption_by_value"],
        "monthly_usage_value",
        "Consumption Value",
        story,
    )
    _pdf_top_table(
        "Top 10 Materials \u2014 ME2M Procurement Value",
        summary["top_procurement_by_value"],
        "effective_value",
        "Procurement Value",
        story,
    )
    _pdf_top_table(
        "Top 10 Materials \u2014 MC.9 Closing Stock Value",
        summary["top_stock_by_value"],
        "closing_stock_value",
        "Closing Stock Value",
        story,
    )
    story.append(_RLPageBreak())


def _pdf_source_files_page(story: list, summary: dict[str, Any]) -> None:
    story.append(_RLSpacer(1, 5 * _RL_MM))
    story.append(_RLPara("Source Files & Notes", _PS_SECTION_H))
    story.append(_pdf_rule(_PDF_NAVY))
    story.append(_RLSpacer(1, 3 * _RL_MM))

    rc = summary["row_counts"]
    sf = summary["source_files"]
    story.append(
        _pdf_compact_table(
            ["Source", "File Name", "Rows in Month", "Excluded / Note"],
            [
                ["MB51 \u2014 Consumption", sf.get("mb51", "\u2013"), str(rc.get("mb51_in_month", 0)), str(rc.get("mb51_excluded_other_months", 0))],
                ["ME2M \u2014 Procurement", sf.get("me2m", "\u2013"), str(rc.get("me2m_snapshot", 0)), "Snapshot (all rows)"],
                ["MC.9 \u2014 Stock", sf.get("mc9", "\u2013"), str(rc.get("mc9_in_month", 0)), str(rc.get("mc9_excluded_other_months", 0))],
            ],
            [36 * _RL_MM, _PDF_CW - 84 * _RL_MM, 24 * _RL_MM, 24 * _RL_MM],
        )
    )
    story.append(_RLSpacer(1, 7 * _RL_MM))
    story.append(
        _RLPara(
            "<b>Note:</b> All monetary values are in Indian Rupees (\u20b9). "
            "Consumption data is sourced from SAP MB51 goods movement records, procurement data from ME2M purchase-order snapshots, "
            "and stock data from MC.9 plant stock reports. This report is system-generated for internal management review.",
            _PS_SMALL,
        )
    )


def _build_pdf_report(summary: dict[str, Any]) -> bytes:
    """Render an enterprise Chairman-level PDF using ReportLab Platypus."""
    if _REPORTLAB_IMPORT_ERROR is not None:
        raise ValueError(
            "PDF generation requires the `reportlab` package. Install the project requirements and try again."
        ) from _REPORTLAB_IMPORT_ERROR

    buf = BytesIO()
    doc = _RLDoc(
        buf,
        pagesize=_RL_A4,
        leftMargin=_PDF_ML,
        rightMargin=_PDF_MR,
        topMargin=_PDF_MT + 14 * _RL_MM,   # reserve space for the header band
        bottomMargin=_PDF_MB + 9 * _RL_MM,  # reserve space for the footer band
        title=f"ONGC Inventory Monthly Executive Report \u2014 {summary.get('report_label', '')}",
        author="ONGC Inventory Intelligence System",
        subject="Inventory Monthly Update",
        creator="ONGC Inventory Intelligence",
    )
    story: list = []
    _pdf_cover_page(story, summary)
    _pdf_contents_page(story)
    _pdf_reading_guide_page(story)
    _pdf_executive_summary_page(story, summary)
    _pdf_anomalies_page(story, summary)
    _pdf_stock_health_page(story, summary)
    _pdf_risk_page(story, summary)
    _pdf_top_materials_page(story, summary)
    _pdf_source_files_page(story, summary)
    doc.build(story, onFirstPage=_pdf_header_footer, onLaterPages=_pdf_header_footer)
    return buf.getvalue()


def build_monthly_inventory_report_package(
    *,
    report_year: int,
    report_month: int,
    mb51_bytes: bytes,
    mb51_filename: str,
    me2m_bytes: bytes,
    me2m_filename: str,
    mc9_bytes: bytes,
    mc9_filename: str,
    consumption_history: pd.DataFrame | None = None,
    mc9_history: pd.DataFrame | None = None,
) -> MonthlyInventoryReportPackage:
    _validate_report_period(report_year, report_month)

    mb51_raw = _read_uploaded_seed_dataframe(mb51_bytes, mb51_filename)
    _validate_seed_frame(mb51_raw, "consumption")
    if _detect_frame_variant([str(column) for column in mb51_raw.columns], "consumption") != "mb51":
        raise ValueError("Upload the MB51 consumption export for the monthly update module.")
    consumption_all = _normalize_consumption_frame(mb51_raw)
    consumption_history_all = consumption_history.copy() if consumption_history is not None else consumption_all.copy()
    consumption_frame = _filter_month(consumption_all, report_year, report_month)
    if consumption_frame.empty:
        raise ValueError(
            f"MB51 file does not contain rows for {_month_label(report_year, report_month)}. "
            f"Available months: {_available_month_labels(consumption_all)}."
        )

    me2m_raw = _read_uploaded_seed_dataframe(me2m_bytes, me2m_filename)
    _validate_seed_frame(me2m_raw, "procurement")
    procurement_frame = _normalize_procurement_frame(me2m_raw)
    if procurement_frame.empty:
        raise ValueError("ME2M file does not contain any usable procurement snapshot rows.")

    mc9_raw = _read_uploaded_seed_dataframe(mc9_bytes, mc9_filename)
    _validate_seed_frame(mc9_raw, "mc9")
    if _detect_frame_variant([str(column) for column in mc9_raw.columns], "mc9") != "mc9":
        raise ValueError("Upload the MC.9 stock-and-consumption export for the monthly update module.")
    mc9_all = _normalize_mc9_frame(mc9_raw)
    mc9_history_all = mc9_history.copy() if mc9_history is not None else mc9_all.copy()
    mc9_frame = _filter_month(mc9_all, report_year, report_month)
    if mc9_frame.empty:
        raise ValueError(
            f"MC.9 file does not contain rows for {_month_label(report_year, report_month)}. "
            f"Available months: {_available_month_labels(mc9_all)}."
        )

    opening_stock_summary = _summarize_opening_stock(mc9_history_all, report_year, report_month)
    consumption_summary = _summarize_consumption(consumption_frame)
    procurement_summary = _summarize_procurement(procurement_frame)
    mc9_summary = _summarize_mc9(mc9_frame)
    material_rollup = _build_material_rollup(opening_stock_summary, consumption_summary, mc9_summary, procurement_summary)

    summary = _build_summary(
        report_year=report_year,
        report_month=report_month,
        mb51_filename=mb51_filename,
        me2m_filename=me2m_filename,
        mc9_filename=mc9_filename,
        consumption_all=consumption_history_all,
        consumption_frame=consumption_frame,
        procurement_frame=procurement_frame,
        mc9_frame=mc9_frame,
        consumption_summary=consumption_summary,
        procurement_summary=procurement_summary,
        mc9_summary=mc9_summary,
        material_rollup=material_rollup,
        excluded_mb51_rows=max(len(consumption_history_all) - len(consumption_frame), 0),
        excluded_mc9_rows=max(len(mc9_history_all) - len(mc9_frame), 0),
    )
    history_metrics = _build_consumption_history_metrics(consumption_history_all, report_year, report_month)
    abc_map, _ = _apply_abc_classification(consumption_summary, "monthly_usage_value")
    if not history_metrics.empty:
        material_rollup = material_rollup.merge(
            history_metrics[
                [
                    "material_code",
                    "material_desc",
                    "avg_monthly_usage_mt",
                ]
            ],
            on=_material_key_columns(),
            how="left",
        )
    else:
        material_rollup["avg_monthly_usage_mt"] = 0.0
    material_rollup["days_on_hand"] = material_rollup.apply(
        lambda row: round((_safe_float(row.get("closing_stock_qty_mt")) / _safe_float(row.get("avg_monthly_usage_mt")) * 30.0), 1)
        if _safe_float(row.get("avg_monthly_usage_mt")) > 0
        else None,
        axis=1,
    )
    material_rollup["stock_band"] = material_rollup["days_on_hand"].map(_days_on_hand_band)
    material_rollup["abc_class"] = material_rollup.apply(
        lambda row: abc_map.get(_history_key(row.get("material_code"), row.get("material_desc")), ""),
        axis=1,
    )

    token = _month_token(report_year, report_month)
    excel_bytes = _build_excel_report(
        summary=summary,
        material_rollup=material_rollup,
        opening_stock_summary=opening_stock_summary,
        consumption_summary=consumption_summary,
        procurement_summary=procurement_summary,
        mc9_summary=mc9_summary,
        consumption_frame=consumption_frame,
        procurement_frame=procurement_frame,
        mc9_frame=mc9_frame,
    )
    pdf_bytes = _build_pdf_report(summary)

    return MonthlyInventoryReportPackage(
        report_year=report_year,
        report_month=report_month,
        report_label=summary["report_label"],
        mb51_filename=mb51_filename,
        me2m_filename=me2m_filename,
        mc9_filename=mc9_filename,
        mb51_row_count=int(len(consumption_frame)),
        me2m_row_count=int(len(procurement_frame)),
        mc9_row_count=int(len(mc9_frame)),
        excel_filename=f"inventory-monthly-report-{token}.xlsx",
        excel_bytes=excel_bytes,
        pdf_filename=f"inventory-monthly-executive-summary-{token}.pdf",
        pdf_bytes=pdf_bytes,
        summary=summary,
    )


def upsert_monthly_inventory_upload_batch(
    *,
    package: MonthlyInventoryReportPackage,
    uploaded_by: int | None,
) -> tuple[InventoryMonthlyUploadBatch, str]:
    batch = InventoryMonthlyUploadBatch.query.filter_by(
        report_year=package.report_year,
        report_month=package.report_month,
    ).one_or_none()
    action = "created"
    now = datetime.now(timezone.utc)
    if batch is None:
        batch = InventoryMonthlyUploadBatch(
            report_year=package.report_year,
            report_month=package.report_month,
        )
        action = "created"
        db_session_add = True
    else:
        action = "replaced"
        db_session_add = False

    batch.report_label = package.report_label
    batch.status = "success"
    batch.error_message = None
    batch.mb51_filename = package.mb51_filename
    batch.me2m_filename = package.me2m_filename
    batch.mc9_filename = package.mc9_filename
    batch.mb51_row_count = package.mb51_row_count
    batch.me2m_row_count = package.me2m_row_count
    batch.mc9_row_count = package.mc9_row_count
    batch.summary_json = json.dumps(package.summary, ensure_ascii=True)
    batch.excel_filename = package.excel_filename
    batch.excel_content_type = XLSX_MIME
    batch.excel_file_size = len(package.excel_bytes)
    batch.excel_data = package.excel_bytes
    batch.pdf_filename = package.pdf_filename
    batch.pdf_content_type = PDF_MIME
    batch.pdf_file_size = len(package.pdf_bytes)
    batch.pdf_data = package.pdf_bytes
    batch.uploaded_by = uploaded_by
    batch.updated_at = now
    if not batch.uploaded_at:
        batch.uploaded_at = now

    if db_session_add:
        db.session.add(batch)
    return batch, action


def list_monthly_inventory_upload_batches(limit: int = 18) -> list[dict[str, Any]]:
    batches = (
        InventoryMonthlyUploadBatch.query.order_by(
            InventoryMonthlyUploadBatch.report_year.desc(),
            InventoryMonthlyUploadBatch.report_month.desc(),
            InventoryMonthlyUploadBatch.updated_at.desc(),
        )
        .limit(limit)
        .all()
    )
    results: list[dict[str, Any]] = []
    for batch in batches:
        try:
            summary = json.loads(batch.summary_json or "{}")
        except Exception:
            summary = {}
        results.append(
            {
                "id": batch.id,
                "report_year": batch.report_year,
                "report_month": batch.report_month,
                "report_label": batch.report_label,
                "mb51_filename": batch.mb51_filename,
                "me2m_filename": batch.me2m_filename,
                "mc9_filename": batch.mc9_filename,
                "mb51_row_count": batch.mb51_row_count,
                "me2m_row_count": batch.me2m_row_count,
                "mc9_row_count": batch.mc9_row_count,
                "uploaded_at": batch.uploaded_at,
                "updated_at": batch.updated_at,
                "summary": summary,
            }
        )
    return results


def get_monthly_inventory_upload_batch(
    batch_id: int,
    *,
    include_excel: bool = False,
    include_pdf: bool = False,
) -> InventoryMonthlyUploadBatch | None:
    query = InventoryMonthlyUploadBatch.query
    if include_excel:
        query = query.options(undefer(InventoryMonthlyUploadBatch.excel_data))
    if include_pdf:
        query = query.options(undefer(InventoryMonthlyUploadBatch.pdf_data))
    return query.filter_by(id=batch_id).one_or_none()
