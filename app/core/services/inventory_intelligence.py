"""Inventory Intelligence seed-data analytics engine."""

from __future__ import annotations

import calendar
import json
import logging
import os
import pickle
import re
import tempfile
from datetime import date
from io import BytesIO

try:
    import numpy as np
    import pandas as pd
except ImportError:  # pragma: no cover – present only when ENABLE_INVENTORY=False
    np = None  # type: ignore[assignment]
    pd = None  # type: ignore[assignment]
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, inspect

from app.extensions import db
from app.core.services.inventory_forecast import (
    fit_holt_winters,
    forecast_hw,
    fit_seasonal_naive,
    forecast_seasonal_naive,
    fit_wma as fit_wma_exp,
    forecast_wma as forecast_wma_exp,
    hw_forecast_list,
    hw_one_step,
    seasonal_naive_forecast_list,
    seasonal_naive_one_step,
    wma_exp_forecast_list,
    wma_exp_one_step,
    bootstrap_ci,
    select_best_model,
    compute_bootstrap_bands,
    walk_forward_validate,
    MIN_NZ_HW,
    MIN_NZ_SEAS,
    SEAS_PERIOD as _FC_SEAS_PERIOD,
)

logger = logging.getLogger(__name__)

_BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
CONS_FILE = os.path.join(_BASE_DIR, "Consumption_data.xlsx")
PROC_FILE = os.path.join(_BASE_DIR, "Procurement_data.xlsx")
MC9_FILE = os.path.join(_BASE_DIR, "MC9_data.xlsx")
PLANT_GROUPING_FILE = os.path.join(_BASE_DIR, "inventory_plant_grouping.json")
_CACHE_DIR = os.path.join(_BASE_DIR, ".cache", "inventory")
_FORECAST_CACHE_FILE = os.path.join(_CACHE_DIR, "forecast_precomputed.pkl")

_PERIOD_RE = re.compile(r"^\s*(\d{1,2})\.(\d{4})\s*$")
_PRICE_SWING_THRESHOLD_PCT = 15.0
_FORECAST_WEIGHT_PATTERNS = [
    (0.35, 0.25, 0.15, 0.10, 0.10, 0.05),
    (0.30, 0.25, 0.18, 0.12, 0.10, 0.05),
    (0.25, 0.22, 0.18, 0.15, 0.12, 0.08),
    (0.20, 0.20, 0.18, 0.16, 0.14, 0.12),
    (1 / 6, 1 / 6, 1 / 6, 1 / 6, 1 / 6, 1 / 6),
]
_CONSUMPTION_MOVEMENT_SIGNS = {
    "201": 1.0,
    "221": 1.0,
    "261": 1.0,
    "202": -1.0,
    "222": -1.0,
    "262": -1.0,
}
_QUANTITY_UNIT_FACTORS = {
    "G": ("KG", 0.001),
    "GM": ("KG", 0.001),
    "GRAM": ("KG", 0.001),
    "GRAMS": ("KG", 0.001),
    "KG": ("KG", 1.0),
    "KGS": ("KG", 1.0),
    "MT": ("KG", 1000.0),
    "T": ("KG", 1000.0),
    "TON": ("KG", 1000.0),
    "TONS": ("KG", 1000.0),
    "TONNE": ("KG", 1000.0),
    "TONNES": ("KG", 1000.0),
    "KL": ("L", 1000.0),
    "KLS": ("L", 1000.0),
    "L": ("L", 1.0),
    "LT": ("L", 1.0),
    "LTR": ("L", 1.0),
    "LTRS": ("L", 1.0),
    "LTS": ("L", 1.0),
    "ML": ("L", 0.001),
}
_REPORTING_PLANT_GROUPS = {
    "11A1": {"11A1", "11F1", "11F2", "11F3", "11F4", "11F5", "11F9", "11FA"},
    "12A1": {"12A1", "12F1", "12F5"},
    "13A1": {"13A1", "13F1", "13F2", "13F3", "13F9", "13FB", "13FC", "13U9"},
    "22A1": {"22A1", "22W1"},
    "25A1": {"25A1", "25W1"},
    "50A1": {"50A1", "50W1"},
    "51A1": {"51A1", "51W1"},
}
_REPORTING_PLANT_LOOKUP = {
    child: parent
    for parent, children in _REPORTING_PLANT_GROUPS.items()
    for child in children
}
_LEGACY_CONSUMPTION_ALIASES = {
    "plant": "plant",
    "material": "material",
    "storage location": "storage_location",
    "month": "month_raw",
    "usage quantity": "usage_qty",
    "usage uom": "uom",
    "usage value": "usage_value",
    "usage currency": "currency",
    "currency": "currency",
}
_MB51_CONSUMPTION_ALIASES = {
    "material": "material_code",
    "plant": "plant",
    "storage location": "storage_location",
    "movement type": "movement_type",
    "material document": "material_document",
    "base unit of measure": "base_uom",
    "material doc item": "material_document_item",
    "material doc item": "material_document_item",
    "material description": "material_desc",
    "posting date": "posting_date",
    "qty in unit of entry": "entry_qty",
    "unit of entry": "uom",
    "amt in loc cur": "usage_value",
    "receiving plant": "receiving_plant",
    "purchase order": "po_number",
}
_LEGACY_PROCUREMENT_ALIASES = {
    "plant": "plant",
    "net price": "net_price",
    "purchasing document": "po_number",
    "document date": "doc_date",
    "material": "material_code",
    "item": "item",
    "supplier supplying plant": "vendor",
    "short text": "material_desc",
    "order quantity": "order_qty",
    "still to be delivered qty": "still_to_be_delivered_qty",
    "still to be delivered quantity": "still_to_be_delivered_qty",
    "still to be delivered": "still_to_be_delivered_qty",
    "open quantity": "still_to_be_delivered_qty",
    "open qty": "still_to_be_delivered_qty",
    "order unit": "order_unit",
    "currency": "currency",
    "price unit": "price_unit",
    "effective value": "effective_value",
    "release indicator": "release_indicator",
}
_ME2M_PROCUREMENT_ALIASES = dict(_LEGACY_PROCUREMENT_ALIASES)
_MC9_ALIASES = {
    "plant": "plant_label",
    "material": "material_label",
    "storage location": "storage_location",
    "mrp controller": "mrp_controller",
    "mrp type": "mrp_type",
    "material type": "material_type",
    "material group": "material_group",
    "business area": "business_area",
    "division": "division",
    "month": "month_raw",
    "valstckval": "stock_value",
    "valstckval 1": "stock_value_currency",
    "val stock": "stock_qty",
    "val stock 1": "stock_uom",
    "tot usage": "usage_qty",
    "tot usage 1": "uom",
    "tot us val": "usage_value",
    "tot us val 1": "currency",
}
_CANONICAL_CONSUMPTION_COLUMNS = [
    "Material",
    "Plant",
    "Storage Location",
    "Movement Type",
    "Material Document",
    "Base Unit of Measure",
    "Material Doc.Item",
    "Material Description",
    "Posting Date",
    "Qty in unit of entry",
    "Unit of Entry",
    "Amt.in Loc.Cur.",
    "Receiving plant",
    "Purchase order",
]
_CANONICAL_PROCUREMENT_COLUMNS = [
    "Plant",
    "Net Price",
    "Purchasing Document",
    "Document Date",
    "Material",
    "Item",
    "Supplier/Supplying Plant",
    "Short Text",
    "Order Quantity",
    "Still to be delivered (qty)",
    "Order Unit",
    "Currency",
    "Price Unit",
    "Effective value",
    "Release indicator",
]
_CANONICAL_MC9_COLUMNS = [
    "Plant",
    "Material",
    "Storage Location",
    "MRP Controller",
    "MRP Type",
    "Material Type",
    "Material Group",
    "Business Area",
    "Division",
    "Month",
    "ValStckVal",
    "ValStckVal.1",
    "Val. stock",
    "Val. stock.1",
    "Tot. usage",
    "Tot. usage.1",
    "Tot.us.val",
    "Tot.us.val.1",
]
_DEFAULT_PLANT_GROUPING_CONFIG = {
    "prefix_groups": {
        "11": "11A1",
        "12": "12A1",
        "13": "13A1",
    },
    "explicit_groups": {
        "22W1": "22A1",
        "25W1": "25A1",
        "50W1": "50A1",
        "51W1": "51A1",
    },
    "group_members": {
        "11A1": ["11A1", "11F1", "11F2", "11F3", "11F4", "11F5", "11F9", "11FA"],
        "12A1": ["12A1", "12F1", "12F5"],
        "13A1": ["13A1", "13F1", "13F2", "13F3", "13F9", "13FB", "13FC", "13U9"],
        "22A1": ["22A1", "22W1"],
        "25A1": ["25A1", "25W1"],
        "50A1": ["50A1", "50W1"],
        "51A1": ["51A1", "51W1"],
    },
}
_PLANT_GROUPING_CACHE = {
    "mtime": None,
    "config": None,
}
_DB_SEED_TABLES_READY: bool | None = None

# ---------------------------------------------------------------------------
# Adaptive multi-strategy forecasting engine
# ---------------------------------------------------------------------------
# Material demand profiles: each needs a different prediction strategy.
#   - STEADY:   CV < 0.5, few zeros  -> WMA / exponential smoothing
#   - VARIABLE: CV 0.5-1.0           -> ensemble of WMA + seasonal decomp
#   - SPORADIC: many zero months     -> Croston's method (separate demand
#                                       occurrence & demand size models)
#   - LUMPY:    high CV + zeros      -> Syntetos-Boylan Approximation (SBA)
# ---------------------------------------------------------------------------

_ZERO_FRACTION_THRESHOLD = 0.30   # >30% zeros = intermittent demand
_HIGH_CV_THRESHOLD = 0.50         # coefficient of variation threshold
_MIN_HISTORY_MONTHS = 7           # unchanged from original
_MIN_RECENT_ACTIVE = 3            # relaxed from 5 to handle sporadic items
_SEASONAL_CYCLE = 12              # months in a seasonal cycle


def _classify_demand(quantities: np.ndarray) -> str:
    """Classify demand pattern into steady / variable / sporadic / lumpy."""
    if len(quantities) < 4:
        return "insufficient"
    nonzero = quantities[quantities > 0]
    zero_frac = 1.0 - (len(nonzero) / len(quantities))
    cv = float(np.std(nonzero) / np.mean(nonzero)) if len(nonzero) > 1 and np.mean(nonzero) > 0 else 0.0
    if zero_frac > _ZERO_FRACTION_THRESHOLD and cv > _HIGH_CV_THRESHOLD:
        return "lumpy"
    if zero_frac > _ZERO_FRACTION_THRESHOLD:
        return "sporadic"
    if cv > _HIGH_CV_THRESHOLD:
        return "variable"
    return "steady"


def _exponential_smoothing(quantities: np.ndarray, alpha: float = 0.3) -> np.ndarray:
    """Simple exponential smoothing with optimisable alpha."""
    n = len(quantities)
    smoothed = np.zeros(n)
    smoothed[0] = quantities[0]
    for i in range(1, n):
        smoothed[i] = alpha * quantities[i] + (1 - alpha) * smoothed[i - 1]
    return smoothed


def _double_exponential_smoothing(
    quantities: np.ndarray, alpha: float = 0.3, beta: float = 0.1
) -> tuple[np.ndarray, np.ndarray]:
    """Holt's linear trend method."""
    n = len(quantities)
    level = np.zeros(n)
    trend = np.zeros(n)
    level[0] = quantities[0]
    trend[0] = quantities[1] - quantities[0] if n > 1 else 0.0
    for i in range(1, n):
        level[i] = alpha * quantities[i] + (1 - alpha) * (level[i - 1] + trend[i - 1])
        trend[i] = beta * (level[i] - level[i - 1]) + (1 - beta) * trend[i - 1]
    return level, trend


def _seasonal_indices(quantities: np.ndarray, cycle: int = 12) -> np.ndarray:
    """Compute multiplicative seasonal indices from full cycles."""
    n = len(quantities)
    if n < cycle:
        return np.ones(cycle)
    full_cycles = n // cycle
    usable = quantities[: full_cycles * cycle].reshape(full_cycles, cycle)
    row_means = usable.mean(axis=1, keepdims=True)
    row_means[row_means == 0] = 1.0  # avoid division by zero
    ratios = usable / row_means
    indices = ratios.mean(axis=0)
    # normalise so they sum to cycle
    total = indices.sum()
    if total > 0:
        indices = indices * (cycle / total)
    indices[indices == 0] = 0.01  # floor to prevent multiply-by-zero
    return indices


def _holt_winters_additive(
    quantities: np.ndarray, cycle: int = 12, alpha: float = 0.3,
    beta: float = 0.1, gamma: float = 0.2, horizon: int = 6,
) -> list[float]:
    """Holt-Winters additive seasonal model with forecast."""
    n = len(quantities)
    if n < cycle + 2:
        # fall back to simple ES forecast
        sm = _exponential_smoothing(quantities, alpha)
        return [max(float(sm[-1]), 0.0)] * horizon

    # initialise
    level = np.zeros(n)
    trend = np.zeros(n)
    season = np.zeros(n + horizon)

    # initial seasonal component from first cycle
    first_cycle_mean = quantities[:cycle].mean() or 1.0
    for i in range(cycle):
        season[i] = quantities[i] - first_cycle_mean

    level[0] = first_cycle_mean
    trend[0] = (quantities[cycle:2 * cycle].mean() - quantities[:cycle].mean()) / cycle if n >= 2 * cycle else 0.0

    for i in range(1, n):
        l_prev = level[i - 1]
        t_prev = trend[i - 1]
        s_prev = season[i - cycle] if i >= cycle else season[i]
        level[i] = alpha * (quantities[i] - s_prev) + (1 - alpha) * (l_prev + t_prev)
        trend[i] = beta * (level[i] - l_prev) + (1 - beta) * t_prev
        season[i] = gamma * (quantities[i] - level[i]) + (1 - gamma) * s_prev

    forecasts = []
    for h in range(1, horizon + 1):
        s_idx = n - cycle + ((h - 1) % cycle)
        fcast = level[n - 1] + h * trend[n - 1] + season[s_idx]
        forecasts.append(max(float(fcast), 0.0))
    return forecasts


def _croston_forecast(
    quantities: np.ndarray, alpha: float = 0.15, horizon: int = 6,
) -> list[float]:
    """Croston's method for intermittent demand."""
    demands = []
    intervals = []
    period_since_last = 0
    for val in quantities:
        period_since_last += 1
        if val > 0:
            demands.append(val)
            intervals.append(period_since_last)
            period_since_last = 0

    if not demands:
        return [0.0] * horizon

    # smooth demand sizes and intervals separately
    z = float(demands[0])  # demand size estimate
    p = float(intervals[0])  # inter-demand interval estimate
    for i in range(1, len(demands)):
        z = alpha * demands[i] + (1 - alpha) * z
        p = alpha * intervals[i] + (1 - alpha) * p

    p = max(p, 1.0)
    forecast_per_period = z / p
    return [max(float(forecast_per_period), 0.0)] * horizon


def _sba_forecast(
    quantities: np.ndarray, alpha: float = 0.15, horizon: int = 6,
) -> list[float]:
    """Syntetos-Boylan Approximation – bias-corrected Croston's."""
    demands = []
    intervals = []
    period_since_last = 0
    for val in quantities:
        period_since_last += 1
        if val > 0:
            demands.append(val)
            intervals.append(period_since_last)
            period_since_last = 0

    if not demands:
        return [0.0] * horizon

    z = float(demands[0])
    p = float(intervals[0])
    for i in range(1, len(demands)):
        z = alpha * demands[i] + (1 - alpha) * z
        p = alpha * intervals[i] + (1 - alpha) * p

    p = max(p, 1.0)
    # SBA bias correction factor
    forecast_per_period = (1 - alpha / 2) * (z / p)
    return [max(float(forecast_per_period), 0.0)] * horizon


def _optimize_alpha(quantities: np.ndarray, method: str, alphas: list[float] | None = None) -> tuple[float, float]:
    """Find best smoothing parameter via walk-forward validation."""
    if alphas is None:
        alphas = [0.05, 0.1, 0.15, 0.2, 0.25, 0.3, 0.4, 0.5, 0.6, 0.7]
    best_alpha = 0.3
    best_mape = float("inf")
    n = len(quantities)
    train_end = max(n - 6, int(n * 0.7))
    if train_end < 4:
        return best_alpha, best_mape

    for a in alphas:
        errors = []
        for t in range(train_end, n):
            train = quantities[:t]
            actual = float(quantities[t])
            if method == "ses":
                sm = _exponential_smoothing(train, alpha=a)
                pred = float(sm[-1])
            elif method == "croston":
                preds = _croston_forecast(train, alpha=a, horizon=1)
                pred = preds[0]
            elif method == "sba":
                preds = _sba_forecast(train, alpha=a, horizon=1)
                pred = preds[0]
            else:
                pred = float(np.mean(train[-6:])) if len(train) >= 6 else float(np.mean(train))
            if actual > 0:
                errors.append(abs(actual - pred) / actual)
            else:
                errors.append(0.0 if pred < 0.5 else 1.0)
        mape = float(np.mean(errors)) * 100 if errors else float("inf")
        if mape < best_mape:
            best_mape = mape
            best_alpha = a
    return best_alpha, best_mape


def _ensemble_forecast(
    quantities: np.ndarray, labels: list[str], horizon: int = 6,
) -> tuple[list[float], dict]:
    """Adaptive ensemble: classifies demand, picks best strategy, ensembles top-2."""
    demand_type = _classify_demand(quantities)
    n = len(quantities)
    candidates: dict[str, list[float]] = {}
    meta: dict = {"demand_type": demand_type, "candidate_methods": []}

    # Candidate 1: Best WMA (original logic)
    if n >= _MIN_HISTORY_MONTHS:
        best_weights, best_mae, _, _, _, _ = _backtest_forecast_weight_patterns(quantities, labels)
        series = list(quantities.astype(float))
        wma_forecasts: list[float] = []
        lookback = len(best_weights)
        for _ in range(horizon):
            window = np.asarray(series[-lookback:][::-1], dtype=float)
            predicted = max(_weighted_moving_average(window, best_weights), 0.0)
            wma_forecasts.append(predicted)
            series.append(predicted)
        candidates["wma"] = wma_forecasts

    # Candidate 2: Optimised exponential smoothing
    best_alpha_ses, _ = _optimize_alpha(quantities, "ses")
    sm = _exponential_smoothing(quantities, alpha=best_alpha_ses)
    candidates["ses"] = [max(float(sm[-1]), 0.0)] * horizon

    # Candidate 3: Holt's (trend)
    if n >= 6:
        level, trend = _double_exponential_smoothing(quantities, alpha=0.3, beta=0.1)
        holt_forecasts = [max(float(level[-1] + (h + 1) * trend[-1]), 0.0) for h in range(horizon)]
        candidates["holt"] = holt_forecasts

    # Candidate 4: Optimised Holt-Winters (L-BFGS-B / coordinate-descent)
    nz_count = int((quantities > 0).sum())
    if nz_count >= MIN_NZ_HW and n >= _SEASONAL_CYCLE + 2:
        try:
            candidates["holt_winters"] = hw_forecast_list(quantities, horizon)
        except Exception:
            # fall back to the simple additive HW if optimisation fails
            if n >= _SEASONAL_CYCLE + 2:
                candidates["holt_winters"] = _holt_winters_additive(quantities, cycle=_SEASONAL_CYCLE, horizon=horizon)
    elif n >= _SEASONAL_CYCLE + 2:
        # Not enough non-zero months for optimised HW, use simple additive
        hw_forecasts = _holt_winters_additive(quantities, cycle=_SEASONAL_CYCLE, horizon=horizon)
        candidates["holt_winters"] = hw_forecasts

    # Candidate 4b: Seasonal Naive with trend correction (robust for sparse series)
    if nz_count >= MIN_NZ_SEAS:
        try:
            candidates["seasonal_naive_trend"] = seasonal_naive_forecast_list(quantities, horizon)
        except Exception:
            pass

    # Candidate 4c: Exponential-decay WMA with seasonal index
    if n >= 6:
        try:
            candidates["wma_exp"] = wma_exp_forecast_list(quantities, horizon)
        except Exception:
            pass

    # Candidate 5: Croston's (intermittent)
    if demand_type in ("sporadic", "lumpy"):
        best_alpha_cr, _ = _optimize_alpha(quantities, "croston")
        candidates["croston"] = _croston_forecast(quantities, alpha=best_alpha_cr, horizon=horizon)

    # Candidate 6: SBA (intermittent, bias-corrected)
    if demand_type in ("sporadic", "lumpy"):
        best_alpha_sba, _ = _optimize_alpha(quantities, "sba")
        candidates["sba"] = _sba_forecast(quantities, alpha=best_alpha_sba, horizon=horizon)

    # Candidate 7: Seasonal Naive (same month last year)
    if n >= 13:
        seasonal_naive = []
        for h in range(horizon):
            idx = n - _SEASONAL_CYCLE + h
            if 0 <= idx < n:
                seasonal_naive.append(max(float(quantities[idx]), 0.0))
            else:
                seasonal_naive.append(max(float(np.mean(quantities[-6:])), 0.0))
        candidates["seasonal_naive"] = seasonal_naive

    # Walk-forward backtest each candidate to find best
    eval_start = max(n - 6, _MIN_HISTORY_MONTHS)
    method_scores: list[dict] = []
    for method_name, _ in candidates.items():
        errors = []
        for t in range(eval_start, n):
            actual = float(quantities[t])
            # Generate a 1-step forecast using data up to t
            train = quantities[:t]
            if len(train) < 3:
                continue
            if method_name == "wma" and len(train) >= _MIN_HISTORY_MONTHS:
                lbl = labels[:t]
                bw, _, _, _, _, _ = _backtest_forecast_weight_patterns(train, lbl)
                window = np.asarray(list(train[-len(bw):][::-1]), dtype=float)
                pred = max(_weighted_moving_average(window, bw), 0.0)
            elif method_name == "ses":
                sm_t = _exponential_smoothing(train, alpha=best_alpha_ses)
                pred = max(float(sm_t[-1]), 0.0)
            elif method_name == "holt" and len(train) >= 6:
                lv, tr = _double_exponential_smoothing(train)
                pred = max(float(lv[-1] + tr[-1]), 0.0)
            elif method_name == "holt_winters":
                if (train > 0).sum() >= MIN_NZ_HW and len(train) >= _SEASONAL_CYCLE + 2:
                    try:
                        pred = hw_one_step(train)
                    except Exception:
                        if len(train) >= _SEASONAL_CYCLE + 2:
                            pred = _holt_winters_additive(train, cycle=_SEASONAL_CYCLE, horizon=1)[0]
                        else:
                            continue
                elif len(train) >= _SEASONAL_CYCLE + 2:
                    pred = _holt_winters_additive(train, cycle=_SEASONAL_CYCLE, horizon=1)[0]
                else:
                    continue
            elif method_name == "seasonal_naive_trend":
                if (train > 0).sum() >= MIN_NZ_SEAS:
                    try:
                        pred = seasonal_naive_one_step(train)
                    except Exception:
                        continue
                else:
                    continue
            elif method_name == "wma_exp" and len(train) >= 6:
                try:
                    pred = wma_exp_one_step(train)
                except Exception:
                    continue
            elif method_name == "croston":
                pred = _croston_forecast(train, alpha=best_alpha_cr if 'best_alpha_cr' in dir() else 0.15, horizon=1)[0]
            elif method_name == "sba":
                pred = _sba_forecast(train, alpha=best_alpha_sba if 'best_alpha_sba' in dir() else 0.15, horizon=1)[0]
            elif method_name == "seasonal_naive" and len(train) >= 13:
                pred = max(float(train[t - _SEASONAL_CYCLE]) if t >= _SEASONAL_CYCLE else float(np.mean(train[-6:])), 0.0)
            else:
                continue
            abs_err = abs(actual - pred)
            errors.append({"actual": actual, "pred": pred, "abs_error": abs_err})

        if errors:
            total_actual = sum(e["actual"] for e in errors)
            total_error = sum(e["abs_error"] for e in errors)
            gap_pct = (total_error / total_actual * 100) if total_actual > 0 else 100.0
            # For zero-actual periods, penalise overforecasting mildly
            score = max(0.0, 100.0 - min(gap_pct, 100.0))
        else:
            score = 0.0
            gap_pct = 100.0
        method_scores.append({
            "method": method_name,
            "confidence": round(score, 2),
            "gap_pct": round(gap_pct, 2),
            "eval_points": len(errors),
        })

    method_scores.sort(key=lambda x: x["confidence"], reverse=True)
    meta["candidate_methods"] = method_scores

    # Ensemble: weighted blend of top-2 methods (70/30)
    if len(method_scores) >= 2:
        top1_name = method_scores[0]["method"]
        top2_name = method_scores[1]["method"]
        top1_conf = method_scores[0]["confidence"]
        top2_conf = method_scores[1]["confidence"]
        total_conf = top1_conf + top2_conf
        if total_conf > 0:
            w1 = top1_conf / total_conf
            w2 = top2_conf / total_conf
        else:
            w1, w2 = 0.5, 0.5
        f1 = candidates[top1_name]
        f2 = candidates[top2_name]
        blended = [max(w1 * f1[i] + w2 * f2[i], 0.0) for i in range(horizon)]
        meta["ensemble_weights"] = {top1_name: round(w1, 3), top2_name: round(w2, 3)}
        meta["selected_method"] = f"ensemble({top1_name}+{top2_name})"
    elif method_scores:
        blended = candidates[method_scores[0]["method"]]
        meta["selected_method"] = method_scores[0]["method"]
    else:
        blended = [max(float(np.mean(quantities[-6:])), 0.0)] * horizon
        meta["selected_method"] = "simple_average"

    return blended, meta


def _compute_adaptive_confidence(
    quantities: np.ndarray,
    labels: list[str],
    forecast_func,
    months: int = 6,
) -> tuple[list[dict], str, float | None, float | None]:
    """Procurement-grade confidence via cumulative-horizon walk-forward.

    For inventory planning, what matters most is whether the *cumulative*
    forecast over the planning horizon tracks the actual cumulative demand.
    Individual months will always fluctuate – but procurement decisions
    are based on aggregate quantities over weeks/months.

    Scoring pillars
    ===============
    1. **Cumulative horizon accuracy** (40%) – slide a window of size
       `horizon` across evaluation months.  At each step compare
       *predicted total* vs *actual total*.  This is the single most
       important metric for procurement planning.

    2. **Demand-rate accuracy** (25%) – compare the *average predicted
       monthly rate* vs *average actual rate*.  Tolerates per-month
       noise while rewarding correct magnitude.

    3. **Directional accuracy** (15%) – did the model correctly predict
       the sign of change vs. the prior month?

    4. **Data quality & coverage** (20%) – more evaluation points,
       fewer missing months, and stable error variance.
    """
    n = len(quantities)
    demand_type = _classify_demand(quantities)

    # Evaluation window: use more history for intermittent items
    eval_months = months
    if demand_type in ("sporadic", "lumpy"):
        eval_months = min(n - 4, 18)  # up to 18 months
    elif demand_type == "variable":
        eval_months = min(n - 4, 12)
    eval_start = max(n - eval_months, 4)

    results: list[dict] = []
    pred_history: list[float] = []
    actual_history: list[float] = []
    direction_correct = direction_total = 0

    for t in range(eval_start, n):
        train = quantities[:t]
        train_labels = labels[:t]
        actual = float(quantities[t])

        if len(train) < 4:
            continue

        forecasted, _ = forecast_func(train, train_labels, horizon=1)
        pred = forecasted[0] if forecasted else 0.0
        abs_err = abs(actual - pred)
        results.append({
            "period": labels[t],
            "forecast_qty": round(pred, 2),
            "actual_qty": round(actual, 2),
            "abs_error": round(abs_err, 2),
        })
        pred_history.append(pred)
        actual_history.append(actual)

        if t >= eval_start + 1 and t - 1 >= 0:
            direction_total += 1
            ad = actual - float(quantities[t - 1])
            pd_ = pred - float(quantities[t - 1])
            if (ad >= 0 and pd_ >= 0) or (ad < 0 and pd_ < 0):
                direction_correct += 1

    if not results:
        return [], "", None, None

    # ================================================================
    # Pillar 1: Cumulative horizon accuracy (40%)
    # Slide windows of various sizes (3, 6, full) and measure how
    # close predicted totals track actual totals.
    # ================================================================
    horizon_scores: list[float] = []
    for win_size in [3, 6, len(actual_history)]:
        ws = min(win_size, len(actual_history))
        if ws < 2:
            continue
        for i in range(ws, len(actual_history) + 1):
            actual_sum = sum(actual_history[i - ws : i])
            pred_sum = sum(pred_history[i - ws : i])
            max_val = max(abs(actual_sum), abs(pred_sum), 1e-9)
            # Relative accuracy: 1 - |error|/max (bounded 0..1)
            rel_acc = 1.0 - (abs(actual_sum - pred_sum) / max_val)
            horizon_scores.append(max(rel_acc, 0.0))

    cumulative_score = float(np.mean(horizon_scores)) * 100.0 if horizon_scores else 50.0

    # ================================================================
    # Pillar 2: Demand-rate accuracy (25%)
    # How close is predicted avg monthly rate to actual avg rate?
    # ================================================================
    actual_rate = float(np.mean(actual_history)) if actual_history else 0.0
    pred_rate = float(np.mean(pred_history)) if pred_history else 0.0
    max_rate = max(abs(actual_rate), abs(pred_rate), 1e-9)
    rate_accuracy = 1.0 - (abs(actual_rate - pred_rate) / max_rate)
    rate_score = max(rate_accuracy, 0.0) * 100.0

    # ================================================================
    # Pillar 3: Directional accuracy (15%)
    # ================================================================
    if direction_total > 0:
        direction_score = (direction_correct / direction_total) * 100.0
    else:
        direction_score = 60.0

    # ================================================================
    # Pillar 4: Data quality & coverage (20%)
    # ================================================================
    coverage = min(len(results) / max(months, 1), 1.0)
    # Bonus for having lots of evaluation points
    depth_bonus = min(len(results) / 12.0, 1.0)  # up to 12 months
    # Penalize if model errors are wildly inconsistent
    pct_errors = []
    for a, p in zip(actual_history, pred_history):
        d = max(abs(a), abs(p), 1e-9)
        pct_errors.append(abs(a - p) / d)
    if len(pct_errors) > 1:
        err_consistency = 1.0 - min(float(np.std(pct_errors)), 1.0)
    else:
        err_consistency = 0.5
    quality_score = (0.4 * coverage + 0.3 * depth_bonus + 0.3 * err_consistency) * 100.0

    # ================================================================
    # Composite confidence
    # ================================================================
    confidence_pct = (
        0.40 * cumulative_score
        + 0.25 * rate_score
        + 0.15 * direction_score
        + 0.20 * quality_score
    )
    confidence_pct = max(0.0, min(100.0, confidence_pct))

    window_label = (
        f"{results[0]['period']} to {results[-1]['period']}"
        if results
        else ""
    )
    total_actual = sum(_safe_float(r["actual_qty"]) for r in results)
    total_error = sum(_safe_float(r["abs_error"]) for r in results)
    gap_pct = (total_error / total_actual * 100.0) if total_actual > 0 else None

    return (
        results,
        window_label,
        round(gap_pct, 2) if gap_pct is not None else None,
        round(confidence_pct, 2),
    )

SEED_UPLOAD_SCHEMA = {
    "consumption": {
        "title": "Consumption MB51 File",
        "filename": os.path.basename(CONS_FILE),
        "required": [
            ("Plant", "SAP plant code"),
            ("Material", "SAP material code"),
            ("Material Description", "Material description from MB51"),
            ("Movement Type", "Consumption movement type such as 201/221/261 or reversal 202/222/262"),
            ("Posting Date", "Transaction posting date used to derive financial year"),
            ("Qty in unit of entry", "Movement quantity"),
            ("Unit of Entry", "Entry UoM"),
            ("Amt.in Loc.Cur.", "Movement value in local currency"),
        ],
        "optional": [
            ("Storage Location", "Storage location within the plant"),
            ("Material Document", "Material document number"),
            ("Material Doc.Item", "Material document item"),
            ("Receiving plant", "Receiving plant for transfer-style movements"),
            ("Purchase order", "Linked purchase order if present"),
        ],
        "allowed_exts": {".xlsx", ".xls", ".csv"},
    },
    "procurement": {
        "title": "Procurement ME2M File",
        "filename": os.path.basename(PROC_FILE),
        "required": [
            ("Plant", "SAP plant code"),
            ("Net Price", "SAP source field retained in the upload; analytics use calculated Unit Price from Effective value / Order Quantity"),
            ("Purchasing Document", "PO number"),
            ("Document Date", "PO document date"),
            ("Material", "Material code"),
            ("Supplier/Supplying Plant", "Vendor name or supplying plant"),
            ("Short Text", "Material description"),
            ("Order Quantity", "Ordered quantity"),
            ("Still to be delivered (qty)", "Undelivered balance used to derive procured quantity"),
            ("Order Unit", "Order unit of measure"),
            ("Currency", "Currency code, typically INR"),
            ("Price Unit", "Price basis unit"),
            ("Effective value", "Procurement value"),
        ],
        "optional": [
            ("Item", "PO item number"),
            ("Release indicator", "Release status"),
        ],
        "allowed_exts": {".xlsx", ".xls", ".csv"},
    },
    "mc9": {
        "title": "Monthly Stock And Consumption MC.9 File",
        "filename": os.path.basename(MC9_FILE),
        "required": [
            ("Plant", "Plant label with the SAP plant code prefix"),
            ("Material", "Material code and description"),
            ("Storage Location", "Storage location"),
            ("Month", "Monthly reporting period such as 4.2023"),
            ("ValStckVal", "Closing stock value"),
            ("ValStckVal.1", "Closing stock currency"),
            ("Val. stock", "Closing stock quantity"),
            ("Val. stock.1", "Closing stock unit of measure"),
            ("Tot. usage", "Monthly consumption quantity"),
            ("Tot. usage.1", "Monthly consumption unit of measure"),
            ("Tot.us.val", "Monthly consumption value"),
            ("Tot.us.val.1", "Monthly consumption currency"),
        ],
        "optional": [
            ("MRP Controller", "MRP controller"),
            ("MRP Type", "MRP type"),
            ("Material Type", "Material type"),
            ("Material Group", "Material group"),
            ("Business Area", "Business area"),
            ("Division", "Division"),
        ],
        "allowed_exts": {".xlsx", ".xls", ".csv"},
    },
}


def _safe_float(value, default: float = 0.0) -> float:
    try:
        if pd.isna(value):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _normalize_plant(plant: str | None) -> str | None:
    if plant is None:
        return None
    value = str(plant).strip().upper()
    return value or None


def _normalize_grouping_config(raw_config: dict | None) -> dict:
    config = raw_config if isinstance(raw_config, dict) else {}
    prefix_groups = {
        _normalize_code(prefix): _normalize_code(parent)
        for prefix, parent in (config.get("prefix_groups") or {}).items()
        if _normalize_code(prefix) and _normalize_code(parent)
    }
    explicit_groups = {
        _normalize_code(child): _normalize_code(parent)
        for child, parent in (config.get("explicit_groups") or {}).items()
        if _normalize_code(child) and _normalize_code(parent)
    }
    group_members: dict[str, list[str]] = {}
    for parent, members in (config.get("group_members") or {}).items():
        normalized_parent = _normalize_code(parent)
        normalized_members = [
            _normalize_code(member)
            for member in (members or [])
            if _normalize_code(member)
        ]
        if normalized_parent:
            group_members[normalized_parent] = normalized_members
    return {
        "prefix_groups": prefix_groups,
        "explicit_groups": explicit_groups,
        "group_members": group_members,
    }


def _load_plant_grouping_config() -> dict:
    try:
        mtime = os.path.getmtime(PLANT_GROUPING_FILE)
    except OSError:
        mtime = None

    cache_hit = (
        _PLANT_GROUPING_CACHE["config"] is not None
        and _PLANT_GROUPING_CACHE["mtime"] == mtime
    )
    if cache_hit:
        return _PLANT_GROUPING_CACHE["config"]

    config = _normalize_grouping_config(_DEFAULT_PLANT_GROUPING_CONFIG)
    if mtime is not None:
        try:
            with open(PLANT_GROUPING_FILE, "r", encoding="utf-8") as handle:
                loaded = json.load(handle)
            config = _normalize_grouping_config(loaded)
        except Exception:
            logger.exception("Failed to load plant grouping config from %s", PLANT_GROUPING_FILE)

    _PLANT_GROUPING_CACHE["mtime"] = mtime
    _PLANT_GROUPING_CACHE["config"] = config
    return config


def _reporting_plant(plant: str | None) -> str | None:
    normalized = _normalize_plant(plant)
    if not normalized:
        return None
    config = _load_plant_grouping_config()

    member_lookup = {}
    for parent, members in config.get("group_members", {}).items():
        member_lookup[parent] = parent
        for member in members:
            member_lookup[member] = parent
    if normalized in member_lookup:
        return member_lookup[normalized]

    explicit_groups = config.get("explicit_groups", {})
    if normalized in explicit_groups:
        return explicit_groups[normalized]

    prefix_groups = sorted(
        config.get("prefix_groups", {}).items(),
        key=lambda item: len(item[0]),
        reverse=True,
    )
    for prefix, parent in prefix_groups:
        if normalized.startswith(prefix):
            return parent

    return normalized


def _normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _normalize_quantity_series(quantity: pd.Series, uom: pd.Series) -> tuple[pd.Series, pd.Series]:
    normalized_uom = uom.map(_normalize_code)
    conversion = normalized_uom.map(_QUANTITY_UNIT_FACTORS)
    factors = conversion.map(lambda item: item[1] if item else np.nan).astype(float)
    canonical_uom = conversion.map(lambda item: item[0] if item else None)
    normalized_qty = quantity.where(factors.isna(), quantity * factors)
    normalized_uom = normalized_uom.where(canonical_uom.isna(), canonical_uom)
    return normalized_qty, normalized_uom


def _normalize_code(value) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    return text.strip().upper()


def _extract_prefixed_code(value) -> str:
    text = _normalize_text(value).upper()
    if not text:
        return ""
    match = re.match(r"^([A-Z0-9]{3,12})\b", text)
    return _normalize_code(match.group(1)) if match else ""


def _split_material_label(value) -> tuple[str, str]:
    text = _normalize_text(value).upper()
    if not text:
        return "", ""
    match = re.match(r"^([0-9]{6,18})\s+(.+)$", text)
    if not match:
        return "", text
    return _normalize_code(match.group(1)), match.group(2).strip()


def _header_key(value) -> str:
    text = _normalize_text(value).lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _apply_header_aliases(df: pd.DataFrame, aliases: dict[str, str]) -> pd.DataFrame:
    frame = df.copy()
    frame.columns = [str(column).strip() for column in frame.columns]
    col_map = {}
    for column in frame.columns:
        alias = aliases.get(_header_key(column))
        if alias:
            col_map[column] = alias
    return frame.rename(columns=col_map)


def _detect_frame_variant(columns: list[str], seed_type: str) -> str:
    keys = {_header_key(column) for column in columns}
    if seed_type == "consumption":
        if "movement type" in keys and "posting date" in keys:
            return "mb51"
        return "legacy"
    if seed_type == "mc9":
        if {"plant", "material", "month", "tot usage", "tot us val"}.issubset(keys):
            return "mc9"
        return "legacy"
    if "purchasing document" in keys and "document date" in keys:
        return "me2m"
    return "legacy"


def _format_period_label(year: int, month: int) -> str:
    return f"{calendar.month_abbr[int(month)]} {int(year)}"


def _looks_like_period(value) -> bool:
    return bool(_PERIOD_RE.match(_normalize_text(value)))


def _parse_period(value) -> tuple[pd.Series, pd.Series]:
    parts = value.fillna("").astype(str).str.extract(_PERIOD_RE)
    month = pd.to_numeric(parts[0], errors="coerce").astype("Int64")
    year = pd.to_numeric(parts[1], errors="coerce").astype("Int64")
    return month, year


def _clean_currency(value, fallback: str = "INR") -> str:
    text = _normalize_text(value).upper()
    if re.fullmatch(r"[A-Z]{3,5}", text):
        return text
    return fallback


def _financial_year_label(year: int, month: int) -> str:
    start_year = int(year) if int(month) >= 4 else int(year) - 1
    end_year = start_year + 1
    return f"FY{str(start_year)[-2:]}-{str(end_year)[-2:]}"


def _apply_financial_year_columns(df: pd.DataFrame, date_column: str) -> pd.DataFrame:
    frame = df.copy()
    frame["financial_year"] = ""
    frame["financial_year_start"] = pd.Series(dtype="Int64")
    frame["financial_year_end"] = pd.Series(dtype="Int64")
    frame["fiscal_month_index"] = pd.Series(dtype="Int64")

    valid = frame[date_column].notna()
    if not valid.any():
        return frame

    year = frame.loc[valid, date_column].dt.year
    month = frame.loc[valid, date_column].dt.month
    fy_start = year.where(month >= 4, year - 1).astype("Int64")
    frame.loc[valid, "financial_year_start"] = fy_start
    frame.loc[valid, "financial_year_end"] = fy_start + 1
    frame.loc[valid, "financial_year"] = [
        f"FY{str(int(start))[-2:]}-{str(int(start + 1))[-2:]}"
        for start in fy_start
    ]
    frame.loc[valid, "fiscal_month_index"] = ((month - 4) % 12 + 1).astype("Int64")
    return frame


def _build_seed_row_financial_years(df: pd.DataFrame, seed_type: str) -> pd.Series:
    frame = df.copy()
    if frame.empty:
        return pd.Series(dtype="object")

    variant = _detect_frame_variant([str(column) for column in frame.columns], seed_type)
    if seed_type == "consumption":
        if variant == "mb51":
            frame = _apply_header_aliases(frame, _MB51_CONSUMPTION_ALIASES)
            posting_dates = pd.to_datetime(frame.get("posting_date"), errors="coerce")
            return posting_dates.map(
                lambda value: _financial_year_label(value.year, value.month) if pd.notna(value) else ""
            )

        frame = _apply_header_aliases(frame, _LEGACY_CONSUMPTION_ALIASES)
        month_raw = frame.get("month_raw", pd.Series(index=frame.index, dtype="object")).fillna("").astype(str)
        month, year = _parse_period(month_raw)
        result = pd.Series("", index=frame.index, dtype="object")
        valid = month.notna() & year.notna()
        result.loc[valid] = [
            _financial_year_label(int(y), int(m))
            for y, m in zip(year.loc[valid], month.loc[valid])
        ]
        return result

    if seed_type == "mc9":
        frame = _apply_header_aliases(frame, _MC9_ALIASES)
        month_raw = frame.get("month_raw", pd.Series(index=frame.index, dtype="object")).fillna("").astype(str)
        month, year = _parse_period(month_raw)
        result = pd.Series("", index=frame.index, dtype="object")
        valid = month.notna() & year.notna()
        result.loc[valid] = [
            _financial_year_label(int(y), int(m))
            for y, m in zip(year.loc[valid], month.loc[valid])
        ]
        return result

    aliases = _ME2M_PROCUREMENT_ALIASES if variant == "me2m" else _LEGACY_PROCUREMENT_ALIASES
    frame = _apply_header_aliases(frame, aliases)
    doc_dates = pd.to_datetime(frame.get("doc_date"), errors="coerce")
    return doc_dates.map(
        lambda value: _financial_year_label(value.year, value.month) if pd.notna(value) else ""
    )


def _build_period_index(df: pd.DataFrame) -> pd.PeriodIndex:
    return pd.PeriodIndex.from_fields(
        year=df["year"].astype(int),
        month=df["month"].astype(int),
        freq="M",
    )


def _repair_shifted_consumption_rows(df: pd.DataFrame) -> pd.DataFrame:
    repaired = df.copy()
    repaired[["month_raw", "usage_qty", "uom", "usage_value", "currency"]] = (
        repaired[["month_raw", "usage_qty", "uom", "usage_value", "currency"]].astype("object")
    )
    shifted_mask = (
        repaired["month_raw"].fillna("").astype(str).str.strip().eq("")
        & repaired["usage_qty"].apply(_looks_like_period)
    )
    if not shifted_mask.any():
        return repaired

    logger.info("Repairing %d shifted consumption rows", int(shifted_mask.sum()))
    shifted_values = repaired.loc[shifted_mask, ["usage_qty", "uom", "usage_value", "currency"]].copy()
    repaired.loc[shifted_mask, "month_raw"] = shifted_values["usage_qty"].to_numpy()
    repaired.loc[shifted_mask, "usage_qty"] = shifted_values["uom"].to_numpy()
    repaired.loc[shifted_mask, "uom"] = shifted_values["usage_value"].to_numpy()
    repaired.loc[shifted_mask, "usage_value"] = shifted_values["currency"].to_numpy()
    repaired.loc[shifted_mask, "currency"] = "INR"
    return repaired


def _normalize_legacy_consumption_frame(df: pd.DataFrame) -> pd.DataFrame:
    frame = _apply_header_aliases(df, _LEGACY_CONSUMPTION_ALIASES)
    frame["source_excel_row"] = frame.index + 2
    for column in ["plant", "material", "storage_location", "month_raw", "usage_qty", "uom", "usage_value", "currency"]:
        if column not in frame.columns:
            frame[column] = pd.NA

    frame = _repair_shifted_consumption_rows(frame)
    frame["plant"] = frame["plant"].map(_normalize_code)
    frame["material"] = frame["material"].map(_normalize_text).str.upper()
    frame["storage_location"] = frame["storage_location"].map(_normalize_text)
    frame["month_raw"] = frame["month_raw"].map(_normalize_text)
    frame["uom"] = frame["uom"].map(_normalize_code)
    frame["currency"] = frame["currency"].map(_clean_currency)
    frame["usage_qty"] = pd.to_numeric(frame["usage_qty"], errors="coerce")
    frame["usage_value"] = pd.to_numeric(frame["usage_value"], errors="coerce")
    frame["usage_qty"], frame["uom"] = _normalize_quantity_series(frame["usage_qty"], frame["uom"])

    month, year = _parse_period(frame["month_raw"])
    frame["month"] = month
    frame["year"] = year
    frame["posting_date"] = pd.to_datetime(
        dict(
            year=pd.to_numeric(frame["year"], errors="coerce"),
            month=pd.to_numeric(frame["month"], errors="coerce"),
            day=1,
        ),
        errors="coerce",
    )

    def _split_material(value: str) -> tuple[str, str]:
        match = re.match(r"^(\d{6,12})\s+(.+)$", value.strip())
        if not match:
            return "", value.strip()
        return match.group(1), match.group(2).strip()

    split_values = frame["material"].apply(_split_material)
    frame["material_code"] = split_values.apply(lambda item: item[0])
    frame["material_desc"] = split_values.apply(lambda item: item[1])
    frame["movement_type"] = ""
    frame["movement_sign"] = np.nan
    frame["material_document"] = ""
    frame["material_document_item"] = ""
    frame["receiving_plant"] = ""
    frame["po_number"] = ""
    frame["reporting_plant"] = frame["plant"].map(_reporting_plant)
    frame = _apply_financial_year_columns(frame, "posting_date")
    return frame[frame["material_desc"] != ""].copy()


def _normalize_mb51_consumption_frame(df: pd.DataFrame) -> pd.DataFrame:
    frame = _apply_header_aliases(df, _MB51_CONSUMPTION_ALIASES)
    frame["source_excel_row"] = frame.index + 2
    for column in [
        "material_code",
        "plant",
        "storage_location",
        "movement_type",
        "material_document",
        "base_uom",
        "material_document_item",
        "material_desc",
        "posting_date",
        "entry_qty",
        "uom",
        "usage_value",
        "receiving_plant",
        "po_number",
    ]:
        if column not in frame.columns:
            frame[column] = pd.NA

    frame["plant"] = frame["plant"].map(_normalize_code)
    frame["storage_location"] = frame["storage_location"].map(_normalize_text)
    frame["movement_type"] = frame["movement_type"].map(_normalize_code)
    frame["material_code"] = frame["material_code"].map(_normalize_code)
    frame["material_desc"] = frame["material_desc"].map(_normalize_text).str.upper()
    frame["material_document"] = frame["material_document"].map(_normalize_code)
    frame["material_document_item"] = frame["material_document_item"].map(_normalize_code)
    frame["uom"] = frame["uom"].map(_normalize_code)
    frame["base_uom"] = frame["base_uom"].map(_normalize_code)
    frame["receiving_plant"] = frame["receiving_plant"].map(_normalize_code)
    frame["po_number"] = frame["po_number"].map(_normalize_code)
    frame["posting_date"] = pd.to_datetime(frame["posting_date"], errors="coerce")
    frame["entry_qty"] = pd.to_numeric(frame["entry_qty"], errors="coerce")
    frame["usage_value"] = pd.to_numeric(frame["usage_value"], errors="coerce")
    frame["movement_sign"] = frame["movement_type"].map(_CONSUMPTION_MOVEMENT_SIGNS)
    frame["usage_qty_raw"] = frame["entry_qty"]
    frame["usage_value_raw"] = frame["usage_value"]
    frame["usage_qty"] = frame["entry_qty"].abs() * frame["movement_sign"]
    frame["usage_value"] = frame["usage_value"].abs() * frame["movement_sign"]
    frame["usage_qty"], frame["uom"] = _normalize_quantity_series(frame["usage_qty"], frame["uom"])
    _, frame["base_uom"] = _normalize_quantity_series(pd.Series(np.nan, index=frame.index), frame["base_uom"])
    frame["month"] = frame["posting_date"].dt.month.astype("Int64")
    frame["year"] = frame["posting_date"].dt.year.astype("Int64")
    frame["month_raw"] = frame["posting_date"].dt.strftime("%m.%Y").fillna("")
    frame["currency"] = "INR"
    frame["reporting_plant"] = frame["plant"].map(_reporting_plant)
    frame = _apply_financial_year_columns(frame, "posting_date")
    frame = frame[frame["material_desc"] != ""].copy()
    frame = frame[frame["movement_sign"].notna()].copy()
    return frame


def _normalize_consumption_frame(df: pd.DataFrame) -> pd.DataFrame:
    variant = _detect_frame_variant([str(column) for column in df.columns], "consumption")
    if variant == "mb51":
        return _normalize_mb51_consumption_frame(df)
    return _normalize_legacy_consumption_frame(df)


def _load_consumption() -> pd.DataFrame:
    return _load_normalized_cache(
        source_path=CONS_FILE,
        cache_name="consumption_normalized.pkl",
        normalizer=_normalize_consumption_frame,
    )


def _normalize_mc9_frame(df: pd.DataFrame) -> pd.DataFrame:
    frame = _apply_header_aliases(df, _MC9_ALIASES)
    frame["source_excel_row"] = frame.index + 2
    for column in [
        "plant_label",
        "material_label",
        "storage_location",
        "mrp_controller",
        "mrp_type",
        "material_type",
        "material_group",
        "business_area",
        "division",
        "month_raw",
        "stock_value",
        "stock_value_currency",
        "stock_qty",
        "stock_uom",
        "usage_qty",
        "uom",
        "usage_value",
        "currency",
    ]:
        if column not in frame.columns:
            frame[column] = pd.NA

    frame["plant_label"] = frame["plant_label"].map(_normalize_text)
    frame["material_label"] = frame["material_label"].map(_normalize_text).str.upper()
    frame["storage_location"] = frame["storage_location"].map(_normalize_text)
    frame["mrp_controller"] = frame["mrp_controller"].map(_normalize_text)
    frame["mrp_type"] = frame["mrp_type"].map(_normalize_text)
    frame["material_type"] = frame["material_type"].map(_normalize_text)
    frame["material_group"] = frame["material_group"].map(_normalize_text)
    frame["business_area"] = frame["business_area"].map(_normalize_text)
    frame["division"] = frame["division"].map(_normalize_text)
    frame["month_raw"] = frame["month_raw"].map(_normalize_text)
    frame["plant"] = frame["plant_label"].map(_extract_prefixed_code)
    split_material = frame["material_label"].apply(_split_material_label)
    frame["material_code"] = split_material.apply(lambda item: item[0])
    frame["material_desc"] = split_material.apply(lambda item: item[1])
    frame["stock_value_currency"] = frame["stock_value_currency"].map(_clean_currency)
    frame["stock_uom"] = frame["stock_uom"].map(_normalize_code)
    frame["uom"] = frame["uom"].map(_normalize_code)
    frame["currency"] = frame["currency"].map(_clean_currency)
    for column in ["stock_value", "stock_qty", "usage_qty", "usage_value"]:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    frame["stock_qty"], frame["stock_uom"] = _normalize_quantity_series(frame["stock_qty"], frame["stock_uom"])
    frame["usage_qty"], frame["uom"] = _normalize_quantity_series(frame["usage_qty"], frame["uom"])

    month, year = _parse_period(frame["month_raw"])
    frame["month"] = month
    frame["year"] = year
    frame["posting_date"] = pd.to_datetime(
        dict(
            year=pd.to_numeric(frame["year"], errors="coerce"),
            month=pd.to_numeric(frame["month"], errors="coerce"),
            day=1,
        ),
        errors="coerce",
    )
    frame["reporting_plant"] = frame["plant"].map(_reporting_plant)
    frame = _apply_financial_year_columns(frame, "posting_date")
    return frame[(frame["material_desc"] != "") | (frame["material_code"] != "")].copy()


def _load_mc9() -> pd.DataFrame:
    return _load_normalized_cache(
        source_path=MC9_FILE,
        cache_name="mc9_normalized.pkl",
        normalizer=_normalize_mc9_frame,
    )


def _normalize_legacy_procurement_frame(df: pd.DataFrame) -> pd.DataFrame:
    frame = _apply_header_aliases(df, _LEGACY_PROCUREMENT_ALIASES)
    frame["source_excel_row"] = frame.index + 2
    for column in [
        "plant",
        "net_price",
        "po_number",
        "doc_date",
        "material_code",
        "item",
        "vendor",
        "material_desc",
        "order_qty",
        "still_to_be_delivered_qty",
        "order_unit",
        "currency",
        "price_unit",
        "effective_value",
        "release_indicator",
    ]:
        if column not in frame.columns:
            frame[column] = pd.NA

    frame["plant"] = frame["plant"].map(_normalize_code)
    frame["po_number"] = frame["po_number"].map(_normalize_code)
    frame["material_code"] = frame["material_code"].map(_normalize_code)
    frame["item"] = frame["item"].map(_normalize_code)
    frame["vendor"] = frame["vendor"].map(_normalize_text)
    frame["material_desc"] = frame["material_desc"].map(_normalize_text).str.upper()
    frame["order_unit"] = frame["order_unit"].map(_normalize_code)
    frame["currency"] = frame["currency"].map(_clean_currency)
    frame["release_indicator"] = frame["release_indicator"].map(_normalize_text)
    frame["doc_date"] = pd.to_datetime(frame["doc_date"], errors="coerce")
    for column in ["net_price", "order_qty", "still_to_be_delivered_qty", "price_unit", "effective_value"]:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")

    positive_qty = frame["order_qty"].where(frame["order_qty"].fillna(0) > 0)
    derived_unit_price = frame["effective_value"] / positive_qty
    frame["unit_price"] = derived_unit_price.replace([np.inf, -np.inf], np.nan)
    frame["net_price"] = frame["unit_price"]
    frame["procured_qty"] = frame["order_qty"] - frame["still_to_be_delivered_qty"].fillna(0)
    frame["reporting_plant"] = frame["plant"].map(_reporting_plant)
    frame = _apply_financial_year_columns(frame, "doc_date")
    return frame[frame["material_desc"] != ""].copy()


def _normalize_procurement_frame(df: pd.DataFrame) -> pd.DataFrame:
    return _normalize_legacy_procurement_frame(df)


def _load_procurement() -> pd.DataFrame:
    return _load_normalized_cache(
        source_path=PROC_FILE,
        cache_name="procurement_normalized.pkl",
        normalizer=_normalize_procurement_frame,
    )


def _source_signature(path: str) -> tuple[int, int] | None:
    if not os.path.exists(path):
        return None
    stat = os.stat(path)
    return (stat.st_mtime_ns, stat.st_size)


def _cache_path(cache_name: str) -> str:
    os.makedirs(_CACHE_DIR, exist_ok=True)
    return os.path.join(_CACHE_DIR, cache_name)


def _load_normalized_cache(
    source_path: str,
    cache_name: str,
    normalizer,
) -> pd.DataFrame:
    signature = _source_signature(source_path)
    cache_path = _cache_path(cache_name)

    if signature and os.path.exists(cache_path):
        try:
            with open(cache_path, "rb") as fh:
                payload = pickle.load(fh)
            if (
                isinstance(payload, dict)
                and payload.get("signature") == signature
                and isinstance(payload.get("frame"), pd.DataFrame)
            ):
                logger.info("Loaded normalized inventory cache from %s", cache_path)
                return payload["frame"]
        except Exception:
            logger.warning("Failed to read inventory cache %s; rebuilding", cache_path, exc_info=True)

    df = pd.read_excel(source_path, sheet_name=0)
    frame = normalizer(df)

    if signature:
        try:
            with open(cache_path, "wb") as fh:
                pickle.dump({"signature": signature, "frame": frame}, fh, protocol=pickle.HIGHEST_PROTOCOL)
        except Exception:
            logger.warning("Failed to write inventory cache %s", cache_path, exc_info=True)

    return frame


def _forecast_cache_payload_signature() -> dict[str, tuple[int, int] | None]:
    return {
        "consumption": _source_signature(CONS_FILE),
        "procurement": _source_signature(PROC_FILE),
        "plant_grouping": _source_signature(PLANT_GROUPING_FILE),
    }


def _forecast_cache_entry_key(material: str, plant: str | None) -> str:
    return f"{_normalize_plant(plant) or '__all__'}::{material.upper()}"


def _load_precomputed_forecast_cache() -> dict[str, dict]:
    signature = _forecast_cache_payload_signature()
    if os.path.exists(_FORECAST_CACHE_FILE):
        try:
            with open(_FORECAST_CACHE_FILE, "rb") as handle:
                payload = pickle.load(handle)
            if (
                isinstance(payload, dict)
                and payload.get("signature") == signature
                and isinstance(payload.get("forecasts"), dict)
            ):
                logger.info("Loaded precomputed forecast cache from %s", _FORECAST_CACHE_FILE)
                return payload["forecasts"]
        except Exception:
            logger.warning("Failed to read forecast cache %s; rebuilding lazily", _FORECAST_CACHE_FILE, exc_info=True)
    return {}


def _write_precomputed_forecast_cache(forecasts: dict[str, dict]) -> None:
    os.makedirs(_CACHE_DIR, exist_ok=True)
    payload = {
        "signature": _forecast_cache_payload_signature(),
        "forecasts": forecasts,
    }
    with tempfile.NamedTemporaryFile(dir=_CACHE_DIR, suffix=".tmp", delete=False) as handle:
        temp_path = handle.name
    try:
        with open(temp_path, "wb") as handle:
            pickle.dump(payload, handle, protocol=pickle.HIGHEST_PROTOCOL)
        os.replace(temp_path, _FORECAST_CACHE_FILE)
    except Exception:
        try:
            os.remove(temp_path)
        except OSError:
            pass
        raise


def precompute_forecast_cache(
    consumption_frame: pd.DataFrame | None = None,
    procurement_frame: pd.DataFrame | None = None,
    top_n: int = 3,
) -> dict[str, int | float]:
    cons = consumption_frame if consumption_frame is not None else _load_consumption()
    if cons.empty:
        _write_precomputed_forecast_cache({})
        return {"materials": 0, "seconds": 0.0}

    started = pd.Timestamp.utcnow()
    proc = procurement_frame if procurement_frame is not None else _load_procurement()
    material_groups = {material: frame for material, frame in cons.groupby("material_desc") if material}
    if not material_groups:
        _write_precomputed_forecast_cache({})
        return {"materials": 0, "seconds": 0.0}

    ranked_materials: list[str]
    if not proc.empty and "material_desc" in proc and "effective_value" in proc:
        ranked_materials = [
            material
            for material in (
                proc.groupby("material_desc")["effective_value"]
                .sum()
                .sort_values(ascending=False)
                .head(top_n)
                .index
                .tolist()
            )
            if material in material_groups
        ]
    else:
        ranked_materials = []

    if not ranked_materials:
        ranked_materials = [
            material
            for material in (
                cons.groupby("material_desc")["usage_qty"]
                .sum()
                .sort_values(ascending=False)
                .head(top_n)
                .index
                .tolist()
            )
            if material in material_groups
        ]

    forecasts: dict[str, dict] = {}
    for material_name in ranked_materials:
        frame = material_groups[material_name]
        monthly = _monthly_series(frame)
        forecasts[_forecast_cache_entry_key(material_name, None)] = _build_forecast(monthly)

    _write_precomputed_forecast_cache(forecasts)
    elapsed = (pd.Timestamp.utcnow() - started).total_seconds()
    logger.info(
        "Precomputed %d all-plant material forecasts in %.2fs",
        len(forecasts),
        elapsed,
    )
    return {"materials": len(forecasts), "seconds": round(elapsed, 2)}


def _read_uploaded_seed_dataframe(file_bytes: bytes, filename: str) -> pd.DataFrame:
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        return pd.read_csv(BytesIO(file_bytes))
    if ext in {".xlsx", ".xls"}:
        return pd.read_excel(BytesIO(file_bytes), sheet_name=0)
    raise ValueError(f"Unsupported file type: {ext or 'unknown'}")


def _detect_consumption_fields(columns: list[str]) -> set[str]:
    aliases = _MB51_CONSUMPTION_ALIASES if _detect_frame_variant(columns, "consumption") == "mb51" else _LEGACY_CONSUMPTION_ALIASES
    return {
        aliases[key]
        for key in (_header_key(column) for column in columns)
        if key in aliases
    }


def _detect_procurement_fields(columns: list[str]) -> set[str]:
    aliases = _ME2M_PROCUREMENT_ALIASES if _detect_frame_variant(columns, "procurement") == "me2m" else _LEGACY_PROCUREMENT_ALIASES
    return {
        aliases[key]
        for key in (_header_key(column) for column in columns)
        if key in aliases
    }


def _detect_mc9_fields(columns: list[str]) -> set[str]:
    return {
        _MC9_ALIASES[key]
        for key in (_header_key(column) for column in columns)
        if key in _MC9_ALIASES
    }


def _validate_seed_frame(df: pd.DataFrame, seed_type: str) -> None:
    columns = [str(column).strip() for column in df.columns]
    if seed_type == "consumption":
        detected = _detect_consumption_fields(columns)
        if _detect_frame_variant(columns, seed_type) == "mb51":
            required_fields = {"plant", "material_code", "material_desc", "movement_type", "posting_date", "entry_qty", "uom", "usage_value"}
            display_names = {
                "plant": "Plant",
                "material_code": "Material",
                "material_desc": "Material Description",
                "movement_type": "Movement Type",
                "posting_date": "Posting Date",
                "entry_qty": "Qty in unit of entry",
                "uom": "Unit of Entry",
                "usage_value": "Amt.in Loc.Cur.",
            }
        else:
            required_fields = {"plant", "material", "month_raw", "usage_qty", "uom", "usage_value", "currency"}
            display_names = {
                "plant": "Plant",
                "material": "Material",
                "month_raw": "Month",
                "usage_qty": "Usage Quantity",
                "uom": "Usage UoM",
                "usage_value": "Usage Value",
                "currency": "Usage Currency",
            }
    elif seed_type == "mc9":
        detected = _detect_mc9_fields(columns)
        required_fields = {
            "plant_label",
            "material_label",
            "storage_location",
            "month_raw",
            "stock_value",
            "stock_value_currency",
            "stock_qty",
            "stock_uom",
            "usage_qty",
            "uom",
            "usage_value",
            "currency",
        }
        display_names = {
            "plant_label": "Plant",
            "material_label": "Material",
            "storage_location": "Storage Location",
            "month_raw": "Month",
            "stock_value": "ValStckVal",
            "stock_value_currency": "ValStckVal.1",
            "stock_qty": "Val. stock",
            "stock_uom": "Val. stock.1",
            "usage_qty": "Tot. usage",
            "uom": "Tot. usage.1",
            "usage_value": "Tot.us.val",
            "currency": "Tot.us.val.1",
        }
    else:
        detected = _detect_procurement_fields(columns)
        required_fields = {
            "plant",
            "po_number",
            "doc_date",
            "material_code",
            "vendor",
            "material_desc",
            "order_qty",
            "still_to_be_delivered_qty",
            "order_unit",
            "currency",
            "price_unit",
            "effective_value",
        }
        display_names = {
            "plant": "Plant",
            "net_price": "Net Price (SAP raw)",
            "po_number": "Purchasing Document",
            "doc_date": "Document Date",
            "material_code": "Material",
            "vendor": "Supplier/Supplying Plant",
            "material_desc": "Short Text",
            "order_qty": "Order Quantity",
            "still_to_be_delivered_qty": "Still to be delivered (qty)",
            "order_unit": "Order Unit",
            "currency": "Currency",
            "price_unit": "Price Unit",
            "effective_value": "Effective value",
        }
    missing = sorted(required_fields - detected)
    if missing:
        missing_labels = [display_names.get(field, field) for field in missing]
        raise ValueError(
            "Missing required columns for "
            f"{seed_type} seed file: {', '.join(missing_labels)}."
        )


def get_seed_upload_schema() -> dict:
    return SEED_UPLOAD_SCHEMA


def _current_financial_year_start(today: date | None = None) -> int:
    today = today or date.today()
    return today.year if today.month >= 4 else today.year - 1


def _format_financial_year_from_start(start_year: int) -> str:
    return f"FY{str(start_year)[-2:]}-{str(start_year + 1)[-2:]}"


def get_seed_upload_status(years: int = 5) -> list[dict]:
    expected_starts = [
        _current_financial_year_start() - offset
        for offset in range(years)
    ]
    expected_labels = [_format_financial_year_from_start(start) for start in expected_starts]

    cons_status: dict[str, dict] = {}
    proc_status: dict[str, dict] = {}
    mc9_status: dict[str, dict] = {}

    if os.path.exists(CONS_FILE):
        cons = _load_consumption()
        if "financial_year" in cons.columns:
            grouped = cons.groupby("financial_year", as_index=False).agg(
                rows=("material_desc", "size"),
                materials=("material_code", "nunique"),
                plants=("reporting_plant", "nunique"),
            )
            cons_status = {
                row["financial_year"]: {
                    "rows": int(row["rows"]),
                    "materials": int(row["materials"]),
                    "plants": int(row["plants"]),
                }
                for _, row in grouped.iterrows()
            }

    if os.path.exists(PROC_FILE):
        proc = _load_procurement()
        if "financial_year" in proc.columns:
            grouped = proc.groupby("financial_year", as_index=False).agg(
                rows=("material_desc", "size"),
                materials=("material_code", "nunique"),
                plants=("reporting_plant", "nunique"),
            )
            proc_status = {
                row["financial_year"]: {
                    "rows": int(row["rows"]),
                    "materials": int(row["materials"]),
                    "plants": int(row["plants"]),
                }
                for _, row in grouped.iterrows()
            }

    if os.path.exists(MC9_FILE):
        mc9 = _load_mc9()
        if "financial_year" in mc9.columns:
            grouped = mc9.groupby("financial_year", as_index=False).agg(
                rows=("material_desc", "size"),
                materials=("material_code", "nunique"),
                plants=("reporting_plant", "nunique"),
            )
            mc9_status = {
                row["financial_year"]: {
                    "rows": int(row["rows"]),
                    "materials": int(row["materials"]),
                    "plants": int(row["plants"]),
                }
                for _, row in grouped.iterrows()
            }

    results: list[dict] = []
    for label in expected_labels:
        consumption = cons_status.get(label)
        procurement = proc_status.get(label)
        mc9 = mc9_status.get(label)
        if consumption and procurement and mc9:
            status_label = "Complete"
            status_tone = "neutral"
            status_detail = "MB51, ME2M, and MC.9 uploaded"
        elif consumption and procurement and not mc9:
            status_label = "Partial"
            status_tone = "warning"
            status_detail = "MB51 and ME2M uploaded, MC.9 missing"
        elif consumption and mc9 and not procurement:
            status_label = "Partial"
            status_tone = "warning"
            status_detail = "MB51 and MC.9 uploaded, ME2M missing"
        elif procurement and mc9 and not consumption:
            status_label = "Partial"
            status_tone = "warning"
            status_detail = "ME2M and MC.9 uploaded, MB51 missing"
        elif consumption and not procurement and not mc9:
            status_label = "Partial"
            status_tone = "warning"
            status_detail = "MB51 uploaded, ME2M and MC.9 missing"
        elif procurement and not consumption and not mc9:
            status_label = "Partial"
            status_tone = "warning"
            status_detail = "ME2M uploaded, MB51 and MC.9 missing"
        elif mc9 and not consumption and not procurement:
            status_label = "Partial"
            status_tone = "warning"
            status_detail = "MC.9 uploaded, MB51 and ME2M missing"
        else:
            status_label = "Missing"
            status_tone = "danger"
            status_detail = "MB51, ME2M, and MC.9 missing"

        results.append(
            {
                "label": label,
                "short_label": label.replace("FY", ""),
                "consumption": consumption,
                "procurement": procurement,
                "mc9": mc9,
                "is_complete": bool(consumption) and bool(procurement) and bool(mc9),
                "status_label": status_label,
                "status_tone": status_tone,
                "status_detail": status_detail,
            }
        )
    return results


def _canonicalize_seed_dataframe(df: pd.DataFrame, seed_type: str) -> pd.DataFrame:
    variant = _detect_frame_variant([str(column) for column in df.columns], seed_type)
    frame = df.copy()

    if seed_type == "consumption":
        if variant == "mb51":
            frame = _apply_header_aliases(frame, _MB51_CONSUMPTION_ALIASES)
            rename_map = {
                "material_code": "Material",
                "plant": "Plant",
                "storage_location": "Storage Location",
                "movement_type": "Movement Type",
                "material_document": "Material Document",
                "base_uom": "Base Unit of Measure",
                "material_document_item": "Material Doc.Item",
                "material_desc": "Material Description",
                "posting_date": "Posting Date",
                "entry_qty": "Qty in unit of entry",
                "uom": "Unit of Entry",
                "usage_value": "Amt.in Loc.Cur.",
                "receiving_plant": "Receiving plant",
                "po_number": "Purchase order",
            }
            frame = frame.rename(columns=rename_map)
            frame = frame.loc[:, ~frame.columns.duplicated()]
            for column in _CANONICAL_CONSUMPTION_COLUMNS:
                if column not in frame.columns:
                    frame[column] = pd.NA
            frame = frame.reindex(columns=_CANONICAL_CONSUMPTION_COLUMNS)
            return frame

        return frame

    if seed_type == "mc9":
        frame = _apply_header_aliases(frame, _MC9_ALIASES)
        rename_map = {
            "plant_label": "Plant",
            "material_label": "Material",
            "storage_location": "Storage Location",
            "mrp_controller": "MRP Controller",
            "mrp_type": "MRP Type",
            "material_type": "Material Type",
            "material_group": "Material Group",
            "business_area": "Business Area",
            "division": "Division",
            "month_raw": "Month",
            "stock_value": "ValStckVal",
            "stock_value_currency": "ValStckVal.1",
            "stock_qty": "Val. stock",
            "stock_uom": "Val. stock.1",
            "usage_qty": "Tot. usage",
            "uom": "Tot. usage.1",
            "usage_value": "Tot.us.val",
            "currency": "Tot.us.val.1",
        }
        frame = frame.rename(columns=rename_map)
        frame = frame.loc[:, ~frame.columns.duplicated()]
        for column in _CANONICAL_MC9_COLUMNS:
            if column not in frame.columns:
                frame[column] = pd.NA
        return frame.reindex(columns=_CANONICAL_MC9_COLUMNS)

    frame = _apply_header_aliases(frame, _ME2M_PROCUREMENT_ALIASES)
    rename_map = {
        "plant": "Plant",
        "net_price": "Net Price",
        "po_number": "Purchasing Document",
        "doc_date": "Document Date",
        "material_code": "Material",
        "item": "Item",
        "vendor": "Supplier/Supplying Plant",
        "material_desc": "Short Text",
        "order_qty": "Order Quantity",
        "still_to_be_delivered_qty": "Still to be delivered (qty)",
        "order_unit": "Order Unit",
        "currency": "Currency",
        "price_unit": "Price Unit",
        "effective_value": "Effective value",
        "release_indicator": "Release indicator",
    }
    frame = frame.rename(columns=rename_map)
    frame = frame.loc[:, ~frame.columns.duplicated()]
    for column in _CANONICAL_PROCUREMENT_COLUMNS:
        if column not in frame.columns:
            frame[column] = pd.NA
    return frame.reindex(columns=_CANONICAL_PROCUREMENT_COLUMNS)


def _merge_seed_collection_frame(existing_df: pd.DataFrame, uploaded_df: pd.DataFrame, seed_type: str) -> pd.DataFrame:
    uploaded_variant = _detect_frame_variant([str(column) for column in uploaded_df.columns], seed_type)
    uploaded_canonical = _canonicalize_seed_dataframe(uploaded_df, seed_type)
    uploaded_fys = {
        fy
        for fy in _build_seed_row_financial_years(uploaded_canonical, seed_type).tolist()
        if fy
    }
    if not uploaded_fys:
        raise ValueError(
            "The uploaded seed file did not contain any usable financial-year rows after parsing."
        )

    if existing_df is None or existing_df.empty:
        return uploaded_canonical.reset_index(drop=True)

    existing_variant = _detect_frame_variant([str(column) for column in existing_df.columns], seed_type)
    if existing_variant != uploaded_variant and uploaded_variant in {"mb51", "me2m"}:
        logger.info(
            "Replacing legacy %s seed history with canonical %s upload.",
            seed_type,
            uploaded_variant,
        )
        return uploaded_canonical.reset_index(drop=True)

    existing_canonical = _canonicalize_seed_dataframe(existing_df, seed_type)
    existing_fys = _build_seed_row_financial_years(existing_canonical, seed_type)
    keep_mask = ~existing_fys.isin(uploaded_fys)
    merged = pd.concat(
        [
            existing_canonical.loc[keep_mask].copy(),
            uploaded_canonical.copy(),
        ],
        ignore_index=True,
    )
    return merged.reset_index(drop=True)


def _write_temp_seed_workbook(df: pd.DataFrame, target: str, sheet_name: str) -> str:
    suffix = os.path.splitext(target)[1] or ".xlsx"
    directory = os.path.dirname(target) or None
    with tempfile.NamedTemporaryFile(suffix=suffix, dir=directory, delete=False) as handle:
        temp_path = handle.name

    try:
        with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
    except Exception:
        try:
            os.remove(temp_path)
        except OSError:
            pass
        raise

    return temp_path


def save_seed_workbook(file_bytes: bytes, filename: str, seed_type: str) -> str:
    from app.core.services.inventory_seed_db import sync_inventory_seed_tables

    df = _read_uploaded_seed_dataframe(file_bytes, filename)
    _validate_seed_frame(df, seed_type)

    target_map = {
        "consumption": CONS_FILE,
        "procurement": PROC_FILE,
        "mc9": MC9_FILE,
    }
    sheet_name_map = {
        "consumption": "Sheet1",
        "procurement": "Data",
        "mc9": "Sheet1",
    }
    target = target_map[seed_type]
    sheet_name = sheet_name_map[seed_type]
    existing_df = pd.read_excel(target, sheet_name=0) if os.path.exists(target) else pd.DataFrame()
    merged_df = _merge_seed_collection_frame(existing_df, df, seed_type)
    temp_path = _write_temp_seed_workbook(merged_df, target, sheet_name)
    os.replace(temp_path, target)
    cons_df = (
        _normalize_consumption_frame(merged_df)
        if seed_type == "consumption"
        else (_load_consumption() if os.path.exists(CONS_FILE) else pd.DataFrame())
    )
    proc_df = (
        _normalize_procurement_frame(merged_df)
        if seed_type == "procurement"
        else (_load_procurement() if os.path.exists(PROC_FILE) else pd.DataFrame())
    )
    sync_inventory_seed_tables(
        consumption_frame=cons_df,
        procurement_frame=proc_df,
        consumption_filename=os.path.basename(CONS_FILE),
        procurement_filename=os.path.basename(PROC_FILE),
    )
    try:
        precompute_forecast_cache(consumption_frame=cons_df, procurement_frame=proc_df)
    except Exception:
        logger.warning("Failed to precompute forecast cache after %s seed update", seed_type, exc_info=True)
    return target


def save_seed_workbooks(
    consumption_bytes: bytes,
    consumption_filename: str,
    procurement_bytes: bytes,
    procurement_filename: str,
    mc9_bytes: bytes,
    mc9_filename: str,
) -> dict:
    from app.core.services.inventory_seed_db import sync_inventory_seed_tables

    cons_df = _read_uploaded_seed_dataframe(consumption_bytes, consumption_filename)
    proc_df = _read_uploaded_seed_dataframe(procurement_bytes, procurement_filename)
    mc9_df = _read_uploaded_seed_dataframe(mc9_bytes, mc9_filename)
    _validate_seed_frame(cons_df, "consumption")
    _validate_seed_frame(proc_df, "procurement")
    _validate_seed_frame(mc9_df, "mc9")

    existing_cons = pd.read_excel(CONS_FILE, sheet_name=0) if os.path.exists(CONS_FILE) else pd.DataFrame()
    existing_proc = pd.read_excel(PROC_FILE, sheet_name=0) if os.path.exists(PROC_FILE) else pd.DataFrame()
    existing_mc9 = pd.read_excel(MC9_FILE, sheet_name=0) if os.path.exists(MC9_FILE) else pd.DataFrame()
    merged_cons = _merge_seed_collection_frame(existing_cons, cons_df, "consumption")
    merged_proc = _merge_seed_collection_frame(existing_proc, proc_df, "procurement")
    merged_mc9 = _merge_seed_collection_frame(existing_mc9, mc9_df, "mc9")
    normalized_cons = _normalize_consumption_frame(merged_cons)
    normalized_proc = _normalize_procurement_frame(merged_proc)

    targets = {
        "consumption": CONS_FILE,
        "procurement": PROC_FILE,
        "mc9": MC9_FILE,
    }
    temp_files = {
        "consumption": _write_temp_seed_workbook(merged_cons, CONS_FILE, "Sheet1"),
        "procurement": _write_temp_seed_workbook(merged_proc, PROC_FILE, "Data"),
        "mc9": _write_temp_seed_workbook(merged_mc9, MC9_FILE, "Sheet1"),
    }

    os.replace(temp_files["consumption"], targets["consumption"])
    os.replace(temp_files["procurement"], targets["procurement"])
    os.replace(temp_files["mc9"], targets["mc9"])
    sync_inventory_seed_tables(
        consumption_frame=normalized_cons,
        procurement_frame=normalized_proc,
        consumption_filename=consumption_filename,
        procurement_filename=procurement_filename,
    )
    try:
        precompute_forecast_cache(
            consumption_frame=normalized_cons,
            procurement_frame=normalized_proc,
        )
    except Exception:
        logger.warning("Failed to precompute forecast cache after staged seed apply", exc_info=True)
    return targets


def _monthly_series(cons_df: pd.DataFrame) -> pd.DataFrame:
    monthly = (
        cons_df.dropna(subset=["year", "month"])
        .groupby(["year", "month"], as_index=False)
        .agg(qty=("usage_qty", "sum"), value=("usage_value", "sum"))
        .sort_values(["year", "month"])
    )
    if monthly.empty:
        return monthly.assign(period=pd.Series(dtype="object"), label=pd.Series(dtype="object"))

    monthly["period"] = _build_period_index(monthly)
    full_periods = pd.period_range(monthly["period"].min(), monthly["period"].max(), freq="M")
    monthly = monthly.set_index("period").reindex(full_periods)
    monthly.index.name = "period"
    monthly["qty"] = monthly["qty"].fillna(0.0)
    monthly["value"] = monthly["value"].fillna(0.0)
    monthly["year"] = monthly.index.year
    monthly["month"] = monthly.index.month
    monthly["label"] = [_format_period_label(period.year, period.month) for period in monthly.index]
    return monthly.reset_index()


def _monthly_procurement_series(proc_df: pd.DataFrame) -> pd.DataFrame:
    monthly = (
        proc_df.dropna(subset=["doc_date"])
        .assign(year=lambda df: df["doc_date"].dt.year, month=lambda df: df["doc_date"].dt.month)
        .groupby(["year", "month"], as_index=False)
        .agg(value=("effective_value", "sum"))
        .sort_values(["year", "month"])
    )
    if monthly.empty:
        return monthly.assign(period=pd.Series(dtype="object"), label=pd.Series(dtype="object"))

    monthly["period"] = _build_period_index(monthly)
    full_periods = pd.period_range(monthly["period"].min(), monthly["period"].max(), freq="M")
    monthly = monthly.set_index("period").reindex(full_periods)
    monthly.index.name = "period"
    monthly["value"] = monthly["value"].fillna(0.0)
    monthly["year"] = monthly.index.year
    monthly["month"] = monthly.index.month
    monthly["label"] = [_format_period_label(period.year, period.month) for period in monthly.index]
    return monthly.reset_index()


def _merge_monthly_views(cons_monthly: pd.DataFrame, proc_monthly: pd.DataFrame) -> pd.DataFrame:
    if cons_monthly.empty and proc_monthly.empty:
        return pd.DataFrame(columns=["period", "year", "month", "label", "qty", "proc_value"])

    periods = []
    if not cons_monthly.empty:
        periods.extend(cons_monthly["period"].tolist())
    if not proc_monthly.empty:
        periods.extend(proc_monthly["period"].tolist())
    full_periods = pd.period_range(min(periods), max(periods), freq="M")

    merged = pd.DataFrame({"period": full_periods})
    if not cons_monthly.empty:
        merged = merged.merge(cons_monthly[["period", "qty"]], on="period", how="left")
    else:
        merged["qty"] = 0.0
    if not proc_monthly.empty:
        merged = merged.merge(proc_monthly[["period", "value"]].rename(columns={"value": "proc_value"}), on="period", how="left")
    else:
        merged["proc_value"] = 0.0

    merged["qty"] = merged["qty"].fillna(0.0)
    merged["proc_value"] = merged["proc_value"].fillna(0.0)
    merged["year"] = merged["period"].dt.year
    merged["month"] = merged["period"].dt.month
    merged["label"] = [_format_period_label(period.year, period.month) for period in merged["period"]]
    return merged


def _actual_monthly_rows(cons_df: pd.DataFrame) -> pd.DataFrame:
    return (
        cons_df.dropna(subset=["year", "month"])
        .groupby(["year", "month"], as_index=False)
        .agg(qty=("usage_qty", "sum"), value=("usage_value", "sum"), uom=("uom", "first"))
        .sort_values(["year", "month"], ascending=[False, False])
    )


def _format_weight_pattern(weights: tuple[float, ...]) -> str:
    return ", ".join(f"{round(weight * 100, 2):g}%" for weight in weights)


def _weighted_moving_average(window: np.ndarray, weights: tuple[float, ...]) -> float:
    return float(np.dot(window, np.asarray(weights, dtype=float)))


def _build_backtest_results(
    quantities: np.ndarray,
    labels: list[str],
    weights: tuple[float, ...],
) -> list[dict]:
    lookback = len(weights)
    results: list[dict] = []
    for idx in range(lookback, len(quantities)):
        window = quantities[idx - lookback:idx][::-1]
        forecast_qty = _weighted_moving_average(window, weights)
        actual_qty = float(quantities[idx])
        abs_error = abs(actual_qty - forecast_qty)
        results.append(
            {
                "period": labels[idx],
                "forecast_qty": round(_safe_float(forecast_qty), 2),
                "actual_qty": round(_safe_float(actual_qty), 2),
                "abs_error": round(_safe_float(abs_error), 2),
            }
        )
    return results


def _summarize_recent_backtest(results: list[dict], months: int = 6) -> tuple[list[dict], str, float | None, float | None]:
    recent_results = results[-months:] if results else []
    if not recent_results:
        return [], "", None, None

    actual_total = sum(_safe_float(row.get("actual_qty")) for row in recent_results)
    error_total = sum(_safe_float(row.get("abs_error")) for row in recent_results)
    gap_pct = ((error_total / actual_total) * 100) if actual_total > 0 else None
    confidence_pct = max(0.0, 100.0 - min(_safe_float(gap_pct), 100.0)) if gap_pct is not None else None
    window_label = f"{recent_results[0]['period']} to {recent_results[-1]['period']}" if recent_results else ""
    return recent_results, window_label, round(_safe_float(gap_pct), 2) if gap_pct is not None else None, round(_safe_float(confidence_pct), 2) if confidence_pct is not None else None


def _backtest_forecast_weight_patterns(
    quantities: np.ndarray,
    labels: list[str],
) -> tuple[tuple[float, ...], float, int, str, str, list[dict]]:
    lookback = len(_FORECAST_WEIGHT_PATTERNS[0])
    evaluation_points = len(quantities) - lookback
    start_idx = len(quantities) - evaluation_points
    best_weights = _FORECAST_WEIGHT_PATTERNS[0]
    best_mae = float("inf")
    candidate_scores: list[dict] = []

    for weights in _FORECAST_WEIGHT_PATTERNS:
        errors: list[float] = []
        for idx in range(start_idx, len(quantities)):
            window = quantities[idx - len(weights):idx][::-1]
            predicted = _weighted_moving_average(window, weights)
            errors.append(abs(float(quantities[idx]) - predicted))
        mae = float(np.mean(errors)) if errors else float("inf")
        candidate_scores.append(
            {
                "weights": [round(weight, 6) for weight in weights],
                "weights_label": _format_weight_pattern(weights),
                "mae": round(_safe_float(mae), 2),
                "evaluation_points": int(len(errors)),
            }
        )
        if mae < best_mae:
            best_weights = weights
            best_mae = mae
    backtest_start_label = labels[start_idx] if evaluation_points > 0 else ""
    backtest_end_label = labels[-1] if evaluation_points > 0 else ""
    candidate_scores = sorted(candidate_scores, key=lambda item: item["mae"])
    for idx, item in enumerate(candidate_scores, start=1):
        item["rank"] = idx
        item["is_selected"] = item["weights_label"] == _format_weight_pattern(best_weights)
    return best_weights, best_mae, evaluation_points, backtest_start_label, backtest_end_label, candidate_scores


def _compute_prediction_bands(
    quantities: np.ndarray,
    labels: list[str],
    forecast_func,
    forecast_values: list[float],
    horizon: int = 6,
) -> dict:
    """Select either P10/P90 or P5/P95 using walk-forward backtesting."""
    n = len(quantities)
    eval_start = max(n - 18, 4)  # use up to 18 months of backtest data

    pct_errors: list[float] = []

    for t in range(eval_start, n):
        train = quantities[:t]
        train_labels = labels[:t]
        actual = float(quantities[t])
        if len(train) < 4:
            continue
        try:
            forecasted, _ = forecast_func(train, train_labels, horizon=1)
            pred = forecasted[0] if forecasted else 0.0
        except Exception:
            continue
        if pred > 0:
            pct_errors.append((actual - pred) / pred)
        elif actual > 0:
            pct_errors.append(1.0)  # underforecast
        else:
            pct_errors.append(0.0)

    p50 = [max(_safe_float(v), 0.0) for v in forecast_values]
    if len(pct_errors) < 3:
        p5_vals  = [max(v * 0.60, 0.0) for v in p50]
        p10_vals = [max(v * 0.70, 0.0) for v in p50]
        p90_vals = [max(v * 1.30, 0.0) for v in p50]
        p95_vals = [max(v * 1.40, 0.0) for v in p50]
        return {
            "selected_band_label": "P10/P90",
            "selected_lower_label": "P10",
            "selected_upper_label": "P90",
            "selected_lower_quantities": [round(v, 2) for v in p10_vals],
            "selected_upper_quantities": [round(v, 2) for v in p90_vals],
            "p5_quantities":  [round(v, 2) for v in p5_vals],
            "p10_quantities": [round(v, 2) for v in p10_vals],
            "p50_quantities": [round(v, 2) for v in p50],
            "p90_quantities": [round(v, 2) for v in p90_vals],
            "p95_quantities": [round(v, 2) for v in p95_vals],
            "selected_band_coverage_pct": None,
            "selected_band_target_coverage_pct": 80.0,
            "selected_band_fit_score": None,
            "band_selection_window_label": "",
            "candidate_bands": [],
        }

    err_array = np.array(pct_errors)
    p50_adj = float(np.percentile(err_array, 50))
    p50_vals = [max(v * (1.0 + p50_adj), 0.0) for v in p50]

    candidate_specs = [
        ("P10/P90", "P10", "P90", 10, 90, 80.0),
        ("P5/P95", "P5", "P95", 5, 95, 90.0),
    ]
    candidate_bands: list[dict] = []

    for band_label, lower_label, upper_label, lower_q, upper_q, target_coverage in candidate_specs:
        lower_adj = float(np.percentile(err_array, lower_q))
        upper_adj = float(np.percentile(err_array, upper_q))

        lower_vals = [max(v * (1.0 + lower_adj), 0.0) for v in p50]
        upper_vals = [max(v * (1.0 + upper_adj), 0.0) for v in p50]
        for idx in range(len(p50_vals)):
            lower_vals[idx] = min(lower_vals[idx], p50_vals[idx])
            upper_vals[idx] = max(upper_vals[idx], p50_vals[idx])

        in_band = 0
        miss_gap_pcts: list[float] = []
        width_pcts: list[float] = []
        backtest_rows = 0
        for t in range(eval_start, n):
            train = quantities[:t]
            train_labels = labels[:t]
            actual = float(quantities[t])
            if len(train) < 4:
                continue
            try:
                forecasted, _ = forecast_func(train, train_labels, horizon=1)
                pred = max(_safe_float(forecasted[0] if forecasted else 0.0), 0.0)
            except Exception:
                continue
            p50_bt = max(pred * (1.0 + p50_adj), 0.0)
            low = min(max(pred * (1.0 + lower_adj), 0.0), p50_bt)
            high = max(max(pred * (1.0 + upper_adj), 0.0), p50_bt)
            denom = max(actual, p50_bt, 1e-9)
            width_pcts.append(((high - low) / denom) * 100.0)
            if low <= actual <= high:
                in_band += 1
                miss_gap_pcts.append(0.0)
            elif actual < low:
                miss_gap_pcts.append(((low - actual) / denom) * 100.0)
            else:
                miss_gap_pcts.append(((actual - high) / denom) * 100.0)
            backtest_rows += 1

        coverage_pct = (in_band / backtest_rows * 100.0) if backtest_rows else 0.0
        avg_width_pct = float(np.mean(width_pcts)) if width_pcts else 0.0
        avg_miss_gap_pct = float(np.mean(miss_gap_pcts)) if miss_gap_pcts else 0.0
        fit_score = abs(coverage_pct - target_coverage) + (0.10 * avg_width_pct) + avg_miss_gap_pct
        candidate_bands.append(
            {
                "label": band_label,
                "lower_label": lower_label,
                "upper_label": upper_label,
                "lower_quantities": [round(v, 2) for v in lower_vals],
                "upper_quantities": [round(v, 2) for v in upper_vals],
                "target_coverage_pct": target_coverage,
                "coverage_pct": round(coverage_pct, 1),
                "avg_width_pct": round(avg_width_pct, 2),
                "avg_miss_gap_pct": round(avg_miss_gap_pct, 2),
                "fit_score": round(fit_score, 2),
            }
        )

    candidate_bands = sorted(candidate_bands, key=lambda item: (item["fit_score"], item["avg_width_pct"]))
    selected = candidate_bands[0]
    for band in candidate_bands:
        band["is_selected"] = band["label"] == selected["label"]

    # Extract all 5 percentiles explicitly from the two candidate bands
    band_map: dict[str, dict] = {b["label"]: b for b in candidate_bands}
    b_p10p90 = band_map.get("P10/P90", candidate_bands[0])
    b_p5p95  = band_map.get("P5/P95",  candidate_bands[-1])
    p10_quantities = b_p10p90["lower_quantities"]
    p90_quantities = b_p10p90["upper_quantities"]
    p5_quantities  = b_p5p95["lower_quantities"]
    p95_quantities = b_p5p95["upper_quantities"]
    p50_list = [round(v, 2) for v in p50_vals]

    window_label = f"{labels[eval_start]} to {labels[-1]}" if n > eval_start else ""
    return {
        "selected_band_label": selected["label"],
        "selected_lower_label": selected["lower_label"],
        "selected_upper_label": selected["upper_label"],
        "selected_lower_quantities": selected["lower_quantities"],
        "selected_upper_quantities": selected["upper_quantities"],
        "p5_quantities":  p5_quantities,
        "p10_quantities": p10_quantities,
        "p50_quantities": p50_list,
        "p90_quantities": p90_quantities,
        "p95_quantities": p95_quantities,
        "selected_band_coverage_pct": selected["coverage_pct"],
        "selected_band_target_coverage_pct": selected["target_coverage_pct"],
        "selected_band_fit_score": selected["fit_score"],
        "band_selection_window_label": window_label,
        "candidate_bands": candidate_bands,
    }


def _build_forecast(monthly: pd.DataFrame) -> dict:
    history_labels = monthly.get("label", pd.Series(dtype=str)).tolist()
    history_quantities = [round(_safe_float(v), 2) for v in monthly.get("qty", pd.Series(dtype=float)).tolist()]

    _empty_forecast_stub = {
        "forecast_labels": [],
        "forecast_quantities": [],
        "projected_total_qty": 0.0,
        "avg_forecast_qty": 0.0,
        "selected_band_label": "",
        "selected_lower_label": "",
        "selected_upper_label": "",
        "selected_lower_quantities": [],
        "selected_upper_quantities": [],
        "p5_quantities": [],
        "p10_quantities": [],
        "p50_quantities": [],
        "p90_quantities": [],
        "p95_quantities": [],
        "lower_total_qty": 0.0,
        "p50_total_qty": 0.0,
        "upper_total_qty": 0.0,
        "p5_total_qty": 0.0,
        "p95_total_qty": 0.0,
        "buffer_total_qty": 0.0,
        "band_selection_window_label": "",
        "selected_band_coverage_pct": None,
        "selected_band_target_coverage_pct": None,
        "selected_band_fit_score": None,
        "candidate_bands": [],
        "forecast_rows": [],
    }

    if monthly.empty or len(monthly) < _MIN_HISTORY_MONTHS:
        return {
            "method": "Insufficient history",
            "history_labels": history_labels,
            "history_quantities": history_quantities,
            "consistency_message": f"At least {_MIN_HISTORY_MONTHS} months of history are required for forecasting.",
            "recent_active_months": int((monthly.get("qty", pd.Series(dtype=float)).tail(6) > 0).sum()) if not monthly.empty else 0,
            **_empty_forecast_stub,
        }

    quantities = monthly["qty"].to_numpy(dtype=float)
    recent_active_months = int((monthly["qty"].tail(6) > 0).sum())

    # Relaxed gate: allow sporadic/lumpy items through (>=3 active months)
    if recent_active_months < _MIN_RECENT_ACTIVE:
        return {
            "method": "No consistent data for forecasting",
            "history_labels": history_labels,
            "history_quantities": history_quantities,
            "consistency_message": f"No consistent data for forecasting. At least {_MIN_RECENT_ACTIVE} of the last 6 months must have active consumption.",
            "recent_active_months": recent_active_months,
            **_empty_forecast_stub,
        }

    # --- Adaptive ensemble forecasting ---
    demand_type = _classify_demand(quantities)

    # Generate ensemble forecast
    def _ensemble_func(q, lbl, horizon=6):
        return _ensemble_forecast(q, lbl, horizon=horizon)

    forecast_values, ensemble_meta = _ensemble_forecast(quantities, history_labels, horizon=6)

    # Also run the new engine's model selector for bootstrap-based bands
    try:
        bootstrap_result = compute_bootstrap_bands(quantities, forecast_values, horizon=6)
        new_engine_validation = bootstrap_result.get("validation", {})
        new_engine_model = bootstrap_result.get("model_type", "")
    except Exception:
        bootstrap_result = None
        new_engine_validation = {}
        new_engine_model = ""

    forecast_array = np.asarray(forecast_values, dtype=float)
    last_period = monthly["period"].iloc[-1]
    forecast_periods = pd.period_range(last_period + 1, periods=6, freq="M")
    prediction_bands = _compute_prediction_bands(
        quantities, history_labels, _ensemble_func,
        forecast_values, horizon=6,
    )

    # Merge bootstrap bands into prediction_bands when they provide tighter coverage
    if bootstrap_result is not None:
        for key in ("p5_quantities", "p10_quantities", "p50_quantities", "p90_quantities", "p95_quantities"):
            bs_vals = bootstrap_result.get(key)
            if bs_vals and len(bs_vals) == 6:
                existing = prediction_bands.get(key, [])
                # Use bootstrap values if existing percentile bands fell back to fixed multipliers
                if not existing or prediction_bands.get("selected_band_coverage_pct") is None:
                    prediction_bands[key] = bs_vals
    lower_quantities = prediction_bands["selected_lower_quantities"]
    p50_quantities   = prediction_bands["p50_quantities"]
    upper_quantities = prediction_bands["selected_upper_quantities"]
    p5_quantities    = prediction_bands["p5_quantities"]
    p10_quantities   = prediction_bands["p10_quantities"]
    p90_quantities   = prediction_bands["p90_quantities"]
    p95_quantities   = prediction_bands["p95_quantities"]
    forecast_rows = []
    for idx, label in enumerate([_format_period_label(period.year, period.month) for period in forecast_periods]):
        p5_qty  = _safe_float(p5_quantities[idx])  if idx < len(p5_quantities)  else 0.0
        p10_qty = _safe_float(p10_quantities[idx]) if idx < len(p10_quantities) else 0.0
        p50_qty = _safe_float(p50_quantities[idx]) if idx < len(p50_quantities) else 0.0
        p90_qty = _safe_float(p90_quantities[idx]) if idx < len(p90_quantities) else 0.0
        p95_qty = _safe_float(p95_quantities[idx]) if idx < len(p95_quantities) else 0.0
        forecast_rows.append(
            {
                "period": label,
                "p5_qty":  round(p5_qty,  2),
                "p10_qty": round(p10_qty, 2),
                "p50_qty": round(p50_qty, 2),
                "p90_qty": round(p90_qty, 2),
                "p95_qty": round(p95_qty, 2),
                "buffer_qty": round(max(p95_qty - p5_qty, 0.0), 2),
                # legacy aliases for any existing callers
                "lower_qty": round(p10_qty, 2),
                "upper_qty": round(p90_qty, 2),
            }
        )

    method_label = f"Adaptive ensemble ({ensemble_meta.get('selected_method', 'blend')})"
    if new_engine_model:
        method_label += f" + {new_engine_model}"
    if demand_type in ("sporadic", "lumpy"):
        method_label += f" [intermittent: {demand_type}]"

    # Attach walk-forward validation from new engine if available
    validation_info = {}
    if new_engine_validation:
        validation_info = {
            "wf_mape": new_engine_validation.get("mape"),
            "wf_bias_pct": new_engine_validation.get("bias_pct"),
            "wf_coverage": new_engine_validation.get("coverage"),
            "wf_conf_score": new_engine_validation.get("conf_score"),
            "wf_model": new_engine_validation.get("model_used"),
        }

    return {
        "method": method_label,
        "history_labels": monthly["label"].tolist(),
        "history_quantities": [round(_safe_float(v), 2) for v in monthly["qty"].tolist()],
        "forecast_labels": [_format_period_label(period.year, period.month) for period in forecast_periods],
        "forecast_quantities": [round(_safe_float(v), 2) for v in p50_quantities],
        "projected_total_qty": round(sum(p50_quantities), 2),
        "avg_forecast_qty": round((sum(p50_quantities) / len(p50_quantities)) if p50_quantities else 0.0, 2),
        "consistency_message": f"Adaptive {demand_type} forecast using the backtested {prediction_bands.get('selected_band_label', '')} interval.",
        "recent_active_months": recent_active_months,
        "selected_band_label": prediction_bands["selected_band_label"],
        "selected_lower_label": prediction_bands["selected_lower_label"],
        "selected_upper_label": prediction_bands["selected_upper_label"],
        "selected_lower_quantities": lower_quantities,
        "selected_upper_quantities": upper_quantities,
        "p5_quantities":  p5_quantities,
        "p10_quantities": p10_quantities,
        "p50_quantities": p50_quantities,
        "p90_quantities": p90_quantities,
        "p95_quantities": p95_quantities,
        "lower_total_qty": round(sum(lower_quantities), 2),
        "p50_total_qty": round(sum(p50_quantities), 2),
        "upper_total_qty": round(sum(upper_quantities), 2),
        "p5_total_qty":  round(sum(_safe_float(v) for v in p5_quantities), 2),
        "p95_total_qty": round(sum(_safe_float(v) for v in p95_quantities), 2),
        "buffer_total_qty": round(sum(max(_safe_float(p95_quantities[i]) - _safe_float(p5_quantities[i]), 0.0) for i in range(len(forecast_periods))), 2),
        "band_selection_window_label": prediction_bands["band_selection_window_label"],
        "selected_band_coverage_pct": prediction_bands["selected_band_coverage_pct"],
        "selected_band_target_coverage_pct": prediction_bands["selected_band_target_coverage_pct"],
        "selected_band_fit_score": prediction_bands["selected_band_fit_score"],
        "candidate_bands": prediction_bands["candidate_bands"],
        "forecast_rows": forecast_rows,
        "demand_type": demand_type,
        "ensemble_meta": ensemble_meta,
        **validation_info,
    }


def _build_yoy(cons_monthly: pd.DataFrame, proc_monthly: pd.DataFrame) -> dict:
    if cons_monthly.empty and proc_monthly.empty:
        return {"records": []}

    qty_fy = cons_monthly.copy()
    if not qty_fy.empty:
        qty_fy["financial_year"] = qty_fy.apply(
            lambda row: _financial_year_label(int(row["year"]), int(row["month"])),
            axis=1,
        )
        qty_fy["financial_year_start"] = qty_fy["year"].where(qty_fy["month"] >= 4, qty_fy["year"] - 1).astype(int)
        qty_yearly = (
            qty_fy.groupby(["financial_year", "financial_year_start"], as_index=False)
            .agg(
                total_qty=("qty", "sum"),
                months_observed=("month", "nunique"),
                active_months=("qty", lambda values: int((values > 0).sum())),
            )
            .sort_values("financial_year_start")
        )
    else:
        qty_yearly = pd.DataFrame(
            columns=["financial_year", "financial_year_start", "total_qty", "months_observed", "active_months"]
        )

    proc_fy = proc_monthly.copy()
    if not proc_fy.empty:
        proc_fy["financial_year"] = proc_fy.apply(
            lambda row: _financial_year_label(int(row["year"]), int(row["month"])),
            axis=1,
        )
        proc_fy["financial_year_start"] = proc_fy["year"].where(proc_fy["month"] >= 4, proc_fy["year"] - 1).astype(int)
        proc_yearly = (
            proc_fy.groupby(["financial_year", "financial_year_start"], as_index=False)
            .agg(total_procurement_value=("value", "sum"))
            .sort_values("financial_year_start")
        )
    else:
        proc_yearly = pd.DataFrame(
            columns=["financial_year", "financial_year_start", "total_procurement_value"]
        )

    if qty_yearly.empty:
        yearly = proc_yearly.copy()
        yearly["months_observed"] = 0
        yearly["active_months"] = 0
        yearly["total_qty"] = 0.0
    else:
        yearly = qty_yearly.copy()

    yearly = yearly.merge(
        proc_yearly,
        on=["financial_year", "financial_year_start"],
        how="outer",
    ).fillna(
        {
            "months_observed": 0,
            "active_months": 0,
            "total_qty": 0.0,
            "total_procurement_value": 0.0,
        }
    ).sort_values("financial_year_start")
    yearly["avg_monthly_qty"] = yearly["total_qty"] / yearly["months_observed"].replace(0, pd.NA)
    yearly["avg_monthly_qty"] = yearly["avg_monthly_qty"].fillna(0.0)

    records: list[dict] = []
    previous = None
    for _, row in yearly.iterrows():
        qty_change = None
        value_change = None
        if previous is not None and int(previous["months_observed"]) == int(row["months_observed"]) and int(row["months_observed"]) > 0:
            if _safe_float(previous["total_qty"]) > 0:
                qty_change = ((_safe_float(row["total_qty"]) - _safe_float(previous["total_qty"])) / _safe_float(previous["total_qty"])) * 100
        if previous is not None and _safe_float(previous["total_procurement_value"]) > 0 and _safe_float(row["total_procurement_value"]) > 0:
            value_change = (
                (_safe_float(row["total_procurement_value"]) - _safe_float(previous["total_procurement_value"]))
                / _safe_float(previous["total_procurement_value"])
            ) * 100
        records.append(
            {
                "financial_year": str(row["financial_year"]),
                "months_observed": int(row["months_observed"]),
                "active_months": int(row["active_months"]),
                "total_qty": round(_safe_float(row["total_qty"]), 2),
                "total_value": round(_safe_float(row["total_procurement_value"]), 2),
                "avg_monthly_qty": round(_safe_float(row["avg_monthly_qty"]), 2),
                "yoy_qty_pct": round(_safe_float(qty_change), 2) if qty_change is not None else None,
                "yoy_value_pct": round(_safe_float(value_change), 2) if value_change is not None else None,
            }
        )
        previous = row

    return {"records": records}


def _build_financial_year_summary(cons_monthly: pd.DataFrame, proc_monthly: pd.DataFrame) -> dict:
    if cons_monthly.empty and proc_monthly.empty:
        return {"records": []}

    qty_fy = cons_monthly.copy()
    if not qty_fy.empty:
        qty_fy["financial_year"] = qty_fy.apply(
            lambda row: _financial_year_label(int(row["year"]), int(row["month"])),
            axis=1,
        )
        qty_fy["financial_year_start"] = qty_fy["year"].where(qty_fy["month"] >= 4, qty_fy["year"] - 1).astype(int)
        qty_yearly = (
            qty_fy.groupby(["financial_year", "financial_year_start"], as_index=False)
            .agg(
                months_observed=("month", "nunique"),
                active_months=("qty", lambda values: int((values > 0).sum())),
                total_qty=("qty", "sum"),
                total_consumption_value=("value", "sum"),
            )
            .sort_values("financial_year_start")
        )
    else:
        qty_yearly = pd.DataFrame(
            columns=[
                "financial_year",
                "financial_year_start",
                "months_observed",
                "active_months",
                "total_qty",
                "total_consumption_value",
            ]
        )

    proc_fy = proc_monthly.copy()
    if not proc_fy.empty:
        proc_fy["financial_year"] = proc_fy.apply(
            lambda row: _financial_year_label(int(row["year"]), int(row["month"])),
            axis=1,
        )
        proc_fy["financial_year_start"] = proc_fy["year"].where(proc_fy["month"] >= 4, proc_fy["year"] - 1).astype(int)
        proc_yearly = (
            proc_fy.groupby(["financial_year", "financial_year_start"], as_index=False)
            .agg(total_procurement_value=("value", "sum"))
            .sort_values("financial_year_start")
        )
    else:
        proc_yearly = pd.DataFrame(
            columns=["financial_year", "financial_year_start", "total_procurement_value"]
        )

    if qty_yearly.empty:
        yearly = proc_yearly.copy()
        yearly["months_observed"] = 0
        yearly["active_months"] = 0
        yearly["total_qty"] = 0.0
        yearly["total_consumption_value"] = 0.0
    else:
        yearly = qty_yearly.copy()

    yearly = yearly.merge(
        proc_yearly,
        on=["financial_year", "financial_year_start"],
        how="outer",
    ).fillna(
        {
            "months_observed": 0,
            "active_months": 0,
            "total_qty": 0.0,
            "total_consumption_value": 0.0,
            "total_procurement_value": 0.0,
        }
    ).sort_values("financial_year_start")

    yearly["avg_monthly_qty"] = yearly["total_qty"] / yearly["months_observed"].replace(0, pd.NA)
    yearly["avg_monthly_qty"] = yearly["avg_monthly_qty"].fillna(0.0)

    records: list[dict] = []
    for _, row in yearly.iterrows():
        records.append(
            {
                "financial_year": str(row["financial_year"]),
                "months_observed": int(row["months_observed"]),
                "active_months": int(row["active_months"]),
                "total_qty": round(_safe_float(row["total_qty"]), 2),
                "total_consumption_value": round(_safe_float(row["total_consumption_value"]), 2),
                "total_procurement_value": round(_safe_float(row["total_procurement_value"]), 2),
                "avg_monthly_qty": round(_safe_float(row["avg_monthly_qty"]), 2),
            }
        )

    return {"records": records}


def _quick_forecast_total(monthly: pd.DataFrame, horizon: int = 6) -> float:
    if monthly.empty:
        return 0.0

    last_period = monthly["period"].iloc[-1]
    forecast_periods = pd.period_range(last_period + 1, periods=horizon, freq="M")
    history_by_month = {
        int(month_number): group["qty"].to_numpy(dtype=float)
        for month_number, group in monthly.groupby("month")
    }
    all_history = monthly["qty"].to_numpy(dtype=float)
    total = 0.0
    for period in forecast_periods:
        sample = history_by_month.get(int(period.month))
        values = sample if sample is not None and len(sample) else all_history
        total += max(_safe_float(np.median(values)), 0.0)
    return round(total, 2)


def _build_vendor_scores(proc_df: pd.DataFrame) -> list[dict]:
    pricing = proc_df.dropna(subset=["vendor", "unit_price", "doc_date"]).copy()
    if pricing.empty:
        return []

    grouped = pricing.groupby("vendor", as_index=False)
    summary = grouped.agg(
        avg_unit_price=("unit_price", "mean"),
        lowest_unit_price=("unit_price", "min"),
        highest_unit_price=("unit_price", "max"),
        latest_doc_date=("doc_date", "max"),
        transactions=("po_number", "count"),
        spend=("effective_value", "sum"),
    )
    latest_prices = (
        pricing.sort_values("doc_date")
        .groupby("vendor")
        .tail(1)[["vendor", "unit_price"]]
        .rename(columns={"unit_price": "latest_unit_price"})
    )
    volatility = (
        pricing.groupby("vendor")["unit_price"]
        .std(ddof=0)
        .fillna(0.0)
        .reset_index()
        .rename(columns={"unit_price": "price_std"})
    )
    summary = summary.merge(latest_prices, on="vendor", how="left")
    summary = summary.merge(volatility, on="vendor", how="left")

    best_avg_price = summary["avg_unit_price"].min()
    total_spend = summary["spend"].sum() or 1.0

    records: list[dict] = []
    for _, row in summary.sort_values(["avg_unit_price", "transactions", "vendor"]).iterrows():
        avg_unit_price = _safe_float(row["avg_unit_price"])
        price_std = _safe_float(row["price_std"])
        volatility_pct = (price_std / avg_unit_price * 100) if avg_unit_price else 0.0
        competitiveness = (best_avg_price / avg_unit_price) if avg_unit_price else 0.0
        reliability = min(_safe_float(row["transactions"]) / 5.0, 1.0)
        consistency = max(0.0, 1.0 - min(volatility_pct / 100.0, 1.0))
        score = 100.0 * ((0.7 * competitiveness) + (0.2 * consistency) + (0.1 * reliability))
        premium_pct = ((avg_unit_price - best_avg_price) / best_avg_price * 100) if best_avg_price else 0.0

        records.append(
            {
                "vendor": row["vendor"],
                "transactions": int(row["transactions"]),
                "avg_unit_price": round(avg_unit_price, 2),
                "latest_unit_price": round(_safe_float(row["latest_unit_price"]), 2),
                "lowest_unit_price": round(_safe_float(row["lowest_unit_price"]), 2),
                "highest_unit_price": round(_safe_float(row["highest_unit_price"]), 2),
                "volatility_pct": round(volatility_pct, 2),
                "premium_vs_best_pct": round(premium_pct, 2),
                "spend_share_pct": round((_safe_float(row["spend"]) / total_spend) * 100, 2),
                "latest_doc_date": row["latest_doc_date"].strftime("%d %b %Y") if pd.notna(row["latest_doc_date"]) else "",
                "score": round(max(0.0, min(score, 100.0)), 1),
            }
        )
    return records

def _build_cost_variance(proc_df: pd.DataFrame) -> dict:
    timeline = proc_df.dropna(subset=["doc_date", "unit_price"]).copy()
    if timeline.empty:
        return {
            "threshold_pct": _PRICE_SWING_THRESHOLD_PCT,
            "latest_unit_price": 0.0,
            "avg_unit_price": 0.0,
            "max_increase_pct": 0.0,
            "max_decrease_pct": 0.0,
            "significant_events": [],
            "significant_event_count": 0,
        }

    timeline = (
        timeline.groupby(["doc_date", "po_number", "vendor"], as_index=False)
        .agg(
            unit_price=("unit_price", "mean"),
            procured_qty=("procured_qty", "sum"),
            effective_value=("effective_value", "sum"),
        )
        .sort_values(["doc_date", "po_number", "vendor"])
    )

    events: list[dict] = []
    previous = None
    for _, row in timeline.iterrows():
        if previous is not None:
            prev_price = _safe_float(previous["unit_price"])
            current_price = _safe_float(row["unit_price"])
            pct_change = 0.0
            if prev_price > 0:
                pct_change = ((current_price - prev_price) / prev_price) * 100
            events.append(
                {
                    "from_date": previous["doc_date"].strftime("%d %b %Y"),
                    "to_date": row["doc_date"].strftime("%d %b %Y"),
                    "from_vendor": previous["vendor"],
                    "to_vendor": row["vendor"],
                    "from_price": round(prev_price, 2),
                    "to_price": round(current_price, 2),
                    "pct_change": round(pct_change, 2),
                    "direction": "Increase" if pct_change >= 0 else "Decrease",
                    "significant": abs(pct_change) >= _PRICE_SWING_THRESHOLD_PCT,
                }
            )
        previous = row

    significant_events = sorted(
        [event for event in events if event["significant"]],
        key=lambda event: abs(event["pct_change"]),
        reverse=True,
    )
    pct_changes = [event["pct_change"] for event in events]

    return {
        "threshold_pct": _PRICE_SWING_THRESHOLD_PCT,
        "latest_unit_price": round(_safe_float(timeline["unit_price"].iloc[-1]), 2),
        "avg_unit_price": round(_safe_float(timeline["unit_price"].mean()), 2),
        "max_increase_pct": round(max([change for change in pct_changes if change > 0] or [0.0]), 2),
        "max_decrease_pct": round(min([change for change in pct_changes if change < 0] or [0.0]), 2),
        "significant_events": significant_events[:8],
        "significant_event_count": len(significant_events),
    }


def _build_storage_breakdown(cons_df: pd.DataFrame, plant: str | None) -> list[dict]:
    if cons_df.empty:
        return []

    group_fields = ["storage_location"] if plant else ["reporting_plant", "storage_location"]

    # ── Step 1: aggregate to monthly net-consumption per group ─────────────
    # For MB51 data each group can have many transaction rows per month.
    # First net them to monthly totals (cancelling reversals), then aggregate
    # the monthly totals to a single row per storage location so that
    # active_months counts months with non-zero net consumption.
    month_group_fields = group_fields + ["year", "month"]
    valid = cons_df.dropna(subset=["year", "month"])
    if valid.empty:
        # Fall back to raw row counts if date columns are missing
        monthly_net = cons_df.copy()
        monthly_net["_net_qty"] = monthly_net["usage_qty"]
        location_monthly = (
            monthly_net.groupby(group_fields, as_index=False)
            .agg(total_qty=("_net_qty", "sum"), active_months=("month_raw", "nunique"))
        )
    else:
        # Net each (location × month) to a single signed quantity
        per_month = (
            valid.groupby(month_group_fields, as_index=False)
            .agg(net_qty=("usage_qty", "sum"))
        )
        # Now roll up to storage-location level
        location_monthly = (
            per_month.groupby(group_fields, as_index=False)
            .agg(
                total_qty=("net_qty", "sum"),
                # active months = months where net consumption is non-zero
                active_months=("net_qty", lambda s: int((s != 0).sum())),
            )
        )

    grouped = location_monthly.sort_values("total_qty", ascending=False)
    total_qty = grouped["total_qty"].sum() or 1.0

    records: list[dict] = []
    for _, row in grouped.head(12).iterrows():
        storage_location = _normalize_text(row.get("storage_location"))
        plant_code = _normalize_text(row.get("reporting_plant")) if not plant else plant
        label = storage_location if plant else f"{plant_code} / {storage_location}"
        records.append(
            {
                "plant": plant_code,
                "storage_location": storage_location,
                "label": label,
                "total_qty": round(_safe_float(row["total_qty"]), 2),
                "active_months": int(row["active_months"]),
                "share_qty_pct": round((_safe_float(row["total_qty"]) / total_qty) * 100, 2),
            }
        )
    return records


def _filter_material_rows(materials: list[dict], query: str = "") -> list[dict]:
    query = _normalize_text(query).lower()
    if not query:
        return list(materials)
    filtered = []
    for material in materials:
        haystacks = [
            _normalize_text(material.get("material")).lower(),
            _normalize_text(material.get("material_code")).lower(),
            " ".join(material.get("vendors", [])).lower(),
        ]
        if any(query in haystack for haystack in haystacks):
            filtered.append(material)
    return filtered


class _DataStore:
    """Lazy-loaded, cached data store for inventory intelligence."""

    def __init__(self):
        self._cons: pd.DataFrame | None = None
        self._proc: pd.DataFrame | None = None
        self._overview_cache: dict[str, dict] = {}
        self._materials_cache: dict[str, list[dict]] = {}
        self._dashboard_cache: dict[str, dict] = {}
        self._material_context_cache: dict[tuple[str, str], dict] = {}
        self._analytics_overview_cache: dict[tuple[str, str], dict] = {}
        self._analytics_section_cache: dict[tuple[str, str, str], dict] = {}
        self._consumption_detail_cache: dict[tuple[str, str], list[dict]] = {}
        self._procurement_history_cache: dict[tuple[str, str], list[dict]] = {}
        self._persistent_forecast_cache: dict[str, dict] | None = None

    def _ensure_loaded(self):
        if self._cons is None:
            logger.info("Loading consumption seed data from %s", CONS_FILE)
            self._cons = _load_consumption()
            logger.info("Loaded %d consumption records", len(self._cons))
        if self._proc is None:
            logger.info("Loading procurement seed data from %s", PROC_FILE)
            self._proc = _load_procurement()
            logger.info("Loaded %d procurement records", len(self._proc))

    @property
    def cons(self) -> pd.DataFrame:
        self._ensure_loaded()
        return self._cons

    @property
    def proc(self) -> pd.DataFrame:
        self._ensure_loaded()
        return self._proc

    def reload(self):
        self._cons = None
        self._proc = None
        self._overview_cache = {}
        self._materials_cache = {}
        self._dashboard_cache = {}
        self._material_context_cache = {}
        self._analytics_overview_cache = {}
        self._analytics_section_cache = {}
        self._consumption_detail_cache = {}
        self._procurement_history_cache = {}
        self._persistent_forecast_cache = None

    def _cache_key(self, plant: str | None) -> str:
        return plant or "__all__"

    def _material_cache_key(self, material: str, plant: str | None) -> tuple[str, str]:
        return self._cache_key(_normalize_plant(plant)), material.upper()

    def _get_material_context(self, material: str, plant: str | None = None) -> dict:
        normalized_material = material.upper()
        normalized_plant = _normalize_plant(plant)
        cache_key = self._material_cache_key(normalized_material, normalized_plant)
        if cache_key in self._material_context_cache:
            return self._material_context_cache[cache_key]

        cons = self._filtered_consumption(normalized_plant)
        proc = self._filtered_procurement(normalized_plant)
        cons_df = cons[cons["material_desc"] == normalized_material].copy()
        proc_df = proc[proc["material_desc"] == normalized_material].copy()
        monthly = _monthly_series(cons_df)
        proc_monthly = _monthly_procurement_series(proc_df)

        if not monthly.empty:
            total_qty = round(_safe_float(monthly["qty"].sum()), 2)
            active_months = int((monthly["qty"] != 0).sum())
        else:
            total_qty = 0.0
            active_months = 0

        context = {
            "material": normalized_material,
            "plant": normalized_plant or "",
            "cons_df": cons_df,
            "proc_df": proc_df,
            "monthly": monthly,
            "proc_monthly": proc_monthly,
            "summary": {
                "total_qty": total_qty,
                "total_value": round(_safe_float(proc_df["effective_value"].sum()), 2),
                "active_months": active_months,
                "storage_location_count": int(cons_df["storage_location"].nunique()) if not cons_df.empty else 0,
                "vendor_count": int(proc_df["vendor"].nunique()) if not proc_df.empty else 0,
                "last_unit_price": round(
                    _safe_float(proc_df.sort_values("doc_date")["unit_price"].iloc[-1]),
                    2,
                ) if not proc_df.empty else 0.0,
            },
            "consumption": {
                "labels": monthly.get("label", pd.Series(dtype=str)).tolist(),
                "quantities": [
                    round(_safe_float(value), 2)
                    for value in monthly.get("qty", pd.Series(dtype=float)).tolist()
                ],
                "values": [
                    round(_safe_float(value), 2)
                    for value in monthly.get("value", pd.Series(dtype=float)).tolist()
                ],
            },
        }
        self._material_context_cache[cache_key] = context
        return context

    def _get_persistent_forecast_cache(self) -> dict[str, dict]:
        if self._persistent_forecast_cache is None:
            self._persistent_forecast_cache = _load_precomputed_forecast_cache()
        return self._persistent_forecast_cache

    def _store_persistent_forecast(self, material: str, plant: str | None, forecast: dict) -> None:
        forecasts = self._get_persistent_forecast_cache()
        forecasts[_forecast_cache_entry_key(material, plant)] = forecast
        _write_precomputed_forecast_cache(forecasts)

    def _db_seed_tables_available(self) -> bool:
        global _DB_SEED_TABLES_READY
        try:
            if _DB_SEED_TABLES_READY is True:
                return True
            inspector = inspect(db.engine)
            ready = (
                inspector.has_table("inventory_consumption_seed_rows")
                and inspector.has_table("inventory_procurement_seed_rows")
            )
            if ready:
                _DB_SEED_TABLES_READY = True
            return ready
        except Exception:
            logger.warning("Failed to inspect inventory seed tables; falling back to Excel reads", exc_info=True)
            return False

    def _db_seed_rows_available(self) -> bool:
        if not self._db_seed_tables_available():
            return False
        try:
            from app.models.inventory.inventory_consumption_seed import InventoryConsumptionSeed

            row_count = db.session.query(func.count(InventoryConsumptionSeed.id)).scalar() or 0
            if row_count:
                return True
            return self._bootstrap_db_seed_tables_if_empty()
        except Exception:
            logger.warning("Failed to query inventory seed tables; falling back to Excel reads", exc_info=True)
            return False

    def _bootstrap_db_seed_tables_if_empty(self) -> bool:
        from app.core.services.inventory_seed_db import sync_inventory_seed_tables

        if not (os.path.exists(CONS_FILE) and os.path.exists(PROC_FILE)):
            return False
        logger.info("Bootstrapping normalized inventory seed tables from seed workbooks")
        sync_inventory_seed_tables(
            consumption_frame=_load_consumption(),
            procurement_frame=_load_procurement(),
            consumption_filename=os.path.basename(CONS_FILE),
            procurement_filename=os.path.basename(PROC_FILE),
        )
        from app.models.inventory.inventory_consumption_seed import InventoryConsumptionSeed

        return bool(db.session.query(func.count(InventoryConsumptionSeed.id)).scalar())

    def _db_consumption_filter(self, query, model, plant: str | None):
        if plant:
            query = query.filter(model.reporting_plant == plant)
        return query

    def _db_procurement_filter(self, query, model, plant: str | None):
        if plant:
            query = query.filter(model.reporting_plant == plant)
        return query

    def _get_db_filter_options(self) -> dict:
        from app.models.inventory.inventory_consumption_seed import InventoryConsumptionSeed

        query = (
            db.session.query(InventoryConsumptionSeed.reporting_plant)
            .filter(InventoryConsumptionSeed.reporting_plant.isnot(None))
            .distinct()
            .order_by(InventoryConsumptionSeed.reporting_plant.asc())
        )
        return {"plants": [row[0] for row in query.all() if row[0]]}

    def _build_db_top_consumers(self, plant: str | None, limit: int = 10) -> list[dict]:
        from app.models.inventory.inventory_consumption_seed import InventoryConsumptionSeed
        from app.models.inventory.inventory_procurement_seed import InventoryProcurementSeed

        cons_query = self._db_consumption_filter(
            db.session.query(
                InventoryConsumptionSeed.material_desc.label("material"),
                func.min(InventoryConsumptionSeed.material_code).label("material_code"),
                func.sum(InventoryConsumptionSeed.usage_qty).label("total_qty"),
                func.min(InventoryConsumptionSeed.uom).label("uom"),
            )
            .filter(InventoryConsumptionSeed.material_desc.isnot(None))
            .group_by(InventoryConsumptionSeed.material_desc),
            InventoryConsumptionSeed,
            plant,
        )
        cons_map = {
            row.material: {
                "material": row.material,
                "material_code": _normalize_text(row.material_code),
                "total_qty": round(_safe_float(row.total_qty), 2),
                "uom": _normalize_text(row.uom),
            }
            for row in cons_query.all()
            if row.material
        }

        proc_query = self._db_procurement_filter(
            db.session.query(
                InventoryProcurementSeed.material_desc.label("material"),
                func.sum(InventoryProcurementSeed.effective_value).label("total_value"),
            )
            .filter(InventoryProcurementSeed.material_desc.isnot(None))
            .group_by(InventoryProcurementSeed.material_desc),
            InventoryProcurementSeed,
            plant,
        )
        for row in proc_query.all():
            if row.material not in cons_map:
                cons_map[row.material] = {
                    "material": row.material,
                    "material_code": "",
                    "total_qty": 0.0,
                    "uom": "",
                }
            cons_map[row.material]["total_value"] = round(_safe_float(row.total_value), 2)

        top_rows = sorted(
            cons_map.values(),
            key=lambda item: item.get("total_value", 0.0),
            reverse=True,
        )[:limit]
        for item in top_rows:
            item.setdefault("total_value", 0.0)
        return top_rows

    def _build_db_kpis(self, plant: str | None) -> dict:
        from app.models.inventory.inventory_consumption_seed import InventoryConsumptionSeed
        from app.models.inventory.inventory_procurement_seed import InventoryProcurementSeed

        cons_base = self._db_consumption_filter(
            db.session.query(InventoryConsumptionSeed),
            InventoryConsumptionSeed,
            plant,
        )
        proc_base = self._db_procurement_filter(
            db.session.query(InventoryProcurementSeed),
            InventoryProcurementSeed,
            plant,
        )

        total_materials = (
            self._db_consumption_filter(
                db.session.query(func.count(func.distinct(InventoryConsumptionSeed.material_desc))),
                InventoryConsumptionSeed,
                plant,
            ).scalar()
            or 0
        )
        total_vendors = (
            self._db_procurement_filter(
                db.session.query(func.count(func.distinct(InventoryProcurementSeed.vendor))),
                InventoryProcurementSeed,
                plant,
            )
            .filter(InventoryProcurementSeed.vendor.isnot(None))
            .scalar()
            or 0
        )
        total_procurement_value = (
            self._db_procurement_filter(
                db.session.query(func.sum(InventoryProcurementSeed.effective_value)),
                InventoryProcurementSeed,
                plant,
            ).scalar()
            or 0
        )
        unique_plants = (
            self._db_consumption_filter(
                db.session.query(func.count(func.distinct(InventoryConsumptionSeed.reporting_plant))),
                InventoryConsumptionSeed,
                plant,
            ).scalar()
            or 0
        )
        data_points = cons_base.count()
        period_subq = (
            self._db_consumption_filter(
                db.session.query(
                    InventoryConsumptionSeed.year.label("year"),
                    InventoryConsumptionSeed.month.label("month"),
                ),
                InventoryConsumptionSeed,
                plant,
            )
            .filter(
                InventoryConsumptionSeed.year.isnot(None),
                InventoryConsumptionSeed.month.isnot(None),
            )
            .distinct()
            .subquery()
        )
        total_months = db.session.query(func.count()).select_from(period_subq).scalar() or 0

        min_period = (
            self._db_consumption_filter(
                db.session.query(InventoryConsumptionSeed.year, InventoryConsumptionSeed.month),
                InventoryConsumptionSeed,
                plant,
            )
            .filter(
                InventoryConsumptionSeed.year.isnot(None),
                InventoryConsumptionSeed.month.isnot(None),
            )
            .order_by(InventoryConsumptionSeed.year.asc(), InventoryConsumptionSeed.month.asc())
            .first()
        )
        max_period = (
            self._db_consumption_filter(
                db.session.query(InventoryConsumptionSeed.year, InventoryConsumptionSeed.month),
                InventoryConsumptionSeed,
                plant,
            )
            .filter(
                InventoryConsumptionSeed.year.isnot(None),
                InventoryConsumptionSeed.month.isnot(None),
            )
            .order_by(InventoryConsumptionSeed.year.desc(), InventoryConsumptionSeed.month.desc())
            .first()
        )

        period_from = _format_period_label(int(min_period.year), int(min_period.month)) if min_period else ""
        period_to = _format_period_label(int(max_period.year), int(max_period.month)) if max_period else ""

        return {
            "total_materials": int(total_materials),
            "total_vendors": int(total_vendors),
            "total_procurement_value": round(_safe_float(total_procurement_value), 2),
            "period_from": period_from,
            "period_to": period_to,
            "unique_plants": int(unique_plants),
            "data_points": int(data_points),
            "total_months": int(total_months),
            "selected_plant": plant or "",
        }

    def _build_db_materials_table(self, plant: str | None) -> list[dict]:
        from app.models.inventory.inventory_consumption_seed import InventoryConsumptionSeed
        from app.models.inventory.inventory_procurement_seed import InventoryProcurementSeed

        total_months_subq = (
            self._db_consumption_filter(
                db.session.query(
                    InventoryConsumptionSeed.year.label("year"),
                    InventoryConsumptionSeed.month.label("month"),
                ),
                InventoryConsumptionSeed,
                plant,
            )
            .filter(
                InventoryConsumptionSeed.year.isnot(None),
                InventoryConsumptionSeed.month.isnot(None),
            )
            .distinct()
            .subquery()
        )
        total_distinct_months = db.session.query(func.count()).select_from(total_months_subq).scalar() or 1

        cons_rows = self._db_consumption_filter(
            db.session.query(
                InventoryConsumptionSeed.material_desc.label("material"),
                func.min(InventoryConsumptionSeed.material_code).label("material_code"),
                func.min(InventoryConsumptionSeed.uom).label("uom"),
                func.sum(InventoryConsumptionSeed.usage_qty).label("total_qty"),
                func.count(func.distinct(InventoryConsumptionSeed.month_raw)).label("n_months"),
                func.count(func.distinct(InventoryConsumptionSeed.reporting_plant)).label("plant_count"),
                func.count(func.distinct(InventoryConsumptionSeed.storage_location)).label("storage_location_count"),
                func.min(InventoryConsumptionSeed.currency).label("currency"),
            )
            .filter(InventoryConsumptionSeed.material_desc.isnot(None))
            .group_by(InventoryConsumptionSeed.material_desc),
            InventoryConsumptionSeed,
            plant,
        ).all()

        materials: dict[str, dict] = {}
        for row in cons_rows:
            avg_monthly = _safe_float(row.total_qty) / max(total_distinct_months, 1)
            materials[row.material] = {
                "material": row.material,
                "material_code": _normalize_text(row.material_code),
                "uom": _normalize_text(row.uom),
                "avg_monthly_consumption": round(avg_monthly, 2),
                "avg_annual_consumption": round(avg_monthly * 12, 2),
                "total_qty": round(_safe_float(row.total_qty), 2),
                "total_value": 0.0,
                "n_months": int(row.n_months or 0),
                "currency": _clean_currency(row.currency),
                "last_proc_qty": 0.0,
                "last_proc_value": 0.0,
                "last_po_number": "",
                "last_proc_date": "",
                "last_vendor": "",
                "last_proc_unit": "",
                "last_unit_price": 0.0,
                "vendors": [],
                "vendor_count": 0,
                "plants": [],
                "plant_count": int(row.plant_count or 0),
                "storage_location_count": int(row.storage_location_count or 0),
            }

        proc_totals = self._db_procurement_filter(
            db.session.query(
                InventoryProcurementSeed.material_desc.label("material"),
                func.sum(InventoryProcurementSeed.effective_value).label("total_value"),
                func.min(InventoryProcurementSeed.currency).label("currency"),
            )
            .filter(InventoryProcurementSeed.material_desc.isnot(None))
            .group_by(InventoryProcurementSeed.material_desc),
            InventoryProcurementSeed,
            plant,
        ).all()
        for row in proc_totals:
            if row.material not in materials:
                materials[row.material] = {
                    "material": row.material,
                    "material_code": "",
                    "uom": "",
                    "avg_monthly_consumption": 0.0,
                    "avg_annual_consumption": 0.0,
                    "total_qty": 0.0,
                    "total_value": 0.0,
                    "n_months": 0,
                    "currency": _clean_currency(row.currency),
                    "last_proc_qty": 0.0,
                    "last_proc_value": 0.0,
                    "last_po_number": "",
                    "last_proc_date": "",
                    "last_vendor": "",
                    "last_proc_unit": "",
                    "last_unit_price": 0.0,
                    "vendors": [],
                    "vendor_count": 0,
                    "plants": [],
                    "plant_count": 0,
                    "storage_location_count": 0,
                }
            materials[row.material]["total_value"] = round(_safe_float(row.total_value), 2)
            if not materials[row.material]["currency"]:
                materials[row.material]["currency"] = _clean_currency(row.currency)

        vendor_rows = self._db_procurement_filter(
            db.session.query(
                InventoryProcurementSeed.material_desc,
                InventoryProcurementSeed.vendor,
            )
            .filter(
                InventoryProcurementSeed.material_desc.isnot(None),
                InventoryProcurementSeed.vendor.isnot(None),
            )
            .distinct(),
            InventoryProcurementSeed,
            plant,
        ).all()
        for material_name, vendor in vendor_rows:
            record = materials.get(material_name)
            if not record or not vendor:
                continue
            record["vendors"].append(_normalize_text(vendor))
        for record in materials.values():
            record["vendors"] = sorted(v for v in set(record["vendors"]) if v)
            record["vendor_count"] = len(record["vendors"])

        last_proc_rows = self._db_procurement_filter(
            db.session.query(InventoryProcurementSeed)
            .filter(InventoryProcurementSeed.material_desc.isnot(None))
            .order_by(
                InventoryProcurementSeed.material_desc.asc(),
                InventoryProcurementSeed.doc_date.desc(),
                InventoryProcurementSeed.id.desc(),
            ),
            InventoryProcurementSeed,
            plant,
        ).all()
        seen = set()
        for row in last_proc_rows:
            material_name = row.material_desc
            if not material_name or material_name in seen:
                continue
            seen.add(material_name)
            record = materials.setdefault(material_name, {
                "material": material_name,
                "material_code": _normalize_text(row.material_code),
                "uom": "",
                "avg_monthly_consumption": 0.0,
                "avg_annual_consumption": 0.0,
                "total_qty": 0.0,
                "total_value": 0.0,
                "n_months": 0,
                "currency": _clean_currency(row.currency),
                "last_proc_qty": 0.0,
                "last_proc_value": 0.0,
                "last_po_number": "",
                "last_proc_date": "",
                "last_vendor": "",
                "last_proc_unit": "",
                "last_unit_price": 0.0,
                "vendors": [],
                "vendor_count": 0,
                "plants": [],
                "plant_count": 0,
                "storage_location_count": 0,
            })
            record["last_proc_qty"] = round(_safe_float(row.procured_qty), 2)
            record["last_proc_value"] = round(_safe_float(row.effective_value), 2)
            record["last_po_number"] = _normalize_text(row.po_number)
            record["last_proc_date"] = row.doc_date.strftime("%d %b %Y") if row.doc_date else ""
            record["last_vendor"] = _normalize_text(row.vendor)
            record["last_proc_unit"] = _normalize_text(row.order_unit)
            record["last_unit_price"] = round(_safe_float(row.unit_price), 2)

        return sorted(materials.values(), key=lambda item: item["total_value"], reverse=True)

    def get_filter_options(self) -> dict:
        if self._db_seed_rows_available():
            return self._get_db_filter_options()
        self._ensure_loaded()
        plants = sorted(self._cons["reporting_plant"].dropna().astype(str).unique().tolist())
        return {"plants": plants}

    def _filtered_consumption(self, plant: str | None = None) -> pd.DataFrame:
        cons = self.cons
        normalized_plant = _normalize_plant(plant)
        if not normalized_plant:
            return cons
        return cons[cons["reporting_plant"] == normalized_plant].copy()

    def _filtered_procurement(self, plant: str | None = None) -> pd.DataFrame:
        proc = self.proc
        normalized_plant = _normalize_plant(plant)
        if not normalized_plant:
            return proc
        return proc[proc["reporting_plant"] == normalized_plant].copy()

    def _build_kpis(
        self,
        cons: pd.DataFrame,
        proc: pd.DataFrame,
        normalized_plant: str | None,
    ) -> dict:
        min_year = cons["year"].min() if not cons.empty else None
        min_month = cons.loc[cons["year"] == min_year, "month"].min() if pd.notna(min_year) else None
        max_year = cons["year"].max() if not cons.empty else None
        max_month = cons.loc[cons["year"] == max_year, "month"].max() if pd.notna(max_year) else None
        period_from = _format_period_label(int(min_year), int(min_month)) if pd.notna(min_year) and pd.notna(min_month) else ""
        period_to = _format_period_label(int(max_year), int(max_month)) if pd.notna(max_year) and pd.notna(max_month) else ""
        vendor_count = 0
        if not proc.empty and "vendor" in proc:
            vendor_count = int(
                proc["vendor"]
                .fillna("")
                .astype(str)
                .str.strip()
                .replace("", np.nan)
                .dropna()
                .nunique()
            )

        return {
            "total_materials": int(cons["material_desc"].nunique()) if not cons.empty else 0,
            "total_vendors": vendor_count,
            "total_procurement_value": round(_safe_float(proc["effective_value"].sum()), 2) if not proc.empty else 0.0,
            "period_from": period_from,
            "period_to": period_to,
            "unique_plants": int(cons["reporting_plant"].nunique()) if not cons.empty else 0,
            "data_points": int(len(cons)),
            "total_months": int(cons[["year", "month"]].drop_duplicates().shape[0]) if not cons.empty else 0,
            "selected_plant": normalized_plant or "",
        }

    def _build_top_consumers(
        self,
        cons: pd.DataFrame,
        proc: pd.DataFrame,
        limit: int = 10,
    ) -> list[dict]:
        if cons.empty:
            return []

        cons_totals = (
            cons.groupby("material_desc", as_index=False)
            .agg(
                material_code=("material_code", "first"),
                total_qty=("usage_qty", "sum"),
                uom=("uom", "first"),
            )
            .rename(columns={"material_desc": "material"})
        )
        proc_totals = (
            proc.groupby("material_desc", as_index=False)
            .agg(total_value=("effective_value", "sum"))
            .rename(columns={"material_desc": "material"})
        ) if not proc.empty else pd.DataFrame(columns=["material", "total_value"])

        merged = cons_totals.merge(proc_totals, on="material", how="left")
        merged["total_value"] = merged["total_value"].fillna(0.0)
        merged = merged.sort_values("total_value", ascending=False).head(limit)

        return [
            {
                "material": row["material"],
                "material_code": _normalize_text(row.get("material_code")),
                "total_value": round(_safe_float(row.get("total_value")), 2),
                "total_qty": round(_safe_float(row.get("total_qty")), 2),
                "uom": _normalize_text(row.get("uom")),
            }
            for _, row in merged.iterrows()
        ]

    def _build_materials_summary_table(self, cons: pd.DataFrame, proc: pd.DataFrame) -> list[dict]:
        if cons.empty:
            return []

        total_distinct_months = cons[["year", "month"]].drop_duplicates().shape[0] or 1
        proc_sorted = proc.sort_values("doc_date", ascending=False, na_position="last")

        mat_agg = (
            cons.groupby("material_desc", as_index=False)
            .agg(
                material_code=("material_code", "first"),
                uom=("uom", "first"),
                total_qty=("usage_qty", "sum"),
                n_months=("month_raw", "nunique"),
                plant_count=("reporting_plant", "nunique"),
                plants=("reporting_plant", lambda values: sorted(values.dropna().unique().tolist())),
                storage_location_count=("storage_location", "nunique"),
                currency=("currency", "first"),
            )
            .rename(columns={"material_desc": "material"})
        )
        mat_agg["avg_monthly_consumption"] = mat_agg["total_qty"] / total_distinct_months
        mat_agg["avg_annual_consumption"] = mat_agg["avg_monthly_consumption"] * 12

        last_proc = (
            proc_sorted.groupby("material_desc", as_index=False)
            .first()[
                [
                    "material_desc",
                    "procured_qty",
                    "effective_value",
                    "po_number",
                    "doc_date",
                    "vendor",
                    "order_unit",
                    "unit_price",
                ]
            ]
            .rename(
                columns={
                    "material_desc": "material",
                    "procured_qty": "last_proc_qty",
                    "effective_value": "last_proc_value",
                    "po_number": "last_po_number",
                    "doc_date": "last_proc_date",
                    "vendor": "last_vendor",
                    "order_unit": "last_proc_unit",
                    "unit_price": "last_unit_price",
                }
            )
        ) if not proc.empty else pd.DataFrame(
            columns=[
                "material",
                "last_proc_qty",
                "last_proc_value",
                "last_po_number",
                "last_proc_date",
                "last_vendor",
                "last_proc_unit",
                "last_unit_price",
            ]
        )
        vendor_list = (
            proc.groupby("material_desc")["vendor"]
            .apply(lambda values: sorted([value for value in values.dropna().unique().tolist() if value]))
            .reset_index()
            .rename(columns={"material_desc": "material", "vendor": "vendors"})
        ) if not proc.empty else pd.DataFrame(columns=["material", "vendors"])
        proc_totals = (
            proc.groupby("material_desc", as_index=False)
            .agg(
                total_value=("effective_value", "sum"),
                currency_proc=("currency", "first"),
            )
            .rename(columns={"material_desc": "material"})
        ) if not proc.empty else pd.DataFrame(columns=["material", "total_value", "currency_proc"])

        merged = mat_agg.merge(last_proc, on="material", how="left")
        merged = merged.merge(vendor_list, on="material", how="left")
        merged = merged.merge(proc_totals, on="material", how="left")
        merged["total_value"] = merged["total_value"].fillna(0.0)
        merged["currency"] = merged["currency_proc"].fillna(merged["currency"])
        merged = merged.sort_values("total_value", ascending=False)

        materials: list[dict] = []
        for _, row in merged.iterrows():
            materials.append(
                {
                    "material": row["material"],
                    "material_code": _normalize_text(row.get("material_code")),
                    "uom": _normalize_text(row.get("uom")),
                    "avg_monthly_consumption": round(_safe_float(row.get("avg_monthly_consumption")), 2),
                    "avg_annual_consumption": round(_safe_float(row.get("avg_annual_consumption")), 2),
                    "total_qty": round(_safe_float(row.get("total_qty")), 2),
                    "total_value": round(_safe_float(row.get("total_value")), 2),
                    "n_months": int(row.get("n_months", 0)),
                    "currency": _clean_currency(row.get("currency")),
                    "last_proc_qty": round(_safe_float(row.get("last_proc_qty")), 2),
                    "last_proc_value": round(_safe_float(row.get("last_proc_value")), 2),
                    "last_po_number": _normalize_text(row.get("last_po_number")),
                    "last_proc_date": row["last_proc_date"].strftime("%d %b %Y") if pd.notna(row.get("last_proc_date")) else "",
                    "last_vendor": _normalize_text(row.get("last_vendor")),
                    "last_proc_unit": _normalize_text(row.get("last_proc_unit")),
                    "last_unit_price": round(_safe_float(row.get("last_unit_price")), 2),
                    "vendors": row.get("vendors", []) if isinstance(row.get("vendors"), list) else [],
                    "vendor_count": len(row.get("vendors", []) if isinstance(row.get("vendors"), list) else []),
                    "plants": row.get("plants", []) if isinstance(row.get("plants"), list) else [],
                    "plant_count": int(row.get("plant_count", 0)),
                    "storage_location_count": int(row.get("storage_location_count", 0)),
                }
            )
        return materials

    def _build_materials_table(self, cons: pd.DataFrame, proc: pd.DataFrame) -> list[dict]:
        if cons.empty:
            return []

        total_distinct_months = cons[["year", "month"]].drop_duplicates().shape[0] or 1
        proc_sorted = proc.sort_values("doc_date", ascending=False, na_position="last")

        mat_agg = (
            cons.groupby("material_desc", as_index=False)
            .agg(
                material_code=("material_code", "first"),
                uom=("uom", "first"),
                total_qty=("usage_qty", "sum"),
                n_months=("month_raw", "nunique"),
                plant_count=("reporting_plant", "nunique"),
                plants=("reporting_plant", lambda values: sorted(values.dropna().unique().tolist())),
                storage_location_count=("storage_location", "nunique"),
                currency=("currency", "first"),
            )
            .rename(columns={"material_desc": "material"})
        )
        mat_agg["avg_monthly_consumption"] = mat_agg["total_qty"] / total_distinct_months
        mat_agg["avg_annual_consumption"] = mat_agg["avg_monthly_consumption"] * 12

        last_proc = (
            proc_sorted.groupby("material_desc", as_index=False)
            .first()[
                [
                    "material_desc",
                    "procured_qty",
                    "effective_value",
                    "po_number",
                    "doc_date",
                    "vendor",
                    "order_unit",
                    "unit_price",
                ]
            ]
            .rename(
                columns={
                    "material_desc": "material",
                    "procured_qty": "last_proc_qty",
                    "effective_value": "last_proc_value",
                    "po_number": "last_po_number",
                    "doc_date": "last_proc_date",
                    "vendor": "last_vendor",
                    "order_unit": "last_proc_unit",
                    "unit_price": "last_unit_price",
                }
            )
        )
        vendor_list = (
            proc.groupby("material_desc")["vendor"]
            .apply(lambda values: sorted([value for value in values.dropna().unique().tolist() if value]))
            .reset_index()
            .rename(columns={"material_desc": "material", "vendor": "vendors"})
        )
        proc_totals = (
            proc.groupby("material_desc", as_index=False)
            .agg(
                total_value=("effective_value", "sum"),
                currency_proc=("currency", "first"),
            )
            .rename(columns={"material_desc": "material"})
        )

        merged = mat_agg.merge(last_proc, on="material", how="left")
        merged = merged.merge(vendor_list, on="material", how="left")
        merged = merged.merge(proc_totals, on="material", how="left")
        merged["total_value"] = merged["total_value"].fillna(0.0)
        merged["currency"] = merged["currency_proc"].fillna(merged["currency"])
        merged = merged.sort_values("total_value", ascending=False)
        cons_groups = {material: frame for material, frame in cons.groupby("material_desc")}
        proc_groups = {material: frame for material, frame in proc.groupby("material_desc")}

        materials: list[dict] = []
        for _, row in merged.iterrows():
            material_name = row["material"]
            cons_mat = cons_groups.get(material_name, cons.iloc[0:0])
            proc_mat = proc_groups.get(material_name, proc.iloc[0:0])
            monthly_preview = _monthly_series(cons_mat)
            forecast_total = _quick_forecast_total(monthly_preview)
            cost_variance = _build_cost_variance(proc_mat)
            dominant_swing = max(
                cost_variance["significant_events"],
                key=lambda event: abs(event["pct_change"]),
                default=None,
            )
            price_swing_pct = abs(_safe_float(dominant_swing["pct_change"])) if dominant_swing else 0.0
            price_swing_direction = ""
            if dominant_swing:
                if _safe_float(dominant_swing["pct_change"]) > 0:
                    price_swing_direction = "up"
                elif _safe_float(dominant_swing["pct_change"]) < 0:
                    price_swing_direction = "down"

            materials.append(
                {
                    "material": material_name,
                    "material_code": _normalize_text(row.get("material_code")),
                    "uom": _normalize_text(row.get("uom")),
                    "avg_monthly_consumption": round(_safe_float(row.get("avg_monthly_consumption")), 2),
                    "avg_annual_consumption": round(_safe_float(row.get("avg_annual_consumption")), 2),
                    "total_qty": round(_safe_float(row.get("total_qty")), 2),
                    "total_value": round(_safe_float(row.get("total_value")), 2),
                    "n_months": int(row.get("n_months", 0)),
                    "currency": _clean_currency(row.get("currency")),
                    "last_proc_qty": round(_safe_float(row.get("last_proc_qty")), 2),
                    "last_proc_value": round(_safe_float(row.get("last_proc_value")), 2),
                    "last_po_number": _normalize_text(row.get("last_po_number")),
                    "last_proc_date": row["last_proc_date"].strftime("%d %b %Y") if pd.notna(row.get("last_proc_date")) else "",
                    "last_vendor": _normalize_text(row.get("last_vendor")),
                    "last_proc_unit": _normalize_text(row.get("last_proc_unit")),
                    "last_unit_price": round(_safe_float(row.get("last_unit_price")), 2),
                    "vendors": row.get("vendors", []) if isinstance(row.get("vendors"), list) else [],
                    "vendor_count": len(row.get("vendors", []) if isinstance(row.get("vendors"), list) else []),
                    "plants": row.get("plants", []) if isinstance(row.get("plants"), list) else [],
                    "plant_count": int(row.get("plant_count", 0)),
                    "storage_location_count": int(row.get("storage_location_count", 0)),
                    "forecast_6m_qty": round(_safe_float(forecast_total), 2),
                    "forecast_method": "Seasonal median preview",
                    "avg_vendor_score": 0.0,
                    "max_price_swing_pct": round(price_swing_pct, 2),
                    "max_price_swing_direction": price_swing_direction,
                    "significant_price_swings": int(cost_variance["significant_event_count"]),
                }
            )
        return materials

    def get_materials_table(self, plant: str | None = None) -> list[dict]:
        if self._db_seed_rows_available():
            normalized_plant = _normalize_plant(plant)
            cache_key = self._cache_key(normalized_plant)
            if cache_key not in self._materials_cache:
                self._materials_cache[cache_key] = self._build_db_materials_table(normalized_plant)
            return self._materials_cache[cache_key]
        self._ensure_loaded()
        normalized_plant = _normalize_plant(plant)
        cache_key = self._cache_key(normalized_plant)
        if cache_key in self._materials_cache:
            return self._materials_cache[cache_key]

        cons = self._filtered_consumption(normalized_plant)
        proc = self._filtered_procurement(normalized_plant)
        materials = self._build_materials_summary_table(cons, proc)
        self._materials_cache[cache_key] = materials
        return materials

    def get_top_consumers(self, plant: str | None = None, limit: int = 10) -> list[dict]:
        overview = self.get_overview_payload(plant=plant)
        return overview["top_consumers"][:limit]

    def get_kpi_summary(self, plant: str | None = None) -> dict:
        if self._db_seed_rows_available():
            return self.get_overview_payload(plant=plant)["kpis"]
        return self.get_overview_payload(plant=plant)["kpis"]

    def get_overview_payload(self, plant: str | None = None) -> dict:
        if self._db_seed_rows_available():
            normalized_plant = _normalize_plant(plant)
            cache_key = self._cache_key(normalized_plant)
            if cache_key in self._overview_cache:
                return self._overview_cache[cache_key]
            payload = {
                "kpis": self._build_db_kpis(normalized_plant),
                "top_consumers": self._build_db_top_consumers(normalized_plant),
                "filters": self._get_db_filter_options(),
            }
            self._overview_cache[cache_key] = payload
            return payload
        self._ensure_loaded()
        normalized_plant = _normalize_plant(plant)
        cache_key = self._cache_key(normalized_plant)
        if cache_key in self._overview_cache:
            return self._overview_cache[cache_key]

        cons = self._filtered_consumption(normalized_plant)
        proc = self._filtered_procurement(normalized_plant)
        payload = {
            "kpis": self._build_kpis(cons, proc, normalized_plant),
            "top_consumers": self._build_top_consumers(cons, proc),
            "filters": self.get_filter_options(),
        }
        self._overview_cache[cache_key] = payload
        return payload

    def get_dashboard_payload(self, plant: str | None = None) -> dict:
        self._ensure_loaded()
        normalized_plant = _normalize_plant(plant)
        cache_key = self._cache_key(normalized_plant)
        if cache_key in self._dashboard_cache:
            return self._dashboard_cache[cache_key]

        cons = self._filtered_consumption(normalized_plant)
        proc = self._filtered_procurement(normalized_plant)
        materials = self._build_materials_table(cons, proc)

        kpis = self._build_kpis(cons, proc, normalized_plant)

        payload = {
            "kpis": kpis,
            "materials": materials,
            "top_consumers": self._build_top_consumers(cons, proc),
            "filters": self.get_filter_options(),
        }
        self._dashboard_cache[cache_key] = payload
        return payload

    def get_top_consumers_from_materials(self, materials: list[dict], limit: int = 10) -> list[dict]:
        return [
            {
                "material": material["material"],
                "material_code": material["material_code"],
                "total_value": material["total_value"],
                "total_qty": material["total_qty"],
                "uom": material["uom"],
            }
            for material in materials[:limit]
        ]

    def get_monthly_consumption(self, material: str, plant: str | None = None) -> dict:
        return dict(self._get_material_context(material, plant)["consumption"])

    def get_monthly_consumption_detail(self, material: str, plant: str | None = None) -> list[dict]:
        cache_key = self._material_cache_key(material, plant)
        if cache_key in self._consumption_detail_cache:
            return self._consumption_detail_cache[cache_key]

        context = self._get_material_context(material, plant)
        df = context["cons_df"]
        qty_rows = _actual_monthly_rows(df).rename(columns={"value": "consumption_value"})
        proc_monthly = context["proc_monthly"].rename(columns={"value": "procurement_value"})
        monthly = qty_rows.merge(
            proc_monthly[["year", "month", "procurement_value"]],
            on=["year", "month"],
            how="outer",
        ).sort_values(["year", "month"], ascending=[False, False])
        monthly["qty"] = monthly["qty"].fillna(0.0)
        monthly["procurement_value"] = monthly["procurement_value"].fillna(0.0)
        monthly["uom"] = monthly["uom"].fillna(df["uom"].iloc[0] if not df.empty else "")
        detail = [
            {
                "period": _format_period_label(int(row["year"]), int(row["month"])),
                "year": int(row["year"]),
                "month": int(row["month"]),
                "quantity": round(_safe_float(row["qty"]), 2),
                "value": round(_safe_float(row["procurement_value"]), 2),
                "uom": _normalize_text(row["uom"]),
            }
            for _, row in monthly.iterrows()
        ]
        self._consumption_detail_cache[cache_key] = detail
        return detail

    def get_procurement_history(self, material: str, plant: str | None = None) -> list[dict]:
        cache_key = self._material_cache_key(material, plant)
        if cache_key in self._procurement_history_cache:
            return self._procurement_history_cache[cache_key]

        df = self._get_material_context(material, plant)["proc_df"]
        if df.empty:
            return []
        df = df.sort_values("doc_date", ascending=False, na_position="last")
        history = [
            {
                "po_number": _normalize_text(row.get("po_number")),
                "doc_date": row["doc_date"].strftime("%d %b %Y") if pd.notna(row.get("doc_date")) else "",
                "vendor": _normalize_text(row.get("vendor")),
                "order_qty": round(_safe_float(row.get("order_qty")), 2),
                "still_to_be_delivered_qty": round(_safe_float(row.get("still_to_be_delivered_qty")), 2),
                "procured_qty": round(_safe_float(row.get("procured_qty")), 2),
                "order_unit": _normalize_text(row.get("order_unit")),
                "unit_price": round(_safe_float(row.get("unit_price")), 2),
                "price_unit": round(_safe_float(row.get("price_unit"), 1.0), 2),
                "effective_value": round(_safe_float(row.get("effective_value")), 2),
                "plant": _normalize_text(row.get("reporting_plant") or row.get("plant")),
                "currency": _clean_currency(row.get("currency")),
            }
            for _, row in df.iterrows()
        ]
        self._procurement_history_cache[cache_key] = history
        return history

    def get_material_analytics_overview(self, material: str, plant: str | None = None) -> dict:
        cache_key = self._material_cache_key(material, plant)
        if cache_key in self._analytics_overview_cache:
            return self._analytics_overview_cache[cache_key]

        context = self._get_material_context(material, plant)
        payload = {
            "material": context["material"],
            "plant": context["plant"],
            "summary": dict(context["summary"]),
            "consumption": dict(context["consumption"]),
            "financial_years": _build_financial_year_summary(
                context["monthly"],
                context["proc_monthly"],
            ),
        }
        self._analytics_overview_cache[cache_key] = payload
        return payload

    def get_material_analytics_section(
        self,
        material: str,
        section: str,
        plant: str | None = None,
    ) -> dict:
        normalized_section = (section or "").strip().lower()
        cache_key = (*self._material_cache_key(material, plant), normalized_section)
        if cache_key in self._analytics_section_cache:
            return self._analytics_section_cache[cache_key]

        context = self._get_material_context(material, plant)
        if normalized_section == "forecast":
            forecast_key = _forecast_cache_entry_key(material, plant)
            forecast = self._get_persistent_forecast_cache().get(forecast_key)
            if forecast is None:
                forecast = _build_forecast(context["monthly"])
                try:
                    self._store_persistent_forecast(material, plant, forecast)
                except Exception:
                    logger.warning(
                        "Failed to persist forecast cache for material=%s plant=%s",
                        material,
                        plant,
                        exc_info=True,
                    )
            payload = {"forecast": forecast}
        elif normalized_section == "yoy":
            payload = {"yoy": _build_yoy(context["monthly"], context["proc_monthly"])}
        elif normalized_section == "vendors":
            payload = {"vendor_scores": _build_vendor_scores(context["proc_df"])}
        elif normalized_section == "storage":
            payload = {"storage_breakdown": _build_storage_breakdown(context["cons_df"], context["plant"] or None)}
        elif normalized_section == "variance":
            payload = {"cost_variance": _build_cost_variance(context["proc_df"])}
        else:
            raise ValueError(f"Unsupported analytics section: {section}")

        self._analytics_section_cache[cache_key] = payload
        return payload

    def get_material_analytics(self, material: str, plant: str | None = None) -> dict:
        payload = dict(self.get_material_analytics_overview(material, plant=plant))
        for section in ("forecast", "yoy", "vendors", "storage", "variance"):
            payload.update(self.get_material_analytics_section(material, section, plant=plant))
        return payload

    def build_export_workbook(self, plant: str | None = None, query: str = "") -> tuple[BytesIO, str]:
        """Build and return the professional MI report workbook as a BytesIO stream.

        Delegates to ``generate_mi_report.generate_workbook_from_dataframes()``
        which produces the same multi-sheet, chart-rich workbook as the
        standalone ``generate_mi_report.py`` CLI script.
        """
        import sys as _sys

        # ── Dynamically import the standalone report-generator module ─────────
        # _BASE_DIR already points to the workspace root where the script lives.
        _report_dir = _BASE_DIR
        if _report_dir not in _sys.path:
            _sys.path.insert(0, _report_dir)
        try:
            import generate_mi_report as _rpt  # type: ignore
            _gen = _rpt.generate_workbook_from_dataframes
        except Exception as _exc:
            logger.warning("Could not import generate_mi_report (%s); falling back to basic export", _exc)
            return self._build_basic_export_workbook(plant=plant, query=query)

        # ── Get filtered consumption and procurement DataFrames ───────────────
        cons = self._filtered_consumption(plant)
        proc = self._filtered_procurement(plant)

        # Optionally narrow to the search-query subset
        if query:
            payload = self.get_dashboard_payload(plant=plant)
            materials = _filter_material_rows(payload["materials"], query)
            material_names = {m["material"] for m in materials}
            if material_names:
                cons = cons[cons["material_desc"].isin(material_names)].copy()
                proc = proc[proc["material_desc"].isin(material_names)].copy()
            else:
                cons = cons.iloc[0:0].copy()
                proc = proc.iloc[0:0].copy()

        # ── Load master data (density / physical state) from DB ───────────────
        master_df = None
        try:
            from app.models.inventory.material_master import MaterialMaster
            from app.extensions import db as _db
            import json as _json

            rows_db = _db.session.query(
                MaterialMaster.material,
                MaterialMaster.short_text,
                MaterialMaster.physical_state,
                MaterialMaster.extra_data,
            ).all()

            records = []
            for mat, stext, phys, extra_raw in rows_db:
                ed: dict = {}
                if extra_raw:
                    try:
                        ed = _json.loads(extra_raw) if isinstance(extra_raw, str) else (extra_raw or {})
                    except Exception:
                        pass
                density_raw = ed.get("Density") or ed.get("density")
                try:
                    density = float(density_raw) if density_raw not in (None, "", "nan", "NaN") else None
                except Exception:
                    density = None
                records.append({
                    "material_code":  str(mat).strip(),
                    "short_text":     str(stext or "").strip(),
                    "physical_state": str(phys or "").strip() or None,
                    "density":        density,
                })
            master_df = pd.DataFrame(records)
            logger.info("Loaded %d master records for MI report export", len(master_df))
        except Exception as _exc:
            logger.warning("Could not load master data for export (%s); skipping density conversion", _exc)

        # ── Generate the professional workbook ────────────────────────────────
        try:
            stream = _gen(cons, proc, master_df=master_df, plant_filter=plant)
        except Exception as _exc:
            logger.error("generate_workbook_from_dataframes failed (%s); falling back to basic export", _exc, exc_info=True)
            return self._build_basic_export_workbook(plant=plant, query=query)

        plant_suffix = (_normalize_plant(plant) or "all-plants").lower()
        filename = f"mi-report-{plant_suffix}.xlsx"
        return stream, filename

    def _build_basic_export_workbook(self, plant: str | None = None, query: str = "") -> tuple[BytesIO, str]:
        """Fallback: the original simple 3-sheet export workbook."""
        payload = self.get_dashboard_payload(plant=plant)
        materials = _filter_material_rows(payload["materials"], query)
        material_names = {material["material"] for material in materials}

        cons = self._filtered_consumption(plant)
        proc = self._filtered_procurement(plant)
        if material_names:
            cons = cons[cons["material_desc"].isin(material_names)].copy()
            proc = proc[proc["material_desc"].isin(material_names)].copy()
        else:
            cons = cons.iloc[0:0].copy()
            proc = proc.iloc[0:0].copy()

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Materials Summary"

        summary_rows = [
            ["Plant Filter", _normalize_plant(plant) or "All Plants"],
            ["Search Filter", _normalize_text(query) or "None"],
            ["Value Basis", "Procurement effective value from SAP procurement export"],
            ["Records Exported", len(materials)],
            ["Consumption Rows", len(cons)],
            ["Procurement Rows", len(proc)],
            [],
        ]
        for row in summary_rows:
            summary_sheet.append(row)

        material_headers = [
            "Material", "Material Code", "UoM",
            "Avg Monthly Consumption", "Avg Annual Consumption",
            "Total Quantity", "Total Procurement Value",
            "Last Procurement Qty", "Last Procurement Value",
            "Last Vendor", "Last PO Number", "Last Procurement Date",
            "Vendor Score", "Forecast 6M Qty",
            "Max Price Swing %", "Price Swing Direction", "Plants",
        ]
        summary_sheet.append(material_headers)
        for material in materials:
            summary_sheet.append([
                material["material"], material["material_code"], material["uom"],
                material["avg_monthly_consumption"], material["avg_annual_consumption"],
                material["total_qty"], material["total_value"],
                material["last_proc_qty"], material["last_proc_value"],
                material["last_vendor"], material["last_po_number"], material["last_proc_date"],
                material["avg_vendor_score"], material["forecast_6m_qty"],
                material["max_price_swing_pct"], material.get("max_price_swing_direction", ""),
                ", ".join(material.get("plants", [])),
            ])

        cons_sheet = workbook.create_sheet("Consumption Detail")
        cons_sheet.append(["Plant", "Source Plant", "Material Code", "Material",
                            "Storage Location", "Financial Year", "Month", "Usage Quantity", "UoM"])
        for _, row in cons.sort_values(["reporting_plant", "plant", "material_desc", "year", "month"]).iterrows():
            cons_sheet.append([
                _normalize_text(row.get("reporting_plant")),
                _normalize_text(row.get("plant")),
                _normalize_text(row.get("material_code")),
                _normalize_text(row.get("material_desc")),
                _normalize_text(row.get("storage_location")),
                _normalize_text(row.get("financial_year")),
                _normalize_text(row.get("month_raw")),
                round(_safe_float(row.get("usage_qty")), 2),
                _normalize_text(row.get("uom")),
            ])

        proc_sheet = workbook.create_sheet("Procurement Detail")
        proc_sheet.append(["Plant", "Source Plant", "Material Code", "Material",
                            "PO Number", "Document Date", "Financial Year", "Vendor",
                            "Ordered Quantity", "Still to be Delivered Qty", "Procured Quantity", "Order Unit", "Price Unit",
                            "Unit Price", "Effective Value", "Currency"])
        for _, row in proc.sort_values(["reporting_plant", "plant", "material_desc", "doc_date"],
                                        ascending=[True, True, True, False]).iterrows():
            proc_sheet.append([
                _normalize_text(row.get("reporting_plant")),
                _normalize_text(row.get("plant")),
                _normalize_text(row.get("material_code")),
                _normalize_text(row.get("material_desc")),
                _normalize_text(row.get("po_number")),
                row["doc_date"].strftime("%d %b %Y") if pd.notna(row.get("doc_date")) else "",
                _normalize_text(row.get("financial_year")),
                _normalize_text(row.get("vendor")),
                round(_safe_float(row.get("order_qty")), 2),
                round(_safe_float(row.get("still_to_be_delivered_qty")), 2),
                round(_safe_float(row.get("procured_qty")), 2),
                _normalize_text(row.get("order_unit")),
                round(_safe_float(row.get("price_unit"), 1.0), 2),
                round(_safe_float(row.get("unit_price")), 2),
                round(_safe_float(row.get("effective_value")), 2),
                _clean_currency(row.get("currency")),
            ])

        self._format_export_sheet(summary_sheet, header_row=8)
        self._format_export_sheet(cons_sheet, header_row=1)
        self._format_export_sheet(proc_sheet, header_row=1)

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        plant_suffix = (_normalize_plant(plant) or "all-plants").lower()
        filename = f"inventory-intelligence-{plant_suffix}.xlsx"
        return stream, filename

    def _format_export_sheet(self, sheet, header_row: int) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"

        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 32)


_store = _DataStore()


def get_data_store() -> _DataStore:
    return _store


# ── Master Data helpers ────────────────────────────────────────────────────────
# Convenience accessors so any analysis module can import from here and enrich
# its output with material master attributes without a direct DB dependency.

def get_master_record(material: str) -> dict | None:
    """Return the MaterialMaster dict for *material*, or None.

    The *material* key matches the SAP ``material`` field used in
    inventory_records and procurement data.

    Usage in analysis code::

        from app.core.services.inventory_intelligence import get_master_record

        record = get_master_record("90001025")
        if record:
            group = record["group"]
            centralization = record["centralization"]
    """
    from app.core.services.master_data import get_master_record as _get
    return _get(material)


def get_all_master_data() -> list[dict]:
    """Return all MaterialMaster rows as a list of dicts, sorted by material.

    Usage::

        from app.core.services.inventory_intelligence import get_all_master_data

        master_map = {r["material"]: r for r in get_all_master_data()}
        # Then join: row["group"] = master_map.get(row["material"], {}).get("group", "")
    """
    from app.core.services.master_data import get_all_master_data as _get_all
    return _get_all()
