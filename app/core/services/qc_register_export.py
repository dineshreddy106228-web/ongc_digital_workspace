"""Excel export of the SAP sample register.

The register page caps what it renders, so the download exists to be the
complete answer: it applies the same filters and then exports every match.

The sheet layout follows how the laboratories read the register rather than
how the data is stored.  A record carrying an SAP work center is one a bench
is actually holding, so those are split by Corporate Specification sub-group —
DFC, PC and the rest — because that is the axis a chemist works along.  The
two source-completeness sheets are the exceptions worth chasing: a
notification with no inspection lot, and an inspection lot with no
notification.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

from app.core.services.csc_utils import SPEC_SUBSET_LABELS, SPEC_SUBSET_ORDER
from app.core.services.sap_quality_control import (
    CORPORATE_SPECIFICATION_UNMATCHED_KEY,
    CORPORATE_SPECIFICATION_UNMATCHED_LABEL,
    sap_sample_register_data,
    source_completeness_label,
)

NAVY = "FF1F3864"
HEADERS = [
    "Laboratory", "Plant", "SAP as of", "Inspection lot", "Notification no",
    "Material code", "Material description", "Corporate Specification",
    "Sub-group", "Work center", "SAP status", "Usage decision",
    "SAP receipt", "Notification date", "STT days", "STT due", "Days over STT",
    "Completion date", "Pairing", "Lab follow-up",
]
# Identifiers, not quantities — a 12-digit lot left numeric renders as 8.9E+11.
IDENTIFIER_HEADERS = {
    "Plant", "Inspection lot", "Notification no", "Material code",
}


def _row(item: dict[str, Any]) -> list[Any]:
    record = item["record"]
    update = item.get("lab_update")
    return [
        item["laboratory"]["name"],
        record.plant_code,
        item["batch"].as_of_date,
        record.inspection_lot_number,
        record.notification_no,
        record.material_code,
        record.material_description,
        item["specification_no"] if item["specification_match"] else None,
        item["subgroup_label"],
        record.work_center,
        (record.official_status or "").title(),
        record.usage_decision_code,
        record.start_inspection_date,
        record.notification_start_date,
        item["stt_days"],
        item["stt_due_date"],
        item["stt_variance_days"] if item["stt_overdue"] else None,
        record.completion_date,
        source_completeness_label(record),
        (
            "Excluded from monitoring" if item["is_excluded"]
            else "Exclusion review required" if item["exclusion_requires_review"]
            else item["reconciliation_label"]
        ),
    ]


def _latest_first(item: dict[str, Any]) -> tuple:
    """Newest notification first, with undated rows last rather than first."""
    record = item["record"]
    stamp = record.notification_start_date or record.start_inspection_date
    return (stamp is None, -(stamp.toordinal() if stamp else 0), record.id)


def _sheet(workbook, title: str, subtitle: str, items: list[dict[str, Any]]):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet(title)
    thin = Side(style="thin", color="FFD9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.cell(row=1, column=1, value=subtitle).font = Font(
        name="Arial", size=10, bold=True, color=NAVY,
    )
    for column, label in enumerate(HEADERS, 1):
        cell = sheet.cell(row=2, column=column, value=label)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    identifiers = {index for index, label in enumerate(HEADERS) if label in IDENTIFIER_HEADERS}
    for item in items:
        sheet.append([
            "" if value is None
            else str(value) if index in identifiers
            else value
            for index, value in enumerate(_row(item))
        ])
    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            if isinstance(cell.value, date):
                cell.number_format = "DD-MMM-YYYY"
                cell.alignment = Alignment(horizontal="center")
            elif isinstance(cell.value, str) and cell.value.isdigit():
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="left")
    for column, label in enumerate(HEADERS, 1):
        longest = max([len(label)] + [
            len(str(sheet.cell(row=row, column=column).value or ""))
            for row in range(3, sheet.max_row + 1)
        ])
        sheet.column_dimensions[get_column_letter(column)].width = min(34, max(11, longest + 2))
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{max(sheet.max_row, 2)}"
    return sheet


def build_sample_register_workbook(
    lab_code: str = "", search: str = "", status: str = "", subgroup: str = "",
) -> tuple[BytesIO, str]:
    """Export the filtered register, uncapped, split the way the labs read it."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    register = sap_sample_register_data(lab_code, search, status, subgroup, limit=None)
    entries = register["entries"]
    if not entries:
        raise ValueError("No current SAP records match these filters, so there is nothing to export.")

    workbook = Workbook()
    workbook.remove(workbook.active)

    # A record with a work center is one a bench is holding right now.
    under_testing = [item for item in entries if item["record"].work_center]
    by_subgroup: dict[str, list[dict[str, Any]]] = {}
    for item in under_testing:
        key = item["subgroup_key"] or CORPORATE_SPECIFICATION_UNMATCHED_KEY
        by_subgroup.setdefault(key, []).append(item)

    order = [key for key in SPEC_SUBSET_ORDER if key in by_subgroup]
    order += sorted(set(by_subgroup) - set(SPEC_SUBSET_ORDER) - {CORPORATE_SPECIFICATION_UNMATCHED_KEY})
    if CORPORATE_SPECIFICATION_UNMATCHED_KEY in by_subgroup:
        order.append(CORPORATE_SPECIFICATION_UNMATCHED_KEY)

    sheet_index: list[tuple[str, str, int]] = []
    for key in order:
        items = sorted(by_subgroup[key], key=_latest_first)
        label = (
            CORPORATE_SPECIFICATION_UNMATCHED_LABEL
            if key == CORPORATE_SPECIFICATION_UNMATCHED_KEY
            else SPEC_SUBSET_LABELS.get(key, f"{key} chemicals")
        )
        # The code is the sheet name because Excel allows 31 characters and
        # "CCA · Cement and Cement Additives" is 33; the label leads the sheet.
        title = "Not on register" if key == CORPORATE_SPECIFICATION_UNMATCHED_KEY else key
        _sheet(
            workbook, title,
            f"{label} — {len(items)} sample(s) with an SAP work center, newest first",
            items,
        )
        sheet_index.append((title, f"{label} · under testing", len(items)))

    def _completeness_sheet(title: str, completeness: str, description: str, *, without_lot: bool = False):
        """One source-completeness bucket.

        ``without_lot`` narrows the notification sheet to the records that
        state no inspection lot at all — the ones with a stated lot are a
        different question, answered by the pairing column.  It never applies
        to inspection-lot-only records, which have a lot by definition.
        """
        items = [
            item for item in entries
            if item["record"].source_completeness == completeness
            and (not without_lot or not item["record"].inspection_lot_number)
        ]
        # Laboratory-wise, and newest notification first inside each laboratory.
        items.sort(key=lambda item: (item["laboratory"]["name"].casefold(), *_latest_first(item)))
        _sheet(workbook, title, f"{description} — {len(items)} sample(s), laboratory-wise, newest first", items)
        sheet_index.append((title, description, len(items)))

    _completeness_sheet(
        "Notification no lot", "notification_only",
        "Notification raised with no inspection lot stated", without_lot=True,
    )
    _completeness_sheet(
        "Inspection lot only", "inspection_lot_only",
        "Inspection lot with no notification in SAP",
    )

    summary = workbook.create_sheet("Summary", 0)
    summary["A1"] = "SAP sample register export"
    summary["A1"].font = Font(name="Arial", size=14, bold=True, color=NAVY)

    applied = [
        label for label, value in (
            (f"laboratory = {lab_code}", lab_code), (f"search = {search}", search),
            (f"SAP position = {status}", status), (f"sub-group = {subgroup}", subgroup),
        ) if value
    ]
    notes = [
        f"{register['total_matching']} current SAP record(s) exported — every match, not the register page's first {register['visible_limit'] or 'n'}.",
        "Filters applied: " + (", ".join(applied) if applied else "none — the entire register."),
        "",
        "Sub-group sheets hold samples carrying an SAP work center, which means a bench is holding them now.",
        f"Those {len(under_testing)} samples plus the {sum(1 for item in entries if not item['record'].work_center)} with no work center account for all {register['total_matching']} records.",
        "",
        "The 'Notification no lot' sheet is a subset of the sub-group sheets, not a further slice: every such record still carries a work center. Do not add its count to the totals above.",
    ]
    for index, line in enumerate(notes):
        cell = summary.cell(row=3 + index, column=1, value=line)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    header_row = 3 + len(notes) + 1
    for column, label in enumerate(("Sheet", "Contents", "Records"), 1):
        cell = summary.cell(row=header_row, column=column, value=label)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        from openpyxl.styles import PatternFill
        cell.fill = PatternFill("solid", fgColor=NAVY)
    for index, (title, description, _count) in enumerate(sheet_index):
        row = header_row + 1 + index
        summary.cell(row=row, column=1, value=title).font = Font(name="Arial", size=10)
        summary.cell(row=row, column=2, value=description).font = Font(name="Arial", size=10)
        # Counted from the sheet so the figure cannot drift from its rows.
        summary.cell(row=row, column=3, value=f"=COUNTA('{title}'!A3:A100000)").font = Font(name="Arial", size=10)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 62
    summary.column_dimensions["C"].width = 11

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    scope = lab_code or "All laboratories"
    return output, f"QC SAP Sample Register {scope} {date.today():%d %b %Y}.xlsx"
