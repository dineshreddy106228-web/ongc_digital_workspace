"""Inventory Intelligence — Forecast Excel Export Builder.

Generates a professionally formatted multi-sheet Excel workbook from the
forecast data produced by the Demand Forecast module.  Adapted from the
ONGC Xylene Pilot Excel builder.

Sheets
------
1. Forecast Summary   – one row per material with 6-month point + bands
2. Forecast Detail    – monthly breakdown per material with history + forecast
3. Confidence & Quality – scoring breakdown and data-quality metrics
"""

from __future__ import annotations

import logging
from io import BytesIO

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Styling constants ──────────────────────────────────────────────────
_FN = "Arial"
_C_HDR = "1A4A6B"
_C_GRN = "1D7A55"
_C_AMB = "BA7517"
_C_RED = "A32D2D"
_C_ALT = "F7FAFA"
_TH = Side(style="thin", color="BFBFBF")
_BD = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)
_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_L = Alignment(horizontal="left", vertical="center", wrap_text=False)
_R = Alignment(horizontal="right", vertical="center")


def _fnt(bold: bool = False, sz: int = 10, col: str = "000000") -> Font:
    return Font(name=_FN, bold=bold, size=sz, color=col)


def _fl(colour: str) -> PatternFill:
    c = str(colour)[:6]
    return PatternFill("solid", fgColor=c, start_color=c)


def _score_color(score: float) -> str:
    if score >= 80:
        return _C_GRN
    if score >= 70:
        return _C_AMB
    return _C_RED


def _score_label(score: float) -> str:
    if score >= 80:
        return "High confidence — proceed"
    if score >= 70:
        return "Medium — review advised"
    return "Low — manual validation required"


def _slugify(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "-" for ch in str(value or "").strip().lower())
    compact = "-".join(part for part in cleaned.split("-") if part)
    return compact or "value"


def _title_row(ws, row: int, text: str, ncols: int, bg: str = _C_HDR, sz: int = 13, h: int = 28):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = Font(name=_FN, bold=True, size=sz, color="FFFFFF")
    c.fill = _fl(bg)
    c.alignment = _C
    ws.row_dimensions[row].height = h


def _hdr_row(ws, row: int, values: list, bg: str = _C_HDR, fc: str = "FFFFFF", h: int = 28, sc: int = 1):
    ws.row_dimensions[row].height = h
    for i, v in enumerate(values, sc):
        c = ws.cell(row, i, v)
        c.font = Font(name=_FN, bold=True, size=10, color=fc)
        c.fill = _fl(bg)
        c.alignment = _C
        c.border = _BD


# ═══════════════════════════════════════════════════════════════════════
# PUBLIC API
# ═══════════════════════════════════════════════════════════════════════

def build_forecast_workbook(
    materials_forecast: list[dict],
    plant_label: str = "All Plants",
) -> tuple[BytesIO, str]:
    """Build a multi-sheet forecast Excel workbook.

    Parameters
    ----------
    materials_forecast
        A list of dicts, each containing at least::

            {
                "material": str,
                "forecast": { ... },        # from _build_forecast()
                "consumption": { ... },     # from get_material_analytics()
                "summary": { ... },         # from get_material_analytics()
            }
    plant_label
        Human-readable plant filter description.

    Returns
    -------
    (BytesIO stream, filename)
    """
    wb = Workbook()

    # ── Sheet 1: Forecast Summary ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Forecast Summary"
    ws1.freeze_panes = "C4"

    ncols = 2 + 3 * 6 + 3  # unit + desc + 6*(point,lo,hi) + score + rating + method
    _title_row(ws1, 1, f"Inventory Intelligence — 6-Month Demand Forecast  |  {plant_label}", ncols, h=28)

    # Sub-header row for month groups
    ws1.row_dimensions[2].height = 22
    # Build header
    h_vals = ["Material", "UoM"]
    forecast_periods_all: list[str] = []
    if materials_forecast:
        f0 = materials_forecast[0].get("forecast", {})
        forecast_periods_all = f0.get("forecast_labels", [])
    for period in forecast_periods_all:
        h_vals += [f"{period} Point", f"{period} Lower", f"{period} Upper"]
    h_vals += ["Conf Score", "Rating", "Method"]
    _hdr_row(ws1, 3, h_vals, h=30)

    for ri, mf in enumerate(materials_forecast):
        fc = mf.get("forecast", {})
        er = ri + 4
        ws1.cell(er, 1, mf.get("material", "")).font = _fnt(bold=True, sz=10)
        ws1.cell(er, 1).border = _BD
        ws1.cell(er, 1).alignment = _L
        ws1.cell(er, 2, fc.get("history_quantities", [""])[0] if False else mf.get("summary", {}).get("uom", "")).font = _fnt(sz=9)
        ws1.cell(er, 2).border = _BD

        p50_vals = fc.get("p50_quantities", [])
        lo_vals = fc.get("p10_quantities", [])
        hi_vals = fc.get("p90_quantities", [])
        for fi in range(min(6, len(p50_vals))):
            base_col = 3 + fi * 3
            for ci_off, val in enumerate([p50_vals[fi], lo_vals[fi] if fi < len(lo_vals) else 0, hi_vals[fi] if fi < len(hi_vals) else 0]):
                cell = ws1.cell(er, base_col + ci_off, round(val, 0) if val else 0)
                cell.number_format = "#,##0"
                cell.font = _fnt(bold=(ci_off == 0), sz=10)
                cell.alignment = _R
                cell.border = _BD
                if ci_off == 0:
                    cell.fill = _fl("EAF6EE")
                elif ci_off == 1:
                    cell.fill = _fl("EBF5FB")
                else:
                    cell.fill = _fl("FEF9E7")

        # Confidence score from walk-forward validation
        conf = fc.get("wf_conf_score") or fc.get("selected_band_coverage_pct") or 0
        conf_val = float(conf) if conf else 0.0
        sc = _score_color(conf_val)
        meta_start = 3 + 6 * 3

        cell_s = ws1.cell(er, meta_start, round(conf_val, 1))
        cell_s.font = Font(name=_FN, bold=True, size=11, color="FFFFFF")
        cell_s.fill = _fl(sc)
        cell_s.alignment = _C
        cell_s.border = _BD

        cell_r = ws1.cell(er, meta_start + 1, _score_label(conf_val))
        cell_r.font = _fnt(bold=True, sz=9)
        cell_r.fill = _fl({"1D7A55": "C7E6D5", "BA7517": "FFF3CD", "A32D2D": "FFCDD2"}.get(sc, "F5F5F5"))
        cell_r.alignment = _C
        cell_r.border = _BD

        cell_m = ws1.cell(er, meta_start + 2, fc.get("method", ""))
        cell_m.font = _fnt(sz=9)
        cell_m.alignment = _L
        cell_m.border = _BD

        ws1.row_dimensions[er].height = 22

    # Column widths
    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 10
    for i in range(3, ncols + 1):
        ws1.column_dimensions[get_column_letter(i)].width = 14

    # ── Sheet 2: Forecast Detail ──────────────────────────────────────
    ws2 = wb.create_sheet("Forecast Detail")
    row = 1
    for mf in materials_forecast:
        fc = mf.get("forecast", {})
        hist_labels = fc.get("history_labels", [])
        hist_qty = fc.get("history_quantities", [])
        fc_labels = fc.get("forecast_labels", [])
        fc_rows = fc.get("forecast_rows", [])

        # Material header
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        c = ws2.cell(row, 1, f"  {mf.get('material', '')}  |  {fc.get('method', '')}  |  {fc.get('demand_type', '')}")
        c.font = Font(name=_FN, bold=True, size=11, color="FFFFFF")
        c.fill = _fl(_C_HDR)
        c.alignment = _L
        ws2.row_dimensions[row].height = 24
        row += 1

        _hdr_row(ws2, row, ["Month", "Actual", "", "Forecast Month", "P5", "P50", "P95", "Buffer", "Range Width"], _C_HDR, h=24)
        row += 1

        n_hist = len(hist_labels)
        n_fc = len(fc_rows)
        max_rows = max(n_hist, n_fc)
        for i in range(max_rows):
            if i < n_hist:
                ws2.cell(row, 1, hist_labels[i]).font = _fnt(sz=9)
                ws2.cell(row, 1).border = _BD
                ws2.cell(row, 1).alignment = _C
                cell_v = ws2.cell(row, 2, round(hist_qty[i], 0) if i < len(hist_qty) else 0)
                cell_v.number_format = "#,##0"
                cell_v.font = _fnt(sz=9)
                cell_v.border = _BD
                cell_v.alignment = _R
            ws2.cell(row, 3, "").border = _BD

            if i < n_fc:
                fr = fc_rows[i]
                ws2.cell(row, 4, fr.get("period", "")).font = _fnt(sz=9)
                ws2.cell(row, 4).border = _BD
                ws2.cell(row, 4).alignment = _C
                for ci, key in enumerate(["p5_qty", "p50_qty", "p95_qty", "buffer_qty", "range_width_qty"], 5):
                    cell_f = ws2.cell(row, ci, round(fr.get(key, 0), 0))
                    cell_f.number_format = "#,##0"
                    cell_f.font = _fnt(bold=(key == "p50_qty"), sz=9)
                    cell_f.border = _BD
                    cell_f.alignment = _R
                    if key == "p50_qty":
                        cell_f.fill = _fl("EAF6EE")
            ws2.row_dimensions[row].height = 16
            row += 1
        row += 2

    for c_letter, w in zip("ABCDEFGHI", [12, 14, 3, 12, 14, 14, 14, 14, 14]):
        ws2.column_dimensions[c_letter].width = w

    # ── Sheet 3: Confidence & Quality ─────────────────────────────────
    ws3 = wb.create_sheet("Confidence & Quality")
    _title_row(ws3, 1, "Forecast Confidence & Reliability", 10, h=26)
    _hdr_row(
        ws3,
        2,
        ["Material", "Demand Type", "Active Months", "Method", "Op Percentile", "Coverage %", "Bias", "Directional %", "Conf Score", "Conf Band"],
        _C_HDR,
        h=30,
    )

    for ri, mf in enumerate(materials_forecast):
        fc = mf.get("forecast", {})
        confidence = fc.get("forecast_confidence_record") or {}
        er = ri + 3
        ws3.cell(er, 1, mf.get("material", "")).font = _fnt(bold=True, sz=9)
        ws3.cell(er, 1).border = _BD
        ws3.cell(er, 1).alignment = _L
        ws3.cell(er, 2, fc.get("demand_type", "")).font = _fnt(sz=9)
        ws3.cell(er, 2).border = _BD
        ws3.cell(er, 2).alignment = _C
        ws3.cell(er, 3, fc.get("recent_active_months", 0)).font = _fnt(sz=9)
        ws3.cell(er, 3).border = _BD
        ws3.cell(er, 3).alignment = _C
        ws3.cell(er, 4, fc.get("method", "")).font = _fnt(sz=9)
        ws3.cell(er, 4).border = _BD
        ws3.cell(er, 4).alignment = _L

        op_percentile = fc.get("selected_percentile_label", "P50")
        coverage = confidence.get("coverage_pct", "")
        bias = confidence.get("bias", "")
        directional = confidence.get("directional_accuracy_pct", "")
        score = confidence.get("confidence_score", "")
        band = confidence.get("confidence_band", "")
        for ci, val in enumerate([op_percentile, coverage, bias, directional, score, band], 5):
            cell_q = ws3.cell(er, ci, val if val != "" else "—")
            cell_q.font = _fnt(sz=10)
            cell_q.border = _BD
            cell_q.alignment = _C
            if ci == 9 and isinstance(val, (int, float)):
                sc = _score_color(float(val))
                cell_q.font = Font(name=_FN, bold=True, size=11, color="FFFFFF")
                cell_q.fill = _fl(sc)
        ws3.row_dimensions[er].height = 20

    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 38
    for i in range(5, 11):
        ws3.column_dimensions[get_column_letter(i)].width = 14

    # ── Save ──────────────────────────────────────────────────────────
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    plant_suffix = plant_label.lower().replace(" ", "-") if plant_label else "all-plants"
    filename = f"demand-forecast-{plant_suffix}.xlsx"
    return stream, filename


def build_forecast_backtest_workbook(
    material: str,
    forecast: dict,
    plant_label: str = "All Plants",
) -> tuple[BytesIO, str]:
    """Build a detailed walk-forward backtest workbook for one material."""
    wb = Workbook()

    selected_percentile = forecast.get("selected_percentile_label", "P50")
    confidence = forecast.get("forecast_confidence_record") or {}
    backtest_rows = forecast.get("walk_forward_backtest_rows", []) or []
    percentile_summary = forecast.get("percentile_backtest_summary", []) or []
    confidence_by_percentile = forecast.get("confidence_by_percentile_table", []) or []
    production_rows = forecast.get("production_forecast_rows", []) or []

    ws1 = wb.active
    ws1.title = "Walk-Forward Detail"
    _title_row(ws1, 1, f"Walk-Forward Backtest | {material} | {plant_label}", 27, h=28)
    _hdr_row(
        ws1,
        2,
        [
            "Selected Percentile",
            "Confidence Score",
            "Confidence Band",
            "Demand Type",
            "Method",
            "Backtest Start",
            "Backtest End",
            "Evaluation Points",
        ],
        _C_HDR,
        h=24,
    )
    meta_values = [
        selected_percentile,
        confidence.get("confidence_score", forecast.get("confidence_score", "—")),
        confidence.get("confidence_band", forecast.get("confidence_band", "—")),
        forecast.get("demand_type", ""),
        forecast.get("method", ""),
        confidence.get("backtest_window_start", ""),
        confidence.get("backtest_window_end", ""),
        confidence.get("evaluation_points", len(backtest_rows)),
    ]
    for col_idx, value in enumerate(meta_values, 1):
        cell = ws1.cell(3, col_idx, value)
        cell.font = _fnt(sz=10, bold=(col_idx == 1))
        cell.border = _BD
        cell.alignment = _C if col_idx != 5 else _L

    detail_headers = [
        "Backtest Month",
        "Prev Actual Consumption",
        "Actual Consumption",
        "P50 Forecast",
        "P55 Forecast",
        "P60 Forecast",
        "P65 Forecast",
        "P70 Forecast",
        "P75 Forecast",
        "P80 Forecast",
        "P85 Forecast",
        "P90 Forecast",
        "P95 Forecast",
        "Selected Percentile",
        "Selected Forecast",
        "Coverage Lower",
        "Coverage Upper",
        "Within Coverage",
        "Forecast Error",
        "Abs Error",
        "Underforecast",
        "Overforecast",
        "Buffer",
        "Adjusted Buffer",
        "Actual Direction",
        "Forecast Direction",
        "Directional Hit",
    ]
    _hdr_row(ws1, 5, detail_headers, _C_HDR, h=28)

    row_idx = 6
    for row in backtest_rows:
        values = [
            row.get("backtest_period", ""),
            row.get("prev_actual_consumption_qty", ""),
            row.get("actual_consumption_qty", ""),
            row.get("p50_forecast_qty", ""),
            row.get("p55_forecast_qty", ""),
            row.get("p60_forecast_qty", ""),
            row.get("p65_forecast_qty", ""),
            row.get("p70_forecast_qty", ""),
            row.get("p75_forecast_qty", ""),
            row.get("p80_forecast_qty", ""),
            row.get("p85_forecast_qty", ""),
            row.get("p90_forecast_qty", ""),
            row.get("p95_forecast_qty", ""),
            row.get("selected_percentile_label", ""),
            row.get("selected_forecast_qty", ""),
            row.get("coverage_lower_qty", ""),
            row.get("coverage_upper_qty", ""),
            "Yes" if row.get("within_coverage_band") else "No",
            row.get("forecast_error_qty", ""),
            row.get("abs_error_qty", ""),
            row.get("underforecast_qty", ""),
            row.get("overforecast_qty", ""),
            row.get("buffer_qty", ""),
            row.get("adjusted_buffer_qty", ""),
            row.get("actual_direction", ""),
            row.get("forecast_direction", ""),
            "Yes" if row.get("directional_hit") else "No" if row.get("directional_hit") is not None else "—",
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws1.cell(row_idx, col_idx, value)
            cell.border = _BD
            cell.alignment = _C if col_idx in (1, 14, 18, 25, 26, 27) else _R
            if col_idx == 1:
                cell.alignment = _L
            if row_idx % 2 == 0:
                cell.fill = _fl(_C_ALT)
        row_idx += 1

    widths = [16, 16, 16, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 14, 14, 14, 12, 12, 12, 12, 12, 12, 14, 12, 14, 12]
    for idx, width in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(idx)].width = width

    ws2 = wb.create_sheet("Percentile Summary")
    _title_row(ws2, 1, f"Percentile Backtest Summary | {material}", 10, h=28)
    _hdr_row(
        ws2,
        2,
        ["Percentile", "MAE", "RMSE", "Bias", "Underforecast %", "Overforecast %", "Underforecast Penalty", "Weighted Score", "Eval Points", "Selected"],
        _C_HDR,
        h=24,
    )
    for row_idx, row in enumerate(percentile_summary, 3):
        is_selected = int(row.get("percentile", -1)) == int(forecast.get("selected_percentile", 50))
        values = [
            f"P{int(row.get('percentile', 0))}",
            row.get("mae", ""),
            row.get("rmse", ""),
            row.get("bias", ""),
            row.get("underforecast_pct", ""),
            row.get("overforecast_pct", ""),
            row.get("underforecast_penalty", ""),
            row.get("weighted_score", ""),
            row.get("evaluation_points", ""),
            "Yes" if is_selected else "",
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws2.cell(row_idx, col_idx, value)
            cell.border = _BD
            cell.alignment = _C if col_idx in (1, 9, 10) else _R
            if is_selected:
                cell.fill = _fl("EAF6EE")
                cell.font = _fnt(bold=True, sz=10)
            else:
                cell.font = _fnt(sz=10)
    for idx, width in enumerate([12, 12, 12, 12, 16, 16, 20, 16, 12, 10], 1):
        ws2.column_dimensions[get_column_letter(idx)].width = width

    ws3 = wb.create_sheet("Confidence by Percentile")
    _title_row(ws3, 1, f"Forecast Confidence by Percentile | {material}", 9, h=28)
    _hdr_row(
        ws3,
        2,
        ["Percentile", "Coverage %", "Error Volatility", "Bias", "Directional %", "Confidence Score", "Confidence Band", "Eval Points", "Selected"],
        _C_HDR,
        h=24,
    )
    for row_idx, row in enumerate(confidence_by_percentile, 3):
        is_selected = int(row.get("percentile", -1)) == int(forecast.get("selected_percentile", 50))
        values = [
            f"P{int(row.get('percentile', 0))}",
            row.get("coverage_pct", ""),
            row.get("error_volatility", ""),
            row.get("bias", ""),
            row.get("directional_accuracy_pct", ""),
            row.get("confidence_score", ""),
            row.get("confidence_band", ""),
            row.get("evaluation_points", ""),
            "Yes" if is_selected else "",
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws3.cell(row_idx, col_idx, value)
            cell.border = _BD
            cell.alignment = _C if col_idx in (1, 7, 8, 9) else _R
            if col_idx == 6 and isinstance(value, (int, float)):
                cell.fill = _fl(_score_color(float(value)))
                cell.font = Font(name=_FN, bold=True, size=10, color="FFFFFF")
            elif is_selected:
                cell.fill = _fl("EAF6EE")
                cell.font = _fnt(bold=True, sz=10)
            else:
                cell.font = _fnt(sz=10)
    for idx, width in enumerate([12, 12, 16, 12, 14, 16, 18, 12, 10], 1):
        ws3.column_dimensions[get_column_letter(idx)].width = width

    ws4 = wb.create_sheet("Operational Forecast")
    _title_row(ws4, 1, f"Live Operational Forecast | {material}", 9, h=28)
    _hdr_row(
        ws4,
        2,
        ["Forecast Month", "Baseline P50", "Selected Percentile", "Selected Forecast", "Lower Bound", "Upper Bound", "Confidence Score", "Confidence Band", "Adj Buffer Total"],
        _C_HDR,
        h=24,
    )
    for row_idx, row in enumerate(production_rows, 3):
        values = [
            row.get("forecast_month", ""),
            row.get("baseline_p50", ""),
            f"P{int(row.get('selected_percentile', forecast.get('selected_percentile', 50)))}",
            row.get("selected_forecast", ""),
            row.get("lower_bound", ""),
            row.get("upper_bound", ""),
            row.get("confidence_score", ""),
            row.get("confidence_band", ""),
            forecast.get("confidence_adjusted_buffer_total_qty", ""),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws4.cell(row_idx, col_idx, value)
            cell.border = _BD
            cell.alignment = _C if col_idx in (1, 3, 7, 8) else _R
            cell.font = _fnt(sz=10, bold=(col_idx == 4))
    for idx, width in enumerate([16, 14, 14, 14, 14, 14, 16, 18, 16], 1):
        ws4.column_dimensions[get_column_letter(idx)].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    material_slug = _slugify(material or "material")
    plant_slug = _slugify(plant_label or "all-plants")
    filename = f"forecast-backtest-{material_slug}-{plant_slug}.xlsx"
    return stream, filename
