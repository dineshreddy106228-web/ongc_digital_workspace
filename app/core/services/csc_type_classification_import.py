"""Helpers for importing CSC workbook-driven workflow drafts."""

from __future__ import annotations

from dataclasses import dataclass, field
import re
from typing import Callable, Mapping

from openpyxl import load_workbook

from app.core.services.csc_utils import (
    normalize_parameter_type_label,
    normalize_test_procedure_type,
    parse_spec_number,
    sanitize_multiline_text,
)

TYPE_CLASSIFICATION_STREAM = "type_classification"
MATERIAL_HANDLING_STREAM = "material_handling"
IMPORTED_JUSTIFICATION_START = "[Imported Type Classification Notes]"
IMPORTED_JUSTIFICATION_END = "[/Imported Type Classification Notes]"
MATERIAL_HANDLING_BACKGROUND_START = "[Imported Material Handling Background]"
MATERIAL_HANDLING_BACKGROUND_END = "[/Imported Material Handling Background]"
MATERIAL_HANDLING_JUSTIFICATION_START = "[Imported Material Handling Justification]"
MATERIAL_HANDLING_JUSTIFICATION_END = "[/Imported Material Handling Justification]"

_HEADER_ALIASES = {
    "specnumber": "spec_number",
    "specno": "spec_number",
    "chemicalname": "chemical_name",
    "materialcode": "material_code",
    "sno": "parameter_serial",
    "serialno": "parameter_serial",
    "parametersno": "parameter_serial",
    "parameterserialno": "parameter_serial",
    "parametername": "parameter_name",
    "requirement": "requirement",
    "currenttype": "current_type",
    "type": "target_type",
    "vitalreason1": "vital_reason_1",
    "vitalreason2": "vital_reason_2",
    "vitalreason3": "vital_reason_3",
    "vitalreason4": "vital_reason_4",
    "testproceduretype": "test_procedure_type",
    "testproceduretext": "test_procedure_text",
}
_MATERIAL_HANDLING_HEADER_ALIASES = {
    "sno": "serial_number",
    "chemicalname": "chemical_name",
    "materialcode": "material_code",
    "specificationno": "source_spec_number",
    "proposednewspecificationno": "proposed_spec_number",
    "unit": "unit",
    "officersresponsible": "officers_responsible",
    "physicalstate": "physical_state",
    "volatility": "volatility",
    "sunlightsensitivity": "sunlight_sensitivity",
    "moisturesensitivity": "moisture_sensitivity",
    "temperaturesensitivity": "temperature_sensitivity",
    "reactivity": "reactivity",
    "flammable": "flammable",
    "toxic": "toxic",
    "corrosive": "corrosive",
    "storageconditionsgeneral": "storage_conditions_general",
    "storageconditionsspecial": "storage_conditions_special",
    "containertypepacking": "container_type",
    "containercapacitypacking": "container_capacity",
    "containerdescription": "container_description",
    "primarystorageclassification": "primary_storage_classification",
    "remarksifany": "remarks",
}
_REQUIRED_HEADERS = {"spec_number", "material_code", "parameter_serial", "target_type"}
_MATERIAL_HANDLING_REQUIRED_HEADERS = {
    "chemical_name",
    "material_code",
    "source_spec_number",
    "proposed_spec_number",
}
_SPEC_SLASH_RE = re.compile(r"^\s*ONGC/([^/]+)/([^/]+)/(\d{4})\s*$", re.IGNORECASE)
_SPEC_DASH_RE = re.compile(r"^\s*ONGC-([^-]+)-([^-]+)-(\d{4})\s*$", re.IGNORECASE)
_IMPORT_BLOCK_RE = re.compile(
    rf"{re.escape(IMPORTED_JUSTIFICATION_START)}.*?{re.escape(IMPORTED_JUSTIFICATION_END)}",
    re.DOTALL,
)
_MATERIAL_HANDLING_BACKGROUND_RE = re.compile(
    rf"{re.escape(MATERIAL_HANDLING_BACKGROUND_START)}.*?{re.escape(MATERIAL_HANDLING_BACKGROUND_END)}",
    re.DOTALL,
)
_MATERIAL_HANDLING_JUSTIFICATION_RE = re.compile(
    rf"{re.escape(MATERIAL_HANDLING_JUSTIFICATION_START)}.*?{re.escape(MATERIAL_HANDLING_JUSTIFICATION_END)}",
    re.DOTALL,
)
_SEP_NORMALIZE_RE = re.compile(r"\s*([/-])\s*")


@dataclass
class TypeClassificationWorkbookRow:
    sheet_name: str
    row_number: int
    spec_number: str
    chemical_name: str
    material_code: str
    parameter_serial: int | None
    parameter_name: str
    requirement: str
    current_type: str
    target_type: str
    vital_reasons: list[str] = field(default_factory=list)
    test_procedure_type: str = ""
    test_procedure_text: str = ""


@dataclass
class TypeClassificationWorkbookSheet:
    sheet_name: str
    spec_number: str
    chemical_name: str
    material_code: str
    rows: list[TypeClassificationWorkbookRow] = field(default_factory=list)

    @property
    def match_key(self) -> tuple[str, str]:
        return (
            canonicalize_spec_number(self.spec_number),
            material_code_lookup_key(self.material_code),
        )


@dataclass
class TypeClassificationSheetIssue:
    sheet_name: str
    spec_number: str
    material_code: str
    reason: str


@dataclass
class TypeClassificationRowIssue:
    sheet_name: str
    row_number: int
    spec_number: str
    material_code: str
    parameter_serial: int | None
    parameter_name: str
    reason: str


@dataclass
class ParsedTypeClassificationWorkbook:
    workbook_name: str
    sheets: list[TypeClassificationWorkbookSheet] = field(default_factory=list)
    invalid_sheets: list[TypeClassificationSheetIssue] = field(default_factory=list)
    blank_sheets: list[str] = field(default_factory=list)

    @property
    def subset_codes(self) -> list[str]:
        codes: list[str] = []
        seen: set[str] = set()
        for sheet in self.sheets:
            subset_code, _, _ = parse_spec_number(sheet.spec_number or "")
            if subset_code and subset_code not in seen:
                seen.add(subset_code)
                codes.append(subset_code)
        return codes


@dataclass
class TypeClassificationDraftUpdate:
    parameters_matched: int = 0
    parameters_updated: int = 0
    justification_entries_created: int = 0
    justification_changed: bool = False
    justification_text: str = ""
    unmatched_parameter_rows: list[TypeClassificationRowIssue] = field(default_factory=list)


@dataclass
class TypeClassificationImportDetail:
    sheet_name: str
    spec_number: str
    material_code: str
    chemical_name: str
    parent_draft_id: int | None
    draft_id: int | None
    created_new: bool
    parameters_matched: int
    parameters_updated: int
    justification_entries_created: int


@dataclass
class TypeClassificationImportSummary:
    workbook_name: str
    subset_codes: list[str] = field(default_factory=list)
    sheets_processed: int = 0
    specs_matched: int = 0
    drafts_created: int = 0
    drafts_reused: int = 0
    parameters_matched: int = 0
    parameters_updated: int = 0
    justification_entries_created: int = 0
    unmatched_sheets: list[TypeClassificationSheetIssue] = field(default_factory=list)
    unmatched_parameter_rows: list[TypeClassificationRowIssue] = field(default_factory=list)
    details: list[TypeClassificationImportDetail] = field(default_factory=list)


@dataclass
class MaterialHandlingWorkbookRow:
    row_number: int
    source_spec_number: str
    proposed_spec_number: str
    chemical_name: str
    material_code: str
    unit: str = ""
    officers_responsible: str = ""
    physical_state: str = ""
    volatility: str = ""
    sunlight_sensitivity: str = ""
    moisture_sensitivity: str = ""
    temperature_sensitivity: str = ""
    reactivity: str = ""
    flammable: str = ""
    toxic: str = ""
    corrosive: str = ""
    storage_conditions_general: str = ""
    storage_conditions_special: str = ""
    container_type: str = ""
    container_capacity: str = ""
    container_description: str = ""
    primary_storage_classification: str = ""
    remarks: str = ""

    @property
    def subset_code(self) -> str | None:
        subset_code, _, _ = parse_spec_number(self.proposed_spec_number or "")
        return subset_code


@dataclass
class MaterialHandlingRowIssue:
    row_number: int
    source_spec_number: str
    proposed_spec_number: str
    material_code: str
    chemical_name: str
    reason: str


@dataclass
class ParsedMaterialHandlingWorkbook:
    workbook_name: str
    rows: list[MaterialHandlingWorkbookRow] = field(default_factory=list)
    invalid_rows: list[MaterialHandlingRowIssue] = field(default_factory=list)

    @property
    def subset_codes(self) -> list[str]:
        codes: list[str] = []
        seen: set[str] = set()
        for row in self.rows:
            if row.subset_code and row.subset_code not in seen:
                seen.add(row.subset_code)
                codes.append(row.subset_code)
        return codes


@dataclass
class MaterialHandlingDraftUpdate:
    row_updated: bool = False
    background_text: str = ""
    justification_text: str = ""
    staged_master_values: dict[str, str] = field(default_factory=dict)


@dataclass
class MaterialHandlingImportDetail:
    row_number: int
    source_spec_number: str
    proposed_spec_number: str
    material_code: str
    chemical_name: str
    parent_draft_id: int | None
    draft_id: int | None
    created_new: bool
    staged_field_count: int


@dataclass
class MaterialHandlingImportSummary:
    workbook_name: str
    subset_codes: list[str] = field(default_factory=list)
    rows_processed: int = 0
    specs_matched: int = 0
    drafts_created: int = 0
    drafts_reused: int = 0
    rows_updated: int = 0
    staged_field_updates: int = 0
    unmatched_rows: list[MaterialHandlingRowIssue] = field(default_factory=list)
    details: list[MaterialHandlingImportDetail] = field(default_factory=list)


def canonicalize_spec_number(raw_value: object) -> str:
    text = str(raw_value or "").strip().upper()
    if not text:
        return ""
    text = _SEP_NORMALIZE_RE.sub(r"\1", text)
    dash_match = _SPEC_DASH_RE.match(text)
    if dash_match:
        return f"ONGC/{dash_match.group(1).upper()}/{dash_match.group(2).upper()}/{dash_match.group(3)}"
    slash_match = _SPEC_SLASH_RE.match(text)
    if slash_match:
        return f"ONGC/{slash_match.group(1).upper()}/{slash_match.group(2).upper()}/{slash_match.group(3)}"
    return text.replace("\\", "/")


def normalize_material_code(raw_value: object) -> str:
    if raw_value is None:
        return ""
    if isinstance(raw_value, float) and raw_value.is_integer():
        return str(int(raw_value))
    if isinstance(raw_value, int):
        return str(raw_value)
    text = str(raw_value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def material_code_lookup_key(raw_value: object) -> str:
    text = normalize_material_code(raw_value)
    if not text:
        return ""
    stripped = text.lstrip("0")
    return stripped or "0"


def parse_type_classification_workbook(file_obj, workbook_name: str) -> ParsedTypeClassificationWorkbook:
    workbook = load_workbook(file_obj, data_only=True)
    parsed = ParsedTypeClassificationWorkbook(workbook_name=workbook_name or "classification.xlsx")

    for worksheet in workbook.worksheets:
        if _worksheet_is_blank(worksheet):
            parsed.blank_sheets.append(worksheet.title)
            continue

        header_row_number, header_map = _find_header_map(worksheet)
        if header_row_number is None or header_map is None:
            parsed.invalid_sheets.append(
                TypeClassificationSheetIssue(
                    sheet_name=worksheet.title,
                    spec_number="",
                    material_code="",
                    reason="Required workbook columns were not found in this sheet.",
                )
            )
            continue

        rows: list[TypeClassificationWorkbookRow] = []
        for row_number in range(header_row_number + 1, worksheet.max_row + 1):
            row_values = {
                key: worksheet.cell(row=row_number, column=column_index).value
                for key, column_index in header_map.items()
            }
            if _row_is_blank(row_values.values()):
                continue
            spec_number = canonicalize_spec_number(
                row_values.get("spec_number") or worksheet.title
            )
            material_code = normalize_material_code(row_values.get("material_code"))
            chemical_name = sanitize_multiline_text(
                row_values.get("chemical_name") or "",
                max_length=255,
            )
            parameter_name = sanitize_multiline_text(
                row_values.get("parameter_name") or "",
                max_length=2000,
            )
            requirement = sanitize_multiline_text(
                row_values.get("requirement") or "",
                max_length=4000,
            )
            current_type = sanitize_multiline_text(
                row_values.get("current_type") or "",
                max_length=100,
            )
            target_type = sanitize_multiline_text(
                row_values.get("target_type") or "",
                max_length=100,
            )
            reasons = [
                _normalize_reason_text(
                    sanitize_multiline_text(row_values.get(reason_key) or "", max_length=2000)
                )
                for reason_key in (
                    "vital_reason_1",
                    "vital_reason_2",
                    "vital_reason_3",
                    "vital_reason_4",
                )
            ]
            rows.append(
                TypeClassificationWorkbookRow(
                    sheet_name=worksheet.title,
                    row_number=row_number,
                    spec_number=spec_number,
                    chemical_name=chemical_name,
                    material_code=material_code,
                    parameter_serial=_coerce_parameter_serial(row_values.get("parameter_serial")),
                    parameter_name=parameter_name,
                    requirement=requirement,
                    current_type=current_type,
                    target_type=target_type,
                    vital_reasons=[reason for reason in reasons if reason],
                    test_procedure_type=sanitize_multiline_text(
                        row_values.get("test_procedure_type") or "",
                        max_length=500,
                    ),
                    test_procedure_text=sanitize_multiline_text(
                        row_values.get("test_procedure_text") or "",
                        max_length=4000,
                    ),
                )
            )

        if not rows:
            parsed.blank_sheets.append(worksheet.title)
            continue

        parsed.sheets.append(
            TypeClassificationWorkbookSheet(
                sheet_name=worksheet.title,
                spec_number=rows[0].spec_number,
                chemical_name=rows[0].chemical_name,
                material_code=rows[0].material_code,
                rows=rows,
            )
        )

    return parsed


def build_parent_lookup(parent_drafts: list[object]) -> dict[tuple[str, str], object]:
    lookup: dict[tuple[str, str], object] = {}
    for parent_draft in parent_drafts:
        key = (
            canonicalize_spec_number(getattr(parent_draft, "spec_number", "")),
            material_code_lookup_key(getattr(parent_draft, "material_code", "")),
        )
        if key[0] and key[1] and key not in lookup:
            lookup[key] = parent_draft
    return lookup


def build_parent_lookup_by_material_code(parent_drafts: list[object]) -> dict[str, object]:
    lookup: dict[str, object] = {}
    for parent_draft in parent_drafts:
        material_code = material_code_lookup_key(getattr(parent_draft, "material_code", ""))
        if material_code and material_code not in lookup:
            lookup[material_code] = parent_draft
    return lookup


def parse_material_handling_workbook(file_obj, workbook_name: str) -> ParsedMaterialHandlingWorkbook:
    workbook = load_workbook(file_obj, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row_number, header_map = _find_material_handling_header_map(worksheet)
    parsed = ParsedMaterialHandlingWorkbook(workbook_name=workbook_name or "material_handling.xlsx")
    if header_row_number is None or header_map is None:
        parsed.invalid_rows.append(
            MaterialHandlingRowIssue(
                row_number=1,
                source_spec_number="",
                proposed_spec_number="",
                material_code="",
                chemical_name="",
                reason="Required material handling workbook columns were not found.",
            )
        )
        return parsed

    for row_number in range(header_row_number + 1, worksheet.max_row + 1):
        row_values = {
            key: worksheet.cell(row=row_number, column=column_index).value
            for key, column_index in header_map.items()
        }
        if _row_is_blank(row_values.values()):
            continue
        material_code = normalize_material_code(row_values.get("material_code"))
        source_spec_number = canonicalize_spec_number(row_values.get("source_spec_number"))
        proposed_spec_number = canonicalize_spec_number(row_values.get("proposed_spec_number"))
        chemical_name = sanitize_multiline_text(row_values.get("chemical_name") or "", max_length=255)
        if not material_code:
            parsed.invalid_rows.append(
                MaterialHandlingRowIssue(
                    row_number=row_number,
                    source_spec_number=source_spec_number,
                    proposed_spec_number=proposed_spec_number,
                    material_code="",
                    chemical_name=chemical_name,
                    reason="Material Code is required for material handling import.",
                )
            )
            continue
        if not proposed_spec_number:
            parsed.invalid_rows.append(
                MaterialHandlingRowIssue(
                    row_number=row_number,
                    source_spec_number=source_spec_number,
                    proposed_spec_number="",
                    material_code=material_code,
                    chemical_name=chemical_name,
                    reason="Proposed New Specification No. is required for material handling import.",
                )
            )
            continue
        parsed.rows.append(
            MaterialHandlingWorkbookRow(
                row_number=row_number,
                source_spec_number=source_spec_number,
                proposed_spec_number=proposed_spec_number,
                chemical_name=chemical_name,
                material_code=material_code,
                unit=sanitize_multiline_text(row_values.get("unit") or "", max_length=100),
                officers_responsible=sanitize_multiline_text(row_values.get("officers_responsible") or "", max_length=1000),
                physical_state=sanitize_multiline_text(row_values.get("physical_state") or "", max_length=255),
                volatility=sanitize_multiline_text(row_values.get("volatility") or "", max_length=255),
                sunlight_sensitivity=sanitize_multiline_text(row_values.get("sunlight_sensitivity") or "", max_length=255),
                moisture_sensitivity=sanitize_multiline_text(row_values.get("moisture_sensitivity") or "", max_length=255),
                temperature_sensitivity=sanitize_multiline_text(row_values.get("temperature_sensitivity") or "", max_length=255),
                reactivity=sanitize_multiline_text(row_values.get("reactivity") or "", max_length=255),
                flammable=sanitize_multiline_text(row_values.get("flammable") or "", max_length=255),
                toxic=sanitize_multiline_text(row_values.get("toxic") or "", max_length=255),
                corrosive=sanitize_multiline_text(row_values.get("corrosive") or "", max_length=255),
                storage_conditions_general=sanitize_multiline_text(row_values.get("storage_conditions_general") or "", max_length=4000),
                storage_conditions_special=sanitize_multiline_text(row_values.get("storage_conditions_special") or "", max_length=4000),
                container_type=sanitize_multiline_text(row_values.get("container_type") or "", max_length=255),
                container_capacity=sanitize_multiline_text(row_values.get("container_capacity") or "", max_length=255),
                container_description=sanitize_multiline_text(row_values.get("container_description") or "", max_length=4000),
                primary_storage_classification=sanitize_multiline_text(row_values.get("primary_storage_classification") or "", max_length=255),
                remarks=_normalize_reason_text(
                    sanitize_multiline_text(row_values.get("remarks") or "", max_length=4000)
                ),
            )
        )
    return parsed


def import_type_classification_workbook(
    workbook: ParsedTypeClassificationWorkbook,
    parent_lookup: Mapping[tuple[str, str], object],
    open_draft_for_parent: Callable[[object], tuple[object | None, bool, str | None]],
    get_justification_text: Callable[[object], str],
    set_justification_text: Callable[[object, str], None],
) -> TypeClassificationImportSummary:
    summary = TypeClassificationImportSummary(
        workbook_name=workbook.workbook_name,
        subset_codes=list(workbook.subset_codes),
        unmatched_sheets=list(workbook.invalid_sheets),
    )

    for sheet in workbook.sheets:
        summary.sheets_processed += 1
        parent_draft = parent_lookup.get(sheet.match_key)
        if parent_draft is None:
            summary.unmatched_sheets.append(
                TypeClassificationSheetIssue(
                    sheet_name=sheet.sheet_name,
                    spec_number=sheet.spec_number,
                    material_code=sheet.material_code,
                    reason="No published parent specification matched this sheet's Spec No. and Material Code.",
                )
            )
            continue

        try:
            draft, created_new, error = open_draft_for_parent(parent_draft)
        except Exception as exc:  # pragma: no cover - defensive guard around callback behavior
            draft, created_new, error = None, False, str(exc)
        if error or draft is None:
            summary.unmatched_sheets.append(
                TypeClassificationSheetIssue(
                    sheet_name=sheet.sheet_name,
                    spec_number=sheet.spec_number,
                    material_code=sheet.material_code,
                    reason=error or "Unable to open workflow draft for this specification.",
                )
            )
            continue

        summary.specs_matched += 1
        if created_new:
            summary.drafts_created += 1
        else:
            summary.drafts_reused += 1

        draft_update = apply_sheet_to_draft(
            sheet,
            draft,
            existing_justification_text=get_justification_text(draft),
            workbook_name=workbook.workbook_name,
        )
        set_justification_text(draft, draft_update.justification_text)

        summary.parameters_matched += draft_update.parameters_matched
        summary.parameters_updated += draft_update.parameters_updated
        summary.justification_entries_created += draft_update.justification_entries_created
        summary.unmatched_parameter_rows.extend(draft_update.unmatched_parameter_rows)
        summary.details.append(
            TypeClassificationImportDetail(
                sheet_name=sheet.sheet_name,
                spec_number=sheet.spec_number,
                material_code=sheet.material_code,
                chemical_name=sheet.chemical_name,
                parent_draft_id=getattr(parent_draft, "id", None),
                draft_id=getattr(draft, "id", None),
                created_new=created_new,
                parameters_matched=draft_update.parameters_matched,
                parameters_updated=draft_update.parameters_updated,
                justification_entries_created=draft_update.justification_entries_created,
            )
        )

    return summary


def import_material_handling_workbook(
    workbook: ParsedMaterialHandlingWorkbook,
    parent_lookup_by_material_code: Mapping[str, object],
    open_draft_for_parent: Callable[[object], tuple[object | None, bool, str | None]],
    get_background_text: Callable[[object], str],
    set_background_text: Callable[[object, str], None],
    get_justification_text: Callable[[object], str],
    set_justification_text: Callable[[object, str], None],
    stage_master_values: Callable[[object, dict[str, str]], int],
) -> MaterialHandlingImportSummary:
    summary = MaterialHandlingImportSummary(
        workbook_name=workbook.workbook_name,
        subset_codes=list(workbook.subset_codes),
        unmatched_rows=list(workbook.invalid_rows),
    )

    for row in workbook.rows:
        summary.rows_processed += 1
        parent_draft = parent_lookup_by_material_code.get(material_code_lookup_key(row.material_code))
        if parent_draft is None:
            summary.unmatched_rows.append(
                MaterialHandlingRowIssue(
                    row_number=row.row_number,
                    source_spec_number=row.source_spec_number,
                    proposed_spec_number=row.proposed_spec_number,
                    material_code=row.material_code,
                    chemical_name=row.chemical_name,
                    reason="No published parent specification matched this Material Code.",
                )
            )
            continue
        try:
            draft, created_new, error = open_draft_for_parent(parent_draft)
        except Exception as exc:  # pragma: no cover
            draft, created_new, error = None, False, str(exc)
        if error or draft is None:
            summary.unmatched_rows.append(
                MaterialHandlingRowIssue(
                    row_number=row.row_number,
                    source_spec_number=row.source_spec_number,
                    proposed_spec_number=row.proposed_spec_number,
                    material_code=row.material_code,
                    chemical_name=row.chemical_name,
                    reason=error or "Unable to open workflow draft for this specification.",
                )
            )
            continue

        summary.specs_matched += 1
        if created_new:
            summary.drafts_created += 1
        else:
            summary.drafts_reused += 1

        update = apply_material_handling_row_to_draft(
            row,
            draft,
            existing_background_text=get_background_text(draft),
            existing_justification_text=get_justification_text(draft),
        )
        set_background_text(draft, update.background_text)
        set_justification_text(draft, update.justification_text)
        staged_count = stage_master_values(draft, update.staged_master_values)

        summary.rows_updated += 1
        summary.staged_field_updates += staged_count
        summary.details.append(
            MaterialHandlingImportDetail(
                row_number=row.row_number,
                source_spec_number=row.source_spec_number,
                proposed_spec_number=row.proposed_spec_number,
                material_code=row.material_code,
                chemical_name=row.chemical_name,
                parent_draft_id=getattr(parent_draft, "id", None),
                draft_id=getattr(draft, "id", None),
                created_new=created_new,
                staged_field_count=staged_count,
            )
        )

    return summary


def apply_sheet_to_draft(
    sheet: TypeClassificationWorkbookSheet,
    draft: object,
    *,
    existing_justification_text: str,
    workbook_name: str,
) -> TypeClassificationDraftUpdate:
    parameter_lookup = _build_parameter_lookup(draft)
    update = TypeClassificationDraftUpdate(justification_text=existing_justification_text or "")
    justification_entries: list[str] = []

    for row in sheet.rows:
        if row.parameter_serial is None:
            update.unmatched_parameter_rows.append(
                TypeClassificationRowIssue(
                    sheet_name=row.sheet_name,
                    row_number=row.row_number,
                    spec_number=row.spec_number,
                    material_code=row.material_code,
                    parameter_serial=row.parameter_serial,
                    parameter_name=row.parameter_name,
                    reason="Parameter S.No. is blank or invalid.",
                )
            )
            continue

        parameter = parameter_lookup.get(row.parameter_serial)
        if parameter is None:
            update.unmatched_parameter_rows.append(
                TypeClassificationRowIssue(
                    sheet_name=row.sheet_name,
                    row_number=row.row_number,
                    spec_number=row.spec_number,
                    material_code=row.material_code,
                    parameter_serial=row.parameter_serial,
                    parameter_name=row.parameter_name,
                    reason="No draft parameter matched this Parameter S.No.",
                )
            )
            continue

        update.parameters_matched += 1
        if _apply_row_updates_to_parameter(parameter, row):
            update.parameters_updated += 1

        justification_entry = build_parameter_justification_entry(row)
        if justification_entry:
            justification_entries.append(justification_entry)
            update.justification_entries_created += 1

    justification_text = merge_imported_justification_block(
        existing_justification_text,
        justification_entries,
        workbook_name=workbook_name,
        sheet_name=sheet.sheet_name,
    )
    update.justification_changed = justification_text != (existing_justification_text or "")
    update.justification_text = justification_text
    return update


def apply_material_handling_row_to_draft(
    row: MaterialHandlingWorkbookRow,
    draft: object,
    *,
    existing_background_text: str,
    existing_justification_text: str,
) -> MaterialHandlingDraftUpdate:
    update = MaterialHandlingDraftUpdate(row_updated=False)

    if row.chemical_name and sanitize_multiline_text(getattr(draft, "chemical_name", "") or "", max_length=255) != row.chemical_name:
        setattr(draft, "chemical_name", row.chemical_name)
        update.row_updated = True
    if row.proposed_spec_number and canonicalize_spec_number(getattr(draft, "spec_number", "")) != row.proposed_spec_number:
        setattr(draft, "spec_number", row.proposed_spec_number)
        update.row_updated = True

    update.background_text = merge_material_handling_section_block(
        existing_background_text,
        build_material_handling_background_text(row),
        start_marker=MATERIAL_HANDLING_BACKGROUND_START,
        end_marker=MATERIAL_HANDLING_BACKGROUND_END,
    )
    update.justification_text = merge_material_handling_section_block(
        existing_justification_text,
        build_material_handling_justification_text(row),
        start_marker=MATERIAL_HANDLING_JUSTIFICATION_START,
        end_marker=MATERIAL_HANDLING_JUSTIFICATION_END,
    )
    update.staged_master_values = build_material_handling_master_payload(row)
    return update


def build_parameter_justification_entry(row: TypeClassificationWorkbookRow) -> str:
    reasons = [reason for reason in row.vital_reasons if reason]
    if not reasons:
        return ""

    parameter_label = row.parameter_name or f"Parameter {row.parameter_serial or ''}".strip()
    lines = [
        f"Parameter {row.parameter_serial} - {parameter_label}",
        f"Type classification: {normalize_parameter_type_label(row.target_type)}",
    ]

    normalized_type = normalize_test_procedure_type(row.test_procedure_type)
    if normalized_type:
        lines.append(f"Test Procedure Type: {normalized_type}")
    if row.test_procedure_text:
        lines.append(f"Test Procedure Text: {row.test_procedure_text}")

    for index, reason in enumerate(reasons, start=1):
        lines.append(f'Change reason {index} - "{reason}"')
    return "\n".join(lines)


def build_material_handling_background_text(row: MaterialHandlingWorkbookRow) -> str:
    lines = [
        MATERIAL_HANDLING_BACKGROUND_START,
        f"Material Code: {row.material_code}",
        f"Current Specification Number: {row.source_spec_number or '—'}",
        f"Proposed New Specification Number: {row.proposed_spec_number or '—'}",
        f"Chemical Name: {row.chemical_name or '—'}",
    ]
    if row.unit:
        lines.append(f"Unit: {row.unit}")
    if row.officers_responsible:
        lines.append(f"Officers Responsible: {row.officers_responsible}")
    if row.remarks:
        lines.append(f"Material Handling Remarks: {row.remarks}")
    lines.append(MATERIAL_HANDLING_BACKGROUND_END)
    return "\n".join(lines)


def build_material_handling_justification_text(row: MaterialHandlingWorkbookRow) -> str:
    detail_lines = [
        f"Proposed specification number - {row.proposed_spec_number or '—'}",
        f"Chemical name - {row.chemical_name or '—'}",
        f"Physical state - {row.physical_state or '—'}",
        f"Storage conditions (general) - {row.storage_conditions_general or '—'}",
        f"Storage conditions (special) - {row.storage_conditions_special or '—'}",
        f"Packing - {row.container_type or '—'} / {row.container_capacity or '—'}",
        f"Primary storage classification - {row.primary_storage_classification or '—'}",
    ]
    if row.remarks:
        detail_lines.append(f"Justification note - {row.remarks}")
    return "\n".join(
        [
            MATERIAL_HANDLING_JUSTIFICATION_START,
            "\n".join(detail_lines),
            MATERIAL_HANDLING_JUSTIFICATION_END,
        ]
    )


def build_material_handling_master_payload(row: MaterialHandlingWorkbookRow) -> dict[str, str]:
    return {
        "physical_state": row.physical_state,
        "volatility": row.volatility,
        "sunlight_sensitivity": row.sunlight_sensitivity,
        "moisture_sensitivity": row.moisture_sensitivity,
        "temperature_sensitivity": row.temperature_sensitivity,
        "reactivity": row.reactivity,
        "flammable": row.flammable,
        "toxic": row.toxic,
        "corrosive": row.corrosive,
        "storage_conditions_general": row.storage_conditions_general,
        "storage_conditions_special": row.storage_conditions_special,
        "container_type": row.container_type,
        "container_capacity": row.container_capacity,
        "container_description": row.container_description,
        "primary_storage_classification": row.primary_storage_classification,
    }


def apply_material_handling_core_updates(target_draft: object, source_draft: object) -> bool:
    changed = False
    proposed_spec_number = canonicalize_spec_number(getattr(source_draft, "spec_number", ""))
    proposed_chemical_name = sanitize_multiline_text(
        getattr(source_draft, "chemical_name", "") or "",
        max_length=255,
    )
    if proposed_spec_number and canonicalize_spec_number(getattr(target_draft, "spec_number", "")) != proposed_spec_number:
        setattr(target_draft, "spec_number", proposed_spec_number)
        changed = True
    if proposed_chemical_name and sanitize_multiline_text(getattr(target_draft, "chemical_name", "") or "", max_length=255) != proposed_chemical_name:
        setattr(target_draft, "chemical_name", proposed_chemical_name)
        changed = True
    return changed


def merge_imported_justification_block(
    existing_text: str,
    justification_entries: list[str],
    *,
    workbook_name: str,
    sheet_name: str,
) -> str:
    manual_text = _IMPORT_BLOCK_RE.sub("", existing_text or "").strip()
    if not justification_entries:
        return manual_text

    import_lines = [
        IMPORTED_JUSTIFICATION_START,
        f"Workbook: {workbook_name}",
        f"Sheet: {sheet_name}",
        "",
        "\n\n".join(justification_entries),
        IMPORTED_JUSTIFICATION_END,
    ]
    imported_block = "\n".join(import_lines).strip()
    if not manual_text:
        return imported_block
    return f"{manual_text}\n\n{imported_block}"


def merge_material_handling_section_block(
    existing_text: str,
    imported_text: str,
    *,
    start_marker: str,
    end_marker: str,
) -> str:
    pattern = re.compile(
        rf"{re.escape(start_marker)}.*?{re.escape(end_marker)}",
        re.DOTALL,
    )
    manual_text = pattern.sub("", existing_text or "").strip()
    if not imported_text:
        return manual_text
    if not manual_text:
        return imported_text.strip()
    return f"{manual_text}\n\n{imported_text.strip()}"


def _apply_row_updates_to_parameter(parameter: object, row: TypeClassificationWorkbookRow) -> bool:
    changed = False

    target_type = normalize_parameter_type_label(row.target_type)
    if normalize_parameter_type_label(getattr(parameter, "parameter_type", "")) != target_type:
        setattr(parameter, "parameter_type", target_type)
        changed = True

    if row.test_procedure_type:
        normalized_procedure_type = normalize_test_procedure_type(row.test_procedure_type)
        current_procedure_type = normalize_test_procedure_type(
            getattr(parameter, "test_procedure_type", "") or getattr(parameter, "test_method", "")
        )
        if current_procedure_type != normalized_procedure_type:
            setattr(parameter, "test_procedure_type", normalized_procedure_type)
            setattr(parameter, "test_method", normalized_procedure_type)
            changed = True

    if row.test_procedure_text:
        current_text = sanitize_multiline_text(getattr(parameter, "test_procedure_text", "") or "", max_length=4000)
        if current_text != row.test_procedure_text:
            setattr(parameter, "test_procedure_text", row.test_procedure_text)
            changed = True

    parameter_justification = build_parameter_justification_entry(row)
    if parameter_justification:
        current_justification = sanitize_multiline_text(
            getattr(parameter, "justification", "") or "",
            max_length=20000,
        )
        if current_justification != parameter_justification:
            setattr(parameter, "justification", parameter_justification)
            changed = True

    return changed


def _worksheet_is_blank(worksheet) -> bool:
    for row in worksheet.iter_rows(values_only=True):
        if not _row_is_blank(row):
            return False
    return True


def _find_header_map(worksheet) -> tuple[int | None, dict[str, int] | None]:
    scan_limit = min(worksheet.max_row, 10)
    for row_number in range(1, scan_limit + 1):
        header_map: dict[str, int] = {}
        for column_number in range(1, worksheet.max_column + 1):
            normalized_header = _normalize_header(worksheet.cell(row=row_number, column=column_number).value)
            field_name = _HEADER_ALIASES.get(normalized_header)
            if field_name and field_name not in header_map:
                header_map[field_name] = column_number
        if _REQUIRED_HEADERS.issubset(header_map):
            return row_number, header_map
    return None, None


def _find_material_handling_header_map(worksheet) -> tuple[int | None, dict[str, int] | None]:
    scan_limit = min(worksheet.max_row, 10)
    for row_number in range(1, scan_limit + 1):
        header_map: dict[str, int] = {}
        for column_number in range(1, worksheet.max_column + 1):
            normalized_header = _normalize_header(worksheet.cell(row=row_number, column=column_number).value)
            field_name = _MATERIAL_HANDLING_HEADER_ALIASES.get(normalized_header)
            if field_name and field_name not in header_map:
                header_map[field_name] = column_number
        if _MATERIAL_HANDLING_REQUIRED_HEADERS.issubset(header_map):
            return row_number, header_map
    return None, None


def _normalize_header(raw_value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(raw_value or "").strip().lower())


def _row_is_blank(values) -> bool:
    for value in values:
        if value is None:
            continue
        if str(value).strip():
            return False
    return True


def _coerce_parameter_serial(raw_value: object) -> int | None:
    if raw_value is None:
        return None
    if isinstance(raw_value, int):
        return raw_value
    if isinstance(raw_value, float) and raw_value.is_integer():
        return int(raw_value)
    text = str(raw_value).strip()
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    try:
        return int(text)
    except ValueError:
        return None


def _normalize_reason_text(raw_text: str) -> str:
    if not raw_text:
        return ""
    return re.sub(r"\bessential\b", "Desirable", raw_text, flags=re.IGNORECASE)


def _build_parameter_lookup(draft: object) -> dict[int, object]:
    parameters_attr = getattr(draft, "parameters", [])
    if hasattr(parameters_attr, "all"):
        parameters = parameters_attr.all()
    else:
        parameters = list(parameters_attr)

    lookup: dict[int, object] = {}
    for parameter in parameters:
        serial = _coerce_parameter_serial(getattr(parameter, "sort_order", None))
        if serial is not None and serial not in lookup:
            lookup[serial] = parameter
    return lookup
