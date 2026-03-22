"""Admin blueprint routes – User Management with User Governance V2.

Governance V2 additions:
  - Reporting hierarchy assignment
  - Module-level access permissions (checkboxes)
  - Audit logging for governance changes
  - User category reflected via role assignment
"""

import logging
import os
from pathlib import Path
import tempfile
from datetime import datetime
from io import BytesIO
from flask import render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import joinedload
from app.modules.admin import admin_bp
from app.extensions import db
from app.features import get_admin_module_options
from app.models.core.backup_snapshot import BackupSnapshot
from app.models.core.user import User
from app.models.core.role import Role
from app.models.core.module_admin_assignment import ModuleAdminAssignment
from app.models.office.office import Office
from app.models.core.audit_log import AuditLog
from app.models.core.activity_log import ActivityLog
from app.models.core.announcement import (
    Announcement,
    AnnouncementRecipient,
    AnnouncementVote,
)
from app.models.core.notification import Notification
from app.models.csc.draft import CSCDraft
from app.models.csc.revision import CSCRevision
from app.models.inventory.inventory_consumption_seed import InventoryConsumptionSeed
from app.models.inventory.inventory_procurement_seed import InventoryProcurementSeed
from app.models.inventory.inventory_upload import InventoryUpload
from app.models.inventory.material_master import MaterialMaster
from app.models.tasks.recurring_task_collaborator import RecurringTaskCollaborator
from app.models.tasks.recurring_task_template import RecurringTaskTemplate
from app.models.tasks.task import Task
from app.models.tasks.task_collaborator import TaskCollaborator
from app.models.tasks.task_update import TaskUpdate
from app.models.core.user_module_permission import (
    UserModulePermission,
    SUPER_USER_MODULES,
)
from app.core.utils.decorators import roles_required
from app.core.utils.request_meta import get_client_ip, get_user_agent
from app.core.utils.activity import log_activity
from app.core.permissions import can_manage_offices
from app.core.module_registry import invalidate_user_module_access_cache
from app.core.services.notifications import create_notification
from app.core.services.backups import (
    BackupError,
    build_backup_filename,
    create_full_backup_bundle,
    restore_database_backup,
    get_runtime_environment_name,
    validate_backup_file,
)
from app.core.roles import (
    ADMIN_ROLE,
    ROLE_DESCRIPTIONS,
    SUPERUSER_ROLE,
    USER_ROLE,
    ROLE_REGISTRY,
    canonicalize_role_name,
)

logger = logging.getLogger(__name__)


# ── Helper ──────────────────────────────────────────────────────
def _client_ip():
    return get_client_ip()


def _set_module_permissions(user: User, selected_codes: list, role: Role):
    """
    Clear existing module permissions for *user* and set fresh ones.

    superuser → all business modules
    admin     → admin_users only
    user      → exactly what is in selected_codes (admin-controlled)
    """
    UserModulePermission.query.filter_by(user_id=user.id).delete()
    db.session.flush()

    canonical_role = canonicalize_role_name(role.name) if role else None

    if canonical_role == SUPERUSER_ROLE:
        codes_to_save = list(SUPER_USER_MODULES)
    elif canonical_role == ADMIN_ROLE:
        codes_to_save = ["admin_users"]
    else:
        codes_to_save = list(selected_codes)

    for code in codes_to_save:
        db.session.add(
            UserModulePermission(
                user_id=user.id,
                module_code=code,
                can_access=True,
            )
        )
    db.session.flush()


def _module_options():
    """Expose module permissions with app-level feature status for admin forms."""
    return get_admin_module_options()


def _manageable_module_options():
    """Return business modules that can have workspace-assigned module admins."""
    return [
        option
        for option in _module_options()
        if option["code"] not in {"dashboard", "admin_users"}
    ]


def _user_can_admin_module(user: User, module_code: str) -> bool:
    """Check whether a user has underlying access for a module-admin assignment."""
    canonical_role = canonicalize_role_name(user.role.name) if user.role else None

    if canonical_role == SUPERUSER_ROLE:
        return module_code in SUPER_USER_MODULES

    return (
        UserModulePermission.query.filter_by(
            user_id=user.id,
            module_code=module_code,
            can_access=True,
        ).first()
        is not None
    )


def _assignable_roles():
    """Return active role choices using canonical labels with legacy fallback."""
    roles = Role.query.filter(Role.is_active == True).order_by(Role.name).all()
    canonical_roles = {}

    for role in roles:
        canonical_name = canonicalize_role_name(role.name)
        if canonical_name not in ROLE_REGISTRY:
            continue

        existing = canonical_roles.get(canonical_name)
        if existing is None or role.name == canonical_name:
            canonical_roles[canonical_name] = {
                "id": role.id,
                "name": canonical_name,
                "description": ROLE_DESCRIPTIONS.get(canonical_name, role.description),
            }

    return [
        canonical_roles[role_name]
        for role_name in ROLE_REGISTRY
        if role_name in canonical_roles
    ]


def _display_user_name(user) -> str:
    if user is None:
        return "Unassigned"

    full_name = (user.full_name or "").strip()
    if full_name:
        return full_name

    username = (user.username or "").strip()
    return username or "Unassigned"


def _user_sort_key(user: User):
    return (_display_user_name(user).lower(), (user.username or "").lower(), user.id)


def _resolve_officer_selection(raw_id, label, target_user_id=None):
    raw_value = (raw_id or "").strip()
    if not raw_value:
        return None, None

    if not raw_value.isdigit():
        return None, f"Select a valid {label}."

    officer = User.query.filter_by(id=int(raw_value), is_active=True).first()
    if officer is None:
        return None, f"Selected {label} is invalid or inactive."

    if target_user_id is not None and officer.id == target_user_id:
        return None, f"A user cannot be their own {label}."

    return officer, None


def _parse_power_user_limit(raw_value):
    cleaned = (raw_value or "").strip()
    if cleaned == "":
        return 0, None
    if not cleaned.isdigit():
        return 0, "Power user slots must be a non-negative whole number."

    value = int(cleaned)
    if value < 0:
        return 0, "Power user slots cannot be negative."
    if value > 25:
        return 0, "Power user slots cannot exceed 25 for one office."
    return value, None


def _power_user_candidates_for_office(office_id: int) -> list[User]:
    return (
        User.query
        .options(joinedload(User.role))
        .filter(User.office_id == office_id, User.is_active.is_(True))
        .order_by(User.full_name, User.username)
        .all()
    )


def _resolve_power_user_selection(office_id: int, raw_ids: list[str]) -> tuple[list[int], list[str]]:
    errors: list[str] = []
    selected_ids: list[int] = []

    if not raw_ids:
        return selected_ids, errors

    candidates = {user.id: user for user in _power_user_candidates_for_office(office_id)}

    for raw_id in raw_ids:
        value = (raw_id or "").strip()
        if not value:
            continue
        if not value.isdigit():
            errors.append("Select valid power users from the selected office.")
            continue
        user_id = int(value)
        if user_id not in candidates:
            errors.append("Power users must be active users mapped to this office.")
            continue
        if user_id not in selected_ids:
            selected_ids.append(user_id)

    return selected_ids, errors


def _record_dependency_change(changes: list[str], label: str, count: int) -> None:
    if count:
        suffix = "" if count == 1 else "s"
        changes.append(f"{count} {label}{suffix}")


def _format_excel_datetime(value) -> str:
    if value is None:
        return ""
    return value.strftime("%d %b %Y %I:%M %p")


def _autosize_worksheet(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def _write_report_sheet(sheet, headers: list[str], rows: list[list[object]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F7")
    header_font = Font(bold=True, color="1F2937")

    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        sheet.append(row)

    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _autosize_worksheet(sheet)


def _build_user_report_workbook(
    users: list[User],
    roles: list[Role],
    offices: list[Office],
    permissions: list[UserModulePermission],
    module_admin_assignments: list[ModuleAdminAssignment],
    hierarchy_summary: dict,
):
    from openpyxl import Workbook

    workbook = Workbook()
    module_label_map = {option["code"]: option["label"] for option in _module_options()}
    user_by_id = {user.id: user for user in users}
    permission_map: dict[int, list[UserModulePermission]] = {}
    for permission in permissions:
        permission_map.setdefault(permission.user_id, []).append(permission)
    for permission_rows in permission_map.values():
        permission_rows.sort(key=lambda item: (item.module_code or "").lower())

    module_admin_map: dict[int, list[ModuleAdminAssignment]] = {}
    for assignment in module_admin_assignments:
        module_admin_map.setdefault(assignment.user_id, []).append(assignment)
    for assignment_rows in module_admin_map.values():
        assignment_rows.sort(key=lambda item: (item.module_code or "").lower())

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_rows = [
        ["Generated At", _format_excel_datetime(datetime.now())],
        ["Exported By", current_user.username],
        ["Total Users", hierarchy_summary["total_users"]],
        ["Mapped Users", hierarchy_summary["mapped_users"]],
        ["Fully Mapped Users", hierarchy_summary["fully_mapped_users"]],
        ["Office Heads", hierarchy_summary["office_heads"]],
        ["Power Users", hierarchy_summary.get("power_users", 0)],
        ["Configured Power User Slots", hierarchy_summary.get("power_user_slots", 0)],
        ["Active Users", sum(1 for user in users if user.is_active)],
        ["Inactive Users", sum(1 for user in users if not user.is_active)],
        ["Roles", len(roles)],
        ["Offices", len(offices)],
        ["Module Permission Rows", len(permissions)],
        ["Module Admin Assignments", len(module_admin_assignments)],
    ]
    _write_report_sheet(summary_sheet, ["Metric", "Value"], summary_rows)

    users_sheet = workbook.create_sheet("Users")
    user_rows = []
    for user in sorted(users, key=_user_sort_key):
        permission_labels = [
            module_label_map.get(permission.module_code, permission.module_code)
            for permission in permission_map.get(user.id, [])
            if permission.can_access
        ]
        module_admin_labels = [
            module_label_map.get(assignment.module_code, assignment.module_code)
            for assignment in module_admin_map.get(user.id, [])
        ]
        user_rows.append(
            [
                user.id,
                user.username,
                user.full_name or "",
                user.email or "",
                user.employee_code or "",
                user.designation or "",
                user.role.name if user.role else "",
                user.office.office_name if user.office else "",
                user.office.office_code if user.office else "",
                "Yes" if user.is_power_user else "No",
                _display_user_name(user.controlling_officer) if user.controlling_officer else "",
                _display_user_name(user.reviewing_officer) if user.reviewing_officer else "",
                _display_user_name(user.accepting_officer) if user.accepting_officer else "",
                "Active" if user.is_active else "Inactive",
                "Yes" if user.must_change_password else "No",
                _format_excel_datetime(user.last_login_at),
                _format_excel_datetime(user.created_at),
                _format_excel_datetime(user.updated_at),
                ", ".join(permission_labels),
                ", ".join(module_admin_labels),
            ]
        )
    _write_report_sheet(
        users_sheet,
        [
            "User ID",
            "Username",
            "Full Name",
            "Email",
            "Employee Code",
            "Designation",
            "Role",
            "Office",
            "Office Code",
            "Power User",
            "Controlling Officer",
            "Reviewing Officer",
            "Accepting Officer",
            "Status",
            "Must Change Password",
            "Last Login",
            "Created At",
            "Updated At",
            "Module Access",
            "Module Admin Ownership",
        ],
        user_rows,
    )

    roles_sheet = workbook.create_sheet("Roles")
    role_rows = []
    for role in sorted(roles, key=lambda item: (item.name or "").lower()):
        role_rows.append(
            [
                role.id,
                role.name,
                role.description or "",
                "Active" if role.is_active else "Inactive",
                role.users.count(),
                _format_excel_datetime(role.created_at),
                _format_excel_datetime(role.updated_at),
            ]
        )
    _write_report_sheet(
        roles_sheet,
        ["Role ID", "Role", "Description", "Status", "User Count", "Created At", "Updated At"],
        role_rows,
    )

    offices_sheet = workbook.create_sheet("Offices")
    office_rows = []
    for office in sorted(offices, key=lambda item: (item.office_name or "").lower()):
        office_users = [user for user in users if user.office_id == office.id]
        office_rows.append(
            [
                office.id,
                office.office_code,
                office.office_name,
                office.location or "",
                "Active" if office.is_active else "Inactive",
                office.power_user_limit,
                sum(1 for user in office_users if user.is_power_user and user.is_active),
                len(office_users),
                sum(1 for user in office_users if user.is_active),
                _format_excel_datetime(office.created_at),
                _format_excel_datetime(office.updated_at),
            ]
        )
    _write_report_sheet(
        offices_sheet,
        [
            "Office ID",
            "Office Code",
            "Office Name",
            "Location",
            "Status",
            "Power User Slots",
            "Assigned Power Users",
            "Total Users",
            "Active Users",
            "Created At",
            "Updated At",
        ],
        office_rows,
    )

    permissions_sheet = workbook.create_sheet("Module Access")
    permission_rows = []
    for permission in sorted(
        permissions,
        key=lambda item: (
            (user_by_id.get(item.user_id).username if user_by_id.get(item.user_id) else "").lower(),
            (item.module_code or "").lower(),
        ),
    ):
        user = user_by_id.get(permission.user_id)
        permission_rows.append(
            [
                user.username if user else permission.user_id,
                user.full_name if user else "",
                user.role.name if user and user.role else "",
                module_label_map.get(permission.module_code, permission.module_code),
                permission.module_code,
                "Yes" if permission.can_access else "No",
                _format_excel_datetime(permission.created_at),
                _format_excel_datetime(permission.updated_at),
            ]
        )
    _write_report_sheet(
        permissions_sheet,
        ["Username", "Full Name", "Role", "Module", "Module Code", "Can Access", "Created At", "Updated At"],
        permission_rows,
    )

    module_admins_sheet = workbook.create_sheet("Module Admins")
    admin_rows = []
    for assignment in sorted(
        module_admin_assignments,
        key=lambda item: (
            module_label_map.get(item.module_code, item.module_code).lower(),
            item.user.username.lower() if item.user else "",
        ),
    ):
        admin_rows.append(
            [
                module_label_map.get(assignment.module_code, assignment.module_code),
                assignment.module_code,
                assignment.user.username if assignment.user else "",
                assignment.user.full_name if assignment.user else "",
                assignment.user.role.name if assignment.user and assignment.user.role else "",
                assignment.user.office.office_name if assignment.user and assignment.user.office else "",
                _format_excel_datetime(assignment.created_at),
                _format_excel_datetime(assignment.updated_at),
            ]
        )
    _write_report_sheet(
        module_admins_sheet,
        ["Module", "Module Code", "Username", "Full Name", "Role", "Office", "Created At", "Updated At"],
        admin_rows,
    )

    return workbook


ORGANOGRAM_LEVEL_META = {
    "head": {"label": "Office Head", "order": 0},
    "reviewing": {"label": "Reviewing Officer", "order": 1},
    "controlling": {"label": "Controlling Officer", "order": 2},
    "user": {"label": "User", "order": 3},
}


def _new_organogram_node(user: User) -> dict:
    meta = ORGANOGRAM_LEVEL_META["user"]
    return {
        "key": f"user:{user.id}",
        "id": user.id,
        "level": "user",
        "level_label": meta["label"],
        "level_order": meta["order"],
        "name": _display_user_name(user),
        "username": user.username,
        "role_name": user.role.name if user.role else "No role",
        "office_name": user.office.office_name if user.office else "No office",
        "designation": (user.designation or "").strip(),
        "is_active": user.is_active,
        "tags": [{"level": "user", "label": meta["label"], "order": meta["order"]}],
        "_tag_levels": {"user"},
        "children": [],
        "_child_index": {},
    }


def _apply_organogram_tag(node: dict, level: str) -> None:
    if level not in ORGANOGRAM_LEVEL_META or level in node["_tag_levels"]:
        return

    meta = ORGANOGRAM_LEVEL_META[level]
    node["_tag_levels"].add(level)
    node["tags"].append(
        {
            "level": level,
            "label": meta["label"],
            "order": meta["order"],
        }
    )
    if meta["order"] < node["level_order"]:
        node["level"] = level
        node["level_label"] = meta["label"]
        node["level_order"] = meta["order"]


def _organogram_node_sort_key(node: dict):
    return (
        node["level_order"],
        (node["name"] or "").lower(),
        (node["username"] or "").lower(),
        node["id"],
    )


def _build_user_organogram(users):
    organogram_users = [
        user
        for user in users
        if canonicalize_role_name(user.role.name if user.role else None) != ADMIN_ROLE
    ]
    users_by_id = {user.id: user for user in organogram_users}
    office_sections = {}

    for user in organogram_users:
        office_key = user.office_id or 0
        office_name = user.office.office_name if user.office else "Unassigned Office"
        section = office_sections.setdefault(
            office_key,
            {
                "office_id": user.office_id,
                "office_name": office_name,
                "users": [],
            },
        )
        section["users"].append(user)

    def _finalize(nodes: list[dict]) -> list[dict]:
        nodes.sort(key=_organogram_node_sort_key)
        finalized = []
        for node in nodes:
            node["children"] = _finalize(node["children"])
            sorted_tags = sorted(node["tags"], key=lambda tag: (tag["order"], tag["label"]))
            non_user_tags = [tag for tag in sorted_tags if tag["level"] != "user"]
            node["display_tags"] = non_user_tags or sorted_tags
            node.pop("tags", None)
            node.pop("_tag_levels", None)
            node.pop("_child_index", None)
            finalized.append(node)
        return finalized

    organogram = []

    for section in sorted(
        office_sections.values(),
        key=lambda item: ((item["office_name"] or "").lower(), item["office_id"] or 0),
    ):
        office_users = section["users"]
        office_user_ids = {user.id for user in office_users}
        single_user_office = len(office_users) == 1
        reviewing_reference_ids = {
            int(user.reviewing_officer_id)
            for user in office_users
            if user.reviewing_officer_id is not None and int(user.reviewing_officer_id) in office_user_ids
        }
        controlling_reference_ids = {
            int(user.controlling_officer_id)
            for user in office_users
            if user.controlling_officer_id is not None and int(user.controlling_officer_id) in office_user_ids
        }
        roots = []
        node_index = {}
        inferred_parent_map = {}

        def _ensure_node(user: User) -> dict:
            node = node_index.get(user.id)
            if node is None:
                node = _new_organogram_node(user)
                node_index[user.id] = node
            return node

        def _valid_officer(officer_id):
            if officer_id is None:
                return None
            officer = users_by_id.get(officer_id)
            if officer is None or officer.id not in office_user_ids:
                return None
            return officer

        for user in sorted(office_users, key=_user_sort_key):
            node = _ensure_node(user)
            if user.id in reviewing_reference_ids:
                _apply_organogram_tag(node, "reviewing")
            if user.id in controlling_reference_ids:
                _apply_organogram_tag(node, "controlling")
            if not single_user_office and not _valid_officer(user.controlling_officer_id):
                _apply_organogram_tag(node, "head")

            reviewing_officer = _valid_officer(user.reviewing_officer_id)
            controlling_officer = _valid_officer(user.controlling_officer_id)
            if reviewing_officer and controlling_officer and reviewing_officer.id != controlling_officer.id:
                inferred_parent_map.setdefault(controlling_officer.id, reviewing_officer.id)

        parent_map = {}
        for user in sorted(office_users, key=_user_sort_key):
            controlling_officer = _valid_officer(user.controlling_officer_id)
            reviewing_officer = _valid_officer(user.reviewing_officer_id)
            if controlling_officer and controlling_officer.id != user.id:
                parent_map[user.id] = controlling_officer.id
            elif reviewing_officer and reviewing_officer.id != user.id:
                parent_map[user.id] = reviewing_officer.id
            else:
                inferred_parent_id = inferred_parent_map.get(user.id)
                parent_map[user.id] = (
                    inferred_parent_id
                    if inferred_parent_id is not None and inferred_parent_id != user.id
                    else None
                )

        def _creates_cycle(user_id: int, parent_id) -> bool:
            seen_ids = {user_id}
            current_id = parent_id
            while current_id is not None:
                if current_id in seen_ids:
                    return True
                seen_ids.add(current_id)
                current_id = parent_map.get(current_id)
            return False

        for user_id, parent_id in list(parent_map.items()):
            if parent_id is not None and _creates_cycle(user_id, parent_id):
                parent_map[user_id] = None

        attached_ids = set()
        for user in sorted(office_users, key=_user_sort_key):
            node = _ensure_node(user)
            parent_id = parent_map.get(user.id)
            if parent_id is None:
                continue
            parent = node_index.get(parent_id)
            if parent is None:
                continue
            if user.id not in parent["_child_index"]:
                parent["_child_index"][user.id] = node
                parent["children"].append(node)
                attached_ids.add(user.id)

        for user in sorted(office_users, key=_user_sort_key):
            node = _ensure_node(user)
            if user.id not in attached_ids:
                roots.append(node)

        finalized_roots = _finalize(roots)
        organogram.append(
            {
                "office_id": section["office_id"],
                "office_name": section["office_name"],
                "total_users": len(office_users),
                "mapped_users": sum(
                    1
                    for user in office_users
                    if any(
                        [
                            user.controlling_officer_id,
                            user.reviewing_officer_id,
                            user.accepting_officer_id,
                        ]
                    )
                ),
                "office_heads": 0 if single_user_office else sum(
                    1
                    for node in finalized_roots
                    if any(tag["level"] == "head" for tag in node["display_tags"])
                ),
                "unmapped_users": sum(
                    1
                    for user in office_users
                    if not any(
                        [
                            user.controlling_officer_id,
                            user.reviewing_officer_id,
                            user.accepting_officer_id,
                        ]
                    )
                ),
                "nodes": finalized_roots,
            }
        )

    return organogram


# ── Users List ───────────────────────────────────────────────────
@admin_bp.route("/users")
@login_required
@roles_required(ADMIN_ROLE)
def users():
    all_users = (
        User.query
        .options(
            joinedload(User.role),
            joinedload(User.office),
            joinedload(User.controlling_officer),
            joinedload(User.reviewing_officer),
            joinedload(User.accepting_officer),
        )
        .order_by(User.created_at.desc())
        .all()
    )
    organogram = _build_user_organogram(all_users)
    hierarchy_summary = {
        "total_users": len(all_users),
        "mapped_users": sum(
            1
            for user in all_users
            if any(
                [
                    user.controlling_officer_id,
                    user.reviewing_officer_id,
                    user.accepting_officer_id,
                ]
            )
        ),
        "fully_mapped_users": sum(
            1
            for user in all_users
            if user.controlling_officer_id
            and user.reviewing_officer_id
            and user.accepting_officer_id
        ),
        "office_heads": sum(section["office_heads"] for section in organogram),
        "power_users": sum(1 for user in all_users if user.is_power_user and user.is_active),
        "power_user_slots": sum(office.power_user_limit for office in Office.query.all()),
    }
    return render_template(
        "admin/users.html",
        users=all_users,
        organogram=organogram,
        hierarchy_summary=hierarchy_summary,
    )


@admin_bp.route("/users/export")
@login_required
@roles_required(ADMIN_ROLE)
def export_users_report():
    users = (
        User.query
        .options(
            joinedload(User.role),
            joinedload(User.office),
            joinedload(User.controlling_officer),
            joinedload(User.reviewing_officer),
            joinedload(User.accepting_officer),
        )
        .order_by(User.created_at.desc())
        .all()
    )
    roles = Role.query.order_by(Role.name.asc()).all()
    offices = Office.query.order_by(Office.office_name.asc()).all()
    permissions = (
        UserModulePermission.query
        .order_by(UserModulePermission.user_id.asc(), UserModulePermission.module_code.asc())
        .all()
    )
    module_admin_assignments = (
        ModuleAdminAssignment.query
        .options(
            joinedload(ModuleAdminAssignment.user).joinedload(User.role),
            joinedload(ModuleAdminAssignment.user).joinedload(User.office),
        )
        .order_by(ModuleAdminAssignment.module_code.asc(), ModuleAdminAssignment.user_id.asc())
        .all()
    )

    organogram = _build_user_organogram(users)
    hierarchy_summary = {
        "total_users": len(users),
        "mapped_users": sum(
            1
            for user in users
            if any(
                [
                    user.controlling_officer_id,
                    user.reviewing_officer_id,
                    user.accepting_officer_id,
                ]
            )
        ),
        "fully_mapped_users": sum(
            1
            for user in users
            if user.controlling_officer_id
            and user.reviewing_officer_id
            and user.accepting_officer_id
        ),
        "office_heads": sum(section["office_heads"] for section in organogram),
        "power_users": sum(1 for user in users if user.is_power_user and user.is_active),
        "power_user_slots": sum(office.power_user_limit for office in offices),
    }

    workbook = _build_user_report_workbook(
        users=users,
        roles=roles,
        offices=offices,
        permissions=permissions,
        module_admin_assignments=module_admin_assignments,
        hierarchy_summary=hierarchy_summary,
    )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = f"user_management_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    try:
        db.session.add(
            AuditLog(
                action="USER_REPORT_EXPORTED",
                user_id=current_user.id,
                entity_type="UserReport",
                entity_id=filename,
                details=(
                    f"Admin '{current_user.username}' exported the user management report "
                    f"'{filename}'."
                ),
                ip_address=AuditLog._normalize_ip(_client_ip()),
                user_agent=AuditLog._normalize_user_agent(get_user_agent()),
            )
        )
        log_activity(
            current_user.username,
            "user_report_exported",
            "user_management",
            filename,
        )
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()

    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
    )


@admin_bp.route("/modules", methods=["GET", "POST"])
@login_required
@roles_required(ADMIN_ROLE)
def module_management():
    """Assign up to two module admins for each business module."""
    module_options = _manageable_module_options()
    active_users = User.query.filter_by(is_active=True).order_by(User.full_name, User.username).all()

    if request.method == "POST":
        touched_user_ids = set()

        try:
            for module in module_options:
                module_code = module["code"]
                raw_values = [
                    request.form.get(f"{module_code}_admin_1", "").strip(),
                    request.form.get(f"{module_code}_admin_2", "").strip(),
                ]
                selected_user_ids = []
                for value in raw_values:
                    if not value:
                        continue
                    try:
                        selected_user_ids.append(int(value))
                    except ValueError:
                        raise ValueError(f"Invalid user selection for {module['label']}.")

                if len(selected_user_ids) != len(set(selected_user_ids)):
                    raise ValueError(f"Select distinct module admins for {module['label']}.")
                if len(selected_user_ids) > 2:
                    raise ValueError(f"{module['label']} can have a maximum of two module admins.")

                selected_users = (
                    User.query.filter(User.id.in_(selected_user_ids)).all()
                    if selected_user_ids
                    else []
                )
                if len(selected_users) != len(selected_user_ids):
                    raise ValueError(f"One or more selected users for {module['label']} no longer exist.")

                for user in selected_users:
                    if not _user_can_admin_module(user, module_code):
                        raise ValueError(
                            f"{user.username} does not currently hold {module['label']} module access."
                        )

                existing_assignments = ModuleAdminAssignment.query.filter_by(module_code=module_code).all()
                touched_user_ids.update(assignment.user_id for assignment in existing_assignments)
                ModuleAdminAssignment.query.filter_by(module_code=module_code).delete()
                db.session.flush()

                for user_id in selected_user_ids:
                    db.session.add(
                        ModuleAdminAssignment(
                            module_code=module_code,
                            user_id=user_id,
                        )
                    )
                    touched_user_ids.add(user_id)

            db.session.commit()

            for user_id in touched_user_ids:
                invalidate_user_module_access_cache(user_id)

            AuditLog.log(
                action="MODULE_ADMINS_UPDATED",
                user_id=current_user.id,
                entity_type="ModuleAdminAssignment",
                entity_id="workspace",
                details=f"Workspace admin '{current_user.username}' updated module admin assignments.",
                ip_address=_client_ip(),
                user_agent=get_user_agent(),
            )
            log_activity(current_user.username, "module_admins_updated", "module_management", "workspace")
            flash("Module admin assignments updated.", "success")
            return redirect(url_for("admin.module_management"))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")

    assignments_by_module = {}
    for assignment in (
        ModuleAdminAssignment.query
        .order_by(ModuleAdminAssignment.module_code.asc(), ModuleAdminAssignment.user_id.asc())
        .all()
    ):
        assignments_by_module.setdefault(assignment.module_code, []).append(assignment.user_id)

    modules = []
    for module in module_options:
        eligible_users = [
            user for user in active_users
            if _user_can_admin_module(user, module["code"])
        ]
        modules.append(
            {
                **module,
                "eligible_users": eligible_users,
                "assigned_user_ids": assignments_by_module.get(module["code"], []),
            }
        )

    return render_template(
        "admin/module_management.html",
        modules=modules,
    )


# ── Create User ──────────────────────────────────────────────────
@admin_bp.route("/users/create", methods=["GET", "POST"])
@login_required
@roles_required(ADMIN_ROLE)
def create_user():
    roles = _assignable_roles()
    offices = Office.query.filter_by(is_active=True).order_by(Office.office_name).all()
    # Officers dropdown – all active users except the one being created
    officers = User.query.filter_by(is_active=True).order_by(User.full_name, User.username).all()

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        full_name = request.form.get("full_name", "").strip()
        email = request.form.get("email", "").strip().lower()
        role_id = request.form.get("role_id", "").strip()
        office_id = request.form.get("office_id", "").strip()
        designation = request.form.get("designation", "").strip()
        employee_code = request.form.get("employee_code", "").strip()
        temp_password = request.form.get("temporary_password", "")
        is_active = request.form.get("is_active") == "on"
        controlling_officer_id_raw = request.form.get("controlling_officer_id", "").strip()
        reviewing_officer_id_raw = request.form.get("reviewing_officer_id", "").strip()
        accepting_officer_id_raw = request.form.get("accepting_officer_id", "").strip()
        selected_modules = request.form.getlist("module_access")

        # ── Validation ────────────────────────────────────────────
        errors = []

        if not username:
            errors.append("Username is required.")
        if not full_name:
            errors.append("Full name is required.")
        if not email:
            errors.append("Email is required.")
        if not role_id:
            errors.append("Role is required.")
        if not office_id:
            errors.append("Office is required.")
        if not temp_password:
            errors.append("Temporary password is required.")
        elif len(temp_password) < 6:
            errors.append("Password must be at least 6 characters.")

        if username and User.query.filter_by(username=username).first():
            errors.append(f"Username '{username}' is already taken.")
        if email and User.query.filter(User.email == email).first():
            errors.append(f"Email '{email}' is already registered.")

        role = None
        office = None
        if role_id:
            role = Role.query.filter_by(id=role_id, is_active=True).first()
            if not role:
                errors.append("Selected role is invalid or inactive.")
        if office_id:
            office = Office.query.filter_by(id=office_id, is_active=True).first()
            if not office:
                errors.append("Selected office is invalid or inactive.")

        # Validate officer references
        controlling_officer_id = None
        reviewing_officer_id = None
        accepting_officer_id = None
        co, co_error = _resolve_officer_selection(controlling_officer_id_raw, "controlling officer")
        ro, ro_error = _resolve_officer_selection(reviewing_officer_id_raw, "reviewing officer")
        ao, ao_error = _resolve_officer_selection(accepting_officer_id_raw, "accepting officer")
        if co_error:
            errors.append(co_error)
        elif co:
            controlling_officer_id = co.id
        if ro_error:
            errors.append(ro_error)
        elif ro:
            reviewing_officer_id = ro.id
        if ao_error:
            errors.append(ao_error)
        elif ao:
            accepting_officer_id = ao.id

        if errors:
            for err in errors:
                flash(err, "danger")
            return render_template(
                "admin/create_user.html",
                roles=roles,
                offices=offices,
                officers=officers,
                supported_modules=_module_options(),
                form_data=request.form,
                selected_modules=selected_modules,
            )

        # ── Create user ────────────────────────────────────────────
        new_user = User(
            username=username,
            full_name=full_name,
            email=email,
            role_id=int(role_id),
            office_id=int(office_id),
            designation=designation,
            employee_code=employee_code,
            is_active=is_active,
            must_change_password=True,
            controlling_officer_id=controlling_officer_id,
            reviewing_officer_id=reviewing_officer_id,
            accepting_officer_id=accepting_officer_id,
        )
        new_user.set_password(temp_password)
        db.session.add(new_user)
        db.session.flush()  # get new_user.id

        # ── Set module permissions ─────────────────────────────────
        _set_module_permissions(new_user, selected_modules, role)

        # ── Audit log ──────────────────────────────────────────────
        AuditLog.log(
            action="USER_CREATED",
            user_id=current_user.id,
            entity_type="User",
            entity_id=str(new_user.id),
            details=(
                f"Admin '{current_user.username}' created user '{username}' "
                f"(role={role.name}, office={office.office_name}, "
                f"modules={','.join(selected_modules) or 'auto'})"
            ),
            ip_address=_client_ip(),
            user_agent=get_user_agent(),
        )

        log_activity(current_user.username, "user_created", "user", username,
                     details=f"role={role.name}, office={office.office_name}")
        create_notification(
            user_id=new_user.id,
            title="Your account is ready",
            message=(
                f"{current_user.full_name or current_user.username} created your account "
                f"with the role '{role.name}' for {office.office_name}."
            ),
            severity="success",
            link="/dashboard",
        )
        invalidate_user_module_access_cache(new_user.id)
        db.session.commit()

        flash(
            f"User '{username}' created successfully. "
            "They must change password on first login.",
            "success",
        )
        return redirect(url_for("admin.users"))

    return render_template(
        "admin/create_user.html",
        roles=roles,
        offices=offices,
        officers=officers,
        supported_modules=_module_options(),
        form_data={},
        selected_modules=[],
    )


# ── Edit User ────────────────────────────────────────────────────
@admin_bp.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@login_required
@roles_required(ADMIN_ROLE)
def edit_user(user_id):
    target = User.query.get_or_404(user_id)
    roles = _assignable_roles()
    offices = Office.query.filter_by(is_active=True).order_by(Office.office_name).all()
    # Exclude the user being edited from officer dropdowns
    officers = (
        User.query.filter(User.is_active == True, User.id != user_id)
        .order_by(User.full_name, User.username)
        .all()
    )
    # Current module permissions for this user
    current_module_codes = [
        p.module_code
        for p in UserModulePermission.query.filter_by(user_id=user_id, can_access=True).all()
    ]

    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        email = request.form.get("email", "").strip().lower()
        role_id = request.form.get("role_id", "").strip()
        office_id = request.form.get("office_id", "").strip()
        designation = request.form.get("designation", "").strip()
        employee_code = request.form.get("employee_code", "").strip()
        is_active = request.form.get("is_active") == "on"
        controlling_officer_id_raw = request.form.get("controlling_officer_id", "").strip()
        reviewing_officer_id_raw = request.form.get("reviewing_officer_id", "").strip()
        accepting_officer_id_raw = request.form.get("accepting_officer_id", "").strip()
        selected_modules = request.form.getlist("module_access")

        # ── Validation ────────────────────────────────────────────
        errors = []

        if not full_name:
            errors.append("Full name is required.")
        if not email:
            errors.append("Email is required.")
        if not role_id:
            errors.append("Role is required.")
        if not office_id:
            errors.append("Office is required.")

        if email:
            conflict = User.query.filter(User.email == email, User.id != user_id).first()
            if conflict:
                errors.append(f"Email '{email}' is already registered to another user.")

        role = None
        office = None
        if role_id:
            role = Role.query.filter_by(id=role_id, is_active=True).first()
            if not role:
                errors.append("Selected role is invalid or inactive.")
        if office_id:
            office = Office.query.filter_by(id=office_id, is_active=True).first()
            if not office:
                errors.append("Selected office is invalid or inactive.")

        # Validate officer references
        controlling_officer_id = None
        reviewing_officer_id = None
        accepting_officer_id = None
        co, co_error = _resolve_officer_selection(
            controlling_officer_id_raw, "controlling officer", target_user_id=user_id
        )
        ro, ro_error = _resolve_officer_selection(
            reviewing_officer_id_raw, "reviewing officer", target_user_id=user_id
        )
        ao, ao_error = _resolve_officer_selection(
            accepting_officer_id_raw, "accepting officer", target_user_id=user_id
        )
        if co_error:
            errors.append(co_error)
        elif co:
            controlling_officer_id = co.id
        if ro_error:
            errors.append(ro_error)
        elif ro:
            reviewing_officer_id = ro.id
        if ao_error:
            errors.append(ao_error)
        elif ao:
            accepting_officer_id = ao.id

        if errors:
            for err in errors:
                flash(err, "danger")
            return render_template(
                "admin/edit_user.html",
                target=target,
                roles=roles,
                offices=offices,
                officers=officers,
                supported_modules=_module_options(),
                current_module_codes=selected_modules,
                form_data=request.form,
            )

        # ── Apply changes ──────────────────────────────────────────
        changed_fields = []
        previous_role_name = target.role.name if target.role else "No role"
        role_changed = str(target.role_id) != str(role_id)
        if target.full_name != full_name:
            changed_fields.append(f"full_name: '{target.full_name}' → '{full_name}'")
            target.full_name = full_name
        if target.email != email:
            changed_fields.append(f"email: '{target.email}' → '{email}'")
            target.email = email
        if str(target.role_id) != str(role_id):
            changed_fields.append(f"role_id: {target.role_id} → {role_id}")
            target.role_id = int(role_id)
        if str(target.office_id) != str(office_id):
            changed_fields.append(f"office_id: {target.office_id} → {office_id}")
            target.office_id = int(office_id)
            if target.is_power_user:
                changed_fields.append("power_user: cleared after office reassignment")
                target.is_power_user = False
        if target.designation != designation:
            changed_fields.append(f"designation: '{target.designation}' → '{designation}'")
            target.designation = designation
        if target.employee_code != employee_code:
            changed_fields.append(f"employee_code: '{target.employee_code}' → '{employee_code}'")
            target.employee_code = employee_code
        if target.is_active != is_active:
            changed_fields.append(f"is_active: {target.is_active} → {is_active}")
            target.is_active = is_active
            if not is_active and target.is_power_user:
                changed_fields.append("power_user: cleared because account was deactivated")
                target.is_power_user = False

        # Hierarchy changes
        old_co = target.controlling_officer_id
        old_ro = target.reviewing_officer_id
        old_ao = target.accepting_officer_id
        target.controlling_officer_id = controlling_officer_id
        target.reviewing_officer_id = reviewing_officer_id
        target.accepting_officer_id = accepting_officer_id
        if old_co != controlling_officer_id:
            changed_fields.append(f"controlling_officer_id: {old_co} → {controlling_officer_id}")
        if old_ro != reviewing_officer_id:
            changed_fields.append(f"reviewing_officer_id: {old_ro} → {reviewing_officer_id}")
        if old_ao != accepting_officer_id:
            changed_fields.append(f"accepting_officer_id: {old_ao} → {accepting_officer_id}")

        db.session.flush()

        # ── Update module permissions ──────────────────────────────
        _set_module_permissions(target, selected_modules, role)

        # ── Audit log ──────────────────────────────────────────────
        AuditLog.log(
            action="USER_UPDATED",
            user_id=current_user.id,
            entity_type="User",
            entity_id=str(target.id),
            details=(
                f"Admin '{current_user.username}' updated user '{target.username}'. "
                f"Changes: {'; '.join(changed_fields) if changed_fields else 'none'}"
            ),
            ip_address=_client_ip(),
            user_agent=get_user_agent(),
        )

        # Separate audit entries for governance changes
        if (
            old_co != controlling_officer_id
            or old_ro != reviewing_officer_id
            or old_ao != accepting_officer_id
        ):
            AuditLog.log(
                action="USER_HIERARCHY_UPDATED",
                user_id=current_user.id,
                entity_type="User",
                entity_id=str(target.id),
                details=(
                    f"Hierarchy updated for '{target.username}': "
                    f"controlling_officer_id={controlling_officer_id}, "
                    f"reviewing_officer_id={reviewing_officer_id}, "
                    f"accepting_officer_id={accepting_officer_id}"
                ),
                ip_address=_client_ip(),
                user_agent=get_user_agent(),
            )

        if sorted(current_module_codes) != sorted(selected_modules):
            AuditLog.log(
                action="USER_MODULE_ACCESS_UPDATED",
                user_id=current_user.id,
                entity_type="User",
                entity_id=str(target.id),
                details=(
                    f"Module access updated for '{target.username}': "
                    f"modules={','.join(selected_modules) or 'none'}"
                ),
                ip_address=_client_ip(),
                user_agent=get_user_agent(),
            )

        if changed_fields:
            log_activity(current_user.username, "user_updated", "user",
                         target.username,
                         details="; ".join(changed_fields))
        if sorted(current_module_codes) != sorted(selected_modules):
            log_activity(current_user.username, "role_changed", "user",
                         target.username,
                         details=f"modules={','.join(selected_modules) or 'none'}")
        if role_changed:
            create_notification(
                user_id=target.id,
                title="Your role has been updated",
                message=(
                    f"{current_user.full_name or current_user.username} changed your role "
                    f"from '{previous_role_name}' to '{role.name}'."
                ),
                severity="warning",
                link="/dashboard",
            )
        invalidate_user_module_access_cache(target.id)
        db.session.commit()

        flash(f"User '{target.username}' updated successfully.", "success")
        return redirect(url_for("admin.users"))

    return render_template(
        "admin/edit_user.html",
        target=target,
        roles=roles,
        offices=offices,
        officers=officers,
        supported_modules=_module_options(),
        current_module_codes=current_module_codes,
        form_data={},
    )


# ── Reset Password ───────────────────────────────────────────────
@admin_bp.route("/users/<int:user_id>/reset-password", methods=["GET", "POST"])
@login_required
@roles_required(ADMIN_ROLE)
def reset_user_password(user_id):
    target = User.query.get_or_404(user_id)

    if target.id == current_user.id:
        flash("Use /change-password to update your own password.", "warning")
        return redirect(url_for("admin.users"))

    if request.method == "POST":
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        errors = []
        if not new_password:
            errors.append("New password is required.")
        elif len(new_password) < 6:
            errors.append("Password must be at least 6 characters.")
        if new_password and new_password != confirm_password:
            errors.append("Passwords do not match.")

        if errors:
            for err in errors:
                flash(err, "danger")
            return render_template("admin/reset_user_password.html", target=target)

        target.set_password(new_password)
        target.must_change_password = True
        db.session.flush()

        AuditLog.log(
            action="USER_PASSWORD_RESET",
            user_id=current_user.id,
            entity_type="User",
            entity_id=str(target.id),
            details=(
                f"Admin '{current_user.username}' reset password for user '{target.username}'. "
                f"must_change_password set to True."
            ),
            ip_address=_client_ip(),
            user_agent=get_user_agent(),
        )

        log_activity(current_user.username, "password_reset", "user",
                     target.username)
        db.session.commit()

        flash(
            f"Password for '{target.username}' has been reset. "
            "They will be required to change it on next login.",
            "success",
        )
        return redirect(url_for("admin.users"))

    return render_template("admin/reset_user_password.html", target=target)


# ── Toggle Active ────────────────────────────────────────────────
@admin_bp.route("/users/<int:user_id>/toggle-active", methods=["POST"])
@login_required
@roles_required(ADMIN_ROLE)
def toggle_user_active(user_id):
    target = User.query.get_or_404(user_id)

    if target.id == current_user.id:
        flash("You cannot deactivate your own account.", "warning")
        return redirect(url_for("admin.users"))

    if target.is_active:
        target.is_active = False
        if target.is_power_user:
            target.is_power_user = False
        action = "USER_DEACTIVATED"
        verb = "deactivated"
    else:
        target.is_active = True
        action = "USER_ACTIVATED"
        verb = "activated"

    db.session.flush()

    AuditLog.log(
        action=action,
        user_id=current_user.id,
        entity_type="User",
        entity_id=str(target.id),
        details=f"Admin '{current_user.username}' {verb} user '{target.username}'.",
        ip_address=_client_ip(),
        user_agent=get_user_agent(),
    )
    invalidate_user_module_access_cache(target.id)

    activity_action = "user_deactivated" if verb == "deactivated" else "user_activated"
    log_activity(current_user.username, activity_action, "user", target.username)
    db.session.commit()

    flash(f"User '{target.username}' has been {verb}.", "success")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
@roles_required(ADMIN_ROLE)
def delete_user(user_id):
    target = User.query.get_or_404(user_id)

    if target.id == current_user.id:
        flash("You cannot delete your own account.", "warning")
        return redirect(url_for("admin.users"))

    if target.is_active:
        flash("Deactivate the user before deleting the account.", "warning")
        return redirect(url_for("admin.users"))

    target_username = target.username

    try:
        dependency_changes = []

        affected = User.query.filter_by(controlling_officer_id=target.id).update(
            {User.controlling_officer_id: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "controlling-officer mapping cleared", affected)

        affected = User.query.filter_by(reviewing_officer_id=target.id).update(
            {User.reviewing_officer_id: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "reviewing-officer mapping cleared", affected)

        affected = User.query.filter_by(accepting_officer_id=target.id).update(
            {User.accepting_officer_id: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "accepting-officer mapping cleared", affected)

        affected = Task.query.filter_by(owner_id=target.id).update(
            {Task.owner_id: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "task owner reference cleared", affected)

        affected = Task.query.filter_by(created_by=target.id).update(
            {Task.created_by: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "task creator reference cleared", affected)

        affected = RecurringTaskTemplate.query.filter_by(owner_id=target.id).update(
            {RecurringTaskTemplate.owner_id: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "recurring-task owner reference cleared", affected)

        affected = RecurringTaskTemplate.query.filter_by(created_by=target.id).update(
            {RecurringTaskTemplate.created_by: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "recurring-task creator reference cleared", affected)

        affected = TaskUpdate.query.filter_by(updated_by=target.id).update(
            {TaskUpdate.updated_by: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "task update author reference cleared", affected)

        affected = TaskCollaborator.query.filter_by(user_id=target.id).delete(
            synchronize_session=False
        )
        _record_dependency_change(dependency_changes, "task collaborator link removed", affected)

        affected = RecurringTaskCollaborator.query.filter_by(user_id=target.id).delete(
            synchronize_session=False
        )
        _record_dependency_change(dependency_changes, "recurring collaborator link removed", affected)

        affected = UserModulePermission.query.filter_by(user_id=target.id).delete(
            synchronize_session=False
        )
        _record_dependency_change(dependency_changes, "module permission removed", affected)

        affected = ModuleAdminAssignment.query.filter_by(user_id=target.id).delete(
            synchronize_session=False
        )
        _record_dependency_change(dependency_changes, "module-admin assignment removed", affected)

        affected = Notification.query.filter_by(user_id=target.id).delete(
            synchronize_session=False
        )
        _record_dependency_change(dependency_changes, "notification removed", affected)

        affected = AnnouncementRecipient.query.filter_by(user_id=target.id).delete(
            synchronize_session=False
        )
        _record_dependency_change(dependency_changes, "announcement receipt removed", affected)

        affected = AnnouncementVote.query.filter_by(user_id=target.id).delete(
            synchronize_session=False
        )
        _record_dependency_change(dependency_changes, "announcement vote removed", affected)

        created_announcements = Announcement.query.filter_by(created_by=target.id).all()
        if created_announcements:
            for announcement in created_announcements:
                db.session.delete(announcement)
            _record_dependency_change(
                dependency_changes,
                "announcement created by user removed",
                len(created_announcements),
            )

        affected = CSCDraft.query.filter_by(created_by_id=target.id).update(
            {CSCDraft.created_by_id: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "CSC draft creator reference cleared", affected)

        affected = CSCRevision.query.filter_by(committee_head_user_id=target.id).update(
            {CSCRevision.committee_head_user_id: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "CSC committee-head reference cleared", affected)

        affected = CSCRevision.query.filter_by(module_admin_user_id=target.id).update(
            {CSCRevision.module_admin_user_id: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "CSC module-admin reference cleared", affected)

        affected = InventoryUpload.query.filter_by(uploaded_by=target.id).update(
            {InventoryUpload.uploaded_by: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "inventory upload reference cleared", affected)

        affected = MaterialMaster.query.filter_by(updated_by=target.id).update(
            {MaterialMaster.updated_by: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "material-master editor reference cleared", affected)

        affected = InventoryConsumptionSeed.query.filter_by(imported_by=target.id).update(
            {InventoryConsumptionSeed.imported_by: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "inventory consumption import reference cleared", affected)

        affected = InventoryProcurementSeed.query.filter_by(imported_by=target.id).update(
            {InventoryProcurementSeed.imported_by: None},
            synchronize_session=False,
        )
        _record_dependency_change(dependency_changes, "inventory procurement import reference cleared", affected)

        db.session.delete(target)
        db.session.flush()

        db.session.add(
            AuditLog(
                action="USER_DELETED",
                user_id=current_user.id,
                entity_type="User",
                entity_id=str(user_id),
                details=(
                    f"Admin '{current_user.username}' deleted user "
                    f"'{target_username}' after deactivation. "
                    f"Affected dependencies: {'; '.join(dependency_changes) if dependency_changes else 'none'}"
                ),
                ip_address=AuditLog._normalize_ip(_client_ip()),
                user_agent=AuditLog._normalize_user_agent(get_user_agent()),
            )
        )
        log_activity(current_user.username, "user_deleted", "user", target_username)
        invalidate_user_module_access_cache(user_id)
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        logger.exception("Failed to delete user '%s' due to dependent database rows.", target_username)
        flash(f"Could not delete user '{target_username}' due to a database error.", "danger")
        return redirect(url_for("admin.users"))

    flash(f"User '{target_username}' has been deleted.", "success")
    if dependency_changes:
        flash(
            "Dependencies cleared before deletion: " + "; ".join(dependency_changes) + ".",
            "info",
        )
    return redirect(url_for("admin.users"))


# ── Activity History ────────────────────────────────────────────
@admin_bp.route("/activity")
@login_required
@roles_required(ADMIN_ROLE, SUPERUSER_ROLE)
def activity_history():
    page = request.args.get("page", 1, type=int)
    per_page = 30
    action_filter = request.args.get("action", "").strip()
    actor_filter = request.args.get("actor", "").strip()

    query = ActivityLog.query

    if action_filter:
        query = query.filter(ActivityLog.action_type == action_filter)
    if actor_filter:
        query = query.filter(ActivityLog.actor_username == actor_filter)

    pagination = (
        query
        .order_by(ActivityLog.created_at.desc())
        .paginate(page=page, per_page=per_page, error_out=False)
    )

    # Distinct action types and actors for filter dropdowns
    action_types = [
        r[0] for r in
        db.session.query(ActivityLog.action_type)
        .distinct()
        .order_by(ActivityLog.action_type)
        .all()
    ]
    actors = [
        r[0] for r in
        db.session.query(ActivityLog.actor_username)
        .filter(ActivityLog.actor_username.isnot(None))
        .distinct()
        .order_by(ActivityLog.actor_username)
        .all()
    ]

    return render_template(
        "admin/activity_history.html",
        pagination=pagination,
        activities=pagination.items,
        action_types=action_types,
        actors=actors,
        filters={"action": action_filter, "actor": actor_filter},
    )


@admin_bp.route("/backups")
@login_required
@roles_required(ADMIN_ROLE)
def backup_center():
    snapshots = []
    backup_events = []
    history_available = True
    try:
        snapshots = (
            BackupSnapshot.query
            .order_by(BackupSnapshot.created_at.desc())
            .limit(20)
            .all()
        )
        backup_events = (
            AuditLog.query
            .filter(
                AuditLog.action.in_([
                    "DATABASE_BACKUP_EXPORTED",
                    "DATABASE_BACKUP_IMPORTED",
                ])
            )
            .order_by(AuditLog.created_at.desc())
            .limit(20)
            .all()
        )
    except SQLAlchemyError:
        db.session.rollback()
        history_available = False

    return render_template(
        "admin/backups.html",
        snapshots=snapshots,
        backup_events=backup_events,
        history_available=history_available,
        environment_name=get_runtime_environment_name(),
        next_backup_filename=build_backup_filename(),
        restore_phrase="RESTORE BACKUP",
    )


@admin_bp.route("/backups/export", methods=["POST"])
@login_required
@roles_required(ADMIN_ROLE)
def export_backup():
    try:
        artifact = create_full_backup_bundle()
    except BackupError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("admin.backup_center"))

    try:
        db.session.add(
            BackupSnapshot(
                filename=artifact.download_name,
                created_by_username=current_user.username,
                environment=get_runtime_environment_name(),
                notes="Generated full backup bundle from the admin Backup Center and streamed immediately.",
            )
        )
        db.session.add(
            AuditLog(
                action="DATABASE_BACKUP_EXPORTED",
                user_id=current_user.id,
                entity_type="BackupSnapshot",
                entity_id=artifact.download_name,
                details=(
                    f"Admin '{current_user.username}' exported full backup bundle "
                    f"'{artifact.download_name}'."
                ),
                ip_address=AuditLog._normalize_ip(_client_ip()),
                user_agent=AuditLog._normalize_user_agent(get_user_agent()),
            )
        )
        log_activity(
            current_user.username,
            "backup_exported",
            "backup",
            artifact.download_name,
            details=f"environment={get_runtime_environment_name()}",
        )
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()

    response = send_file(
        artifact.temp_path,
        as_attachment=True,
        download_name=artifact.download_name,
        mimetype="application/x-tar",
        max_age=0,
    )
    response.call_on_close(artifact.cleanup)
    return response


@admin_bp.route("/backups/import", methods=["POST"])
@login_required
@roles_required(ADMIN_ROLE)
def import_backup():
    """Ingest/Restore a full backup bundle or legacy SQL backup file."""
    environment_name = get_runtime_environment_name()
    required_restore_phrase = "RESTORE BACKUP"

    if "backup_file" not in request.files:
        flash("No file part provided.", "danger")
        return redirect(url_for("admin.backup_center"))

    file = request.files["backup_file"]
    if not file or not file.filename:
        flash("No file selected.", "danger")
        return redirect(url_for("admin.backup_center"))

    if request.form.get("ack_downloaded_backup") != "on":
        flash("Confirm that you downloaded a fresh backup of the current database before ingesting a new one.", "danger")
        return redirect(url_for("admin.backup_center"))

    if request.form.get("ack_destructive_restore") != "on":
        flash("Confirm that you understand the ingest will overwrite the current database.", "danger")
        return redirect(url_for("admin.backup_center"))

    confirm_environment = (request.form.get("confirm_environment") or "").strip()
    if confirm_environment.casefold() != environment_name.casefold():
        flash(
            f"Environment confirmation failed. Type the current environment name exactly: {environment_name}.",
            "danger",
        )
        return redirect(url_for("admin.backup_center"))

    confirm_phrase = (request.form.get("confirm_phrase") or "").strip()
    if confirm_phrase.upper() != required_restore_phrase:
        flash(
            f"Restore confirmation phrase mismatch. Type '{required_restore_phrase}' to continue.",
            "danger",
        )
        return redirect(url_for("admin.backup_center"))

    # ── Temporary storage for the upload ──────────────────────────
    suffix = "".join(Path(file.filename).suffixes) or ".upload"
    fd, temp_path = tempfile.mkstemp(prefix="restore-upload-", suffix=suffix)
    try:
        with os.fdopen(fd, "wb") as f:
            file.save(f)

        validation = validate_backup_file(temp_path)

        # ── Trigger restore ───────────────────────────────────────
        result = restore_database_backup(temp_path)

        # ── Audit log ──────────────────────────────────────────────
        AuditLog.log(
            action="DATABASE_BACKUP_IMPORTED",
            user_id=current_user.id,
            entity_type="BackupRestore",
            entity_id=file.filename,
            details=(
                f"Admin '{current_user.username}' restored database from "
                f"'{file.filename}' into {result['database']} at {result['host']} "
                f"using {result.get('format', 'sql')} backup format."
            ),
            ip_address=_client_ip(),
            user_agent=get_user_agent(),
        )
        log_activity(
            current_user.username,
            "backup_restored",
            "backup",
            file.filename,
            details=f"database={result['database']}",
        )
        db.session.commit()

        if result.get("format") == "bundle":
            flash(
                f"Full backup bundle '{file.filename}' restored successfully. "
                f"Database tables were refreshed and committee attachments were synchronized for {environment_name}.",
                "success",
            )
        else:
            flash(
                f"Legacy SQL backup '{file.filename}' restored successfully "
                f"({validation['size_bytes']} bytes). Database tables were refreshed for {environment_name}. "
                "Filesystem attachments were not changed.",
                "success",
            )
    except BackupError as exc:
        flash(f"Restore failed: {exc}", "danger")
    except Exception as exc:
        flash(f"An unexpected error occurred during restore: {exc}", "danger")
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

    return redirect(url_for("admin.backup_center"))


# ══════════════════════════════════════════════════════════════════
#  OFFICE MANAGEMENT
# ══════════════════════════════════════════════════════════════════

def _require_office_management():
    """Abort 403 if current_user cannot manage offices."""
    if not can_manage_offices(current_user):
        from flask import abort
        abort(403)


# ── List Offices ─────────────────────────────────────────────────
@admin_bp.route("/offices")
@admin_bp.route("/offices/")
@login_required
@roles_required(ADMIN_ROLE)
def offices():
    _require_office_management()
    all_offices = Office.query.order_by(Office.office_name).all()
    return render_template("admin/offices.html", offices=all_offices)


# ── Add Office ───────────────────────────────────────────────────
@admin_bp.route("/offices/add", methods=["GET", "POST"])
@login_required
@roles_required(ADMIN_ROLE)
def add_office():
    _require_office_management()

    if request.method == "POST":
        office_code = request.form.get("office_code", "").strip().upper()
        office_name = request.form.get("office_name", "").strip()
        location = request.form.get("location", "").strip()
        power_user_limit_raw = request.form.get("power_user_limit", "").strip()
        power_user_limit, limit_error = _parse_power_user_limit(power_user_limit_raw)

        errors = []
        if not office_code:
            errors.append("Office code is required.")
        elif len(office_code) > 50:
            errors.append("Office code cannot exceed 50 characters.")
        if not office_name:
            errors.append("Office name is required.")
        elif len(office_name) > 150:
            errors.append("Office name cannot exceed 150 characters.")
        if len(location) > 150:
            errors.append("Location cannot exceed 150 characters.")
        if limit_error:
            errors.append(limit_error)

        if office_code and Office.query.filter_by(office_code=office_code).first():
            errors.append(f"Office code '{office_code}' already exists.")

        if errors:
            for err in errors:
                flash(err, "danger")
            return render_template(
                "admin/office_form.html",
                mode="add",
                form_data=request.form,
            )

        try:
            new_office = Office(
                office_code=office_code,
                office_name=office_name,
                location=location,
                power_user_limit=power_user_limit,
                is_active=True,
            )
            db.session.add(new_office)
            db.session.flush()

            AuditLog.log(
                action="OFFICE_CREATED",
                user_id=current_user.id,
                entity_type="Office",
                entity_id=str(new_office.id),
                details=(
                    f"Admin '{current_user.username}' created office "
                    f"'{office_name}' ({office_code}) with {power_user_limit} power-user slot(s)."
                ),
                ip_address=_client_ip(),
                user_agent=get_user_agent(),
            )
            log_activity(
                current_user.username,
                "office_created",
                "office",
                office_name,
                details=(
                    f"code={office_code}, location={location or '-'}, "
                    f"power_user_limit={power_user_limit}"
                ),
            )
            db.session.commit()
        except SQLAlchemyError:
            db.session.rollback()
            flash("Could not create office due to a database error.", "danger")
            return render_template(
                "admin/office_form.html",
                mode="add",
                form_data=request.form,
                power_user_candidates=[],
                selected_power_user_ids=[],
            )

        flash(f"Office '{office_name}' created successfully.", "success")
        return redirect(url_for("admin.offices"))

    return render_template(
        "admin/office_form.html",
        mode="add",
        form_data={},
        power_user_candidates=[],
        selected_power_user_ids=[],
    )


# ── Edit Office ──────────────────────────────────────────────────
@admin_bp.route("/offices/<int:office_id>/edit", methods=["GET", "POST"])
@login_required
@roles_required(ADMIN_ROLE)
def edit_office(office_id):
    _require_office_management()
    office = Office.query.get_or_404(office_id)
    office_users = (
        User.query
        .options(joinedload(User.role))
        .filter(User.office_id == office.id)
        .order_by(User.full_name, User.username)
        .all()
    )
    power_user_candidates = _power_user_candidates_for_office(office.id)
    current_power_user_ids = [
        str(user.id)
        for user in office_users
        if user.is_power_user
    ]

    if request.method == "POST":
        office_name = request.form.get("office_name", "").strip()
        location = request.form.get("location", "").strip()
        power_user_limit_raw = request.form.get("power_user_limit", "").strip()
        selected_power_user_ids, selection_errors = _resolve_power_user_selection(
            office.id,
            request.form.getlist("power_user_ids"),
        )
        power_user_limit, limit_error = _parse_power_user_limit(power_user_limit_raw)

        errors = []
        if not office_name:
            errors.append("Office name is required.")
        elif len(office_name) > 150:
            errors.append("Office name cannot exceed 150 characters.")
        if len(location) > 150:
            errors.append("Location cannot exceed 150 characters.")
        if limit_error:
            errors.append(limit_error)
        errors.extend(selection_errors)
        if len(selected_power_user_ids) > power_user_limit:
            errors.append(
                f"Select at most {power_user_limit} power user"
                f"{'' if power_user_limit == 1 else 's'} for this office."
            )

        if errors:
            for err in errors:
                flash(err, "danger")
            return render_template(
                "admin/office_form.html",
                mode="edit",
                office=office,
                form_data=request.form,
                power_user_candidates=power_user_candidates,
                selected_power_user_ids=[str(user_id) for user_id in selected_power_user_ids],
            )

        changed_fields = []
        if office.office_name != office_name:
            changed_fields.append(f"name '{office.office_name}' → '{office_name}'")
            office.office_name = office_name
        if (office.location or "") != location:
            changed_fields.append(f"location '{office.location or '-'}' → '{location or '-'}'")
            office.location = location
        if office.power_user_limit != power_user_limit:
            changed_fields.append(
                f"power_user_limit '{office.power_user_limit}' → '{power_user_limit}'"
            )
            office.power_user_limit = power_user_limit

        previous_power_user_ids = {user.id for user in office_users if user.is_power_user}
        next_power_user_ids = set(selected_power_user_ids)
        if previous_power_user_ids != next_power_user_ids:
            previous_labels = [
                _display_user_name(user)
                for user in office_users
                if user.id in previous_power_user_ids
            ]
            next_labels = [
                _display_user_name(user)
                for user in office_users
                if user.id in next_power_user_ids
            ]
            changed_fields.append(
                "power_users "
                f"'{', '.join(previous_labels) or '-'}' → "
                f"'{', '.join(next_labels) or '-'}'"
            )

        for office_user in office_users:
            office_user.is_power_user = office_user.id in next_power_user_ids

        try:
            db.session.flush()

            AuditLog.log(
                action="OFFICE_UPDATED",
                user_id=current_user.id,
                entity_type="Office",
                entity_id=str(office.id),
                details=(
                    f"Admin '{current_user.username}' updated office '{office.office_code}'. "
                    f"Changes: {'; '.join(changed_fields) if changed_fields else 'none'}"
                ),
                ip_address=_client_ip(),
                user_agent=get_user_agent(),
            )
            if changed_fields:
                log_activity(
                    current_user.username,
                    "office_updated",
                    "office",
                    office.office_name,
                    details="; ".join(changed_fields),
                )
            db.session.commit()
        except SQLAlchemyError:
            db.session.rollback()
            flash("Could not update office due to a database error.", "danger")
            return render_template(
                "admin/office_form.html",
                mode="edit",
                office=office,
                form_data=request.form,
                power_user_candidates=power_user_candidates,
                selected_power_user_ids=[str(user_id) for user_id in selected_power_user_ids],
            )

        flash(f"Office '{office.office_name}' updated successfully.", "success")
        return redirect(url_for("admin.offices"))

    return render_template(
        "admin/office_form.html",
        mode="edit",
        office=office,
        form_data={},
        power_user_candidates=power_user_candidates,
        selected_power_user_ids=current_power_user_ids,
    )


# ── Toggle Office Active/Inactive ───────────────────────────────
@admin_bp.route("/offices/<int:office_id>/toggle", methods=["POST"])
@login_required
@roles_required(ADMIN_ROLE)
def toggle_office(office_id):
    _require_office_management()
    office = Office.query.get_or_404(office_id)

    if office.is_active:
        active_user_count = User.query.filter_by(
            office_id=office.id, is_active=True
        ).count()
        if active_user_count > 0:
            flash(
                f"Cannot deactivate '{office.office_name}' — "
                f"{active_user_count} active user(s) are assigned to it. "
                "Reassign them first.",
                "danger",
            )
            return redirect(url_for("admin.offices"))

        office.is_active = False
        action = "OFFICE_DEACTIVATED"
        verb = "deactivated"
    else:
        office.is_active = True
        action = "OFFICE_ACTIVATED"
        verb = "activated"

    try:
        db.session.flush()
        AuditLog.log(
            action=action,
            user_id=current_user.id,
            entity_type="Office",
            entity_id=str(office.id),
            details=(
                f"Admin '{current_user.username}' {verb} office "
                f"'{office.office_name}' ({office.office_code})."
            ),
            ip_address=_client_ip(),
            user_agent=get_user_agent(),
        )
        log_activity(
            current_user.username,
            f"office_{verb}",
            "office",
            office.office_name,
        )
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        flash(f"Could not {verb[:-1]}e office due to a database error.", "danger")
        return redirect(url_for("admin.offices"))

    flash(f"Office '{office.office_name}' has been {verb}.", "success")
    return redirect(url_for("admin.offices"))


# ── Users in Office ──────────────────────────────────────────────
@admin_bp.route("/offices/<int:office_id>/users")
@login_required
@roles_required(ADMIN_ROLE)
def office_users(office_id):
    _require_office_management()
    office = Office.query.get_or_404(office_id)
    users_in_office = (
        User.query
        .options(
            joinedload(User.role),
            joinedload(User.controlling_officer),
            joinedload(User.reviewing_officer),
            joinedload(User.accepting_officer),
        )
        .filter_by(office_id=office.id)
        .order_by(User.is_power_user.desc(), User.is_active.desc(), User.full_name, User.username)
        .all()
    )
    return render_template(
        "admin/office_users.html",
        office=office,
        users=users_in_office,
    )


# ── Assign User to Office ───────────────────────────────────────
@admin_bp.route("/users/<int:user_id>/assign-office", methods=["POST"])
@login_required
@roles_required(ADMIN_ROLE)
def assign_user_office(user_id):
    _require_office_management()
    target = User.query.get_or_404(user_id)
    office_id_raw = request.form.get("office_id", "").strip()

    if not office_id_raw or not office_id_raw.isdigit():
        flash("Please select a valid office.", "danger")
        return redirect(url_for("admin.edit_user", user_id=user_id))

    office = Office.query.filter_by(id=int(office_id_raw), is_active=True).first()
    if not office:
        flash("Selected office is invalid or inactive.", "danger")
        return redirect(url_for("admin.edit_user", user_id=user_id))

    old_office_name = target.office.office_name if target.office else "None"
    target.office_id = office.id

    try:
        db.session.flush()
        AuditLog.log(
            action="USER_OFFICE_ASSIGNED",
            user_id=current_user.id,
            entity_type="User",
            entity_id=str(target.id),
            details=(
                f"Admin '{current_user.username}' reassigned '{target.username}' "
                f"from '{old_office_name}' to '{office.office_name}'."
            ),
            ip_address=_client_ip(),
            user_agent=get_user_agent(),
        )
        log_activity(
            current_user.username,
            "user_office_assigned",
            "user",
            target.username,
            details=f"office={office.office_name}",
        )
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        flash("Could not assign office due to a database error.", "danger")
        return redirect(url_for("admin.edit_user", user_id=user_id))

    flash(
        f"'{target.username}' has been assigned to '{office.office_name}'.",
        "success",
    )
    return redirect(url_for("admin.edit_user", user_id=user_id))


# ── Set Officers (Controlling / Reviewing / Accepting) ───────────
@admin_bp.route("/users/<int:user_id>/set-officers", methods=["POST"])
@login_required
@roles_required(ADMIN_ROLE)
def set_user_officers(user_id):
    _require_office_management()
    target = User.query.get_or_404(user_id)

    co_raw = request.form.get("controlling_officer_id", "").strip()
    ro_raw = request.form.get("reviewing_officer_id", "").strip()
    ao_raw = request.form.get("accepting_officer_id", "").strip()

    co, co_error = _resolve_officer_selection(co_raw, "controlling officer", target_user_id=target.id)
    ro, ro_error = _resolve_officer_selection(ro_raw, "reviewing officer", target_user_id=target.id)
    ao, ao_error = _resolve_officer_selection(ao_raw, "accepting officer", target_user_id=target.id)

    for error in (co_error, ro_error, ao_error):
        if error:
            flash(error, "danger")
            return redirect(url_for("admin.edit_user", user_id=user_id))

    changed = []
    old_co = target.controlling_officer_id
    old_ro = target.reviewing_officer_id
    old_ao = target.accepting_officer_id

    target.controlling_officer_id = co.id if co else None
    target.reviewing_officer_id = ro.id if ro else None
    target.accepting_officer_id = ao.id if ao else None

    if old_co != target.controlling_officer_id:
        changed.append(f"controlling_officer: {old_co} → {target.controlling_officer_id}")
    if old_ro != target.reviewing_officer_id:
        changed.append(f"reviewing_officer: {old_ro} → {target.reviewing_officer_id}")
    if old_ao != target.accepting_officer_id:
        changed.append(f"accepting_officer: {old_ao} → {target.accepting_officer_id}")

    if not changed:
        flash("No changes to officer assignments.", "info")
        return redirect(url_for("admin.edit_user", user_id=user_id))

    try:
        db.session.flush()
        AuditLog.log(
            action="USER_OFFICERS_UPDATED",
            user_id=current_user.id,
            entity_type="User",
            entity_id=str(target.id),
            details=(
                f"Admin '{current_user.username}' updated officers for "
                f"'{target.username}'. Changes: {'; '.join(changed)}"
            ),
            ip_address=_client_ip(),
            user_agent=get_user_agent(),
        )
        log_activity(
            current_user.username,
            "user_officers_updated",
            "user",
            target.username,
            details="; ".join(changed),
        )
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        flash("Could not update officer assignments due to a database error.", "danger")
        return redirect(url_for("admin.edit_user", user_id=user_id))

    flash(
        f"Officer assignments for '{target.username}' updated successfully.",
        "success",
    )
    return redirect(url_for("admin.edit_user", user_id=user_id))
