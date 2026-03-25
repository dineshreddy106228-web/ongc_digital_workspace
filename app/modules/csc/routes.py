"""CSC (Chemical Specification Committee) Workflow Module – Routes.

Endpoints:
    GET  /csc/                                  → CSC landing page with KPI cards
    GET  /csc/api/overview                      → JSON: KPI data
    GET  /csc/workspace                         → Draft selection/list page
    GET  /csc/api/drafts                        → JSON: filtered draft list
    GET  /csc/workspace/<int:draft_id>          → 9-section tabbed editor
    GET  /csc/api/draft/<int:draft_id>          → JSON: full draft data
    POST /csc/api/draft/<int:draft_id>/section/<section_name>   → Save section text
    POST /csc/api/draft/<int:draft_id>/issues   → Save issue flags
    POST /csc/api/draft/<int:draft_id>/parameters              → Save parameters (Phase 1)
    POST /csc/api/draft/<int:draft_id>/parameters/proposed     → Save proposed values (Phase 2)
    POST /csc/api/draft/<int:draft_id>/impact  → Save impact analysis
    POST /csc/api/draft/<int:draft_id>/flag   → Update single impact flag (4-state AJAX)
    POST /csc/api/draft/<int:draft_id>/submit  → Submit revision for approval
    POST /csc/api/draft/<int:draft_id>/delete  → Delete draft
    GET  /csc/draft/<int:draft_id>/export/docx → Download Word document
    POST /csc/workspace/new-revision/<int:parent_id> → Create revision draft
    GET  /csc/admin                             → Admin dashboard [superuser]
    GET  /csc/admin/ingest                      → PDF/DOCX ingest page [superuser]
    POST /csc/admin/ingest/pdf                  → Upload PDF [superuser]
    POST /csc/admin/ingest/docx                 → Upload DOCX [superuser]
    POST /csc/admin/ingest/create-draft         → Create draft from extraction [superuser]
    GET  /csc/admin/drafts                      → Manage all drafts [superuser]
    GET  /csc/admin/draft/<int:draft_id>/edit   → Edit draft [superuser]
    POST /csc/admin/draft/<int:draft_id>/update → Update draft metadata [superuser]
    POST /csc/admin/draft/<int:draft_id>/lock-phase1 → Toggle Phase 1 lock [superuser]
    POST /csc/admin/draft/<int:draft_id>/set-stage → Set admin stage [superuser]
    DELETE /csc/admin/draft/<int:draft_id>/delete → Delete draft (admin) [superuser]
    GET  /csc/admin/revisions                   → Committee workbench [superuser]
    GET  /csc/admin/revision/<int:revision_id>/review → Review revision [superuser]
    POST /csc/admin/revision/<int:revision_id>/approve  → Approve revision [superuser]
    POST /csc/admin/revision/<int:revision_id>/reject   → Reject revision [superuser]
    GET  /csc/admin/export                      → Export page [superuser]
    GET  /csc/admin/export/master-docx          → Download master spec document [superuser]
    GET  /csc/admin/export/<int:export_id>/download → Download export file [superuser]
"""

from __future__ import annotations

from dataclasses import asdict
import json
import logging
import io
import os
import re
import shutil
import tempfile
import uuid
from pathlib import Path
from datetime import datetime, timedelta, timezone
from typing import Optional
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.exc import OperationalError, ProgrammingError

from flask import (
    abort,
    current_app,
    flash,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import login_required, current_user

from app.modules.csc import csc_bp
from app.extensions import db
from app.core.utils.decorators import module_access_required, module_admin_required
from app.models.csc import (
    CSCDraft,
    CSCParameter,
    CSCSection,
    CSCIssueFlag,
    CSCImpactAnalysis,
    CSCAudit,
    CSCRevision,
    CSCSpecVersion,
)
from app.models.core.user import User
from app.core.services.notifications import create_notification
from app.core.services.csc_utils import (
    ADMIN_STAGE_DRAFTING,
    ADMIN_STAGE_PUBLISHED,
    ALL_FLAG_IDS,
    VALID_FLAG_VALUES,
    IMPACT_CHECKLIST_FLAGS,
    SPEC_SUBSET_LABELS,
    SPEC_SUBSET_ORDER,
    build_default_impact_checklist_state,
    build_impact_legacy_payload,
    compute_impact_classification,
    deserialize_impact_checklist_state,
    format_required_value,
    format_spec_version,
    increment_spec_version,
    normalize_spec_version,
    normalize_parameter_type_label,
    normalize_test_procedure_type,
    parse_spec_number,
    spec_sort_key,
    summarize_impact_checklist_state,
    PARAMETER_TYPES,
    TEST_PROCEDURE_OPTIONS,
    WORKFLOW_DRAFTING_APPROVED,
    WORKFLOW_DRAFTING_STAGING,
    WORKFLOW_DRAFTING_HEAD_APPROVED,
    WORKFLOW_DRAFTING_OPEN,
    WORKFLOW_DRAFTING_REJECTED,
    WORKFLOW_DRAFTING_RETURNED,
    WORKFLOW_DRAFTING_STATE_OPTIONS,
    WORKFLOW_DRAFTING_SUBMITTED,
)
from app.core.services.csc_master_data import (
    ADMIN_MASTER_EXTRA_FIELDS,
    ADMIN_MASTER_FIELDS,
    ADMIN_MASTER_FIELD_CONFIGS,
    CSC_ADMIN_DRAFT_FIELDS,
    MASTER_DATA_FIELDS,
    MASTER_EXTRA_FIELDS,
    MASTER_IMPACT_FIELDS,
    STAGED_MASTER_SECTION_NAME,
    _load_staged_master_payload,
    _write_staged_master_payload,
    backfill_draft_material_codes,
    clear_staged_master_payload,
    get_master_record_for_draft,
    get_material_properties_fields,
    get_material_properties_values,
    get_storage_handling_fields,
    get_storage_handling_values,
    get_master_form_values,
    sync_impact_fields_to_master_record,
    update_material_properties_from_form,
    update_storage_handling_from_form,
    upsert_master_record_from_form,
)
from app.core.services.csc_committee_directory import (
    get_committee_access_for_username,
    get_committee_config_payload,
    get_committee_directory,
    get_committee_tree,
    get_office_orders,
    normalize_committee_config_payload,
)
from app.core.services.csc_pdf_extractor import (
    ExtractedParameter,
    SpecDocument,
    extract_spec_from_pdf,
)
from app.core.services.csc_docx_extractor import extract_all_specs_from_docx
from app.core.services.csc_export import (
    build_flask_review_document,
    build_master_spec_document,
)
from app.core.services.csc_type_classification_import import (
    MATERIAL_HANDLING_BACKGROUND_END,
    MATERIAL_HANDLING_BACKGROUND_START,
    MATERIAL_HANDLING_JUSTIFICATION_END,
    MATERIAL_HANDLING_JUSTIFICATION_START,
    MATERIAL_HANDLING_STREAM,
    TYPE_CLASSIFICATION_STREAM,
    MaterialHandlingImportSummary,
    MaterialHandlingRowIssue,
    MaterialHandlingWorkbookRow,
    ParsedMaterialHandlingWorkbook,
    TypeClassificationImportSummary,
    apply_material_handling_core_updates,
    build_material_handling_background_text,
    build_material_handling_justification_text,
    build_parent_lookup_by_material_code,
    canonicalize_spec_number,
    import_type_classification_workbook,
    import_material_handling_workbook,
    material_code_lookup_key,
    merge_material_handling_section_block,
    parse_type_classification_workbook,
    parse_material_handling_workbook,
)
from app.core.services.master_data import get_all_master_data

logger = logging.getLogger(__name__)


def _batch_fetch_users_by_usernames(usernames: list[str]) -> dict[str, User]:
    """Return {username: User} for active users in one query."""
    clean = [u for u in (name.strip() for name in usernames) if u]
    if not clean:
        return {}
    users = User.query.filter(
        User.username.in_(clean), User.is_active.is_(True)
    ).all()
    return {u.username: u for u in users}


WORKFLOW_SCOPE_SECTION_NAME = "__workflow_scope_json__"
WORKFLOW_COMMITTEE_SECTION_NAME = "__workflow_committee_slug__"
WORKFLOW_STREAM_SECTION_NAME = "__workflow_stream_name__"
DRAFT_ORIGIN_SECTION_NAME = "__draft_origin_json__"
MATERIAL_HANDLING_IMPORT_STAGING_ROOT = os.path.join(
    tempfile.gettempdir(),
    "csc_material_handling_import_staging",
)
MATERIAL_HANDLING_IMPORT_REPORT_ROOT = os.path.join(
    tempfile.gettempdir(),
    "csc_material_handling_import_reports",
)
SYSTEM_BRACKET_LINE_RE = re.compile(r"^\s*\[[^\n\]]+\]\s*$", re.MULTILINE)
WORKFLOW_TRACK_SUBSET = "subset"
WORKFLOW_TRACK_MATERIAL_HANDLING = "material_handling"
WORKFLOW_SCOPE_DEFAULT = {
    "material_properties": True,
    "storage_handling": True,
    "parameters": True,
    "impact": True,
    "parameter_type": True,
    "parameter_other": True,
}
EDITOR_ISSUE_LABELS = [
    "Specification Ambiguity",
    "Parameter Non-compliance",
    "Test Procedure Gaps",
    "Safety/Environmental Concern",
]
EDITOR_LOCK_SECTION_NAME = "__editor_lock_json__"
EDITOR_LOCK_TIMEOUT = timedelta(minutes=3)


def _coerce_admin_boolean(value) -> bool:
    """Coerce HTML/JSON boolean-like input into a real bool."""
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _get_revision_for_child(child_draft_id: int | None) -> CSCRevision | None:
    if not child_draft_id:
        return None
    return CSCRevision.query.filter_by(child_draft_id=child_draft_id).first()


def _get_active_revision_for_parent(parent_draft_id: int | None) -> CSCRevision | None:
    if not parent_draft_id:
        return None
    return (
        CSCRevision.query.filter_by(parent_draft_id=parent_draft_id)
        .filter(
            CSCRevision.status.in_(
                [
                    WORKFLOW_DRAFTING_STAGING,
                    WORKFLOW_DRAFTING_OPEN,
                    WORKFLOW_DRAFTING_SUBMITTED,
                    WORKFLOW_DRAFTING_HEAD_APPROVED,
                    WORKFLOW_DRAFTING_RETURNED,
                ]
            )
        )
        .order_by(CSCRevision.updated_at.desc(), CSCRevision.id.desc())
        .first()
    )


def _get_active_revisions_for_parent(parent_draft_id: int | None) -> list[CSCRevision]:
    if not parent_draft_id:
        return []
    return (
        CSCRevision.query.filter_by(parent_draft_id=parent_draft_id)
        .filter(
            CSCRevision.status.in_(
                [
                    WORKFLOW_DRAFTING_STAGING,
                    WORKFLOW_DRAFTING_OPEN,
                    WORKFLOW_DRAFTING_SUBMITTED,
                    WORKFLOW_DRAFTING_HEAD_APPROVED,
                    WORKFLOW_DRAFTING_RETURNED,
                ]
            )
        )
        .order_by(CSCRevision.updated_at.desc(), CSCRevision.id.desc())
        .all()
    )


def _is_open_revision_state(status: str | None) -> bool:
    return (status or "").strip().lower() in {
        WORKFLOW_DRAFTING_OPEN,
        WORKFLOW_DRAFTING_RETURNED,
    }


def _is_secretary_staging_state(status: str | None) -> bool:
    return (status or "").strip().lower() == WORKFLOW_DRAFTING_STAGING


def _can_submit_revision(revision: CSCRevision | None) -> bool:
    return revision is not None and _is_open_revision_state(revision.status)


def _can_user_edit_workflow_draft(draft: CSCDraft) -> bool:
    if (
        draft.parent_draft_id is not None
        and not draft.is_admin_draft
        and draft.created_by_id == current_user.id
    ):
        revision = _get_revision_for_child(draft.id)
        return revision is not None and _is_open_revision_state(revision.status)
    revision = _get_revision_for_child(draft.id)
    if revision is None:
        return False
    return _is_open_revision_state(revision.status)


def _can_current_user_edit_draft_as_committee_head(draft: CSCDraft) -> bool:
    """Allow Committee Heads to edit submitted drafts in their covered subsets."""
    if draft.parent_draft_id is None or draft.is_admin_draft:
        return False
    revision = _get_revision_for_child(draft.id)
    if revision is None or revision.status != WORKFLOW_DRAFTING_SUBMITTED:
        return False
    return _revision_in_current_user_committee_head_scope(revision)


def _notify_material_master_admins(
    title: str,
    message: str,
    link: str | None = None,
    severity: str = "info",
) -> None:
    """Notify only users explicitly assigned as CSC module admins."""
    _notify_users(
        [user.id for user in _get_material_master_admin_users()],
        title=title,
        message=message,
        severity=severity,
        link=link,
    )


def _notify_users(user_ids: list[int | None], title: str, message: str, link: str | None = None, severity: str = "info") -> None:
    """Create notifications for the provided unique user ids."""
    created = False
    seen = set()
    for user_id in user_ids:
        if not user_id or user_id in seen:
            continue
        seen.add(user_id)
        create_notification(
            user_id=user_id,
            title=title,
            message=message,
            severity=severity,
            link=link,
        )
        created = True
    if created:
        db.session.commit()


def _copy_draft_content(source: CSCDraft, target: CSCDraft) -> None:
    target.spec_number = source.spec_number
    target.chemical_name = source.chemical_name
    target.committee_name = source.committee_name
    target.meeting_date = source.meeting_date
    target.prepared_by = source.prepared_by
    target.reviewed_by = source.reviewed_by
    target.material_code = source.material_code

    target.parameters.delete()
    for parameter in source.parameters.order_by(CSCParameter.sort_order).all():
        db.session.add(
            CSCParameter(
                draft_id=target.id,
                parameter_name=parameter.parameter_name,
                parameter_type=parameter.parameter_type,
                unit_of_measure=parameter.unit_of_measure,
                existing_value=parameter.existing_value,
                proposed_value=parameter.proposed_value,
                test_method=parameter.test_method,
                test_procedure_type=parameter.test_procedure_type,
                test_procedure_text=parameter.test_procedure_text,
                parameter_conditions=parameter.parameter_conditions,
                required_value_type=parameter.required_value_type,
                required_value_text=parameter.required_value_text,
                required_value_operator_1=parameter.required_value_operator_1,
                required_value_value_1=parameter.required_value_value_1,
                required_value_operator_2=parameter.required_value_operator_2,
                required_value_value_2=parameter.required_value_value_2,
                justification=parameter.justification,
                remarks=parameter.remarks,
                sort_order=parameter.sort_order,
            )
        )

    target.sections.delete()
    for section in source.sections.order_by(CSCSection.sort_order).all():
        if section.section_name in {
            STAGED_MASTER_SECTION_NAME,
            WORKFLOW_SCOPE_SECTION_NAME,
            WORKFLOW_STREAM_SECTION_NAME,
        }:
            continue
        section_text = section.section_text or ""
        if _is_exact_testing_placeholder(section_text):
            section_text = ""
        db.session.add(
            CSCSection(
                draft_id=target.id,
                section_name=section.section_name,
                section_text=section_text,
                sort_order=section.sort_order,
            )
        )

    target.issue_flags.delete()
    for flag in source.issue_flags.order_by(CSCIssueFlag.sort_order).all():
        db.session.add(
            CSCIssueFlag(
                draft_id=target.id,
                issue_type=flag.issue_type,
                is_present=flag.is_present,
                note=flag.note,
                sort_order=flag.sort_order,
            )
        )

    source_impact = source.impact_analysis
    target_impact = target.impact_analysis
    if source_impact:
        if target_impact is None:
            target_impact = CSCImpactAnalysis(draft_id=target.id)
            db.session.add(target_impact)
        target_impact.operational_impact_score = source_impact.operational_impact_score
        target_impact.safety_environment_score = source_impact.safety_environment_score
        target_impact.supply_risk_score = source_impact.supply_risk_score
        target_impact.no_substitute_flag = source_impact.no_substitute_flag
        target_impact.impact_score_total = source_impact.impact_score_total
        target_impact.impact_grade = source_impact.impact_grade
        target_impact.operational_note = source_impact.operational_note
        target_impact.safety_environment_note = source_impact.safety_environment_note
        target_impact.supply_risk_note = source_impact.supply_risk_note
        target_impact.checklist_state_json = source_impact.checklist_state_json
    elif target_impact is not None:
        db.session.delete(target_impact)


# ──────────────────────────────────────────────────────────────────────────────
# Helper Functions
# ──────────────────────────────────────────────────────────────────────────────


def _get_committee_user_assignments(committee: dict[str, object] | None) -> list[dict[str, object]]:
    assignments = []
    seen_usernames = set()
    source_list = []
    if isinstance(committee, dict):
        source_list = committee.get("committee_users") or []
        if not source_list and str(committee.get("committee_user") or "").strip():
            source_list = [{"username": str(committee.get("committee_user") or "").strip()}]

    for entry in source_list:
        if isinstance(entry, dict):
            username = str(entry.get("username") or "").strip()
            raw_subset_codes = entry.get("subset_codes") or []
        else:
            username = str(entry or "").strip()
            raw_subset_codes = []
        username_key = username.lower()
        if not username or username_key in seen_usernames:
            continue
        seen_usernames.add(username_key)
        subset_codes = []
        for code in raw_subset_codes:
            normalized_code = str(code or "").strip().upper()
            if normalized_code and normalized_code not in subset_codes:
                subset_codes.append(normalized_code)
        assignments.append({"username": username, "subset_codes": subset_codes})
    return assignments


def _get_committee_user_assignment_for_username(
    committee: dict[str, object] | None,
    username: str | None,
) -> dict[str, object] | None:
    username_key = str(username or "").strip().lower()
    if not username_key:
        return None
    for assignment in _get_committee_user_assignments(committee):
        if str(assignment.get("username") or "").strip().lower() == username_key:
            return assignment
    return None


def _resolve_committee_user_subset_codes(
    committee: dict[str, object] | None,
    assignment: dict[str, object] | None = None,
) -> list[str]:
    if not isinstance(committee, dict):
        return []
    if str(committee.get("slug") or "").strip() == "material-handling" and assignment is not None:
        return [
            str(code or "").strip().upper()
            for code in (assignment.get("subset_codes") or [])
            if str(code or "").strip()
        ]
    return [
        str((subset or {}).get("code") or "").strip().upper()
        for subset in (committee.get("subsets") or [])
        if str((subset or {}).get("code") or "").strip()
    ]


def _parse_editor_lock_timestamp(value: str | None) -> datetime | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        parsed = datetime.fromisoformat(raw)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _get_editor_lock_section(draft: CSCDraft) -> CSCSection | None:
    if not getattr(draft, "id", None):
        return None
    return draft.sections.filter_by(section_name=EDITOR_LOCK_SECTION_NAME).first()


def _load_editor_lock_payload(draft: CSCDraft) -> dict[str, object]:
    section = _get_editor_lock_section(draft)
    if section is None or not (section.section_text or "").strip():
        return {}
    try:
        payload = json.loads(section.section_text)
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _clear_editor_lock(draft: CSCDraft) -> None:
    section = _get_editor_lock_section(draft)
    if section is not None:
        db.session.delete(section)


def _save_editor_lock_payload(draft: CSCDraft, payload: dict[str, object]) -> dict[str, object]:
    section = _get_editor_lock_section(draft)
    serialized = json.dumps(payload)
    if section is None:
        section = CSCSection(
            draft_id=draft.id,
            section_name=EDITOR_LOCK_SECTION_NAME,
            section_text=serialized,
            sort_order=-999,
        )
        db.session.add(section)
    else:
        section.section_text = serialized
    return payload


def _get_active_editor_lock_payload(draft: CSCDraft) -> dict[str, object] | None:
    payload = _load_editor_lock_payload(draft)
    if not payload:
        return None
    last_seen_at = _parse_editor_lock_timestamp(payload.get("last_seen_at"))
    if last_seen_at is None:
        return None
    if datetime.now(timezone.utc) - last_seen_at > EDITOR_LOCK_TIMEOUT:
        return None
    return payload


def _editor_lock_holder_label(payload: dict[str, object] | None) -> str:
    if not payload:
        return ""
    display_name = str(payload.get("display_name") or "").strip()
    username = str(payload.get("username") or "").strip()
    if display_name and username and display_name.lower() != username.lower():
        return f"{display_name} ({username})"
    return display_name or username


def _editor_lock_conflict_message(payload: dict[str, object] | None) -> str:
    holder_label = _editor_lock_holder_label(payload) or "another user"
    return f"{holder_label} is currently editing this draft. Please try again once they finish."


def _json_editor_lock_conflict_response(
    draft: CSCDraft,
    editor_mode: str | None = None,
):
    blocking_lock = _get_blocking_editor_lock(draft, editor_mode)
    if not blocking_lock:
        return None
    holder_label = _editor_lock_holder_label(blocking_lock) or "another user"
    return jsonify(
        {
            "error": _editor_lock_conflict_message(blocking_lock),
            "locked_by": holder_label,
            "lock_conflict": True,
        }
    ), 409


def _committee_editor_lock_enabled_for_draft(draft: CSCDraft, editor_mode: str | None = None) -> bool:
    mode = (editor_mode or "").strip().lower()
    if current_user.is_super_user():
        return False
    if mode == "committee_head":
        return False
    if draft.parent_draft_id is None or draft.is_admin_draft:
        return False
    revision = _get_revision_for_child(draft.id)
    return revision is not None and _is_open_revision_state(revision.status)


def _get_blocking_editor_lock(draft: CSCDraft, editor_mode: str | None = None) -> dict[str, object] | None:
    if not _committee_editor_lock_enabled_for_draft(draft, editor_mode):
        return None
    payload = _get_active_editor_lock_payload(draft)
    if not payload:
        return None
    if int(payload.get("user_id") or 0) == int(current_user.id):
        return None
    return payload


def _acquire_editor_lock(draft: CSCDraft) -> tuple[bool, dict[str, object] | None]:
    active_payload = _get_active_editor_lock_payload(draft)
    if active_payload and int(active_payload.get("user_id") or 0) not in {0, int(current_user.id)}:
        return False, active_payload

    now_iso = datetime.now(timezone.utc).isoformat()
    payload = {
        "user_id": int(current_user.id),
        "username": current_user.username,
        "display_name": (current_user.full_name or current_user.username or "").strip(),
        "locked_at": active_payload.get("locked_at") if active_payload else now_iso,
        "last_seen_at": now_iso,
    }
    _save_editor_lock_payload(draft, payload)
    return True, payload


def _refresh_editor_lock(draft: CSCDraft) -> tuple[bool, dict[str, object] | None]:
    active_payload = _get_active_editor_lock_payload(draft)
    if active_payload and int(active_payload.get("user_id") or 0) not in {0, int(current_user.id)}:
        return False, active_payload
    if not active_payload:
        return _acquire_editor_lock(draft)
    active_payload["last_seen_at"] = datetime.now(timezone.utc).isoformat()
    _save_editor_lock_payload(draft, active_payload)
    return True, active_payload


def _release_editor_lock_if_owned(draft: CSCDraft) -> None:
    active_payload = _get_active_editor_lock_payload(draft)
    if active_payload and int(active_payload.get("user_id") or 0) == int(current_user.id):
        _clear_editor_lock(draft)


def _can_edit_draft(draft: CSCDraft) -> bool:
    """Check if current user can edit this draft."""
    revision = _get_revision_for_child(draft.id)
    if _can_current_user_edit_draft_as_committee_head(draft):
        return True
    if (
        draft.parent_draft_id is not None
        and revision is not None
        and _is_secretary_staging_state(revision.status)
        and (_current_user_is_material_master_admin() or current_user.is_super_user())
    ):
        return True
    if (
        draft.parent_draft_id is not None
        and draft.is_admin_draft
        and draft.created_by_id == current_user.id
        and (_current_user_is_material_master_admin() or current_user.is_super_user())
        and revision is not None
        and _is_open_revision_state(revision.status)
    ):
        return True
    if not _draft_in_current_user_committee_scope(draft):
        return False
    if current_user.is_super_user():
        if (
            draft.parent_draft_id is not None
            and draft.is_admin_draft
            and draft.created_by_id == current_user.id
            and revision is not None
            and _is_open_revision_state(revision.status)
        ):
            return True
    return _can_user_edit_workflow_draft(draft)


def _can_delete_draft(draft: CSCDraft) -> bool:
    """Check if current user can delete this draft."""
    revision = _get_revision_for_child(draft.id)
    if draft.parent_draft_id is None or revision is None:
        return False
    if not (_current_user_is_material_master_admin() or current_user.is_super_user()):
        return False
    return (revision.status or "").strip().lower() in {
        WORKFLOW_DRAFTING_STAGING,
        WORKFLOW_DRAFTING_OPEN,
        WORKFLOW_DRAFTING_SUBMITTED,
        WORKFLOW_DRAFTING_HEAD_APPROVED,
        WORKFLOW_DRAFTING_RETURNED,
    }


def _get_current_user_committee_access() -> dict[str, object]:
    """Return committee assignment data for the current user."""
    explicit_access = get_committee_access_for_username(getattr(current_user, "username", None))
    if explicit_access.get("subset_codes") or explicit_access.get("committee_slugs"):
        return explicit_access
    if current_user.is_super_user():
        return {"committee_slugs": [], "committee_titles": [], "subset_codes": []}
    return explicit_access


def _get_current_user_committee_subset_codes() -> list[str]:
    """Return the subset codes available to the current user."""
    return list(_get_current_user_committee_access().get("subset_codes") or [])


def _get_effective_committee_user_subset_codes() -> list[str]:
    """Return subset scope for the Committee User Workbench only."""
    return _get_current_user_committee_user_subset_codes()


def _get_current_user_committee_user_subset_codes_for_slug(committee_slug: str | None) -> list[str]:
    """Return subset codes assigned to the current user for one committee-user role."""
    slug = str(committee_slug or "").strip()
    if not slug or slug in {"coordination", "corporate-specification-committee"}:
        return []

    subset_codes = set()
    for record in _get_current_user_committee_records():
        committee = record["committee"]
        if "committee_user" not in record["roles"]:
            continue
        if str(committee.get("slug") or "").strip() != slug:
            continue
        for code in _resolve_committee_user_subset_codes(
            committee,
            record.get("committee_user_assignment"),
        ):
            if code:
                subset_codes.add(code)

    ordered_codes = [code for code in SPEC_SUBSET_ORDER if code in subset_codes]
    ordered_codes.extend(sorted(code for code in subset_codes if code not in SPEC_SUBSET_ORDER))
    return ordered_codes


def _get_committee_user_upload_subset_codes() -> list[str]:
    """Return subset codes eligible for committee-user new specification upload."""
    subset_codes = set()
    for record in _get_current_user_committee_records():
        committee = record["committee"]
        committee_slug = str(committee.get("slug") or "").strip()
        if "committee_user" not in record["roles"]:
            continue
        if committee_slug in {"coordination", "corporate-specification-committee", "material-handling"}:
            continue
        for subset in committee.get("subsets") or []:
            code = str((subset or {}).get("code") or "").strip().upper()
            if code:
                subset_codes.add(code)
    ordered_codes = [code for code in SPEC_SUBSET_ORDER if code in subset_codes]
    ordered_codes.extend(sorted(code for code in subset_codes if code not in SPEC_SUBSET_ORDER))
    return ordered_codes


def _require_committee_user_scope() -> None:
    """Abort when the current user has no committee-user subset assignment."""
    if not _get_current_user_committee_user_slugs():
        abort(403)


def _require_committee_user_upload_scope() -> None:
    """Abort when the current user cannot create subset-committee uploads."""
    if not _get_committee_user_upload_subset_codes():
        abort(403)


def _workflow_subset_code_for_draft(
    draft: CSCDraft | None,
    revision: CSCRevision | None = None,
) -> str | None:
    """Return the stable workflow subset code for the draft, preferring the published parent."""
    if draft is None:
        return None

    parent_draft = revision.parent_draft if revision is not None else None
    if parent_draft is None and getattr(draft, "parent_draft_id", None):
        parent_draft = db.session.get(CSCDraft, draft.parent_draft_id)

    subset_code = getattr(parent_draft or draft, "subset", None)
    normalized = str(subset_code or "").strip().upper()
    return normalized or None


def _subset_code_is_within_scope(subset_code: str | None, allowed_subset_codes: list[str]) -> bool:
    """Return True when the subset is covered or no explicit subset scope is configured."""
    if not allowed_subset_codes:
        return True
    normalized = str(subset_code or "").strip().upper()
    return bool(normalized and normalized in set(allowed_subset_codes))


def _draft_in_current_user_committee_scope(
    draft: CSCDraft,
    revision: CSCRevision | None = None,
) -> bool:
    """Return True when the draft falls inside the current user's covered subsets."""
    allowed_committee_slugs = set(_get_current_user_committee_user_slugs())
    if current_user.is_super_user() and not allowed_committee_slugs:
        return True
    revision = revision or _get_revision_for_child(draft.id)
    committee_slug = _get_revision_committee_slug(revision) if revision else None
    if not committee_slug or committee_slug not in allowed_committee_slugs:
        return False
    allowed_subset_codes = _get_current_user_committee_user_subset_codes_for_slug(committee_slug)
    return _subset_code_is_within_scope(
        _workflow_subset_code_for_draft(draft, revision),
        allowed_subset_codes,
    )


def _ensure_current_user_can_access_subset(subset_code: str | None) -> None:
    """Raise PermissionError when a committee-scoped user targets an uncovered subset."""
    allowed_subsets = set(_get_effective_committee_user_subset_codes())
    if current_user.is_super_user() and not allowed_subsets:
        return
    normalized_subset = (subset_code or "").strip().upper()
    if not normalized_subset:
        return
    if normalized_subset and normalized_subset in allowed_subsets:
        return
    raise PermissionError(
        "You are not assigned to this subset in Governance Settings."
    )


def _get_current_user_committee_records() -> list[dict[str, object]]:
    """Return normalized committee records where the current user is assigned."""
    username_key = str(getattr(current_user, "username", "") or "").strip().lower()
    if not username_key:
        return []

    payload = get_committee_config_payload()
    records = []
    for committee in [payload["ROOT_COMMITTEE"], *payload["CHILD_COMMITTEES"]]:
        roles = []
        assignment = _get_committee_user_assignment_for_username(committee, username_key)
        if assignment is not None:
            roles.append("committee_user")
        if str(committee.get("committee_head") or "").strip().lower() == username_key:
            roles.append("committee_head")
        if roles:
            records.append(
                {
                    "committee": committee,
                    "roles": roles,
                    "committee_user_assignment": assignment,
                }
            )
    return records


def _get_current_user_committee_slugs_for_role(role_name: str) -> list[str]:
    """Return committee slugs where the current user is assigned for the given role."""
    slugs = []
    for record in _get_current_user_committee_records():
        if role_name in record["roles"]:
            slug = str(record["committee"].get("slug") or "").strip()
            if slug:
                slugs.append(slug)
    return list(dict.fromkeys(slugs))


def _get_current_user_committee_user_slugs() -> list[str]:
    return _get_current_user_committee_slugs_for_role("committee_user")


def _get_current_user_committee_head_slugs() -> list[str]:
    return _get_current_user_committee_slugs_for_role("committee_head")


def _get_current_user_committee_head_subset_codes() -> list[str]:
    """Return subset codes where the current user is configured as committee head."""
    subset_codes = set()
    for record in _get_current_user_committee_records():
        committee = record["committee"]
        if "committee_head" not in record["roles"]:
            continue
        if committee.get("slug") == "coordination":
            continue
        for subset in committee.get("subsets") or []:
            code = str((subset or {}).get("code") or "").strip().upper()
            if code:
                subset_codes.add(code)
    ordered_codes = [code for code in SPEC_SUBSET_ORDER if code in subset_codes]
    ordered_codes.extend(sorted(code for code in subset_codes if code not in SPEC_SUBSET_ORDER))
    return ordered_codes


def _get_current_user_committee_head_subset_codes_for_slug(committee_slug: str | None) -> list[str]:
    """Return subset codes assigned to the current user for one committee-head role."""
    slug = str(committee_slug or "").strip()
    if not slug or slug == "coordination":
        return []

    subset_codes = set()
    for record in _get_current_user_committee_records():
        committee = record["committee"]
        if "committee_head" not in record["roles"]:
            continue
        if str(committee.get("slug") or "").strip() != slug:
            continue
        for subset in committee.get("subsets") or []:
            code = str((subset or {}).get("code") or "").strip().upper()
            if code:
                subset_codes.add(code)

    ordered_codes = [code for code in SPEC_SUBSET_ORDER if code in subset_codes]
    ordered_codes.extend(sorted(code for code in subset_codes if code not in SPEC_SUBSET_ORDER))
    return ordered_codes


def _get_current_user_committee_user_subset_codes() -> list[str]:
    """Return subset codes where the current user is configured as committee user."""
    subset_codes = set()
    for record in _get_current_user_committee_records():
        committee = record["committee"]
        committee_slug = str(committee.get("slug") or "").strip()
        if "committee_user" not in record["roles"]:
            continue
        if committee_slug in {"coordination", "corporate-specification-committee"}:
            continue
        for code in _resolve_committee_user_subset_codes(
            committee,
            record.get("committee_user_assignment"),
        ):
            if code:
                subset_codes.add(code)
    ordered_codes = [code for code in SPEC_SUBSET_ORDER if code in subset_codes]
    ordered_codes.extend(sorted(code for code in subset_codes if code not in SPEC_SUBSET_ORDER))
    return ordered_codes


def _resolve_subset_committee_slug_for_subset(subset_code: str | None) -> str | None:
    """Return the configured subset committee slug responsible for the subset."""
    normalized_subset = str(subset_code or "").strip().upper()
    if not normalized_subset:
        return None
    payload = get_committee_config_payload()
    for committee in payload.get("CHILD_COMMITTEES", []):
        slug = str(committee.get("slug") or "").strip()
        if slug in {"coordination", "material-handling"}:
            continue
        subset_codes = {
            str((subset or {}).get("code") or "").strip().upper()
            for subset in (committee.get("subsets") or [])
        }
        if normalized_subset in subset_codes:
            return slug
    return None


def _committee_slug_to_track(committee_slug: str | None) -> str | None:
    slug = str(committee_slug or "").strip()
    if not slug:
        return None
    if slug == "material-handling":
        return WORKFLOW_TRACK_MATERIAL_HANDLING
    if slug in {"coordination", "corporate-specification-committee"}:
        return None
    return WORKFLOW_TRACK_SUBSET


def _workflow_track_to_stream_name(workflow_track: str | None) -> str:
    track = str(workflow_track or "").strip().lower()
    if track == WORKFLOW_TRACK_MATERIAL_HANDLING:
        return MATERIAL_HANDLING_STREAM
    return TYPE_CLASSIFICATION_STREAM


def _get_workflow_stream_section(draft: CSCDraft) -> CSCSection | None:
    if not getattr(draft, "id", None) or not getattr(draft, "parent_draft_id", None):
        return None
    return CSCSection.query.filter_by(
        draft_id=draft.id,
        section_name=WORKFLOW_STREAM_SECTION_NAME,
    ).first()


def _write_workflow_stream_name(draft: CSCDraft, stream_name: str | None) -> str:
    normalized = str(stream_name or "").strip().lower() or TYPE_CLASSIFICATION_STREAM
    section = _get_workflow_stream_section(draft)
    if section is None:
        section = CSCSection(
            draft_id=draft.id,
            section_name=WORKFLOW_STREAM_SECTION_NAME,
            sort_order=9996,
        )
        db.session.add(section)
    section.section_text = normalized
    return normalized


def _infer_workflow_stream_name(draft: CSCDraft | None) -> str:
    if not draft:
        return TYPE_CLASSIFICATION_STREAM
    section = _get_workflow_stream_section(draft)
    if section and (section.section_text or "").strip():
        return (section.section_text or "").strip().lower()
    committee_slug = _infer_workflow_committee_slug(draft)
    if committee_slug == "material-handling":
        return MATERIAL_HANDLING_STREAM
    return TYPE_CLASSIFICATION_STREAM


def _master_data_editable_for_stream(stream_name: str | None) -> bool:
    return (stream_name or "").strip().lower() == MATERIAL_HANDLING_STREAM


def _master_data_editable_for_draft(draft: CSCDraft | None) -> bool:
    return _master_data_editable_for_stream(_infer_workflow_stream_name(draft))


def _get_workflow_committee_section(draft: CSCDraft) -> CSCSection | None:
    if not getattr(draft, "id", None) or not getattr(draft, "parent_draft_id", None):
        return None
    return CSCSection.query.filter_by(
        draft_id=draft.id,
        section_name=WORKFLOW_COMMITTEE_SECTION_NAME,
    ).first()


def _write_workflow_committee_slug(draft: CSCDraft, committee_slug: str | None) -> str:
    slug = str(committee_slug or "").strip()
    section = _get_workflow_committee_section(draft)
    if section is None:
        section = CSCSection(
            draft_id=draft.id,
            section_name=WORKFLOW_COMMITTEE_SECTION_NAME,
            sort_order=9997,
        )
        db.session.add(section)
    section.section_text = slug
    return slug


def _infer_workflow_committee_slug(draft: CSCDraft | None) -> str | None:
    if not draft:
        return None
    section = _get_workflow_committee_section(draft)
    if section and (section.section_text or "").strip():
        return (section.section_text or "").strip()
    scope = _load_workflow_scope(draft)
    if (
        (_scope_allows_material_properties(scope) or _scope_allows_storage_handling(scope))
        and not _scope_allows_impact(scope)
    ):
        return "material-handling"
    return _resolve_subset_committee_slug_for_subset(draft.subset)


def _get_revision_committee_slug(revision: CSCRevision | None) -> str | None:
    if revision is None:
        return None
    return _infer_workflow_committee_slug(revision.child_draft)


def _get_committee_title_by_slug(committee_slug: str | None) -> str:
    slug = str(committee_slug or "").strip()
    payload = get_committee_config_payload()
    for committee in [payload["ROOT_COMMITTEE"], *payload["CHILD_COMMITTEES"]]:
        if str(committee.get("slug") or "").strip() == slug:
            return str(committee.get("title") or slug or "Committee")
    return slug or "Committee"


def _resolve_requested_workflow_committee_slug(parent_draft: CSCDraft, requested_track: str | None) -> str | None:
    track = str(requested_track or "").strip().lower()
    if track == WORKFLOW_TRACK_MATERIAL_HANDLING:
        return "material-handling"
    if track == WORKFLOW_TRACK_SUBSET:
        return _resolve_subset_committee_slug_for_subset(parent_draft.subset)
    return None


def _current_user_is_governance_committee_member() -> bool:
    """Return True when the current user is the Governance Committee user/head."""
    return any(
        (record["committee"].get("slug") == "coordination")
        for record in _get_current_user_committee_records()
    )


def _can_current_user_access_committee_user_workbench() -> bool:
    """Return True when the current user can open the committee-user workbench."""
    return (
        current_user.is_super_user()
        or bool(_get_current_user_committee_user_slugs())
    )


def _current_user_is_material_master_admin() -> bool:
    """Return True when the current user is a module admin for CSC."""
    return current_user.is_module_admin("csc")


def _can_current_user_view_revision_snapshot(revision: CSCRevision) -> bool:
    """Return True when the current user can open a read-only revision snapshot."""
    if _current_user_is_material_master_admin() or current_user.is_super_user():
        return True
    if _current_user_is_governance_committee_member():
        return True
    if _revision_in_current_user_committee_head_scope(revision):
        return True
    child_draft = revision.child_draft
    return bool(child_draft and _draft_in_current_user_committee_scope(child_draft, revision))


def _revision_in_current_user_committee_head_scope(revision: CSCRevision | None) -> bool:
    """Return True when the revision falls inside the current committee-head assignment."""
    if revision is None:
        return False

    committee_slug = _get_revision_committee_slug(revision)
    if not committee_slug or committee_slug not in set(_get_current_user_committee_head_slugs()):
        return False

    allowed_subset_codes = _get_current_user_committee_head_subset_codes_for_slug(committee_slug)
    return _subset_code_is_within_scope(
        _workflow_subset_code_for_draft(revision.child_draft, revision),
        allowed_subset_codes,
    )


def _can_current_user_decide_revision_as_committee_head(revision: CSCRevision) -> bool:
    """Return True when the current user can take committee-head decisions."""
    return revision.status == WORKFLOW_DRAFTING_SUBMITTED and _revision_in_current_user_committee_head_scope(revision)


def _can_current_user_decide_revision_as_module_admin(revision: CSCRevision) -> bool:
    """Return True when the current user can take module-admin decisions."""
    return (
        (_current_user_is_material_master_admin() or current_user.is_super_user())
        and revision.status == WORKFLOW_DRAFTING_HEAD_APPROVED
    )


def _require_revision_snapshot_access(revision: CSCRevision) -> None:
    """Abort when the current user cannot inspect the revision snapshot."""
    if not _can_current_user_view_revision_snapshot(revision):
        abort(403)


def _get_material_master_admin_users() -> list[User]:
    """Return active module admins for the CSC module."""
    try:
        from app.models.core.module_admin_assignment import ModuleAdminAssignment
        admin_user_ids_subq = (
            db.session.query(ModuleAdminAssignment.user_id)
            .filter_by(module_code="csc")
            .subquery()
        )
        return sorted(
            User.query
            .filter(User.id.in_(admin_user_ids_subq), User.is_active.is_(True))
            .all(),
            key=lambda u: ((u.full_name or "").strip().lower(), u.username.lower()),
        )
    except (ProgrammingError, OperationalError) as exc:
        logger.warning("CSC module admin lookup skipped: %s", exc)
        db.session.rollback()
        return []


def _get_material_master_admin_labels() -> list[str]:
    """Return display labels for assigned CSC module admins."""
    return [
        f"{user.full_name} ({user.username})"
        if (user.full_name or "").strip()
        else user.username
        for user in _get_material_master_admin_users()
    ]


@csc_bp.context_processor
def inject_csc_workflow_navigation():
    """Expose CSC workflow-navigation visibility flags to templates."""
    return {
        "csc_can_access_committee_user_workbench": _can_current_user_access_committee_user_workbench(),
        "csc_can_upload_new_specs": bool(_get_committee_user_upload_subset_codes()),
        "csc_can_access_review_workbench": (
            _current_user_is_governance_committee_member()
            or bool(_get_current_user_committee_head_slugs())
        ),
        "csc_is_material_master_admin": _current_user_is_material_master_admin(),
    }


def _get_committee_head_user_ids_for_subset(subset_code: str | None) -> list[int]:
    """Return active configured committee-head user ids for the subset."""
    normalized_subset = str(subset_code or "").strip().upper()
    if not normalized_subset:
        return []

    candidate_usernames: list[str] = []
    for committee in get_committee_config_payload().get("CHILD_COMMITTEES", []):
        if committee.get("slug") == "coordination":
            continue
        subset_codes = {
            str((subset or {}).get("code") or "").strip().upper()
            for subset in (committee.get("subsets") or [])
        }
        if normalized_subset not in subset_codes:
            continue
        username = str(committee.get("committee_head") or "").strip()
        if username:
            candidate_usernames.append(username)

    user_map = _batch_fetch_users_by_usernames(candidate_usernames)
    user_ids: list[int] = []
    for username in candidate_usernames:
        user = user_map.get(username)
        if user:
            user_ids.append(user.id)
    return list(dict.fromkeys(user_ids))


def _get_committee_user_ids_for_committee_slug(committee_slug: str | None, role_field: str) -> list[int]:
    """Return active configured committee user/head ids for the committee slug."""
    slug = str(committee_slug or "").strip()
    if role_field not in {"committee_user", "committee_head"} or not slug:
        return []
    for committee in get_committee_config_payload().get("CHILD_COMMITTEES", []):
        if str(committee.get("slug") or "").strip() != slug:
            continue
        usernames = []
        if role_field == "committee_user":
            usernames = [
                str((assignment or {}).get("username") or "").strip()
                for assignment in _get_committee_user_assignments(committee)
            ]
        else:
            usernames = [str(committee.get("committee_head") or "").strip()]

        user_map = _batch_fetch_users_by_usernames(usernames)
        user_ids = []
        for username in usernames:
            if not username:
                continue
            user = user_map.get(username)
            if user:
                user_ids.append(user.id)
        return list(dict.fromkeys(user_ids))
    return []


def _get_committee_user_ids_for_revision(revision: CSCRevision | None) -> list[int]:
    """Return committee-user ids eligible to receive workflow notifications for a revision."""
    if revision is None:
        return []

    committee_slug = _get_revision_committee_slug(revision)
    if not committee_slug:
        return []

    subset_code = _workflow_subset_code_for_draft(revision.child_draft, revision)
    if committee_slug != "material-handling":
        return _get_committee_user_ids_for_committee_slug(committee_slug, "committee_user")

    for committee in get_committee_config_payload().get("CHILD_COMMITTEES", []):
        if str(committee.get("slug") or "").strip() != committee_slug:
            continue

        candidate_usernames: list[str] = []
        for assignment in _get_committee_user_assignments(committee):
            allowed_subset_codes = [
                str(code or "").strip().upper()
                for code in (assignment.get("subset_codes") or [])
                if str(code or "").strip()
            ]
            if allowed_subset_codes and not _subset_code_is_within_scope(subset_code, allowed_subset_codes):
                continue
            username = str(assignment.get("username") or "").strip()
            if username:
                candidate_usernames.append(username)
        user_map = _batch_fetch_users_by_usernames(candidate_usernames)
        user_ids: list[int] = []
        for username in candidate_usernames:
            user = user_map.get(username)
            if user:
                user_ids.append(user.id)
        return list(dict.fromkeys(user_ids))

    return []


def _get_committee_user_display_label_for_committee_slug(committee_slug: str | None, role_field: str) -> str:
    """Return display label for the configured committee user/head on a committee slug."""
    slug = str(committee_slug or "").strip()
    if role_field not in {"committee_user", "committee_head"} or not slug:
        return ""
    for committee in get_committee_config_payload().get("CHILD_COMMITTEES", []):
        if str(committee.get("slug") or "").strip() != slug:
            continue
        if role_field == "committee_user":
            usernames = [
                str((assignment or {}).get("username") or "").strip()
                for assignment in _get_committee_user_assignments(committee)
            ]
        else:
            usernames = [str(committee.get("committee_head") or "").strip()]

        user_map = _batch_fetch_users_by_usernames(usernames)
        labels = []
        for username in usernames:
            if not username:
                continue
            user = user_map.get(username)
            if not user:
                labels.append(username)
                continue
            if (user.full_name or "").strip():
                labels.append(f"{user.full_name} ({user.username})")
            else:
                labels.append(user.username)
        return ", ".join(labels)
    return ""


def _load_draft_origin_metadata(draft: CSCDraft) -> dict[str, object]:
    """Return hidden draft-origin metadata persisted in a system section."""
    if not getattr(draft, "id", None):
        return {}
    section = draft.sections.filter_by(section_name=DRAFT_ORIGIN_SECTION_NAME).first()
    if section is None or not (section.section_text or "").strip():
        return {}
    try:
        payload = json.loads(section.section_text)
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _write_draft_origin_metadata(draft: CSCDraft, payload: dict[str, object]) -> None:
    """Persist hidden draft-origin metadata on a draft."""
    section = draft.sections.filter_by(section_name=DRAFT_ORIGIN_SECTION_NAME).first()
    text = json.dumps(payload)
    if section is None:
        db.session.add(
            CSCSection(
                draft_id=draft.id,
                section_name=DRAFT_ORIGIN_SECTION_NAME,
                section_text=text,
                sort_order=9990,
            )
        )
        return
    section.section_text = text


def _draft_type_label(draft: CSCDraft) -> str:
    """Return the user-facing workflow type for a draft."""
    origin = _load_draft_origin_metadata(draft)
    if (origin.get("draft_kind") or "").strip() == "new_specification":
        return "New Specification"
    return "Revision"


def _calculate_impact_grade(
    operational: int, safety: int, supply: int, no_substitute: bool
) -> str:
    """Calculate impact grade based on weighted score.

    Weight formula: (operational * 0.5) + (safety * 0.3) + (supply * 0.2)
    Grades: LOW (< 2.0), MODERATE (2.0-2.99), HIGH (3.0-3.99), CRITICAL (≥ 4.0)
    Override: if no_substitute_flag and grade in (LOW, MODERATE) → HIGH
    """
    weighted_score = (operational * 0.5) + (safety * 0.3) + (supply * 0.2)

    if weighted_score < 2.0:
        grade = "LOW"
    elif weighted_score < 3.0:
        grade = "MODERATE"
    elif weighted_score < 4.0:
        grade = "HIGH"
    else:
        grade = "CRITICAL"

    # Override if no substitute flag and grade is LOW or MODERATE
    if no_substitute and grade in ("LOW", "MODERATE"):
        grade = "HIGH"

    return grade


def _status_filter_to_db_value(status: str | None) -> str | None:
    """Map UI status tokens to stored CSC draft status values."""
    value = (status or "").strip().lower()
    if not value:
        return None

    mapping = {
        "drafting": "Drafting",
        "published": "Published",
    }
    return mapping.get(value)


def _status_db_to_ui_value(status: str | None) -> str:
    """Map stored CSC draft status values to UI tokens."""
    value = (status or "").strip().lower().replace(" ", "_")
    allowed = {"drafting", "published"}
    return value if value in allowed else "published"


def _subset_display(code: str | None) -> str | None:
    """Return subset code plus human-readable label."""
    if not code:
        return None
    label = SPEC_SUBSET_LABELS.get(code, code)
    return f"{code} - {label}" if label else code


def _draft_sort_key(draft: CSCDraft) -> tuple:
    """Sort drafts by canonical subset order first, then sequence."""
    return spec_sort_key(draft.spec_number or "")


def _find_conflicting_root_spec(spec_number: str) -> CSCDraft | None:
    """Return an existing root draft if the subset/serial is already in use."""
    subset_code, sequence, _ = parse_spec_number(spec_number or "")
    if not subset_code or not sequence:
        return None
    return CSCDraft.query.filter(
        CSCDraft.parent_draft_id.is_(None),
        CSCDraft.spec_number.like(f"{subset_code}-{sequence}%"),
    ).first()


def _seed_sections_for_new_spec(draft: CSCDraft, source_label: str) -> None:
    """Create initial narrative sections for a newly ingested specification."""
    section_rows = list(_default_section_rows(TYPE_CLASSIFICATION_STREAM))
    for section_name, section_text, sort_order in section_rows:
        db.session.add(
            CSCSection(
                draft_id=draft.id,
                section_name=section_name,
                section_text=section_text,
                sort_order=sort_order,
            )
        )


def _default_background_section_text() -> str:
    return "\n".join(
        [
            "Background /",
            "Usage of Chemical :",
            "Current constraints in Storage Conditions :",
        ]
    )


def _default_existing_spec_section_text() -> str:
    return "\n".join(
        [
            "Existing Specification Summary/",
            "Operational Issues:",
            "",
            "Quality Control Issues:",
        ]
    )


def _default_proposed_changes_section_text(stream_name: str | None = None) -> str:
    if stream_name == MATERIAL_HANDLING_STREAM:
        return "\n".join(
            [
                "Proposed Changes /",
                "Update the material handling, storage conditions, and packing controls as per the uploaded workbook.",
                "Replace the current specification number with the proposed new specification number after committee approval.",
            ]
        )

    return (
        "Proposed Changes /\n"
        "Complete the proposed changes narrative for this specification before submission."
    )


def _default_justification_section_text(stream_name: str | None = None) -> str:
    if stream_name == MATERIAL_HANDLING_STREAM:
        return "\n".join(
            [
                "Justification/",
                "as per Safety Data Sheet",
            ]
        )

    return "\n".join(
        [
            "Justification/",
            "parameter wise justification for rationale adapted to classify the parameter as vital",
        ]
    )


def _default_recommendation_section_text(stream_name: str | None = None) -> str:
    if stream_name == MATERIAL_HANDLING_STREAM:
        return "Migration to 2026 version with definition of material properties, handling and storage conditions"
    return (
        "Version Change Reason/\n"
        "Migration to 2026 version with type classification of all parameters into Vital and Desirable. "
        "Material Handling and Storage Conditions also defined"
    )


def _is_exact_testing_placeholder(text: str | None) -> bool:
    """Identify the seeded placeholder text that should not appear in drafts."""
    return str(text or "").strip().lower() == "testing"


def _canonical_section_name(section_name: str | None) -> str:
    normalized = str(section_name or "").strip()
    if normalized == "changes":
        return "proposed_changes"
    return normalized


def _default_section_rows(stream_name: str | None = None) -> list[tuple[str, str, int]]:
    return [
        ("background", _default_background_section_text(), 10),
        ("existing_spec", _default_existing_spec_section_text(), 20),
        ("proposed_changes", _default_proposed_changes_section_text(stream_name), 30),
        ("justification", _default_justification_section_text(stream_name), 60),
        ("recommendation", _default_recommendation_section_text(stream_name), 70),
    ]


def _ensure_default_draft_sections(draft: CSCDraft, stream_name: str | None = None) -> None:
    section_defaults = {
        section_name: (default_text, sort_order)
        for section_name, default_text, sort_order in _default_section_rows(stream_name)
    }
    legacy_material_handling_recommendation = (
        "Version Change Reason/\n"
        "Migration to 2026 version with revised material handling, storage conditions, "
        "packing controls, and updated specification numbering."
    )
    previous_material_handling_recommendation = (
        "Migration to 2026 version with defined material handling and storage conditions"
    )
    material_handling_recommendation_placeholders = {
        legacy_material_handling_recommendation,
        previous_material_handling_recommendation,
        _default_recommendation_section_text(),
    }

    existing_sections: dict[str, CSCSection] = {}
    for section in draft.sections.order_by(CSCSection.sort_order).all():
        canonical_name = _canonical_section_name(section.section_name)
        existing = existing_sections.get(canonical_name)
        if canonical_name != (section.section_name or ""):
            section.section_name = canonical_name
        if existing is None:
            existing_sections[canonical_name] = section
            continue

        section_text = (section.section_text or "").strip()
        existing_text = (existing.section_text or "").strip()
        if (
            (not existing_text or _is_exact_testing_placeholder(existing_text))
            and section_text
            and not _is_exact_testing_placeholder(section_text)
        ):
            existing.section_text = section.section_text
        if not existing.sort_order and section.sort_order:
            existing.sort_order = section.sort_order
        db.session.delete(section)

    for section_name, (default_text, sort_order) in section_defaults.items():
        section = existing_sections.get(section_name)
        if section is None:
            db.session.add(
                CSCSection(
                    draft_id=draft.id,
                    section_name=section_name,
                    section_text=default_text,
                    sort_order=sort_order,
                )
            )
            continue
        section_text = section.section_text or ""
        if (
            not section_text.strip()
            or _is_exact_testing_placeholder(section_text)
            or (
                stream_name == MATERIAL_HANDLING_STREAM
                and section_name == "recommendation"
                and section_text.strip() in material_handling_recommendation_placeholders
            )
        ):
            section.section_text = default_text
        if section.sort_order != sort_order:
            section.sort_order = sort_order


def _section_name_aliases(section_name: str | None) -> list[str]:
    normalized = _canonical_section_name(section_name)
    if normalized == "proposed_changes":
        return ["proposed_changes", "changes"]
    return [normalized] if normalized else []


def _seed_parameters_from_extracted_spec(draft: CSCDraft, spec: SpecDocument) -> None:
    """Create CSC parameter rows from an extracted CSC Format A spec document."""
    for index, parameter in enumerate(spec.parameters, start=1):
        parameter_name = (parameter.name or "").strip() or f"Parameter {index}"
        if (parameter.group or "").strip():
            parameter_name = f"[{parameter.group.strip()}] {parameter_name}"
        existing_value = (parameter.existing_value or "").strip()
        db.session.add(
            CSCParameter(
                draft_id=draft.id,
                parameter_name=parameter_name,
                parameter_type=normalize_parameter_type_label(parameter.parameter_type or "Desirable"),
                unit_of_measure="",
                existing_value=existing_value,
                required_value_type="text",
                required_value_text=existing_value,
                parameter_conditions=(parameter.unit_condition or "").strip(),
                test_procedure_type="",
                test_procedure_text="",
                test_method="",
                sort_order=index,
            )
        )


def _build_ingest_preview_spec(spec: SpecDocument) -> dict[str, object]:
    """Serialize an extracted spec for the committee ingest review screen."""
    return {
        "chemical_name": (spec.chemical_name or "").strip(),
        "spec_number": (spec.spec_number or "").strip(),
        "material_code": (spec.material_code or "").strip(),
        "test_procedure": (spec.test_procedure or "").strip(),
        "parse_warnings": list(spec.parse_warnings or []),
        "parameters": [
            {
                "number": parameter.number,
                "name": (parameter.name or "").strip(),
                "parameter_type": normalize_parameter_type_label(
                    parameter.parameter_type or "Desirable"
                ),
                "unit_condition": (parameter.unit_condition or "").strip(),
                "existing_value": (parameter.existing_value or "").strip(),
                "group": (parameter.group or "").strip(),
            }
            for parameter in spec.parameters
        ],
    }


def _spec_document_from_ingest_form() -> SpecDocument:
    """Rebuild an extracted spec from the committee review form."""
    spec = SpecDocument(
        chemical_name=(request.form.get("chemical_name") or "").strip(),
        spec_number=(request.form.get("spec_number") or "").strip(),
        material_code=(request.form.get("material_code") or "").strip(),
        test_procedure=(request.form.get("test_procedure") or "").strip(),
    )

    numbers = request.form.getlist("parameter_number[]")
    names = request.form.getlist("parameter_name[]")
    types = request.form.getlist("parameter_type[]")
    conditions = request.form.getlist("parameter_condition[]")
    values = request.form.getlist("parameter_value[]")
    groups = request.form.getlist("parameter_group[]")

    row_count = max(
        len(numbers),
        len(names),
        len(types),
        len(conditions),
        len(values),
        len(groups),
    )
    for index in range(row_count):
        number_raw = numbers[index] if index < len(numbers) else ""
        name = (names[index] if index < len(names) else "").strip()
        parameter_type = normalize_parameter_type_label(
            (types[index] if index < len(types) else "Desirable").strip()
        )
        unit_condition = (conditions[index] if index < len(conditions) else "").strip()
        existing_value = (values[index] if index < len(values) else "").strip()
        group = (groups[index] if index < len(groups) else "").strip()

        if not any([name, unit_condition, existing_value, group]):
            continue

        try:
            number = int((number_raw or "").strip())
        except ValueError:
            number = len(spec.parameters) + 1

        spec.parameters.append(
            ExtractedParameter(
                number=number,
                name=name or f"Parameter {number}",
                unit_condition=unit_condition,
                existing_value=existing_value,
                parameter_type=parameter_type,
                group=group,
                raw_left=name,
                raw_right=existing_value,
            )
        )

    return spec


def _create_new_spec_workflow_from_extraction(spec: SpecDocument, source_label: str) -> CSCDraft:
    """Create a committee workflow draft for a brand-new specification upload."""
    spec_number = (spec.spec_number or "").strip()
    chemical_name = (spec.chemical_name or "").strip()
    material_code = (spec.material_code or "").strip()

    subset_code, sequence, year = parse_spec_number(spec_number)
    if not subset_code or not sequence or year is None:
        raise ValueError("The uploaded file does not contain a valid CSC Format A specification number.")

    conflict = _find_conflicting_root_spec(spec_number)
    if conflict is not None:
        raise ValueError(
            f"Serial {subset_code}/{sequence} already exists as {conflict.spec_number}. "
            "Use the replacement workflow for the existing specification instead."
        )

    parent_draft = CSCDraft(
        spec_number=spec_number,
        chemical_name=chemical_name or "New Specification",
        committee_name="Corporate Specifiction Committee",
        prepared_by=current_user.username,
        reviewed_by=None,
        material_code=material_code or None,
        status="Drafting",
        created_by_role=current_user.role.name if current_user.role else "User",
        is_admin_draft=False,
        phase1_locked=False,
        parent_draft_id=None,
        admin_stage=ADMIN_STAGE_DRAFTING,
        spec_version=0,
        created_by_id=current_user.id,
    )
    db.session.add(parent_draft)
    db.session.flush()

    _seed_sections_for_new_spec(parent_draft, source_label)
    _seed_parameters_from_extracted_spec(parent_draft, spec)
    _write_draft_origin_metadata(
        parent_draft,
        {
            "draft_kind": "new_specification",
            "source_label": source_label,
            "ingested_by": current_user.username,
            "spec_number": spec_number,
        },
    )

    child_draft = _create_workflow_child_draft(
        parent_draft,
        created_by_id=current_user.id,
        created_by_role=current_user.role.name if current_user.role else "User",
        is_admin_draft=False,
    )
    _write_workflow_scope(child_draft, dict(WORKFLOW_SCOPE_DEFAULT))
    _write_workflow_committee_slug(
        child_draft,
        _resolve_subset_committee_slug_for_subset(subset_code),
    )
    _write_workflow_stream_name(child_draft, TYPE_CLASSIFICATION_STREAM)
    _ensure_default_draft_sections(child_draft, TYPE_CLASSIFICATION_STREAM)
    _write_draft_origin_metadata(
        child_draft,
        {
            "draft_kind": "new_specification",
            "source_label": source_label,
            "ingested_by": current_user.username,
            "spec_number": spec_number,
        },
    )

    db.session.add(
        CSCRevision(
            parent_draft_id=parent_draft.id,
            child_draft_id=child_draft.id,
            status=WORKFLOW_DRAFTING_STAGING,
        )
    )
    parent_draft.updated_at = datetime.now(timezone.utc)
    child_draft.updated_at = datetime.now(timezone.utc)
    return child_draft


def _create_audit_entry(
    draft_id: int,
    action: str,
    old_value: Optional[str] = None,
    new_value: Optional[str] = None,
    remarks: Optional[str] = None,
) -> None:
    """Create an audit log entry."""
    try:
        audit = CSCAudit(
            draft_id=draft_id,
            action=action,
            user_name=current_user.username,
            action_time=datetime.now(timezone.utc),
            old_value=old_value,
            new_value=new_value,
            remarks=remarks,
        )
        db.session.add(audit)
        db.session.commit()
    except Exception:
        logger.exception("Failed to create audit entry")
        db.session.rollback()


def _display_review_value(value: object, default: str = "—") -> str:
    """Normalize values for read-only admin review rendering."""
    if value is None:
        return default
    if isinstance(value, bool):
        return "Yes" if value else "No"
    text = str(value).strip()
    return text or default


def _normalize_review_compare_value(value: object) -> str:
    """Normalize review text for stable published-vs-draft comparisons."""
    text = _display_review_value(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in text.split("\n")]
    return "\n".join(lines).strip()


def _normalize_editor_text(value: object) -> str:
    """Match the editor text normalization used by the browser UI."""
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _normalize_editor_proposed_summary_text(value: object) -> str:
    """Strip auto-generated checklist wrappers from the changes summary baseline."""
    text = _normalize_editor_text(value)
    if text.startswith("Sections updated:\n"):
        marker = "\n\nChange summary:\n"
        return _normalize_editor_text(text.split(marker, 1)[1] if marker in text else "")
    if text.startswith("Change types: "):
        segments = text.split("\n\n")
        return _normalize_editor_text("\n\n".join(segments[1:]) if len(segments) > 1 else "")
    return text


def _editor_grid_snapshot(
    config: list[dict[str, object]] | None,
    state: dict[str, object] | None,
    *,
    include_defaults: bool = True,
) -> str:
    """Serialize a material-properties/storage grid the same way as the editor UI."""
    config = config or []
    snapshot_state = {str(key): _normalize_editor_text(value) for key, value in (state or {}).items()}
    snapshot: dict[str, str] = {}

    for field in config:
        field_name = str(field.get("field_name") or "").strip()
        if not field_name:
            continue
        default_value = _normalize_editor_text(field.get("default_value"))
        current_value = snapshot_state.get(field_name, "")
        if include_defaults and not current_value and default_value:
            current_value = default_value
        snapshot_state[field_name] = current_value

        show_if = field.get("show_if") if isinstance(field, dict) else None
        if isinstance(show_if, dict):
            dependency_name = str(show_if.get("field_name") or "").strip()
            dependency_value = snapshot_state.get(dependency_name, "")
            expected_value = _normalize_editor_text(show_if.get("equals"))
            if dependency_value != expected_value:
                snapshot_state[field_name] = ""
                snapshot[field_name] = ""
                continue

        if include_defaults and not snapshot_state[field_name] and default_value:
            snapshot_state[field_name] = default_value
        snapshot[field_name] = snapshot_state[field_name]

    return json.dumps(snapshot)


def _serialize_editor_issues(issues: list[dict[str, object]]) -> str:
    """Serialize issue flags using the same shape as the editor snapshot."""
    return json.dumps(
        [
            {
                "flagged": bool(issue.get("flagged")),
                "notes": _normalize_editor_text(issue.get("notes")),
            }
            for issue in issues
        ]
    )


def _serialize_editor_impact_state(checklist_state: dict[str, object] | None, chemical_name: str) -> str:
    """Serialize impact checklist state using the editor snapshot shape."""
    state = dict(checklist_state or {})
    state["chemical_name"] = chemical_name or ""
    return json.dumps(
        {
            "version": state.get("version") or 2,
            "chemical_name": state.get("chemical_name") or "",
            "flags": state.get("flags") or {},
        }
    )


def _build_editor_issue_rows(draft: CSCDraft | None) -> list[dict[str, object]]:
    """Return the editor issue rows for a draft or a blank issue baseline."""
    issue_rows = (
        draft.issue_flags.order_by(CSCIssueFlag.sort_order).all()
        if draft is not None
        else []
    )
    issues: list[dict[str, object]] = []
    for index, label in enumerate(EDITOR_ISSUE_LABELS):
        issue = issue_rows[index] if index < len(issue_rows) else None
        issues.append(
            {
                "issue_type": label,
                "flagged": bool(issue.is_present) if issue else False,
                "notes": issue.note if issue else "",
            }
        )
    return issues


def _build_editor_parameter_baseline(
    draft: CSCDraft,
    parent_draft: CSCDraft | None,
    *,
    phase1_locked: bool,
) -> str:
    """Serialize the published baseline for editor-side parameter change detection."""
    parent_by_sort: dict[int, CSCParameter] = {}
    parent_by_name: dict[str, CSCParameter] = {}
    if parent_draft is not None:
        parent_parameters = parent_draft.parameters.order_by(CSCParameter.sort_order).all()
        parent_by_sort = {parameter.sort_order: parameter for parameter in parent_parameters}
        parent_by_name = {
            (parameter.parameter_name or "").strip().lower(): parameter
            for parameter in parent_parameters
            if (parameter.parameter_name or "").strip()
        }

    baseline_rows: list[dict[str, object]] = []
    for parameter in draft.parameters.order_by(CSCParameter.sort_order).all():
        parent_parameter = _match_parent_parameter(parameter, parent_by_sort, parent_by_name)
        source_parameter = parent_parameter or parameter
        if phase1_locked:
            baseline_rows.append(
                {
                    "id": str(parameter.id or ""),
                    "proposed_value": "",
                }
            )
            continue

        baseline_rows.append(
            {
                "parameter_name": _normalize_editor_text(source_parameter.parameter_name),
                "parameter_type": _normalize_editor_text(source_parameter.parameter_type or "Desirable"),
                "unit_of_measure": _normalize_editor_text(source_parameter.unit_of_measure),
                "required_value_type": _normalize_editor_text(source_parameter.required_value_type or "text"),
                "required_value_text": _normalize_editor_text(
                    source_parameter.required_value_text or source_parameter.existing_value or ""
                ),
                "required_value_operator_1": _normalize_editor_text(source_parameter.required_value_operator_1),
                "required_value_value_1": _normalize_editor_text(source_parameter.required_value_value_1),
                "required_value_operator_2": _normalize_editor_text(source_parameter.required_value_operator_2),
                "required_value_value_2": _normalize_editor_text(source_parameter.required_value_value_2),
                "parameter_conditions": _normalize_editor_text(source_parameter.parameter_conditions),
                "test_procedure_type": _normalize_editor_text(
                    normalize_test_procedure_type(
                        source_parameter.test_procedure_type or source_parameter.test_method or ""
                    )
                ),
                "test_procedure_text": _normalize_editor_text(
                    source_parameter.test_procedure_text or source_parameter.test_method or ""
                ),
            }
        )

    return json.dumps(baseline_rows)


def _build_editor_comparison_baseline(
    draft: CSCDraft,
    parent_draft: CSCDraft | None,
    *,
    phase1_locked: bool,
) -> dict[str, str]:
    """Build the published baseline used by the editor Changes tab."""
    source_draft = parent_draft or draft
    blank_supporting_baseline = _should_blank_legacy_supporting_baseline(parent_draft)
    source_sections = {
        "background": _strip_system_bracket_lines(_get_draft_section_text(source_draft, "background")),
        "existing_spec": _strip_system_bracket_lines(_get_draft_section_text(source_draft, "existing_spec")),
        "proposed_changes": _normalize_editor_proposed_summary_text(
            _strip_system_bracket_lines(_get_draft_section_text(source_draft, "proposed_changes"))
        ),
        "justification": _strip_system_bracket_lines(_get_draft_section_text(source_draft, "justification")),
        "recommendation": _strip_system_bracket_lines(_get_draft_section_text(source_draft, "recommendation")),
        "recommendation_remarks": _strip_system_bracket_lines(
            _get_draft_section_text(source_draft, "recommendation_remarks")
        ),
    }

    material_properties_state = (
        {}
        if blank_supporting_baseline and parent_draft is not None
        else get_material_properties_values(source_draft)
    )
    storage_handling_state = (
        {}
        if blank_supporting_baseline and parent_draft is not None
        else get_storage_handling_values(source_draft)
    )

    if blank_supporting_baseline and parent_draft is not None:
        impact_state = build_default_impact_checklist_state(draft.chemical_name)
    elif source_draft.impact_analysis and (source_draft.impact_analysis.checklist_state_json or "").strip():
        impact_state = deserialize_impact_checklist_state(
            source_draft.impact_analysis.checklist_state_json,
            draft.chemical_name,
        )
    else:
        impact_state = build_default_impact_checklist_state(draft.chemical_name)

    return {
        "background": _normalize_editor_text(source_sections["background"]),
        "existing": _normalize_editor_text(source_sections["existing_spec"]),
        "issues": _serialize_editor_issues(_build_editor_issue_rows(source_draft)),
        "materialProperties": _editor_grid_snapshot(
            get_material_properties_fields(),
            material_properties_state,
            include_defaults=not (blank_supporting_baseline and parent_draft is not None),
        ),
        "storageHandling": _editor_grid_snapshot(
            get_storage_handling_fields(),
            storage_handling_state,
            include_defaults=not (blank_supporting_baseline and parent_draft is not None),
        ),
        "parameters": _build_editor_parameter_baseline(
            draft,
            parent_draft,
            phase1_locked=phase1_locked,
        ),
        "impact": _serialize_editor_impact_state(impact_state, draft.chemical_name or ""),
        "changes": _normalize_editor_text(source_sections["proposed_changes"]),
        "justification": _normalize_editor_text(source_sections["justification"]),
        "recommendation": _normalize_editor_text(source_sections["recommendation"]),
        "recommendationRemarks": _normalize_editor_text(source_sections["recommendation_remarks"]),
    }


def _format_parameter_requirement_for_review(parameter: CSCParameter) -> str:
    """Render the current requirement text for review/export comparisons."""
    return _display_review_value(
        format_required_value(
            parameter.required_value_type or "text",
            parameter.required_value_text or parameter.existing_value or "",
            parameter.required_value_operator_1,
            parameter.required_value_value_1,
            parameter.required_value_operator_2,
            parameter.required_value_value_2,
        )
    )


def _format_impact_checklist_summary_for_review(summary: object) -> str:
    """Render checklist summary dicts as a concise human-readable line."""
    if not isinstance(summary, dict):
        return _display_review_value(summary)
    flags = summary.get("flags") or []
    flag_bits = []
    for flag in flags:
        if not isinstance(flag, dict):
            continue
        order = flag.get("order")
        answer = flag.get("answer_label") or "—"
        flag_bits.append(f"F{order}: {answer}")
    confidence = summary.get("confidence", "")
    total = summary.get("total_flags", 10)
    parts = [
        f"Classification: {summary.get('classification') or '—'}" + (f" ({confidence})" if confidence else ""),
        f"Rule: {summary.get('rule') or '—'}",
        f"Red YES: {summary.get('red_yes_count', 0)}/4",
        f"Amber YES: {summary.get('amber_yes_count', 0)}/5",
        f"Answered: {summary.get('answered_count', 0)}/{total}",
    ]
    if flag_bits:
        parts.append("Flags: " + ", ".join(flag_bits))
    return " | ".join(parts)


def _build_impact_review_rows(
    draft: CSCDraft,
    master_values: dict[str, object] | None = None,
) -> list[dict[str, str]]:
    """Build impact rows from the Flask checklist model, with master fallback."""
    impact_analysis = draft.impact_analysis
    if impact_analysis and (impact_analysis.checklist_state_json or "").strip():
        summary = summarize_impact_checklist_state(
            deserialize_impact_checklist_state(
                impact_analysis.checklist_state_json,
                draft.chemical_name,
            )
        )
        confidence = summary.get("confidence", "")
        classification_str = str(summary.get("classification") or "—")
        if confidence:
            classification_str += f" ({confidence})"
        rows = [
            {
                "label": "Impact Classification",
                "value": _display_review_value(classification_str),
            },
            {
                "label": "Decision Rule",
                "value": _display_review_value(summary.get("rule")),
            },
            {
                "label": "Red YES Count",
                "value": _display_review_value(f"{summary.get('red_yes_count', 0)}/4"),
            },
            {
                "label": "Amber YES Count",
                "value": _display_review_value(f"{summary.get('amber_yes_count', 0)}/5"),
            },
            {
                "label": "Flags Answered",
                "value": _display_review_value(
                    f"{summary.get('answered_count', 0)}/{summary.get('total_flags', 10)}"
                ),
            },
        ]
        for flag in summary.get("flags") or []:
            if not isinstance(flag, dict):
                continue
            rows.append(
                {
                    "label": _display_review_value(flag.get("dimension"), "Impact Flag"),
                    "value": _display_review_value(flag.get("answer_label")),
                }
            )
        return rows

    fallback_rows = []
    if master_values:
        fallback_rows = [
            row
            for row in _build_labeled_value_rows(MASTER_IMPACT_FIELDS, master_values)
            if row.get("value") != "—"
        ]
    if fallback_rows:
        return fallback_rows

    if impact_analysis:
        return [
            {
                "label": "Impact Grade",
                "value": _display_review_value(impact_analysis.impact_grade),
            }
        ]

    return []


def _build_comparison_value_rows(
    current_rows: list[dict[str, object]] | None,
    source_rows: list[dict[str, object]] | None = None,
) -> list[dict[str, str]]:
    """Annotate label/value rows with published baseline and change status."""
    current_rows = current_rows or []
    if source_rows is None:
        source_rows = current_rows

    current_map = {
        str(row.get("label") or "").strip(): _display_review_value(row.get("value"))
        for row in current_rows
        if isinstance(row, dict) and str(row.get("label") or "").strip()
    }
    source_map = {
        str(row.get("label") or "").strip(): _display_review_value(row.get("value"))
        for row in source_rows
        if isinstance(row, dict) and str(row.get("label") or "").strip()
    }

    ordered_labels: list[str] = []
    for rows in (current_rows, source_rows):
        for row in rows:
            if not isinstance(row, dict):
                continue
            label = str(row.get("label") or "").strip()
            if label and label not in ordered_labels:
                ordered_labels.append(label)

    comparison_rows: list[dict[str, str]] = []
    for label in ordered_labels:
        value = current_map.get(label, "—")
        source_value = source_map.get(label, "—")
        if value == "—" and source_value == "—":
            continue
        comparison_rows.append(
            {
                "label": label,
                "value": value,
                "source_value": source_value,
                "change_status": (
                    "Revised"
                    if _normalize_review_compare_value(value) != _normalize_review_compare_value(source_value)
                    else "Retained"
                ),
            }
        )
    return comparison_rows


def _blank_comparison_source_labels(
    source_rows: list[dict[str, object]] | None,
    labels_to_blank: set[str] | None,
) -> list[dict[str, object]] | None:
    if source_rows is None:
        return None
    labels_to_blank = {str(label or "").strip() for label in (labels_to_blank or set()) if str(label or "").strip()}
    if not labels_to_blank:
        return list(source_rows)

    normalized_rows: list[dict[str, object]] = []
    for row in source_rows:
        if not isinstance(row, dict):
            continue
        normalized = dict(row)
        if str(normalized.get("label") or "").strip() in labels_to_blank:
            normalized["value"] = "—"
        normalized_rows.append(normalized)
    return normalized_rows


def _review_section_label(section_name: str | None) -> str:
    labels = {
        "background": "Background Context",
        "existing_spec": "Existing Specification Summary",
        "changes": "Proposed Changes",
        "proposed_changes": "Proposed Changes",
        "justification": "Justification",
        "recommendation": "Version Change Reason",
    }
    key = str(section_name or "").strip()
    return labels.get(key, key.replace("_", " ").strip().title() or "Section")


def _build_section_review_rows_for_draft(
    draft: CSCDraft,
    parent_draft: CSCDraft | None = None,
) -> list[dict[str, str]]:
    """Build narrative comparison rows against the published parent draft."""
    if parent_draft is None and getattr(draft, "parent_draft_id", None):
        parent_draft = db.session.get(CSCDraft, draft.parent_draft_id)

    excluded_section_names = {
        STAGED_MASTER_SECTION_NAME,
        WORKFLOW_SCOPE_SECTION_NAME,
        WORKFLOW_STREAM_SECTION_NAME,
        DRAFT_ORIGIN_SECTION_NAME,
    }

    current_sections: dict[str, str] = {}
    ordered_names: list[str] = []
    for section in draft.sections.order_by(CSCSection.sort_order).all():
        section_name = str(section.section_name or "").strip()
        if not section_name or section_name in excluded_section_names:
            continue
        current_sections[section_name] = (section.section_text or "").strip()
        if section_name not in ordered_names:
            ordered_names.append(section_name)

    source_sections: dict[str, str] = {}
    if parent_draft is not None:
        for section in parent_draft.sections.order_by(CSCSection.sort_order).all():
            section_name = str(section.section_name or "").strip()
            if not section_name or section_name in excluded_section_names:
                continue
            source_sections[section_name] = (section.section_text or "").strip()
            if section_name not in ordered_names:
                ordered_names.append(section_name)

    rows: list[dict[str, str]] = []
    for section_name in ordered_names:
        current_text = current_sections.get(section_name, "")
        source_text = source_sections.get(section_name, "")
        text = current_text or "—"
        source_value = source_text or "—"
        if text == "—" and source_value == "—":
            continue
        rows.append(
            {
                "label": _review_section_label(section_name),
                "text": text,
                "source_text": source_value,
                "change_status": (
                    "Revised"
                    if _normalize_review_compare_value(text) != _normalize_review_compare_value(source_value)
                    else "Retained"
                ),
            }
        )
    return rows


def _should_blank_legacy_supporting_baseline(parent_draft: CSCDraft | None) -> bool:
    """Legacy published specs at v0 do not carry supporting baseline sections."""
    return parent_draft is not None and int(getattr(parent_draft, "spec_version", 0) or 0) == 0


def _should_hide_legacy_supporting_sections(draft: CSCDraft | None) -> bool:
    """Published legacy v0 specs should not expose supporting sections as live baseline."""
    return (
        draft is not None
        and getattr(draft, "parent_draft_id", None) is None
        and int(getattr(draft, "spec_version", 0) or 0) == 0
    )


def _match_parent_parameter(
    child_parameter: CSCParameter,
    parent_by_sort: dict[int, CSCParameter],
    parent_by_name: dict[str, CSCParameter],
) -> CSCParameter | None:
    """Find the comparable parent parameter for a revision child row."""
    parent = parent_by_sort.get(child_parameter.sort_order)
    if parent is not None:
        return parent
    name_key = (child_parameter.parameter_name or "").strip().lower()
    if not name_key:
        return None
    return parent_by_name.get(name_key)


def _build_parameter_review_rows_for_draft(
    draft: CSCDraft,
    parent_draft: CSCDraft | None = None,
) -> list[dict[str, str]]:
    """Build review/export parameter rows using parent-vs-child comparisons when available."""
    if parent_draft is None and getattr(draft, "parent_draft_id", None):
        parent_draft = db.session.get(CSCDraft, draft.parent_draft_id)

    parent_by_sort: dict[int, CSCParameter] = {}
    parent_by_name: dict[str, CSCParameter] = {}
    if parent_draft is not None:
        parent_parameters = parent_draft.parameters.order_by(CSCParameter.sort_order).all()
        parent_by_sort = {parameter.sort_order: parameter for parameter in parent_parameters}
        parent_by_name = {
            (parameter.parameter_name or "").strip().lower(): parameter
            for parameter in parent_parameters
            if (parameter.parameter_name or "").strip()
        }

    rows: list[dict[str, str]] = []
    for parameter in draft.parameters.order_by(CSCParameter.sort_order).all():
        parent_parameter = _match_parent_parameter(parameter, parent_by_sort, parent_by_name)
        source_requirement = (
            _format_parameter_requirement_for_review(parent_parameter)
            if parent_parameter is not None
            else _format_parameter_requirement_for_review(parameter)
        )
        revised_requirement = _format_parameter_requirement_for_review(parameter)
        source_type = normalize_parameter_type_label(
            (parent_parameter.parameter_type if parent_parameter is not None else parameter.parameter_type) or ""
        )
        revised_type = normalize_parameter_type_label(parameter.parameter_type or "")
        source_unit = _display_review_value(
            parent_parameter.unit_of_measure if parent_parameter is not None else parameter.unit_of_measure
        )
        revised_unit = _display_review_value(parameter.unit_of_measure)
        source_test_procedure_type = _display_review_value(
            normalize_test_procedure_type(
                (parent_parameter.test_procedure_type if parent_parameter is not None else parameter.test_procedure_type)
                or ""
            )
        )
        revised_test_procedure_type = _display_review_value(
            normalize_test_procedure_type(parameter.test_procedure_type or "")
        )
        source_test_method = _display_review_value(
            (parent_parameter.test_procedure_text or parent_parameter.test_method)
            if parent_parameter is not None
            else (parameter.test_procedure_text or parameter.test_method)
        )
        revised_test_method = _display_review_value(parameter.test_procedure_text or parameter.test_method)
        source_conditions = _display_review_value(
            parent_parameter.parameter_conditions if parent_parameter is not None else parameter.parameter_conditions
        )
        revised_conditions = _display_review_value(parameter.parameter_conditions)
        requirement_changed = source_requirement != revised_requirement

        changed = any(
            [
                requirement_changed,
                source_type != revised_type,
                source_unit != revised_unit,
                source_test_procedure_type != revised_test_procedure_type,
                source_test_method != revised_test_method,
                source_conditions != revised_conditions,
            ]
        )

        rows.append(
            {
                "parameter_name": parameter.parameter_name or "Untitled Parameter",
                "parameter_type": revised_type or "—",
                "source_parameter_type": source_type or "—",
                "unit_of_measure": revised_unit,
                "source_unit_of_measure": source_unit,
                "existing_value": _display_review_value(
                    (parent_parameter.existing_value if parent_parameter is not None else parameter.existing_value)
                ),
                "required_value": source_requirement,
                "revised_requirement": revised_requirement,
                "proposed_value": (
                    revised_requirement
                    if requirement_changed
                    else ("No value change submitted" if changed else "No change submitted")
                ),
                "proposed_value_raw": revised_requirement if requirement_changed else "",
                "change_status": "Revised" if changed else "Retained",
                "test_procedure_type": revised_test_procedure_type,
                "source_test_procedure_type": source_test_procedure_type,
                "test_method": revised_test_method,
                "procedure_text": revised_test_method,
                "source_test_method": source_test_method,
                "conditions": revised_conditions,
                "source_conditions": source_conditions,
                "remarks": _display_review_value(parameter.remarks),
                "final_requirement": revised_requirement,
            }
        )
    return rows


def _normalize_workflow_scope_payload(raw: object | None) -> dict[str, bool]:
    """Normalize workflow scope JSON stored on workflow drafts."""
    payload = dict(WORKFLOW_SCOPE_DEFAULT)
    data = raw if isinstance(raw, dict) else {}
    explicit_parameter_main = False
    explicit_parameter_suboptions = False

    for key in ("material_properties", "storage_handling", "parameter_type", "parameter_other"):
        if key in data:
            payload[key] = _coerce_admin_boolean(data.get(key))
            if key in {"parameter_type", "parameter_other"}:
                explicit_parameter_suboptions = True

    if "parameters" in data:
        payload["parameters"] = _coerce_admin_boolean(data.get("parameters"))
        explicit_parameter_main = True
    elif "parameters_impact" in data:
        payload["parameters"] = _coerce_admin_boolean(data.get("parameters_impact"))
        explicit_parameter_main = True

    if "impact" in data:
        payload["impact"] = _coerce_admin_boolean(data.get("impact"))
    elif "parameters_impact" in data:
        payload["impact"] = _coerce_admin_boolean(data.get("parameters_impact"))

    if explicit_parameter_main and not payload["parameters"]:
        payload["parameter_type"] = False
        payload["parameter_other"] = False
    elif explicit_parameter_suboptions:
        payload["parameters"] = bool(payload["parameter_type"] or payload["parameter_other"])
    elif explicit_parameter_main and payload["parameters"]:
        payload["parameter_type"] = True
        payload["parameter_other"] = True

    return payload


def _get_workflow_scope_section(draft: CSCDraft) -> CSCSection | None:
    if not getattr(draft, "id", None) or not getattr(draft, "parent_draft_id", None):
        return None
    return CSCSection.query.filter_by(
        draft_id=draft.id,
        section_name=WORKFLOW_SCOPE_SECTION_NAME,
    ).first()


def _load_workflow_scope(draft: CSCDraft) -> dict[str, bool]:
    """Return per-draft workflow scope, defaulting to all tabs open."""
    section = _get_workflow_scope_section(draft)
    if section is None or not (section.section_text or "").strip():
        return dict(WORKFLOW_SCOPE_DEFAULT)
    try:
        payload = json.loads(section.section_text or "{}")
    except Exception:
        return dict(WORKFLOW_SCOPE_DEFAULT)
    return _normalize_workflow_scope_payload(payload)


def _write_workflow_scope(draft: CSCDraft, raw_scope: object | None) -> dict[str, bool]:
    """Persist workflow scope on a child workflow draft."""
    scope = _normalize_workflow_scope_payload(raw_scope)
    section = _get_workflow_scope_section(draft)
    if section is None:
        section = CSCSection(
            draft_id=draft.id,
            section_name=WORKFLOW_SCOPE_SECTION_NAME,
            sort_order=9998,
        )
        db.session.add(section)
    section.section_text = json.dumps(scope)
    return scope


def _validate_workflow_scope(scope: dict[str, bool]) -> str | None:
    if not any(scope.get(key) for key in ("material_properties", "storage_handling", "parameters", "impact")):
        return "Select at least one revision tab before creating the workflow."
    return None


def _scope_allows_material_properties(scope: dict[str, bool]) -> bool:
    return bool(scope.get("material_properties"))


def _scope_allows_storage_handling(scope: dict[str, bool]) -> bool:
    return bool(scope.get("storage_handling"))


def _scope_allows_impact(scope: dict[str, bool]) -> bool:
    return bool(scope.get("impact"))


def _scope_allows_parameter_tab(scope: dict[str, bool]) -> bool:
    return bool(scope.get("parameters") and (scope.get("parameter_type") or scope.get("parameter_other")))


def _summarize_workflow_scope(scope: dict[str, bool]) -> list[str]:
    summary: list[str] = []
    if _scope_allows_material_properties(scope):
        summary.append("Material Properties")
    if _scope_allows_storage_handling(scope):
        summary.append("Storage and Handling")
    if _scope_allows_parameter_tab(scope):
        if scope.get("parameter_type") and scope.get("parameter_other"):
            summary.append("Parameters — Full Revision")
        elif scope.get("parameter_type"):
            summary.append("Parameters — Type Only")
        elif scope.get("parameter_other"):
            summary.append("Parameters — Structure and Values")
    if _scope_allows_impact(scope):
        summary.append("Impact")
    return summary


def _build_labeled_value_rows(
    fields: list[tuple[str, str]],
    values: dict[str, object],
) -> list[dict[str, str]]:
    """Build non-empty label/value rows for review cards."""
    rows: list[dict[str, str]] = []
    for field_name, label in fields:
        value = _display_review_value(values.get(field_name))
        if value == "—":
            continue
        rows.append({"label": label, "value": value})
    return rows


def _create_workflow_child_draft(
    parent_draft: CSCDraft,
    *,
    created_by_id: int | None,
    created_by_role: str,
    is_admin_draft: bool,
) -> CSCDraft:
    """Create a workflow child draft by copying the published parent snapshot."""
    child_draft = CSCDraft(
        spec_number=parent_draft.spec_number,
        chemical_name=parent_draft.chemical_name,
        committee_name=parent_draft.committee_name,
        meeting_date=parent_draft.meeting_date,
        prepared_by=parent_draft.prepared_by,
        reviewed_by=parent_draft.reviewed_by,
        material_code=parent_draft.material_code,
        status="Drafting",
        created_by_role=created_by_role,
        is_admin_draft=is_admin_draft,
        phase1_locked=False,
        parent_draft_id=parent_draft.id,
        admin_stage=ADMIN_STAGE_DRAFTING,
        spec_version=parent_draft.spec_version,
        created_by_id=created_by_id,
    )
    db.session.add(child_draft)
    db.session.flush()
    _copy_draft_content(parent_draft, child_draft)
    child_draft.status = "Drafting"
    child_draft.created_by_role = created_by_role
    child_draft.is_admin_draft = is_admin_draft
    child_draft.phase1_locked = False
    child_draft.parent_draft_id = parent_draft.id
    child_draft.admin_stage = ADMIN_STAGE_DRAFTING
    child_draft.spec_version = parent_draft.spec_version
    child_draft.created_by_id = created_by_id
    return child_draft


def _seed_child_draft_master_snapshot(
    parent_draft: CSCDraft,
    child_draft: CSCDraft,
) -> None:
    """Ensure child drafts retain the parent's master-data snapshot, including impact fields."""
    if not getattr(child_draft, "parent_draft_id", None):
        return

    parent_values = dict(get_master_form_values(parent_draft))
    child_staged_values = dict(_load_staged_master_payload(child_draft))
    merged_values = {**parent_values, **child_staged_values}

    if merged_values != child_staged_values:
        _write_staged_master_payload(child_draft, merged_values)


def _get_active_revision_for_parent_committee(
    parent_draft_id: int | None,
    committee_slug: str | None,
) -> CSCRevision | None:
    slug = str(committee_slug or "").strip()
    if not slug:
        return None
    for revision in _get_active_revisions_for_parent(parent_draft_id):
        if _get_revision_committee_slug(revision) == slug:
            return revision
    return None


def _open_committee_workflow_draft_for_parent(
    parent_draft: CSCDraft,
    workflow_scope: dict[str, bool],
    workflow_track: str,
) -> tuple[CSCDraft | None, bool, str | None]:
    """Create or reuse a committee-facing workflow draft for the selected track."""
    committee_slug = _resolve_requested_workflow_committee_slug(parent_draft, workflow_track)
    if not committee_slug:
        return None, False, "No committee is configured for the selected workflow track."

    active_revisions = _get_active_revisions_for_parent(parent_draft.id)
    existing_for_committee = _get_active_revision_for_parent_committee(parent_draft.id, committee_slug)
    if existing_for_committee and existing_for_committee.child_draft:
        _seed_child_draft_master_snapshot(parent_draft, existing_for_committee.child_draft)
        _write_workflow_scope(existing_for_committee.child_draft, workflow_scope)
        _write_workflow_committee_slug(existing_for_committee.child_draft, committee_slug)
        stream_name = _write_workflow_stream_name(
            existing_for_committee.child_draft,
            _workflow_track_to_stream_name(workflow_track),
        )
        _ensure_default_draft_sections(existing_for_committee.child_draft, stream_name)
        return existing_for_committee.child_draft, False, None

    if len(active_revisions) >= 2:
        return None, False, "Only two active workflow drafts are allowed per specification at a time."

    child_draft = _create_workflow_child_draft(
        parent_draft,
        created_by_id=current_user.id,
        created_by_role=current_user.role.name if current_user.role else "Admin",
        is_admin_draft=True,
    )
    _seed_child_draft_master_snapshot(parent_draft, child_draft)
    _write_workflow_scope(child_draft, workflow_scope)
    _write_workflow_committee_slug(child_draft, committee_slug)
    stream_name = _write_workflow_stream_name(child_draft, _workflow_track_to_stream_name(workflow_track))
    _ensure_default_draft_sections(child_draft, stream_name)

    revision = CSCRevision(
        parent_draft_id=parent_draft.id,
        child_draft_id=child_draft.id,
        status=WORKFLOW_DRAFTING_STAGING,
    )
    db.session.add(revision)
    parent_draft.status = "Drafting"
    parent_draft.admin_stage = ADMIN_STAGE_DRAFTING
    parent_draft.updated_at = datetime.now(timezone.utc)
    return child_draft, True, None


def _build_workflow_meta_lines(
    parent_draft: CSCDraft,
    child_draft: CSCDraft | None,
    revision: CSCRevision | None,
) -> list[str]:
    """Return high-signal workflow metadata lines for Secretary notifications/popups."""
    committee_slug = _get_revision_committee_slug(revision) if revision else _infer_workflow_committee_slug(child_draft)
    committee_title = _get_committee_title_by_slug(committee_slug)
    committee_user_label = _get_committee_user_display_label_for_committee_slug(committee_slug, "committee_user")
    drafting_status = (revision.status if revision else "") or "open"
    lines = [
        f"Specification: {parent_draft.spec_number} — {parent_draft.chemical_name}",
        f"Committee: {committee_title}",
        f"Draft ID: {child_draft.id if child_draft else '—'}",
        f"Workflow Status: {drafting_status.replace('_', ' ').title()}",
    ]
    if committee_user_label:
        lines.append(f"Assigned User(s): {committee_user_label}")
    return lines


def _mark_response_no_store(response):
    """Disable browser caching for live workflow surfaces."""
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


def _open_admin_self_draft_for_parent(
    parent_draft: CSCDraft,
    workflow_scope: dict[str, bool],
) -> tuple[CSCDraft | None, bool, str | None]:
    """Create or reuse an admin-owned self draft for a published parent spec."""
    active_revision = _get_active_revision_for_parent(parent_draft.id)
    if active_revision and active_revision.child_draft:
        child_draft = active_revision.child_draft
        if (
            child_draft.is_admin_draft
            and child_draft.created_by_id == current_user.id
            and _is_open_revision_state(active_revision.status)
        ):
            _write_workflow_scope(child_draft, workflow_scope)
            return child_draft, False, None

        if active_revision.status in {WORKFLOW_DRAFTING_SUBMITTED, WORKFLOW_DRAFTING_HEAD_APPROVED}:
            return None, False, (
                "A committee workflow draft is already under review. "
                "Close that workflow before creating your own admin draft."
            )

        owner_name = (
            child_draft.creator.username
            if child_draft.creator and child_draft.creator.username
            else "another user"
        )
        return None, False, (
            f"An active workflow draft already exists for this specification under {owner_name}. "
            "Admin edits must use a self draft after that workflow is closed."
        )

    child_draft = _create_workflow_child_draft(
        parent_draft,
        created_by_id=current_user.id,
        created_by_role=current_user.role.name if current_user.role else "Admin",
        is_admin_draft=True,
    )
    _write_workflow_scope(child_draft, workflow_scope)
    _ensure_default_draft_sections(child_draft, TYPE_CLASSIFICATION_STREAM)

    revision = CSCRevision(
        parent_draft_id=parent_draft.id,
        child_draft_id=child_draft.id,
        status=WORKFLOW_DRAFTING_STAGING,
    )
    db.session.add(revision)
    parent_draft.status = "Drafting"
    parent_draft.admin_stage = ADMIN_STAGE_DRAFTING
    parent_draft.updated_at = datetime.now(timezone.utc)
    return child_draft, True, None


def _compose_approval_notes(
    reviewer_notes: str,
    approved_by_name: str,
    disha_file_number: str,
) -> str:
    """Persist approval note plus approval reference details in one field."""
    lines = [
        reviewer_notes.strip(),
        f"Approved By: {approved_by_name.strip()}",
        f"Disha File Number: {disha_file_number.strip()}",
    ]
    return "\n".join(line for line in lines if line)


def _workflow_stream_label(stream_name: str | None) -> str:
    return "Material Handling" if stream_name == MATERIAL_HANDLING_STREAM else "Type Classification"


def _workflow_scope_display_label(draft: CSCDraft | None) -> str:
    stream_name = _infer_workflow_stream_name(draft)
    if stream_name == MATERIAL_HANDLING_STREAM:
        return "Material Handling"

    if draft is None:
        return "Type Classification"

    workflow_scope = _load_workflow_scope(draft)
    if workflow_scope.get("parameter_type") and workflow_scope.get("parameter_other"):
        return "Full Revision"
    if workflow_scope.get("parameter_other"):
        return "Structure and Values"
    if workflow_scope.get("parameter_type"):
        return "Type Classification"
    return "Type Classification"


def _stream_section_markers(stream_name: str | None) -> tuple[str, str]:
    label = _workflow_stream_label(stream_name)
    return (f"[Approved {label} Stream]", f"[/Approved {label} Stream]")


def _merge_stream_section_text(
    existing_text: str,
    incoming_text: str,
    stream_name: str | None,
) -> str:
    incoming = (incoming_text or "").strip()
    if not incoming:
        return (existing_text or "").strip()
    start_marker, end_marker = _stream_section_markers(stream_name)
    pattern = re.compile(
        rf"{re.escape(start_marker)}.*?{re.escape(end_marker)}",
        re.DOTALL,
    )
    manual_text = pattern.sub("", existing_text or "").strip()
    merged_block = f"{start_marker}\n{incoming}\n{end_marker}"
    if not manual_text:
        return merged_block
    return f"{manual_text}\n\n{merged_block}"


def _replace_parent_parameters_from_child(parent_draft: CSCDraft, child_draft: CSCDraft) -> None:
    parent_draft.parameters.delete()
    for parameter in child_draft.parameters.order_by(CSCParameter.sort_order).all():
        db.session.add(
            CSCParameter(
                draft_id=parent_draft.id,
                parameter_name=parameter.parameter_name,
                parameter_type=parameter.parameter_type,
                unit_of_measure=parameter.unit_of_measure,
                existing_value=parameter.existing_value,
                proposed_value=parameter.proposed_value,
                test_method=parameter.test_method,
                test_procedure_type=parameter.test_procedure_type,
                test_procedure_text=parameter.test_procedure_text,
                parameter_conditions=parameter.parameter_conditions,
                required_value_type=parameter.required_value_type,
                required_value_text=parameter.required_value_text,
                required_value_operator_1=parameter.required_value_operator_1,
                required_value_value_1=parameter.required_value_value_1,
                required_value_operator_2=parameter.required_value_operator_2,
                required_value_value_2=parameter.required_value_value_2,
                justification=parameter.justification,
                remarks=parameter.remarks,
                sort_order=parameter.sort_order,
            )
        )


def _merge_child_sections_into_parent(
    parent_draft: CSCDraft,
    child_draft: CSCDraft,
    *,
    section_names: list[str],
    stream_name: str,
) -> None:
    for section_name in section_names:
        child_text = _get_draft_section_text(child_draft, section_name)
        if not (child_text or "").strip():
            continue
        parent_text = _get_draft_section_text(parent_draft, section_name)
        _set_draft_section_text(
            parent_draft,
            section_name,
            _merge_stream_section_text(parent_text, child_text, stream_name),
        )


def _apply_approved_revision_stream_to_parent(
    parent_draft: CSCDraft,
    child_draft: CSCDraft,
    revision: CSCRevision,
) -> dict[str, object]:
    stream_name = _infer_workflow_stream_name(child_draft)
    master_payload = dict(get_master_form_values(child_draft))

    if stream_name == MATERIAL_HANDLING_STREAM:
        apply_material_handling_core_updates(parent_draft, child_draft)
        _merge_child_sections_into_parent(
            parent_draft,
            child_draft,
            section_names=["background", "proposed_changes", "justification", "recommendation"],
            stream_name=stream_name,
        )
        master_payload["chemical_name"] = parent_draft.chemical_name or ""
        master_payload["short_text"] = parent_draft.chemical_name or ""
    else:
        _replace_parent_parameters_from_child(parent_draft, child_draft)
        _merge_child_sections_into_parent(
            parent_draft,
            child_draft,
            section_names=["existing_spec", "proposed_changes", "justification", "recommendation"],
            stream_name=stream_name,
        )

    parent_draft.updated_at = datetime.now(timezone.utc)

    for sibling in _get_active_revisions_for_parent(parent_draft.id):
        if sibling.id == revision.id or sibling.child_draft is None:
            continue
        if stream_name == MATERIAL_HANDLING_STREAM:
            sibling.child_draft.spec_number = parent_draft.spec_number
            sibling.child_draft.chemical_name = parent_draft.chemical_name
            sibling.child_draft.updated_at = datetime.now(timezone.utc)

    return master_payload


def _build_revision_review_context(revision: CSCRevision) -> dict[str, object]:
    """Assemble a read-only snapshot for the admin review modal."""
    draft = revision.child_draft
    parent_draft = revision.parent_draft
    master_values = get_master_form_values(draft)
    material_values = get_material_properties_values(draft)
    storage_values = get_storage_handling_values(draft)
    parent_master_values = get_master_form_values(parent_draft) if parent_draft is not None else {}
    parent_material_values = get_material_properties_values(parent_draft) if parent_draft is not None else {}
    parent_storage_values = get_storage_handling_values(parent_draft) if parent_draft is not None else {}
    blank_supporting_baseline = _should_blank_legacy_supporting_baseline(parent_draft)
    blank_master_labels = (
        {"Storage Conditions - General", "Primary Storage Classification"}
        if blank_supporting_baseline else set()
    )
    workflow_scope = _load_workflow_scope(draft)
    workflow_stream = _infer_workflow_stream_name(draft)

    summary_rows = [
        {"label": "Specification", "value": _display_review_value(draft.spec_number)},
        {"label": "Chemical", "value": _display_review_value(draft.chemical_name)},
        {"label": "Subset", "value": _display_review_value(draft.subset_display)},
        {"label": "Workflow Stream", "value": _workflow_scope_display_label(draft)},
        {"label": "Draft Type", "value": _draft_type_label(draft)},
        {"label": "Version", "value": f"v{draft.version}"},
        {"label": "Material Code", "value": _display_review_value(draft.material_code)},
        {
            "label": "Physical State",
            "value": _display_review_value(master_values.get("physical_state")),
        },
        {"label": "Submitted By", "value": _display_review_value(revision.submitted_by)},
        {
            "label": "Submitted At",
            "value": revision.submitted_at,
            "is_datetime": True,
        },
        {
            "label": "Authorized By",
            "value": _display_review_value(revision.authorized_by_name),
        },
        {
            "label": "Subcommittee Head",
            "value": _display_review_value(revision.subcommittee_head_name),
        },
        {
            "label": "Authorization Confirmed",
            "value": _display_review_value(revision.authorization_confirmed),
        },
        {
            "label": "Current Status",
            "value": _display_review_value((revision.status or "").replace("_", " ").title()),
        },
        {
            "label": "Open for Revision",
            "value": _display_review_value(", ".join(_summarize_workflow_scope(workflow_scope))),
        },
        {
            "label": "Committee Head Review",
            "value": _display_review_value(
                revision.committee_head_user.username
                if revision.committee_head_user
                else None
            ),
        },
        {
            "label": "Committee Head Reviewed At",
            "value": revision.committee_head_reviewed_at,
            "is_datetime": True,
        },
        {
            "label": "Material Master Admin Review",
            "value": _display_review_value(
                revision.module_admin_user.username
                if revision.module_admin_user
                else None
            ),
        },
        {
            "label": "Material Master Admin Reviewed At",
            "value": revision.module_admin_reviewed_at,
            "is_datetime": True,
        },
    ]

    sections = _build_section_review_rows_for_draft(draft, parent_draft=parent_draft)

    parameter_rows = _build_parameter_review_rows_for_draft(draft, parent_draft=parent_draft)

    impact_rows = _build_comparison_value_rows(
        _build_impact_review_rows(draft, master_values),
        ([] if blank_supporting_baseline else _build_impact_review_rows(parent_draft, parent_master_values))
        if parent_draft is not None else None,
    )

    return {
        "draft": draft,
        "summary_rows": summary_rows,
        "section_rows": sections,
        "parameter_rows": parameter_rows,
        "master_rows": _build_comparison_value_rows(
            _build_labeled_value_rows(
                [("material_code", "Material Code"), *MASTER_DATA_FIELDS, *MASTER_EXTRA_FIELDS],
                master_values,
            ),
            _blank_comparison_source_labels(
                _build_labeled_value_rows(
                    [("material_code", "Material Code"), *MASTER_DATA_FIELDS, *MASTER_EXTRA_FIELDS],
                    parent_master_values,
                ) if parent_draft is not None else None,
                blank_master_labels,
            ),
        ),
        "material_property_rows": _build_comparison_value_rows(
            [
                {
                    "label": field["label"],
                    "value": _display_review_value(material_values.get(field["field_name"])),
                }
                for field in get_material_properties_fields()
                if _display_review_value(material_values.get(field["field_name"])) != "—"
            ],
            ([] if blank_supporting_baseline else [
                {
                    "label": field["label"],
                    "value": _display_review_value(parent_material_values.get(field["field_name"])),
                }
                for field in get_material_properties_fields()
                if _display_review_value(parent_material_values.get(field["field_name"])) != "—"
            ]) if parent_draft is not None else None,
        ),
        "storage_rows": _build_comparison_value_rows(
            [
                {
                    "label": field["label"],
                    "value": _display_review_value(storage_values.get(field["field_name"])),
                }
                for field in get_storage_handling_fields()
                if _display_review_value(storage_values.get(field["field_name"])) != "—"
            ],
            ([] if blank_supporting_baseline else [
                {
                    "label": field["label"],
                    "value": _display_review_value(parent_storage_values.get(field["field_name"])),
                }
                for field in get_storage_handling_fields()
                if _display_review_value(parent_storage_values.get(field["field_name"])) != "—"
            ]) if parent_draft is not None else None,
        ),
        "impact_rows": impact_rows,
    }


def _build_draft_export_review_context(draft: CSCDraft) -> dict[str, object]:
    """Build a Flask-style review snapshot for workflow Word export."""
    revision = _get_revision_for_child(draft.id)
    if revision is not None:
        context = _build_revision_review_context(revision)
        context["revision"] = revision
        return context

    master_values = get_master_form_values(draft)
    material_values = get_material_properties_values(draft)
    storage_values = get_storage_handling_values(draft)
    workflow_scope = _load_workflow_scope(draft)
    workflow_stream = _infer_workflow_stream_name(draft)
    parent_draft = db.session.get(CSCDraft, draft.parent_draft_id) if draft.parent_draft_id else None
    parent_master_values = get_master_form_values(parent_draft) if parent_draft is not None else {}
    parent_material_values = get_material_properties_values(parent_draft) if parent_draft is not None else {}
    parent_storage_values = get_storage_handling_values(parent_draft) if parent_draft is not None else {}
    blank_supporting_baseline = _should_blank_legacy_supporting_baseline(parent_draft)
    hide_legacy_supporting_sections = _should_hide_legacy_supporting_sections(draft)
    blank_master_labels = (
        {"Storage Conditions - General", "Primary Storage Classification"}
        if blank_supporting_baseline else set()
    )

    summary_rows = [
        {"label": "Specification", "value": _display_review_value(draft.spec_number)},
        {"label": "Chemical", "value": _display_review_value(draft.chemical_name)},
        {"label": "Subset", "value": _display_review_value(draft.subset_display)},
        {"label": "Workflow Stream", "value": _workflow_scope_display_label(draft)},
        {"label": "Draft Type", "value": _draft_type_label(draft)},
        {"label": "Version", "value": f"v{draft.version}"},
        {"label": "Material Code", "value": _display_review_value(draft.material_code)},
        {
            "label": "Physical State",
            "value": _display_review_value(master_values.get("physical_state")),
        },
        {"label": "Prepared By", "value": _display_review_value(draft.prepared_by)},
        {"label": "Reviewed By", "value": _display_review_value(draft.reviewed_by)},
        {"label": "Current Status", "value": _display_review_value(draft.status)},
        {
            "label": "Open for Revision",
            "value": _display_review_value(", ".join(_summarize_workflow_scope(workflow_scope))),
        },
    ]

    section_rows = _build_section_review_rows_for_draft(draft, parent_draft=parent_draft)

    parameter_rows = _build_parameter_review_rows_for_draft(draft, parent_draft=parent_draft)

    current_impact_rows = [] if hide_legacy_supporting_sections else _build_impact_review_rows(draft, master_values)
    impact_rows = _build_comparison_value_rows(
        current_impact_rows,
        ([] if blank_supporting_baseline else _build_impact_review_rows(parent_draft, parent_master_values))
        if parent_draft is not None else None,
    )

    return {
        "draft": draft,
        "summary_rows": summary_rows,
        "section_rows": section_rows,
        "parameter_rows": parameter_rows,
        "master_rows": _build_comparison_value_rows(
            _build_labeled_value_rows(
                [("material_code", "Material Code"), *MASTER_DATA_FIELDS, *MASTER_EXTRA_FIELDS],
                master_values,
            ),
            _blank_comparison_source_labels(
                _build_labeled_value_rows(
                    [("material_code", "Material Code"), *MASTER_DATA_FIELDS, *MASTER_EXTRA_FIELDS],
                    parent_master_values,
                ) if parent_draft is not None else None,
                blank_master_labels,
            ),
        ),
        "material_property_rows": _build_comparison_value_rows(
            ([] if hide_legacy_supporting_sections else [
                {
                    "label": field["label"],
                    "value": _display_review_value(material_values.get(field["field_name"])),
                }
                for field in get_material_properties_fields()
                if _display_review_value(material_values.get(field["field_name"])) != "—"
            ]),
            ([] if blank_supporting_baseline else [
                {
                    "label": field["label"],
                    "value": _display_review_value(parent_material_values.get(field["field_name"])),
                }
                for field in get_material_properties_fields()
                if _display_review_value(parent_material_values.get(field["field_name"])) != "—"
            ]) if parent_draft is not None else None,
        ),
        "storage_rows": _build_comparison_value_rows(
            ([] if hide_legacy_supporting_sections else [
                {
                    "label": field["label"],
                    "value": _display_review_value(storage_values.get(field["field_name"])),
                }
                for field in get_storage_handling_fields()
                if _display_review_value(storage_values.get(field["field_name"])) != "—"
            ]),
            ([] if blank_supporting_baseline else [
                {
                    "label": field["label"],
                    "value": _display_review_value(parent_storage_values.get(field["field_name"])),
                }
                for field in get_storage_handling_fields()
                if _display_review_value(parent_storage_values.get(field["field_name"])) != "—"
            ]) if parent_draft is not None else None,
        ),
        "impact_rows": impact_rows,
        "revision": revision,
        "latest_review_notes": (
            revision.module_admin_notes or revision.committee_head_notes or revision.reviewer_notes
        ) if revision is not None else "",
    }


def _get_admin_stats() -> dict:
    """Get admin dashboard statistics."""
    try:
        return {
            "published": CSCDraft.query.filter_by(status="Published", parent_draft_id=None).count(),
            "drafting": CSCDraft.query.filter_by(status="Drafting", parent_draft_id=None).count(),
            "submitted_to_head": CSCRevision.query.filter_by(status=WORKFLOW_DRAFTING_SUBMITTED).count(),
            "pending_admin": CSCRevision.query.filter_by(status=WORKFLOW_DRAFTING_HEAD_APPROVED).count(),
            "returned": CSCRevision.query.filter_by(status=WORKFLOW_DRAFTING_RETURNED).count(),
        }
    except Exception:
        return {"published": 0, "drafting": 0, "submitted_to_head": 0, "pending_admin": 0, "returned": 0}


def _get_export_stats() -> dict:
    """Get export page statistics."""
    try:
        published_drafts = (
            CSCDraft.query
            .filter_by(status="Published", parent_draft_id=None)
            .all()
        )
        published_count = len(published_drafts)
        subsets = {
            parse_spec_number(draft.spec_number or "")[0]
            for draft in published_drafts
            if parse_spec_number(draft.spec_number or "")[0]
        }
        total_parameters = db.session.query(
            db.func.count(CSCParameter.id)
        ).join(CSCDraft).filter(CSCDraft.status == "Published", CSCDraft.parent_draft_id.is_(None)).scalar() or 0
        last_export_at = (
            db.session.query(db.func.max(CSCAudit.action_time))
            .filter(CSCAudit.action.in_(["DOCX_EXPORT", "MASTER_DOCX_EXPORT"]))
            .scalar()
        )
        return {
            "published_count": published_count,
            "subset_count": len(subsets),
            "total_parameters": total_parameters,
            "last_export_at": last_export_at,
        }
    except Exception:
        return {
            "published_count": 0,
            "subset_count": 0,
            "total_parameters": 0,
            "last_export_at": None,
        }


def _get_export_subset_options() -> list[dict[str, object]]:
    """Return published subset choices with counts in canonical order."""
    published_drafts = (
        CSCDraft.query
        .filter_by(status="Published", parent_draft_id=None)
        .all()
    )
    counts: dict[str, int] = {}
    for draft in published_drafts:
        subset_code, _, _ = parse_spec_number(draft.spec_number or "")
        if not subset_code:
            continue
        counts[subset_code] = counts.get(subset_code, 0) + 1

    options = []
    for code in SPEC_SUBSET_ORDER:
        if code not in counts:
            continue
        options.append(
            {
                "code": code,
                "label": SPEC_SUBSET_LABELS.get(code, code),
                "count": counts[code],
            }
        )
    return options

def _serialize_draft_for_export(draft: CSCDraft) -> dict[str, object]:
    payload = draft.to_dict()
    payload["version_display"] = f"v{format_spec_version(draft.spec_version)}"
    return payload


def _serialize_parameters_for_export(draft: CSCDraft) -> list[dict[str, object]]:
    rows = []
    for parameter in draft.parameters.order_by(CSCParameter.sort_order).all():
        existing_value = format_required_value(
            parameter.required_value_type or "text",
            parameter.required_value_text or parameter.existing_value or "",
            parameter.required_value_operator_1,
            parameter.required_value_value_1,
            parameter.required_value_operator_2,
            parameter.required_value_value_2,
        ) or (parameter.existing_value or "")
        rows.append(
            {
                "id": parameter.id,
                "parameter_name": parameter.parameter_name or "",
                "parameter_type": normalize_parameter_type_label(parameter.parameter_type or ""),
                "existing_value": existing_value,
                "proposed_value": parameter.proposed_value or "",
                "test_method": parameter.test_procedure_text or parameter.test_method or "",
                "sort_order": parameter.sort_order,
            }
        )
    return rows


def _serialize_sections_for_export(draft: CSCDraft) -> dict[str, str]:
    sections: dict[str, str] = {}
    for section in draft.sections.order_by(CSCSection.sort_order).all():
        canonical_name = _canonical_section_name(section.section_name)
        section_text = section.section_text or ""
        existing_text = sections.get(canonical_name, "")
        if (
            canonical_name not in sections
            or ((not existing_text.strip() or _is_exact_testing_placeholder(existing_text)) and section_text.strip())
        ):
            sections[canonical_name] = section_text
    return sections


def _serialize_flags_for_export(draft: CSCDraft) -> list[dict[str, object]]:
    return [
        {
            "issue_type": flag.issue_type,
            "is_present": bool(flag.is_present),
            "note": flag.note or "",
            "sort_order": flag.sort_order,
        }
        for flag in draft.issue_flags.order_by(CSCIssueFlag.sort_order).all()
    ]


def _serialize_impact_for_export(draft: CSCDraft) -> dict[str, object] | None:
    if not draft.impact_analysis:
        return None
    return draft.impact_analysis.to_dict()


def _build_export_bundle(draft: CSCDraft) -> dict[str, object]:
    return {
        "draft": _serialize_draft_for_export(draft),
        "sections": _serialize_sections_for_export(draft),
        "parameters": _serialize_parameters_for_export(draft),
        "flags": _serialize_flags_for_export(draft),
        "impact_analysis": _serialize_impact_for_export(draft),
    }


def _serialize_review_context_for_snapshot(review_context: dict[str, object]) -> dict[str, object]:
    """Convert a live review context into a JSON-safe payload for version snapshots."""
    draft = review_context.get("draft") or {}
    if hasattr(draft, "to_dict"):
        draft_payload = draft.to_dict()
    elif isinstance(draft, dict):
        draft_payload = dict(draft)
    else:
        draft_payload = {}

    summary_rows = []
    for row in review_context.get("summary_rows", []) or []:
        if not isinstance(row, dict):
            continue
        normalized_row = dict(row)
        value = normalized_row.get("value")
        if isinstance(value, datetime):
            normalized_row["value"] = value.isoformat()
        summary_rows.append(normalized_row)

    return {
        "draft": draft_payload,
        "summary_rows": summary_rows,
        "section_rows": list(review_context.get("section_rows", []) or []),
        "parameter_rows": list(review_context.get("parameter_rows", []) or []),
        "master_rows": list(review_context.get("master_rows", []) or []),
        "material_property_rows": list(review_context.get("material_property_rows", []) or []),
        "storage_rows": list(review_context.get("storage_rows", []) or []),
        "impact_rows": list(review_context.get("impact_rows", []) or []),
        "latest_review_notes": (
            review_context.get("latest_review_notes")
            or getattr(review_context.get("revision"), "reviewer_notes", "")
            or ""
        ),
    }


def _override_review_context_summary_value(
    review_context: dict[str, object],
    label: str,
    value: object,
) -> None:
    """Update a summary-row value in place when capturing published snapshots."""
    summary_rows = review_context.get("summary_rows")
    if not isinstance(summary_rows, list):
        return
    for row in summary_rows:
        if isinstance(row, dict) and row.get("label") == label:
            row["value"] = value
            return


def _load_spec_version_snapshot_payload(snapshot: CSCSpecVersion | None) -> dict[str, object]:
    """Parse a version snapshot payload safely."""
    if snapshot is None or not (snapshot.payload_json or "").strip():
        return {}
    try:
        payload = json.loads(snapshot.payload_json or "{}")
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _extract_review_context_from_snapshot_payload(payload: dict[str, object] | None) -> dict[str, object] | None:
    """Return stored review-context payload when available."""
    if not isinstance(payload, dict):
        return None
    review_context = payload.get("review_context")
    return review_context if isinstance(review_context, dict) else None


def _snapshot_has_exact_published_dossier(snapshot: CSCSpecVersion | None) -> bool:
    """Return True only when the snapshot stores the original published review dossier."""
    return _extract_review_context_from_snapshot_payload(
        _load_spec_version_snapshot_payload(snapshot)
    ) is not None


def _build_published_version_map(draft: CSCDraft) -> list[dict[str, object]]:
    """Build ordered version-history nodes for the published spec detail page."""
    rows: list[dict[str, object]] = []
    seen_versions: set[int] = set()
    snapshots = draft.spec_versions.order_by(
        CSCSpecVersion.spec_version.asc(),
        CSCSpecVersion.created_at.asc(),
    ).all()

    for snapshot in snapshots:
        has_review_context = _snapshot_has_exact_published_dossier(snapshot)
        is_current = snapshot.spec_version == draft.spec_version
        rows.append(
            {
                "spec_version": snapshot.spec_version,
                "version_label": format_spec_version(snapshot.spec_version),
                "created_at": snapshot.created_at,
                "created_by": snapshot.created_by or "system",
                "source_action": snapshot.source_action or "",
                "remarks": snapshot.remarks or "",
                "is_current": is_current,
                "can_download_dossier": has_review_context,
                "availability_note": (
                    "Exact published dossier available"
                    if has_review_context
                    else "Legacy version snapshot does not include a dossier payload"
                ),
            }
        )
        seen_versions.add(snapshot.spec_version)

    if draft.spec_version not in seen_versions:
        rows.append(
            {
                "spec_version": draft.spec_version,
                "version_label": format_spec_version(draft.spec_version),
                "created_at": draft.updated_at,
                "created_by": draft.reviewed_by or draft.prepared_by or "system",
                "source_action": "current_published_state",
                "remarks": "",
                "is_current": True,
                "can_download_dossier": False,
                "availability_note": "Exact published dossier was not retained for this version",
            }
        )

    rows.sort(key=lambda row: int(row["spec_version"]))
    return rows


def _get_version_changelog() -> list[dict[str, object]]:
    rows = (
        db.session.query(
            CSCAudit.draft_id,
            CSCDraft.spec_number,
            CSCDraft.chemical_name,
            CSCDraft.spec_version,
            CSCAudit.new_value.label("version_label"),
            CSCAudit.action,
            CSCAudit.action_time,
            CSCAudit.user_name,
            CSCAudit.remarks,
        )
        .join(CSCDraft, CSCDraft.id == CSCAudit.draft_id)
        .filter(
            CSCAudit.action.in_(
                [
                    "VERSION_INCREMENT",
                    "ADMIN_STAGE_PUBLISHED",
                    "REVISION_APPROVED",
                    "RESTORE_VERSION",
                    "ADMIN_PDF_IMPORT",
                    "CREATE_DRAFT",
                ]
            ),
            CSCDraft.is_admin_draft.is_(True),
            CSCDraft.parent_draft_id.is_(None),
        )
        .order_by(CSCAudit.action_time.desc())
        .all()
    )
    return [
        {
            "draft_id": row.draft_id,
            "spec_number": row.spec_number,
            "chemical_name": row.chemical_name,
            "spec_version": row.spec_version,
            "version_label": row.version_label,
            "action": row.action,
            "action_time": row.action_time.isoformat() if row.action_time else "",
            "user_name": row.user_name,
            "remarks": row.remarks,
        }
        for row in rows
    ]


def _record_export_audit(action: str, draft_id: int | None, new_value: str) -> None:
    try:
        audit = CSCAudit(
            draft_id=draft_id,
            action=action,
            user_name=current_user.username,
            action_time=datetime.now(timezone.utc),
            new_value=new_value,
        )
        db.session.add(audit)
        db.session.commit()
    except Exception as exc:
        logger.warning("Failed to record CSC export audit: %s", exc)
        db.session.rollback()


def _next_identity_value(max_id: int | None) -> int:
    return max(1, int(max_id or 0) + 1)


def _csc_reset_identity_targets(
    remaining_max_draft_id: int | None,
    remaining_max_revision_id: int | None = None,
    remaining_max_audit_id: int | None = None,
) -> dict[str, int]:
    return {
        "csc_drafts": _next_identity_value(remaining_max_draft_id),
        "csc_revisions": _next_identity_value(remaining_max_revision_id),
        "csc_audit": _next_identity_value(remaining_max_audit_id),
    }


def _reset_table_identity(table_name: str, next_id: int) -> None:
    """Reset the next auto-increment / sequence value for supported databases."""
    bind = db.session.get_bind()
    if bind is None:
        return

    dialect_name = (bind.dialect.name or "").lower()
    target_id = max(1, int(next_id or 1))

    if dialect_name in {"mysql", "mariadb"}:
        db.session.execute(
            text(f"ALTER TABLE {table_name} AUTO_INCREMENT = :next_id"),
            {"next_id": target_id},
        )
        return

    if dialect_name == "postgresql":
        db.session.execute(
            text(
                "SELECT setval(pg_get_serial_sequence(:table_name, 'id'), :next_value, false)"
            ),
            {"table_name": table_name, "next_value": target_id},
        )
        return

    if dialect_name == "sqlite":
        try:
            db.session.execute(
                text("DELETE FROM sqlite_sequence WHERE name = :table_name"),
                {"table_name": table_name},
            )
            if target_id > 1:
                db.session.execute(
                    text(
                        "INSERT INTO sqlite_sequence(name, seq) VALUES (:table_name, :seq)"
                    ),
                    {"table_name": table_name, "seq": target_id - 1},
                )
        except Exception:
            logger.debug(
                "SQLite sequence reset skipped for table %s",
                table_name,
                exc_info=True,
            )


def _reset_csc_test_workflow_state() -> dict[str, int]:
    """Clear test-only workflow state while preserving published parent drafts."""
    child_drafts = (
        CSCDraft.query.filter(CSCDraft.parent_draft_id.isnot(None))
        .order_by(CSCDraft.id.desc())
        .all()
    )
    touched_parent_ids = {
        child.parent_draft_id
        for child in child_drafts
        if child.parent_draft_id is not None
    }

    audit_count = CSCAudit.query.count()
    revision_count = CSCRevision.query.count()
    child_draft_count = len(child_drafts)

    if revision_count:
        CSCRevision.query.delete(synchronize_session=False)
    if audit_count:
        CSCAudit.query.delete(synchronize_session=False)
    for child_draft in child_drafts:
        db.session.delete(child_draft)

    db.session.flush()

    for parent_id in touched_parent_ids:
        parent_draft = db.session.get(CSCDraft, parent_id)
        if parent_draft is None or parent_draft.parent_draft_id is not None:
            continue
        parent_draft.status = "Published"
        parent_draft.admin_stage = ADMIN_STAGE_PUBLISHED
        parent_draft.updated_at = datetime.now(timezone.utc)

    db.session.flush()

    remaining_max_draft_id = db.session.query(db.func.max(CSCDraft.id)).scalar()
    identity_targets = _csc_reset_identity_targets(remaining_max_draft_id)
    for table_name, next_id in identity_targets.items():
        _reset_table_identity(table_name, next_id)

    db.session.commit()
    return {
        "audit_entries_deleted": audit_count,
        "revisions_deleted": revision_count,
        "child_drafts_deleted": child_draft_count,
        "next_draft_id": identity_targets["csc_drafts"],
    }


def _get_spec_subsets() -> list[dict[str, str]]:
    """Get distinct spec subsets for filter dropdowns in canonical order."""
    try:
        rows = db.session.query(CSCDraft.spec_number).all()
        subsets = set()
        for (spec_num,) in rows:
            subset_code, _, _ = parse_spec_number(spec_num or "")
            if subset_code:
                subsets.add(subset_code)

        ordered_codes = [code for code in SPEC_SUBSET_ORDER if code in subsets]
        extra_codes = sorted(code for code in subsets if code not in SPEC_SUBSET_ORDER)
        return [
            {
                "code": code,
                "label": SPEC_SUBSET_LABELS.get(code, code),
                "display": _subset_display(code) or code,
            }
            for code in [*ordered_codes, *extra_codes]
        ]
    except Exception:
        return []


# ──────────────────────────────────────────────────────────────────────────────
# Landing & Overview
# ──────────────────────────────────────────────────────────────────────────────


@csc_bp.route("/")
@login_required
@module_access_required("csc")
def index():
    """CSC landing page with KPI cards and navigation."""
    return render_template(
        "csc/landing.html",
        committee_directory=get_committee_directory(),
        committee_tree=get_committee_tree(),
        office_orders=get_office_orders(),
        material_master_admin_labels=_get_material_master_admin_labels(),
        can_access_committee_user_workbench=(
            bool(_get_current_user_committee_user_slugs())
        ),
        can_access_review_workbench=(
            _current_user_is_governance_committee_member()
            or bool(_get_current_user_committee_head_slugs())
        ),
        is_material_master_admin=_current_user_is_material_master_admin(),
    )


# Alias so templates using url_for('csc.landing') also work
@csc_bp.route("/landing")
@login_required
@module_access_required("csc")
def landing():
    """Alias for index – CSC landing page."""
    return render_template(
        "csc/landing.html",
        committee_directory=get_committee_directory(),
        committee_tree=get_committee_tree(),
        office_orders=get_office_orders(),
        material_master_admin_labels=_get_material_master_admin_labels(),
        can_access_committee_user_workbench=(
            bool(_get_current_user_committee_user_slugs())
        ),
        can_access_review_workbench=(
            _current_user_is_governance_committee_member()
            or bool(_get_current_user_committee_head_slugs())
        ),
        is_material_master_admin=_current_user_is_material_master_admin(),
    )


def _row_material_code(row) -> str:
    if isinstance(row, dict):
        return str(row.get("material") or "").strip()
    return str(getattr(row, "material", "") or "").strip()


def _row_short_text(row) -> str:
    if isinstance(row, dict):
        return str(row.get("short_text") or "").strip()
    return str(getattr(row, "short_text", "") or "").strip()


def _build_msds_material_selector_options(rows) -> list[dict[str, str]]:
    """Return MSDS selector options ordered by published spec subset order."""
    material_rows = {
        _row_material_code(row): row
        for row in rows
        if _row_material_code(row)
    }

    options: list[dict[str, str]] = []
    seen_codes: set[str] = set()

    try:
        published_specs = sorted(
            CSCDraft.query.filter(
                CSCDraft.parent_draft_id.is_(None),
                CSCDraft.admin_stage == ADMIN_STAGE_PUBLISHED,
                CSCDraft.material_code.isnot(None),
                CSCDraft.material_code != "",
            ).all(),
            key=_draft_sort_key,
        )
    except Exception:
        logger.exception("Failed to build published-spec selector options for MSDS Center")
        published_specs = []

    for draft in published_specs:
        material_code = (draft.material_code or "").strip()
        if not material_code or material_code in seen_codes:
            continue

        row = material_rows.get(material_code)
        short_text = _row_short_text(row) if row is not None else ""
        description = short_text or (draft.chemical_name or "").strip() or "—"

        options.append(
            {
                "value": material_code,
                "spec_number": (draft.spec_number or "").strip(),
                "material_code": material_code,
                "description": description,
            }
        )
        seen_codes.add(material_code)

    for material_code in sorted(code for code in material_rows if code not in seen_codes):
        row = material_rows.get(material_code)
        options.append(
            {
                "value": material_code,
                "spec_number": "",
                "material_code": material_code,
                "description": _row_short_text(row) or "—",
            }
        )

    return options


@csc_bp.route("/msds")
@login_required
@module_access_required("csc")
def msds_page():
    """MSDS center within Material Master Management."""
    from app.core.services.msds_service import (
        MSDSError,
        get_msds_material_index,
        get_msds_slot_options,
    )

    rows = []
    try:
        rows = get_all_master_data()
    except Exception:
        logger.exception("Failed to load material master rows for CSC MSDS Center")
        flash("Material master rows could not be loaded right now.", "warning")
    try:
        msds_by_material = get_msds_material_index(
            [_row_material_code(row) for row in rows if _row_material_code(row)]
        )
    except MSDSError as exc:
        flash(str(exc), "warning")
        msds_by_material = {}
    return render_template(
        "csc/msds.html",
        rows=rows,
        material_selector_options=_build_msds_material_selector_options(rows),
        msds_slot_options=get_msds_slot_options(),
        total=len(rows),
        msds_by_material=msds_by_material,
        msds_count=sum(len(files) for files in msds_by_material.values()),
        msds_material_total=len(msds_by_material),
        prefill_material_code=(request.args.get("material_code") or "").strip(),
        prefill_slot_code=(request.args.get("slot_code") or "").strip().lower() or "standard",
    )


@csc_bp.route("/msds/upload", methods=["POST"])
@login_required
@module_access_required("csc")
def upload_msds():
    """Upload an MSDS PDF from Material Master Management."""
    from app.core.services.msds_service import MSDSError, store_msds_document

    material_code = request.form.get("material_code", "").strip()
    slot_code = request.form.get("slot_code", "").strip().lower()
    file_obj = request.files.get("msds_file")

    try:
        msds_file = store_msds_document(
            material_code=material_code,
            file_obj=file_obj,
            slot_code=slot_code,
        )
        db.session.commit()
    except MSDSError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
        return redirect(url_for("csc.msds_page", material_code=material_code, slot_code=slot_code))
    except Exception:
        db.session.rollback()
        logger.exception("Failed to upload MSDS PDF for material=%s", material_code)
        flash("Could not upload the MSDS PDF.", "danger")
        return redirect(url_for("csc.msds_page", material_code=material_code, slot_code=slot_code))

    action = "replaced" if getattr(msds_file, "storage_action", "") == "replaced" else "stored"
    flash(
        f"{msds_file.slot_label} {action} for material '{material_code}'.",
        "success",
    )
    return redirect(url_for("csc.msds_page", material_code=material_code, slot_code=slot_code))


@csc_bp.route("/msds/<int:file_id>")
@login_required
@module_access_required("csc")
def open_msds(file_id: int):
    """Open or download an MSDS PDF from Material Master Management."""
    from app.core.services.msds_service import MSDSError, MSDSNotFoundError, get_msds_file

    download = (request.args.get("download") or "").strip().lower() in {"1", "true", "yes"}
    try:
        document = get_msds_file(file_id, include_data=True)
    except MSDSNotFoundError as exc:
        abort(404, description=str(exc))
    except MSDSError as exc:
        abort(500, description=str(exc))

    return send_file(
        io.BytesIO(document.data),
        mimetype=document.content_type or "application/pdf",
        as_attachment=download,
        download_name=document.filename,
        max_age=0,
    )


@csc_bp.route("/msds/<int:file_id>/delete", methods=["POST"])
@login_required
@module_access_required("csc")
def delete_msds(file_id: int):
    """Delete a stored MSDS PDF from Material Master Management."""
    from app.core.services.msds_service import MSDSError, delete_msds_file, get_msds_file

    try:
        msds_file = get_msds_file(file_id)
        deleted = delete_msds_file(file_id)
        if not deleted:
            flash("The selected MSDS file no longer exists.", "info")
            return redirect(url_for("csc.msds_page"))
        db.session.commit()
    except MSDSError as exc:
        db.session.rollback()
        flash(str(exc), "danger")
        return redirect(url_for("csc.msds_page"))
    except Exception:
        db.session.rollback()
        logger.exception("Failed to delete MSDS PDF for file_id=%s", file_id)
        flash("Could not delete the MSDS PDF.", "danger")
        return redirect(url_for("csc.msds_page"))

    flash(
        f"{msds_file.slot_label} '{msds_file.filename}' deleted for material '{msds_file.material_code}'.",
        "success",
    )
    return redirect(url_for("csc.msds_page", material_code=msds_file.material_code))


@csc_bp.route("/office-orders/<slug>")
@login_required
@module_access_required("csc")
def office_order(slug: str):
    """Stream office-order PDFs stored in the database for committee governance."""
    download = request.args.get("download", "").strip().lower() in {"1", "true", "yes"}

    try:
        from app.models.csc.governance import CSCOfficeOrderFile
        order_file = db.session.get(CSCOfficeOrderFile, slug)
        if order_file:
            from io import BytesIO
            return send_file(
                BytesIO(order_file.file_data),
                as_attachment=download,
                download_name=order_file.file_name,
                mimetype="application/pdf",
            )
    except Exception:
        logger.exception("Failed to load office order from database for slug=%s", slug)
        return redirect(url_for("csc.index"))

    flash("Office order PDF is not uploaded in Governance Settings.", "danger")
    return redirect(url_for("csc.index"))


@csc_bp.route("/api/overview")
@login_required
@module_access_required("csc")
def api_overview():
    """JSON KPIs for landing page."""
    try:
        total_specs = CSCDraft.query.filter_by(
            parent_draft_id=None,
        ).count()
        open_drafting = CSCRevision.query.filter(
            CSCRevision.status.in_([WORKFLOW_DRAFTING_OPEN, WORKFLOW_DRAFTING_RETURNED])
        ).count()
        submitted = CSCRevision.query.filter(
            CSCRevision.status.in_([WORKFLOW_DRAFTING_SUBMITTED, WORKFLOW_DRAFTING_HEAD_APPROVED])
        ).count()
        published = CSCDraft.query.filter_by(
            parent_draft_id=None,
            status="Published",
        ).count()

        return jsonify({
            "total_specs": total_specs,
            "under_review": submitted,
            "pending_revisions": open_drafting,
            "published": published,
        })
    except Exception:
        logger.exception("Error fetching overview")
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


# ──────────────────────────────────────────────────────────────────────────────
# Workspace (Committee Members)
# ──────────────────────────────────────────────────────────────────────────────


def _render_committee_user_workbench():
    """Render the committee-user drafting workbench."""
    try:
        if not _can_current_user_access_committee_user_workbench():
            abort(403)

        global_visibility = current_user.is_super_user()
        allowed_subset_codes = _get_effective_committee_user_subset_codes()
        subset_options = _get_spec_subsets()
        if allowed_subset_codes and not global_visibility:
            subset_options = [
                subset for subset in subset_options
                if subset["code"] in set(allowed_subset_codes)
            ]
        response = make_response(
            render_template(
                "csc/workspace.html",
                drafts=[],
                q=request.args.get("q", "").strip(),
                subset_filter=request.args.get("subset", "").strip(),
                status_filter=request.args.get("status", "").strip().lower(),
                subset_options=subset_options,
                allowed_subset_codes=allowed_subset_codes,
                allowed_subset_labels=[
                    _subset_display(code) or code for code in allowed_subset_codes
                ],
                user_committee_titles=_get_current_user_committee_access().get("committee_titles", []),
                can_create_scoped_draft=bool(_get_committee_user_upload_subset_codes()),
                committee_workbench_has_global_visibility=global_visibility,
                committee_workspace_result_label=(
                    "active drafts in workbench" if global_visibility else "active allotted drafts"
                ),
            )
        )
        return _mark_response_no_store(response)
    except Exception:
        logger.exception("Error loading workspace")
        flash("Error loading workspace.", "danger")
        return redirect(url_for("csc.index"))


@csc_bp.route("/committee-user-workbench")
@login_required
@module_access_required("csc")
def committee_user_workbench():
    """Committee-user drafting workbench filtered by subset coverage."""
    return _render_committee_user_workbench()


@csc_bp.route("/workspace")
@login_required
@module_access_required("csc")
def workspace():
    """Legacy alias for the committee-user drafting workbench."""
    return _render_committee_user_workbench()


@csc_bp.route("/ingest")
@login_required
@module_access_required("csc")
def ingest():
    """Committee-facing ingest page for brand-new specifications."""
    _require_committee_user_upload_scope()
    return render_template("csc/ingest.html", staged_spec=None)


@csc_bp.route("/ingest/pdf", methods=["POST"])
@login_required
@module_access_required("csc")
def ingest_pdf():
    """Extract a CSC Format A PDF and stage it for committee review."""
    _require_committee_user_upload_scope()
    try:
        file = request.files.get("pdf_file")
        if file is None or file.filename == "":
            flash("Select a PDF file to ingest.", "danger")
            return redirect(url_for("csc.ingest"))

        spec = extract_spec_from_pdf(file.read())
        _ensure_current_user_can_access_subset(parse_spec_number(spec.spec_number or "")[0])
        return render_template(
            "csc/ingest.html",
            staged_spec=_build_ingest_preview_spec(spec),
            staged_source_label="PDF",
            staged_filename=(file.filename or "").strip(),
        )
    except PermissionError as e:
        flash(str(e), "danger")
        return redirect(url_for("csc.ingest"))
    except Exception:
        logger.exception("Error ingesting committee PDF")
        db.session.rollback()
        flash(
            str(e) if isinstance(e, (ValueError, ImportError)) else "Error processing PDF file.",
            "danger",
        )
        return redirect(url_for("csc.ingest"))


@csc_bp.route("/ingest/docx", methods=["POST"])
@login_required
@module_access_required("csc")
def ingest_docx():
    """Extract a single CSC Format A DOCX and stage it for committee review."""
    _require_committee_user_upload_scope()
    try:
        file = request.files.get("docx_file")
        if file is None or file.filename == "":
            flash("Select a DOCX file to ingest.", "danger")
            return redirect(url_for("csc.ingest"))

        specs = extract_all_specs_from_docx(file.read())
        if len(specs) != 1:
            raise ValueError(
                "Upload a single CSC Format A specification DOCX. Master/bulk DOCX ingestion remains an admin-only task."
            )

        _ensure_current_user_can_access_subset(parse_spec_number(specs[0].spec_number or "")[0])
        return render_template(
            "csc/ingest.html",
            staged_spec=_build_ingest_preview_spec(specs[0]),
            staged_source_label="DOCX",
            staged_filename=(file.filename or "").strip(),
        )
    except PermissionError as e:
        flash(str(e), "danger")
        return redirect(url_for("csc.ingest"))
    except Exception:
        logger.exception("Error ingesting committee DOCX")
        db.session.rollback()
        flash(str(e) if isinstance(e, ValueError) else "Error processing DOCX file.", "danger")
        return redirect(url_for("csc.ingest"))


@csc_bp.route("/ingest/create-draft", methods=["POST"])
@login_required
@module_access_required("csc")
def create_ingested_draft():
    """Create the workflow draft from the reviewed extraction payload."""
    _require_committee_user_upload_scope()
    try:
        source_label = (request.form.get("source_label") or "Upload").strip()
        spec = _spec_document_from_ingest_form()
        _ensure_current_user_can_access_subset(parse_spec_number(spec.spec_number or "")[0])
        child_draft = _create_new_spec_workflow_from_extraction(spec, source_label)
        db.session.commit()
        _create_audit_entry(
            child_draft.parent_draft_id,
            "New specification draft ingested",
            new_value=json.dumps(
                {
                    "child_draft_id": child_draft.id,
                    "source_label": source_label,
                    "draft_kind": "new_specification",
                }
            ),
        )
        flash(
            f"New specification draft created for {child_draft.spec_number} and moved to Secretary staging.",
            "success",
        )
        return redirect(url_for("csc.ingest"))
    except PermissionError as e:
        flash(str(e), "danger")
        return redirect(url_for("csc.ingest"))
    except Exception:
        logger.exception("Error creating committee draft from staged extraction")
        db.session.rollback()
        flash(str(e) if isinstance(e, ValueError) else "Error creating draft from extracted specification.", "danger")
        return redirect(url_for("csc.ingest"))


@csc_bp.route("/api/drafts")
@login_required
@module_access_required("csc")
def api_drafts():
    """JSON endpoint for filtered draft list."""
    try:
        subset = request.args.get("subset", "").strip()
        search = request.args.get("search", request.args.get("q", "")).strip()
        allowed_committee_slugs = set(_get_current_user_committee_user_slugs())
        can_view_all_drafts = current_user.is_super_user()

        query = CSCDraft.query.filter(CSCDraft.parent_draft_id.isnot(None))
        query = query.join(CSCRevision, CSCRevision.child_draft_id == CSCDraft.id)
        query = query.filter(
            CSCRevision.status.in_(
                [
                    WORKFLOW_DRAFTING_OPEN,
                    WORKFLOW_DRAFTING_RETURNED,
                    WORKFLOW_DRAFTING_SUBMITTED,
                    WORKFLOW_DRAFTING_HEAD_APPROVED,
                ]
            )
        )
        if not allowed_committee_slugs and not can_view_all_drafts:
            return _mark_response_no_store(jsonify({"drafts": [], "total": 0}))

        if search:
            query = query.filter(
                db.or_(
                    CSCDraft.chemical_name.ilike(f"%{search}%"),
                    CSCDraft.spec_number.ilike(f"%{search}%"),
                )
            )

        drafts = []
        for draft in query.all():
            revision = _get_revision_for_child(draft.id)
            committee_slug = _get_revision_committee_slug(revision)
            workflow_subset_code = _workflow_subset_code_for_draft(draft, revision)
            if not can_view_all_drafts and not _draft_in_current_user_committee_scope(draft, revision):
                continue
            if subset and subset != "all" and workflow_subset_code != subset.upper():
                continue
            drafts.append((draft, revision, committee_slug, workflow_subset_code))
        drafts = sorted(drafts, key=lambda item: _draft_sort_key(item[0]))

        response = jsonify({
            "drafts": [
                (
                    lambda blocking_lock, draft, revision, committee_slug, workflow_subset_code: {
                        "id": draft.id,
                        "spec_number": draft.spec_number,
                        "chemical_name": draft.chemical_name,
                        "status": _status_db_to_ui_value(draft.status),
                        "drafting_status": revision.status if revision else "",
                        "draft_type": _draft_type_label(draft),
                        "subset": workflow_subset_code,
                        "subset_label": SPEC_SUBSET_LABELS.get(workflow_subset_code, workflow_subset_code),
                        "subset_display": _subset_display(workflow_subset_code),
                        "committee_title": _get_committee_title_by_slug(committee_slug),
                        "version": draft.version,
                        "can_edit": _can_edit_draft(draft) and not blocking_lock,
                        "can_view_snapshot": bool(revision and _can_current_user_view_revision_snapshot(revision)),
                        "review_url": (
                            url_for("csc.review_revision", revision_id=revision.id)
                            if revision and _can_current_user_view_revision_snapshot(revision)
                            else ""
                        ),
                        "locked_by_other": bool(blocking_lock),
                        "locked_by_label": _editor_lock_holder_label(blocking_lock),
                        "locked_message": _editor_lock_conflict_message(blocking_lock) if blocking_lock else "",
                        "created_at": draft.created_at.isoformat(),
                        "updated_at": draft.updated_at.isoformat(),
                    }
                )(
                    _get_blocking_editor_lock(draft, "committee_user"),
                    draft,
                    revision,
                    committee_slug,
                    workflow_subset_code,
                )
                for draft, revision, committee_slug, workflow_subset_code in drafts
            ],
            "total": len(drafts),
        })
        return _mark_response_no_store(response)
    except Exception:
        logger.exception("Error fetching drafts")
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/workspace/<int:draft_id>")
@login_required
@module_access_required("csc")
def editor(draft_id: int):
    """9-section tabbed editor for a single draft."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            abort(404)

        if not _can_edit_draft(draft):
            flash("You do not have permission to edit this draft.", "danger")
            return redirect(url_for("csc.workspace"))

        editor_mode = "committee_head" if _can_current_user_edit_draft_as_committee_head(draft) else "committee_user"
        if _committee_editor_lock_enabled_for_draft(draft, editor_mode):
            acquired, blocking_lock = _acquire_editor_lock(draft)
            if not acquired:
                flash(_editor_lock_conflict_message(blocking_lock), "info")
                db.session.rollback()
                return redirect(url_for("csc.workspace"))
            db.session.commit()

        workflow_stream = _infer_workflow_stream_name(draft)
        _ensure_default_draft_sections(draft, workflow_stream)
        try:
            db.session.commit()
        except Exception as commit_err:
            logger.warning(f"Could not persist default section setup for draft {draft_id}: {commit_err}")
            db.session.rollback()
        section_defaults = {
            section_name: default_text
            for section_name, default_text, _sort_order in _default_section_rows(workflow_stream)
        }
        section_rows = draft.sections.order_by(CSCSection.sort_order).all()
        sections: dict[str, str] = {}
        for section in section_rows:
            canonical_name = _canonical_section_name(section.section_name)
            section_text = section.section_text or ""
            existing_text = sections.get(canonical_name, "")
            if (
                canonical_name not in sections
                or ((not existing_text.strip() or _is_exact_testing_placeholder(existing_text)) and section_text.strip())
            ):
                sections[canonical_name] = section_text
        for section_name, default_text in section_defaults.items():
            sections[section_name] = sections.get(section_name, "").strip() or default_text
        sections = {
            section_name: _strip_system_bracket_lines(section_text or "")
            for section_name, section_text in sections.items()
        }

        parent_by_sort: dict[int, CSCParameter] = {}
        parent_by_name: dict[str, CSCParameter] = {}
        parent_draft = db.session.get(CSCDraft, draft.parent_draft_id) if draft.parent_draft_id else None
        issues = _build_editor_issue_rows(draft)
        if parent_draft is not None:
            parent_parameters = parent_draft.parameters.order_by(CSCParameter.sort_order).all()
            parent_by_sort = {parameter.sort_order: parameter for parameter in parent_parameters}
            parent_by_name = {
                (parameter.parameter_name or "").strip().lower(): parameter
                for parameter in parent_parameters
                if (parameter.parameter_name or "").strip()
            }

        parameters = []
        for parameter in draft.parameters.order_by(CSCParameter.sort_order).all():
            parent_parameter = _match_parent_parameter(parameter, parent_by_sort, parent_by_name)
            parameters.append(
                {
                    "id": parameter.id,
                    "parameter_name": parameter.parameter_name,
                    "parameter_type": parameter.parameter_type or "Desirable",
                    "source_parameter_type": normalize_parameter_type_label(
                        (parent_parameter.parameter_type if parent_parameter is not None else parameter.parameter_type)
                        or "Desirable"
                    ),
                    "unit_of_measure": parameter.unit_of_measure or "",
                    "specification": parameter.existing_value or "",
                    "required_value_type": parameter.required_value_type or "text",
                    "required_value_text": parameter.required_value_text or parameter.existing_value or "",
                    "required_value_operator_1": parameter.required_value_operator_1 or "",
                    "required_value_value_1": parameter.required_value_value_1 or "",
                    "required_value_operator_2": parameter.required_value_operator_2 or "",
                    "required_value_value_2": parameter.required_value_value_2 or "",
                    "parameter_conditions": parameter.parameter_conditions or "",
                    "test_method": parameter.test_method or "",
                    "test_procedure_type": normalize_test_procedure_type(
                        parameter.test_procedure_type or parameter.test_method or ""
                    ),
                    "test_procedure_text": parameter.test_procedure_text or parameter.test_method or "",
                    "proposed_value": parameter.proposed_value or "",
                    "approved": False,
                }
            )

        checklist_state = build_default_impact_checklist_state(draft.chemical_name)
        impact = {
            "checklist_state": checklist_state,
            "checklist_summary": summarize_impact_checklist_state(checklist_state),
            "legacy_only": False,
        }
        if draft.impact_analysis:
            checklist_state = deserialize_impact_checklist_state(
                draft.impact_analysis.checklist_state_json,
                draft.chemical_name,
            )
            impact = {
                "cost_impact": draft.impact_analysis.operational_impact_score,
                "quality_impact": draft.impact_analysis.safety_environment_score,
                "process_impact": draft.impact_analysis.supply_risk_score,
                "no_substitute": draft.impact_analysis.no_substitute_flag,
                "weighted_score": draft.impact_analysis.impact_score_total,
                "grade": draft.impact_analysis.impact_grade,
                "checklist_state": checklist_state,
                "checklist_summary": summarize_impact_checklist_state(checklist_state),
                "legacy_only": not bool((draft.impact_analysis.checklist_state_json or "").strip()),
            }
        workflow_scope = _load_workflow_scope(draft)
        master_data_editable = _master_data_editable_for_draft(draft)
        comparison_baseline = _build_editor_comparison_baseline(
            draft,
            parent_draft,
            phase1_locked=bool(draft.phase1_locked),
        )

        return render_template(
            "csc/editor.html",
            draft=draft,
            editor_mode=editor_mode,
            can_delete_draft=_can_delete_draft(draft),
            sections=sections,
            issues=issues,
            parameters=parameters,
            parameter_type_options=PARAMETER_TYPES,
            test_procedure_options=TEST_PROCEDURE_OPTIONS,
            impact=impact,
            impact_checklist_flags=IMPACT_CHECKLIST_FLAGS,
            master_data=get_master_form_values(draft),
            material_properties_fields=get_material_properties_fields(),
            material_properties_values=get_material_properties_values(draft),
            storage_handling_fields=get_storage_handling_fields(),
            storage_handling_values=get_storage_handling_values(draft),
            master_fields=MASTER_DATA_FIELDS,
            master_extra_fields=MASTER_EXTRA_FIELDS,
            master_impact_fields=MASTER_IMPACT_FIELDS,
            editable_master_fields=ADMIN_MASTER_FIELD_CONFIGS,
            editable_master_extra_fields=ADMIN_MASTER_EXTRA_FIELDS,
            comparison_baseline=comparison_baseline,
            workflow_scope=workflow_scope,
            workflow_scope_summary=_summarize_workflow_scope(workflow_scope),
            master_data_editable=master_data_editable,
            draft_type_label=_draft_type_label(draft),
            editor_lock_enabled=_committee_editor_lock_enabled_for_draft(draft, editor_mode),
        )
    except Exception:
        logger.exception("Error loading editor")
        flash("Error loading draft editor.", "danger")
        return redirect(url_for("csc.workspace"))


@csc_bp.route("/api/draft/<int:draft_id>/editor-lock/ping", methods=["POST"])
@login_required
@module_access_required("csc")
def ping_editor_lock(draft_id: int):
    """Refresh the live committee-user editor lock for an open draft."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404
        if not _can_edit_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403
        if not _committee_editor_lock_enabled_for_draft(draft, "committee_user"):
            return jsonify({"success": True, "lock_enabled": False})

        acquired, payload = _refresh_editor_lock(draft)
        if not acquired:
            db.session.rollback()
            return _json_editor_lock_conflict_response(draft, "committee_user")

        db.session.commit()
        return jsonify({"success": True, "lock_enabled": True})
    except Exception:
        logger.exception("Error refreshing editor lock")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>/editor-lock/release", methods=["POST"])
@login_required
@module_access_required("csc")
def release_editor_lock(draft_id: int):
    """Release the live committee-user editor lock when the editor closes."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"success": True})
        _release_editor_lock_if_owned(draft)
        db.session.commit()
        return jsonify({"success": True})
    except Exception:
        logger.exception("Error releasing editor lock")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>")
@login_required
@module_access_required("csc")
def api_draft_detail(draft_id: int):
    """Full draft data as JSON."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        if not _can_edit_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403

        parameters = [
            p.to_dict() for p in draft.parameters.order_by(CSCParameter.sort_order).all()
        ]
        sections = [
            s.to_dict() for s in draft.sections.order_by(CSCSection.sort_order).all()
        ]
        issue_flags = [
            f.to_dict() for f in draft.issue_flags.order_by(CSCIssueFlag.sort_order).all()
        ]
        if draft.impact_analysis:
            impact = draft.impact_analysis.to_dict()
        else:
            checklist_state = build_default_impact_checklist_state(draft.chemical_name)
            impact = {
                "checklist_state": checklist_state,
                "checklist_summary": summarize_impact_checklist_state(checklist_state),
            }

        return jsonify({
            "draft": draft.to_dict(),
            "parameters": parameters,
            "sections": sections,
            "issue_flags": issue_flags,
            "impact_analysis": impact,
            "master_data": get_master_form_values(draft),
            "material_properties": get_material_properties_values(draft),
            "storage_handling": get_storage_handling_values(draft),
        })
    except Exception:
        logger.exception("Error fetching draft detail")
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>/master-data", methods=["POST"])
@login_required
@module_access_required("csc")
def save_master_data(draft_id: int):
    """Save linked master-data fields directly from the CSC drafting space."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        if not _can_edit_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403
        if not _master_data_editable_for_draft(draft):
            return jsonify({
                "error": "Master data is read-only for Type Classification drafts.",
            }), 403
        lock_response = _json_editor_lock_conflict_response(draft, "committee_user")
        if lock_response is not None:
            return lock_response

        data = request.get_json() or {}

        if "chemical_name" in data:
            draft.chemical_name = (data.get("chemical_name") or "").strip()

        if (draft.material_code or "").strip():
            data["material_code"] = draft.material_code

        record = upsert_master_record_from_form(draft, data, user_id=current_user.id)
        if record is not None and record not in db.session:
            db.session.add(record)

        draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        _create_audit_entry(draft_id, "Master data saved from drafting workspace")

        return jsonify({
            "success": True,
            "draft": draft.to_dict(),
            "master_data": get_master_form_values(draft),
        })
    except Exception:
        logger.exception("Error saving master data")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>/material-properties", methods=["POST"])
@login_required
@module_access_required("csc")
def save_material_properties(draft_id: int):
    """Save workbook-driven material properties fields only."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        if not _can_edit_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403
        lock_response = _json_editor_lock_conflict_response(draft, "committee_user")
        if lock_response is not None:
            return lock_response
        if not _scope_allows_material_properties(_load_workflow_scope(draft)):
            return jsonify({"error": "This draft is not open for revision in Material Properties."}), 403

        if not (draft.material_code or "").strip():
            return jsonify({"error": "Material code is required before material properties can be saved."}), 400

        data = request.get_json() or {}
        record = update_material_properties_from_form(draft, data, user_id=current_user.id)
        if record is not None and record not in db.session:
            db.session.add(record)

        draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        _create_audit_entry(draft_id, "Material properties saved from drafting workspace")

        return jsonify({
            "success": True,
            "draft": draft.to_dict(),
            "master_data": get_master_form_values(draft),
            "material_properties": get_material_properties_values(draft),
            "storage_handling": get_storage_handling_values(draft),
        })
    except Exception:
        logger.exception("Error saving material properties")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>/storage-handling", methods=["POST"])
@login_required
@module_access_required("csc")
def save_storage_handling(draft_id: int):
    """Save storage and handling fields only."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        if not _can_edit_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403
        lock_response = _json_editor_lock_conflict_response(draft, "committee_user")
        if lock_response is not None:
            return lock_response
        if not _scope_allows_storage_handling(_load_workflow_scope(draft)):
            return jsonify({"error": "This draft is not open for revision in Storage and Handling."}), 403

        if not (draft.material_code or "").strip():
            return jsonify({"error": "Material code is required before storage and handling can be saved."}), 400

        data = request.get_json() or {}
        record = update_storage_handling_from_form(draft, data, user_id=current_user.id)
        if record is not None and record not in db.session:
            db.session.add(record)

        draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        _create_audit_entry(draft_id, "Storage and handling saved from drafting workspace")

        return jsonify({
            "success": True,
            "draft": draft.to_dict(),
            "master_data": get_master_form_values(draft),
            "material_properties": get_material_properties_values(draft),
            "storage_handling": get_storage_handling_values(draft),
        })
    except Exception:
        logger.exception("Error saving storage and handling")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>/section/<section_name>", methods=["POST"])
@login_required
@module_access_required("csc")
def save_section(draft_id: int, section_name: str):
    """Save section text content via AJAX."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        if not _can_edit_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403
        lock_response = _json_editor_lock_conflict_response(draft, "committee_user")
        if lock_response is not None:
            return lock_response

        if draft.phase1_locked and section_name in ["properties", "specifications", "test_methods"]:
            return jsonify({"error": "Phase 1 is locked"}), 403

        data = request.get_json() or {}
        section_text = data.get("section_text", "")

        _set_draft_section_text(draft, section_name, section_text)

        draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        section = None
        for candidate_name in _section_name_aliases(section_name):
            section = CSCSection.query.filter_by(
                draft_id=draft_id, section_name=candidate_name
            ).first()
            if section is not None:
                break

        _create_audit_entry(draft_id, f"Section saved: {section_name}")

        return jsonify({"success": True, "section": section.to_dict()})
    except Exception:
        logger.exception("Error saving section")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>/issues", methods=["POST"])
@login_required
@module_access_required("csc")
def save_issues(draft_id: int):
    """Save issue flags via AJAX."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        if not _can_edit_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403
        lock_response = _json_editor_lock_conflict_response(draft, "committee_user")
        if lock_response is not None:
            return lock_response

        data = request.get_json() or []

        CSCIssueFlag.query.filter_by(draft_id=draft_id).delete()

        for idx, issue_data in enumerate(data):
            flag = CSCIssueFlag(
                draft_id=draft_id,
                issue_type=issue_data.get("issue_type", ""),
                is_present=issue_data.get("is_present", False),
                note=issue_data.get("note", ""),
                sort_order=idx,
            )
            db.session.add(flag)

        draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        _create_audit_entry(draft_id, "Issue flags saved")

        return jsonify({"success": True})
    except Exception:
        logger.exception("Error saving issues for draft_id=%s", draft_id)
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>/parameters", methods=["POST"])
@login_required
@module_access_required("csc")
def save_parameters(draft_id: int):
    """Save parameters (Phase 1 - full edit)."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        if not _can_edit_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403
        lock_response = _json_editor_lock_conflict_response(draft, "committee_user")
        if lock_response is not None:
            return lock_response

        if draft.phase1_locked:
            return jsonify({"error": "Phase 1 is locked"}), 403

        workflow_scope = _load_workflow_scope(draft)
        parameter_type_allowed = bool(workflow_scope.get("parameter_type"))
        parameter_other_allowed = bool(workflow_scope.get("parameter_other"))
        if not _scope_allows_parameter_tab(workflow_scope):
            return jsonify({"error": "This draft is not open for revision in Parameters."}), 403

        data = request.get_json() or []
        existing_parameters = draft.parameters.order_by(CSCParameter.sort_order).all()
        existing_by_id = {str(parameter.id): parameter for parameter in existing_parameters}

        if parameter_type_allowed and not parameter_other_allowed:
            if len(data) != len(existing_parameters):
                return jsonify({"error": "Parameter structure is locked. Only Type can be revised in this workflow."}), 400
            for idx, param_data in enumerate(data):
                param_id = str(param_data.get("parameter_id") or "")
                parameter = existing_by_id.get(param_id)
                if parameter is None and idx < len(existing_parameters):
                    parameter = existing_parameters[idx]
                if parameter is None:
                    return jsonify({"error": "Parameter structure is locked. Refresh and try again."}), 400
                parameter.parameter_type = normalize_parameter_type_label(
                    param_data.get("parameter_type", parameter.parameter_type or "Desirable")
                )
            draft.updated_at = datetime.now(timezone.utc)
            db.session.commit()

            _create_audit_entry(draft_id, f"Parameter types saved: {len(data)} items")

            return jsonify({"success": True})

        CSCParameter.query.filter_by(draft_id=draft_id).delete()

        for idx, param_data in enumerate(data):
            param_id = str(param_data.get("parameter_id") or "")
            existing_parameter = existing_by_id.get(param_id)
            parameter_type = normalize_parameter_type_label(
                param_data.get(
                    "parameter_type",
                    (existing_parameter.parameter_type if existing_parameter else "Desirable"),
                )
            ) if parameter_type_allowed else normalize_parameter_type_label(
                existing_parameter.parameter_type if existing_parameter else "Desirable"
            )

            param = CSCParameter(
                draft_id=draft_id,
                parameter_name=param_data.get("parameter_name", ""),
                parameter_type=parameter_type,
                unit_of_measure=(param_data.get("unit_of_measure", "") or "").strip() or None,
                existing_value=format_required_value(
                    param_data.get("required_value_type", "text"),
                    param_data.get("required_value_text", param_data.get("existing_value", "")),
                    param_data.get("required_value_operator_1", ""),
                    param_data.get("required_value_value_1", ""),
                    param_data.get("required_value_operator_2", ""),
                    param_data.get("required_value_value_2", ""),
                ),
                proposed_value=param_data.get("proposed_value", ""),
                test_method=normalize_test_procedure_type(
                    param_data.get("test_procedure_type", param_data.get("test_method", ""))
                ),
                test_procedure_type=normalize_test_procedure_type(
                    param_data.get("test_procedure_type", param_data.get("test_method", ""))
                ),
                test_procedure_text=param_data.get("test_procedure_text", ""),
                parameter_conditions=param_data.get("parameter_conditions", ""),
                required_value_type=(
                    "comparison"
                    if str(param_data.get("required_value_type", "text")).strip().lower() == "comparison"
                    else "text"
                ),
                required_value_text=param_data.get("required_value_text", ""),
                required_value_operator_1=param_data.get("required_value_operator_1", ""),
                required_value_value_1=param_data.get("required_value_value_1", ""),
                required_value_operator_2=param_data.get("required_value_operator_2", ""),
                required_value_value_2=param_data.get("required_value_value_2", ""),
                justification=param_data.get("justification", ""),
                remarks=param_data.get("remarks", ""),
                sort_order=idx,
            )
            db.session.add(param)

        draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        _create_audit_entry(draft_id, f"Parameters saved (Phase 1): {len(data)} items")

        return jsonify(
            {
                "success": True,
                "title": "Draft Returned to Committee User",
                "message": (
                    f"{revision.parent_draft.spec_number} — {revision.parent_draft.chemical_name} "
                    "was returned to the Committee User for updates."
                ),
            }
        )
    except Exception:
        logger.exception("Error saving parameters")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>/parameters/proposed", methods=["POST"])
@login_required
@module_access_required("csc")
def save_proposed_values(draft_id: int):
    """Save proposed values only (Phase 2)."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        if not _can_edit_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403
        lock_response = _json_editor_lock_conflict_response(draft, "committee_user")
        if lock_response is not None:
            return lock_response
        workflow_scope = _load_workflow_scope(draft)
        if not bool(workflow_scope.get("parameter_other")):
            return jsonify({"error": "This draft is not open for revision in parameter values."}), 403

        data = request.get_json() or []

        for item in data:
            param_id = item.get("parameter_id")
            proposed_value = item.get("proposed_value", "")

            param = CSCParameter.query.filter_by(
                id=param_id, draft_id=draft_id
            ).first()

            if param:
                param.proposed_value = proposed_value

        draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        _create_audit_entry(draft_id, f"Proposed values saved (Phase 2): {len(data)} items")

        return jsonify({"success": True})
    except Exception:
        logger.exception("Error saving proposed values for draft_id=%s", draft_id)
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>/impact", methods=["POST"])
@login_required
@module_access_required("csc")
def save_impact(draft_id: int):
    """Save impact analysis scores."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        if not _can_edit_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403
        lock_response = _json_editor_lock_conflict_response(draft, "committee_user")
        if lock_response is not None:
            return lock_response
        if not _scope_allows_impact(_load_workflow_scope(draft)):
            return jsonify({"error": "This draft is not open for revision in Impact."}), 403

        data = request.get_json() or {}

        checklist_state = deserialize_impact_checklist_state(
            data.get("checklist_state"),
            data.get("chemical_name") or draft.chemical_name,
        )
        legacy_payload = build_impact_legacy_payload(checklist_state)

        impact = draft.impact_analysis
        if not impact:
            impact = CSCImpactAnalysis(draft_id=draft_id)
            db.session.add(impact)

        impact.operational_impact_score = legacy_payload["operational_impact_score"]
        impact.safety_environment_score = legacy_payload["safety_environment_score"]
        impact.supply_risk_score = legacy_payload["supply_risk_score"]
        impact.no_substitute_flag = legacy_payload["no_substitute_flag"]
        impact.impact_score_total = legacy_payload["impact_score_total"]
        impact.impact_grade = legacy_payload["impact_grade"]
        impact.checklist_state_json = legacy_payload["checklist_state_json"]
        impact.operational_note = legacy_payload["operational_note"]
        impact.safety_environment_note = legacy_payload["safety_environment_note"]
        impact.supply_risk_note = legacy_payload["supply_risk_note"]
        record = sync_impact_fields_to_master_record(
            draft,
            checklist_state,
            user_id=current_user.id,
        )
        if record is not None and record not in db.session:
            db.session.add(record)

        draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        _create_audit_entry(
            draft_id,
            "Impact analysis saved",
            new_value=json.dumps({
                "classification": legacy_payload["checklist_summary"]["classification"],
                "rule": legacy_payload["checklist_summary"]["rule"],
            }),
        )

        return jsonify({
            "success": True,
            "impact": impact.to_dict(),
            "checklist_summary": legacy_payload["checklist_summary"],
            "master_data": get_master_form_values(draft),
        })
    except Exception:
        logger.exception("Error saving impact")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>/flag", methods=["POST"])
@login_required
@module_access_required("csc")
def update_flag(draft_id: int):
    """Update a single impact flag value via AJAX (4-state: YES/PROVISIONAL/REVIEW/NO)."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404
        if not _can_edit_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403
        lock_response = _json_editor_lock_conflict_response(draft, "committee_user")
        if lock_response is not None:
            return lock_response
        if not _scope_allows_impact(_load_workflow_scope(draft)):
            return jsonify({"error": "This draft is not open for revision in Impact."}), 403

        data = request.get_json() or {}
        flag_name = data.get("flag_name", "")
        new_value = data.get("new_value", "")

        if flag_name not in ALL_FLAG_IDS:
            return jsonify({"error": f"Invalid flag name: {flag_name}"}), 400
        if new_value not in VALID_FLAG_VALUES:
            return jsonify({"error": f"Invalid flag value: {new_value}"}), 400

        display_name = (
            (current_user.full_name or current_user.username or "").strip()
            or current_user.username
        )
        answered_on = datetime.now(timezone.utc).isoformat()

        # Load and migrate existing state
        impact = draft.impact_analysis
        if impact is None:
            impact = CSCImpactAnalysis(draft_id=draft_id)
            db.session.add(impact)

        checklist_state = deserialize_impact_checklist_state(
            impact.checklist_state_json,
            draft.chemical_name,
        )

        # Update the specific flag with full metadata
        fin_reason = data.get("fin_reason", "")
        checklist_state["flags"][flag_name] = {
            "value": new_value,
            "source_type": "COMMITTEE",
            "answered_by": display_name,
            "answered_on": answered_on,
            "fin_reason": fin_reason,
        }

        # Recompute classification
        flat_flags = {
            fid: entry.get("value", "REVIEW")
            for fid, entry in checklist_state["flags"].items()
        }
        classification = compute_impact_classification(flat_flags)

        # Build legacy payload and persist
        legacy_payload = build_impact_legacy_payload(checklist_state)
        impact.operational_impact_score = legacy_payload["operational_impact_score"]
        impact.safety_environment_score = legacy_payload["safety_environment_score"]
        impact.supply_risk_score = legacy_payload["supply_risk_score"]
        impact.no_substitute_flag = legacy_payload["no_substitute_flag"]
        impact.impact_score_total = legacy_payload["impact_score_total"]
        impact.impact_grade = legacy_payload["impact_grade"]
        impact.checklist_state_json = json.dumps(checklist_state)
        impact.operational_note = legacy_payload["operational_note"]
        impact.safety_environment_note = legacy_payload["safety_environment_note"]
        impact.supply_risk_note = legacy_payload["supply_risk_note"]

        draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        return jsonify({
            "success": True,
            "classification": classification,
            "flag_meta": checklist_state["flags"][flag_name],
        })
    except Exception:
        logger.exception("Error updating flag")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>/submit", methods=["POST"])
@login_required
@module_access_required("csc")
def submit_revision(draft_id: int):
    """Submit an open drafting copy for admin approval."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        if not _can_edit_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403
        lock_response = _json_editor_lock_conflict_response(draft, "committee_user")
        if lock_response is not None:
            return lock_response

        revision = _get_revision_for_child(draft_id)
        if draft.parent_draft_id is None or revision is None:
            return jsonify({"error": "Only drafting copies can be submitted"}), 400
        if not _can_submit_revision(revision):
            return jsonify({"error": "This drafting copy is not open for submission"}), 400

        data = request.get_json(silent=True) or {}
        authorization_confirmed = bool(data.get("authorization_confirmed", False))
        authorized_by_name = (data.get("authorized_by_name") or "").strip()
        subcommittee_head_name = (data.get("subcommittee_head_name") or "").strip()

        if not authorization_confirmed:
            return jsonify({
                "error": "Authorization confirmation is required before submission."
            }), 400

        if not authorized_by_name or not subcommittee_head_name:
            return jsonify({
                "error": "Enter your name and the Sub Committee Head name before submission."
            }), 400

        draft.status = "Drafting"
        draft.created_by_id = current_user.id
        draft.created_by_role = current_user.role.name if current_user.role else draft.created_by_role
        draft.updated_at = datetime.now(timezone.utc)
        revision.submitted_at = datetime.now(timezone.utc)
        revision.status = WORKFLOW_DRAFTING_SUBMITTED
        revision.authorization_confirmed = authorization_confirmed
        revision.authorized_by_name = authorized_by_name
        revision.subcommittee_head_name = subcommittee_head_name
        db.session.add(
            CSCAudit(
                draft_id=draft_id,
                action="Revision submitted for approval",
                user_name=current_user.username,
                action_time=datetime.now(timezone.utc),
                new_value=json.dumps(
                    {
                        "authorization_confirmed": authorization_confirmed,
                        "authorized_by_name": authorized_by_name,
                        "subcommittee_head_name": subcommittee_head_name,
                        "subset": draft.subset,
                    }
                ),
            )
        )
        head_user_ids = []
        governance_user_ids = []
        committee_payload = get_committee_config_payload()
        committee_slug = _get_revision_committee_slug(revision)
        head_user_ids.extend(
            _get_committee_user_ids_for_committee_slug(committee_slug, "committee_head")
        )
        for record in committee_payload.get("CHILD_COMMITTEES", []):
            if record.get("slug") != "coordination":
                continue
            governance_user_ids.extend(
                _get_committee_user_ids_for_committee_slug("coordination", "committee_user")
            )
            governance_user_ids.extend(
                _get_committee_user_ids_for_committee_slug("coordination", "committee_head")
            )
            break

        _notify_users(
            list(dict.fromkeys(head_user_ids + governance_user_ids)),
            title="Draft submitted for committee-head review",
            message=f"{draft.spec_number} — {draft.chemical_name} was submitted for committee-head review.",
            severity="info",
            link=url_for("csc.review_workbench"),
        )

        db.session.commit()
        return jsonify(
            {
                "success": True,
                "redirect": url_for("csc.workspace"),
                "title": "Draft Submitted for Approval",
                "message": (
                    f"{draft.spec_number} — {draft.chemical_name} has been submitted "
                    "to the Committee Head for review."
                ),
            }
        )
    except Exception:
        logger.exception("Error submitting revision")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/api/draft/<int:draft_id>/delete", methods=["POST", "DELETE"])
@login_required
@module_access_required("csc")
def delete_draft(draft_id: int):
    """Delete a draft (only own drafts or admin)."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        if not _can_delete_draft(draft):
            return jsonify({"error": "Unauthorized"}), 403

        _create_audit_entry(draft_id, "Draft deleted")
        spec_label = draft.spec_number or "Draft"
        chemical_label = draft.chemical_name or "Untitled specification"

        db.session.delete(draft)
        db.session.commit()

        return jsonify(
            {
                "success": True,
                "title": "Draft Deleted",
                "message": f"{spec_label} — {chemical_label} was deleted successfully.",
            }
        )
    except Exception:
        logger.exception("Error deleting draft")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/draft/<int:draft_id>/export/docx")
@login_required
@module_access_required("csc")
def export_docx(draft_id: int):
    """Generate and download Word document."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            abort(404)

        if not _can_edit_draft(draft):
            flash("You do not have permission to export this draft.", "danger")
            return redirect(url_for("csc.workspace"))

        review_context = _build_draft_export_review_context(draft)
        doc_bytes = build_flask_review_document(review_context)
        safe_spec_number = "".join(
            char if char.isalnum() or char in "-_" else "_"
            for char in (draft.spec_number or f"draft_{draft.id}")
        )
        filename = f"CSC_REVIEW_DOSSIER_{safe_spec_number}_{datetime.now().strftime('%Y%m%d_%H%M')}.docx"
        _record_export_audit("DOCX_EXPORT", draft.id, filename)
        return send_file(
            io.BytesIO(doc_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            as_attachment=True,
            download_name=filename,
        )

    except Exception:
        logger.exception("Error exporting draft")
        flash("Error exporting draft.", "danger")
        return redirect(url_for("csc.workspace"))


@csc_bp.route("/draft/<int:draft_id>/dossier-preview")
@login_required
@module_access_required("csc")
def draft_dossier_preview(draft_id: int):
    """Return an in-app dossier preview for the current draft."""
    draft = db.session.get(CSCDraft, draft_id)
    if not draft:
        abort(404)

    if not _can_edit_draft(draft):
        return jsonify({"success": False, "error": "You do not have permission to preview this draft."}), 403

    review_context = _build_draft_export_review_context(draft)
    html = render_template(
        "csc/_draft_dossier_preview.html",
        draft=draft,
        review_context=review_context,
        current_version=(draft.version if draft.version is not None else 1),
        last_updated=draft.updated_at,
    )
    return jsonify(
        {
            "success": True,
            "title": draft.spec_number or "Draft Dossier",
            "subtitle": draft.chemical_name or "Draft dossier preview",
            "html": html,
        }
    )


@csc_bp.route("/workspace/new-revision/<int:parent_id>", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def create_revision(parent_id: int):
    """Open a drafting copy from a published specification."""
    try:
        parent_draft = db.session.get(CSCDraft, parent_id)
        if not parent_draft:
            return jsonify({"error": "Parent draft not found"}), 404

        request_payload = request.get_json(silent=True) or {}
        workflow_scope = _normalize_workflow_scope_payload(
            request_payload.get("workflow_scope")
        )
        workflow_track = str(request_payload.get("workflow_track") or "").strip().lower()
        scope_error = _validate_workflow_scope(workflow_scope)
        if scope_error:
            return jsonify({"error": scope_error}), 400
        if workflow_track not in {WORKFLOW_TRACK_SUBSET, WORKFLOW_TRACK_MATERIAL_HANDLING}:
            return jsonify({"error": "Choose Subset Committee or Material Handling Committee."}), 400

        child_draft, created_new, error = _open_committee_workflow_draft_for_parent(
            parent_draft,
            workflow_scope,
            workflow_track,
        )
        if error:
            return jsonify({"error": error}), 409
        if child_draft is None:
            return jsonify({"error": "Unable to open workflow draft"}), 500

        committee_slug = _infer_workflow_committee_slug(child_draft)
        revision = _get_revision_for_child(child_draft.id)
        committee_user_ids = _get_committee_user_ids_for_committee_slug(
            committee_slug,
            "committee_user",
        )
        committee_user_label = _get_committee_user_display_label_for_committee_slug(
            committee_slug,
            "committee_user",
        )
        workflow_meta_lines = _build_workflow_meta_lines(parent_draft, child_draft, revision)

        db.session.commit()

        _create_audit_entry(
            parent_id,
            "Workflow draft staged",
            new_value=json.dumps(
                {
                    "child_draft_id": child_draft.id,
                    "workflow_scope": workflow_scope,
                    "workflow_track": workflow_track,
                    "committee_slug": _infer_workflow_committee_slug(child_draft),
                }
            ),
        )
        return jsonify({
            "success": True,
            "draft_id": child_draft.id,
            "redirect": url_for("csc.editor", draft_id=child_draft.id),
            "created_new": created_new,
            "workflow_meta_lines": workflow_meta_lines,
            "title": "Draft Staged for Secretary" if created_new else "Workflow Already Exists",
            "message": (
                "Draft created in Secretary staging. Review, structure, and open it for committee user(s) when ready - "
                + (committee_user_label or "Committee Users not assigned")
                if created_new
                else f"A workflow already exists for {parent_draft.spec_number}."
            ),
        })
    except Exception:
        logger.exception("Error creating revision")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/admin/draft/<int:parent_id>/open-self-draft", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_open_self_draft(parent_id: int):
    """Create or reopen an admin-owned workflow draft for the published spec."""
    wants_json = request.is_json or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    try:
        parent_draft = db.session.get(CSCDraft, parent_id)
        if not parent_draft or parent_draft.parent_draft_id is not None:
            if wants_json:
                return jsonify({"error": "Published specification not found"}), 404
            flash("Published specification not found.", "danger")
            return redirect(url_for("csc.admin_revision_generation"))
        request_payload = request.get_json(silent=True) or {}
        workflow_scope = _normalize_workflow_scope_payload(
            request_payload.get("workflow_scope") if request_payload else WORKFLOW_SCOPE_DEFAULT
        )
        scope_error = _validate_workflow_scope(workflow_scope)
        if scope_error:
            if wants_json:
                return jsonify({"error": scope_error}), 400
            flash(scope_error, "danger")
            return redirect(url_for("csc.admin_revision_generation"))

        child_draft, _, error = _open_admin_self_draft_for_parent(parent_draft, workflow_scope)
        if error:
            if wants_json:
                return jsonify({"error": error}), 409
            flash(error, "warning")
            return redirect(url_for("csc.admin_revision_generation"))
        if child_draft is None:
            if wants_json:
                return jsonify({"error": "Unable to open workflow draft"}), 500
            flash("Unable to open workflow draft.", "danger")
            return redirect(url_for("csc.admin_revision_generation"))
        db.session.commit()

        _create_audit_entry(
            parent_id,
            "Admin self draft opened",
            new_value=json.dumps(
                {
                    "child_draft_id": child_draft.id,
                    "workflow_scope": workflow_scope,
                }
            ),
        )

        if wants_json:
            return jsonify(
                {
                    "success": True,
                    "draft_id": child_draft.id,
                    "redirect": url_for("csc.editor", draft_id=child_draft.id),
                }
            )

        flash("Admin self draft opened.", "success")
        return redirect(url_for("csc.editor", draft_id=child_draft.id))
    except Exception:
        logger.exception("Error opening admin self draft")
        db.session.rollback()
        if wants_json:
            return jsonify({"error": "An internal error occurred. Please try again."}), 500
        flash("Error opening admin self draft.", "danger")
        return redirect(url_for("csc.admin_revision_generation"))

@csc_bp.route("/admin/draft-generation/open-self-drafts", methods=["POST"])
@csc_bp.route("/admin/drafts/open-self-drafts", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_open_self_drafts_bulk():
    """Create admin workflow drafts for a selected set of published specs."""
    try:
        data = request.get_json(silent=True) or {}
        raw_ids = data.get("parent_ids") or []
        parent_ids: list[int] = []
        for value in raw_ids:
            try:
                parent_ids.append(int(value))
            except (TypeError, ValueError):
                continue
        parent_ids = list(dict.fromkeys(parent_ids))
        if not parent_ids:
            return jsonify({"error": "Select at least one specification to open for revision."}), 400

        workflow_scope = _normalize_workflow_scope_payload(data.get("workflow_scope"))
        workflow_track = str(data.get("workflow_track") or "").strip().lower()
        scope_error = _validate_workflow_scope(workflow_scope)
        if scope_error:
            return jsonify({"error": scope_error}), 400
        if workflow_track not in {WORKFLOW_TRACK_SUBSET, WORKFLOW_TRACK_MATERIAL_HANDLING}:
            return jsonify({"error": "Choose Subset Committee or Material Handling Committee."}), 400

        created_rows: list[dict[str, object]] = []
        conflicts: list[str] = []
        notified_labels: list[str] = []
        for parent_id in parent_ids:
            parent_draft = db.session.get(CSCDraft, parent_id)
            if not parent_draft or parent_draft.parent_draft_id is not None:
                conflicts.append(f"Specification {parent_id} is not available.")
                continue
            child_draft, created_new, error = _open_committee_workflow_draft_for_parent(
                parent_draft,
                workflow_scope,
                workflow_track,
            )
            if error:
                conflicts.append(f"{parent_draft.spec_number}: {error}")
                continue
            if child_draft is None:
                conflicts.append(f"{parent_draft.spec_number}: unable to open workflow draft.")
                continue
            committee_slug = _infer_workflow_committee_slug(child_draft)
            revision = _get_revision_for_child(child_draft.id)
            committee_user_ids = _get_committee_user_ids_for_committee_slug(
                committee_slug,
                "committee_user",
            )
            committee_user_label = _get_committee_user_display_label_for_committee_slug(
                committee_slug,
                "committee_user",
            )
            if committee_user_label:
                notified_labels.append(committee_user_label)
            created_rows.append(
                {
                    "parent_id": parent_id,
                    "child_draft_id": child_draft.id,
                    "created_new": created_new,
                    "spec_number": parent_draft.spec_number,
                    "chemical_name": parent_draft.chemical_name,
                    "workflow_meta_lines": _build_workflow_meta_lines(parent_draft, child_draft, revision),
                }
            )

        if not created_rows:
            db.session.rollback()
            return jsonify({"error": conflicts[0] if conflicts else "No workflow drafts were opened."}), 409

        db.session.commit()

        for row in created_rows:
            _create_audit_entry(
                int(row["parent_id"]),
                "Workflow draft staged",
                new_value=json.dumps(
                {
                    "child_draft_id": row["child_draft_id"],
                    "workflow_scope": workflow_scope,
                    "workflow_track": workflow_track,
                }
            ),
        )

        redirect = (
            url_for("csc.editor", draft_id=created_rows[0]["child_draft_id"])
            if len(created_rows) == 1
            else url_for("csc.admin_revision_generation")
        )
        created_count = sum(1 for row in created_rows if row.get("created_new"))
        existing_count = len(created_rows) - created_count
        if len(created_rows) == 1:
            title = "Draft Staged for Secretary" if created_count else "Workflow Already Exists"
            if created_count:
                message = (
                    "Draft created in Secretary staging for user(s) - "
                    + (notified_labels[0] if notified_labels else "Committee User not assigned")
                )
            else:
                message = f"A workflow already exists for {created_rows[0]['spec_number']}."
        elif existing_count == 0:
            title = "Workflow Drafts Staged"
            message = (
                f"Opened {len(created_rows)} workflow draft(s) into Secretary staging for "
                + (", ".join(sorted(set(notified_labels))) if notified_labels else "the assigned committee users")
                + "."
            )
        elif created_count == 0:
            title = "Workflow Drafts Already Exist"
            message = f"All {len(created_rows)} selected specification(s) already have active workflow drafts."
        else:
            title = "Workflow Drafts Staged"
            message = (
                f"Created {created_count} new staged workflow draft(s). "
                f"{existing_count} specification(s) already had active workflow drafts."
            )
        return jsonify(
            {
                "success": True,
                "redirect": redirect,
                "created_count": len(created_rows),
                "new_created_count": created_count,
                "existing_count": existing_count,
                "workflow_details": created_rows,
                "title": title,
                "message": message,
                "conflicts": conflicts,
            }
        )
    except Exception:
        logger.exception("Error opening admin workflow drafts")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


# ──────────────────────────────────────────────────────────────────────────────
# Admin Routes
# ──────────────────────────────────────────────────────────────────────────────


@csc_bp.route("/admin")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_index():
    """Admin dashboard with KPI stats."""
    stats = _get_admin_stats()
    return render_template("csc/admin/index.html", stats=stats)


@csc_bp.route("/admin/ingest")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_ingest():
    """PDF/DOCX ingest page (GET only – shows upload forms)."""
    return render_template("csc/admin/ingest.html")


@csc_bp.route("/admin/ingest/pdf", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_ingest_pdf():
    """Process uploaded PDF file."""
    try:
        if "pdf_file" not in request.files:
            flash("No PDF file provided.", "danger")
            return redirect(url_for("csc.admin_ingest"))

        file = request.files["pdf_file"]
        if file.filename == "":
            flash("No file selected.", "danger")
            return redirect(url_for("csc.admin_ingest"))

        # TODO: Implement PDF parsing with csc_pdf_extractor
        draft = CSCDraft(
            spec_number=request.form.get("spec_number", "TBD"),
            chemical_name=request.form.get("chemical_name", file.filename or "TBD"),
            committee_name="Corporate Specifiction Committee",
            prepared_by=current_user.username,
            status="Published",
            created_by_role="Admin",
            is_admin_draft=True,
            admin_stage=ADMIN_STAGE_PUBLISHED,
            created_by_id=current_user.id,
        )
        db.session.add(draft)
        db.session.commit()

        flash(f"PDF ingested – Draft created (ID: {draft.id})", "success")
        return redirect(url_for("csc.admin_revision_generation"))
    except Exception:
        logger.exception("Error ingesting PDF")
        flash("Error processing PDF file.", "danger")
        return redirect(url_for("csc.admin_ingest"))


@csc_bp.route("/admin/ingest/docx", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_ingest_docx():
    """Process uploaded DOCX master document."""
    try:
        if "docx_file" not in request.files:
            flash("No DOCX file provided.", "danger")
            return redirect(url_for("csc.admin_ingest"))

        file = request.files["docx_file"]
        if file.filename == "":
            flash("No file selected.", "danger")
            return redirect(url_for("csc.admin_ingest"))

        # TODO: Implement DOCX parsing with csc_docx_extractor
        draft = CSCDraft(
            spec_number=request.form.get("spec_number", "TBD"),
            chemical_name=request.form.get("chemical_name", file.filename or "TBD"),
            committee_name="Corporate Specifiction Committee",
            prepared_by=current_user.username,
            status="Published",
            created_by_role="Admin",
            is_admin_draft=True,
            admin_stage=ADMIN_STAGE_PUBLISHED,
            created_by_id=current_user.id,
        )
        db.session.add(draft)
        db.session.commit()

        flash(f"DOCX ingested – Draft created (ID: {draft.id})", "success")
        return redirect(url_for("csc.admin_revision_generation"))
    except Exception:
        logger.exception("Error ingesting DOCX")
        flash("Error processing DOCX file.", "danger")
        return redirect(url_for("csc.admin_ingest"))


@csc_bp.route("/admin/ingest/create-draft", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_create_draft_from_extraction():
    """Create draft from extraction preview."""
    try:
        extraction_id = request.form.get("extraction_id")
        spec_index = request.form.get("spec_index", 0, type=int)

        # TODO: look up extraction result from session/cache and create proper draft
        draft = CSCDraft(
            spec_number=f"EXT-{extraction_id}-{spec_index}",
            chemical_name="Extracted Specification",
            committee_name="Corporate Specifiction Committee",
            prepared_by=current_user.username,
            status="Published",
            created_by_role="Admin",
            is_admin_draft=True,
            admin_stage=ADMIN_STAGE_PUBLISHED,
            created_by_id=current_user.id,
        )
        db.session.add(draft)
        db.session.commit()

        flash(f"Draft created from extraction (ID: {draft.id})", "success")
        return redirect(url_for("csc.admin_revision_generation"))
    except Exception:
        logger.exception("Error creating draft from extraction")
        flash("Error creating draft from extraction.", "danger")
        return redirect(url_for("csc.admin_ingest"))


@csc_bp.route("/published-specs")
@login_required
@module_access_required("csc")
def published_specs():
    """List current published specifications for all CSC module users."""
    try:
        drafts = sorted(
            CSCDraft.query.filter_by(parent_draft_id=None).all(),
            key=_draft_sort_key,
        )
        subsets = _get_spec_subsets()
        published_rows = []
        for draft in drafts:
            active_revision = _get_active_revision_for_parent(draft.id)
            current_snapshot = (
                draft.spec_versions.filter_by(spec_version=draft.spec_version)
                .order_by(CSCSpecVersion.created_at.desc())
                .first()
            )
            snapshot_count = draft.spec_versions.count()
            if current_snapshot is None:
                snapshot_count += 1

            published_rows.append(
                {
                    "draft": draft,
                    "active_revision": active_revision,
                    "current_version": format_spec_version(draft.spec_version),
                    "version_count": snapshot_count,
                    "latest_published_at": (
                        current_snapshot.created_at
                        if current_snapshot is not None
                        else draft.updated_at
                    ),
                }
            )
        response = make_response(
            render_template(
                "csc/published_specs.html",
                published_rows=published_rows,
                subsets=subsets,
            )
        )
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response
    except Exception:
        logger.exception("Error loading published specs")
        flash("Error loading published specifications.", "danger")
        return redirect(url_for("csc.index"))


@csc_bp.route("/published-specs/<int:draft_id>")
@login_required
@module_access_required("csc")
def published_spec_detail(draft_id: int):
    """Render a read-only published specification snapshot for all CSC users."""
    draft = db.session.get(CSCDraft, draft_id)
    if not draft or draft.parent_draft_id is not None:
        abort(404)

    review_context = _build_draft_export_review_context(draft)
    return render_template(
        "csc/published_spec.html",
        draft=draft,
        review_context=review_context,
        active_revision=_get_active_revision_for_parent(draft.id),
        version_map=_build_published_version_map(draft),
    )


@csc_bp.route("/published-specs/<int:draft_id>/preview")
@login_required
@module_access_required("csc")
def published_spec_preview(draft_id: int):
    """Return an in-app CSC Format A preview for a published specification."""
    draft = db.session.get(CSCDraft, draft_id)
    if not draft or draft.parent_draft_id is not None:
        abort(404)

    review_context = _build_draft_export_review_context(draft)
    current_snapshot = (
        draft.spec_versions.filter_by(spec_version=draft.spec_version)
        .order_by(CSCSpecVersion.created_at.desc())
        .first()
    )
    latest_published_at = (
        current_snapshot.created_at
        if current_snapshot is not None
        else draft.updated_at
    )

    html = render_template(
        "csc/_format_a_preview.html",
        draft=draft,
        review_context=review_context,
        active_revision=_get_active_revision_for_parent(draft.id),
        latest_published_at=latest_published_at,
        current_version=format_spec_version(draft.spec_version),
    )
    return jsonify(
        {
            "success": True,
            "title": draft.spec_number,
            "subtitle": draft.chemical_name or "Published specification preview",
            "html": html,
        }
    )


@csc_bp.route("/published-specs/<int:draft_id>/versions/<int:spec_version>/dossier.docx")
@login_required
@module_access_required("csc")
def published_spec_version_dossier(draft_id: int, spec_version: int):
    """Download the stored published dossier for a version-map node."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft or draft.parent_draft_id is not None:
            abort(404)

        snapshot = (
            draft.spec_versions.filter_by(spec_version=spec_version)
            .order_by(CSCSpecVersion.created_at.desc())
            .first()
        )
        payload = _load_spec_version_snapshot_payload(snapshot)
        review_context = _extract_review_context_from_snapshot_payload(payload)

        if review_context is None:
            flash(
                "The exact published dossier is unavailable for this version because a full review snapshot was not stored at publish time.",
                "warning",
            )
            return redirect(url_for("csc.published_spec_detail", draft_id=draft.id))

        doc_bytes = build_flask_review_document(review_context)
        filename = (
            f"{draft.spec_number}_v{format_spec_version(spec_version)}_published_dossier.docx"
        )
        _record_export_audit("PUBLISHED_DOSSIER_EXPORT", draft.id, filename)
        return send_file(
            io.BytesIO(doc_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            as_attachment=True,
            download_name=filename,
        )
    except Exception:
        logger.exception("Error exporting published dossier")
        flash("Error exporting the published dossier.", "danger")
        return redirect(url_for("csc.published_specs"))


@csc_bp.route("/admin/draft-generation")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_revision_generation():
    """Admin-only surface to generate workflow drafts and self drafts."""
    try:
        drafts = sorted(
            CSCDraft.query.filter_by(parent_draft_id=None).all(),
            key=_draft_sort_key,
        )
        subsets = _get_spec_subsets()
        workflow_rows = []
        for draft in drafts:
            display_status = draft.status if draft.status in {"Published", "Drafting"} else "Published"
            active_revisions = _get_active_revisions_for_parent(draft.id)
            if active_revisions:
                for revision in active_revisions:
                    workflow_draft = revision.child_draft if revision.child_draft else draft
                    master_values = get_master_form_values(workflow_draft)
                    committee_slug = _get_revision_committee_slug(revision)
                    workflow_rows.append(
                        {
                            "draft": draft,
                            "workflow_draft": workflow_draft,
                            "active_revision": revision,
                            "display_status": display_status,
                            "physical_state": master_values.get("physical_state", ""),
                            "drafting_status": revision.status,
                            "submitted_at": revision.submitted_at,
                            "committee_slug": committee_slug,
                            "committee_title": _get_committee_title_by_slug(committee_slug),
                            "committee_track": _committee_slug_to_track(committee_slug),
                            "active_revision_count": len(active_revisions),
                            "can_generate_more": len(active_revisions) < 2,
                        }
                    )
                continue

            master_values = get_master_form_values(draft)
            workflow_rows.append(
                {
                    "draft": draft,
                    "workflow_draft": draft,
                    "active_revision": None,
                    "display_status": display_status,
                    "physical_state": master_values.get("physical_state", ""),
                    "drafting_status": "",
                    "submitted_at": None,
                    "committee_slug": "",
                    "committee_title": "No active workflow",
                    "committee_track": "",
                    "active_revision_count": 0,
                    "can_generate_more": True,
                }
            )
        workflow_create_groups_map: dict[str, dict[str, object]] = {}
        for draft in drafts:
            subset_code = draft.subset or "OTHER"
            subset_display = draft.subset_display or "Other / Unmapped"
            group = workflow_create_groups_map.setdefault(
                subset_code,
                {
                    "subset_code": subset_code,
                    "subset_display": subset_display,
                    "rows": [],
                },
            )
            group["rows"].append(
                {
                    "parent_id": draft.id,
                    "spec_number": draft.spec_number,
                    "chemical_name": draft.chemical_name or "—",
                    "material_code": draft.material_code or "",
                    "active_revision_count": len(_get_active_revisions_for_parent(draft.id)),
                }
            )
        workflow_create_groups = list(workflow_create_groups_map.values())
        response = make_response(
            render_template(
                "csc/admin/drafts.html",
                workflow_rows=workflow_rows,
                subsets=subsets,
                workflow_create_groups=workflow_create_groups,
            )
        )
        # Safari's back/forward cache can restore an older drafts table after
        # navigating back from child pages. Mark this view as non-cacheable so
        # the browser always re-requests the current table layout.
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response
    except Exception:
        logger.exception("Error loading admin drafts")
        flash("Error loading drafts.", "danger")
        return redirect(url_for("csc.admin_index"))


@csc_bp.route("/admin/drafts")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_drafts():
    """Backward-compatible alias for the revision-generation page."""
    return redirect(url_for("csc.admin_revision_generation"))


@csc_bp.route("/admin/draft/<int:draft_id>/published-view")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_view_published_spec(draft_id: int):
    """Legacy admin alias redirected to the shared published-spec page."""
    return redirect(url_for("csc.published_spec_detail", draft_id=draft_id))


@csc_bp.route("/admin/draft/<int:draft_id>/edit")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_edit_draft(draft_id: int):
    """Legacy route kept only to redirect admins into the draft-generation flow."""
    flash("Direct admin material master editing has been removed. Open a self draft from Draft Generation for Revision in Secretary Workspace to make changes.", "info")
    return redirect(url_for("csc.admin_revision_generation"))


@csc_bp.route("/admin/draft/<int:draft_id>/update", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_update_draft(draft_id: int):
    """Legacy route kept only to block direct admin master-data edits."""
    message = (
        "Direct admin material master editing has been removed. "
        "Create or open your self draft from Draft Generation for Revision in Secretary Workspace instead."
    )
    if request.is_json:
        return jsonify({"error": message}), 403
    flash(message, "warning")
    return redirect(url_for("csc.admin_revision_generation"))


@csc_bp.route("/admin/draft/<int:draft_id>/lock-phase1", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_lock_phase1(draft_id: int):
    """Toggle Phase 1 lock."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        draft.phase1_locked = not draft.phase1_locked
        draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        _create_audit_entry(
            draft_id,
            f"Phase 1 {'locked' if draft.phase1_locked else 'unlocked'} (admin)",
        )

        return jsonify({
            "success": True,
            "phase1_locked": draft.phase1_locked,
        })
    except Exception:
        logger.exception("Error toggling phase1 lock")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


# Alias: templates use url_for('csc.api_admin_toggle_phase1_lock', draft_id=...)
@csc_bp.route("/admin/api/draft/<int:draft_id>/toggle-phase1-lock", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def api_admin_toggle_phase1_lock(draft_id: int):
    """Alias for admin_lock_phase1."""
    return admin_lock_phase1(draft_id)


@csc_bp.route("/admin/draft/<int:draft_id>/set-stage", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_set_stage(draft_id: int):
    """Set workflow stage (drafting/published)."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        data = request.get_json() or {}

        new_stage = data.get("admin_stage", ADMIN_STAGE_PUBLISHED)
        if new_stage not in (ADMIN_STAGE_DRAFTING, ADMIN_STAGE_PUBLISHED):
            return jsonify({"error": "Invalid stage"}), 400

        old_stage = draft.admin_stage
        draft.admin_stage = new_stage

        if new_stage == ADMIN_STAGE_PUBLISHED and draft.status != "Published":
            draft.status = "Published"
        elif new_stage == ADMIN_STAGE_DRAFTING:
            draft.status = "Drafting"

        draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        _create_audit_entry(
            draft_id,
            f"Admin stage changed: {old_stage} → {new_stage}",
        )

        return jsonify({"success": True, "new_stage": new_stage})
    except Exception:
        logger.exception("Error setting stage for draft_id=%s", draft_id)
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


# Alias: templates use url_for('csc.api_admin_set_stage', draft_id=...)
@csc_bp.route("/admin/api/draft/<int:draft_id>/set-stage", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def api_admin_set_stage(draft_id: int):
    """Alias for admin_set_stage."""
    return admin_set_stage(draft_id)


@csc_bp.route("/admin/draft/<int:draft_id>/delete", methods=["POST", "DELETE"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_delete_draft(draft_id: int):
    """Admin delete draft."""
    try:
        draft = db.session.get(CSCDraft, draft_id)
        if not draft:
            return jsonify({"error": "Draft not found"}), 404

        _create_audit_entry(draft_id, "Draft deleted (admin)")

        db.session.delete(draft)
        db.session.commit()

        return jsonify({"success": True})
    except Exception:
        logger.exception("Error deleting draft for draft_id=%s", draft_id)
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


# Alias: templates use url_for('csc.api_admin_delete_draft', draft_id=...)
@csc_bp.route("/admin/api/draft/<int:draft_id>/delete", methods=["POST", "DELETE"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def api_admin_delete_draft(draft_id: int):
    """Alias for admin_delete_draft."""
    return admin_delete_draft(draft_id)


@csc_bp.route("/admin/settings", methods=["GET"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_settings():
    """Settings to configure committee directory and upload office orders."""
    from app.models.csc.governance import CSCOfficeOrderFile
    payload = get_committee_config_payload()
    root_c = payload["ROOT_COMMITTEE"]
    child_c = payload["CHILD_COMMITTEES"]
    office_orders_raw = payload["OFFICE_ORDERS"]

    all_committees = [root_c] + list(child_c)
    username_options = []
    for user in (
        User.query.filter_by(is_active=True)
        .order_by(User.full_name.asc(), User.username.asc())
        .all()
    ):
        if not user.has_module_access("csc"):
            continue
        username_options.append(
            {
                "username": user.username,
                "label": f"{user.full_name} ({user.username})" if (user.full_name or "").strip() else user.username,
            }
        )

    # ── Office order upload status ────────────────────────────────────────
    try:
        db_slugs = {row.slug: row for row in db.session.query(CSCOfficeOrderFile).all()}
    except Exception:
        db_slugs = {}

    office_order_statuses = []
    for o in office_orders_raw:
        slug = o["slug"]
        in_db = slug in db_slugs
        office_order_statuses.append({
            "slug": slug,
            "title": o["title"],
            "in_db": in_db,
            "file_name": db_slugs[slug].file_name if in_db else None,
        })

    # ── Data blob for JS serialiser ────────────────────────────────────────
    config_data = {
        "all_committees": all_committees,
        "office_orders": office_orders_raw,
    }

    # ── Master subset catalogue (from root committee defaults) ────────────
    all_subsets = root_c["subsets"]

    return render_template(
        "csc/settings.html",
        all_committees=all_committees,
        all_subsets=all_subsets,
        office_order_statuses=office_order_statuses,
        config_data=config_data,
        username_options=username_options,
        username_values=[option["username"] for option in username_options],
    )


@csc_bp.route("/admin/settings/update", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_settings_update():
    from app.models.csc.governance import CSCConfig
    try:
        data_text = request.form.get("directory_json", "").strip()
        parsed = json.loads(data_text)
        if not isinstance(parsed, dict):
            raise ValueError("Must be a JSON object.")
        normalized = normalize_committee_config_payload(parsed)

        config = db.session.query(CSCConfig).first()
        if not config:
            config = CSCConfig()
            db.session.add(config)

        config.directory_json = json.dumps(normalized, indent=2)
        db.session.commit()
        flash("Committee configuration saved.", "success")
    except Exception:
        flash(f"Invalid JSON: {e}", "danger")
    
    return redirect(url_for("csc.admin_settings"))


@csc_bp.route("/admin/office-order/upload", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_upload_office_order():
    from app.models.csc.governance import CSCOfficeOrderFile
    slug = request.form.get("slug", "").strip()
    file_obj = request.files.get("file")

    if not slug or not file_obj:
        flash("Slug and file are required.", "danger")
        return redirect(url_for("csc.admin_settings"))

    file_data = file_obj.read()
    if not file_data:
        flash("Uploaded file is empty.", "danger")
        return redirect(url_for("csc.admin_settings"))

    order = db.session.get(CSCOfficeOrderFile, slug)
    if not order:
        order = CSCOfficeOrderFile(slug=slug)
        db.session.add(order)
    
    order.file_name = file_obj.filename or "office_order.pdf"
    order.file_data = file_data
    db.session.commit()

    flash(f"Office Order '{slug}' mapped and saved to database successfully.", "success")
    return redirect(url_for("csc.admin_settings"))


@csc_bp.route("/admin/revisions")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_revisions():
    """Material Master Admin workbench for committee-head-approved drafts."""
    return _render_admin_revisions_workbench()


def _render_admin_revisions_workbench(
    latest_import_summary: TypeClassificationImportSummary | MaterialHandlingImportSummary | None = None,
    latest_import_kind: str | None = None,
    pending_material_handling_preview: dict[str, object] | None = None,
    latest_import_report_url: str | None = None,
):
    """Render the Secretary workbench with an optional workbook import summary."""
    try:
        revisions = CSCRevision.query.order_by(
            CSCRevision.status != WORKFLOW_DRAFTING_HEAD_APPROVED,
            CSCRevision.updated_at.desc(),
            CSCRevision.id.desc(),
        ).all()
        revision_rows = [
            {
                "revision": revision,
                "draft_type": _draft_type_label(revision.child_draft) if revision.child_draft else "Revision",
                "is_new_spec": (_draft_type_label(revision.child_draft) == "New Specification") if revision.child_draft else False,
                "latest_updated_at": (
                    revision.child_draft.updated_at if revision.child_draft and revision.child_draft.updated_at else revision.updated_at
                ),
                "review_url": url_for("csc.review_revision", revision_id=revision.id),
                "workflow_stream": _infer_workflow_stream_name(revision.child_draft),
                "workflow_scope_label": _workflow_scope_display_label(revision.child_draft),
                "can_decide": _can_current_user_decide_revision_as_module_admin(revision),
                "can_edit": bool(
                    revision.child_draft
                    and _is_secretary_staging_state(revision.status)
                    and _can_edit_draft(revision.child_draft)
                ),
                "edit_url": (
                    url_for("csc.editor", draft_id=revision.child_draft.id)
                    if revision.child_draft and _is_secretary_staging_state(revision.status)
                    else ""
                ),
                "can_open_for_committee": bool(
                    revision.child_draft
                    and _is_secretary_staging_state(revision.status)
                ),
                "open_for_committee_url": (
                    url_for("csc.admin_open_revision_for_committee", revision_id=revision.id)
                    if revision.child_draft and _is_secretary_staging_state(revision.status)
                    else ""
                ),
                "can_force_delete": revision.status in {
                    WORKFLOW_DRAFTING_STAGING,
                    WORKFLOW_DRAFTING_OPEN,
                    WORKFLOW_DRAFTING_SUBMITTED,
                    WORKFLOW_DRAFTING_HEAD_APPROVED,
                    WORKFLOW_DRAFTING_RETURNED,
                },
                "force_delete_url": url_for("csc.admin_reject_revision", revision_id=revision.id),
                "decision_label": "Review & Publish" if revision.status == WORKFLOW_DRAFTING_HEAD_APPROVED else "View Snapshot",
                "decision_note": (
                    "Secretary staging: edit and send this draft to the committee workspace when ready."
                    if _is_secretary_staging_state(revision.status)
                    else
                    "Decision pending Material Master Management Admin action"
                    if revision.status == WORKFLOW_DRAFTING_HEAD_APPROVED
                    else "Delete remains available from the snapshot modal"
                ),
            }
            for revision in revisions
        ]

        response = make_response(
            render_template(
            "csc/admin/revisions.html",
            revision_rows=revision_rows,
            can_import_workbooks=True,
            workflow_nav_active="admin",
            workbench_title="Secretary Workbench",
            workbench_subtitle="Stage newly created drafts, edit and structure them as Secretary, then send them to committee users. Committee-head-approved drafts can then be published, returned, or deleted.",
            workbench_kicker_label="Secretary",
            workbench_kicker_url=url_for("csc.admin_index"),
            intro_message="Only the assigned Material Master Management Admins can access this dashboard. New workflow drafts enter Secretary staging first. After Secretary edits are complete, send them to committee users by moving them to Open. Publish remains limited to committee-head-approved drafts.",
            latest_import_summary=latest_import_summary,
            latest_import_kind=latest_import_kind,
            pending_material_handling_preview=pending_material_handling_preview,
            latest_import_report_url=latest_import_report_url,
            decision_stage=WORKFLOW_DRAFTING_HEAD_APPROVED,
            decision_mode="module_admin",
            review_subtitle_pending="Review the verified snapshot below, then publish, return, or delete without editing the submitted draft.",
            review_subtitle_readonly="Revision snapshot for audit history. Material Master Management Admin can still delete the active draft from this modal.",
            decision_note="Publish updates the master specification. Return sends the draft back for committee changes. Delete removes the active workflow draft.",
            primary_action_label="Publish",
            return_action_label="Return",
            reject_action_label="Delete",
            review_endpoint_name="csc.review_revision",
            approve_endpoint_name="csc.admin_approve_revision",
            return_endpoint_name="csc.admin_return_revision",
            reject_endpoint_name="csc.admin_reject_revision",
            require_publish_fields=True,
            )
        )
        return _mark_response_no_store(response)
    except Exception:
        logger.exception("Error loading revisions")
        flash("Error loading revisions.", "danger")
        return redirect(url_for("csc.admin_index"))


def _revision_can_be_force_deleted_by_secretary(revision: CSCRevision | None) -> bool:
    """Return whether a revision is deletable from the Secretary workbench."""
    return bool(
        revision
        and revision.status in {
            WORKFLOW_DRAFTING_STAGING,
            WORKFLOW_DRAFTING_OPEN,
            WORKFLOW_DRAFTING_SUBMITTED,
            WORKFLOW_DRAFTING_HEAD_APPROVED,
            WORKFLOW_DRAFTING_RETURNED,
        }
    )


def _build_workbook_created_notification_message(
    stream_label: str,
    subset_codes: list[str] | set[str] | None,
    draft_labels: list[str] | None,
) -> str:
    """Format grouped workbook-import notifications with a numbered draft list."""
    normalized_subset_codes = sorted(
        {
            str(code or "").strip().upper()
            for code in (subset_codes or [])
            if str(code or "").strip()
        }
    )
    subset_phrase = ""
    if normalized_subset_codes:
        subset_phrase = f" under your subset coverage ({', '.join(normalized_subset_codes)})"

    cleaned_labels = [
        str(label or "").strip()
        for label in (draft_labels or [])
        if str(label or "").strip()
    ]
    if not cleaned_labels:
        return (
            f"{stream_label} drafts have been created{subset_phrase}. "
            "Review them in your committee workspace."
        )

    numbered_lines = [f"{index}. {label}" for index, label in enumerate(cleaned_labels, start=1)]
    return (
        f"{stream_label} drafts have been created for the following chemicals{subset_phrase}:\n"
        + "\n".join(numbered_lines)
        + "\nReview them in your committee workspace."
    )


def _delete_revision_as_module_admin(
    revision: CSCRevision,
    reviewer_notes: str,
    *,
    commit_changes: bool = True,
) -> dict[str, object]:
    """Delete one active workflow revision using the Secretary workbench rules."""
    if not revision:
        raise ValueError("Revision not found")
    if not _revision_can_be_force_deleted_by_secretary(revision):
        raise ValueError("Only staged or active drafts can be deleted")
    if revision.child_draft is None or revision.parent_draft is None:
        raise ValueError("Revision draft is incomplete")

    child_draft = revision.child_draft
    parent_draft = revision.parent_draft
    is_new_spec = _draft_type_label(child_draft) == "New Specification"

    parent_draft_id = parent_draft.id
    parent_spec_number = parent_draft.spec_number
    parent_chemical_name = parent_draft.chemical_name
    proposer_id = child_draft.created_by_id
    child_draft_id = child_draft.id
    committee_slug = _get_revision_committee_slug(revision)
    committee_head_user_ids = _get_committee_user_ids_for_committee_slug(
        committee_slug,
        "committee_head",
    )
    if revision.committee_head_user_id:
        committee_head_user_ids.append(revision.committee_head_user_id)

    revision.status = WORKFLOW_DRAFTING_REJECTED
    revision.reviewed_at = datetime.now(timezone.utc)
    revision.reviewer_notes = reviewer_notes
    revision.module_admin_reviewed_at = datetime.now(timezone.utc)
    revision.module_admin_notes = reviewer_notes
    revision.module_admin_user_id = current_user.id
    if not is_new_spec:
        parent_draft.status = "Published"
        parent_draft.admin_stage = ADMIN_STAGE_PUBLISHED
        parent_draft.updated_at = datetime.now(timezone.utc)

    if not is_new_spec:
        db.session.add(
            CSCAudit(
                draft_id=parent_draft_id,
                action="Drafting rejected and deleted",
                user_name=current_user.username,
                action_time=datetime.now(timezone.utc),
                new_value=json.dumps({"submitted_draft_id": child_draft_id}),
                remarks=reviewer_notes,
            )
        )

    seen_user_ids: set[int] = set()
    for user_id in [proposer_id, *committee_head_user_ids]:
        if not user_id or user_id in seen_user_ids:
            continue
        seen_user_ids.add(user_id)
        create_notification(
            user_id=user_id,
            title="Draft deleted",
            message=(
                f"{parent_spec_number} — {parent_chemical_name} was deleted by the Material Master Management Admin. "
                f"Note: {reviewer_notes}"
            ),
            severity="danger",
            link=url_for("csc.workspace"),
        )

    clear_staged_master_payload(child_draft)
    db.session.delete(revision)
    db.session.delete(child_draft)
    if is_new_spec:
        db.session.delete(parent_draft)
    if commit_changes:
        db.session.commit()

    return {
        "parent_draft_id": parent_draft_id,
        "child_draft_id": child_draft_id,
        "spec_number": parent_spec_number,
        "chemical_name": parent_chemical_name,
        "is_new_spec": is_new_spec,
    }


def _material_handling_import_staging_paths(token: str) -> dict[str, str]:
    root = os.path.join(MATERIAL_HANDLING_IMPORT_STAGING_ROOT, token)
    return {
        "root": root,
        "payload": os.path.join(root, "payload.json"),
    }


def _material_handling_import_report_paths(token: str) -> dict[str, str]:
    root = os.path.join(MATERIAL_HANDLING_IMPORT_REPORT_ROOT, token)
    return {
        "root": root,
        "report": os.path.join(root, "material_handling_import_report.xlsx"),
    }


def _write_material_handling_import_report(
    import_summary: MaterialHandlingImportSummary,
    *,
    opened_count: int,
    already_open_count: int,
) -> str:
    token = uuid.uuid4().hex
    paths = _material_handling_import_report_paths(token)
    os.makedirs(paths["root"], exist_ok=True)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Workbook", import_summary.workbook_name])
    summary_sheet.append(["Rows Processed", import_summary.rows_processed])
    summary_sheet.append(["Specs Matched", import_summary.specs_matched])
    summary_sheet.append(["Drafts Created", import_summary.drafts_created])
    summary_sheet.append(["Drafts Reused", import_summary.drafts_reused])
    summary_sheet.append(["Rows Updated", import_summary.rows_updated])
    summary_sheet.append(["Staged Field Updates", import_summary.staged_field_updates])
    summary_sheet.append(["Drafts Pushed To Committee", opened_count])
    summary_sheet.append(["Drafts Already Open", already_open_count])
    summary_sheet.append(["Rows Without Draft Creation", len(import_summary.unmatched_rows)])

    success_sheet = workbook.create_sheet("Draft Workflow Rows")
    success_sheet.append(
        [
            "Row Number",
            "Current Spec No.",
            "Proposed Spec No.",
            "Chemical Name",
            "Material Code",
            "Parent Draft ID",
            "Draft ID",
            "Workflow Action",
            "Staged Field Count",
        ]
    )
    for detail in import_summary.details:
        success_sheet.append(
            [
                detail.row_number,
                detail.source_spec_number,
                detail.proposed_spec_number,
                detail.chemical_name,
                detail.material_code,
                detail.parent_draft_id,
                detail.draft_id,
                "Created" if detail.created_new else "Reused",
                detail.staged_field_count,
            ]
        )

    failure_sheet = workbook.create_sheet("Drafts Not Created")
    failure_sheet.append(
        [
            "Row Number",
            "Current Spec No.",
            "Proposed Spec No.",
            "Chemical Name",
            "Material Code",
            "Reason",
        ]
    )
    for item in import_summary.unmatched_rows:
        failure_sheet.append(
            [
                item.row_number,
                item.source_spec_number,
                item.proposed_spec_number,
                item.chemical_name,
                item.material_code,
                item.reason,
            ]
        )

    workbook.save(paths["report"])
    return token


def _stage_material_handling_workbook_payload(workbook: ParsedMaterialHandlingWorkbook) -> str:
    token = uuid.uuid4().hex
    paths = _material_handling_import_staging_paths(token)
    os.makedirs(paths["root"], exist_ok=True)
    payload = {
        "workbook_name": workbook.workbook_name,
        "rows": [asdict(row) for row in workbook.rows],
        "invalid_rows": [asdict(row) for row in workbook.invalid_rows],
    }
    with open(paths["payload"], "w", encoding="utf-8") as handle:
        json.dump(payload, handle)
    return token


def _load_staged_material_handling_workbook_payload(token: str) -> ParsedMaterialHandlingWorkbook:
    paths = _material_handling_import_staging_paths(token)
    if not os.path.exists(paths["payload"]):
        raise ValueError("The staged material handling review is no longer available. Upload the workbook again.")
    with open(paths["payload"], "r", encoding="utf-8") as handle:
        payload = json.load(handle)
    return ParsedMaterialHandlingWorkbook(
        workbook_name=str(payload.get("workbook_name") or "material_handling.xlsx"),
        rows=[
            MaterialHandlingWorkbookRow(**row_payload)
            for row_payload in (payload.get("rows") or [])
            if isinstance(row_payload, dict)
        ],
        invalid_rows=[
            MaterialHandlingRowIssue(**row_payload)
            for row_payload in (payload.get("invalid_rows") or [])
            if isinstance(row_payload, dict)
        ],
    )


def _discard_staged_material_handling_workbook_payload(token: str | None) -> None:
    if not token:
        return
    paths = _material_handling_import_staging_paths(token)
    if os.path.isdir(paths["root"]):
        shutil.rmtree(paths["root"], ignore_errors=True)


def _preview_text_value(value: object) -> str:
    return str(value or "").strip()


def _strip_system_bracket_lines(text: str) -> str:
    cleaned = SYSTEM_BRACKET_LINE_RE.sub("", text or "")
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def _resolve_material_handling_row_for_preview(
    row: MaterialHandlingWorkbookRow,
    parent_draft: CSCDraft,
) -> MaterialHandlingWorkbookRow:
    return MaterialHandlingWorkbookRow(
        row_number=row.row_number,
        source_spec_number=row.source_spec_number or canonicalize_spec_number(parent_draft.spec_number),
        proposed_spec_number=row.proposed_spec_number,
        chemical_name=row.chemical_name or _preview_text_value(parent_draft.chemical_name),
        material_code=row.material_code,
        unit=row.unit,
        officers_responsible=row.officers_responsible,
        physical_state=row.physical_state,
        volatility_ambient_temperature=row.volatility_ambient_temperature,
        sunlight_sensitivity_up_to_50c=row.sunlight_sensitivity_up_to_50c,
        moisture_sensitivity=row.moisture_sensitivity,
        refrigeration_required=row.refrigeration_required,
        reactivity=row.reactivity,
        flammable=row.flammable,
        toxic=row.toxic,
        corrosive=row.corrosive,
        storage_conditions_general=row.storage_conditions_general,
        storage_conditions_special=row.storage_conditions_special,
        container_type=row.container_type,
        container_capacity=row.container_capacity,
        container_description=row.container_description,
        primary_storage_classification=row.primary_storage_classification,
        remarks=row.remarks,
    )


def _predict_material_handling_import_target(parent_draft: CSCDraft) -> dict[str, object]:
    committee_slug = _resolve_requested_workflow_committee_slug(
        parent_draft,
        WORKFLOW_TRACK_MATERIAL_HANDLING,
    )
    if not committee_slug:
        return {
            "ready": False,
            "reason": "No committee is configured for the material handling workflow track.",
            "target_draft": None,
            "draft_id": None,
            "workflow_action": "Blocked",
            "created_new": False,
        }

    active_revisions = _get_active_revisions_for_parent(parent_draft.id)
    existing_for_committee = _get_active_revision_for_parent_committee(parent_draft.id, committee_slug)
    if existing_for_committee and existing_for_committee.child_draft:
        return {
            "ready": True,
            "reason": "",
            "target_draft": existing_for_committee.child_draft,
            "draft_id": existing_for_committee.child_draft.id,
            "workflow_action": "Reuse Draft",
            "created_new": False,
        }

    if len(active_revisions) >= 2:
        return {
            "ready": False,
            "reason": "Only two active workflow drafts are allowed per specification at a time.",
            "target_draft": parent_draft,
            "draft_id": None,
            "workflow_action": "Blocked",
            "created_new": False,
        }

    return {
        "ready": True,
        "reason": "",
        "target_draft": parent_draft,
        "draft_id": None,
        "workflow_action": "Create Draft",
        "created_new": True,
    }


def _material_handling_preview_comparisons(
    target_draft: CSCDraft,
    workbook_row: MaterialHandlingWorkbookRow,
    resolved_row: MaterialHandlingWorkbookRow,
) -> list[dict[str, object]]:
    current_master_values = get_master_form_values(target_draft)
    rows = [
        ("Current Spec Reference", canonicalize_spec_number(target_draft.spec_number), workbook_row.source_spec_number, resolved_row.source_spec_number),
        ("Specification No.", canonicalize_spec_number(target_draft.spec_number), workbook_row.proposed_spec_number, resolved_row.proposed_spec_number),
        ("Chemical Name", _preview_text_value(target_draft.chemical_name), workbook_row.chemical_name, resolved_row.chemical_name),
        ("Material Code", _preview_text_value(target_draft.material_code), workbook_row.material_code, resolved_row.material_code),
        ("Unit", "", workbook_row.unit, resolved_row.unit),
        ("Officers Responsible", "", workbook_row.officers_responsible, resolved_row.officers_responsible),
        ("Physical State", _preview_text_value(current_master_values.get("physical_state")), workbook_row.physical_state, resolved_row.physical_state),
        ("Volatility at Ambient Temperature", _preview_text_value(current_master_values.get("volatility_ambient_temperature")), workbook_row.volatility_ambient_temperature, resolved_row.volatility_ambient_temperature),
        ("Sunlight Sensitivity (up to 50degC atmospheric temperature)", _preview_text_value(current_master_values.get("sunlight_sensitivity_up_to_50c")), workbook_row.sunlight_sensitivity_up_to_50c, resolved_row.sunlight_sensitivity_up_to_50c),
        ("Moisture Sensitivity", _preview_text_value(current_master_values.get("moisture_sensitivity")), workbook_row.moisture_sensitivity, resolved_row.moisture_sensitivity),
        ("Refrigeration required (Yes / No)", _preview_text_value(current_master_values.get("refrigeration_required")), workbook_row.refrigeration_required, resolved_row.refrigeration_required),
        ("Reactivity", _preview_text_value(current_master_values.get("reactivity")), workbook_row.reactivity, resolved_row.reactivity),
        ("Flammable", _preview_text_value(current_master_values.get("flammable")), workbook_row.flammable, resolved_row.flammable),
        ("Toxic", _preview_text_value(current_master_values.get("toxic")), workbook_row.toxic, resolved_row.toxic),
        ("Corrosive", _preview_text_value(current_master_values.get("corrosive")), workbook_row.corrosive, resolved_row.corrosive),
        ("Storage Conditions - General", _preview_text_value(current_master_values.get("storage_conditions_general")), workbook_row.storage_conditions_general, resolved_row.storage_conditions_general),
        ("Storage Conditions - Special", _preview_text_value(current_master_values.get("storage_conditions_special")), workbook_row.storage_conditions_special, resolved_row.storage_conditions_special),
        ("Packing Type 1", _preview_text_value(current_master_values.get("container_type")), workbook_row.container_type, resolved_row.container_type),
        ("Packing Size 1", _preview_text_value(current_master_values.get("container_capacity")), workbook_row.container_capacity, resolved_row.container_capacity),
        ("Packing Description 1", _preview_text_value(current_master_values.get("container_description")), workbook_row.container_description, resolved_row.container_description),
        ("Primary Storage Classification", _preview_text_value(current_master_values.get("primary_storage_classification")), workbook_row.primary_storage_classification, resolved_row.primary_storage_classification),
        ("Remarks", "", workbook_row.remarks, resolved_row.remarks),
    ]
    comparisons: list[dict[str, object]] = []
    for label, current_value, workbook_value, staged_value in rows:
        current_text = _preview_text_value(current_value)
        workbook_text = _preview_text_value(workbook_value)
        staged_text = _preview_text_value(staged_value)
        comparisons.append(
            {
                "label": label,
                "current_value": current_text,
                "workbook_value": workbook_text,
                "staged_value": staged_text,
                "changed": current_text != staged_text,
            }
        )
    return comparisons


def _build_material_handling_import_preview(
    workbook: ParsedMaterialHandlingWorkbook,
    parent_lookup: dict[str, CSCDraft],
) -> dict[str, object]:
    preview_rows: list[dict[str, object]] = []
    preview_issues: list[dict[str, object]] = [
        {
            "issue_type": "Invalid Workbook Row",
            "row_number": row.row_number,
            "source_spec_number": row.source_spec_number,
            "proposed_spec_number": row.proposed_spec_number,
            "chemical_name": row.chemical_name,
            "material_code": row.material_code,
            "reason": row.reason,
        }
        for row in workbook.invalid_rows
    ]

    create_candidates = 0
    reuse_candidates = 0
    ready_rows = 0

    for workbook_row in workbook.rows:
        parent_draft = parent_lookup.get(material_code_lookup_key(workbook_row.material_code))
        if parent_draft is None:
            preview_issues.append(
                {
                    "issue_type": "No Parent Match",
                    "row_number": workbook_row.row_number,
                    "source_spec_number": workbook_row.source_spec_number,
                    "proposed_spec_number": workbook_row.proposed_spec_number,
                    "chemical_name": workbook_row.chemical_name,
                    "material_code": workbook_row.material_code,
                    "reason": "No published parent specification matched this Material Code.",
                }
            )
            continue

        target_prediction = _predict_material_handling_import_target(parent_draft)
        target_draft = target_prediction.get("target_draft") or parent_draft
        resolved_row = _resolve_material_handling_row_for_preview(workbook_row, parent_draft)
        current_background = _get_draft_section_text(target_draft, "background")
        current_justification = _get_draft_section_text(target_draft, "justification")
        preview_rows.append(
            {
                "row_number": workbook_row.row_number,
                "parent_draft_id": parent_draft.id,
                "parent_spec_number": parent_draft.spec_number,
                "draft_id": target_prediction.get("draft_id"),
                "material_code": workbook_row.material_code,
                "chemical_name": resolved_row.chemical_name,
                "workflow_action": target_prediction.get("workflow_action"),
                "ready": bool(target_prediction.get("ready")),
                "reason": target_prediction.get("reason") or "",
                "changed_field_count": 0,
                "comparisons": [],
                "background_current": _strip_system_bracket_lines(current_background),
                "background_result": _strip_system_bracket_lines(merge_material_handling_section_block(
                    current_background,
                    build_material_handling_background_text(resolved_row),
                    start_marker=MATERIAL_HANDLING_BACKGROUND_START,
                    end_marker=MATERIAL_HANDLING_BACKGROUND_END,
                )),
                "justification_current": _strip_system_bracket_lines(current_justification),
                "justification_result": _strip_system_bracket_lines(merge_material_handling_section_block(
                    current_justification,
                    build_material_handling_justification_text(resolved_row),
                    start_marker=MATERIAL_HANDLING_JUSTIFICATION_START,
                    end_marker=MATERIAL_HANDLING_JUSTIFICATION_END,
                )),
            }
        )
        preview_rows[-1]["comparisons"] = _material_handling_preview_comparisons(
            target_draft,
            workbook_row,
            resolved_row,
        )
        preview_rows[-1]["changed_field_count"] = sum(
            1 for item in preview_rows[-1]["comparisons"] if item.get("changed")
        )
        if target_prediction.get("created_new"):
            create_candidates += 1
        elif target_prediction.get("ready"):
            reuse_candidates += 1
        if target_prediction.get("ready"):
            ready_rows += 1

    return {
        "token": "",
        "workbook_name": workbook.workbook_name,
        "subset_codes": list(workbook.subset_codes),
        "rows_processed": len(workbook.rows),
        "ready_rows": ready_rows,
        "create_candidates": create_candidates,
        "reuse_candidates": reuse_candidates,
        "issue_count": len(preview_issues) + sum(1 for row in preview_rows if not row.get("ready")),
        "preview_rows": preview_rows,
        "preview_issues": preview_issues,
    }


def _type_classification_import_scope() -> dict[str, bool]:
    """Open imported drafts in the existing subset workflow with parameter edits enabled."""
    return {
        "material_properties": False,
        "storage_handling": False,
        "parameters": True,
        "impact": False,
        "parameter_type": True,
        "parameter_other": True,
    }


def _material_handling_import_scope() -> dict[str, bool]:
    """Open imported drafts in the material handling workflow, including impact."""
    return {
        "material_properties": True,
        "storage_handling": True,
        "parameters": False,
        "impact": True,
        "parameter_type": False,
        "parameter_other": False,
    }


def _get_draft_section_text(draft: CSCDraft, section_name: str) -> str:
    for candidate_name in _section_name_aliases(section_name):
        section = CSCSection.query.filter_by(draft_id=draft.id, section_name=candidate_name).first()
        if section is not None:
            return section.section_text or ""
    return ""


def _set_draft_section_text(draft: CSCDraft, section_name: str, section_text: str) -> None:
    alias_names = _section_name_aliases(section_name)
    canonical_name = alias_names[0] if alias_names else _canonical_section_name(section_name)
    section_sort_orders = {
        section_name: sort_order
        for section_name, _default_text, sort_order in _default_section_rows(_infer_workflow_stream_name(draft))
    }
    section = None
    for candidate_name in alias_names:
        section = CSCSection.query.filter_by(draft_id=draft.id, section_name=candidate_name).first()
        if section is not None:
            break
    if section is None:
        section = CSCSection(
            draft_id=draft.id,
            section_name=canonical_name,
            section_text=section_text,
            sort_order=section_sort_orders.get(canonical_name, 0),
        )
        db.session.add(section)
        return
    section.section_name = canonical_name
    section.sort_order = section_sort_orders.get(canonical_name, section.sort_order)
    section.section_text = section_text
    for candidate_name in alias_names:
        if candidate_name == canonical_name:
            continue
        duplicate = CSCSection.query.filter_by(draft_id=draft.id, section_name=candidate_name).first()
        if duplicate is not None and duplicate.id != section.id:
            db.session.delete(duplicate)


def _stage_material_handling_master_values(draft: CSCDraft, staged_values: dict[str, str]) -> int:
    before_values = get_master_form_values(draft)
    material_values = {
        key: value
        for key, value in staged_values.items()
        if key in {field["field_name"] for field in get_material_properties_fields()}
    }
    storage_values = {
        key: value
        for key, value in staged_values.items()
        if key in {field["field_name"] for field in get_storage_handling_fields()}
    }
    if material_values:
        update_material_properties_from_form(draft, material_values, user_id=current_user.id)
    if storage_values:
        update_storage_handling_from_form(draft, storage_values, user_id=current_user.id)
    after_values = get_master_form_values(draft)
    changed_count = 0
    for key, value in staged_values.items():
        if (before_values.get(key) or "") != (after_values.get(key) or ""):
            changed_count += 1
    return changed_count


def _apply_material_handling_workbook_import(
    workbook: ParsedMaterialHandlingWorkbook,
) -> MaterialHandlingImportSummary:
    material_codes = list(
        {
            row.material_code
            for row in workbook.rows
            if (row.material_code or "").strip()
        }
    )
    parent_drafts: list[CSCDraft] = []
    if material_codes:
        parent_drafts = (
            CSCDraft.query.filter(CSCDraft.parent_draft_id.is_(None))
            .filter(CSCDraft.material_code.in_(material_codes))
            .all()
        )
    parent_lookup = build_parent_lookup_by_material_code(parent_drafts)

    import_summary = import_material_handling_workbook(
        workbook,
        parent_lookup,
        open_draft_for_parent=lambda parent_draft: _open_committee_workflow_draft_for_parent(
            parent_draft,
            _material_handling_import_scope(),
            WORKFLOW_TRACK_MATERIAL_HANDLING,
        ),
        get_background_text=lambda draft: _get_draft_section_text(draft, "background"),
        set_background_text=lambda draft, text: _set_draft_section_text(draft, "background", text),
        get_justification_text=lambda draft: _get_draft_section_text(draft, "justification"),
        set_justification_text=lambda draft, text: _set_draft_section_text(draft, "justification", text),
        stage_master_values=lambda draft, values: _stage_material_handling_master_values(draft, values),
    )

    touched_parent_ids: set[int] = set()
    for detail in import_summary.details:
        if detail.parent_draft_id:
            touched_parent_ids.add(detail.parent_draft_id)
    for parent_id in touched_parent_ids:
        parent_draft = db.session.get(CSCDraft, parent_id)
        if parent_draft is not None:
            parent_draft.updated_at = datetime.now(timezone.utc)
    db.session.commit()

    audit_rows = []
    for detail in import_summary.details:
        if not detail.parent_draft_id:
            continue
        audit_rows.append(
            CSCAudit(
                draft_id=detail.parent_draft_id,
                action="Material handling workbook imported",
                user_name=current_user.username,
                action_time=datetime.now(timezone.utc),
                new_value=json.dumps(
                    {
                        "workbook_name": import_summary.workbook_name,
                        "row_number": detail.row_number,
                        "source_spec_number": detail.source_spec_number,
                        "proposed_spec_number": detail.proposed_spec_number,
                        "child_draft_id": detail.draft_id,
                        "created_new": detail.created_new,
                        "staged_field_count": detail.staged_field_count,
                    }
                ),
            )
        )
    if audit_rows:
        db.session.add_all(audit_rows)
        db.session.commit()

    return import_summary


def _open_revisions_for_committee_workflow_bulk(
    revisions: list[CSCRevision],
) -> tuple[int, int]:
    """Batch-open staged revisions for committee users to avoid timeout-heavy per-row commits."""
    unique_revisions: dict[int, CSCRevision] = {}
    for revision in revisions:
        if revision is None or getattr(revision, "id", None) is None:
            continue
        unique_revisions[int(revision.id)] = revision

    opened_revisions: list[CSCRevision] = []
    already_open_count = 0
    now = datetime.now(timezone.utc)

    for revision in unique_revisions.values():
        if revision.child_draft is None or revision.parent_draft is None:
            continue
        if _is_secretary_staging_state(revision.status):
            revision.status = WORKFLOW_DRAFTING_OPEN
            revision.updated_at = now
            revision.parent_draft.status = "Drafting"
            revision.parent_draft.admin_stage = ADMIN_STAGE_DRAFTING
            revision.parent_draft.updated_at = now
            revision.child_draft.updated_at = now
            opened_revisions.append(revision)
        elif revision.status in {WORKFLOW_DRAFTING_OPEN, WORKFLOW_DRAFTING_RETURNED}:
            already_open_count += 1

    if not opened_revisions:
        return 0, already_open_count

    db.session.flush()

    audit_rows: list[CSCAudit] = []
    notifications_by_user: dict[int, dict[str, object]] = {}
    for revision in opened_revisions:
        committee_slug = _get_revision_committee_slug(revision)
        committee_user_ids = _get_committee_user_ids_for_revision(revision)
        workflow_stream = _infer_workflow_stream_name(revision.child_draft)
        stream_label = "Material Handling" if workflow_stream == MATERIAL_HANDLING_STREAM else "Type Classification"
        subset_code = _workflow_subset_code_for_draft(revision.child_draft, revision)
        for user_id in committee_user_ids:
            payload = notifications_by_user.setdefault(
                user_id,
                {
                    "stream_label": stream_label,
                    "subset_codes": set(),
                    "draft_labels": [],
                },
            )
            payload["stream_label"] = stream_label
            if subset_code:
                payload["subset_codes"].add(subset_code)
            draft_spec_number = str(
                revision.child_draft.spec_number or revision.parent_draft.spec_number or ""
            ).strip()
            chemical_name = str(
                revision.parent_draft.chemical_name or revision.child_draft.chemical_name or ""
            ).strip()
            draft_label = " — ".join(part for part in [draft_spec_number, chemical_name] if part)
            if not draft_label:
                draft_label = f"Draft {revision.child_draft.id}"
            if draft_label not in payload["draft_labels"]:
                payload["draft_labels"].append(draft_label)
        audit_rows.append(
            CSCAudit(
                draft_id=revision.parent_draft.id,
                action="Draft opened for committee user",
                user_name=current_user.username,
                action_time=now,
                new_value=json.dumps(
                    {
                        "revision_id": revision.id,
                        "child_draft_id": revision.child_draft.id,
                        "committee_slug": committee_slug,
                    }
                ),
            )
        )

    for user_id, payload in notifications_by_user.items():
        create_notification(
            user_id=user_id,
            title=f"{payload['stream_label']} drafts created for your subset",
            message=_build_workbook_created_notification_message(
                payload["stream_label"],
                payload["subset_codes"],
                payload["draft_labels"],
            ),
            severity="info",
            link=url_for("csc.workspace"),
        )

    if audit_rows:
        db.session.add_all(audit_rows)
    db.session.commit()
    return len(opened_revisions), already_open_count


@csc_bp.route("/admin/revisions/import-type-classification", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_import_type_classification_workbook():
    """Import one type-classification workbook into Secretary workflow drafts."""
    upload = request.files.get("type_classification_workbook")
    if upload is None or not upload.filename:
        flash("Choose a .xlsx workbook to import.", "danger")
        return _render_admin_revisions_workbench()

    filename = Path(upload.filename).name
    if not filename.lower().endswith(".xlsx"):
        flash("Only .xlsx workbooks are supported for type classification import.", "danger")
        return _render_admin_revisions_workbench()

    try:
        workbook = parse_type_classification_workbook(upload, filename)
        if not workbook.sheets and not workbook.invalid_sheets:
            flash("The uploaded workbook did not contain any populated specification sheets.", "warning")
            return _render_admin_revisions_workbench()

        parent_drafts = CSCDraft.query.filter(CSCDraft.parent_draft_id.is_(None)).all()
        parent_lookup_by_material_code = build_parent_lookup_by_material_code(parent_drafts)

        import_summary = import_type_classification_workbook(
            workbook,
            parent_lookup_by_material_code,
            open_draft_for_parent=lambda parent_draft: _open_committee_workflow_draft_for_parent(
                parent_draft,
                _type_classification_import_scope(),
                WORKFLOW_TRACK_SUBSET,
            ),
            get_justification_text=lambda draft: _get_draft_section_text(draft, "justification"),
            set_justification_text=lambda draft, text: _set_draft_section_text(draft, "justification", text),
        )

        touched_parent_ids: set[int] = set()
        created_draft_ids: list[int] = []
        for detail in import_summary.details:
            if detail.parent_draft_id:
                touched_parent_ids.add(detail.parent_draft_id)
            if detail.created_new and detail.draft_id:
                created_draft_ids.append(detail.draft_id)

        for parent_id in touched_parent_ids:
            parent_draft = db.session.get(CSCDraft, parent_id)
            if parent_draft is not None:
                parent_draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        for detail in import_summary.details:
            if detail.parent_draft_id:
                _create_audit_entry(
                    detail.parent_draft_id,
                    "Type classification workbook imported",
                    new_value=json.dumps(
                        {
                            "workbook_name": import_summary.workbook_name,
                            "sheet_name": detail.sheet_name,
                            "child_draft_id": detail.draft_id,
                            "created_new": detail.created_new,
                            "parameters_matched": detail.parameters_matched,
                            "parameters_updated": detail.parameters_updated,
                            "justification_entries_created": detail.justification_entries_created,
                        }
                    ),
                )

        if import_summary.details:
            flash(
                (
                    f"Imported {import_summary.workbook_name}: "
                    f"{import_summary.specs_matched} specification(s) matched, "
                    f"{import_summary.parameters_updated} parameter row(s) updated in Secretary staging."
                ),
                "success",
            )
        elif import_summary.unmatched_sheets:
            flash(
                "The workbook was parsed, but no specifications could be imported. Review the unmatched-sheet report below.",
                "warning",
            )
        else:
            flash("No changes were applied from the workbook.", "info")

        return _render_admin_revisions_workbench(
            latest_import_summary=import_summary,
            latest_import_kind=TYPE_CLASSIFICATION_STREAM,
        )
    except Exception:
        logger.exception("Error importing type classification workbook")
        db.session.rollback()
        flash("Error importing type classification workbook.", "danger")
        return _render_admin_revisions_workbench()


@csc_bp.route("/admin/revisions/import-material-handling", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_import_material_handling_workbook():
    """Parse one material handling workbook and stage a Secretary review preview."""
    upload = request.files.get("material_handling_workbook")
    if upload is None or not upload.filename:
        flash("Choose a .xlsx workbook to import.", "danger")
        return _render_admin_revisions_workbench()

    filename = Path(upload.filename).name
    if not filename.lower().endswith(".xlsx"):
        flash("Only .xlsx workbooks are supported for material handling import.", "danger")
        return _render_admin_revisions_workbench()

    try:
        workbook = parse_material_handling_workbook(upload, filename)
        if not workbook.rows and not workbook.invalid_rows:
            flash("The uploaded workbook did not contain any populated material handling rows.", "warning")
            return _render_admin_revisions_workbench()

        material_codes = list(
            {
                row.material_code
                for row in workbook.rows
                if (row.material_code or "").strip()
            }
        )
        parent_drafts = []
        if material_codes:
            parent_drafts = (
                CSCDraft.query.filter(CSCDraft.parent_draft_id.is_(None))
                .filter(CSCDraft.material_code.in_(material_codes))
                .all()
            )
        parent_lookup = build_parent_lookup_by_material_code(parent_drafts)
        preview_payload = _build_material_handling_import_preview(workbook, parent_lookup)
        preview_payload["token"] = _stage_material_handling_workbook_payload(workbook)
        if preview_payload["ready_rows"]:
            flash(
                (
                    f"Reviewed {workbook.workbook_name}: "
                    f"{preview_payload['ready_rows']} row(s) are ready for Secretary staging. "
                    "Validate the side-by-side comparison below, then confirm import."
                ),
                "info",
            )
        else:
            flash(
                "The workbook was parsed, but no material handling rows are currently ready for staging. Review the comparison report below.",
                "warning",
            )
        return _render_admin_revisions_workbench(
            pending_material_handling_preview=preview_payload,
        )
    except Exception:
        logger.exception("Error importing material handling workbook")
        db.session.rollback()
        flash("Error importing material handling workbook.", "danger")
        return _render_admin_revisions_workbench()


@csc_bp.route("/admin/revisions/import-material-handling/confirm", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_confirm_material_handling_workbook_import():
    """Apply a previously staged material handling workbook review and push drafts to committee users."""
    token = (request.form.get("preview_token") or "").strip()
    if not token:
        flash("The staged material handling review token is missing. Upload the workbook again.", "danger")
        return _render_admin_revisions_workbench()

    try:
        workbook = _load_staged_material_handling_workbook_payload(token)
        import_summary = _apply_material_handling_workbook_import(workbook)
        _discard_staged_material_handling_workbook_payload(token)

        revisions_to_open: list[CSCRevision] = []
        for detail in import_summary.details:
            if not detail.draft_id:
                continue
            revision = _get_revision_for_child(detail.draft_id)
            if revision is None:
                continue
            revisions_to_open.append(revision)

        opened_count, already_open_count = _open_revisions_for_committee_workflow_bulk(
            revisions_to_open
        )
        report_token = _write_material_handling_import_report(
            import_summary,
            opened_count=opened_count,
            already_open_count=already_open_count,
        )

        if import_summary.details:
            flash(
                (
                    f"Accepted {import_summary.workbook_name}: "
                    f"{import_summary.specs_matched} specification(s) matched, "
                    f"{import_summary.rows_updated} material handling row(s) staged, "
                    f"{opened_count} draft(s) pushed to committee users"
                    + (f", {already_open_count} already open." if already_open_count else ".")
                ),
                "success",
            )
        elif import_summary.unmatched_rows:
            flash(
                "The workbook was parsed, but no material handling rows could be imported. Review the unmatched-row report below.",
                "warning",
            )
        else:
            flash("No changes were applied from the workbook.", "info")

        return _render_admin_revisions_workbench(
            latest_import_summary=import_summary,
            latest_import_kind=MATERIAL_HANDLING_STREAM,
            latest_import_report_url=url_for(
                "csc.admin_download_material_handling_import_report",
                token=report_token,
            ),
        )
    except Exception:
        logger.exception("Error confirming material handling workbook import")
        db.session.rollback()
        flash("Error confirming material handling workbook import.", "danger")
        return _render_admin_revisions_workbench()


@csc_bp.route("/admin/revisions/import-material-handling/report/<token>")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_download_material_handling_import_report(token: str):
    """Download the confirmed material-handling import report as Excel."""
    paths = _material_handling_import_report_paths((token or "").strip())
    if not os.path.exists(paths["report"]):
        flash("The material handling import report is no longer available.", "warning")
        return redirect(url_for("csc.admin_revisions"))

    filename = (
        f"material_handling_import_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    response = send_file(
        paths["report"],
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
        max_age=0,
    )
    return _mark_response_no_store(response)


@csc_bp.route("/admin/revisions/import-material-handling/cancel", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_cancel_material_handling_workbook_import():
    """Discard a staged material handling workbook review."""
    token = (request.form.get("preview_token") or "").strip()
    _discard_staged_material_handling_workbook_payload(token)
    flash("Discarded the staged material handling workbook review.", "info")
    return _render_admin_revisions_workbench()


def _open_revision_for_committee_workflow(revision: CSCRevision) -> tuple[bool, str]:
    if revision.child_draft is None or revision.parent_draft is None:
        return False, "Revision draft is incomplete."
    if _is_secretary_staging_state(revision.status):
        revision.status = WORKFLOW_DRAFTING_OPEN
        revision.updated_at = datetime.now(timezone.utc)
        revision.parent_draft.status = "Drafting"
        revision.parent_draft.admin_stage = ADMIN_STAGE_DRAFTING
        revision.parent_draft.updated_at = datetime.now(timezone.utc)
        revision.child_draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        committee_slug = _get_revision_committee_slug(revision)
        committee_user_ids = _get_committee_user_ids_for_committee_slug(
            committee_slug,
            "committee_user",
        )
        if committee_user_ids:
            _notify_users(
                committee_user_ids,
                title="Draft opened for committee revision",
                message=(
                    f"{revision.parent_draft.spec_number} — {revision.parent_draft.chemical_name} "
                    "has been opened by Secretary and is now available in your committee workspace."
                ),
                severity="info",
                link=url_for("csc.workspace"),
            )

        _create_audit_entry(
            revision.parent_draft.id,
            "Draft opened for committee user",
            new_value=json.dumps(
                {
                    "revision_id": revision.id,
                    "child_draft_id": revision.child_draft.id,
                    "committee_slug": committee_slug,
                }
            ),
        )
        db.session.commit()
        return True, (
            f"{revision.parent_draft.spec_number} — {revision.parent_draft.chemical_name} "
            "is now open in the committee workspace."
        )

    if revision.status in {WORKFLOW_DRAFTING_OPEN, WORKFLOW_DRAFTING_RETURNED}:
        return True, (
            f"{revision.parent_draft.spec_number} — {revision.parent_draft.chemical_name} "
            "is already available in the committee workspace."
        )

    return False, "Only staged drafts can be opened for committee editing."


@csc_bp.route("/revisions")
@login_required
@module_access_required("csc")
def review_workbench():
    """Committee-head, governance, and module-admin workbench for active revisions."""
    head_committee_slugs = set(_get_current_user_committee_head_slugs())
    governance_member = _current_user_is_governance_committee_member()
    if not governance_member and not head_committee_slugs:
        abort(403)

    try:
        revisions = CSCRevision.query.order_by(
            CSCRevision.updated_at.desc(),
            CSCRevision.id.desc(),
        ).all()
        active_statuses = {
            WORKFLOW_DRAFTING_OPEN,
            WORKFLOW_DRAFTING_SUBMITTED,
            WORKFLOW_DRAFTING_HEAD_APPROVED,
            WORKFLOW_DRAFTING_RETURNED,
        }
        revision_rows = [
            {
                "revision": revision,
                "draft_type": _draft_type_label(revision.child_draft) if revision.child_draft else "Revision",
                "is_new_spec": (_draft_type_label(revision.child_draft) == "New Specification") if revision.child_draft else False,
                "latest_updated_at": (
                    revision.child_draft.updated_at if revision.child_draft and revision.child_draft.updated_at else revision.updated_at
                ),
                "review_url": url_for("csc.review_revision", revision_id=revision.id),
                "workflow_stream": _infer_workflow_stream_name(revision.child_draft),
                "workflow_scope_label": _workflow_scope_display_label(revision.child_draft),
                "can_decide": _can_current_user_decide_revision_as_committee_head(revision),
                "can_force_delete": False,
                "force_delete_url": "",
                "can_edit": _can_current_user_edit_draft_as_committee_head(revision.child_draft) if revision.child_draft else False,
                "edit_url": url_for("csc.editor", draft_id=revision.child_draft.id) if revision.child_draft else "",
                "decision_label": (
                    "Review & Decide"
                    if _can_current_user_decide_revision_as_committee_head(revision)
                    else "View Snapshot"
                ),
                "decision_note": (
                    "Review and edit the submitted draft before forwarding it to Secretary"
                    if _can_current_user_decide_revision_as_committee_head(revision)
                    else "Read-only governance snapshot"
                ),
            }
            for revision in revisions
            if revision.status in active_statuses
            and (
                governance_member
                or _revision_in_current_user_committee_head_scope(revision)
            )
        ]

        return render_template(
            "csc/admin/revisions.html",
            revision_rows=revision_rows,
            can_import_workbooks=False,
            workflow_nav_active="review",
            workbench_title="Committee Head Workbench",
            workbench_subtitle="You are the Head of the Committee / Nodal Officer for the Specifications Subset. Review submitted drafts, edit them where needed, then forward them to Secretary.",
            workbench_kicker_label="Material Master Management",
            workbench_kicker_url=url_for("csc.index"),
            intro_message=(
                "Governance Committee members can inspect all active revisions. Subset committee heads can review and edit submitted drafts within their covered subsets, then approve and forward them to Secretary. Draft deletion is reserved for Secretary Workspace."
            ),
            decision_stage=WORKFLOW_DRAFTING_SUBMITTED,
            decision_mode="committee_head",
            review_subtitle_pending="Review the revision snapshot below, then return it or approve and forward it as the Committee Head / Nodal Officer.",
            review_subtitle_readonly="Read-only revision snapshot for governance visibility and decision history.",
            decision_note="Approve forwards the draft to Secretary. Return reopens it for committee updates.",
            primary_action_label="Approve & Forward",
            return_action_label="Return",
            reject_action_label=None,
            review_endpoint_name="csc.review_revision",
            approve_endpoint_name="csc.committee_head_approve_revision",
            return_endpoint_name="csc.committee_head_return_revision",
            reject_endpoint_name="csc.committee_head_delete_revision",
            require_publish_fields=False,
        )
    except Exception:
        logger.exception("Error loading revisions")
        flash("Error loading revisions.", "danger")
        return redirect(url_for("csc.index"))


@csc_bp.route("/admin/msds-diagnostics")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_msds_diagnostics():
    """Report live web-process MSDS diagnostics for Railway debugging."""
    from sqlalchemy import inspect

    from app.core.services.msds_service import (
        get_msds_storage_diagnostics,
        list_msds_files,
    )

    payload: dict[str, object] = {
        "app_environment_name": current_app.config.get("APP_ENVIRONMENT_NAME"),
        "db_host": current_app.config.get("DB_HOST"),
        "db_name": current_app.config.get("DB_NAME"),
        "db_port": current_app.config.get("DB_PORT"),
        "migration_head": None,
        "current_database": None,
        "msds_table_present": False,
        "msds_query_ok": False,
        "msds_count": None,
        "msds_data_column_type": None,
        "msds_data_column_capacity_bytes": None,
        "error": None,
    }
    try:
        current_revision = db.session.execute(
            text("SELECT version_num FROM alembic_version LIMIT 1")
        ).scalar()
        payload["migration_head"] = current_revision
    except Exception as exc:
        payload["migration_head"] = f"error: {exc}"

    try:
        payload["current_database"] = db.engine.url.database
        payload["msds_table_present"] = "msds_files" in inspect(db.engine).get_table_names()
        storage = get_msds_storage_diagnostics()
        payload["msds_data_column_type"] = storage.get("data_column_type")
        payload["msds_data_column_capacity_bytes"] = storage.get("data_column_capacity_bytes")
        documents = list_msds_files()
        payload["msds_query_ok"] = True
        payload["msds_count"] = len(documents)
    except Exception as exc:
        payload["error"] = str(exc)
        return jsonify(payload), 500

    return jsonify(payload)


@csc_bp.route("/revision/<int:revision_id>/review")
@login_required
@module_access_required("csc")
def review_revision(revision_id: int):
    """Return a read-only revision snapshot for workbench modals."""
    revision = db.session.get(CSCRevision, revision_id)
    if not revision:
        return jsonify({"error": "Revision not found"}), 404
    _require_revision_snapshot_access(revision)

    decision_mode = "read_only"
    if _can_current_user_decide_revision_as_module_admin(revision):
        decision_mode = "module_admin"
    elif _can_current_user_decide_revision_as_committee_head(revision):
        decision_mode = "committee_head"

    response = jsonify(
        {
            "success": True,
            "status": revision.status,
            "decision_mode": decision_mode,
            "html": render_template(
                "csc/admin/_revision_detail.html",
                revision=revision,
                **_build_revision_review_context(revision),
            ),
        }
    )
    return _mark_response_no_store(response)


@csc_bp.route("/admin/revision/<int:revision_id>/review")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_review_revision(revision_id: int):
    """Alias for module-admin snapshot review."""
    return review_revision(revision_id)


@csc_bp.route("/admin/revision/<int:revision_id>/open-for-committee", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_open_revision_for_committee(revision_id: int):
    """Move a Secretary-staged workflow draft into committee-open state."""
    try:
        revision = db.session.get(CSCRevision, revision_id)
        if not revision or revision.child_draft is None or revision.parent_draft is None:
            return jsonify({"error": "Revision not found"}), 404
        opened, message = _open_revision_for_committee_workflow(revision)
        if not opened:
            return jsonify({"error": message}), 409

        return jsonify(
            {
                "success": True,
                "message": message,
            }
        )
    except Exception:
        logger.exception("Error opening staged revision for committee")
        db.session.rollback()
        return jsonify({"error": "Unable to open the staged draft for committee editing."}), 500


@csc_bp.route("/revision/<int:revision_id>/approve", methods=["POST"])
@login_required
@module_access_required("csc")
def committee_head_approve_revision(revision_id: int):
    """Approve a submitted draft as committee head and forward it to module admins."""
    try:
        revision = db.session.get(CSCRevision, revision_id)
        if not revision:
            return jsonify({"error": "Revision not found"}), 404
        if revision.status != WORKFLOW_DRAFTING_SUBMITTED:
            return jsonify({"error": "This draft is no longer awaiting committee head approval."}), 409
        if not _revision_in_current_user_committee_head_scope(revision):
            return jsonify({"error": "You are not assigned as the committee head for this draft subset."}), 403

        data = request.get_json() or {}
        reviewer_notes = (data.get("reviewer_notes") or "").strip()
        if not reviewer_notes:
            return jsonify({"error": "Reviewer notes are required"}), 400

        revision.status = WORKFLOW_DRAFTING_HEAD_APPROVED
        revision.committee_head_reviewed_at = datetime.now(timezone.utc)
        revision.committee_head_notes = reviewer_notes
        revision.committee_head_user_id = current_user.id
        revision.reviewer_notes = reviewer_notes
        revision.reviewed_at = datetime.now(timezone.utc)
        if revision.child_draft:
            revision.child_draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        if revision.parent_draft:
            _create_audit_entry(
                revision.parent_draft.id,
                "Draft approved by committee head",
                new_value=json.dumps({"submitted_draft_id": revision.child_draft_id}),
                remarks=reviewer_notes,
            )

        _notify_material_master_admins(
            title="Draft ready for Secretary review",
            message=(
                f"{revision.parent_draft.spec_number} — {revision.parent_draft.chemical_name} "
                "was approved by the committee head and is ready for Secretary review."
            ),
            severity="info",
            link=url_for("csc.admin_revisions"),
        )

        return jsonify(
            {
                "success": True,
                "title": "Draft Approved And Forwarded",
                "message": (
                    f"{revision.parent_draft.spec_number} — {revision.parent_draft.chemical_name} "
                    "was approved by the committee head and forwarded to Secretary."
                ),
            }
        )
    except Exception:
        logger.exception("Error approving revision as committee head")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/revision/<int:revision_id>/return", methods=["POST"])
@login_required
@module_access_required("csc")
def committee_head_return_revision(revision_id: int):
    """Return a submitted draft to committee user editing."""
    try:
        revision = db.session.get(CSCRevision, revision_id)
        if not revision:
            return jsonify({"error": "Revision not found"}), 404
        if revision.status != WORKFLOW_DRAFTING_SUBMITTED:
            return jsonify({"error": "This draft is no longer awaiting committee head review."}), 409
        if not _revision_in_current_user_committee_head_scope(revision):
            return jsonify({"error": "You are not assigned as the committee head for this draft subset."}), 403

        data = request.get_json() or {}
        reviewer_notes = (data.get("reviewer_notes") or "").strip()
        if not reviewer_notes:
            return jsonify({"error": "Reviewer notes are required"}), 400

        revision.status = WORKFLOW_DRAFTING_RETURNED
        revision.committee_head_reviewed_at = datetime.now(timezone.utc)
        revision.committee_head_notes = reviewer_notes
        revision.committee_head_user_id = current_user.id
        revision.reviewer_notes = reviewer_notes
        revision.reviewed_at = datetime.now(timezone.utc)
        if revision.child_draft:
            revision.child_draft.updated_at = datetime.now(timezone.utc)
        if revision.parent_draft:
            revision.parent_draft.status = "Drafting"
            revision.parent_draft.admin_stage = ADMIN_STAGE_DRAFTING
            revision.parent_draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        if revision.parent_draft:
            _create_audit_entry(
                revision.parent_draft.id,
                "Draft returned by committee head",
                new_value=json.dumps({"submitted_draft_id": revision.child_draft_id}),
                remarks=reviewer_notes,
            )

        _notify_users(
            [revision.child_draft.created_by_id if revision.child_draft else None],
            title="Draft returned by committee head",
            message=(
                f"{revision.parent_draft.spec_number} — {revision.parent_draft.chemical_name} "
                f"was returned by the committee head. Note: {reviewer_notes}"
            ),
            severity="warning",
            link=url_for("csc.workspace", q=revision.parent_draft.chemical_name if revision.parent_draft else ""),
        )

        return jsonify({"success": True})
    except Exception:
        logger.exception("Error returning revision as committee head")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/revision/<int:revision_id>/delete", methods=["POST"])
@login_required
@module_access_required("csc")
def committee_head_delete_revision(revision_id: int):
    """Committee Heads cannot delete workflow drafts."""
    return jsonify({"error": "Only Secretary Workspace can delete drafts for revision."}), 403


def _publish_revision_to_parent(
    revision: CSCRevision,
    *,
    version_increment: str,
    reviewer_notes: str,
    approval_authorized_by: str,
    approval_disha_file_number: str,
    allowed_statuses: set[str] | None = None,
    source_action_prefix: str = "admin_approve",
    audit_action: str = "Drafting approved and published",
) -> dict[str, object]:
    """Publish one workflow revision into the parent and stage the snapshot/audit rows."""
    if revision is None or revision.child_draft is None or revision.parent_draft is None:
        raise ValueError("Revision not found")

    allowed = allowed_statuses or {WORKFLOW_DRAFTING_HEAD_APPROVED}
    if revision.status not in allowed:
        raise ValueError("Only committee-head-approved drafts can be published")
    if version_increment not in {"whole", "decimal"}:
        raise ValueError("Choose whole or decimal version increment")
    if not reviewer_notes:
        raise ValueError("Approval notes are required")
    if not approval_authorized_by:
        raise ValueError("Enter the Approved By name before publishing")
    if not approval_disha_file_number:
        raise ValueError("Enter the Disha File Number before publishing")

    child_draft = revision.child_draft
    parent_draft = revision.parent_draft
    stream_name = _infer_workflow_stream_name(child_draft)
    next_spec_version = increment_spec_version(
        parent_draft.spec_version,
        version_increment,
    )
    next_version_label = format_spec_version(next_spec_version)
    approval_notes = _compose_approval_notes(
        reviewer_notes,
        approval_authorized_by,
        approval_disha_file_number,
    )
    published_review_context = _build_revision_review_context(revision)
    _override_review_context_summary_value(
        published_review_context,
        "Version",
        f"v{next_version_label}",
    )
    _override_review_context_summary_value(
        published_review_context,
        "Current Status",
        "Published",
    )
    published_review_context["latest_review_notes"] = approval_notes

    published_master_data = _apply_approved_revision_stream_to_parent(
        parent_draft,
        child_draft,
        revision,
    )
    record = upsert_master_record_from_form(
        parent_draft,
        published_master_data,
        user_id=current_user.id,
    )
    if record is not None and record not in db.session:
        db.session.add(record)

    revision.status = WORKFLOW_DRAFTING_APPROVED
    revision.reviewed_at = datetime.now(timezone.utc)
    revision.reviewer_notes = approval_notes
    revision.module_admin_reviewed_at = datetime.now(timezone.utc)
    revision.module_admin_notes = revision.reviewer_notes
    revision.module_admin_user_id = current_user.id

    parent_draft.status = "Published"
    parent_draft.admin_stage = ADMIN_STAGE_PUBLISHED
    parent_draft.reviewed_by = current_user.username
    parent_draft.spec_version = next_spec_version
    parent_draft.updated_at = datetime.now(timezone.utc)

    snapshot = CSCSpecVersion(
        draft_id=parent_draft.id,
        spec_version=parent_draft.spec_version,
        created_by=current_user.username,
        source_action=f"{source_action_prefix}_{stream_name}_{version_increment}",
        remarks=revision.reviewer_notes or None,
        payload_json=json.dumps(
            {
                "draft": parent_draft.to_dict(),
                "master_data": published_master_data,
                "bundle": _build_export_bundle(parent_draft),
                "review_context": _serialize_review_context_for_snapshot(published_review_context),
            }
        ),
    )
    db.session.add(snapshot)
    clear_staged_master_payload(child_draft)

    proposer_id = child_draft.created_by_id
    committee_head_user_id = revision.committee_head_user_id
    child_draft_id = child_draft.id
    db.session.delete(revision)
    db.session.delete(child_draft)

    _create_audit_entry(
        parent_draft.id,
        audit_action,
        new_value=json.dumps(
            {
                "version": next_version_label,
                "increment_type": version_increment,
                "submitted_draft_id": child_draft_id,
                "workflow_stream": stream_name,
            }
        ),
        remarks=approval_notes,
    )

    return {
        "parent_draft_id": parent_draft.id,
        "spec_number": parent_draft.spec_number,
        "chemical_name": parent_draft.chemical_name,
        "applied_version": next_version_label,
        "proposer_id": proposer_id,
        "committee_head_user_id": committee_head_user_id,
    }


@csc_bp.route("/admin/revision/<int:revision_id>/approve", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_approve_revision(revision_id: int):
    """Approve a submitted drafting request."""
    try:
        revision = db.session.get(CSCRevision, revision_id)
        if not revision:
            return jsonify({"error": "Revision not found"}), 404

        data = request.get_json() or {}
        try:
            publish_result = _publish_revision_to_parent(
                revision,
                version_increment=(data.get("version_increment") or "whole").strip().lower(),
                reviewer_notes=(data.get("reviewer_notes") or "").strip(),
                approval_authorized_by=(data.get("approval_authorized_by") or "").strip(),
                approval_disha_file_number=(data.get("approval_disha_file_number") or "").strip(),
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

        db.session.commit()
        _notify_users(
            [publish_result.get("proposer_id"), publish_result.get("committee_head_user_id")],
            title="Draft published",
            message=(
                f"{publish_result['spec_number']} — {publish_result['chemical_name']} was approved "
                f"and published as v{publish_result['applied_version']}."
            ),
            severity="success",
            link=url_for("csc.workspace"),
        )
        return jsonify({"success": True})
    except Exception:
        logger.exception("Error approving revision")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


# Alias: templates use url_for('csc.api_admin_approve_revision', revision_id=...)
@csc_bp.route("/admin/api/revision/<int:revision_id>/approve", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def api_admin_approve_revision(revision_id: int):
    """Alias for admin_approve_revision."""
    return admin_approve_revision(revision_id)


@csc_bp.route("/admin/revision/<int:revision_id>/reject", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_reject_revision(revision_id: int):
    """Delete an active drafting request as Material Master Management Admin."""
    try:
        revision = db.session.get(CSCRevision, revision_id)
        if not revision:
            return jsonify({"error": "Revision not found"}), 404

        data = request.get_json() or {}
        if not _revision_can_be_force_deleted_by_secretary(revision):
            return jsonify({"error": "Only staged or active drafts can be deleted"}), 400

        reviewer_notes = (data.get("reviewer_notes") or "").strip()
        if not reviewer_notes:
            return jsonify({"error": "Reviewer notes are required"}), 400

        _delete_revision_as_module_admin(revision, reviewer_notes)

        return jsonify({"success": True})
    except Exception:
        logger.exception("Error rejecting revision")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


@csc_bp.route("/admin/revision/<int:revision_id>/return", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_return_revision(revision_id: int):
    """Return a committee-head-approved drafting request to open state with notes."""
    try:
        revision = db.session.get(CSCRevision, revision_id)
        if not revision:
            return jsonify({"error": "Revision not found"}), 404
        if revision.status != WORKFLOW_DRAFTING_HEAD_APPROVED:
            return jsonify({"error": "Only committee-head-approved drafts can be returned"}), 400

        data = request.get_json() or {}
        reviewer_notes = (data.get("reviewer_notes") or "").strip()
        if not reviewer_notes:
            return jsonify({"error": "Reviewer notes are required"}), 400

        child_draft = revision.child_draft
        parent_draft = revision.parent_draft
        revision.status = WORKFLOW_DRAFTING_RETURNED
        revision.reviewed_at = datetime.now(timezone.utc)
        revision.reviewer_notes = reviewer_notes
        revision.module_admin_reviewed_at = datetime.now(timezone.utc)
        revision.module_admin_notes = reviewer_notes
        revision.module_admin_user_id = current_user.id
        parent_draft.status = "Drafting"
        parent_draft.admin_stage = ADMIN_STAGE_DRAFTING
        child_draft.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        _create_audit_entry(
            parent_draft.id,
            "Drafting returned for update",
            new_value=json.dumps({"submitted_draft_id": child_draft.id}),
            remarks=reviewer_notes,
        )
        _notify_users(
            [child_draft.created_by_id, revision.committee_head_user_id],
            title="Draft returned for update",
            message=(
                f"{parent_draft.spec_number} — {parent_draft.chemical_name} was returned by the Material Master Management Admin. "
                f"Note: {reviewer_notes}"
            ),
            severity="warning",
            link=url_for("csc.workspace", q=parent_draft.chemical_name),
        )

        return jsonify({"success": True})
    except Exception:
        logger.exception("Error returning revision")
        db.session.rollback()
        return jsonify({"error": "An internal error occurred. Please try again."}), 500


# Alias: templates use url_for('csc.api_admin_reject_revision', revision_id=...)
@csc_bp.route("/admin/api/revision/<int:revision_id>/reject", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def api_admin_reject_revision(revision_id: int):
    """Alias for admin_reject_revision."""
    return admin_reject_revision(revision_id)


@csc_bp.route("/admin/api/revision/<int:revision_id>/return", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def api_admin_return_revision(revision_id: int):
    """Alias for admin_return_revision."""
    return admin_return_revision(revision_id)


@csc_bp.route("/admin/audit")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_audit_log():
    """Audit log view for CSC workflow actions."""
    entries = (
        CSCAudit.query.order_by(CSCAudit.action_time.desc(), CSCAudit.id.desc())
        .limit(500)
        .all()
    )
    return render_template("csc/admin/audit.html", entries=entries)


@csc_bp.route("/admin/audit/reset-test-state", methods=["POST"])
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_reset_audit_test_state():
    """Audit-log reset action has been removed."""
    abort(404)


@csc_bp.route("/admin/audit/download")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_audit_log_download():
    """Download CSC audit log as CSV."""
    entries = CSCAudit.query.order_by(CSCAudit.action_time.desc(), CSCAudit.id.desc()).all()
    buffer = io.StringIO()
    buffer.write("audit_id,action_time,user_name,draft_id,spec_number,chemical_name,action,remarks\n")
    for entry in entries:
        spec_number = (entry.draft.spec_number if entry.draft else "") or ""
        chemical_name = (entry.draft.chemical_name if entry.draft else "") or ""
        row = [
            str(entry.id),
            entry.action_time.isoformat() if entry.action_time else "",
            entry.user_name or "",
            str(entry.draft_id or ""),
            spec_number,
            chemical_name,
            (entry.action or "").replace('"', "'"),
            (entry.remarks or "").replace('"', "'").replace("\n", " "),
        ]
        buffer.write(",".join(f'"{value}"' for value in row) + "\n")

    response = make_response(buffer.getvalue())
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    response.headers["Content-Disposition"] = "attachment; filename=csc_audit_log.csv"
    return response


@csc_bp.route("/admin/export")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_export():
    """Export page with stats and history."""
    stats = _get_export_stats()
    return render_template(
        "csc/admin/export.html",
        stats=stats,
        subset_options=_get_export_subset_options(),
        export_history=[],  # TODO: track export history in a model
    )


@csc_bp.route("/admin/export/master-docx")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_export_master():
    """Download master spec document."""
    try:
        selected_subsets = [
            value.strip().upper()
            for value in request.args.getlist("subset")
            if value.strip()
        ]
        include_draft_notes = request.args.get("include_draft_notes") == "1"
        include_changelog = request.args.get("include_changelog") == "1"
        include_metadata = request.args.get("include_metadata") == "1"

        query = CSCDraft.query.filter_by(status="Published", parent_draft_id=None)
        published_drafts = sorted(query.all(), key=_draft_sort_key)

        if selected_subsets:
            published_drafts = [
                draft for draft in published_drafts
                if (parse_spec_number(draft.spec_number or "")[0] or "") in selected_subsets
            ]

        if not published_drafts:
            flash("No published specifications matched the selected subsets.", "warning")
            return redirect(url_for("csc.admin_export"))

        all_specs = [_build_export_bundle(draft) for draft in published_drafts]
        changelog = _get_version_changelog() if include_changelog else None
        doc_bytes = build_master_spec_document(
            all_specs,
            include_draft_note=include_draft_notes,
            changelog=changelog,
            include_metadata=include_metadata,
        )

        subset_label = "-".join(selected_subsets) if selected_subsets else "ALL"
        filename = (
            f"ONGC_CSC_MasterSpec_{subset_label}_{datetime.now().strftime('%Y%m%d_%H%M')}.docx"
        )
        _record_export_audit("MASTER_DOCX_EXPORT", None, filename)
        return send_file(
            io.BytesIO(doc_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            as_attachment=True,
            download_name=filename,
        )
    except Exception:
        logger.exception("Error exporting master")
        flash("Error exporting master document.", "danger")
        return redirect(url_for("csc.admin_export"))


@csc_bp.route("/admin/export/<int:export_id>/download")
@login_required
@module_access_required("csc")
@module_admin_required("csc")
def admin_download_export(export_id: int):
    """Download a previously generated export file."""
    # TODO: Implement export file storage and retrieval
    flash("Export file not found.", "warning")
    return redirect(url_for("csc.admin_export"))
